"""Styled Excel (.xlsx) exports — presentation-quality financial reports.

Requires the 'openpyxl' package (pip install openpyxl). Import this module
lazily inside export routes so a missing package degrades to a friendly
flash message instead of breaking the app.
"""
from io import BytesIO

from flask import make_response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NUM_FMT = '#,##0.00'

GREEN_DARK = '107C10'
RED_DARK = 'C42B1C'
BLUE_DARK = '0F6CBD'
INDIGO = '3730A3'
GREY_TEXT = '605E5C'

_thin = Side(style='thin', color='D9D9D9')
BORDER_THIN = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _fill(hex6):
    return PatternFill('solid', fgColor=hex6)


def new_workbook(sheet_title):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    return wb, ws


def title_block(ws, ncols, company, title, subtitle=''):
    """Company name + report title + subtitle, centred across the columns."""
    end = get_column_letter(max(1, ncols))
    ws.merge_cells(f'A1:{end}1')
    ws.merge_cells(f'A2:{end}2')
    ws.merge_cells(f'A3:{end}3')
    c = ws['A1']
    c.value = company or ''
    c.font = Font(size=16, bold=True, color='1A1A2E')
    c.alignment = Alignment(horizontal='center')
    c = ws['A2']
    c.value = title
    c.font = Font(size=12, bold=True, color=GREY_TEXT)
    c.alignment = Alignment(horizontal='center')
    c = ws['A3']
    c.value = subtitle
    c.font = Font(size=10, italic=True, color=GREY_TEXT)
    c.alignment = Alignment(horizontal='center')
    return 5  # first content row


def header_row(ws, row, labels):
    for i, text in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=text)
        c.font = Font(bold=True, size=9, color=GREY_TEXT)
        c.fill = _fill('F3F3F3')
        c.border = BORDER_THIN
        c.alignment = Alignment(horizontal='right' if i > 1 else 'left', wrap_text=True)
    return row + 1


def section_row(ws, row, ncols, text, color=GREEN_DARK, bg='EAF6EA'):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(1, ncols))
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=10, color=color)
    for i in range(1, max(1, ncols) + 1):
        ws.cell(row=row, column=i).fill = _fill(bg)
    return row + 1


def item_row(ws, row, label, amounts, indent=1, color='333333'):
    c = ws.cell(row=row, column=1, value=label)
    c.font = Font(size=10, color=color)
    c.alignment = Alignment(indent=indent)
    for i, v in enumerate(amounts, start=2):
        m = ws.cell(row=row, column=i, value=float(v or 0))
        m.number_format = NUM_FMT
        m.font = Font(size=10, color=color)
        m.alignment = Alignment(horizontal='right')
    return row + 1


def data_row(ws, row, values, num_cols=(), color='333333'):
    """Generic table row: values in order; 1-based column indexes in num_cols are
    formatted as amounts (right-aligned, thousand separators)."""
    for i, v in enumerate(values, start=1):
        if i in num_cols:
            c = ws.cell(row=row, column=i, value=float(v or 0))
            c.number_format = NUM_FMT
            c.alignment = Alignment(horizontal='right')
        else:
            c = ws.cell(row=row, column=i, value=v)
        c.font = Font(size=10, color=color)
    return row + 1


def total_row(ws, row, label, amounts, bg='F3F3F3', color='1A1A2E', size=10):
    c = ws.cell(row=row, column=1, value=label)
    c.font = Font(bold=True, size=size, color=color)
    c.fill = _fill(bg)
    c.alignment = Alignment(horizontal='right')
    n = len(list(amounts))
    for i, v in enumerate(amounts, start=2):
        m = ws.cell(row=row, column=i, value=float(v or 0))
        m.number_format = NUM_FMT
        m.font = Font(bold=True, size=size, color=color)
        m.fill = _fill(bg)
        m.alignment = Alignment(horizontal='right')
    return row + 1


def finish(ws, ncols, first_col_width=48, num_col_width=20):
    ws.column_dimensions['A'].width = first_col_width
    for i in range(2, max(1, ncols) + 1):
        ws.column_dimensions[get_column_letter(i)].width = num_col_width
    ws.sheet_view.showGridLines = False


def workbook_response(wb, filename):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = make_response(buf.getvalue())
    resp.headers['Content-Disposition'] = f'attachment; filename={filename}'
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return resp
