import flask
from flask import render_template, request, redirect, url_for, flash, session, make_response, Response, stream_with_context, jsonify
from database import Database
from datetime import datetime, date
from functools import wraps
from jinja2 import pass_context
from werkzeug.security import generate_password_hash, check_password_hash

def is_safe_db_name(name):
    import re
    return bool(re.match(r'^[a-zA-Z0-9_]+$', str(name)))

import csv
import base64
import io
import json
import os

import difflib
import time
import knowledge_base
import random # For mocking exchange rate
import subprocess
import requests
import string
from datetime import timedelta
import mysql.connector
import urllib.request
import typing
from dataclasses import dataclass

app = flask.Flask(__name__)

# Global cache for exchange rates
exchange_rate_cache = {}
CACHE_DURATION = 3600  # 1 hour
import tempfile

# Global cache for categories
_category_cache = {}

# Global cache for POS items
_pos_items_cache = {}

def get_cached_categories(db):
    global _category_cache
    if not _category_cache:
        bs = db.execute_query("SELECT name_of_category, holding_position FROM balance_sheet_category")
        pl = db.execute_query("SELECT name_of_category, holding_position FROM `p&l_category`")
        cf = db.execute_query("SELECT catogory_name FROM cf_catogory ORDER BY hold_level, catogory_name")
        _category_cache = {'bs_cats': bs, 'pl_cats': pl, 'cf_cats': cf}
    return _category_cache['bs_cats'], _category_cache['pl_cats'], _category_cache['cf_cats']

def clear_category_cache():
    global _category_cache
    _category_cache.clear()

def clear_pos_items_cache():
    global _pos_items_cache
    _pos_items_cache.pop(get_session_db_name(), None)

import services
from num2words import num2words
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()
import logging
import shutil
import re
import os
import PyPDF2
import io

import migrations


# Configure Logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("app.log"),
        logging.StreamHandler()
    ]
)

# Set a secret key for session management.
# In production, this should be set via environment variable.
app.secret_key = os.environ.get('SECRET_KEY')
if not app.secret_key:
    # Use a default development key if not provided, but log a loud warning
    logging.warning("No SECRET_KEY set in environment variables. Falling back to default development key. "
                    "This is unsafe for production. Please set SECRET_KEY in your .env file.")
    app.secret_key = 'default-development-secret-key-change-this-in-production'

app.config['SECRET_KEY'] = app.secret_key

# Theme Configuration
THEMES = {
    'default': {
        'name': 'Professional (Default)',
        'primary': '#0f172a',
        'secondary': '#1e293b',
        'accent': '#2563eb'
    },
    'ocean': {
        'name': 'Ocean Blue',
        'primary': '#003366',
        'secondary': '#004080',
        'accent': '#0073e6'
    },
    'pro_blue': {
        'name': 'Pro Sky Blue',
        'primary': '#4188ff',
        'secondary': '#649eff',
        'accent': '#92bbff'
    },
    'forest': {
        'name': 'Forest Green',
        'primary': '#143d14',
        'secondary': '#1f5c1f',
        'accent': '#2eb82e'
    },
    'ruby': {
        'name': 'Ruby Red',
        'primary': '#4d0000',
        'secondary': '#800000',
        'accent': '#e60000'
    },
    'midnight': {
        'name': 'Midnight Dark',
        'primary': '#000000',
        'secondary': '#1a1a1a',
        'accent': '#ffcc00'
    },
    'sap_neutral_deep': {
        'name': 'SAP Neutral Deep',
        'primary': '#00305D',
        'secondary': '#003B72',
        'accent': '#0055A5'
    },
    'sap_neutral_mid': {
        'name': 'SAP Neutral Mid',
        'primary': '#0065C3',
        'secondary': '#0074E2',
        'accent': '#168EFF'
    },
    'sap_neutral_light': {
        'name': 'SAP Neutral Light',
        'primary': '#3FA2FF',
        'secondary': '#62B3FF',
        'accent': '#8BC7FF'
    }
}

# ── Sidebar Menu Registry ─────────────────────────────────────────────────────
# Single source of truth for all sidebar items. Used by superadmin menu control.
MENU_ITEMS_REGISTRY = [
    # Core Accounting
    {'key': 'pending_approvals',  'label': 'Pending Approvals',   'url': '/approvals',              'icon': 'fas fa-check-double',        'category': 'Core Accounting'},
    {'key': 'sales_invoice',      'label': 'Sales Invoice',        'url': '/invoice_creating',       'icon': 'fas fa-file-invoice-dollar', 'category': 'Core Accounting'},
    {'key': 'customer_receipt',   'label': 'Customer Receipt',     'url': '/customer_receipt',       'icon': 'fas fa-receipt',             'category': 'Core Accounting'},
    {'key': 'cash_payments',      'label': 'Cash Payments',        'url': '/cash_payment',           'icon': 'fas fa-money-bill',          'category': 'Core Accounting'},
    {'key': 'direct_purchase',    'label': 'Direct Purchase',      'url': '/direct_purchasing',      'icon': 'fas fa-hand-holding-usd',    'category': 'Core Accounting'},
    {'key': 'bank_payments',      'label': 'Bank Payments',        'url': '/bank_payment',           'icon': 'fas fa-university',          'category': 'Core Accounting'},
    {'key': 'bank_reconciliation','label': 'Bank Reconciliation',  'url': '/bank_reconciliation',    'icon': 'fas fa-check-double',        'category': 'Core Accounting'},
    {'key': 'cheque_print_setup', 'label': 'Cheque Print Setup',   'url': '/cheque_print_setup',     'icon': 'fas fa-print',               'category': 'Core Accounting'},
    {'key': 'journal_entry',      'label': 'Journal Entry',        'url': '/journal_entry',          'icon': 'fas fa-book',                'category': 'Core Accounting'},
    {'key': 'sub_allocation',     'label': 'Sub-Account Allocation','url': '/sub_account_allocation', 'icon': 'fas fa-tags',                'category': 'Core Accounting'},
    {'key': 'opening_balances',   'label': 'Opening Balances',     'url': '/opening_balances',       'icon': 'fas fa-scale-balanced',      'category': 'Core Accounting'},
    {'key': 'service_entry',      'label': 'Service Entry (SRN)',  'url': '/service_entry',          'icon': 'fas fa-file-invoice',        'category': 'Core Accounting'},
    {'key': 'credit_note',        'label': 'Credit Note',          'url': '/credit_note',            'icon': 'fas fa-file-medical',        'category': 'Core Accounting'},
    {'key': 'debit_note',         'label': 'Debit Note',           'url': '/debit_note',             'icon': 'fas fa-file-invoice-dollar', 'category': 'Core Accounting'},
    {'key': 'fixed_assets',       'label': 'Fixed Assets',         'url': '/fixed_assets',           'icon': 'fas fa-building',            'category': 'Core Accounting'},
    {'key': 'vat_schedule',       'label': 'VAT Schedule',         'url': '/vat_report',             'icon': 'fas fa-file-alt',            'category': 'Core Accounting'},
    # Reversals & Adjustments
    {'key': 'pos_reversal',       'label': 'POS Reversal',         'url': '/pos_reversal',           'icon': 'fas fa-undo',                'category': 'Reversals & Adjustments'},
    {'key': 'cash_pay_reversal',  'label': 'Cash Pay Reversal',    'url': '/cash_payment_reversal',  'icon': 'fas fa-undo-alt',            'category': 'Reversals & Adjustments'},
    {'key': 'bank_pay_reversal',  'label': 'Bank Pay Reversal',    'url': '/bank_payment_reversal',  'icon': 'fas fa-history',             'category': 'Reversals & Adjustments'},
    {'key': 'direct_pay_reversal','label': 'Direct Pay Reversal',  'url': '/direct_payment_reversal','icon': 'fas fa-sync',                'category': 'Reversals & Adjustments'},
    {'key': 'srn_reversal',       'label': 'SRN Reversal',         'url': '/service_entry_reversal', 'icon': 'fas fa-file-invoice',        'category': 'Reversals & Adjustments'},
    {'key': 'grn_reversal',       'label': 'GRN Reversal',         'url': '/grn_reversal',           'icon': 'fas fa-truck-loading',       'category': 'Reversals & Adjustments'},
    {'key': 'reversal_category',  'label': 'Reversal Category',    'url': '/reversal_category',      'icon': 'fas fa-tags',                'category': 'Reversals & Adjustments'},
    {'key': 'postdated_cheques',  'label': 'Postdated Cheques',    'url': '/postdated_cheques',      'icon': 'fas fa-calendar-check',      'category': 'Reversals & Adjustments'},
    # Inventory
    {'key': 'inventory_balance',  'label': 'Inventory Balance',    'url': '/inventory_balance',      'icon': 'fas fa-boxes',               'category': 'Inventory'},
    {'key': 'new_inventory_item', 'label': 'New Inventory Item',   'url': '/add_inventory_item',     'icon': 'fas fa-plus-square',         'category': 'Inventory'},
    {'key': 'grn',                'label': 'GRN',                  'url': '/grn',                    'icon': 'fas fa-truck-loading',       'category': 'Inventory'},
    {'key': 'po_generator',       'label': 'PO Generator',         'url': '/purchase_orders',        'icon': 'fas fa-file-invoice',        'category': 'Inventory'},
    {'key': 'quotation_eval',     'label': 'Quotation Eval',       'url': '/quotation_evaluation',   'icon': 'fas fa-balance-scale-right', 'category': 'Inventory'},
    {'key': 'proforma_invoice',   'label': 'Proforma Invoice',     'url': '/proforma_invoice',       'icon': 'fas fa-file-contract',       'category': 'Inventory'},
    {'key': 'inventory_transfer', 'label': 'Inventory Transfer',   'url': '/inventory_transfer',     'icon': 'fas fa-exchange-alt',        'category': 'Inventory'},
    {'key': 'inventory_issue',    'label': 'Inventory Issue',      'url': '/inventory_issue',        'icon': 'fas fa-dolly',               'category': 'Inventory'},
    {'key': 'manufacturing',      'label': 'Manufacturing',        'url': '/inventory_production',   'icon': 'fas fa-industry',            'category': 'Inventory'},
    {'key': 'trend_analysis',     'label': 'Trend Analysis',       'url': '/inventory_trend_analysis','icon': 'fas fa-chart-line',         'category': 'Inventory'},
    # POS
    {'key': 'pos_system',         'label': 'POS System',           'url': '/pos',                    'icon': 'fas fa-cash-register',       'category': 'POS'},
    # Restaurant / Food & Beverage
    {'key': 'food_costing',       'label': 'Food Costing',         'url': '/food_costing',           'icon': 'fas fa-utensils',            'category': 'Restaurant'},
    # HR & Payroll
    {'key': 'employees',          'label': 'Employees',            'url': '/employees',              'icon': 'fas fa-users',               'category': 'HR & Payroll'},
    {'key': 'leave_applications', 'label': 'Leave Applications',   'url': '/leave_application',      'icon': 'fas fa-calendar-check',      'category': 'HR & Payroll'},
    {'key': 'leave_approvals',    'label': 'Leave Approvals',      'url': '/leave_approvals',        'icon': 'fas fa-check-circle',        'category': 'HR & Payroll'},
    {'key': 'leave_types',        'label': 'Leave Types',          'url': '/leave_types',            'icon': 'fas fa-calendar-alt',        'category': 'HR & Payroll'},
    {'key': 'payroll_run',        'label': 'Payroll Run',          'url': '/payroll_run',            'icon': 'fas fa-money-check-alt',     'category': 'HR & Payroll'},
    {'key': 'salary_components',  'label': 'Salary Components',    'url': '/payroll_components',     'icon': 'fas fa-sliders-h',           'category': 'HR & Payroll'},
    # CRM & Sales
    {'key': 'sales_pipeline',     'label': 'Sales Pipeline',       'url': '/crm',                    'icon': 'fas fa-funnel-dollar',       'category': 'CRM & Sales'},
    {'key': 'add_lead',           'label': 'Add Lead',             'url': '/crm/lead/add',           'icon': 'fas fa-user-plus',           'category': 'CRM & Sales'},
    # Reports
    {'key': 'transaction_search', 'label': 'Transaction Search',    'url': '/transaction_search',     'icon': 'fas fa-search-dollar',       'category': 'Reports'},
    {'key': 'ledger_view',        'label': 'Ledger View',          'url': '/ledger_view',            'icon': 'fas fa-book-open',           'category': 'Reports'},
    {'key': 'trial_balance',      'label': 'Trial Balance',        'url': '/trial_balance',          'icon': 'fas fa-balance-scale-left',  'category': 'Reports'},
    {'key': 'bank_rec_history',   'label': 'Bank Rec. History',    'url': '/bank_reconciliation/history', 'icon': 'fas fa-clipboard-check',  'category': 'Reports'},
    {'key': 'supplier_aging',     'label': 'Supplier Aging',       'url': '/supplier_aging',         'icon': 'fas fa-history',             'category': 'Reports'},
    {'key': 'supplier_payments',  'label': 'Supplier Payments',    'url': '/supplier_payments_report','icon': 'fas fa-money-check-dollar', 'category': 'Reports'},
    {'key': 'grn_payment_method', 'label': 'Supplier Payment Method',   'url': '/grn_payment_method_report','icon': 'fas fa-money-check',       'category': 'Reports'},
    {'key': 'supplier_purchasing_report', 'label': 'Supplier Purchasing Report', 'url': '/supplier_purchasing_report', 'icon': 'fas fa-truck-loading', 'category': 'Reports'},
    {'key': 'grn_payment_ready',  'label': 'Payment Ready List',   'url': '/grn_payment_ready_list',   'icon': 'fas fa-check-circle',       'category': 'Reports'},
    {'key': 'customer_aging',     'label': 'Customer Aging',       'url': '/customer_aging',         'icon': 'fas fa-user-clock',          'category': 'Reports'},
    {'key': 'balance_sheet',      'label': 'Balance Sheet',        'url': '/balance_sheet',          'icon': 'fas fa-balance-scale',       'category': 'Reports'},
    {'key': 'custom_balance_sheet','label': 'Custom Balance Sheet','url': '/balance_sheet_custom',   'icon': 'fas fa-balance-scale',       'category': 'Reports'},
    {'key': 'profit_loss',        'label': 'Profit & Loss',        'url': '/profit_loss',            'icon': 'fas fa-chart-pie',           'category': 'Reports'},
    {'key': 'custom_profit_loss', 'label': 'Custom Profit & Loss', 'url': '/profit_loss_custom',     'icon': 'fas fa-chart-pie',           'category': 'Reports'},
    {'key': 'cash_flow',          'label': 'Cash Flow Statement',  'url': '/cash_flow',              'icon': 'fas fa-water',               'category': 'Reports'},
    {'key': 'job_profit_analysis','label': 'Job Profit Analysis',  'url': '/job_profit_analysis',    'icon': 'fas fa-briefcase',           'category': 'Reports'},
    # Settings
    {'key': 'email_settings',     'label': 'Email Settings',       'url': '/email_settings',         'icon': 'fas fa-envelope-open-text',  'category': 'Settings'},
    {'key': 'period_close',       'label': 'Financial Period Close','url': '/financial_period_close', 'icon': 'fas fa-lock',                'category': 'Settings'},
]

# Reverse lookup used to enforce menu control at the route level.
# Sorted longest-URL first so a more specific path (e.g. /bank_reconciliation/history)
# resolves to its own key instead of a shorter prefix (/bank_reconciliation).
MENU_URL_TO_KEY = sorted(
    ((item['url'], item['key']) for item in MENU_ITEMS_REGISTRY),
    key=lambda pair: len(pair[0]), reverse=True
)

# Database Configuration
# Credentials should be set in .env file or environment variables for security.
db_suport_name = "suwixvkn"

# Force prefix onto database name to handle shared hosting constraints
_raw_db_name = os.environ.get('DB_NAME', 'Book_keeping')
if _raw_db_name.startswith(f"{db_suport_name}_"):
    _final_db_name = _raw_db_name
else:
    _final_db_name = f"{db_suport_name}_{_raw_db_name}"

db_config = {
    'user': os.environ.get('DB_USER', 'root'),
    'password': os.environ.get('DB_PASSWORD', '21219125'),
    'host': os.environ.get('DB_HOST', 'localhost'),
    'database': _final_db_name,
    'raise_on_warnings': True
}


# Ensure critical database configuration is present
if not db_config['user']:
    print("Warning: DB_USER not set in environment variables.")

db = Database(db_config)
MASTER_DB_NAME = f"{db_suport_name}_Book_keeping_Master"

def get_session_db_name():
    """Returns the correct database name based on session."""
    # If standard user
    if 'db_name' in session:
        return session['db_name']
    return db_config['database']

db.set_db_name_getter(get_session_db_name)

# Master DB Connection (Dedicated)
master_db_config = db_config.copy()
master_db_config['database'] = MASTER_DB_NAME
master_db = Database(master_db_config)

def setup_master_db():
    """Ensure the Master DB and its tables exist."""
    try:
        # Connect without DB to create Master DB if needed
        temp_config = db_config.copy()
        if 'database' in temp_config:
            del temp_config['database']

        try:
            if not is_safe_db_name(MASTER_DB_NAME):
                raise ValueError(f"Invalid database name: {MASTER_DB_NAME}")
            conn = mysql.connector.connect(**temp_config)
            cursor = conn.cursor()
            if not is_safe_db_name(MASTER_DB_NAME):
                raise ValueError("Invalid database name")
            cursor.execute(f"CREATE DATABASE IF NOT EXISTS {MASTER_DB_NAME}")
            cursor.close()
            conn.close()
        except mysql.connector.Error as e:
            if e.errno in (1007, 1044, 1045):
                logging.warning(f"Ignored DB creation error {e.errno} for Master DB: {e.msg}")
            else:
                raise e

        # Now create tables in Master DB
        master_db.execute_query("""
            CREATE TABLE IF NOT EXISTS tenants (
                id INT AUTO_INCREMENT PRIMARY KEY,
                company_name VARCHAR(255) NOT NULL UNIQUE,
                db_name VARCHAR(255) NOT NULL UNIQUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                is_active TINYINT DEFAULT 1,
                gb_used DOUBLE DEFAULT 0,
                max_users INT DEFAULT 5,
                sidebar_enabled TINYINT DEFAULT 1,
                db_initialized TINYINT DEFAULT 1
            )
        """)

        # Handle existing deployments by adding the columns if they don't exist
        try:
            master_db.execute_query("ALTER TABLE tenants ADD COLUMN is_active TINYINT DEFAULT 1")
        except Exception:
            pass
        try:
            master_db.execute_query("ALTER TABLE tenants ADD COLUMN gb_used DOUBLE DEFAULT 0")
        except Exception:
            pass
        try:
            master_db.execute_query("ALTER TABLE tenants ADD COLUMN max_users INT DEFAULT 5")
        except Exception:
            pass
        try:
            master_db.execute_query("ALTER TABLE tenants ADD COLUMN sidebar_enabled TINYINT DEFAULT 1")
        except Exception:
            pass
        try:
            master_db.execute_query("ALTER TABLE tenants ADD COLUMN db_initialized TINYINT DEFAULT 1")
        except Exception:
            pass
        try:
            master_db.execute_query("ALTER TABLE tenants ADD COLUMN menu_config TEXT DEFAULT NULL")
        except Exception:
            pass

        # Custom menu items added by superadmin per tenant
        master_db.execute_query("""
            CREATE TABLE IF NOT EXISTS tenant_custom_menu (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                tenant_id   INT          NOT NULL,
                item_label  VARCHAR(200) NOT NULL,
                item_url    VARCHAR(500) NOT NULL,
                item_icon   VARCHAR(100) DEFAULT 'fas fa-circle',
                item_category VARCHAR(100) DEFAULT 'General',
                sort_order  INT          DEFAULT 99,
                is_enabled  TINYINT(1)   DEFAULT 1,
                created_at  DATETIME     DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (tenant_id) REFERENCES tenants(id) ON DELETE CASCADE
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)

        master_db.execute_query("""
            CREATE TABLE IF NOT EXISTS users (
                id INT AUTO_INCREMENT PRIMARY KEY,
                username VARCHAR(255) NOT NULL UNIQUE,
                password VARCHAR(255) NOT NULL,
                email VARCHAR(255),
                tenant_id INT,
                FOREIGN KEY (tenant_id) REFERENCES tenants(id) ON DELETE CASCADE
            )
        """)

        # Add mobile column to users if it doesn't exist
        try:
            master_db.execute_query("ALTER TABLE users ADD COLUMN mobile VARCHAR(50)")
        except Exception:
            pass

        print("Master DB setup complete.")

        # 3. Setup Default/Fallback Database if missing tables (e.g. Login_Table)
        default_db_name = db_config['database']

        # Check if default DB has Login_Table
        try:
            if not is_safe_db_name(default_db_name):
                raise ValueError(f"Invalid database name: {default_db_name}")
            try:
                default_conn = mysql.connector.connect(**temp_config)
                default_cursor = default_conn.cursor()
                if not is_safe_db_name(default_db_name):
                    raise ValueError("Invalid database name")
                default_cursor.execute(f"CREATE DATABASE IF NOT EXISTS `{default_db_name}`")
                default_cursor.close()
                default_conn.close()
            except mysql.connector.Error as e:
                if e.errno in (1007, 1044, 1045):
                    logging.warning(f"Ignored DB creation error {e.errno} for default DB: {e.msg}")
                else:
                    raise e

            # Connect to default DB to check for tables
            check_config = db_config.copy()
            check_conn = mysql.connector.connect(**check_config)
            check_cursor = check_conn.cursor()

            # Try selecting from Login_Table to see if it exists
            table_exists = False
            try:
                check_cursor.execute("SELECT 1 FROM Login_Table LIMIT 1")
                check_cursor.fetchall()
                table_exists = True
            except mysql.connector.errors.ProgrammingError:
                pass

            if not table_exists:
                import re
                print(f"Initializing schema for default database: {default_db_name}")
                if os.path.exists('database_schema.sql'):
                    with open('database_schema.sql', 'r') as f:
                        content = re.sub(r'(?i)Book_keeping', default_db_name, f.read())
                        parse_and_execute_sql(check_cursor, content)

                if os.path.exists('fixed_assets.sql'):
                    with open('fixed_assets.sql', 'r') as f:
                        content = re.sub(r'(?i)Book_keeping', default_db_name, f.read())
                        parse_and_execute_sql(check_cursor, content)

                check_conn.commit()

                # Run application-level migrations on this default database to ensure all dynamic columns are present
                run_schema_migrations(check_conn)

                # Insert default admin user into fallback db
                try:
                    check_cursor.execute("""
                        INSERT INTO Login_Table (User_Name, Password, Email, User_Code, User_Active)
                        VALUES ('admin', 'admin', 'admin@example.com', '1001', 1)
                    """)
                    check_conn.commit()
                    user_id = check_cursor.lastrowid

                    check_cursor.execute("""
                        INSERT INTO User_Rights (
                            Link_To_Loging_Tabke, Add_New_User, OP_Approved, Access_Inventory,
                            Access_POS, Access_Accounting, Access_Reports, Access_Reversals
                        )
                        VALUES (%s, 1, 1, 1, 1, 1, 1, 1)
                    """, (user_id,))
                    check_conn.commit()
                except Exception as e:
                    print(f"Could not insert default admin: {e}")

            check_cursor.close()
            check_conn.close()

        except Exception as default_db_err:
            print(f"Error initializing default DB ({default_db_name}): {default_db_err}")

    except Exception as e:
        if getattr(e, "errno", None) not in (1050, 1007, 1060, 1061, 1146, 1054, 1304, 1305):
            print(f"Error setting up Master DB: {e}")

def parse_and_execute_sql(cursor, content):
    """Parses SQL content with DELIMITER support and executes it."""
    delimiter = ';'
    statement = ""

    lines = content.split('\n')
    for line in lines:
        stripped = line.strip()

        if stripped.upper().startswith('DELIMITER '):
            delimiter = stripped.split()[1]
            continue

        if stripped.startswith('--') or stripped.startswith('#'):
            continue

        if not statement and not stripped:
            continue

        statement += line + "\n"

        if statement.strip().endswith(delimiter):
            sql_to_run = statement.strip()
            if sql_to_run.endswith(delimiter):
                 sql_to_run = sql_to_run[:-len(delimiter)]

            if sql_to_run.strip():
                try:
                    cursor.execute(sql_to_run)
                    while cursor.nextset(): pass
                except mysql.connector.Error as e:
                    # If this is a CREATE SCHEMA/DATABASE statement and we hit 1007/1044, safely ignore
                    upper_sql = sql_to_run.upper()

                    if e.errno in (1007, 1044, 1050, 1305, 1227, 1060, 1061, 1146, 1054):
                        pass # Ignore completely

                    else:
                        print(f"SQL Error: {e} | Statement: {sql_to_run[:50]}...")
                        raise e
                except Exception as e:
                    print(f"SQL Error: {e} | Statement: {sql_to_run[:50]}...")
                    raise e
            statement = ""

def create_tenant_db(company_name, username, password, email, mobile=None):
    """Creates a new tenant DB, runs schema, and registers in Master DB."""
    import re

    safe_name = re.sub(r'[^a-z0-9]', '_', company_name.lower())
    db_name = f"{db_suport_name}_{safe_name}"

    master_conn = master_db.get_connection()
    if not master_conn:
        error_details = master_db.last_error if master_db.last_error else "Unknown error."
        return False, f"System Error: Master database '{MASTER_DB_NAME}' not found or accessible. Details: {error_details}. Please ensure it is created in cPanel and the DB_PASSWORD in .env is correct."
    else:
        master_conn.close()

    existing_user = master_db.execute_query("SELECT id FROM users WHERE username = %s", (username,))
    if existing_user: return False, "Username already exists."

    existing_tenant = master_db.execute_query("SELECT id FROM tenants WHERE company_name = %s", (company_name,))
    if existing_tenant: return False, "Company already registered."

    try:
        # Insert into Master - mark db_initialized = 0 for manual cPanel creation
        tenant_id = master_db.execute_query(
            "INSERT INTO tenants (company_name, db_name, db_initialized) VALUES (%s, %s, 0)",
            (company_name, db_name),
            commit=True
        )

        if tenant_id is None:
            return False, "Database error: Could not insert tenant. Ensure Master DB is created."

        master_db.execute_query("""
            INSERT INTO users (username, password, email, tenant_id, mobile)
            VALUES (%s, %s, %s, %s, %s)
        """, (username, password, email, tenant_id, mobile), commit=True)

        # Notify the superadmin by SMS so a new account request is never missed.
        # Failure to send must not break the registration itself.
        try:
            superadmin_phone = os.getenv('SUPERADMIN_PHONE', '0700000000')
            notify_msg = (
                f"New account request: {company_name} "
                f"(user: {username}, mobile: {mobile or '-'}, email: {email or '-'}). "
                f"Please review and set up the database."
            )
            send_sms(superadmin_phone, notify_msg)
        except Exception as e:
            logging.error(f"Failed to send new-account SMS notification: {e}")

        return True, "Registration successful. Pending database setup by administrator."

    except Exception as e:
        logging.error(f"Registration Error occurred: {e}", exc_info=True)
        return False, f"An error occurred during registration: {e}"

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        company_name = request.form['company_name']
        username = request.form['username']
        password = request.form['password']
        email = request.form['email']
        mobile = request.form.get('mobile')

        success, message = create_tenant_db(company_name, username, password, email, mobile)
        if success:
            flash(message, 'success')
            return redirect(url_for('login'))
        else:
            flash(message, 'danger')

    return render_template('register.html')

# Context Processor for Currency & Theme
@app.context_processor
def inject_globals():
    globals_dict = {}

    # Currency
    try:
        res = db.execute_query("SELECT company_curency FROM company LIMIT 1")
        globals_dict['company_currency'] = res[0]['company_curency'] if res and res[0]['company_curency'] else 'LKR'
    except Exception:
        globals_dict['company_currency'] = 'LKR'

    # Theme
    try:
        res = db.execute_query("SELECT setting_value FROM system_settings WHERE setting_key = 'system_theme'")
        theme_key = res[0]['setting_value'] if res else 'default'
        globals_dict['current_theme'] = THEMES.get(theme_key, THEMES['default'])
        globals_dict['theme_key'] = theme_key
    except Exception:
        globals_dict['current_theme'] = THEMES['default']
        globals_dict['theme_key'] = 'default'

    # Sidebar / Function visibility + per-item menu control (Tenant specific from Master DB)
    try:
        # tenant_id is not always stored in the session, but db_name always is —
        # resolve the tenant by whichever identifier is available.
        if session.get('tenant_id'):
            tenant_res = master_db.execute_query(
                "SELECT sidebar_enabled, menu_config FROM tenants WHERE id = %s",
                (session['tenant_id'],)
            )
        elif session.get('db_name'):
            tenant_res = master_db.execute_query(
                "SELECT sidebar_enabled, menu_config FROM tenants WHERE db_name = %s",
                (session['db_name'],)
            )
        else:
            tenant_res = None

        if tenant_res is not None:
            if tenant_res:
                t = tenant_res[0]
                globals_dict['sidebar_enabled'] = t.get('sidebar_enabled', 1) == 1
                # Parse disabled menu keys from JSON config
                import json as _json
                raw_cfg = t.get('menu_config') or '{}'
                try:
                    cfg = _json.loads(raw_cfg)
                except Exception:
                    cfg = {}
                # cfg = {key: True/False} — False means disabled
                globals_dict['menu_disabled'] = {k for k, v in cfg.items() if not v}
            else:
                globals_dict['sidebar_enabled'] = True
                globals_dict['menu_disabled'] = set()
        else:
            globals_dict['sidebar_enabled'] = True
            globals_dict['menu_disabled'] = set()
    except Exception:
        globals_dict['sidebar_enabled'] = True
        globals_dict['menu_disabled'] = set()

    # Custom menu items added by superadmin for this tenant
    try:
        if 'tenant_id' in session:
            custom = master_db.execute_query(
                "SELECT item_label, item_url, item_icon, item_category FROM tenant_custom_menu "
                "WHERE tenant_id = %s AND is_enabled = 1 ORDER BY item_category, sort_order, id",
                (session['tenant_id'],)
            )
            globals_dict['custom_menu_items'] = custom or []
        else:
            globals_dict['custom_menu_items'] = []
    except Exception:
        globals_dict['custom_menu_items'] = []

    globals_dict['check_permission'] = check_permission

    # Financial period lock date — lets templates disable edit/reverse buttons
    try:
        lock = get_period_lock_date()
        globals_dict['period_lock_date'] = lock.strftime('%Y-%m-%d') if lock else None
    except Exception:
        globals_dict['period_lock_date'] = None

    return globals_dict


@app.before_request
def enforce_menu_disabled():
    """Block a route when the superadmin has disabled that menu item for the
    tenant. Hiding the sidebar link alone left the page reachable by direct URL;
    this closes that gap so a disabled feature is actually disabled."""
    # Only standard logged-in tenant users are subject to menu control.
    # tenant_id is not always in the session, but db_name always is.
    if 'user_id' not in session:
        return
    if session.get('superadmin_logged_in'):
        return
    if not session.get('tenant_id') and not session.get('db_name'):
        return

    path = (request.path or '/').rstrip('/') or '/'
    matched_key = None
    for url, key in MENU_URL_TO_KEY:
        if path == url or path.startswith(url + '/'):
            matched_key = key
            break
    if matched_key is None:
        return  # not a menu-controlled route — no DB lookup needed

    try:
        if session.get('tenant_id'):
            row = master_db.execute_query(
                "SELECT menu_config FROM tenants WHERE id = %s", (session['tenant_id'],)
            )
        else:
            row = master_db.execute_query(
                "SELECT menu_config FROM tenants WHERE db_name = %s", (session['db_name'],)
            )
    except Exception:
        return  # fail open — never lock users out over a DB hiccup
    if not row:
        return

    import json as _json
    try:
        cfg = _json.loads(row[0].get('menu_config') or '{}')
    except Exception:
        cfg = {}

    if cfg.get(matched_key) is False:
        msg = 'This feature has been disabled by your administrator.'
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'error': msg}), 403
        flash(msg, 'warning')
        return redirect(url_for('index'))


# Custom Filter for Currency Formatting
@app.template_filter('currency')
@pass_context
def currency_filter(context, value, symbol=True):
    try:
        if value is None:
            value = 0

        # Format: 1,234.56
        formatted = "{:,.2f}".format(float(value))

        if not symbol:
            return formatted

        # Get symbol from context
        curr_symbol = context.get('company_currency', '')
        if curr_symbol:
            return f"{curr_symbol} {formatted}"
        return formatted
    except (ValueError, TypeError):
        return "0.00"

@app.template_filter('amount_in_words')
def amount_in_words(amount):
    """Converts a numeric amount to words (Rupees and Cents)."""
    try:
        amount = float(amount)
        if amount == 0:
            return "Zero Rupees Only"

        # num2words supports USD which formats as "dollars" and "cents"
        # We will use that and replace the currency names
        words = num2words(amount, to='currency', currency='USD', lang='en')

        # Replace currency names with local context (LKR)
        # Handle singular/plural variations just in case, though usually 'dollars' covers it
        # num2words output example: "one hundred dollars, fifty cents"

        words = words.replace('dollars', 'Rupees')
        words = words.replace('dollar', 'Rupee')
        words = words.replace('cents', 'Cents')
        words = words.replace('cent', 'Cent')

        # Capitalize for Cheque format (Title Case looks better)
        return words.title()
    except Exception as e:
        return f"Error converting amount: {e}"

def parse_float(value):
    """Safely parses a string or number into a float, handling commas and None."""
    try:
        if value is None:
            return 0.0
        if isinstance(value, str):
            value = value.replace(',', '').strip()
            if not value: return 0.0
        return float(value)
    except (ValueError, TypeError):
        return 0.0

def parse_csv_file(file_storage, required_columns=None):
    """
    Parses a file-like object (e.g., Flask FileStorage) into a list of dictionaries.
    Handles multiple encodings and optional column validation.

    Args:
        file_storage: A Flask FileStorage object or similar with a .stream attribute or read() method.
        required_columns: A list of strings representing column headers that must be present.

    Returns:
        A list of dictionaries representing the CSV rows.

    Raises:
        ValueError: If decoding fails or required columns are missing.
    """
    try:
        # Read file bytes. If it's a FileStorage, use .stream.read() or .read()
        if hasattr(file_storage, 'stream'):
            file_bytes = file_storage.stream.read()
        elif hasattr(file_storage, 'read'):
            file_bytes = file_storage.read()
        else:
            raise ValueError("Invalid file object")

        # If file_bytes is empty, decoded_str will be empty, and csv.DictReader might not return headers
        # We need to handle this.
        if not file_bytes:
             if required_columns:
                 raise ValueError("File is empty, missing required columns: " + ", ".join(required_columns))
             return []

        decoded_str = None
        # Try multiple encodings
        for encoding in ['utf-8-sig', 'utf-16', 'utf-8', 'cp1252', 'latin1']:
            try:
                decoded_str = file_bytes.decode(encoding)
                break
            except UnicodeDecodeError:
                continue

        if decoded_str is None:
            raise ValueError("Unable to determine file encoding (tried utf-8, latin1, cp1252, utf-16)")

        stream = io.StringIO(decoded_str, newline=None)
        csv_input = csv.DictReader(stream)

        # Validate Headers
        if required_columns:
            # csv.DictReader reads headers on access to fieldnames or iteration
            headers = csv_input.fieldnames
            if not headers:
                 # Should have been caught by empty check unless file is only newlines?
                 raise ValueError("Missing required columns: " + ", ".join(required_columns))

            # Use strip() on headers for comparison?
            clean_headers = [h.strip() for h in headers if h]
            missing = [col for col in required_columns if col not in clean_headers]
            if missing:
                raise ValueError(f"Missing required columns: {', '.join(missing)}")

        rows = []
        for row in csv_input:
            # Clean keys/values
            clean_row = {k.strip(): (v.strip() if v else '') for k, v in row.items() if k}
            if not clean_row: continue # Skip empty rows

            rows.append(clean_row)

        return rows

    except Exception as e:
        # Re-raise ValueError directly, or wrap other exceptions
        if isinstance(e, ValueError): raise e
        raise ValueError(f"CSV Parsing Error: {str(e)}")

def login_required(f):
    return f

    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def pos_login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # A user can access POS endpoints if they are logged in via the dedicated POS login
        if session.get('pos_logged_in'):
            return f(*args, **kwargs)

        # Or if they are a standard web ERP user who has the 'Access_POS' permission
        if 'user_id' in session and check_permission('Access_POS'):
            return f(*args, **kwargs)

        # Otherwise, redirect to the specific POS login
        return redirect(url_for('pos_login'))
    return decorated_function

def get_current_user_id():
    return session.get('user_id', 0)

def get_current_user_pk():
    try:
        pk = session.get('user_pk', 0)
        return int(pk)
    except (ValueError, TypeError):
        return 0

def check_permission(perm_name):
    """Checks if current user has specific permission."""
    user_pk = session.get('user_pk')
    if not user_pk: return False

    try:
        # Fetch all rights for the user to handle potentially missing columns gracefully
        # This prevents SQL errors if schema isn't fully migrated or perm_name is invalid
        query = "SELECT * FROM User_Rights WHERE Link_To_Loging_Tabke = %s"
        res = db.execute_query(query, (user_pk,))

        # Check if permission exists in the row and is enabled
        if res and res[0].get(perm_name) == 1:
            return True

    except Exception as e:
        logging.error("Permission check error.")
        return False
    return False

def has_permission(perm):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('login'))

            if not check_permission(perm):
                flash(f'Access Denied: Required permission {perm}', 'danger')
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        if not username or not password:
            flash('Please enter both username and password.', 'danger')
            return redirect(url_for('login'))

        # Check credentials
        # 1. Try Login via Master DB (Multi-Tenant)
        try:
            # Query Master DB for user and tenant DB
            master_user_res = master_db.execute_query("""
                SELECT u.username, u.password, t.db_name, t.is_active, t.db_initialized
                FROM users u
                JOIN tenants t ON u.tenant_id = t.id
                WHERE u.username = %s
            """, (username,))

            if master_user_res:
                master_user = master_user_res[0]

                # Check Tenant Active Status
                if master_user['is_active'] == 0:
                    return redirect(url_for('payment_due'))

                # Removed db_initialized check to unblock users

                # Check password (handle both hashed and plaintext for migration)
                is_valid = False
                stored_pwd = master_user['password']
                if stored_pwd.startswith('scrypt:') or stored_pwd.startswith('pbkdf2:'):
                    from werkzeug.security import check_password_hash
                    if check_password_hash(stored_pwd, password):
                        is_valid = True
                elif stored_pwd == password:
                    is_valid = True

                if is_valid:
                    # Login Successful on Master
                    session['db_name'] = master_user['db_name']
                    session['username'] = username

                    # Fetch User Details from Tenant DB (for permissions/FKs)
                    # Note: db instance now points to tenant_db via session
                    tenant_user_res = db.execute_query("SELECT id, User_Code FROM Login_Table WHERE User_Name = %s", (username,))

                    if tenant_user_res:
                        tenant_user = tenant_user_res[0]
                        session['user_id'] = tenant_user['User_Code']
                        session['user_pk'] = tenant_user['id']
                        run_schema_migrations()
                        return redirect(url_for('index'))
                    else:
                        flash('User record missing in tenant database.', 'danger')
                        session.pop('db_name', None)
                        return redirect(url_for('login'))
                else:
                    flash('Incorrect password.', 'danger')
                    return redirect(url_for('login'))
        except Exception as e:
            logging.error(f"Master Login Error occurred: {e}", exc_info=True)
            # Fallthrough to legacy

        # 2. Fallback to Legacy Login (Default DB)
        # Ensure clean session regarding tenant
        session.pop('db_name', None)

        query = "SELECT id, User_Code, Password FROM Login_Table WHERE User_Name = %s"
        users = db.execute_query(query, (username,))

        if users is None:
            error_msg = f"Database connection failed: {db.last_error}" if hasattr(db, 'last_error') and db.last_error else "Database connection failed."
            flash(error_msg, 'danger')
        elif users:
            user = users[0]
            stored_password = user['Password']

            verified = False
            migrated = False

            # 1. Try Hash Verification
            from werkzeug.security import check_password_hash, generate_password_hash
            try:
                if stored_password.startswith('scrypt:') or stored_password.startswith('pbkdf2:'):
                    if check_password_hash(stored_password, password):
                        verified = True
                else:
                    raise ValueError("Not a valid hash")
            except Exception:
                # stored_password might not be a valid hash format (e.g. plain text)
                pass

            # 2. Fallback to Plain Text (Legacy Support & Migration)
            if not verified:
                if stored_password == password:
                    verified = True
                    # Upgrade to Hash
                    try:
                        new_hash = generate_password_hash(password)
                        db.execute_query("UPDATE Login_Table SET Password = %s WHERE id = %s", (new_hash, user['id']), commit=True)
                        migrated = True
                    except Exception as e:
                        logging.error(f"Error migrating password for user: {e}")

            if verified:
                session['user_id'] = user['User_Code']
                session['user_pk'] = user['id']
                session['username'] = username
                run_schema_migrations()
                if migrated:
                    flash('Login successful. Your password security has been upgraded.', 'success')
                return redirect(url_for('index'))
            else:
                flash('Incorrect password.', 'danger')
        else:
            flash('User not found.', 'danger')

        return redirect(url_for('login'))

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
def index():
    if 'user_id' not in session:
        return render_template('landing.html')

    # Check if critical migration table exists, if not, force install page
    # In production, use a more robust check (e.g. system_settings table)
    try:
        conn = db.get_connection()
        if conn:
            cursor = conn.cursor()
            cursor.execute("SHOW TABLES LIKE 'migrations'")
            if not cursor.fetchone():
                return redirect(url_for('installing'))
    except Exception as e:
        logging.error(f"Error checking for migrations table: {e}")

    # Fetch VAT summary for current month for the dashboard
    vat_summary = None
    try:
        from datetime import datetime, timedelta
        import calendar
        now = datetime.now()
        first_day = now.replace(day=1).strftime('%Y-%m-%d')
        last_day = now.replace(day=calendar.monthrange(now.year, now.month)[1]).strftime('%Y-%m-%d')

        from vat_helper import VATReportGenerator
        generator = VATReportGenerator(db, first_day, last_day)
        if generator.check_vat_registered():
            vat_data = generator.generate()
            vat_summary = {
                'total_output': vat_data['summary']['total_output_vat'],
                'total_input': vat_data['summary']['total_input_vat'],
                'net_vat': vat_data['summary']['net_vat'],
                'gl_balance': vat_data['reconciliation']['gl_balance'],
                'difference': vat_data['reconciliation']['difference'],
                'month': now.strftime('%B %Y')
            }
    except Exception as e:
        logging.error(f"Error fetching VAT summary for dashboard: {e}")

    return render_template('index.html', vat_summary=vat_summary)

@app.route('/installing')
def installing():
    return render_template('installing.html')

@app.route('/install_stream')
def install_stream():
    def generate():
        yield f"data: {json.dumps({'message': 'Starting schema migration...', 'progress': 10})}\n\n"

        # We invoke run_schema_migrations logic step-by-step or call it and capture output
        # For simplicity, we assume run_schema_migrations is modified to yield or we check state.
        # But run_schema_migrations is synchronous. We can't yield from it easily without refactoring it into a generator.
        # Let's verify DB connection first.

        try:
            conn = db.get_connection()
            if not conn:
                yield f"data: {json.dumps({'message': 'Database connection failed!', 'status': 'error', 'progress': 100, 'done': True})}\n\n"
                return

            yield f"data: {json.dumps({'message': 'Database connected.', 'progress': 20})}\n\n"

            # Execute Migrations (Sync call, but fast enough for this demo or we refactor)
            # Actually, `run_schema_migrations` prints to console.
            # We will call it here.

            try:
                run_schema_migrations()
                yield f"data: {json.dumps({'message': 'Schema migrations applied successfully.', 'status': 'success', 'progress': 60})}\n\n"
            except Exception as e:
                yield f"data: {json.dumps({'message': f'Migration Error: {str(e)}', 'status': 'error', 'progress': 100, 'done': True})}\n\n"
                return

            # Default User
            create_default_user()
            yield f"data: {json.dumps({'message': 'Default user checked/created.', 'progress': 80})}\n\n"

            # Default Accounts
            ensure_default_accounts()
            yield f"data: {json.dumps({'message': 'Default accounts verified.', 'progress': 90})}\n\n"

            yield f"data: {json.dumps({'message': 'Installation complete!', 'status': 'success', 'progress': 100, 'done': True})}\n\n"

        except Exception as e:
            yield f"data: {json.dumps({'message': f'Critical Error: {str(e)}', 'status': 'error', 'progress': 100, 'done': True})}\n\n"

    return Response(stream_with_context(generate()), mimetype='text/event-stream')

@app.route('/placeholder')
@login_required
def placeholder():
    title = request.args.get('title', 'Feature')
    return render_template('placeholder.html', title=title)

# --- Add Customer (Existing) ---
@app.route('/add_customer', methods=['GET', 'POST'])
@login_required
def add_customer():
    if request.method == 'POST':
        try:
            customer_id = request.form.get('customer_id')
            supplier_name = request.form.get('supplier_name')
            salutation = request.form.get('salutation')
            supplier_code = request.form.get('supplier_code')
            credit_limit = request.form.get('credit_limit', 0)
            vat_no = request.form.get('vat_no')

            address_no = request.form.get('address_no')
            address_line_1 = request.form.get('address_line_1')
            address_line_2 = request.form.get('address_line_2')
            address_line_3 = request.form.get('address_line_3')
            address_line_4 = request.form.get('address_line_4')

            contact_1 = request.form.get('contact_1')
            contact_2 = request.form.get('contact_2')
            email = request.form.get('email')

            if not supplier_name or not supplier_code:
                flash('Supplier Name and Code are required.', 'danger')
                return redirect(url_for('add_customer'))

            current_user = get_current_user_id()
            current_user_pk = get_current_user_pk()
            current_date = datetime.now().date()

            if customer_id:
                # Update existing customer
                query_supplier = """
                    UPDATE suppliers SET
                        supplier_name=%s, supplier_code=%s,
                        supplier_address_1=%s, supplier_address_2=%s, supplier_address_3=%s, supplier_address_4=%s,
                        suppliers_credit_fasility=%s, suppliers_teli_1=%s, suppliers_teli_2=%s,
                        suppliers_last_edit_user=%s, suppliers_last_edit_date=%s,
                        suppliers_e_mail=%s, suppliers_vat_regidter_no=%s, suppliers_salution=%s
                    WHERE sup_id=%s AND Is_Customer=1
                """
                params_supplier = (
                    supplier_name, supplier_code,
                    address_no, address_line_1, address_line_2, address_line_3,
                    parse_float(credit_limit), contact_1, contact_2,
                    current_user_pk, current_date,
                    email, vat_no, salutation,
                    customer_id
                )
            else:
                # Insert new customer
                query_supplier = """
                    INSERT INTO suppliers (
                        sup_id, supplier_name, supplier_code,
                        supplier_address_1, supplier_address_2, supplier_address_3, supplier_address_4,
                        suppliers_credit_fasility, suppliers_teli_1, suppliers_teli_2,
                        supplier_create_date, suppliers_create_user,
                        suppliers_last_edit_user, suppliers_last_edit_date,
                        suppliers_e_mail, suppliers_vat_regidter_no, suppliers_salution,
                        Is_Suplier, Is_Customer
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                params_supplier = (
                    0, supplier_name, supplier_code,
                    address_no, address_line_1, address_line_2, address_line_3,
                    parse_float(credit_limit), contact_1, contact_2,
                    current_date, current_user_pk,
                    current_user_pk, current_date,
                    email, vat_no, salutation,
                    0, 1 # Is_Suplier=0, Is_Customer=1
                )

            try:
                with db.transaction_cursor() as cursor:
                    cursor.execute(query_supplier, params_supplier)

                    if not customer_id:
                        query_sub_account = """
                            INSERT INTO sub_accont_for_new_account (
                                id_sub, sub_sub_accaount_name, sub_new_account,
                                creat_user, creat_date, active, sub_account_code
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                        """
                        cursor.execute(query_sub_account, (
                            0, supplier_name, "Account Receivable",
                            current_user, current_date, 1, 0
                        ))
                        last_sub_id = cursor.lastrowid
                        new_sub_code = last_sub_id + 10001
                        cursor.execute(
                            "UPDATE sub_accont_for_new_account SET sub_account_code = %s WHERE id_sub = %s",
                            (new_sub_code, last_sub_id)
                        )
                if customer_id:
                    flash('Customer updated successfully!', 'success')
                else:
                    flash('Customer added successfully!', 'success')
            except Exception as e:
                print(f"Transaction failed: {e}")
                logging.error(f"Transaction failed: {e}")
                flash(f'Error adding/updating customer: {str(e)}', 'danger')

            return redirect(url_for('add_customer'))

        except Exception as e:
            flash(f'An unexpected error occurred: {str(e)}', 'danger')
            return redirect(url_for('add_customer'))

    salutations = []
    try:
        salutations_data = db.execute_query("SELECT salutation FROM suplier_suporting_1")
        salutations = [row['salutation'] for row in salutations_data]
    except Exception as e:
        logging.error(f"Error loading salutations: {e}")

    customers_list = db.execute_query("""
        SELECT sup_id as id, supplier_name, supplier_code, suppliers_teli_1, suppliers_teli_2,
               suppliers_credit_fasility, suppliers_vat_regidter_no, suppliers_TIN, suppliers_NIC, suppliers_e_mail
        FROM suppliers
        WHERE Is_Customer = 1
        ORDER BY sup_id DESC
    """)

    return render_template('add_customer.html', salutations=salutations, customers_list=customers_list)

# --- Add Supplier (New) ---

@app.route('/api/extract_vat_from_pdf', methods=['POST'])
@login_required
def extract_vat_from_pdf():
    if 'document' not in request.files:
        return jsonify({'success': False, 'message': 'No document uploaded'}), 400

    file = request.files['document']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No selected file'}), 400

    if not file.filename.lower().endswith('.pdf'):
        return jsonify({'success': False, 'message': 'Only PDF files are supported'}), 400

    try:
        # Read PDF content
        pdf_reader = PyPDF2.PdfReader(io.BytesIO(file.read()))
        text = ""
        for page in pdf_reader.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + " "

        # "AI" Regex to find VAT Numbers
        # Matches common VAT formats (e.g., VAT NO: 123456789, VAT: GB123456789, VAT REGISTRATION NO: ...)
        vat_pattern = r'(?i)VAT\s*(?:REGISTRATION\s*)?(?:NO\.|NO|NUMBER|#)?\s*[:\-\s]*([A-Z0-9]{5,15})'
        matches = re.findall(vat_pattern, text)

        if matches:
            # Filter out matches that are purely alphabetic (like "REGISTRATION" or "CERTIFICATE")
            valid_matches = [m.strip() for m in matches if any(char.isdigit() for char in m)]

            if valid_matches:
                # Return first distinct match that contains digits
                vat_no = valid_matches[0]
                return jsonify({'success': True, 'vat_no': vat_no, 'message': 'VAT extracted successfully'})

        return jsonify({'success': False, 'message': 'No VAT number found in the document'})

    except Exception as e:
        app.logger.error(f"Error extracting VAT: {e}")
        return jsonify({'success': False, 'message': 'Failed to process document'}), 500


@app.route('/add_supplier', methods=['GET', 'POST'])
@login_required
def add_supplier():
    if request.method == 'POST':
        try:
            supplier_id = request.form.get('supplier_id')
            supplier_name = request.form.get('supplier_name')
            salutation = request.form.get('salutation')
            supplier_code = request.form.get('supplier_code')
            credit_limit = request.form.get('credit_limit', 0)
            vat_no = request.form.get('vat_no')

            address_no = request.form.get('address_no')
            address_line_1 = request.form.get('address_line_1')
            address_line_2 = request.form.get('address_line_2')
            address_line_3 = request.form.get('address_line_3')
            address_line_4 = request.form.get('address_line_4')

            contact_1 = request.form.get('contact_1')
            contact_2 = request.form.get('contact_2')
            email = request.form.get('email')

            tin = request.form.get('tin_no')
            nic = request.form.get('nic_no')
            default_payment_method = (request.form.get('default_payment_method') or '').strip() or None

            if not supplier_name or not supplier_code:
                flash('Supplier Name and Code are required.', 'danger')
                return redirect(url_for('add_supplier'))

            current_user = get_current_user_id()
            current_user_pk = get_current_user_pk()
            current_date = datetime.now().date()

            if supplier_id:
                # Update existing supplier
                query_supplier = """
                    UPDATE suppliers SET
                        supplier_name=%s, supplier_code=%s,
                        supplier_address_1=%s, supplier_address_2=%s, supplier_address_3=%s, supplier_address_4=%s,
                        suppliers_credit_fasility=%s, suppliers_teli_1=%s, suppliers_teli_2=%s,
                        suppliers_last_edit_user=%s, suppliers_last_edit_date=%s,
                        suppliers_e_mail=%s, suppliers_vat_regidter_no=%s, suppliers_salution=%s,
                        suppliers_TIN=%s, suppliers_NIC=%s, suppliers_default_payment_method=%s
                    WHERE sup_id=%s AND Is_Suplier=1
                """
                params_supplier = (
                    supplier_name, supplier_code,
                    address_no, address_line_1, address_line_2, address_line_3,
                    parse_float(credit_limit), contact_1, contact_2,
                    current_user_pk, current_date,
                    email, vat_no, salutation,
                    tin, nic, default_payment_method,
                    supplier_id
                )
            else:
                # Insert new supplier
                query_supplier = """
                    INSERT INTO suppliers (
                        sup_id, supplier_name, supplier_code,
                        supplier_address_1, supplier_address_2, supplier_address_3, supplier_address_4,
                        suppliers_credit_fasility, suppliers_teli_1, suppliers_teli_2,
                        supplier_create_date, suppliers_create_user,
                        suppliers_last_edit_user, suppliers_last_edit_date,
                        suppliers_e_mail, suppliers_vat_regidter_no, suppliers_salution,
                        Is_Suplier, Is_Customer, suppliers_TIN, suppliers_NIC, suppliers_default_payment_method
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                params_supplier = (
                    0, supplier_name, supplier_code,
                    address_no, address_line_1, address_line_2, address_line_3,
                    parse_float(credit_limit), contact_1, contact_2,
                    current_date, current_user_pk,
                    current_user_pk, current_date,
                    email, vat_no, salutation,
                    1, 0, tin, nic, default_payment_method
                )

            try:
                with db.transaction_cursor() as cursor:
                    cursor.execute(query_supplier, params_supplier)

                    if not supplier_id:
                        query_sub_account = """
                            INSERT INTO sub_accont_for_new_account (
                                id_sub, sub_sub_accaount_name, sub_new_account,
                                creat_user, creat_date, active, sub_account_code
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                        """
                        cursor.execute(query_sub_account, (
                            0, supplier_name, "Account Payable",
                            current_user, current_date, 1, 0
                        ))
                        last_sub_id = cursor.lastrowid
                        new_sub_code = last_sub_id + 10001
                        cursor.execute(
                            "UPDATE sub_accont_for_new_account SET sub_account_code = %s WHERE id_sub = %s",
                            (new_sub_code, last_sub_id)
                        )
                if supplier_id:
                    flash('Supplier updated successfully!', 'success')
                else:
                    flash('Supplier added successfully!', 'success')
            except Exception as e:
                flash(f'Error adding/updating supplier: {str(e)}', 'danger')

            return redirect(url_for('add_supplier'))

        except Exception as e:
            flash(f'An unexpected error occurred: {str(e)}', 'danger')
            return redirect(url_for('add_supplier'))

    salutations = []
    try:
        salutations_data = db.execute_query("SELECT salutation FROM suplier_suporting_1")
        salutations = [row['salutation'] for row in salutations_data]
    except Exception as e:
        logging.error(f"Error loading salutations: {e}")

    suppliers_list = []
    try:
        # Select explicit columns (sup_id aliased to id, no date columns) so the
        # row is safely JSON-serialisable for the Edit button's tojson.
        suppliers_list = db.execute_query("""
            SELECT sup_id AS id, supplier_name, supplier_code,
                   suppliers_teli_1, suppliers_teli_2, suppliers_credit_fasility,
                   suppliers_vat_regidter_no, suppliers_TIN, suppliers_NIC, suppliers_e_mail,
                   suppliers_default_payment_method
            FROM suppliers
            WHERE Is_Suplier = 1
            ORDER BY supplier_name
        """) or []
    except Exception as e:
        logging.error(f"Error loading suppliers: {e}")

    return render_template('add_supplier.html', salutations=salutations, suppliers_list=suppliers_list)

@app.route('/add_salutation', methods=['POST'])
@login_required
def add_salutation():
    new_salutation = request.form.get('new_salutation')
    if new_salutation:
        try:
            db.execute_query(
                "INSERT INTO suplier_suporting_1 (id, salutation) VALUES (%s, %s)",
                (0, new_salutation), commit=True
            )
            flash('Salutation added.', 'success')
        except Exception as e:
            flash(f'Error adding salutation: {e}', 'danger')
    return redirect(url_for('add_customer'))

@app.route('/add_salutation_ajax', methods=['POST'])
@login_required
def add_salutation_ajax():
    new_salutation = request.form.get('new_salutation')
    if new_salutation:
        try:
            db.execute_query(
                "INSERT INTO suplier_suporting_1 (id, salutation) VALUES (%s, %s)",
                (0, new_salutation), commit=True
            )
            return {"success": True}
        except Exception as e:
            return {"success": False, "error": str(e)}
    return {"success": False}

# --- Inventory Category Management ---
@app.route('/inventory_category')
@login_required
@has_permission('Access_Inventory')
def inventory_category():
    main_cats = db.execute_query("SELECT * FROM inventory_carogory WHERE main_catogory IS NOT NULL AND main_catogory != ''")
    sub_cats = db.execute_query("SELECT * FROM inventory_carogory WHERE sub_catogory IS NOT NULL AND sub_catogory != ''")
    return render_template('inventory_category.html', main_categories=main_cats, sub_categories=sub_cats)

@app.route('/inventory_category/main', methods=['POST'])
@login_required
def add_main_category():
    name = (request.form.get('main_category') or '').strip()
    if name:
        try:
            # The old code hard-coded "id = 0" on every insert. That worked
            # once, then every category added after the first hit a
            # duplicate-primary-key error -> uncaught exception -> 500
            # Internal Server Error. Compute the next free id ourselves so
            # it works regardless of whether the column is auto-increment.
            row = db.execute_query("SELECT COALESCE(MAX(id), 0) + 1 AS next_id FROM inventory_carogory")
            next_id = row[0]['next_id'] if row else 1
            db.execute_query("INSERT INTO inventory_carogory (id, main_catogory, sub_catogory) VALUES (%s, %s, NULL)",
                              (next_id, name), commit=True)
            flash('Main category added', 'success')
        except Exception as e:
            flash(f'Could not add main category: {e}', 'danger')
    return redirect(url_for('inventory_category'))

@app.route('/inventory_category/sub', methods=['POST'])
@login_required
def add_sub_category():
    name = (request.form.get('sub_category') or '').strip()
    if name:
        try:
            # Same fix as add_main_category — see comment there.
            row = db.execute_query("SELECT COALESCE(MAX(id), 0) + 1 AS next_id FROM inventory_carogory")
            next_id = row[0]['next_id'] if row else 1
            db.execute_query("INSERT INTO inventory_carogory (id, main_catogory, sub_catogory) VALUES (%s, NULL, %s)",
                              (next_id, name), commit=True)
            flash('Sub category added', 'success')
        except Exception as e:
            flash(f'Could not add sub category: {e}', 'danger')
    return redirect(url_for('inventory_category'))

@app.route('/inventory_category/main/toggle', methods=['POST'])
@login_required
def toggle_main_category():
    try:
        cat_id = request.form.get('id')
        current = int(request.form.get('current_status') or 0)
        new_status = 0 if current == 1 else 1
        db.execute_query("UPDATE inventory_carogory SET dis_continue_main = %s WHERE id = %s", (new_status, cat_id), commit=True)
    except Exception as e:
        flash(f'Could not update category: {e}', 'danger')
    return redirect(url_for('inventory_category'))

@app.route('/inventory_category/sub/toggle', methods=['POST'])
@login_required
def toggle_sub_category():
    try:
        cat_id = request.form.get('id')
        current = int(request.form.get('current_status') or 0)
        new_status = 0 if current == 1 else 1
        db.execute_query("UPDATE inventory_carogory SET dis_continue_sub = %s WHERE id = %s", (new_status, cat_id), commit=True)
    except Exception as e:
        flash(f'Could not update category: {e}', 'danger')
    return redirect(url_for('inventory_category'))

# --- Add Inventory Item ---
@app.route('/add_inventory_item', methods=['GET', 'POST'])
@login_required
@has_permission('Access_Inventory')
def add_inventory_item():
    if request.method == 'POST':
        try:


            # 1. Extract Data
            name = request.form.get('item_name')
            code = request.form.get('item_code')
            supplier_code = request.form.get('supplier_code')
            batch_code = request.form.get('batch_code')
            unit = request.form.get('measurement_unit')
            main_cat = request.form.get('main_category')
            sub_cat = request.form.get('sub_category')
            min_qty = parse_float(request.form.get('min_qty', 0))

            # Prices are now arrays
            cost_prices = request.form.getlist('cost_price[]')
            selling_prices = request.form.getlist('selling_price[]')
            special_prices = request.form.getlist('special_price[]')
            loyalty_prices = request.form.getlist('loyalty_price[]')

            # 2. Handle Image
            img_data = None
            if 'item_image' in request.files:
                file = request.files['item_image']
                if file.filename != '':
                    # C# code saves as JpegBitmapEncoder buffer (bytes)
                    # We store as LONGBLOB or MEDIUMBLOB.
                    # MySQL Connector handles bytes object directly for BLOBs.
                    # Wait, some tables expect base64 string because of C# legacy handling.
                    # Let's save as base64 string to avoid 'Invalid utf8mb4 character string' error
                    # when passing raw bytes to a query that expects text/string.
                    img_data = base64.b64encode(file.read()).decode('utf-8')

            if not name or not code or not unit:
                flash('Name, Code, and Unit are required.', 'danger')
                return redirect(url_for('add_inventory_item'))

            try:
                current_user = get_current_user_id()
                current_user_pk = get_current_user_pk()
                today_date = date.today()

                with db.transaction_cursor() as cursor:
                    # 3. Insert Item
                    query_item = """
                        INSERT INTO inventoy_items (
                            id, inventoy_name, inventoy_code, inventoy_suplier_code, inventoy_bach_code,
                            inventoy_img, inventoy_creat_user_id, inventoy_items_creat_date,
                            inventoy_items_messurment_unit, Main_Catogry, Sub_Catogory, min_qty, active
                        ) VALUES (0, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 1)
                    """
                    cursor.execute(query_item, (
                        name, code, supplier_code, batch_code, img_data,
                        current_user_pk, today_date, unit, main_cat, sub_cat, min_qty
                    ))
                    item_id = cursor.lastrowid

                    # 4. Insert Prices
                    query_price = '''
                        INSERT INTO inventory_price_recod (
                            id, inventory_price_link, inventory_price_purcharsing,
                            inventory_price_selling, inventory_price_profit_marging_comen,
                            inventory_price_for_Loyality_customer, created_date
                        ) VALUES (0, %s, %s, %s, %s, %s, %s)
                    '''

                    # If the user did not add any dynamic rows, the arrays might be empty.
                    # Or there might be 1 default row.
                    if cost_prices:
                        for idx, cp in enumerate(cost_prices):
                            c_val = parse_float(cp)
                            # Handle potential IndexError if arrays are mismatched (shouldn't happen with proper frontend)
                            s_val = parse_float(selling_prices[idx]) if idx < len(selling_prices) else 0.0
                            sp_val = parse_float(special_prices[idx]) if idx < len(special_prices) else 0.0
                            lp_val = parse_float(loyalty_prices[idx]) if idx < len(loyalty_prices) else 0.0

                            cursor.execute(query_price, (item_id, c_val, s_val, sp_val, lp_val, today_date))
                    else:
                        # Fallback if no prices sent, just create a zeroed row
                        cursor.execute(query_price, (item_id, 0.0, 0.0, 0.0, 0.0, today_date))

                flash('Inventory Item created successfully!', 'success')
                clear_pos_items_cache()

            except Exception as e:
                flash(f'Database Error: {str(e)}', 'danger')

        except Exception as e:
            flash(f'System Error: {str(e)}', 'danger')

        return redirect(url_for('add_inventory_item'))

    # GET Request
    main_cats = db.execute_query("SELECT main_catogory FROM inventory_carogory WHERE main_catogory IS NOT NULL AND main_catogory != '' AND dis_continue_main = 0")
    sub_cats = db.execute_query("SELECT sub_catogory FROM inventory_carogory WHERE sub_catogory IS NOT NULL AND sub_catogory != '' AND dis_continue_sub = 0")

    return render_template('add_inventory_item.html', main_categories=main_cats, sub_categories=sub_cats)

# --- GRN (Goods Received Note) Management ---
@app.route('/grn', methods=['GET', 'POST'])
@login_required
@has_permission('Access_Inventory')
def grn():
    if request.method == 'POST':
        try:
            # 1. Extract Data
            supplier_name = request.form.get('supplier')
            items_json = request.form.get('items_json')
            invoice_no = request.form.get('invoice_no')
            invoice_date = request.form.get('invoice_date')
            due_date = request.form.get('due_date')
            narration = request.form.get('narration')
            job_no = request.form.get('job_no')
            location = request.form.get('location')
            po_id = request.form.get('po_id')
            payment_method = (request.form.get('payment_method') or '').strip() or None

            total_value = parse_float(request.form.get('total_value', 0))
            vat_rate = parse_float(request.form.get('vat_rate', 0))
            vat_amount = parse_float(request.form.get('vat_amount', 0))
            grand_total = parse_float(request.form.get('grand_total', 0))

            items = json.loads(items_json) if items_json else []

            if not items:
                flash('No items in GRN', 'warning')
                return redirect(url_for('grn'))

            # 2. Get Supplier Details
            sup_res = db.execute_query("SELECT supplier_code, sup_id FROM suppliers WHERE supplier_name = %s", (supplier_name,))
            if not sup_res:
                flash('Invalid Supplier', 'danger')
                return redirect(url_for('grn'))
            supplier_code = sup_res[0]['supplier_code']
            supplier_id = sup_res[0]['sup_id']

            # 3. Create Transaction using Service Layer
            try:
                current_user = get_current_user_id()
                supplier_info = {'code': supplier_code, 'id': supplier_id}
                invoice_info = {
                    'no': invoice_no,
                    'date': invoice_date,
                    'due_date': due_date,
                    'narration': narration,
                    'job_no': job_no,
                    'location': location,
                    'total_value': total_value,
                    'vat_rate': vat_rate,
                    'vat_amount': vat_amount,
                    'grand_total': grand_total,
                    'payment_method': payment_method
                }

                jv_no = services.create_grn(db, current_user, supplier_info, invoice_info, items)

                # If Auto-Filled from PO, mark PO as completed (status = 2) so it hides from the dropdown
                if po_id:
                    db.execute_query("UPDATE OP_NO_Table SET status = 2 WHERE id = %s", (po_id,))

                current_user_pk = get_current_user_pk()
                flash(f'GRN created successfully. JV No: {jv_no}', 'success')
                return render_template('grn_print.html', grn_no=jv_no, supplier=supplier_name, date=invoice_date, invoice_no=invoice_no, location=location, items=items, total_value=total_value, vat_amount=vat_amount, grand_total=grand_total)

            except Exception as e:
                flash(f'Transaction failed: {str(e)}', 'danger')
                return redirect(url_for('grn'))

        except Exception as e:
            flash(f'Error processing GRN: {str(e)}', 'danger')
            return redirect(url_for('grn'))

    # GET Request: Load Form Data
    suppliers = db.execute_query("SELECT supplier_name, suppliers_default_payment_method FROM suppliers")
    items = db.execute_query("SELECT inventoy_name, inventoy_code, inventoy_items_messurment_unit FROM inventoy_items")
    jobs = db.execute_query("SELECT job_number FROM jobs_unit")
    locations = db.execute_query("SELECT inventory_locations_name FROM inventory_locations")

    return render_template('grn.html',
                           suppliers=suppliers,
                           items=items,
                           jobs=jobs,
                           locations=locations,
                           today_date=date.today().strftime('%Y-%m-%d'))


@app.route('/grn_reversal')
@login_required
@has_permission('Access_Reversals')
def grn_reversal():
    """List recent GRNs (Goods Received Notes) available for deletion/reversal,
    with optional filters by supplier, invoice number and date range."""
    supplier_id = (request.args.get('supplier_id') or '').strip()
    invoice_no = (request.args.get('invoice_no') or '').strip()
    date_from = (request.args.get('date_from') or '').strip()
    date_to = (request.args.get('date_to') or '').strip()

    filters = ["s.suppliers_oustanding_delete = 0", "j.jv_user_code = 'JV FROM GRN'"]
    params = []
    if supplier_id:
        # Filter by the supplier's NAME rather than sup_id: if this supplier
        # has more than one row in `suppliers` (e.g. a duplicate/renumbered
        # entry), a GRN's stored supplier link may point at a different
        # sup_id than the one the dropdown offers for that same name — an
        # ID match would then wrongly come up empty. Matching by name (the
        # same approach the Supplier Payment Method / Ready List reports
        # use) finds the GRN regardless of which duplicate row it's linked to.
        sup_name_row = db.execute_query("SELECT supplier_name FROM suppliers WHERE sup_id = %s", (supplier_id,))
        if sup_name_row:
            filters.append("sup.supplier_name = %s")
            params.append(sup_name_row[0]['supplier_name'])
        else:
            filters.append("sup.sup_id = %s")
            params.append(supplier_id)
    if invoice_no:
        filters.append("s.suppliers_invoice_number LIKE %s")
        params.append(f"%{invoice_no}%")
    if date_from:
        filters.append("s.suppliers_invoice_date >= %s")
        params.append(date_from)
    if date_to:
        filters.append("s.suppliers_invoice_date <= %s")
        params.append(date_to)

    where_clause = " AND ".join(filters)
    query = f"""
        SELECT
            s.suppliers_invoice_JV as jv,
            j.jv_naration,
            sup.supplier_name as SupplierName,
            s.suppliers_invoice_number as InvoiceNo,
            s.suppliers_invoice_date as Date,
            s.suppliers_invoice_total_oustanding as Amount
        FROM suppliers_invoice_data s
        JOIN jv_numbers j ON s.suppliers_invoice_JV = j.jv_id
        LEFT JOIN suppliers sup ON s.suppliers_invoice_buinding_supplier = sup.sup_id
        WHERE {where_clause}
        ORDER BY s.s_i_id DESC
        LIMIT 200
    """
    rows = db.execute_query(query, tuple(params)) if params else db.execute_query(query)

    suppliers = db.execute_query(
        "SELECT sup_id, supplier_name FROM suppliers WHERE Is_Suplier = 1 ORDER BY supplier_name") or []

    return render_template('grn_reversal.html', rows=rows, suppliers=suppliers,
                           f_supplier_id=supplier_id, f_invoice_no=invoice_no,
                           f_date_from=date_from, f_date_to=date_to)


@app.route('/grn_reversal/process', methods=['POST'])
@login_required
@has_permission('Access_Reversals')
def grn_reversal_process():
    """Delete/reverse a GRN: reverses the GL postings, undoes the stock-in
    movement, and marks the supplier invoice as deleted. Blocked if any
    payment has already been made against the GRN's supplier invoice, if the
    GRN has been bank reconciled, or if it falls in a closed accounting
    period — mirrors the guard rails used by service_entry_reversal."""
    jv = request.form.get('jv')
    if not jv:
        flash('No GRN selected', 'danger')
        return redirect(url_for('grn_reversal'))

    if jv_in_closed_period(jv):
        flash(f'Cannot delete: this GRN is in a closed period (locked through {get_period_lock_date()}).', 'danger')
        return redirect(url_for('grn_reversal'))

    conn = None
    cursor = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        # 1. Security Check: Bank Reconciled?
        cursor.execute("SELECT COUNT(*) FROM entry_details WHERE entry_jv = %s AND entry_Rec = 1", (jv,))
        if cursor.fetchone()[0] > 0:
            flash('Cannot delete: Transaction has been Bank Reconciled.', 'danger')
            return redirect(url_for('grn_reversal'))

        # 2. Security Check: Payments Made?
        cursor.execute(
            "SELECT suppliers_invoice_total_payment FROM suppliers_invoice_data "
            "WHERE suppliers_invoice_JV = %s AND suppliers_oustanding_delete = 0", (jv,))
        inv_res = cursor.fetchone()
        if not inv_res:
            flash('GRN not found or already deleted.', 'danger')
            return redirect(url_for('grn_reversal'))

        if inv_res[0] and float(inv_res[0]) > 0:
            flash('Cannot delete: Payments have been made against this GRN. Reverse those payments first.', 'danger')
            return redirect(url_for('grn_reversal'))

        conn.start_transaction()
        current_user_pk = get_current_user_pk()
        today = date.today()

        # 3. Create reversal JV
        cursor.execute(
            "INSERT INTO jv_numbers (jv_user_code, jv_naration, status) VALUES (%s, %s, 1)",
            (f'REV-GRN-{jv}', f'GRN Deletion Reversal of JV-{jv}'))
        rev_jv = cursor.lastrowid

        # 4. Reverse GL entries (swap DR <-> CR): undoes the Inventory DR /
        #    VAT Control DR / Account Payable CR posted when the GRN was created.
        cursor.execute("""
            INSERT INTO entry_details (
                account_name, enty_values_DR, enty_values_CR,
                entry_effective_date, entry_create_date,
                entry_naration, entry_create_user, entry_jv
            )
            SELECT account_name,
                   COALESCE(enty_values_CR, 0),
                   COALESCE(enty_values_DR, 0),
                   %s, %s, %s, %s, %s
            FROM entry_details WHERE entry_jv = %s
        """, (today, today, f'GRN Deletion Reversal of JV-{jv}', current_user_pk, rev_jv, jv))

        # 5. Reverse the stock-in movement (matching OUT movements)
        cursor.execute("""
            INSERT INTO inventory_recod (
                inventoy_name, inventoy_code,
                inventory_recod_action_date,
                inventory_recod_moument_in, inventory_recod_movment_out,
                inventory_recod_mesrmet, inventory_recod_unit_price,
                inventory_recod_account, inventory_recod_user_id,
                JV_No, inventory_recod_location
            )
            SELECT inventoy_name, inventoy_code, %s,
                   0, inventory_recod_moument_in,
                   inventory_recod_mesrmet, inventory_recod_unit_price,
                   'GRN Deletion Reversal', %s,
                   %s, inventory_recod_location
            FROM inventory_recod
            WHERE JV_No = %s AND inventory_recod_moument_in > 0
        """, (today, current_user_pk, jv, jv))

        # 6. Mark the supplier invoice as deleted. Also rename the invoice number
        # (it's UNIQUE at the DB level) so a future GRN can reuse the same
        # invoice number without hitting a duplicate-key error — the deleted
        # row is kept for audit purposes, just under a distinguishable number.
        cursor.execute(
            """UPDATE suppliers_invoice_data
               SET suppliers_oustanding_delete = 1,
                   suppliers_invoice_number = LEFT(CONCAT(COALESCE(suppliers_invoice_number, ''), '-REV', %s), 200)
               WHERE suppliers_invoice_JV = %s""",
            (rev_jv, jv))

        conn.commit()
        flash(f'GRN (JV: {jv}) deleted successfully. Reversal JV: {rev_jv}', 'success')

    except Exception as e:
        if conn:
            conn.rollback()
        flash(f'Error deleting GRN: {str(e)}', 'danger')
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

    return redirect(url_for('grn_reversal'))


@app.route('/grn/edit/<int:jv_no>', methods=['GET', 'POST'])
@login_required
@has_permission('Access_Inventory')
def grn_edit(jv_no):
    """Edit a previously-created GRN. Since a GRN is already posted to the GL
    and to inventory, 'editing' is done the same audit-safe way as everywhere
    else in this app (see grn_reversal_process): the original GRN is reversed
    (GL entries + stock movement) and a brand new, corrected GRN is created in
    its place — both inside one DB transaction, so it's all-or-nothing. Blocked
    by the same guard rails as deletion: payments already made against it,
    bank reconciliation, or a closed accounting period."""
    jvrow = db.execute_query("SELECT jv_naration, jv_user_code FROM jv_numbers WHERE jv_id = %s", (jv_no,))
    if not jvrow or jvrow[0]['jv_user_code'] != 'JV FROM GRN':
        flash('No GRN found to edit.', 'danger')
        return redirect(url_for('grn_reversal'))

    inv_row = db.execute_query("""
        SELECT s.suppliers_invoice_number, s.suppliers_invoice_date, s.suppliers_invoice_final_date,
               s.suppliers_invoice_total_payment, s.suppliers_invoice_buinding_supplier,
               s.suppliers_VAT_rate, s.suppliers_oustanding_delete, s.suppliers_invoice_payment_method, sup.supplier_name
        FROM suppliers_invoice_data s
        LEFT JOIN suppliers sup ON s.suppliers_invoice_buinding_supplier = sup.sup_id
        WHERE s.suppliers_invoice_JV = %s
    """, (jv_no,))
    if not inv_row:
        flash('GRN invoice record not found.', 'danger')
        return redirect(url_for('grn_reversal'))
    inv_row = inv_row[0]

    if inv_row['suppliers_oustanding_delete']:
        flash('This GRN has already been deleted.', 'danger')
        return redirect(url_for('grn_reversal'))

    if inv_row['suppliers_invoice_total_payment'] and float(inv_row['suppliers_invoice_total_payment']) > 0:
        flash('Cannot edit: Payments have been made against this GRN. Reverse those payments first.', 'danger')
        return redirect(url_for('grn_reversal'))

    if jv_in_closed_period(jv_no):
        flash(f'Cannot edit: this GRN is in a closed period (locked through {get_period_lock_date()}).', 'danger')
        return redirect(url_for('grn_reversal'))

    bank_rec = db.execute_query("SELECT COUNT(*) as c FROM entry_details WHERE entry_jv = %s AND entry_Rec = 1", (jv_no,))
    if bank_rec and bank_rec[0]['c'] > 0:
        flash('Cannot edit: Transaction has been Bank Reconciled.', 'danger')
        return redirect(url_for('grn_reversal'))

    if request.method == 'POST':
        supplier_name = request.form.get('supplier')
        items_json = request.form.get('items_json')
        invoice_no = request.form.get('invoice_no')
        invoice_date = request.form.get('invoice_date')
        due_date = request.form.get('due_date')
        narration = request.form.get('narration')
        job_no = request.form.get('job_no') or None
        location = request.form.get('location')
        payment_method = (request.form.get('payment_method') or '').strip() or None

        total_value = parse_float(request.form.get('total_value', 0))
        vat_rate = parse_float(request.form.get('vat_rate', 0))
        vat_amount = parse_float(request.form.get('vat_amount', 0))
        grand_total = parse_float(request.form.get('grand_total', 0))

        items = json.loads(items_json) if items_json else []
        if not items:
            flash('No items in GRN', 'warning')
            return redirect(url_for('grn_edit', jv_no=jv_no))

        sup_res = db.execute_query("SELECT supplier_code, sup_id FROM suppliers WHERE supplier_name = %s", (supplier_name,))
        if not sup_res:
            flash('Invalid Supplier', 'danger')
            return redirect(url_for('grn_edit', jv_no=jv_no))
        supplier_code = sup_res[0]['supplier_code']
        supplier_id = sup_res[0]['sup_id']

        conn = None
        cursor = None
        try:
            conn = db.get_connection()
            cursor = conn.cursor()

            # Re-check guards inside the transaction (state may have changed
            # between loading the edit form and submitting it).
            cursor.execute("SELECT COUNT(*) FROM entry_details WHERE entry_jv = %s AND entry_Rec = 1", (jv_no,))
            if cursor.fetchone()[0] > 0:
                flash('Cannot edit: Transaction has been Bank Reconciled.', 'danger')
                return redirect(url_for('grn_reversal'))

            cursor.execute(
                "SELECT suppliers_invoice_total_payment FROM suppliers_invoice_data "
                "WHERE suppliers_invoice_JV = %s AND suppliers_oustanding_delete = 0", (jv_no,))
            chk = cursor.fetchone()
            if not chk:
                flash('GRN not found or already deleted.', 'danger')
                return redirect(url_for('grn_reversal'))
            if chk[0] and float(chk[0]) > 0:
                flash('Cannot edit: Payments have been made against this GRN. Reverse those payments first.', 'danger')
                return redirect(url_for('grn_reversal'))

            conn.start_transaction()
            current_user_pk = get_current_user_pk()
            current_user = get_current_user_id()
            today = date.today()

            # 1. Reverse the original GRN (mirrors grn_reversal_process step-for-step)
            cursor.execute(
                "INSERT INTO jv_numbers (jv_user_code, jv_naration, status) VALUES (%s, %s, 1)",
                (f'REV-GRN-{jv_no}', f'GRN Edit Reversal of JV-{jv_no}'))
            rev_jv = cursor.lastrowid

            cursor.execute("""
                INSERT INTO entry_details (
                    account_name, enty_values_DR, enty_values_CR,
                    entry_effective_date, entry_create_date,
                    entry_naration, entry_create_user, entry_jv
                )
                SELECT account_name,
                       COALESCE(enty_values_CR, 0),
                       COALESCE(enty_values_DR, 0),
                       %s, %s, %s, %s, %s
                FROM entry_details WHERE entry_jv = %s
            """, (today, today, f'GRN Edit Reversal of JV-{jv_no}', current_user_pk, rev_jv, jv_no))

            cursor.execute("""
                INSERT INTO inventory_recod (
                    inventoy_name, inventoy_code,
                    inventory_recod_action_date,
                    inventory_recod_moument_in, inventory_recod_movment_out,
                    inventory_recod_mesrmet, inventory_recod_unit_price,
                    inventory_recod_account, inventory_recod_user_id,
                    JV_No, inventory_recod_location
                )
                SELECT inventoy_name, inventoy_code, %s,
                       0, inventory_recod_moument_in,
                       inventory_recod_mesrmet, inventory_recod_unit_price,
                       'GRN Edit Reversal', %s,
                       %s, inventory_recod_location
                FROM inventory_recod
                WHERE JV_No = %s AND inventory_recod_moument_in > 0
            """, (today, current_user_pk, jv_no, jv_no))

            # Same rename as grn_reversal_process: free up the invoice number
            # (UNIQUE at the DB level) so the corrected GRN below can reuse it
            # — which is the normal case for an edit that isn't renumbering
            # the invoice, and is exactly what caused the duplicate-key error.
            cursor.execute(
                """UPDATE suppliers_invoice_data
                   SET suppliers_oustanding_delete = 1,
                       suppliers_invoice_number = LEFT(CONCAT(COALESCE(suppliers_invoice_number, ''), '-REV', %s), 200)
                   WHERE suppliers_invoice_JV = %s""",
                (rev_jv, jv_no))

            # 2. Create the corrected GRN (mirrors services.create_grn step-for-step)
            cursor.execute("INSERT INTO jv_numbers (jv_user_code, jv_naration) VALUES (%s, %s)",
                           ('JV FROM GRN', narration))
            new_jv = cursor.lastrowid

            cursor.execute("""
                INSERT INTO suppliers_invoice_data (
                    suppliers_code, suppliers_invoice_number, suppliers_invoice_date,
                    suppliers_invoice_total_oustanding, suppliers_invoice_final_date,
                    suppliers_invoice_buinding_supplier, suppliers_invoice_JV, suppliers_VAT_rate,
                    suppliers_invoice_payment_method, suppliers_invoice_total_payment
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 0)
            """, (supplier_code, invoice_no, invoice_date, grand_total, due_date, supplier_id, new_jv, vat_rate, payment_method))

            cursor.execute("""
                INSERT INTO entry_details (
                    account_name, enty_values_CR, entry_effective_date, entry_create_date,
                    entry_naration, entry_create_user, entry_jv, entry_job_number
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, ('Account Payable', grand_total, invoice_date, today, narration, current_user, new_jv, job_no))

            cursor.execute("""
                INSERT INTO entry_details (
                    account_name, enty_values_DR, entry_effective_date, entry_create_date,
                    entry_naration, entry_create_user, entry_jv, entry_job_number
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, ('Inventory', total_value, invoice_date, today, narration, current_user, new_jv, job_no))

            if vat_amount > 0:
                cursor.execute("""
                    INSERT INTO entry_details (
                        account_name, enty_values_DR, entry_effective_date, entry_create_date,
                        entry_naration, entry_create_user, entry_jv, entry_job_number
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """, ('VAT Control', round(vat_amount, 2), invoice_date, today, narration, current_user, new_jv, job_no))

            for item in items:
                cursor.execute("""
                    INSERT INTO inventory_recod (
                        inventoy_name, inventoy_code, inventory_recod_mesrmet,
                        inventory_recod_unit_price, inventory_recod_moument_in, inventory_recod_movment_out,
                        inventory_recod_suplier_iv_no, inventory_recod_user_id, inventory_recod_user_recod_date,
                        inventory_recod_location, inventory_recod_link_invoice, inventory_recod_action_date, JV_No
                    ) VALUES (%s, %s, %s, %s, %s, 0, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    item['name'], item['code'], item['unit'], item['cost'], item['qty'],
                    invoice_no, current_user, today, location, new_jv, invoice_date, new_jv
                ))

            conn.commit()
            flash(f'GRN updated. Old JV-{jv_no} reversed; corrected GRN created as JV-{new_jv}.', 'success')
            return redirect(url_for('grn_reversal'))

        except Exception as e:
            if conn:
                conn.rollback()
            flash(f'Error updating GRN: {str(e)}', 'danger')
            return redirect(url_for('grn_edit', jv_no=jv_no))
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()

    # GET: Load form data + prefill values from the existing GRN
    suppliers = db.execute_query("SELECT supplier_name, suppliers_default_payment_method FROM suppliers")
    items_master = db.execute_query("SELECT inventoy_name, inventoy_code, inventoy_items_messurment_unit FROM inventoy_items")
    jobs = db.execute_query("SELECT job_number FROM jobs_unit")
    locations = db.execute_query("SELECT inventory_locations_name FROM inventory_locations")

    item_rows = db.execute_query("""
        SELECT inventoy_name, inventoy_code, inventory_recod_mesrmet,
               inventory_recod_unit_price, inventory_recod_moument_in, inventory_recod_location
        FROM inventory_recod WHERE JV_No = %s AND inventory_recod_moument_in > 0
    """, (jv_no,)) or []
    prefill_items = [{
        'name': r['inventoy_name'], 'code': r['inventoy_code'] or '',
        'unit': r['inventory_recod_mesrmet'] or '',
        'qty': float(r['inventory_recod_moument_in'] or 0),
        'cost': float(r['inventory_recod_unit_price'] or 0),
        'total': round(float(r['inventory_recod_unit_price'] or 0) * float(r['inventory_recod_moument_in'] or 0), 2)
    } for r in item_rows]
    prefill_location = item_rows[0]['inventory_recod_location'] if item_rows else ''

    job_row = db.execute_query(
        "SELECT entry_job_number FROM entry_details WHERE entry_jv = %s AND entry_job_number IS NOT NULL LIMIT 1",
        (jv_no,))
    prefill_job_no = job_row[0]['entry_job_number'] if job_row else ''

    vat_row = db.execute_query(
        "SELECT enty_values_DR AS amt FROM entry_details WHERE entry_jv = %s AND account_name = 'VAT Control'",
        (jv_no,))
    prefill_vat_amount = float(vat_row[0]['amt']) if vat_row else 0.0

    return render_template('grn.html',
                           suppliers=suppliers,
                           items=items_master,
                           jobs=jobs,
                           locations=locations,
                           today_date=date.today().strftime('%Y-%m-%d'),
                           edit_mode=True,
                           edit_jv=jv_no,
                           prefill_supplier=inv_row['supplier_name'],
                           prefill_invoice_no=inv_row['suppliers_invoice_number'],
                           prefill_invoice_date=str(inv_row['suppliers_invoice_date']),
                           prefill_due_date=str(inv_row['suppliers_invoice_final_date']),
                           prefill_narration=jvrow[0]['jv_naration'],
                           prefill_job_no=prefill_job_no,
                           prefill_location=prefill_location,
                           prefill_items=prefill_items,
                           prefill_vat_rate=float(inv_row['suppliers_VAT_rate'] or 0),
                           prefill_vat_amount=prefill_vat_amount,
                           prefill_payment_method=inv_row['suppliers_invoice_payment_method'] or '')


@app.route('/grn/view/<int:jv_no>')
@login_required
def grn_view(jv_no):
    """Read-only view of a previously-created GRN, reconstructed from the
    stored supplier invoice, JV and inventory movement records. Used by the
    'View GRN' link on Cash/Bank Payment's outstanding invoices list, so
    users can see what a GRN invoice actually contained before paying it."""
    jvrow = db.execute_query("SELECT jv_naration, jv_user_code FROM jv_numbers WHERE jv_id = %s", (jv_no,))
    if not jvrow or jvrow[0]['jv_user_code'] != 'JV FROM GRN':
        flash('No GRN found for this invoice.', 'danger')
        return redirect(request.referrer or url_for('grn'))

    inv = db.execute_query("""
        SELECT suppliers_invoice_number, suppliers_invoice_date, suppliers_invoice_total_oustanding,
               suppliers_invoice_buinding_supplier
        FROM suppliers_invoice_data WHERE suppliers_invoice_JV = %s
    """, (jv_no,))
    if not inv:
        flash('GRN invoice record not found.', 'danger')
        return redirect(request.referrer or url_for('grn'))
    inv = inv[0]

    sup = db.execute_query("SELECT supplier_name FROM suppliers WHERE sup_id = %s",
                           (inv['suppliers_invoice_buinding_supplier'],))
    supplier_name = sup[0]['supplier_name'] if sup else ''

    # Item lines from the inventory movement records tagged to this GRN
    item_rows = db.execute_query("""
        SELECT inventoy_name, inventoy_code, inventory_recod_mesrmet, inventory_recod_unit_price,
               inventory_recod_moument_in, inventory_recod_location
        FROM inventory_recod WHERE JV_No = %s AND inventory_recod_moument_in > 0
    """, (jv_no,)) or []
    items = [{
        'name': r['inventoy_name'], 'code': r['inventoy_code'] or '',
        'unit': r['inventory_recod_mesrmet'] or '',
        'cost': float(r['inventory_recod_unit_price'] or 0),
        'qty': float(r['inventory_recod_moument_in'] or 0),
        'total': float(r['inventory_recod_unit_price'] or 0) * float(r['inventory_recod_moument_in'] or 0),
    } for r in item_rows]
    location = item_rows[0]['inventory_recod_location'] if item_rows else ''
    total_value = sum(i['total'] for i in items)

    vat_row = db.execute_query(
        "SELECT enty_values_DR AS amt FROM entry_details WHERE entry_jv = %s AND account_name = 'VAT Control'",
        (jv_no,))
    vat_amount = float(vat_row[0]['amt']) if vat_row else 0.0
    grand_total = float(inv['suppliers_invoice_total_oustanding'] or 0) or round(total_value + vat_amount, 2)

    return render_template('grn_print.html', grn_no=jv_no, supplier=supplier_name,
                           date=str(inv['suppliers_invoice_date']), invoice_no=inv['suppliers_invoice_number'],
                           location=location, items=items, total_value=round(total_value, 2),
                           vat_amount=round(vat_amount, 2), grand_total=grand_total, read_only=True)


@app.route('/api/check_duplicate_invoice')
@login_required
def check_duplicate_invoice():
    """Used by the GRN screen while typing the invoice number: tells it
    whether this supplier already has an invoice on file under that same
    number, so the user can jump straight to it (via grn_view) instead of
    accidentally re-entering the same bill twice."""
    supplier_name = (request.args.get('supplier') or '').strip()
    invoice_no = (request.args.get('invoice_no') or '').strip()
    exclude_jv = (request.args.get('exclude_jv') or '').strip()

    if not supplier_name or not invoice_no:
        return jsonify({'duplicate': False})

    query = """
        SELECT s.suppliers_invoice_JV AS jv, s.suppliers_invoice_date AS inv_date,
               s.suppliers_invoice_total_oustanding AS amount
        FROM suppliers_invoice_data s
        JOIN suppliers sup ON s.suppliers_invoice_buinding_supplier = sup.sup_id
        WHERE sup.supplier_name = %s
          AND s.suppliers_invoice_number = %s
          AND COALESCE(s.suppliers_oustanding_delete, 0) = 0
    """
    params = [supplier_name, invoice_no]
    if exclude_jv:
        query += " AND s.suppliers_invoice_JV != %s"
        params.append(exclude_jv)
    query += " ORDER BY s.s_i_id DESC LIMIT 1"

    row = db.execute_query(query, tuple(params))
    if not row:
        return jsonify({'duplicate': False})

    r = row[0]
    return jsonify({
        'duplicate': True,
        'jv': r['jv'],
        'date': str(r['inv_date']) if r['inv_date'] else '',
        'amount': float(r['amount'] or 0),
    })


@app.route('/api/inventory_items')
@login_required
def api_inventory_items():
    """Live inventory item list so a newly created item shows up in an open
    window (e.g. GRN) without a full page refresh."""
    items = db.execute_query(
        "SELECT inventoy_name, inventoy_code, inventoy_items_messurment_unit AS unit FROM inventoy_items") or []
    return jsonify([{'name': i['inventoy_name'], 'code': i['inventoy_code'] or '',
                     'unit': i['unit'] or ''} for i in items])

# --- Inventory Locations ---
@app.route('/inventory_locations', methods=['GET', 'POST'])
@login_required
@has_permission('Access_Inventory')
def inventory_locations():
    if request.method == 'POST':
        name = request.form.get('house_name')
        desc = request.form.get('description')
        if name:
            db.execute_query("INSERT INTO inventory_locations (id, inventory_locations_name, inventory_locations_descriptions) VALUES (0, %s, %s)", (name, desc), commit=True)
            flash('Inventory location added', 'success')
        return redirect(url_for('inventory_locations'))

    locations = db.execute_query("SELECT * FROM inventory_locations")
    return render_template('inventory_locations.html', locations=locations)

# --- Cash Flow Categories ---
@app.route('/cash_flow_categories', methods=['GET', 'POST'])
@login_required
@has_permission('Access_Accounting')
def cash_flow_categories():
    if request.method == 'POST':
        category_id = int(request.form.get('category_id', 0))
        name = request.form.get('category_name')
        level = request.form.get('hold_level')

        if not name:
            flash('Category Name is required', 'danger')
            return redirect(url_for('cash_flow_categories'))

        try:
            if category_id == 0:
                # Insert
                db.execute_query("INSERT INTO cf_catogory (catogory_name, hold_level) VALUES (%s, %s)", (name, level), commit=True)
                clear_category_cache()
                flash('Category added successfully', 'success')
            else:
                # Update
                db.execute_query("UPDATE cf_catogory SET catogory_name = %s, hold_level = %s WHERE id = %s", (name, level, category_id), commit=True)
                clear_category_cache()
                flash('Category updated successfully', 'success')
        except Exception as e:
            flash(f'Error saving category: {str(e)}', 'danger')

        return redirect(url_for('cash_flow_categories'))

    cats = db.execute_query("SELECT * FROM cf_catogory ORDER BY hold_level, catogory_name")
    return render_template('cash_flow_categories.html', categories=cats)

@app.route('/cash_flow_categories/delete', methods=['POST'])
@login_required
@has_permission('Access_Accounting')
def delete_cash_flow_category():
    # Support both single deletion (from old form) and bulk (from new form)
    selected_ids = request.form.getlist('selected_ids')
    single_id = request.form.get('id')

    try:
        if selected_ids:
            placeholders = ', '.join(['%s'] * len(selected_ids))
            query = f"DELETE FROM cf_catogory WHERE id IN ({placeholders})"
            db.execute_query(query, tuple(selected_ids), commit=True)
            clear_category_cache()
            flash(f'{len(selected_ids)} categories deleted', 'success')
        elif single_id:
            db.execute_query("DELETE FROM cf_catogory WHERE id = %s", (single_id,), commit=True)
            clear_category_cache()
            flash('Category deleted', 'success')
        else:
            flash('No items selected', 'info')
    except Exception as e:
        flash(f'Error deleting categories: {str(e)}', 'danger')

    return redirect(url_for('cash_flow_categories'))


# --- Edit Account ---
@app.route('/edit_account/<int:account_id>', methods=['GET', 'POST'])
@login_required
@has_permission('Access_Accounting')
def edit_account(account_id):
    if request.method == 'POST':
        account_name = request.form.get('account_name')
        currency_code = request.form.get('currency_code', 'LKR')

        if not account_name:
            flash('Please enter an account name', 'danger')
            return redirect(url_for('edit_account', account_id=account_id))

        bs_cat_val = request.form.get('bs_category')
        pl_cat_val = request.form.get('income_category')
        cf_cat = request.form.get('cf_category')

        if (not bs_cat_val or bs_cat_val == "") and (not pl_cat_val or pl_cat_val == ""):
            flash('Please select a category', 'danger')
            return redirect(url_for('edit_account', account_id=account_id))

        bs_name = None
        bs_pos = None
        if bs_cat_val:
            parts = bs_cat_val.split(',')
            bs_name = parts[0]
            if len(parts) > 1:
                bs_pos = parts[1]

        pl_name = None
        pl_pos = None
        if pl_cat_val:
            parts = pl_cat_val.split(',')
            pl_name = parts[0]
            if len(parts) > 1:
                pl_pos = parts[1]

        is_income = 1 if 'income' in request.form.getlist('account_type') else 0
        is_expense = 1 if 'expense' in request.form.getlist('account_type') else 0
        is_liability = 1 if 'liability' in request.form.getlist('account_type') else 0
        is_equity = 1 if 'equity' in request.form.getlist('account_type') else 0
        is_asset = 1 if 'asset' in request.form.getlist('account_type') else 0

        query = """
            UPDATE new_account_table SET
                account_name = %s,
                account_hold_possion_PL = %s,
                account_hold_possion_Balace_Sheet = %s,
                account_name_of_catogory_PL = %s,
                account_name_of_catogory_Balace_sheet = %s,
                account_income = %s,
                account_expenses = %s,
                account_assets = %s,
                account_liabilities = %s,
                account_equity = %s,
                cf_catogory = %s,
                currency_code = %s
            WHERE id = %s
        """
        params = (
            account_name, pl_pos, bs_pos, pl_name, bs_name,
            is_income, is_expense, is_asset, is_liability, is_equity,
            cf_cat, currency_code, account_id
        )

        try:
            db.execute_query(query, params, commit=True)
            flash('Account updated successfully!', 'success')
            return redirect(url_for('chart_of_accounts'))
        except Exception as e:
            flash(f'Error updating account: {str(e)}', 'danger')

    # GET request
    account = db.execute_query("SELECT * FROM new_account_table WHERE id = %s", (account_id,))
    if not account:
        flash('Account not found', 'danger')
        return redirect(url_for('chart_of_accounts'))

    account = account[0]

    # Pre-fetch lookup data — use the same correct tables as add_new_account
    bs_categories, pl_categories, cf_categories = get_cached_categories(db)
    currencies = db.execute_query("SELECT currency_code, currency_name FROM currency_table")
    if not currencies:
        currencies = [{'currency_code': 'LKR', 'currency_name': 'Sri Lankan Rupee'}]
    existing_accounts = db.execute_query("SELECT account_name FROM new_account_table WHERE account_active = 1")

    return render_template('edit_account.html',
                           account=account,
                           bs_categories=bs_categories,
                           pl_categories=pl_categories,
                           cf_categories=cf_categories,
                           currencies=currencies,
                           existing_accounts=existing_accounts)

def _company_display_name():
    """Company name for report headers (Excel exports etc.)."""
    try:
        res = db.execute_query("SELECT company_name FROM company LIMIT 1")
        return (res[0]['company_name'] if res else '') or ''
    except Exception:
        return ''


def _category_levels(table):
    """Return {category_name: holding_level} from a categories table.
    This is the AUTHORITATIVE holding level (the per-account copies in
    new_account_table go stale when a category's level is edited later)."""
    try:
        rows = db.execute_query(f"SELECT name_of_category, holding_position FROM {table}") or []
    except Exception:
        return {}
    out = {}
    for r in rows:
        try:
            out[r['name_of_category']] = int(r['holding_position'])
        except (TypeError, ValueError):
            out[r['name_of_category']] = 9999
    return out


# --- Chart of Accounts ---
@app.route('/chart_of_accounts')
@login_required
@has_permission('Access_Accounting')
def chart_of_accounts():
    accounts = db.execute_query("""
        SELECT * FROM new_account_table
        WHERE account_active = 1
        ORDER BY account_name
    """) or []

    pl_levels = _category_levels('`p&l_category`')
    bs_levels = _category_levels('balance_sheet_category')

    def build_groups(accs, cat_field, levels, sort_field):
        groups = {}
        for a in accs:
            cat = a.get(cat_field) or 'Uncategorized'
            g = groups.setdefault(cat, {'name': cat, 'level': levels.get(cat), 'accounts': []})
            g['accounts'].append(a)
        for g in groups.values():
            g['accounts'].sort(key=lambda a: (
                a.get(sort_field) if a.get(sort_field) is not None else 9999,
                a['account_name']))
        return sorted(groups.values(),
                      key=lambda g: (g['level'] if g['level'] is not None else 9999, g['name']))

    pl_accounts = [a for a in accounts if a.get('account_name_of_catogory_PL')]
    bs_accounts = [a for a in accounts
                   if not a.get('account_name_of_catogory_PL')
                   and a.get('account_name_of_catogory_Balace_sheet')]
    other_accounts = [a for a in accounts
                      if not a.get('account_name_of_catogory_PL')
                      and not a.get('account_name_of_catogory_Balace_sheet')]

    pl_groups = build_groups(pl_accounts, 'account_name_of_catogory_PL', pl_levels, 'account_pl_sort')
    bs_groups = build_groups(bs_accounts, 'account_name_of_catogory_Balace_sheet', bs_levels, 'account_bs_sort')

    # Sub-accounts have no listing/edit UI anywhere in the app — add_new_account.html
    # can only create one. List every sub-account here (active and inactive) with an
    # Edit action so an existing one's name / parent account / active flag can be fixed.
    sub_accounts = db.execute_query("""
        SELECT id_sub, sub_sub_accaount_name, sub_new_account, active, sub_account_code
        FROM sub_accont_for_new_account
        ORDER BY sub_new_account, sub_sub_accaount_name
    """) or []
    existing_accounts = db.execute_query(
        "SELECT account_name FROM new_account_table WHERE account_active = 1 ORDER BY account_name") or []

    return render_template('chart_of_accounts.html',
                           pl_groups=pl_groups, bs_groups=bs_groups,
                           other_accounts=other_accounts,
                           sub_accounts=sub_accounts,
                           existing_accounts=existing_accounts,
                           total_accounts=len(accounts),
                           pl_count=len(pl_accounts), bs_count=len(bs_accounts))


@app.route('/edit_sub_account', methods=['POST'])
@login_required
@has_permission('Access_Accounting')
def edit_sub_account():
    sub_id = request.form.get('sub_id')
    name = (request.form.get('sub_account_name') or '').strip()
    main_account = request.form.get('main_account_select') or ''
    active = 1 if request.form.get('active') == 'on' else 0

    if not sub_id or not name or not main_account:
        flash('Sub account name and main account are required', 'danger')
        return redirect(url_for('chart_of_accounts'))

    try:
        db.execute_query(
            "UPDATE sub_accont_for_new_account SET sub_sub_accaount_name = %s, sub_new_account = %s, active = %s WHERE id_sub = %s",
            (name, main_account, active, sub_id), commit=True)
        flash('Sub account updated successfully', 'success')
    except Exception as e:
        flash(f'Error updating sub account: {str(e)}', 'danger')

    return redirect(url_for('chart_of_accounts'))

# --- Add New Account ---
@app.route('/add_new_account', methods=['GET', 'POST'])
@login_required
@has_permission('Access_Accounting')
def add_new_account():
    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'add_account':
            account_name = request.form.get('account_name')
            currency_code = request.form.get('currency_code', 'LKR') # Default LKR

            if not account_name:
                flash('Please enter an account name', 'danger')
                return redirect(url_for('add_new_account'))

            # Process categories
            bs_cat_val = request.form.get('bs_category')
            pl_cat_val = request.form.get('income_category')
            cf_cat = request.form.get('cf_category')

            if (not bs_cat_val or bs_cat_val == "") and (not pl_cat_val or pl_cat_val == ""):
                flash('Please select a category', 'danger')
                return redirect(url_for('add_new_account'))

            if not cf_cat:
                flash('Please select a cash flow category', 'danger')
                return redirect(url_for('add_new_account'))

            bs_name = None
            bs_pos = None
            if bs_cat_val:
                bs_name, bs_pos = bs_cat_val.split(',')

            pl_name = None
            pl_pos = None
            if pl_cat_val:
                pl_name, pl_pos = pl_cat_val.split(',')

            # Account Types
            is_income = 1 if 'income' in request.form.getlist('account_type') else 0
            is_expense = 1 if 'expense' in request.form.getlist('account_type') else 0
            is_liability = 1 if 'liability' in request.form.getlist('account_type') else 0
            is_equity = 1 if 'equity' in request.form.getlist('account_type') else 0
            is_asset = 1 if 'asset' in request.form.getlist('account_type') else 0

            current_user = get_current_user_id()

            query = """
                INSERT INTO new_account_table (
                    account_name, account_hold_possion_PL, account_hold_possion_Balace_Sheet,
                    account_name_of_catogory_PL, account_name_of_catogory_Balace_sheet,
                    account_income, account_expenses, account_assets, account_liabilities, account_equity,
                    cf_catogory, accont_create_date, account_create_user, account_active, account_basment,
                    currency_code
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 1, '', %s)
            """
            params = (
                account_name, pl_pos, bs_pos, pl_name, bs_name,
                is_income, is_expense, is_asset, is_liability, is_equity,
                cf_cat, date.today(), current_user, currency_code
            )

            try:
                db.execute_query(query, params, commit=True)
                flash('New account created successfully', 'success')
            except Exception as e:
                flash(f'Error creating account: {str(e)}', 'danger')

        elif action == 'add_sub_account':
            sub_name = request.form.get('sub_account_name')
            main_account = request.form.get('main_account_select')

            if not sub_name or not main_account:
                flash('Sub account name and main account are required', 'danger')
                return redirect(url_for('add_new_account'))

            current_user = get_current_user_id()

            query = """
                INSERT INTO sub_accont_for_new_account (
                    sub_sub_accaount_name, sub_new_account, creat_user, creat_date, active, sub_account_code
                ) VALUES (%s, %s, %s, %s, 1, 0)
            """
            try:
                new_id = db.execute_query(query, (sub_name, main_account, current_user, date.today()), commit=True)
                # Assign a unique code (same scheme as supplier/customer subs) so that
                # postings to this sub can be attributed and shown under its main account.
                if new_id:
                    db.execute_query(
                        "UPDATE sub_accont_for_new_account SET sub_account_code = %s WHERE id_sub = %s",
                        (int(new_id) + 10001, new_id), commit=True)
                flash('Sub account created successfully', 'success')
            except Exception as e:
                flash(f'Error creating sub account: {str(e)}', 'danger')

        return redirect(url_for('add_new_account'))

    # Load Data for Dropdowns
    bs_cats, pl_cats, cf_cats = get_cached_categories(db)
    existing_accounts = db.execute_query("SELECT account_name FROM new_account_table WHERE account_active = 1")
    currencies = db.execute_query("SELECT currency_code, currency_name FROM currency_table")
    if not currencies: # Fallback if table empty
        currencies = [{'currency_code': 'LKR', 'currency_name': 'Sri Lankan Rupee'}]

    return render_template('add_new_account.html',
                           bs_categories=bs_cats,
                           pl_categories=pl_cats,
                           cf_categories=cf_cats,
                           existing_accounts=existing_accounts,
                           currencies=currencies)

# --- Balance Sheet Category Management ---
@app.route('/balance_sheet_category', methods=['GET', 'POST'])
@login_required
@has_permission('Access_Accounting')
def balance_sheet_category():
    if request.method == 'POST':
        category_id = int(request.form.get('category_id', 0))
        name = request.form.get('category_name')
        level = request.form.get('holding_level')

        current_user = get_current_user_id()
        current_user_pk = get_current_user_pk()
        current_date = date.today()

        if not name or not level:
            flash('Name and Level are required', 'danger')
            return redirect(url_for('balance_sheet_category'))

        try:
            if category_id == 0:
                # Insert
                query = """
                    INSERT INTO balance_sheet_category
                    (id, name_of_category, holding_position, create_date_time, create_user_code)
                    VALUES (0, %s, %s, %s, %s)
                """
                db.execute_query(query, (name, level, current_date, current_user_pk), commit=True)
                clear_category_cache()
                flash('Category created successfully', 'success')
            else:
                # Update
                query = """
                    UPDATE balance_sheet_category
                    SET name_of_category = %s, holding_position = %s,
                        create_date_time = %s, create_user_code = %s
                    WHERE id = %s
                """
                db.execute_query(query, (name, level, current_date, current_user_pk, category_id), commit=True)
                clear_category_cache()
                flash('Category updated successfully', 'success')

        except Exception as e:
            flash(f'Error saving category: {str(e)}', 'danger')

        return redirect(url_for('balance_sheet_category'))

    # GET
    categories = db.execute_query("SELECT * FROM balance_sheet_category ORDER BY holding_position")
    return render_template('balance_sheet_category.html', categories=categories)

@app.route('/balance_sheet_category/delete', methods=['POST'])
@login_required
@has_permission('Access_Accounting')
def delete_balance_sheet_category():
    selected_ids = request.form.getlist('selected_ids')
    if selected_ids:
        try:
            placeholders = ', '.join(['%s'] * len(selected_ids))
            query = f"DELETE FROM balance_sheet_category WHERE id IN ({placeholders})"
            db.execute_query(query, tuple(selected_ids), commit=True)
            clear_category_cache()
            flash(f'{len(selected_ids)} categories deleted', 'success')
        except Exception as e:
            flash(f'Error deleting categories: {str(e)}', 'danger')
    else:
        flash('No items selected', 'info')

    return redirect(url_for('balance_sheet_category'))

# --- Create Bank Account ---
@app.route('/pl_category', methods=['GET', 'POST'])
@login_required
@has_permission('Access_Accounting')
def pl_category():
    if request.method == 'POST':
        category_id = int(request.form.get('category_id', 0))
        name = request.form.get('category_name')
        level = request.form.get('holding_level')

        current_user = get_current_user_id()
        current_user_pk = get_current_user_pk()
        current_date = date.today()

        if not name or not level:
            flash('Name and Level are required', 'danger')
            return redirect(url_for('pl_category'))

        try:
            if category_id == 0:
                # Insert
                query = """
                    INSERT INTO `p&l_category`
                    (id, name_of_category, holding_position, create_date_time, create_user_code)
                    VALUES (0, %s, %s, %s, %s)
                """
                db.execute_query(query, (name, level, current_date, current_user_pk), commit=True)
                clear_category_cache()
                flash('P&L Category created successfully', 'success')
            else:
                # Update
                query = """
                    UPDATE `p&l_category`
                    SET name_of_category = %s, holding_position = %s,
                        create_date_time = %s, create_user_code = %s
                    WHERE id = %s
                """
                db.execute_query(query, (name, level, current_date, current_user_pk, category_id), commit=True)
                clear_category_cache()
                flash('P&L Category updated successfully', 'success')

        except Exception as e:
            flash(f'Error saving category: {str(e)}', 'danger')

        return redirect(url_for('pl_category'))

    # GET
    categories = db.execute_query("SELECT * FROM `p&l_category` ORDER BY holding_position")
    return render_template('pl_category.html', categories=categories)

@app.route('/pl_category/delete', methods=['POST'])
@login_required
@has_permission('Access_Accounting')
def delete_pl_category():
    selected_ids = request.form.getlist('selected_ids')
    if selected_ids:
        try:
            placeholders = ', '.join(['%s'] * len(selected_ids))
            query = f"DELETE FROM `p&l_category` WHERE id IN ({placeholders})"
            db.execute_query(query, tuple(selected_ids), commit=True)
            clear_category_cache()
            flash(f'{len(selected_ids)} categories deleted', 'success')
        except Exception as e:
            flash(f'Error deleting categories: {str(e)}', 'danger')
    else:
        flash('No items selected', 'info')

    return redirect(url_for('pl_category'))

@app.route('/pl_category_correction', methods=['GET', 'POST'])
@login_required
@has_permission('Access_Accounting')
def pl_category_correction():
    if request.method == 'POST':
        account_ids = request.form.getlist('account_id[]')
        selections = request.form.getlist('category_selection[]')

        updates = []
        for i in range(len(account_ids)):
            if i < len(selections) and selections[i]:
                cat_name, hold_pos = selections[i].split(',')
                updates.append((cat_name, hold_pos, account_ids[i]))

        if updates:
            try:
                conn = db.get_connection()
                cursor = conn.cursor()
                conn.start_transaction()

                query = "UPDATE new_account_table SET account_name_of_catogory_PL = %s, account_hold_possion_PL = %s WHERE id = %s"
                cursor.executemany(query, updates)

                conn.commit()
                flash(f'Updated {len(updates)} accounts successfully', 'success')
            except Exception as e:
                flash(f'Error updating accounts: {str(e)}', 'danger')

        return redirect(url_for('pl_category_correction'))

    # GET: Fetch unassigned P&L accounts (Logic from wpf_catigiry_corections)
    # The C# code fetches accounts where (Income=1 OR Expense=1) AND Category IS NULL
    query_acc = """
        SELECT id, account_name, account_name_of_catogory_PL
        FROM new_account_table
        WHERE (account_income = 1 OR account_expenses = 1)
        AND (account_name_of_catogory_PL IS NULL OR account_name_of_catogory_PL = '')
    """
    accounts = db.execute_query(query_acc)

    # Fetch Categories
    query_cat = "SELECT name_of_category, holding_position FROM `p&l_category` ORDER BY holding_position"
    categories = db.execute_query(query_cat)

    return render_template('pl_category_correction.html', accounts=accounts, categories=categories)

@app.route('/bs_category_correction', methods=['GET', 'POST'])
@login_required
@has_permission('Access_Accounting')
def bs_category_correction():
    if request.method == 'POST':
        account_ids = request.form.getlist('account_id[]')
        selections = request.form.getlist('category_selection[]')

        updates = []
        for i in range(len(account_ids)):
            if i < len(selections) and selections[i]:
                cat_name, hold_pos = selections[i].split(',')
                updates.append((cat_name, hold_pos, account_ids[i]))

        if updates:
            try:
                conn = db.get_connection()
                cursor = conn.cursor()
                conn.start_transaction()

                query = "UPDATE new_account_table SET account_name_of_catogory_Balace_sheet = %s, account_hold_possion_Balace_Sheet = %s WHERE id = %s"
                cursor.executemany(query, updates)

                conn.commit()
                flash(f'Updated {len(updates)} accounts successfully', 'success')
            except Exception as e:
                flash(f'Error updating accounts: {str(e)}', 'danger')

        return redirect(url_for('bs_category_correction'))

    # GET: Fetch unassigned BS accounts (Logic from wpf_bs_catogory_corections)
    query_acc = """
        SELECT id, account_name, account_name_of_catogory_Balace_sheet
        FROM new_account_table
        WHERE (account_assets = 1 OR account_liabilities = 1 OR account_equity = 1)
        AND (account_name_of_catogory_Balace_sheet IS NULL OR account_name_of_catogory_Balace_sheet = '')
    """
    accounts = db.execute_query(query_acc)

    # Fetch Categories
    query_cat = "SELECT name_of_category, holding_position FROM balance_sheet_category ORDER BY holding_position"
    categories = db.execute_query(query_cat)

    return render_template('bs_category_correction.html', accounts=accounts, categories=categories)

@app.route('/create_bank_account', methods=['GET', 'POST'])
@login_required
@has_permission('Access_Accounting')
def create_bank_account():
    if request.method == 'POST':
        acc_no = request.form.get('account_number')
        bank_name = request.form.get('bank_name')

        if not acc_no or not bank_name:
            flash('Account number and Bank Name are required', 'danger')
            return redirect(url_for('create_bank_account'))

        current_user = get_current_user_id()
        current_user_pk = get_current_user_pk()
        today_date = date.today()

        conn = db.get_connection()
        cursor = conn.cursor()

        try:
            conn.start_transaction()

            # 1. Check/Create GL Account (Using Account Number as Name)
            cursor.execute("SELECT id FROM new_account_table WHERE account_name = %s", (acc_no,))
            if not cursor.fetchone():
                # Find 'Current assets' or 'Cash & Bank'
                cursor.execute("SELECT holding_position, name_of_category FROM balance_sheet_category WHERE name_of_category LIKE '%Bank%' OR name_of_category LIKE '%Cash%' LIMIT 1")
                res = cursor.fetchone()

                if res:
                    bs_pos = res[0]
                    bs_cat_name = res[1]
                else:
                    # Fallback to the first available category if no match
                    cursor.execute("SELECT holding_position, name_of_category FROM balance_sheet_category LIMIT 1")
                    fallback_res = cursor.fetchone()
                    if fallback_res:
                        bs_pos = fallback_res[0]
                        bs_cat_name = fallback_res[1]
                    else:
                        # Extreme fallback: Create 'Current assets' if table is empty
                        bs_pos = 3
                        bs_cat_name = 'Current assets'
                        cursor.execute("INSERT IGNORE INTO balance_sheet_category (name_of_category, holding_position) VALUES (%s, %s)", (bs_cat_name, bs_pos))

                cursor.execute("""
                    INSERT INTO new_account_table (
                        account_name, account_hold_possion_Balace_Sheet, account_name_of_catogory_Balace_sheet,
                        account_assets, account_basment, accont_create_date, account_create_user, account_active,
                        currency_code
                    ) VALUES (%s, %s, %s, 1, 'DR', %s, %s, 1, 'LKR')
                """, (acc_no, bs_pos, bs_cat_name, today_date, current_user))

            # 2. Insert into Bank Book
            cursor.execute("""
                INSERT INTO bank_book (bank_bookcol_account_number, bank_book_bank_name, bank_book_create_date, bank_book_create_user)
                VALUES (%s, %s, %s, %s)
            """, (acc_no, bank_name, today_date, current_user_pk))

            conn.commit()
            flash('New bank account created', 'success')

        except Exception as e:
            conn.rollback()
            flash(f'Error creating bank account: {str(e)}', 'danger')
        finally:
            cursor.close()
            conn.close()

        return redirect(url_for('create_bank_account'))

    # Fetch existing bank accounts
    bank_accounts = db.execute_query("SELECT * FROM bank_book ORDER BY bank_book_create_date DESC")
    return render_template('create_bank_account.html', bank_accounts=bank_accounts)

# --- Create Cash Account ---
@app.route('/create_cash_account', methods=['GET', 'POST'])
@login_required
@has_permission('Access_Accounting')
def create_cash_account():
    if request.method == 'POST':
        acc_name = request.form.get('account_name')

        if not acc_name:
            flash('Account name is required', 'danger')
            return redirect(url_for('create_cash_account'))

        current_user = get_current_user_id()
        current_user_pk = get_current_user_pk()
        today_date = date.today()

        conn = db.get_connection()
        cursor = conn.cursor()

        try:
            conn.start_transaction()

            # 1. Check if GL Account exists
            cursor.execute("SELECT id FROM new_account_table WHERE account_name = %s", (acc_name,))
            if not cursor.fetchone():
                # Create GL Account (Current Asset)
                # Need to find 'Current assets' category position
                cursor.execute("SELECT holding_position, name_of_category FROM balance_sheet_category WHERE name_of_category LIKE '%Current asset%' LIMIT 1")
                res = cursor.fetchone()

                if res:
                    bs_pos = res[0]
                    bs_cat_name = res[1]
                else:
                    # Fallback to the first available category if no match
                    cursor.execute("SELECT holding_position, name_of_category FROM balance_sheet_category LIMIT 1")
                    fallback_res = cursor.fetchone()
                    if fallback_res:
                        bs_pos = fallback_res[0]
                        bs_cat_name = fallback_res[1]
                    else:
                        # Extreme fallback: Create 'Current assets' if table is empty
                        bs_pos = 3
                        bs_cat_name = 'Current assets'
                        cursor.execute("INSERT IGNORE INTO balance_sheet_category (name_of_category, holding_position) VALUES (%s, %s)", (bs_cat_name, bs_pos))

                cursor.execute("""
                    INSERT INTO new_account_table (
                        account_name, account_hold_possion_Balace_Sheet, account_name_of_catogory_Balace_sheet,
                        account_assets, account_basment, accont_create_date, account_create_user, account_active,
                        currency_code
                    ) VALUES (%s, %s, %s, 1, 'DR', %s, %s, 1, 'LKR')
                """, (acc_name, bs_pos, bs_cat_name, today_date, current_user))

            # 2. Insert into Cash Book
            # cash_book schema: cash_id, cash_book_account_name, cash_creat_date, cash_created_user, Select_As
            cursor.execute("""
                INSERT INTO cash_book (cash_book_account_name, cash_creat_date, cash_created_user, Select_As)
                VALUES (%s, %s, %s, 0)
            """, (acc_name, today_date, current_user_pk))

            conn.commit()
            flash('New cash account created', 'success')

        except Exception as e:
            conn.rollback()
            flash(f'Error creating cash account: {str(e)}', 'danger')
        finally:
            cursor.close()
            conn.close()

        return redirect(url_for('create_cash_account'))

    # Fetch existing cash accounts
    cash_accounts = db.execute_query("SELECT * FROM cash_book ORDER BY cash_creat_date DESC")
    return render_template('create_cash_account.html', cash_accounts=cash_accounts)


@app.route('/edit_bank_account', methods=['POST'])
@login_required
@has_permission('Access_Accounting')
def edit_bank_account():
    """Edit a bank account's name / number. If the number changes, the rename
    is cascaded to the linked GL account and transaction references."""
    bank_id = request.form.get('id')
    new_name = (request.form.get('bank_name') or '').strip()
    new_number = (request.form.get('account_number') or '').strip()

    if not bank_id or not new_name or not new_number:
        flash('Bank name and account number are required.', 'danger')
        return redirect(url_for('create_bank_account'))

    conn = None
    cursor = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        conn.start_transaction()

        # Current number (the linking key)
        cursor.execute("SELECT bank_bookcol_account_number FROM bank_book WHERE id=%s", (bank_id,))
        row = cursor.fetchone()
        if not row:
            flash('Bank account not found.', 'danger')
            return redirect(url_for('create_bank_account'))
        old_number = row[0]

        # Update the bank_book row
        cursor.execute(
            "UPDATE bank_book SET bank_book_bank_name=%s, bank_bookcol_account_number=%s WHERE id=%s",
            (new_name, new_number, bank_id))

        # Cascade the number rename to linked records (each guarded)
        if old_number and old_number != new_number:
            for sql in [
                ("UPDATE new_account_table SET account_name=%s WHERE account_name=%s", (new_number, old_number)),
                ("UPDATE bank_book_recod SET bank_book__accont_name=%s WHERE bank_book__accont_name=%s", (new_number, old_number)),
                ("UPDATE bank_book_voucher_no SET bank_book_voucher_link=%s WHERE bank_book_voucher_link=%s", (new_number, old_number)),
            ]:
                try:
                    cursor.execute(sql[0], sql[1])
                except Exception as ce:
                    logging.warning(f"bank cascade skip: {ce}")

        conn.commit()
        flash('Bank account updated successfully.', 'success')
    except Exception as e:
        if conn: conn.rollback()
        flash(f'Error updating bank account: {str(e)}', 'danger')
    finally:
        if cursor: cursor.close()
        if conn: conn.close()
    return redirect(url_for('create_bank_account'))


@app.route('/edit_cash_account', methods=['POST'])
@login_required
@has_permission('Access_Accounting')
def edit_cash_account():
    """Edit a cash account name. Cascades the rename to the linked GL account
    and transaction references."""
    cash_id = request.form.get('id')
    new_name = (request.form.get('account_name') or '').strip()

    if not cash_id or not new_name:
        flash('Account name is required.', 'danger')
        return redirect(url_for('create_cash_account'))

    conn = None
    cursor = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        conn.start_transaction()

        cursor.execute("SELECT cash_book_account_name FROM cash_book WHERE cash_id=%s", (cash_id,))
        row = cursor.fetchone()
        if not row:
            flash('Cash account not found.', 'danger')
            return redirect(url_for('create_cash_account'))
        old_name = row[0]

        cursor.execute(
            "UPDATE cash_book SET cash_book_account_name=%s WHERE cash_id=%s",
            (new_name, cash_id))

        if old_name and old_name != new_name:
            for sql in [
                ("UPDATE new_account_table SET account_name=%s WHERE account_name=%s", (new_name, old_name)),
                ("UPDATE cash_book_recode SET cash_book_recode_accont_name=%s WHERE cash_book_recode_accont_name=%s", (new_name, old_name)),
                ("UPDATE cash_voucher_no SET cash_voucher_link=%s WHERE cash_voucher_link=%s", (new_name, old_name)),
            ]:
                try:
                    cursor.execute(sql[0], sql[1])
                except Exception as ce:
                    logging.warning(f"cash cascade skip: {ce}")

        conn.commit()
        flash('Cash account updated successfully.', 'success')
    except Exception as e:
        if conn: conn.rollback()
        flash(f'Error updating cash account: {str(e)}', 'danger')
    finally:
        if cursor: cursor.close()
        if conn: conn.close()
    return redirect(url_for('create_cash_account'))

# --- Control Panel (P&L Correction + Settings) ---
@app.route('/control_panel', methods=['GET', 'POST'])
@login_required
@has_permission('Access_Accounting')
def control_panel():
    # 1. Handle Settings (Warranty & Approval)
    if request.method == 'POST':
        # Warranty & Settings
        if 'warranty_enabled' in request.form or 'approval_enabled' in request.form or 'system_theme' in request.form or 'invoice_terms' in request.form:
            # Warranty Logic
            try:
                warranty_enabled = 1 if request.form.get('warranty_enabled') else 0
                count_res = db.execute_query("SELECT COUNT(*) as cnt FROM adding_new")
                if count_res and count_res[0]['cnt'] == 0:
                    # MySQL schema for adding_new has a literal string 'null' as default for ac1, ac2, etc.
                    # which violates FK constraints on new_account_table if 'null' account doesn't exist.
                    # We must explicitly insert Python None (SQL NULL) to bypass the default string.
                    db.execute_query("INSERT INTO adding_new (id, yes, ac1, ac2, ac3, ac4, ac5) VALUES (0, %s, %s, %s, %s, %s, %s)",
                                     (warranty_enabled, None, None, None, None, None), commit=True)
                else:
                    db.execute_query("UPDATE adding_new SET yes = %s", (warranty_enabled,), commit=True)
            except Exception as e:
                print(f"Error updating warranty settings: {e}")

            # Approval Workflow Logic
            approval_enabled = 1 if request.form.get('approval_enabled') else 0
            # Check if setting exists
            check = db.execute_query("SELECT id FROM system_settings WHERE setting_key = 'enable_approval_workflow'")
            if not check:
                db.execute_query("INSERT INTO system_settings (setting_key, setting_value, description) VALUES ('enable_approval_workflow', %s, 'Enable Park & Post Workflow')", (str(approval_enabled),), commit=True)
            else:
                db.execute_query("UPDATE system_settings SET setting_value = %s WHERE setting_key = 'enable_approval_workflow'", (str(approval_enabled),), commit=True)

            # Theme Logic
            new_theme = request.form.get('system_theme')
            if new_theme and new_theme in THEMES:
                check_theme = db.execute_query("SELECT id FROM system_settings WHERE setting_key = 'system_theme'")
                if not check_theme:
                    db.execute_query("INSERT INTO system_settings (setting_key, setting_value, description) VALUES ('system_theme', %s, 'Active System Theme')", (new_theme,), commit=True)
                else:
                    db.execute_query("UPDATE system_settings SET setting_value = %s WHERE setting_key = 'system_theme'", (new_theme,), commit=True)

            # Invoice Terms
            if 'invoice_terms' in request.form:
                try:
                    new_terms = request.form.get('invoice_terms', '')
                    check_terms = db.execute_query("SELECT id FROM system_settings WHERE setting_key = 'invoice_terms_conditions'")
                    if not check_terms:
                        db.execute_query("INSERT INTO system_settings (setting_key, setting_value, description) VALUES ('invoice_terms_conditions', %s, 'Terms and Conditions displayed on Invoices')", (new_terms,), commit=True)
                    else:
                        db.execute_query("UPDATE system_settings SET setting_value = %s WHERE setting_key = 'invoice_terms_conditions'", (new_terms,), commit=True)
                except Exception as e:
                    print(f"Error updating invoice_terms: {e}")

            flash('Settings updated', 'success')
            return redirect(url_for('control_panel'))

    # 2. Fetch Status
    warranty_enabled = False
    try:
        res = db.execute_query("SELECT yes FROM adding_new")
        if res and isinstance(res, list) and len(res) > 0 and res[0].get('yes') == 1:
            warranty_enabled = True
    except Exception as e:
        print(f"Control panel error (warranty): {e}")

    approval_enabled = False
    try:
        res_app = db.execute_query("SELECT setting_value FROM system_settings WHERE setting_key = 'enable_approval_workflow'")
        if res_app and isinstance(res_app, list) and len(res_app) > 0 and res_app[0].get('setting_value') == '1':
            approval_enabled = True
    except Exception as e:
        print(f"Control panel error (approval): {e}")

    current_theme_key = 'default'
    try:
        res_theme = db.execute_query("SELECT setting_value FROM system_settings WHERE setting_key = 'system_theme'")
        if res_theme and isinstance(res_theme, list) and len(res_theme) > 0 and res_theme[0].get('setting_value'):
            current_theme_key = res_theme[0]['setting_value']
    except Exception as e:
        print(f"Control panel error (theme): {e}")

    invoice_terms = ""
    try:
        res_terms = db.execute_query("SELECT setting_value FROM system_settings WHERE setting_key = 'invoice_terms_conditions'")
        if res_terms and isinstance(res_terms, list) and len(res_terms) > 0 and 'setting_value' in res_terms[0]:
            invoice_terms = res_terms[0]['setting_value'] or ""
    except Exception as e:
        print(f"Error fetching invoice_terms: {e}")

    # 3. Fetch Unassigned P&L Accounts (Income or Expense but no P&L Category)
    unassigned_pl = db.execute_query("""
        SELECT id, account_name
        FROM new_account_table
        WHERE (account_income = 1 OR account_expenses = 1)
        AND (account_name_of_catogory_PL IS NULL OR account_name_of_catogory_PL = '')
    """)

    # 4. Fetch Unassigned Balance Sheet Accounts
    unassigned_bs = db.execute_query("""
        SELECT id, account_name
        FROM new_account_table
        WHERE (account_assets = 1 OR account_liabilities = 1 OR account_equity = 1)
        AND (account_name_of_catogory_Balace_sheet IS NULL OR account_name_of_catogory_Balace_sheet = '')
    """)

    # 5. Fetch Categories for Dropdown
    bs_cats, pl_cats, _ = get_cached_categories(db)

    return render_template('control_panel.html',
                           warranty_enabled=warranty_enabled,
                           approval_enabled=approval_enabled,
                           current_theme_key=current_theme_key,
                           themes=THEMES,
                           unassigned_pl=unassigned_pl,
                           unassigned_bs=unassigned_bs,
                           pl_categories=pl_cats,
                           bs_categories=bs_cats,
                           invoice_terms=invoice_terms)

@app.route('/control_panel/update', methods=['POST'])
@login_required
def control_panel_update():
    update_type = request.form.get('update_type')
    updates = []

    # Identify fields based on update type
    if update_type == 'pl':
        prefix = 'category_'
        sql = "UPDATE new_account_table SET account_name_of_catogory_PL = %s, account_hold_possion_PL = %s WHERE id = %s"
    elif update_type == 'bs':
        prefix = 'bscategory_'
        sql = "UPDATE new_account_table SET account_name_of_catogory_Balace_sheet = %s, account_hold_possion_Balace_Sheet = %s WHERE id = %s"
    else:
        flash('Invalid update type', 'danger')
        return redirect(url_for('control_panel'))

    for key, value in request.form.items():
        if key.startswith(prefix) and value:
            acc_id = key.split('_')[1]
            cat_name, hold_pos = value.split(',')
            updates.append((cat_name, hold_pos, acc_id))

    if updates:
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            conn.start_transaction()

            cursor.executemany(sql, updates)

            conn.commit()
            cursor.close()
            conn.close()
            flash(f'Updated {len(updates)} accounts.', 'success')
        except Exception as e:
            flash(f'Error updating accounts: {str(e)}', 'danger')
    else:
        flash('No changes selected.', 'info')

    return redirect(url_for('control_panel'))

# --- Approvals Dashboard ---
@app.route('/approvals', methods=['GET'])
@login_required
@has_permission('OP_Approved') # Assuming this permission covers all approvals for now
def approvals():
    # 1. Pending Purchase Orders (Status = 0)
    pending_pos = db.execute_query("""
        SELECT id, OP_NO_Other as ref_no, Create_Date as date, Sup_Name as party,
               'Purchase Order' as type,
               (SELECT SUM(QTY*Unit_price) FROM PO_Recode_Details WHERE Link_OP_NO_Table=OP_NO_Table.id) as amount
        FROM OP_NO_Table
        WHERE status = 0 AND Delete_PO = 0
    """)

    # 2. Pending JVs (Manual, Payments, Receipts) (Status = 0)
    pending_jvs = db.execute_query("""
        SELECT j.jv_id as id, j.jv_user_code as ref_no, e.entry_effective_date as date,
               'Journal Voucher' as type, j.jv_naration as narration,
               SUM(e.enty_values_DR) as amount
        FROM jv_numbers j
        LEFT JOIN entry_details e ON j.jv_id = e.entry_jv
        WHERE j.status = 0
        GROUP BY j.jv_id, j.jv_user_code, e.entry_effective_date, j.jv_naration
    """)

    # Combine lists
    items = []
    for po in pending_pos:
        items.append({
            'id': po['id'],
            'ref_no': po['ref_no'],
            'date': str(po['date']),
            'party': po['party'],
            'type': 'Purchase Order',
            'amount': float(po['amount'] or 0),
            'source': 'po'
        })

    for jv in pending_jvs:
        items.append({
            'id': jv['id'],
            'ref_no': f"JV-{jv['id']}", # JV User Code might be user ID, using ID for ref
            'date': str(jv['date']),
            'party': jv['narration'], # Use narration as description/party
            'type': 'Journal/Payment',
            'amount': float(jv['amount'] or 0),
            'source': 'jv'
        })

    return render_template('approvals.html', items=items)

@app.route('/approvals/action', methods=['POST'])
@login_required
@has_permission('OP_Approved')
def approval_action():
    item_id = request.form.get('id')
    source = request.form.get('source')
    action = request.form.get('action') # 'approve' or 'reject'

    current_user = get_current_user_id()
    current_user_pk = get_current_user_pk()
    new_status = 1 if action == 'approve' else 2

    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    try:
        if source == 'po':
            db.execute_query("UPDATE OP_NO_Table SET status = %s, Aprove_By = %s, Aproed_Date = %s WHERE id = %s",
                             (new_status, current_user_pk, date.today(), item_id), commit=True)

        elif source == 'jv':
            db.execute_query("UPDATE jv_numbers SET status = %s WHERE jv_id = %s",
                             (new_status, item_id), commit=True)

        if is_ajax:
            return jsonify({'success': True, 'message': f'Item {action}d'})
        flash(f'Item {action}d successfully', 'success')
    except Exception as e:
        if is_ajax:
            return jsonify({'success': False, 'error': str(e)}), 500
        flash(f'Error: {str(e)}', 'danger')

    return redirect(url_for('approvals'))

# --- Bulk Upload Helpers ---
def read_csv_content(file_storage):
    """
    Reads a FileStorage object (CSV), attempts various encodings,
    and returns the decoded string content.
    """
    if not file_storage or file_storage.filename == '':
        raise ValueError("No file provided")

    try:
        file_bytes = file_storage.stream.read()
        decoded_str = None

        # prioritized list of encodings
        encodings = ['utf-8-sig', 'utf-16', 'utf-8', 'cp1252', 'latin1']

        for encoding in encodings:
            try:
                decoded_str = file_bytes.decode(encoding)
                break
            except UnicodeDecodeError:
                continue

        if decoded_str is None:
            raise ValueError(f"Unable to determine file encoding (tried {', '.join(encodings)})")

        return decoded_str

    except Exception as e:
        raise ValueError(f"Error reading file: {str(e)}")

def parse_gl_upload_data(file_storage):
    """
    Parses the GL Upload CSV file and returns a list of dictionaries representing the rows.
    """
    try:
        csv_content = read_csv_content(file_storage)
        stream = io.StringIO(csv_content, newline=None)
        csv_input = csv.DictReader(stream)

        rows = []
        for row in csv_input:
            # Clean keys/values
            clean_row = {k.strip(): v.strip() for k, v in row.items() if k}
            if not clean_row.get('Account Name'): continue
            rows.append(clean_row)

        return rows
    except Exception as e:
        raise ValueError(f"Error parsing CSV data: {str(e)}")

def _parse_gl_category(cat_val):
    cat_name = None
    cat_pos = None
    is_bs = False
    is_pl = False

    if cat_val:
        parts = cat_val.split('|')
        if len(parts) == 2:
            cat_data, cat_type = parts
            cat_name, cat_pos = cat_data.split(',')
            if cat_type == 'BS': is_bs = True
            elif cat_type == 'PL': is_pl = True

    return cat_name, cat_pos, is_bs, is_pl

def _process_bulk_gl_subledgers(cursor, potential_banks, potential_cash, today, current_user):
    # Batch Sub-Ledger (Bank)
    if potential_banks:
        format_strings = ','.join(['%s'] * len(potential_banks))
        cursor.execute(f"SELECT bank_bookcol_account_number FROM bank_book WHERE bank_bookcol_account_number IN ({format_strings})", tuple(potential_banks))
        existing_banks = {row[0] for row in cursor.fetchall()}

        banks_to_insert = []
        for b_name in potential_banks:
            if b_name not in existing_banks:
                banks_to_insert.append((b_name, b_name, today, current_user))

        if banks_to_insert:
            cursor.executemany("""
                INSERT INTO bank_book (bank_bookcol_account_number, bank_book_bank_name, bank_book_create_date, bank_book_create_user)
                VALUES (%s, %s, %s, %s)
            """, banks_to_insert)

    # Batch Sub-Ledger (Cash)
    if potential_cash:
        format_strings = ','.join(['%s'] * len(potential_cash))
        cursor.execute(f"SELECT cash_book_account_name FROM cash_book WHERE cash_book_account_name IN ({format_strings})", tuple(potential_cash))
        existing_cash = {row[0] for row in cursor.fetchall()}

        cash_to_insert = []
        for c_name in potential_cash:
            if c_name not in existing_cash:
                cash_to_insert.append((c_name, today, current_user))

        if cash_to_insert:
            cursor.executemany("""
                INSERT INTO cash_book (cash_book_account_name, cash_creat_date, cash_created_user, Select_As)
                VALUES (%s, %s, %s, 0)
            """, cash_to_insert)

def save_bulk_gl_accounts(form_data, current_user):
    """
    Handles the database insertion logic for bulk GL account upload.
    Returns the count of successfully processed accounts.
    """
    conn = db.get_connection()
    cursor = conn.cursor()
    conn.start_transaction()

    try:
        today = date.today()

        names = form_data.getlist('account_name[]')
        types = form_data.getlist('account_type[]')
        cats = form_data.getlist('category[]')
        cfs = form_data.getlist('cf_category[]')
        actions = form_data.getlist('action[]')

        # Filter valid names to check
        valid_names = [n for i, n in enumerate(names) if actions[i] != 'skip' and n]

        # Batch Fetch Existing Accounts
        existing_map = {}
        if valid_names:
            format_strings = ','.join(['%s'] * len(valid_names))
            cursor.execute(f"SELECT id, account_name FROM new_account_table WHERE account_name IN ({format_strings})", tuple(valid_names))
            existing_rows = cursor.fetchall()
            for row in existing_rows:
                existing_map[row[1]] = row[0] # Name -> ID

        to_update = []
        to_insert = []

        # For Sub-Ledgers
        potential_banks = []
        potential_cash = []

        count = 0
        for i in range(len(names)):
            if actions[i] == 'skip': continue

            name = names[i]
            acc_type = types[i]
            cat_val = cats[i]
            cf = cfs[i]

            # Parse Category
            cat_name, cat_pos, is_bs, is_pl = _parse_gl_category(cat_val)

            is_inc = 1 if acc_type == 'Income' else 0
            is_exp = 1 if acc_type == 'Expense' else 0
            is_ast = 1 if acc_type == 'Asset' else 0
            is_lia = 1 if acc_type == 'Liability' else 0
            is_equ = 1 if acc_type == 'Equity' else 0

            basement = 'DR' if is_ast or is_exp else 'CR'

            if name in existing_map:
                # Prepare Update
                to_update.append((
                    cat_pos if is_pl else None, cat_pos if is_bs else None,
                    cat_name if is_pl else None, cat_name if is_bs else None,
                    is_inc, is_exp, is_ast, is_lia, is_equ,
                    cf, basement, existing_map[name]
                ))
            else:
                # Prepare Insert
                to_insert.append((
                    name, cat_pos if is_pl else None, cat_pos if is_bs else None,
                    cat_name if is_pl else None, cat_name if is_bs else None,
                    is_inc, is_exp, is_ast, is_lia, is_equ,
                    cf, today, current_user, basement
                ))

                # Collect for Sub-Ledger check
                if is_ast:
                    acc_name_lower = name.lower()
                    if 'bank' in acc_name_lower:
                        potential_banks.append(name)
                    elif 'cash' in acc_name_lower:
                        potential_cash.append(name)

            count += 1

        # Process Updates row-by-row to skip failures
        if to_update:
            for row in to_update:
                try:
                    cursor.execute("""
                        UPDATE new_account_table SET
                            account_hold_possion_PL=%s, account_hold_possion_Balace_Sheet=%s,
                            account_name_of_catogory_PL=%s, account_name_of_catogory_Balace_sheet=%s,
                            account_income=%s, account_expenses=%s, account_assets=%s, account_liabilities=%s, account_equity=%s,
                            cf_catogory=%s, account_basment=%s
                        WHERE id=%s
                    """, row)
                except Exception as e:
                    pass

        # Process Inserts row-by-row to skip failures
        if to_insert:
            for row in to_insert:
                try:
                    cursor.execute("""
                        INSERT INTO new_account_table (
                            account_name, account_hold_possion_PL, account_hold_possion_Balace_Sheet,
                            account_name_of_catogory_PL, account_name_of_catogory_Balace_sheet,
                            account_income, account_expenses, account_assets, account_liabilities, account_equity,
                            cf_catogory, accont_create_date, account_create_user, account_active, account_basment, currency_code
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 1, %s, 'LKR')
                    """, row)
                except Exception as e:
                    count -= 1
                    pass

        _process_bulk_gl_subledgers(cursor, potential_banks, potential_cash, today, current_user)

        conn.commit()
        return count

    except Exception as e:
        conn.rollback()
        raise e
    finally:
        cursor.close()
        conn.close()

# --- Bulk Upload Module ---
@app.route('/bulk_upload_gl', methods=['GET', 'POST'])
@login_required
@has_permission('Access_Accounting')
def bulk_upload_gl():
    if request.method == 'POST':
        # Step 2: Process Uploaded File
        if 'file' in request.files:
            file = request.files['file']
            if file.filename == '':
                flash('No file selected', 'danger')
                return redirect(url_for('bulk_upload_gl'))

            try:
                rows = parse_gl_upload_data(file)
                # Use helper to parse and validate
                raw_rows = parse_csv_file(file, required_columns=['Account Name'])

                rows = []
                for row in raw_rows:
                    if not row.get('Account Name'): continue
                    rows.append(row)

                # Fetch Existing Data for Validation/Dropdowns
                existing_accounts = {a['account_name']: a for a in db.execute_query("SELECT account_name, account_basment FROM new_account_table")}
                bs_cats, pl_cats, cf_cats = get_cached_categories(db)

                return render_template('bulk_upload_review.html',
                                       rows=rows,
                                       existing=existing_accounts,
                                       bs_cats=bs_cats,
                                       pl_cats=pl_cats,
                                       cf_cats=cf_cats)

            except Exception as e:
                flash(f'Error processing file: {str(e)}', 'danger')
                return redirect(url_for('bulk_upload_gl'))

        # Step 3: Save Data
        elif 'save_data' in request.form:
            try:
                current_user = get_current_user_id()
                count = save_bulk_gl_accounts(request.form, current_user)
                flash(f'Successfully processed {count} accounts.', 'success')
                return redirect(url_for('chart_of_accounts'))

            except Exception as e:
                flash(f'Transaction failed: {str(e)}', 'danger')
                return redirect(url_for('bulk_upload_gl'))

    return render_template('bulk_upload_gl.html')


@app.route('/bank_statement_analysis', methods=['POST'])
@login_required
@has_permission('Access_Accounting')
def bank_statement_analysis():
    # If the browser already OCR'd the image (Tesseract.js runs client-side
    # in bulk_upload_gl.html), use that text directly — no server-side OCR
    # engine needed at all for this path, and the image is never handed to
    # this server (or any third party) to read.
    browser_text = (request.form.get('ocr_text') or '').strip()

    if browser_text:
        text = browser_text
    else:
        if 'file' not in request.files:
            flash('No file uploaded', 'danger')
            return redirect(url_for('bulk_upload_gl'))

        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'danger')
            return redirect(url_for('bulk_upload_gl'))

        raw_data = file.read()

        if file.filename.lower().endswith('.pdf'):
            # Text-based PDF — read directly via PyPDF2, no OCR engine needed
            # at all (this is the "upload a PDF instead" tip in the OCR
            # message: a text PDF has its content extracted straight away).
            try:
                import PyPDF2
                reader = PyPDF2.PdfReader(io.BytesIO(raw_data))
                text = ''
                for page in reader.pages:
                    t = page.extract_text()
                    if t:
                        text += t + '\n'
            except Exception as e:
                flash(f'Could not read this PDF: {str(e)}', 'danger')
                return redirect(url_for('bulk_upload_gl'))
            if not text.strip():
                flash('No selectable text found in this PDF — it is likely a scanned image saved as a PDF, '
                      'not a text PDF. Try exporting/photographing the page as a PNG or JPG image and '
                      'uploading that instead (it will be read via OCR).', 'warning')
                return redirect(url_for('bulk_upload_gl'))
        else:
            try:
                text = ocr_image_bytes(raw_data)
            except RuntimeError:
                flash(_TESSERACT_HELP, 'danger')
                return redirect(url_for('bulk_upload_gl'))
            except ImportError:
                flash("Error processing image: 'Pillow' is not installed. Install it with 'pip install Pillow'.", "danger")
                return redirect(url_for('bulk_upload_gl'))
            except Exception as e:
                logging.error(f"OCR Error: {e}")
                if 'tesseract' in str(e).lower() or 'not found' in str(e).lower():
                    flash(_TESSERACT_HELP, 'danger')
                else:
                    flash(f'Error processing image: {str(e)}', 'danger')
                return redirect(url_for('bulk_upload_gl'))

    try:
        # Parse extracted text for transactions (Date, Description, Amount).
        #
        # Real bank statements (e.g. Sampath Bank / Sri Lanka style) print one
        # line per transaction as:
        #   03Aug SAMPATH TFRC/CARD5009519   20,016.26   933,240.70Cr
        #   <day><3-letter month>  <particulars>  <amount>  <running balance><Cr|Dr>
        # and print opening/closing balance lines with the year attached and
        # only ONE amount (no separate transaction amount):
        #   03Aug2026 B/F 913,224.44Cr
        #   31Aug2026 C/F 453,124.28Cr
        # There is no "deposit"/"withdrawal" column and no reliable keyword to
        # tell receipts from payments — instead we track the running balance
        # and infer the type from whether it went up (Receipt) or down
        # (Payment) versus the previous line.
        month_map = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
                     'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12}

        # Balance/brought-forward lines: "<day><Mon><yyyy> B/F|C/F <balance><Cr|Dr>"
        bf_cf_re = re.compile(
            r'^(\d{1,2})([A-Za-z]{3})(\d{4})\s+(?:B/F|C/F)\s+([\d,]+\.\d{2})\s*(Cr|Dr)\s*$',
            re.IGNORECASE)
        # Normal transaction lines: "<day><Mon> <particulars> <amount> <balance><Cr|Dr>"
        txn_re = re.compile(
            r'^(\d{1,2})([A-Za-z]{3})\s+(.+?)\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s*(Cr|Dr)\s*$',
            re.IGNORECASE)
        # Fallback default year, e.g. from a "...From 01/08/2026 To 31/08/2026" header.
        year_hdr = re.search(r'From\s+\d{1,2}[/-]\d{1,2}[/-](\d{4})', text, re.IGNORECASE)
        default_year = int(year_hdr.group(1)) if year_hdr else datetime.now().year

        def to_signed_balance(bal_str, sign):
            bal = float(bal_str.replace(',', ''))
            return -bal if sign.lower() == 'dr' else bal

        transactions = []
        month_year = {}   # month number -> year, learned from B/F / C/F lines
        prev_balance = None

        for raw_line in text.split('\n'):
            line = raw_line.strip()
            if not line:
                continue

            m = bf_cf_re.match(line)
            if m:
                mon_key = m.group(2).lower()
                if mon_key in month_map:
                    month_year[month_map[mon_key]] = int(m.group(3))
                prev_balance = to_signed_balance(m.group(4), m.group(5))
                continue

            m = txn_re.match(line)
            if not m:
                continue
            day_str, mon_str, desc, amount_str, balance_str, sign = m.groups()
            mon_key = mon_str.lower()
            if mon_key not in month_map:
                continue
            month_num = month_map[mon_key]
            year_num = month_year.get(month_num, default_year)
            try:
                txn_date = datetime(year_num, month_num, int(day_str))
            except ValueError:
                continue
            try:
                amount = float(amount_str.replace(',', ''))
            except ValueError:
                continue
            balance = to_signed_balance(balance_str, sign)

            if prev_balance is None:
                # No B/F line seen yet to anchor direction — default to Payment,
                # the more conservative guess (won't misfile money as received).
                txn_type = 'Payment'
            else:
                txn_type = 'Receipt' if balance > prev_balance else 'Payment'
            prev_balance = balance

            transactions.append({
                'date': txn_date.strftime('%Y-%m-%d'),
                'description': desc.strip(),
                'amount': amount,
                'type': txn_type
            })

        if not transactions:
            flash('Could not cleanly extract transactions. Please try a clearer image.', 'warning')

        return render_template('bank_statement_review.html', transactions=transactions)

    except Exception as e:
        logging.error(f"OCR Error: {e}")
        if 'tesseract' in str(e).lower() or 'not found' in str(e).lower():
            flash(_TESSERACT_HELP, 'danger')
        else:
            flash(f'Error processing image: {str(e)}', 'danger')
        return redirect(url_for('bulk_upload_gl'))


@app.route('/bulk_upload_tb', methods=['GET', 'POST'])
@login_required
@has_permission('Access_Accounting')
def bulk_upload_tb():
    if request.method == 'POST':
        if 'file' in request.files:
            file = request.files['file']
            if file.filename == '':
                flash('No file selected', 'danger')
                return redirect(url_for('bulk_upload_tb'))

            try:
                # Need to reset stream cursor since read_csv_content consumes it
                # Wait, our first parse_csv_file definition parses directly.
                # Just call parse_csv_file directly and use its output.
                parsed_rows = parse_csv_file(file, required_columns=['Account Name', 'Debit', 'Credit'])

                rows = []
                missing_accounts = []

                # Fetch existing accounts
                existing = {a['account_name'] for a in db.execute_query("SELECT account_name FROM new_account_table")}

                for row in parsed_rows:
                    name = row.get('Account Name', '').strip()
                    dr = parse_float(row.get('Debit', 0) or 0)
                    cr = parse_float(row.get('Credit', 0) or 0)

                    if not name: continue

                    status = 'OK'
                    if name not in existing:
                        status = 'Missing'
                        missing_accounts.append(name)

                    rows.append({'name': name, 'dr': dr, 'cr': cr, 'status': status})

                # Calculate Totals
                total_dr = sum(r['dr'] for r in rows)
                total_cr = sum(r['cr'] for r in rows)

                if missing_accounts:
                    flash(f'Found {len(missing_accounts)} missing accounts. Please create them first.', 'warning')

                    # Fetch categories again for quick create modal
                    bs_cats, pl_cats, cf_cats = get_cached_categories(db)

                    return render_template('bulk_upload_tb_review.html',
                                           rows=rows, total_dr=total_dr, total_cr=total_cr,
                                           bs_cats=bs_cats, pl_cats=pl_cats, cf_cats=cf_cats,
                                           today_date=date.today().strftime('%Y-%m-%d'))

                return render_template('bulk_upload_tb_review.html', rows=rows, total_dr=total_dr, total_cr=total_cr, today_date=date.today().strftime('%Y-%m-%d'))

            except Exception as e:
                flash(f'Error: {str(e)}', 'danger')

        elif 'save_tb' in request.form:
            # Post TB as Opening Balance JV
            names = request.form.getlist('account_name[]')
            drs = request.form.getlist('dr[]')
            crs = request.form.getlist('cr[]')
            opening_date = request.form.get('opening_date')

            if not opening_date:
                flash('Opening Balance Date is required', 'danger')
                return redirect(url_for('bulk_upload_tb'))

            # Verify Totals
            total_dr = sum(parse_float(d) for d in drs)
            total_cr = sum(parse_float(c) for c in crs)

            if abs(total_dr - total_cr) > 0.01:
                flash(f'Totals do not match! Debit: {total_dr}, Credit: {total_cr}. Difference: {total_dr - total_cr}', 'danger')
                return redirect(url_for('bulk_upload_tb'))

            try:
                conn = db.get_connection()
                cursor = conn.cursor()
                conn.start_transaction()

                current_user = get_current_user_id()
                current_user_pk = get_current_user_pk()
                today = date.today()

                # Create JV
                cursor.execute("INSERT INTO jv_numbers (jv_user_code, jv_naration, status) VALUES (%s, %s, 1)",
                               ('OB-UPLOAD', 'Opening Balance Upload'))
                jv_no = cursor.lastrowid

                entries_to_insert = []
                count = 0
                for i in range(len(names)):
                    dr = parse_float(drs[i] or 0)
                    cr = parse_float(crs[i] or 0)
                    if dr == 0 and cr == 0: continue

                    entries_to_insert.append((names[i], dr, cr, opening_date, today, current_user, jv_no))
                    count += 1

                if entries_to_insert:
                    cursor.executemany("""
                        INSERT INTO entry_details (
                            account_name, enty_values_DR, enty_values_CR,
                            entry_effective_date, entry_create_date, entry_naration,
                            entry_create_user, entry_jv
                        ) VALUES (%s, %s, %s, %s, %s, 'Opening Balance', %s, %s)
                    """, entries_to_insert)

                conn.commit()
                flash(f'TB Uploaded successfully. {count} entries posted to JV {jv_no}', 'success')
                return redirect(url_for('trial_balance'))

            except Exception as e:
                if conn: conn.rollback()
                flash(f'Error posting TB: {str(e)}', 'danger')
                return redirect(url_for('bulk_upload_tb'))

    return render_template('bulk_upload_tb.html')

# --- Opening Balances (Suppliers / Customers / Accounts) ---

def ensure_opening_balance_account():
    """Return the name of the 'Opening Balance Equity' contra account, creating it if missing."""
    name = 'Opening Balance Equity'
    existing = db.execute_query("SELECT account_name FROM new_account_table WHERE account_name = %s", (name,))
    if existing:
        return name
    # Mirror the balance-sheet category/position of an existing equity account if one exists
    ref = db.execute_query(
        "SELECT account_name_of_catogory_Balace_sheet AS cat, account_hold_possion_Balace_Sheet AS pos "
        "FROM new_account_table WHERE account_equity = 1 "
        "AND account_name_of_catogory_Balace_sheet IS NOT NULL LIMIT 1"
    )
    bs_cat = ref[0]['cat'] if ref else 'Equity'
    bs_pos = ref[0]['pos'] if ref else 'Liabilities'
    try:
        db.execute_query("""
            INSERT INTO new_account_table (
                account_name, account_hold_possion_PL, account_hold_possion_Balace_Sheet,
                account_name_of_catogory_PL, account_name_of_catogory_Balace_sheet,
                account_income, account_expenses, account_assets, account_liabilities, account_equity,
                cf_catogory, accont_create_date, account_create_user, account_active, account_basment, currency_code
            ) VALUES (%s, %s, %s, %s, %s, 0, 0, 0, 0, 1, %s, %s, %s, 1, '', 'LKR')
        """, (name, None, bs_pos, None, bs_cat, 'Financing', date.today(), get_current_user_id()), commit=True)
    except Exception as e:
        logging.error(f"Could not auto-create Opening Balance Equity account: {e}")
    return name


def _post_ob_gl(cursor, account_name, dr, cr, ob_date, narration, user_pk, jv_no):
    cursor.execute("""
        INSERT INTO entry_details (account_name, enty_values_DR, enty_values_CR,
            entry_effective_date, entry_create_date, entry_naration, entry_create_user, entry_jv)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
    """, (account_name, dr, cr, ob_date, date.today(), narration, user_pk, jv_no))


@app.route('/opening_balances', methods=['GET', 'POST'])
@login_required
@has_permission('Access_Accounting')
def opening_balances():
    if request.method == 'POST':
        action = request.form.get('action')
        contra = (request.form.get('contra_account') or '').strip() or ensure_opening_balance_account()
        ob_date = request.form.get('ob_date') or date.today().strftime('%Y-%m-%d')
        current_user = get_current_user_id()
        user_pk = get_current_user_pk()

        if action == 'supplier_ob':
            supplier_id = request.form.get('supplier_id')
            amount = parse_float(request.form.get('amount'))
            reference = (request.form.get('reference') or '').strip() or f"OB-{ob_date}"
            if not supplier_id or amount <= 0:
                flash('Select a supplier and enter an amount greater than zero.', 'danger')
                return redirect(url_for('opening_balances'))

            conn = db.get_connection(); cursor = conn.cursor(); conn.start_transaction()
            try:
                cursor.execute("SELECT supplier_code FROM suppliers WHERE sup_id = %s", (supplier_id,))
                r = cursor.fetchone(); supplier_code = r[0] if r else ''
                cursor.execute("INSERT INTO jv_numbers (jv_user_code, jv_naration, status) VALUES (%s, %s, 1)",
                               (str(current_user), f'Supplier Opening Balance - {reference}'))
                jv_no = cursor.lastrowid
                # Payable subledger — outstanding invoice so it shows in aging & payments can be applied
                cursor.execute("""
                    INSERT INTO suppliers_invoice_data (
                        suppliers_code, suppliers_invoice_number, suppliers_invoice_date,
                        suppliers_invoice_total_oustanding, suppliers_invoice_total_payment,
                        suppliers_invoice_final_date, suppliers_invoice_buinding_supplier,
                        suppliers_invoice_JV, suppliers_VAT_rate
                    ) VALUES (%s, %s, %s, %s, 0, %s, %s, %s, 0)
                """, (supplier_code, reference, ob_date, amount, ob_date, supplier_id, jv_no))
                # GL: Credit Account Payable ; Debit contra
                _post_ob_gl(cursor, 'Account Payable', 0, amount, ob_date, 'Supplier Opening Balance', user_pk, jv_no)
                _post_ob_gl(cursor, contra, amount, 0, ob_date, 'Supplier Opening Balance', user_pk, jv_no)
                conn.commit()
                flash(f'Supplier opening balance of {amount:,.2f} saved (JV {jv_no}).', 'success')
            except Exception as e:
                conn.rollback(); flash(f'Error saving supplier opening balance: {e}', 'danger')
            finally:
                cursor.close(); conn.close()

        elif action == 'customer_ob':
            customer_id = request.form.get('customer_id')
            amount = parse_float(request.form.get('amount'))
            reference = (request.form.get('reference') or '').strip() or f"OB-{ob_date}"
            if not customer_id or amount <= 0:
                flash('Select a customer and enter an amount greater than zero.', 'danger')
                return redirect(url_for('opening_balances'))

            conn = db.get_connection(); cursor = conn.cursor(); conn.start_transaction()
            try:
                cursor.execute("INSERT INTO jv_numbers (jv_user_code, jv_naration, status) VALUES (%s, %s, 1)",
                               (str(current_user), f'Customer Opening Balance - {reference}'))
                jv_no = cursor.lastrowid
                # Receivable subledger — outstanding invoice so it shows in aging & receipts can be applied
                cursor.execute("""
                    INSERT INTO Invoice_Oustanding (
                        invoice_number, invoice_date, invoice_total_oustanding,
                        invoice_oustanding_Patment, invoice_final_date,
                        invoice_buinding_Customer, invoice_JV, VAT_rate, oustanding_delete
                    ) VALUES (%s, %s, %s, 0, %s, %s, %s, 0, 0)
                """, (reference, ob_date, amount, ob_date, customer_id, jv_no))
                # GL: Debit Account Receivable ; Credit contra
                _post_ob_gl(cursor, 'Account Receivable', amount, 0, ob_date, 'Customer Opening Balance', user_pk, jv_no)
                _post_ob_gl(cursor, contra, 0, amount, ob_date, 'Customer Opening Balance', user_pk, jv_no)
                conn.commit()
                flash(f'Customer opening balance of {amount:,.2f} saved (JV {jv_no}).', 'success')
            except Exception as e:
                conn.rollback(); flash(f'Error saving customer opening balance: {e}', 'danger')
            finally:
                cursor.close(); conn.close()

        elif action == 'account_ob':
            account_name = request.form.get('account_name')
            amount = parse_float(request.form.get('amount'))
            side = request.form.get('side')
            if not account_name or amount <= 0 or side not in ('DR', 'CR'):
                flash('Select an account, choose Debit/Credit and enter an amount.', 'danger')
                return redirect(url_for('opening_balances'))
            if account_name == contra:
                flash('The account and the contra account cannot be the same.', 'danger')
                return redirect(url_for('opening_balances'))

            conn = db.get_connection(); cursor = conn.cursor(); conn.start_transaction()
            try:
                cursor.execute("INSERT INTO jv_numbers (jv_user_code, jv_naration, status) VALUES (%s, %s, 1)",
                               (str(current_user), 'Account Opening Balance'))
                jv_no = cursor.lastrowid
                acc_dr, acc_cr = (amount, 0) if side == 'DR' else (0, amount)
                con_dr, con_cr = (0, amount) if side == 'DR' else (amount, 0)
                _post_ob_gl(cursor, account_name, acc_dr, acc_cr, ob_date, 'Account Opening Balance', user_pk, jv_no)
                _post_ob_gl(cursor, contra, con_dr, con_cr, ob_date, 'Account Opening Balance', user_pk, jv_no)
                conn.commit()
                flash(f'Account opening balance of {amount:,.2f} saved (JV {jv_no}).', 'success')
            except Exception as e:
                conn.rollback(); flash(f'Error saving account opening balance: {e}', 'danger')
            finally:
                cursor.close(); conn.close()

        return redirect(url_for('opening_balances'))

    # GET
    suppliers = db.execute_query("SELECT sup_id, supplier_name, supplier_code FROM suppliers WHERE Is_Suplier = 1 ORDER BY supplier_name") or []
    customers = db.execute_query("SELECT sup_id, supplier_name, supplier_code FROM suppliers WHERE Is_Customer = 1 ORDER BY supplier_name") or []
    accounts = db.execute_query("SELECT account_name FROM new_account_table WHERE account_active = 1 ORDER BY account_name") or []
    recent = db.execute_query(
        "SELECT jv_id, jv_naration FROM jv_numbers WHERE jv_naration LIKE '%Opening Balance%' ORDER BY jv_id DESC LIMIT 30"
    ) or []

    return render_template('opening_balances.html',
                           suppliers=suppliers, customers=customers, accounts=accounts,
                           recent=recent, today=date.today().strftime('%Y-%m-%d'))

# --- Cheque Print Setup ---
@app.route('/cheque_print_setup', methods=['GET', 'POST'])
@login_required
@has_permission('Access_Accounting')
def cheque_print_setup():
    if request.method == 'POST':
        bank_account = request.form.get('bank_account')
        width = request.form.get('paper_width')
        height = request.form.get('paper_height')

        # Coordinates
        date_x = request.form.get('date_x')
        date_y = request.form.get('date_y')
        payee_x = request.form.get('payee_x')
        payee_y = request.form.get('payee_y')
        words_x = request.form.get('words_x')
        words_y = request.form.get('words_y')
        digits_x = request.form.get('digits_x')
        digits_y = request.form.get('digits_y')

        # Font Sizes
        date_fs = request.form.get('date_fs')
        payee_fs = request.form.get('payee_fs')
        words_fs = request.form.get('words_fs')
        digits_fs = request.form.get('digits_fs')

        is_cross = 1 if request.form.get('is_cross') else 0

        try:
            # Check if settings exist for bank account (or global if NULL)
            # Simplification: One setting per bank account, or one global if bank_account is None
            # Here we assume we are saving for a specific bank account or default

            # Delete existing for this bank to overwrite (or update)
            if bank_account:
                db.execute_query("DELETE FROM cheque_print_settings WHERE bank_account = %s", (bank_account,), commit=True)
            else:
                db.execute_query("DELETE FROM cheque_print_settings WHERE bank_account IS NULL", commit=True)

            query = """
                INSERT INTO cheque_print_settings (
                    bank_account, paper_width_mm, paper_height_mm,
                    date_x, date_y, date_font_size,
                    payee_x, payee_y, payee_font_size,
                    amount_words_x, amount_words_y, amount_words_font_size,
                    amount_digits_x, amount_digits_y, amount_digits_font_size,
                    is_cross_cheque
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            params = (
                bank_account if bank_account else None, width, height,
                date_x, date_y, date_fs,
                payee_x, payee_y, payee_fs,
                words_x, words_y, words_fs,
                digits_x, digits_y, digits_fs,
                is_cross
            )
            db.execute_query(query, params, commit=True)
            flash('Cheque settings saved successfully', 'success')

        except Exception as e:
            flash(f'Error saving settings: {str(e)}', 'danger')

        return redirect(url_for('cheque_print_setup', bank_account=bank_account))

    # GET
    selected_bank = request.args.get('bank_account')
    bank_accounts = db.execute_query("SELECT bank_bookcol_account_number FROM bank_book")

    settings = None
    if selected_bank:
        res = db.execute_query("SELECT * FROM cheque_print_settings WHERE bank_account = %s", (selected_bank,))
        if res: settings = res[0]

    # Fallback to global/default if not found for specific bank
    if not settings:
        res = db.execute_query("SELECT * FROM cheque_print_settings WHERE bank_account IS NULL")
        if res: settings = res[0]

    return render_template('cheque_print_setup.html',
                           bank_accounts=bank_accounts,
                           settings=settings,
                           selected_bank=selected_bank)

@app.route('/cheque/print/<int:jv_no>')
@login_required
def print_cheque(jv_no):
    # Fetch Payment Details
    # We need Payee, Date, Amount, Bank Account
    query = """
        SELECT
            b.Bank_Payment_Date as date,
            b.bank_book__suplier_name as payee,
            SUM(b.bank_book__recode_cr) as amount,
            b.bank_book__accont_name as bank_account
        FROM bank_book_recod b
        WHERE b.jv_numbers_jv_id = %s
        GROUP BY b.Bank_Payment_Date, b.bank_book__suplier_name, b.bank_book__accont_name
    """
    res = db.execute_query(query, (jv_no,))
    if not res:
        return "Payment Not Found", 404
    payment = res[0]

    # Fetch Settings for this bank
    settings_res = db.execute_query("SELECT * FROM cheque_print_settings WHERE bank_account = %s", (payment['bank_account'],))
    settings = settings_res[0] if settings_res else None

    if not settings:
        # Fallback to default
        settings_res = db.execute_query("SELECT * FROM cheque_print_settings WHERE bank_account IS NULL")
        settings = settings_res[0] if settings_res else {}

    # Convert amount to words (Placeholder logic or simple implementation)
    # Ideally use `num2words` library. Since I can't install packages easily, simple fallback or dummy.
    # Actually I can `pip install num2words` in bash session if needed, or write a simple function.
    # For now, let's assume numeric.

    return render_template('cheque_print.html', payment=payment, settings=settings)

# --- Tax Settings ---
@app.route('/tax_settings', methods=['GET', 'POST'])
@login_required
@has_permission('Add_New_User') # Assuming admin/setup permission
def tax_settings():
    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'add':
            name = request.form.get('tax_name')
            rate = request.form.get('tax_rate')
            desc = request.form.get('description')

            if name and rate:
                try:
                    db.execute_query("INSERT INTO tax_rates (tax_name, rate, description, active) VALUES (%s, %s, %s, 1)",
                                     (name, rate, desc), commit=True)
                    flash('Tax rate added successfully', 'success')
                except Exception as e:
                    flash(f'Error adding tax rate: {e}', 'danger')

        elif action == 'delete':
            tid = request.form.get('id')
            if tid:
                db.execute_query("DELETE FROM tax_rates WHERE id = %s", (tid,), commit=True)
                flash('Tax rate deleted', 'success')

        return redirect(url_for('tax_settings'))

    rates = db.execute_query("SELECT * FROM tax_rates")
    return render_template('tax_settings.html', rates=rates)

# --- Company Profile ---
@app.route('/company_profile', methods=['GET', 'POST'])
@login_required
@has_permission('Add_New_User')
def company_profile():
    if request.method == 'POST':


        name = request.form.get('company_name')
        addr1 = request.form.get('address_no')
        addr2 = request.form.get('city')
        addr3 = request.form.get('province')
        addr4 = request.form.get('country')
        addr5 = request.form.get('postal_code')
        land = request.form.get('land_no')
        fax = request.form.get('fax_no')
        vat = request.form.get('vat_no')
        curr = request.form.get('currency')
        vat_registered = 1 if request.form.get('vat_registered') else 0

        # Handle Logo Upload
        logo_data = None
        if 'logo' in request.files:
            file = request.files['logo']
            if file.filename != '':
                logo_data = base64.b64encode(file.read()).decode('utf-8')

        # Check if record exists (ID is usually 0 or 1, assuming single row config)
        exists = db.execute_query("SELECT id FROM company")

        try:
            if exists:
                # Update
                query = """
                    UPDATE company SET
                    company_name=%s, company_addras_1=%s, company_addras_2=%s,
                    company_addras_3=%s, company_addras_4=%s, company_addras_5=%s,
                    company_land_line=%s, company_fax_line=%s, company_vate_code=%s,
                    company_curency=%s, vat_registered=%s
                """
                params = [name, addr1, addr2, addr3, addr4, addr5, land, fax, vat, curr, vat_registered]

                if logo_data:
                    query += ", company_log=%s"
                    params.append(logo_data)

                db.execute_query(query, tuple(params), commit=True)
            else:
                # Insert
                query = """
                    INSERT INTO company (
                        id, company_name, company_addras_1, company_addras_2, company_addras_3,
                        company_addras_4, company_addras_5, company_land_line, company_fax_line,
                        company_vate_code, company_curency, company_log, vat_registered
                    ) VALUES (0, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                db.execute_query(query, (name, addr1, addr2, addr3, addr4, addr5, land, fax, vat, curr, logo_data, vat_registered), commit=True)

            flash('Company profile updated successfully', 'success')
        except Exception as e:
            flash(f'Error updating profile: {str(e)}', 'danger')

        return redirect(url_for('company_profile'))

    # Load Data
    res = db.execute_query("SELECT * FROM company LIMIT 1")
    company = res[0] if res else {}

    # If logo exists as bytes, we might need to handle it for display (template expects base64 string)
    # Note: MySQL connector returns bytes for blobs.
    # If it was saved as base64 string (as above), it comes out as string or bytes depending on driver.
    # The template expects a base64 string.
    if company.get('company_log') and isinstance(company['company_log'], bytes):
        try:
            # If it's already base64 bytes
            company['company_log'] = company['company_log'].decode('utf-8')
        except UnicodeDecodeError:
            # If it's raw image bytes, encode it

            company['company_log'] = base64.b64encode(company['company_log']).decode('utf-8')

    currencies = db.execute_query("SELECT currency_code, currency_name FROM currency_table")
    if not currencies: # Fallback if table empty
        currencies = [{'currency_code': 'LKR', 'currency_name': 'Sri Lankan Rupee'}]

    return render_template('company_profile.html', company=company, currencies=currencies)

def _base_currency_and_list():
    """Return (base_currency_code, [currencies]) for multi-currency transaction pages."""
    base_currency = 'LKR'
    try:
        res = db.execute_query("SELECT company_curency FROM company LIMIT 1")
        if res and res[0].get('company_curency'):
            base_currency = res[0]['company_curency']
    except Exception:
        pass
    try:
        currencies = db.execute_query("SELECT currency_code, currency_name FROM currency_table") or []
    except Exception:
        currencies = []
    return base_currency, currencies

# --- Bank Payment Module ---
@app.route('/bank_payment', methods=['GET'])
@login_required
@has_permission('Access_Accounting')
def bank_payment():
    suppliers = db.execute_query("SELECT supplier_name FROM suppliers WHERE Is_Suplier = 1 AND supplier_name != 'Direct Payment'")
    bank_accounts = db.execute_query("SELECT bank_bookcol_account_number FROM bank_book")
    last_jv = request.args.get('last_jv')
    base_currency, currencies = _base_currency_and_list()
    return render_template('bank_payment.html', suppliers=suppliers, bank_accounts=bank_accounts,
                           today_date=date.today().strftime('%Y-%m-%d'), last_jv=last_jv,
                           base_currency=base_currency, currencies=currencies)

# --- Cash Payment Module ---
@app.route('/cash_payment', methods=['GET'])
@login_required
@has_permission('Access_Accounting')
def cash_payment():
    suppliers = db.execute_query("SELECT supplier_name FROM suppliers WHERE Is_Suplier = 1 AND supplier_name != 'Direct Payment'")
    cash_accounts = db.execute_query("SELECT cash_book_account_name FROM cash_book")
    last_jv = request.args.get('last_jv')
    base_currency, currencies = _base_currency_and_list()
    return render_template('cash_payment.html', suppliers=suppliers, cash_accounts=cash_accounts,
                           today_date=date.today().strftime('%Y-%m-%d'), last_jv=last_jv,
                           base_currency=base_currency, currencies=currencies)


@app.route('/api/suppliers_list')
@login_required
def api_suppliers_list():
    """Live supplier list so a newly created supplier shows up in an open
    cash/bank payment window without a full page refresh."""
    rows = db.execute_query(
        "SELECT supplier_name FROM suppliers WHERE Is_Suplier = 1 AND supplier_name != 'Direct Payment' ORDER BY supplier_name") or []
    return jsonify([r['supplier_name'] for r in rows])


@app.route('/api/customers_list')
@login_required
def api_customers_list():
    """Live customer list so a newly created customer shows up in an open
    invoice window without a full page refresh."""
    rows = db.execute_query(
        "SELECT supplier_name AS customer_name FROM suppliers WHERE Is_Customer = 1 ORDER BY supplier_name") or []
    return jsonify([r['customer_name'] for r in rows])


@app.route('/api/invoice_items')
@login_required
def api_invoice_items():
    """Live item list (id, name, code, unit, cost) for the invoice window so a
    newly created item shows up without a full page refresh."""
    rows = db.execute_query("""
        SELECT i.id, i.inventoy_name AS name, i.inventoy_code AS code,
               i.inventoy_items_messurment_unit AS unit,
               (SELECT COALESCE(p.inventory_price_purcharsing, 0)
                  FROM inventory_price_recod p
                 WHERE p.inventory_price_link = i.id
                 ORDER BY p.id DESC LIMIT 1) AS cost
        FROM inventoy_items i
        WHERE i.active = 1
        ORDER BY i.inventoy_name
    """) or []
    return jsonify([{'id': r['id'], 'name': r['name'], 'code': r['code'] or '',
                     'unit': r['unit'] or '', 'cost': float(r['cost'] or 0)} for r in rows])


@app.route('/api/srn_suppliers')
@login_required
def api_srn_suppliers():
    """Live supplier list (id, name, code) for the Service Entry window."""
    rows = db.execute_query(
        "SELECT sup_id, supplier_name, supplier_code FROM suppliers WHERE Is_Suplier = 1 ORDER BY supplier_name") or []
    return jsonify([{'id': r['sup_id'], 'name': r['supplier_name'], 'code': r['supplier_code'] or ''} for r in rows])


@app.route('/api/srn_accounts')
@login_required
def api_srn_accounts():
    """Live P&L (income/expense) account list for the Service Entry line rows."""
    rows = db.execute_query("""
        SELECT account_name FROM new_account_table
        WHERE account_active = 1 AND (account_income = 1 OR account_expenses = 1)
        ORDER BY account_name
    """) or []
    return jsonify([r['account_name'] for r in rows])

def _get_supplier_history_data(supplier_name):
    """Helper to fetch common supplier details, outstanding invoices, and
    payment history for a supplier. History is the COMBINED Cash + Bank
    payment history (tagged per row with 'method': 'Cash'/'Bank') so both
    the Cash Payment and Bank Payment screens show the full picture of what
    was paid to this supplier, regardless of which screen a given payment
    was made on."""
    details, inv_list, sup_id = _get_supplier_base_data(supplier_name)
    if not details:
        return {'error': 'Supplier not found'}, 404

    cash_history = db.execute_query("""
        SELECT cash_book_recod_voucher_no as voucher, Payment_Date as date,
               cash_book_recode_accont_name as account, cash_book_recode_cr as amount,
               User_Enter as user_id, jv_numbers_jv_id as jv_no,
               cash_book_recode_suplier_oustanding_id as inv_id
        FROM cash_book_recode
        WHERE TRIM(cash_book_recode_suplier_name) = TRIM(%s)
        ORDER BY chash_book_recod_id DESC
    """, (supplier_name,)) or []

    bank_history = db.execute_query("""
        SELECT bank_book_recod_voucher_no as voucher, Bank_Payment_Date as date,
               bank_book__accont_name as account, bank_book__recode_cr as amount,
               bank_book__naration as narration, jv_numbers_jv_id as jv_no,
               bank_book__suplier_oustanding_id as inv_id
        FROM bank_book_recod
        WHERE TRIM(bank_book__suplier_name) = TRIM(%s)
        ORDER BY id DESC
    """, (supplier_name,)) or []

    combined = [dict(h, method='Cash') for h in cash_history] + \
               [dict(h, method='Bank') for h in bank_history]
    # Merge-sort the two already-descending lists by date (falling back to
    # jv_no as a tie-breaker so same-day payments still stay in a stable,
    # newest-first order).
    combined.sort(key=lambda h: (str(h.get('date') or ''), h.get('jv_no') or 0), reverse=True)

    # Look up the original invoice (and its GRN, if any) each payment history row settled,
    # so a "View" link can still be shown after the invoice is fully paid off and has
    # dropped out of the Outstanding Invoices list above.
    inv_ids = sorted({h['inv_id'] for h in combined if h.get('inv_id')})
    inv_map = {}
    if inv_ids:
        placeholders = ','.join(['%s'] * len(inv_ids))
        inv_rows = db.execute_query(f"""
            SELECT s.s_i_id, s.suppliers_invoice_number, s.suppliers_invoice_JV, j.jv_user_code
            FROM suppliers_invoice_data s
            LEFT JOIN jv_numbers j ON s.suppliers_invoice_JV = j.jv_id
            WHERE s.s_i_id IN ({placeholders})
        """, tuple(inv_ids)) or []
        for r in inv_rows:
            inv_map[r['s_i_id']] = {
                'invoice_no': r['suppliers_invoice_number'],
                'inv_jv': r['suppliers_invoice_JV'],
                'is_grn': r.get('jv_user_code') == 'JV FROM GRN'
            }

    hist_list = []
    for h in combined:
        inv_info = inv_map.get(h.get('inv_id')) or {}
        hist_list.append({
            'voucher': h['voucher'],
            'date': str(h['date']),
            'account': h['account'],
            'amount': float(h['amount'] or 0),
            'jv_no': h['jv_no'],
            'method': h['method'],
            'invoice_no': inv_info.get('invoice_no'),
            'inv_jv': inv_info.get('inv_jv'),
            'is_grn': inv_info.get('is_grn', False),
            'user_id': h.get('user_id'),
            'narration': h.get('narration'),
        })

    return {'details': details, 'invoices': inv_list, 'history': hist_list}

def _get_supplier_base_data(supplier_name):
    """Helper to fetch common supplier details and outstanding invoices."""
    sup_data = db.execute_query("SELECT * FROM suppliers WHERE supplier_name = %s", (supplier_name,))
    if not sup_data:
        return None, None, None

    s = sup_data[0]
    details = {
        'code': s['supplier_code'],
        'address': f"{s['supplier_address_1']}, {s['supplier_address_2']}",
        'mobile': s['suppliers_teli_1'],
        'email': s['suppliers_e_mail'],
        'vat': s['suppliers_vat_regidter_no']
    }
    sup_id = s['sup_id']

    invoices = db.execute_query("""
        SELECT s.s_i_id, s.suppliers_invoice_number, s.suppliers_invoice_date, s.suppliers_invoice_final_date,
               s.suppliers_invoice_total_oustanding, s.suppliers_invoice_total_payment, s.suppliers_invoice_oustanding,
               s.suppliers_invoice_JV, j.jv_user_code
        FROM suppliers_invoice_data s
        LEFT JOIN jv_numbers j ON s.suppliers_invoice_JV = j.jv_id
        WHERE s.suppliers_invoice_buinding_supplier = %s AND s.suppliers_invoice_oustanding > 0 AND s.suppliers_oustanding_delete = 0
    """, (sup_id,))

    inv_list = []
    for inv in invoices:
        inv_list.append({
            'id': inv['s_i_id'],
            'invoice_no': inv['suppliers_invoice_number'],
            'date': str(inv['suppliers_invoice_date']),
            'due_date': str(inv['suppliers_invoice_final_date']),
            'total': float(inv['suppliers_invoice_total_oustanding']),
            'paid': float(inv['suppliers_invoice_total_payment']),
            'balance': float(inv['suppliers_invoice_oustanding']),
            'jv': inv['suppliers_invoice_JV'],
            # 'View GRN' only makes sense when this invoice was created by a
            # GRN — Direct Purchase / Service Entry invoices have no GRN.
            'is_grn': inv.get('jv_user_code') == 'JV FROM GRN'
        })

    # An invoice already covered by a not-yet-cleared postdated cheque is
    # tracked from the Postdated Cheques page instead — remove it from this
    # Outstanding list entirely as soon as ANY pending cheque references it
    # (regardless of the amount), so it doesn't clutter the list or get paid
    # twice. It automatically reappears here the moment that cheque is
    # Cancelled (status stops being 'pending').
    try:
        pdc_rows = db.execute_query(
            "SELECT id, cheque_no, post_date, payload FROM postdated_cheques "
            "WHERE pdc_type = 'payment' AND status = 'pending' AND party_name = %s",
            (supplier_name,)) or []
    except Exception:
        pdc_rows = []

    covered_ids = set()
    for pdc in pdc_rows:
        try:
            payload = json.loads(pdc['payload'] or '{}')
        except Exception:
            continue
        for p in payload.get('payments', []):
            covered_ids.add(str(p.get('id')))

    inv_list = [item for item in inv_list if str(item['id']) not in covered_ids]

    return details, inv_list, sup_id


@app.route('/cash_payment/get_data')
@login_required
def get_cash_supplier_data():
    supplier_name = request.args.get('name')
    if not supplier_name:
        return {'error': 'No supplier name'}, 400

    return _get_supplier_history_data(supplier_name)

@app.route('/cash_payment/submit', methods=['POST'])
@login_required
def cash_payment_submit():
    supplier_name = request.form.get('supplier')
    cash_account = request.form.get('cash_account')
    payment_date = request.form.get('payment_date')
    narration = request.form.get('narration')
    wht_amount = parse_float(request.form.get('wht_amount', 0))

    # Multi-currency: amounts are entered in this currency; rate converts to base (LKR)
    payment_currency = (request.form.get('payment_currency') or '').strip()
    exchange_rate = parse_float(request.form.get('exchange_rate', 1)) or 1.0
    if exchange_rate <= 0:
        exchange_rate = 1.0

    if not supplier_name or not cash_account:
        flash('Missing supplier or cash account', 'danger')
        return redirect(url_for('cash_payment'))

    # Collect payments (Total Gross Payment to Supplier). 'amount' is in the payment
    # currency; 'base' is the LKR-equivalent posted to the ledger and outstanding.
    payments = []
    total_payment = 0

    # Iterate form to find payment items
    for key in request.form:
        if key.startswith('payment_'):
            inv_id = key.split('_')[1]
            try:
                amount = parse_float(request.form[key])
                if amount > 0:
                    payments.append({'id': inv_id, 'amount': amount, 'base': round(amount * exchange_rate, 2)})
                    total_payment += amount
            except (ValueError, TypeError):
                continue

    if not payments:
        flash('No payment amounts entered', 'warning')
        return redirect(url_for('cash_payment'))

    total_payment_base = round(total_payment * exchange_rate, 2)
    wht_base = round(wht_amount * exchange_rate, 2)
    if payment_currency and exchange_rate != 1.0:
        narration = (narration or '') + f" [{payment_currency} {total_payment:,.2f} @ {exchange_rate:.4f}]"

    _cash_jv = None  # ensure always defined
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        conn.start_transaction()
        current_user = get_current_user_id()
        current_user_pk = get_current_user_pk()

        # 1. Voucher Number — manual override if provided, else auto
        manual_voucher = (request.form.get('manual_voucher') or '').strip()
        if manual_voucher:
            new_voucher = manual_voucher
        else:
            cursor.execute("SELECT MAX(cash_voucher_number) FROM cash_voucher_no WHERE cash_voucher_link = %s", (cash_account,))
            res = cursor.fetchone()
            max_voucher = res[0] if res and res[0] else 0
            new_voucher = max_voucher + 1

        cursor.execute("INSERT INTO cash_voucher_no (id, cash_voucher_link, cash_voucher_number) VALUES (0, %s, %s)",
                       (cash_account, new_voucher))

        # 2. Create Journal Voucher (JV)
        cursor.execute("INSERT INTO jv_numbers (jv_user_code, jv_naration) VALUES (%s, %s)", ('JV FROM PAYMENT', narration))
        jv_no = cursor.lastrowid

        # Get Sub Account Code
        cursor.execute("SELECT sub_account_code FROM sub_accont_for_new_account WHERE sub_sub_accaount_name = %s", (supplier_name,))
        res = cursor.fetchone()
        sub_ac_code = res[0] if res else 0

        # 3. Create GL Entries

        # A. Debit AP (Full Invoice Amount settled) — in base currency
        cursor.execute("""
            INSERT INTO entry_details (
                account_name, enty_values_DR, entry_effective_date, entry_create_date,
                entry_naration, entry_create_user, entry_jv, entry_sub_account_code
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, ('Account Payable', total_payment_base, payment_date, date.today(), narration, current_user_pk, jv_no, sub_ac_code))

        # B. Credit Cash (Net Amount = Total - WHT) — in base currency
        net_payment = total_payment_base - wht_base
        cursor.execute("""
            INSERT INTO entry_details (
                account_name, enty_values_CR, entry_effective_date, entry_create_date,
                entry_naration, entry_create_user, entry_jv
            ) VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (cash_account, net_payment, payment_date, date.today(), narration, current_user_pk, jv_no))

        # C. Credit WHT Payable (If any) — in base currency
        if wht_base > 0:
            cursor.execute("""
                INSERT INTO entry_details (
                    account_name, enty_values_CR, entry_effective_date, entry_create_date,
                    entry_naration, entry_create_user, entry_jv
                ) VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, ('WHT Payable', wht_base, payment_date, date.today(), f"WHT on {narration}", current_user_pk, jv_no))

        # 4. Process Individual Payments (Update Outstanding)
        if payments:
            inv_ids = [p['id'] for p in payments]
            format_strings = ','.join(['%s'] * len(inv_ids))
            cursor.execute(f"SELECT s_i_id, suppliers_invoice_oustanding FROM suppliers_invoice_data WHERE s_i_id IN ({format_strings})", tuple(inv_ids))

            outstanding_map = {}
            for row in cursor.fetchall():
                outstanding_map[str(row[0])] = parse_float(row[1] or 0)

            call_params = []
            insert_params = []

            for p in payments:
                current_outstanding = outstanding_map.get(str(p['id']), 0.0)
                call_params.append((current_outstanding, p['base'], p['id']))

                # Net cash portion of this line (base), after proportional WHT
                net_item_amount = p['base']
                if total_payment_base > 0:
                    net_item_amount = p['base'] * (net_payment / total_payment_base)

                insert_params.append((
                    net_item_amount, cash_account, narration,
                    p['id'], supplier_name, jv_no,
                    p['id'], new_voucher, current_user_pk, payment_date
                ))

            # Update supplier outstanding (in base currency) — settle by the base amount.
            # suppliers_invoice_oustanding is a GENERATED column (total_oustanding -
            # total_payment); only bump the payment and it recomputes itself.
            settle_params = [(p['base'], p['id']) for p in payments]
            cursor.executemany("""
                UPDATE suppliers_invoice_data
                SET suppliers_invoice_total_payment = COALESCE(suppliers_invoice_total_payment, 0) + %s
                WHERE s_i_id = %s
            """, settle_params)

            # Batch insert Cash Book Records
            cursor.executemany("""
                INSERT INTO cash_book_recode (
                    cash_book_recode_dr, cash_book_recode_cr, cash_book_recode_accont_name,
                    cash_book_recode_naration, cash_book_recode_suplier_oustanding_id,
                    cash_book_recode_suplier_name, jv_numbers_jv_id,
                    cash_book_po_no, cash_book_suplier_oustanding_id,
                    cash_book_recod_voucher_no, User_Enter, Payment_Date
                ) VALUES (0, %s, %s, %s, %s, %s, %s, 0, %s, %s, %s, %s)
            """, insert_params)

        conn.commit()
        flash(f'Cash Payment processed successfully. Voucher No: {new_voucher}', 'success')
        _cash_jv = jv_no

    except Exception as e:
        conn.rollback()
        flash(f'Transaction failed: {str(e)}', 'danger')
        logging.error(f"Cash Payment Error: {e}")
        _cash_jv = None
    finally:
        cursor.close()
        conn.close()

    if _cash_jv:
        return redirect(url_for('cash_payment', last_jv=_cash_jv))
    return redirect(url_for('cash_payment'))

@app.route('/cash_payment/delete_invoice', methods=['POST'])
@login_required
@has_permission('Access_Reversals')
def delete_cash_payment_invoice():
    jv_no = request.form.get('jv_no')
    if not jv_no:
        return {'error': 'No JV Number provided'}, 400

    conn = None
    cursor = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        conn.start_transaction()

        current_user_pk = get_current_user_pk()
        today = date.today()

        # 1. Mark supplier invoice as deleted
        cursor.execute(
            "UPDATE suppliers_invoice_data SET suppliers_oustanding_delete = 1 WHERE suppliers_invoice_JV = %s",
            (jv_no,))

        # 2. Create reversal OUT movements for inventory (undo stock-in)
        cursor.execute("""
            INSERT INTO inventory_recod (
                inventoy_name, inventoy_code,
                inventory_recod_action_date,
                inventory_recod_moument_in, inventory_recod_movment_out,
                inventory_recod_mesrmet, inventory_recod_unit_price,
                inventory_recod_account, inventory_recod_user_id,
                JV_No, inventory_recod_location
            )
            SELECT inventoy_name, inventoy_code, %s,
                   0, inventory_recod_moument_in,
                   inventory_recod_mesrmet, inventory_recod_unit_price,
                   'Invoice Deletion', %s,
                   %s, inventory_recod_location
            FROM inventory_recod
            WHERE JV_No = %s AND inventory_recod_moument_in > 0
        """, (today, current_user_pk, jv_no, jv_no))

        conn.commit()
        return {'success': True}

    except Exception as e:
        if conn:
            conn.rollback()
        return {'error': str(e)}, 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/voucher/print/<string:voucher_type>/<int:jv_no>')
@login_required
def print_voucher(voucher_type, jv_no):
    voucher_configs = {
        'cash': {
            'title': "CASH PAYMENT VOUCHER",
            'table': 'cash_book_recode c',
            'columns': {
                'voucher_no': 'c.cash_book_recod_voucher_no',
                'date': 'c.Payment_Date',
                'paid_to': 'c.cash_book_recode_suplier_name',
                'paid_from': 'c.cash_book_recode_accont_name',
                'narration': 'c.cash_book_recode_naration',
                'amount': 'SUM(c.cash_book_recode_cr)',
                'user_id': 'c.User_Enter',
                'is_reversed': 'CASE WHEN MAX(c.User_Revers) IS NOT NULL THEN 1 ELSE 0 END',
                'reversal_id': 'MAX(c.User_Revers)'
            },
            'where': 'c.jv_numbers_jv_id = %s',
            'group_by': ['c.cash_book_recod_voucher_no', 'c.Payment_Date', 'c.cash_book_recode_suplier_name',
                         'c.cash_book_recode_accont_name', 'c.cash_book_recode_naration', 'c.User_Enter']
        },
        'bank': {
            'title': "BANK PAYMENT VOUCHER",
            'table': 'bank_book_recod b',
            'columns': {
                'voucher_no': 'b.bank_book_recod_voucher_no',
                'date': 'b.Bank_Payment_Date',
                'paid_to': 'b.bank_book__suplier_name',
                'paid_from': 'b.bank_book__accont_name',
                'narration': 'b.bank_book__naration',
                'amount': 'SUM(b.bank_book__recode_cr)',
                'user_id': 'b.Bank_User_Id',
                'cheque_no': 'b.bank_book_chque_no',
                'is_reversed': 'CASE WHEN MAX(b.bank_book_book_recode_dr) > 0 THEN 1 ELSE 0 END',
                'reversal_id': 'MAX(b.bank_book_book_recode_dr)'
            },
            'where': 'b.jv_numbers_jv_id = %s',
            'group_by': ['b.bank_book_recod_voucher_no', 'b.Bank_Payment_Date', 'b.bank_book__suplier_name',
                         'b.bank_book__accont_name', 'b.bank_book__naration', 'b.Bank_User_Id', 'b.bank_book_chque_no']
        },
        'direct': {
            'title': "DIRECT PAYMENT VOUCHER",
            'table': 'cash_book_recode c',
            'columns': {
                'voucher_no': 'c.cash_book_recod_voucher_no',
                'date': 'c.Payment_Date',
                'paid_to': "'Direct Purchase'",
                'paid_from': 'c.cash_book_recode_accont_name',
                'narration': 'c.cash_book_recode_naration',
                'amount': 'SUM(c.cash_book_recode_cr)',
                'user_id': 'c.User_Enter',
                'is_reversed': 'CASE WHEN MAX(c.User_Revers) IS NOT NULL THEN 1 ELSE 0 END',
                'reversal_id': 'MAX(c.User_Revers)'
            },
            'where': 'c.jv_numbers_jv_id = %s',
            'group_by': ['c.cash_book_recod_voucher_no', 'c.Payment_Date',
                         'c.cash_book_recode_accont_name', 'c.cash_book_recode_naration', 'c.User_Enter']
        }
    }

    config = voucher_configs.get(voucher_type)
    if not config:
        return "Invalid Voucher Type", 404

    columns_str = ", ".join([f"{v} as {k}" for k, v in config['columns'].items()])
    group_by_str = ", ".join(config['group_by'])
    query = f"SELECT {columns_str} FROM {config['table']} WHERE {config['where']} GROUP BY {group_by_str}"

    res = db.execute_query(query, (jv_no,))
    if not res:
        return "Voucher Not Found", 404
    voucher = res[0]

    # Change title if reversed
    if voucher.get('is_reversed'):
        config['title'] = "REVERSED PAYMENT VOUCHER"
        # Bank uses the amount itself as 'reversal ID' flag currently; we can format it better if we want, or just show the JV.
        if voucher_type == 'bank':
            voucher['reversal_id'] = f"REV-JV-{jv_no}"
        else:
            voucher['reversal_id'] = f"User: {voucher.get('reversal_id')} (JV: {jv_no})"

    # For direct payments: fetch actual GL entries so we show the real expense account,
    # not the hardcoded "Account Payable" placeholder.
    gl_entries = []
    if voucher_type == 'direct':
        gl_rows = db.execute_query("""
            SELECT account_name,
                   COALESCE(enty_values_DR, 0) AS dr,
                   COALESCE(enty_values_CR, 0) AS cr,
                   entry_naration
            FROM entry_details
            WHERE entry_jv = %s
            ORDER BY enty_values_DR DESC
        """, (jv_no,))
        gl_entries = gl_rows or []

    # Fetch Company Info
    company_res = db.execute_query("SELECT * FROM company LIMIT 1")
    company = company_res[0] if company_res else {}

    if company.get('company_log') and isinstance(company['company_log'], bytes):
        try:
            company['company_log'] = company['company_log'].decode('utf-8')
        except UnicodeDecodeError:
            company['company_log'] = base64.b64encode(company['company_log']).decode('utf-8')

    return render_template('payment_voucher_print.html',
                           voucher=voucher,
                           company=company,
                           title=config['title'],
                           gl_entries=gl_entries)

@app.route('/service_entry/print/<int:jv_no>')
@login_required
def print_service_entry(jv_no):
    # Fetch Header
    header_query = """
        SELECT j.jv_user_code, j.jv_naration, MIN(e.entry_effective_date) as entry_date,
               (SELECT suppliers_invoice_number FROM suppliers_invoice_data WHERE suppliers_invoice_JV = j.jv_id LIMIT 1) as inv_no
        FROM jv_numbers j
        LEFT JOIN entry_details e ON j.jv_id = e.entry_jv
        WHERE j.jv_id = %s
        GROUP BY j.jv_id
    """
    header_res = db.execute_query(header_query, (jv_no,))
    if not header_res:
        return "Entry Not Found", 404
    header = header_res[0]

    # Fetch Details (Debit entries are the services/expenses)
    details_query = """
        SELECT account_name, entry_naration, enty_values_DR, entry_job_number
        FROM entry_details
        WHERE entry_jv = %s AND enty_values_DR > 0
    """
    details = db.execute_query(details_query, (jv_no,))

    # Calculate Total
    total = sum(d['enty_values_DR'] for d in details)

    # Fetch Supplier (Credit entry)
    sup_query = """
        SELECT account_name, enty_values_CR
        FROM entry_details
        WHERE entry_jv = %s AND enty_values_CR > 0 AND account_name = 'Account Payable'
    """
    sup_res = db.execute_query(sup_query, (jv_no,))

    company_res = db.execute_query("SELECT * FROM company LIMIT 1")
    company = company_res[0] if company_res else {}

    return render_template('service_entry_print.html',
                           header=header,
                           details=details,
                           total=total,
                           company=company,
                           jv_id=jv_no)

# --- Purchase Orders ---
@app.route('/purchase_orders', methods=['GET'])
@login_required
@has_permission('Access_Inventory')
def purchase_orders():
    suppliers = db.execute_query("SELECT supplier_name FROM suppliers WHERE Is_Suplier = 1")
    items = db.execute_query("""
        SELECT i.inventoy_name, i.inventoy_items_messurment_unit, p.inventory_price_purcharsing
        FROM inventoy_items i
        LEFT JOIN inventory_price_recod p ON i.id = p.inventory_price_link
    """)
    return render_template('purchase_orders.html', suppliers=suppliers, items=items)

@app.route('/purchase_orders/save', methods=['POST'])
@login_required
def save_purchase_order():
    try:
        supplier = request.form.get('supplier')
        po_number = request.form.get('po_number')
        delivery_date = request.form.get('delivery_date')
        if not delivery_date:
            delivery_date = None
        location = request.form.get('location')
        comments = request.form.get('comments')
        vat_rate = parse_float(request.form.get('vat_rate', 0))
        items_json = request.form.get('items_json')
        items = json.loads(items_json) if items_json else []

        if not items:
            flash('No items in Purchase Order', 'danger')
            return redirect(url_for('purchase_orders'))

        current_user = get_current_user_id()
        current_user_pk = get_current_user_pk()

        conn = db.get_connection()
        cursor = conn.cursor()
        conn.start_transaction()

        try:
            # Get Supplier ID
            cursor.execute("SELECT sup_id FROM suppliers WHERE supplier_name = %s", (supplier,))
            res = cursor.fetchone()
            sup_id = res[0] if res else 0

            # Insert Header
            # Use auto-increment ID, but if PO number is manually provided, store it in OP_NO_Other
            # If no manual PO number, generate one? C# logic seems to allow manual.
            if not po_number:
                # Simple generation if empty
                po_number = f"PO-{datetime.now().strftime('%Y%m%d%H%M%S')}"

            query_header = """
                INSERT INTO OP_NO_Table (
                    OP_NO_Other, Creator_Id, Create_Date, Sup_ID, Sup_Name,
                    Special_Instractions, Expecting_Date, Deliver_Location, VAT_Rate,
                    Save_Post, Delete_PO, Aprove_By, Edit_By, status
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 0, 0, 0, 0, 0)
            """
            cursor.execute(query_header, (
                po_number, current_user_pk, date.today(), sup_id, supplier,
                comments, delivery_date, location, vat_rate
            ))
            po_id = cursor.lastrowid

            # Insert Details
            query_detail = """
                INSERT INTO PO_Recode_Details (
                    Link_OP_NO_Table, Item, Discription, QTY, Unit_price, Mesurment
                ) VALUES (%s, %s, %s, %s, %s, %s)
            """
            batch_data = [(
                po_id, item['item'], item['description'],
                parse_float(item.get('qty', 0)), parse_float(item.get('price', 0)), item['unit']
            ) for item in items]

            if batch_data:
                cursor.executemany(query_detail, batch_data)

            conn.commit()
            flash(f'Purchase Order {po_number} created successfully.', 'success')
        except Exception as e:
            conn.rollback()
            flash(f'Error saving PO: {str(e)}', 'danger')
        finally:
            cursor.close()
            conn.close()

    except Exception as e:
        flash(f'System Error: {str(e)}', 'danger')

    return redirect(url_for('purchase_orders'))

@app.route('/purchase_orders/get/<int:po_id>')
@login_required
def get_purchase_order_details(po_id):
    header_res = db.execute_query("SELECT Sup_Name as supplier FROM OP_NO_Table WHERE id = %s", (po_id,))
    if not header_res:
        return {'error': 'Not Found'}, 404

    items = db.execute_query("""
        SELECT Item as item, QTY as qty, Unit_price as price, Mesurment as unit
        FROM PO_Recode_Details WHERE Link_OP_NO_Table = %s
    """, (po_id,))

    return json.dumps({
        'header': header_res[0],
        'items': [{
            'item': i['item'],
            'qty': float(i['qty'] or 0),
            'price': float(i['price'] or 0),
            'unit': i['unit']
        } for i in items]
    })

@app.route('/purchase_orders/list')
@login_required
def list_purchase_orders():
    query = """
        SELECT
            h.id, h.OP_NO_Other as po_number, h.Create_Date as date,
            h.Sup_Name as supplier, h.Save_Post as approved,
            COALESCE(h.status, 1) as status,
            (SELECT SUM(d.QTY * d.Unit_price) FROM PO_Recode_Details d WHERE d.Link_OP_NO_Table = h.id) as subtotal,
            h.VAT_Rate
        FROM OP_NO_Table h
        WHERE h.Delete_PO = 0
        ORDER BY h.id DESC
    """
    rows = db.execute_query(query)

    data = []
    for r in rows:
        subtotal = float(r['subtotal'] or 0)
        vat = float(r['VAT_Rate'] or 0)
        total = subtotal + (subtotal * vat / 100)
        # IMPORTANT: use explicit None check — 0 is falsy so 'or 1' would wrongly convert pending→approved
        raw_st = r.get('status')
        st = int(raw_st) if raw_st is not None else 1
        # status: 0=pending, 1=approved, 2=rejected
        status_label = {0: 'Pending', 1: 'Approved', 2: 'Rejected'}.get(st, 'Approved')

        data.append({
            'id': r['id'],
            'po_number': r['po_number'],
            'date': str(r['date']),
            'supplier': r['supplier'],
            'approved': st == 1,
            'status': st,
            'status_label': status_label,
            'total': total
        })
    return json.dumps(data)

@app.route('/purchase_orders/approve', methods=['POST'])
@login_required
@has_permission('OP_Approved')
def approve_purchase_order():
    po_id = request.form.get('id')
    current_user = get_current_user_id()
    current_user_pk = get_current_user_pk()

    if po_id:
        db.execute_query("""
            UPDATE OP_NO_Table
            SET Save_Post = 1, status = 1, Aprove_By = %s, Aproed_Date = %s
            WHERE id = %s
        """, (current_user_pk, date.today(), po_id), commit=True)
        return jsonify({'success': True})
    return jsonify({'error': 'No ID provided'}), 400

@app.route('/purchase_orders/reject', methods=['POST'])
@login_required
@has_permission('OP_Approved')
def reject_purchase_order():
    po_id = request.form.get('id')
    current_user_pk = get_current_user_pk()
    if po_id:
        db.execute_query("""
            UPDATE OP_NO_Table
            SET status = 2, Aprove_By = %s, Aproed_Date = %s
            WHERE id = %s
        """, (current_user_pk, date.today(), po_id), commit=True)
        return jsonify({'success': True})
    return jsonify({'error': 'No ID provided'}), 400

@app.route('/purchase_orders/print/<int:po_id>')
@login_required
def print_purchase_order(po_id):
    # Fetch Header
    header_res = db.execute_query("SELECT * FROM OP_NO_Table WHERE id = %s", (po_id,))
    if not header_res:
        return "PO Not Found", 404
    header = header_res[0]

    # Fetch Items
    items = db.execute_query("SELECT * FROM PO_Recode_Details WHERE Link_OP_NO_Table = %s", (po_id,))

    # Fetch Company Info
    company_res = db.execute_query("SELECT * FROM company LIMIT 1")
    company = company_res[0] if company_res else {}
    # Decode binary logo to base64 for inline display
    if company.get('company_log') and isinstance(company['company_log'], bytes):
        try:
            company['company_log'] = company['company_log'].decode('utf-8')
        except Exception:
            company['company_log'] = base64.b64encode(company['company_log']).decode('utf-8')

    # Fetch Supplier Address
    supplier_res = db.execute_query("SELECT * FROM suppliers WHERE sup_id = %s", (header['Sup_ID'],))
    supplier = supplier_res[0] if supplier_res else {}

    # Calculate Totals
    subtotal = sum(float(i['QTY'] or 0) * float(i['Unit_price'] or 0) for i in items)
    vat_rate = float(header['VAT_Rate'] or 0)
    vat_amount = subtotal * vat_rate / 100
    grand_total = subtotal + vat_amount

    return render_template('po_print.html',
                           po=header,
                           items=items,
                           company=company,
                           supplier=supplier,
                           subtotal=subtotal,
                           vat_amount=vat_amount,
                           grand_total=grand_total)

# --- Super Admin Panel ---
from functools import wraps

def superadmin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('superadmin_logged_in'):
            flash('Please log in as Super Admin.', 'danger')
            return redirect(url_for('superadmin_login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/superadmin_login', methods=['GET', 'POST'])
def superadmin_login():
    if request.method == 'GET':
        return render_template('superadmin_login.html')

    username = request.form.get('username')
    password = request.form.get('password')

    # Simple hardcoded credentials for Super Admin
    if username == os.getenv('SUPERADMIN_USERNAME', 'superadmin') and password == os.getenv('SUPERADMIN_PASSWORD', 'superadmin_secret'):
        session['superadmin_pending_2fa'] = True

        # Generate OTP
        import random
        otp = str(random.randint(100000, 999999))
        session['superadmin_otp'] = otp

        # Super Admin phone number (hardcoded for now, or use os.getenv)
        superadmin_phone = os.getenv('SUPERADMIN_PHONE', '0700000000')

        # Try to send SMS
        sms_sent = send_sms_otp(superadmin_phone, otp)
        if not sms_sent:
            flash('Failed to send SMS verification code. Please check your SMS Gateway settings.', 'danger')
            return redirect(url_for('superadmin_login'))

        return redirect(url_for('superadmin_verify'))
    else:
        flash('Invalid Super Admin credentials.', 'danger')
        return redirect(url_for('superadmin_login'))

@app.route('/superadmin_verify', methods=['GET', 'POST'])
def superadmin_verify():
    if not session.get('superadmin_pending_2fa'):
        return redirect(url_for('superadmin_login'))

    if request.method == 'GET':
        return render_template('superadmin_2fa.html')

    otp = request.form.get('otp')
    if otp == session.get('superadmin_otp'):
        session.pop('superadmin_pending_2fa', None)
        session.pop('superadmin_otp', None)
        session['superadmin_logged_in'] = True
        return redirect(url_for('superadmin_dashboard'))
    else:
        flash('Invalid OTP.', 'danger')
        return redirect(url_for('superadmin_verify'))

@app.route('/superadmin/logout')
def superadmin_logout():
    session.pop('superadmin_logged_in', None)
    return redirect(url_for('superadmin_login'))

@app.route('/superadmin/initialize_db/<int:tenant_id>', methods=['POST'])
@superadmin_required
def superadmin_initialize_db(tenant_id):
    try:
        # Fetch tenant and admin user info
        tenant = master_db.execute_query("SELECT company_name, db_name, db_initialized FROM tenants WHERE id = %s", (tenant_id,))
        admin_user = master_db.execute_query("SELECT username, password, email, mobile FROM users WHERE tenant_id = %s ORDER BY id ASC LIMIT 1", (tenant_id,))

        if not tenant or not admin_user:
            flash('Tenant or admin user not found.', 'danger')
            return redirect(url_for('superadmin_dashboard'))

        tenant = tenant[0]
        admin_user = admin_user[0]

        if tenant.get('db_initialized') == 1:
            flash('Database is already initialized.', 'info')
            return redirect(url_for('superadmin_dashboard'))

        db_name = tenant['db_name']

        # 1. Connect to the (already created) cPanel Database
        t_config = db_config.copy()
        t_config['database'] = db_name
        t_conn = mysql.connector.connect(**t_config)
        t_cursor = t_conn.cursor()

        # 2. Execute Schemas
        if os.path.exists('database_schema.sql'):
            with open('database_schema.sql', 'r') as f:
                content = re.sub(r'(?i)Book_keeping', db_name, f.read())
                parse_and_execute_sql(t_cursor, content)

        if os.path.exists('fixed_assets.sql'):
            with open('fixed_assets.sql', 'r') as f:
                content = re.sub(r'(?i)Book_keeping', db_name, f.read())
                parse_and_execute_sql(t_cursor, content)

        t_conn.commit()

        # 3. Run application-level migrations
        run_schema_migrations(t_conn)
        t_conn.close()

        # 4. Insert Admin User and setup initial data
        t_db_conf = db_config.copy()
        t_db_conf['database'] = db_name
        t_db = Database(t_db_conf)

        user_id = t_db.execute_query("""
            INSERT INTO Login_Table (User_Name, Password, Email, Mobile_No, User_Code, User_Active)
            VALUES (%s, %s, %s, %s, '1001', 1)
        """, (admin_user['username'], admin_user['password'], admin_user['email'], admin_user.get('mobile')), commit=True)

        t_db.execute_query("""
            INSERT INTO User_Rights (
                Link_To_Loging_Tabke, Add_New_User, OP_Approved, Access_Inventory,
                Access_POS, Access_Accounting, Access_Reports, Access_Reversals
            )
            VALUES (%s, 1, 1, 1, 1, 1, 1, 1)
        """, (user_id,), commit=True)

        t_db.execute_query("INSERT INTO company (id, company_name) VALUES (1, %s)", (tenant['company_name'],), commit=True)

        ensure_default_categories(t_db)
        ensure_default_accounts(t_db)

        # 5. Mark as initialized
        master_db.execute_query("UPDATE tenants SET db_initialized = 1 WHERE id = %s", (tenant_id,), commit=True)

        flash(f'Database {db_name} initialized successfully.', 'success')

    except mysql.connector.Error as e:
        if e.errno == 1049:
            flash(f'Database {tenant.get("db_name")} does not exist. Please create it in cPanel first.', 'danger')
        elif e.errno == 1045:
            flash(f'Access denied. Please ensure the database user is assigned to {tenant.get("db_name")} with all privileges in cPanel.', 'danger')
        else:
            flash(f'Database error during initialization: {str(e)}', 'danger')
    except Exception as e:
        logging.error(f"Error initializing DB for tenant {tenant_id}: {e}", exc_info=True)
        flash(f'Error initializing database: {str(e)}', 'danger')

    return redirect(url_for('superadmin_dashboard'))

@app.route('/superadmin/toggle_tenant/<int:tenant_id>', methods=['POST'])
@superadmin_required
def superadmin_toggle_tenant(tenant_id):
    try:
        current_status = master_db.execute_query("SELECT is_active FROM tenants WHERE id = %s", (tenant_id,))
        if current_status:
            new_status = 0 if current_status[0]['is_active'] == 1 else 1
            master_db.execute_query("UPDATE tenants SET is_active = %s WHERE id = %s", (new_status, tenant_id), commit=True)
            flash('Tenant status updated successfully.', 'success')
    except Exception as e:
        flash(f'Error updating status: {str(e)}', 'danger')
    return redirect(url_for('superadmin_dashboard'))

@app.route('/superadmin/toggle_sidebar/<int:tenant_id>', methods=['POST'])
@superadmin_required
def superadmin_toggle_sidebar(tenant_id):
    try:
        current_status = master_db.execute_query("SELECT sidebar_enabled FROM tenants WHERE id = %s", (tenant_id,))
        if current_status:
            new_status = 0 if current_status[0].get('sidebar_enabled', 1) == 1 else 1
            master_db.execute_query("UPDATE tenants SET sidebar_enabled = %s WHERE id = %s", (new_status, tenant_id), commit=True)
            flash('Tenant sidebar/functions status updated successfully.', 'success')
    except Exception as e:
        flash(f'Error updating sidebar status: {str(e)}', 'danger')
    return redirect(url_for('superadmin_dashboard'))

@app.route('/superadmin/set_max_users/<int:tenant_id>', methods=['POST'])
@superadmin_required
def superadmin_set_max_users(tenant_id):
    try:
        max_users = int(request.form.get('max_users', 5))
        master_db.execute_query("UPDATE tenants SET max_users = %s WHERE id = %s", (max_users, tenant_id), commit=True)
        flash('Max users limit updated successfully.', 'success')
    except Exception as e:
        flash(f'Error updating max users: {str(e)}', 'danger')
    return redirect(url_for('superadmin_dashboard'))


def _cpanel_weighted_len(name):
    """cPanel stores '_', '%' and '\\' as two characters internally (limit 54)."""
    return sum(2 if c in "_%\\" else 1 for c in name)


@app.route('/superadmin/rename_db/<int:tenant_id>', methods=['POST'])
@superadmin_required
def superadmin_rename_db(tenant_id):
    """Edit a tenant's database name — needed when the auto-generated name has too
    many wildcard-sensitive characters for cPanel (each '_' counts as 2, max 54)."""
    new_name = (request.form.get('db_name') or '').strip()

    if not new_name:
        flash('Database name is required.', 'danger')
        return redirect(url_for('superadmin_dashboard'))
    if not re.match(r'^[a-z][a-z0-9_]*$', new_name):
        flash('Database name must start with a lowercase letter and contain only '
              'lowercase letters, digits and underscores.', 'danger')
        return redirect(url_for('superadmin_dashboard'))

    weighted = _cpanel_weighted_len(new_name)
    if weighted > 54:
        flash(f'Database name is still too long ({weighted}/54 internal characters). '
              f'Remove some underscores or shorten it further.', 'danger')
        return redirect(url_for('superadmin_dashboard'))

    dupe = master_db.execute_query(
        "SELECT id FROM tenants WHERE db_name = %s AND id != %s", (new_name, tenant_id))
    if dupe:
        flash('Another company already uses that database name.', 'danger')
        return redirect(url_for('superadmin_dashboard'))

    row = master_db.execute_query(
        "SELECT db_name, db_initialized FROM tenants WHERE id = %s", (tenant_id,))
    if not row:
        flash('Tenant not found.', 'danger')
        return redirect(url_for('superadmin_dashboard'))
    old_name = row[0]['db_name']
    initialized = row[0].get('db_initialized')

    try:
        master_db.execute_query(
            "UPDATE tenants SET db_name = %s WHERE id = %s", (new_name, tenant_id), commit=True)
        if initialized == 1:
            flash(f'Database name changed from "{old_name}" to "{new_name}". WARNING: this '
                  f'tenant was already initialized — you must also rename the actual MySQL '
                  f'database in cPanel to match, or the tenant will fail to connect.', 'warning')
        else:
            flash(f'Database name changed to "{new_name}". Now create this database in cPanel, '
                  f'grant privileges to the user, then click "Init DB".', 'success')
    except Exception as e:
        flash(f'Error renaming database: {str(e)}', 'danger')
    return redirect(url_for('superadmin_dashboard'))

@app.route('/payment_due')
def payment_due():
    return render_template('payment_due.html')

# ── Superadmin: Per-Tenant Menu Control ──────────────────────────────────────
@app.route('/superadmin/menu/<int:tenant_id>', methods=['GET', 'POST'])
@superadmin_required
def superadmin_menu_control(tenant_id):
    import json as _json

    # ── Ensure schema is present (safe to call multiple times) ──────
    try:
        master_db.execute_query("ALTER TABLE tenants ADD COLUMN menu_config TEXT DEFAULT NULL")
    except Exception:
        pass  # column already exists or other harmless error

    try:
        master_db.execute_query("""
            CREATE TABLE IF NOT EXISTS tenant_custom_menu (
                id            INT AUTO_INCREMENT PRIMARY KEY,
                tenant_id     INT          NOT NULL,
                item_label    VARCHAR(200) NOT NULL,
                item_url      VARCHAR(500) NOT NULL,
                item_icon     VARCHAR(100) DEFAULT 'fas fa-circle',
                item_category VARCHAR(100) DEFAULT 'General',
                sort_order    INT          DEFAULT 99,
                is_enabled    TINYINT(1)   DEFAULT 1,
                created_at    DATETIME     DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_tcm_tenant (tenant_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
    except Exception:
        pass

    # ── Load tenant (use safe column list) ──────────────────────────
    try:
        tenant = master_db.execute_query(
            "SELECT id, company_name, menu_config FROM tenants WHERE id = %s", (tenant_id,)
        )
    except Exception:
        # menu_config column may still not exist — fall back without it
        tenant = master_db.execute_query(
            "SELECT id, company_name FROM tenants WHERE id = %s", (tenant_id,)
        )

    if not tenant:
        flash('Tenant not found.', 'danger')
        return redirect(url_for('superadmin_dashboard'))
    tenant = tenant[0]

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'save_menu':
            cfg = {}
            for item in MENU_ITEMS_REGISTRY:
                cfg[item['key']] = request.form.get(f"item_{item['key']}") == '1'
            try:
                master_db.execute_query(
                    "UPDATE tenants SET menu_config = %s WHERE id = %s",
                    (_json.dumps(cfg), tenant_id), commit=True
                )
                flash('Menu configuration saved.', 'success')
            except Exception as e:
                flash(f'Save failed: {e}', 'danger')

        elif action == 'add_custom':
            label    = request.form.get('custom_label', '').strip()
            url_val  = request.form.get('custom_url', '').strip()
            icon     = request.form.get('custom_icon', 'fas fa-circle').strip()
            category = request.form.get('custom_category', 'General').strip()
            if label and url_val:
                try:
                    master_db.execute_query(
                        "INSERT INTO tenant_custom_menu (tenant_id, item_label, item_url, item_icon, item_category) "
                        "VALUES (%s, %s, %s, %s, %s)",
                        (tenant_id, label, url_val, icon, category), commit=True
                    )
                    flash(f'Custom item "{label}" added.', 'success')
                except Exception as e:
                    flash(f'Add failed: {e}', 'danger')
            else:
                flash('Label and URL are required.', 'danger')

        elif action == 'delete_custom':
            custom_id = request.form.get('custom_id')
            if custom_id:
                try:
                    master_db.execute_query(
                        "DELETE FROM tenant_custom_menu WHERE id = %s AND tenant_id = %s",
                        (custom_id, tenant_id), commit=True
                    )
                    flash('Custom item removed.', 'success')
                except Exception as e:
                    flash(f'Delete failed: {e}', 'danger')

        return redirect(url_for('superadmin_menu_control', tenant_id=tenant_id))

    # GET: load current config
    try:
        cfg = _json.loads(tenant.get('menu_config') or '{}')
    except Exception:
        cfg = {}

    try:
        custom_items = master_db.execute_query(
            "SELECT id, item_label, item_url, item_icon, item_category FROM tenant_custom_menu "
            "WHERE tenant_id = %s ORDER BY item_category, sort_order, id",
            (tenant_id,)
        ) or []
    except Exception:
        custom_items = []

    return render_template(
        'superadmin_menu_control.html',
        tenant=tenant,
        menu_registry=MENU_ITEMS_REGISTRY,
        menu_cfg=cfg,
        custom_items=custom_items
    )


@app.route('/superadmin_dashboard')
@superadmin_required
def superadmin_dashboard():
    # In a real app, this should be protected by a Super Admin login or similar authentication.
    # For now, it's accessible directly to demonstrate the functionality.
    try:
        tenants = master_db.execute_query("""
            SELECT
                t.id,
                t.company_name,
                t.db_name,
                t.is_active,
                t.gb_used,
                t.max_users,
                t.sidebar_enabled,
                t.db_initialized,
                (SELECT COUNT(*) FROM users WHERE tenant_id = t.id) as current_users
            FROM tenants t
        """)
        if tenants is None:
            tenants = []
            if master_db.last_error:
                flash(f"Master DB Connection Error: {master_db.last_error}. Please ensure '{MASTER_DB_NAME}' is created in cPanel.", 'danger')
        return render_template('superadmin_dashboard.html', tenants=tenants)
    except Exception as e:
        flash(f"Error loading superadmin dashboard: {str(e)}", 'danger')
        return redirect(url_for('login'))

# --- Admin Panel & User Management ---
@app.route('/admin/users', methods=['GET'])
@login_required
@has_permission('Add_New_User')
def admin_users():
    users = db.execute_query("SELECT * FROM Login_Table")
    return render_template('admin_panel.html', users=users)

@app.route('/admin/users/add', methods=['POST'])
@login_required
@has_permission('Add_New_User')
def add_new_user():
    username = request.form.get('username')
    password = request.form.get('password')
    confirm_password = request.form.get('confirm_password')
    mobile = request.form.get('mobile')
    email = request.form.get('email')

    # Enforce Tenant User Limits
    current_db_name = session.get('db_name')
    if current_db_name:
        try:
            tenant_info = master_db.execute_query("""
                SELECT t.max_users, (SELECT COUNT(*) FROM users WHERE tenant_id = t.id) as current_users
                FROM tenants t WHERE t.db_name = %s
            """, (current_db_name,))
            if tenant_info:
                t_info = tenant_info[0]
                if t_info['current_users'] >= t_info['max_users']:
                    flash(f"Maximum user limit ({t_info['max_users']}) reached. Please contact support to upgrade.", 'danger')
                    return redirect(url_for('admin_users'))
        except Exception as e:
            logging.error(f"Error checking user limit: {e}")

    if not username or not password:
        flash('Username and Password are required', 'danger')
        return redirect(url_for('admin_users'))

    if password != confirm_password:
        flash('Passwords do not match', 'danger')
        return redirect(url_for('admin_users'))

    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        conn.start_transaction()

        # Hash Password
        from werkzeug.security import generate_password_hash
        pw_hash = generate_password_hash(password)

        # Insert User
        cursor.execute("""
            INSERT INTO Login_Table (User_Name, Password, Mobile_No, Email, User_Active)
            VALUES (%s, %s, %s, %s, 1)
        """, (username, pw_hash, mobile, email))
        user_id = cursor.lastrowid

        # Generate User Code (ID + 50000)
        user_code = user_id + 50000
        cursor.execute("UPDATE Login_Table SET User_Code = %s WHERE id = %s", (user_code, user_id))

        # Insert Default Rights
        # Check if extended columns exist, if not, schema migration handles it or we default
        # We assume columns exist or triggers handle it. But per plan, we insert manually.
        # Note: We need to handle potential missing columns safely or migrate schema.
        # For now, inserting basic rights rows. Extended rights updated via update route.
        cursor.execute("""
            INSERT INTO User_Rights (Link_To_Loging_Tabke, Add_New_User, OP_Approved)
            VALUES (%s, 0, 0)
        """, (user_id,))

        conn.commit()

        # Sync to Master DB for multi-tenant login routing
        try:
            current_db_name = get_session_db_name()
            # Find tenant ID
            tenant_res = master_db.execute_query("SELECT id FROM tenants WHERE db_name = %s", (current_db_name,))
            if tenant_res:
                tenant_id = tenant_res[0]['id']
                master_db.execute_query("""
                    INSERT INTO users (username, password, email, tenant_id)
                    VALUES (%s, %s, %s, %s)
                    ON DUPLICATE KEY UPDATE password=VALUES(password), email=VALUES(email)
                """, (username, pw_hash, email, tenant_id), commit=True)
        except Exception as master_e:
            logging.error("Error syncing new user to master DB")

        flash(f'User {username} created successfully.', 'success')
    except Exception as e:
        if 'conn' in locals() and conn:
            conn.rollback()
        logging.error("Database error while creating user.")
        flash(f'Error creating user: {str(e)}', 'danger')
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'conn' in locals() and conn:
            conn.close()

    return redirect(url_for('admin_users'))

@app.route('/admin/users/details/<int:user_id>', methods=['GET'])
@login_required
@has_permission('Add_New_User')
def get_user_details(user_id):
    users = db.execute_query("SELECT id, User_Name, Password, Mobile_No, Email, User_Active FROM Login_Table WHERE id = %s", (user_id,))
    if users:
        return jsonify(users[0])
    return jsonify({'error': 'User not found'}), 404

@app.route('/admin/users/update_details', methods=['POST'])
@login_required
@has_permission('Add_New_User')
def update_user_details():
    user_id = request.form.get('user_id')
    username = request.form.get('username')
    password = request.form.get('password')
    mobile = request.form.get('mobile')
    email = request.form.get('email')
    active = 1 if request.form.get('active') else 0

    if not user_id or not username:
        flash('Username is required', 'danger')
        return redirect(url_for('admin_users'))

    try:
        # If password provided, hash it. Else keep existing.
        if password:
            from werkzeug.security import generate_password_hash
            pw_hash = generate_password_hash(password)
            query = """
                UPDATE Login_Table
                SET User_Name = %s, Password = %s, Mobile_No = %s, Email = %s, User_Active = %s
                WHERE id = %s
            """
            params = (username, pw_hash, mobile, email, active, user_id)
        else:
            query = """
                UPDATE Login_Table
                SET User_Name = %s, Mobile_No = %s, Email = %s, User_Active = %s
                WHERE id = %s
            """
            params = (username, mobile, email, active, user_id)

        db.execute_query(query, params, commit=True)
        # Check if password needs update (if provided)
        # Note: Frontend currently sends password field. If it's intended to be updated only when changed, logic should be handled.
        # However, the provided form logic in `admin_panel.html` (not visible here but implied) likely sends value.
        # If the password field is populated, we assume it's a new password.
        # If we want to support "leave blank to keep existing", we need to check if password is empty.
        # Assuming typical admin panel behavior: empty = no change.

        if password:
            pw_hash = generate_password_hash(password)
            db.execute_query("""
                UPDATE Login_Table
                SET User_Name = %s, Password = %s, Mobile_No = %s, Email = %s, User_Active = %s
                WHERE id = %s
            """, (username, pw_hash, mobile, email, active, user_id), commit=True)
        else:
            db.execute_query("""
                UPDATE Login_Table
                SET User_Name = %s, Mobile_No = %s, Email = %s, User_Active = %s
                WHERE id = %s
            """, (username, mobile, email, active, user_id), commit=True)

        # Sync update to Master DB
        try:
            current_db_name = get_session_db_name()
            tenant_res = master_db.execute_query("SELECT id FROM tenants WHERE db_name = %s", (current_db_name,))
            if tenant_res:
                tenant_id = tenant_res[0]['id']
                if password:
                    master_db.execute_query("""
                        UPDATE users SET username = %s, password = %s, email = %s
                        WHERE username = (SELECT User_Name FROM (SELECT User_Name FROM Login_Table WHERE id = %s) as t) AND tenant_id = %s
                    """, (username, pw_hash, email, user_id, tenant_id), commit=True)
                else:
                    master_db.execute_query("""
                        UPDATE users SET username = %s, email = %s
                        WHERE username = (SELECT User_Name FROM (SELECT User_Name FROM Login_Table WHERE id = %s) as t) AND tenant_id = %s
                    """, (username, email, user_id, tenant_id), commit=True)
        except Exception as master_e:
            logging.error("Error syncing user update to master DB")

        flash('User details updated successfully', 'success')
    except Exception as e:
        logging.error("Database error while updating user.")
        flash('Error updating user. Please try again.', 'danger')

    return redirect(url_for('admin_users'))

@app.route('/admin/users/rights/<int:user_id>', methods=['GET'])
@login_required
@has_permission('Add_New_User')
def get_user_rights(user_id):
    rights = db.execute_query("SELECT * FROM User_Rights WHERE Link_To_Loging_Tabke = %s", (user_id,))
    if rights:
        return jsonify(rights[0])
    return jsonify({})

@app.route('/admin/users/reset_password', methods=['GET', 'POST'])
@login_required
def admin_reset_password():
    if request.method == 'GET':
        return render_template('admin_reset_password.html')
    else:
        # 1. Send OTP to registered mobile number for logged in user
        current_user_id = session.get('user_id')
        user = db.execute_query("SELECT Mobile_No FROM Login_Table WHERE User_Code = %s", (current_user_id,))
        if not user or not user[0]['Mobile_No']:
            flash('No mobile number registered for your account. Please contact admin.', 'danger')
            return redirect(url_for('dashboard'))

        mobile = user[0]['Mobile_No']
        import random
        otp = str(random.randint(100000, 999999))

        # Save OTP to session
        session['pwd_reset_otp'] = otp
        session['pwd_reset_mobile'] = mobile

        # Send SMS
        sms_sent = send_sms_otp(mobile, otp)

        if sms_sent:
            flash(f'An OTP has been sent to your registered mobile number: {mobile[:3]}****{mobile[-3:]}', 'info')
            return render_template('admin_verify_reset.html')
        else:
            flash('Failed to send OTP. Please try again later.', 'danger')
            return redirect(url_for('dashboard'))

@app.route('/admin/users/verify_reset', methods=['POST'])
@login_required
def admin_verify_reset():
    entered_otp = request.form.get('otp')
    new_password = request.form.get('new_password')
    confirm_password = request.form.get('confirm_password')

    saved_otp = session.get('pwd_reset_otp')

    if not saved_otp or entered_otp != saved_otp:
        flash('Invalid OTP. Please try again.', 'danger')
        return render_template('admin_verify_reset.html')

    if new_password != confirm_password:
        flash('Passwords do not match.', 'danger')
        return render_template('admin_verify_reset.html')

    # Update Password
    current_user_id = session.get('user_id')
    from werkzeug.security import generate_password_hash
    pw_hash = generate_password_hash(new_password)

    try:
        db.execute_query("UPDATE Login_Table SET Password = %s WHERE User_Code = %s", (pw_hash, current_user_id), commit=True)

        # Sync password change to Master DB users table
        try:
            current_db_name = get_session_db_name()
            tenant_res = master_db.execute_query("SELECT id FROM tenants WHERE db_name = %s", (current_db_name,))
            if tenant_res:
                tenant_id = tenant_res[0]['id']
                # Get the username to update the correct master DB record
                user_res = db.execute_query("SELECT User_Name FROM Login_Table WHERE User_Code = %s", (current_user_id,))
                if user_res:
                    username = user_res[0]['User_Name']
                    master_db.execute_query(
                        "UPDATE users SET password = %s WHERE username = %s AND tenant_id = %s",
                        (pw_hash, username, tenant_id),
                        commit=True
                    )
        except Exception as e:
            logging.error(f"Error syncing password reset to master DB: {e}")

        # Clear session
        session.pop('pwd_reset_otp', None)
        session.pop('pwd_reset_mobile', None)

        flash('Password reset successfully. Please log in with your new password.', 'success')
        return redirect(url_for('logout'))

    except Exception as e:
        logging.error(f"Error resetting password: {e}")
        flash('Error resetting password. Please try again.', 'danger')
        return render_template('admin_verify_reset.html')

@app.route('/admin/users/rights/update', methods=['POST'])
@login_required
@has_permission('Add_New_User')
def update_user_rights():
    user_id = request.form.get('user_id', '').strip()
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if not user_id:
        if is_ajax: return jsonify({'error': 'No User ID'}), 400
        flash('No user selected', 'danger')
        return redirect(url_for('admin_users'))

    all_perms = [
        'Add_New_User', 'OP_Approved',
        'Access_Sales', 'Access_Purchase', 'Access_Accounting',
        'Access_Inventory', 'Access_POS',
        'Access_Reports', 'Access_Fixed_Assets', 'Access_VAT',
        'Access_Reversals', 'Access_HR', 'Access_CRM', 'Access_Settings'
    ]
    perm_values = {p: 1 if request.form.get(p) else 0 for p in all_perms}

    # Discover which columns exist to skip ones not yet migrated
    try:
        col_rows = db.execute_query("SHOW COLUMNS FROM User_Rights") or []
        # execute_query returns list of dicts; col name is key 'Field'
        existing_cols = set()
        for r in col_rows:
            if isinstance(r, dict):
                existing_cols.add(r.get('Field') or r.get('field') or list(r.values())[0])
            else:
                existing_cols.add(r[0])
    except Exception as e:
        logging.error(f"SHOW COLUMNS failed: {e}")
        existing_cols = set(all_perms)

    valid_perms = [p for p in all_perms if p in existing_cols] or all_perms
    set_clause  = ', '.join([f"`{p}` = %s" for p in valid_perms])
    params      = tuple(perm_values[p] for p in valid_perms) + (user_id,)

    try:
        db.execute_query(
            f"UPDATE User_Rights SET {set_clause} WHERE Link_To_Loging_Tabke = %s",
            params, commit=True)
        if is_ajax:
            return jsonify({'success': True})
        flash('Access rights saved successfully.', 'success')
        return redirect(url_for('admin_users'))
    except Exception as e:
        logging.error(f"Rights Update Error: {e}")
        if is_ajax:
            return jsonify({'error': str(e)}), 500
        flash(f'Error saving rights: {str(e)}', 'danger')
        return redirect(url_for('admin_users'))

# --- Job Management ---
@app.route('/job_management', methods=['GET'])
@login_required
def job_management():
    # Fetch all jobs with status
    jobs = db.execute_query("SELECT * FROM jobs_unit ORDER BY job_number DESC")
    return render_template('job_management.html', jobs=jobs)

@app.route('/create_job', methods=['GET', 'POST'])
@app.route('/jobs/create', methods=['POST'])
@login_required
def create_job():
    if request.method == 'GET':
        return render_template('add_new_job.html')

    job_no = request.form.get('job_no')
    description = request.form.get('job_description')

    if not job_no or not description:
        flash('Job No and Description are required', 'danger')
        return redirect(url_for('job_management'))

    current_user = get_current_user_id()

    try:
        db.execute_query("""
            INSERT INTO jobs_unit (id, job_number, job_description, job_create_date, job_create_user, job_finsh, job_cancell)
            VALUES (0, %s, %s, %s, %s, 0, 0)
        """, (job_no, description, date.today(), current_user), commit=True)
        flash('New job created successfully', 'success')
    except Exception as e:
        flash(f'Error creating job: {str(e)}', 'danger')

    return redirect(url_for('job_management'))

@app.route('/jobs/toggle_status', methods=['POST'])
@login_required
def toggle_job_status():
    job_id = request.form.get('job_id')
    status = request.form.get('status') # 1 = Finish, 0 = Active

    try:
        db.execute_query("UPDATE jobs_unit SET job_finsh = %s WHERE id = %s", (status, job_id), commit=True)
        msg = "Job finished" if str(status) == "1" else "Job re-activated"
        flash(f'{msg} successfully', 'success')
    except Exception as e:
        flash(f'Error updating job: {str(e)}', 'danger')

    return redirect(url_for('job_management'))

# --- Warranty Period Management ---
@app.route('/warranty_period', methods=['GET'])
@login_required
@has_permission('Access_Inventory')
def warranty_period():
    item_name = request.args.get('item_name')

    items = db.execute_query("SELECT inventoy_name FROM inventoy_items WHERE active = 1 ORDER BY inventoy_name")

    query = """
        SELECT wp.id, wp.yeas_, wp.month, wp.date_, ii.inventoy_name as name
        FROM inventory_vorenty_period wp
        LEFT JOIN inventoy_items ii ON wp.name = ii.inventoy_name
    """
    params = None
    if item_name:
        query += " WHERE ii.inventoy_name = %s"
        params = (item_name,)

    query += " ORDER BY ii.inventoy_name"

    warranty_items = db.execute_query(query, params)

    return render_template('warranty_period.html', items=items, warranty_items=warranty_items, selected_item=item_name)

@app.route('/warranty_save', methods=['POST'])
@login_required
def warranty_save():
    ids = request.form.getlist('id[]')
    names = request.form.getlist('name[]')
    years = request.form.getlist('year[]')
    months = request.form.getlist('month[]')
    days = request.form.getlist('day[]')

    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        conn.start_transaction()

        for i in range(len(ids)):
            wid = int(ids[i])
            if wid == 0: # Insert
                cursor.execute("""
                    INSERT INTO inventory_vorenty_period (yeas_, month, date_, name)
                    VALUES (%s, %s, %s, %s)
                """, (years[i], months[i], days[i], names[i]))
            else: # Update
                cursor.execute("""
                    UPDATE inventory_vorenty_period
                    SET yeas_ = %s, month = %s, date_ = %s, name = %s
                    WHERE id = %s
                """, (years[i], months[i], days[i], names[i], wid))

        conn.commit()
        cursor.close()
        conn.close()
        flash('Warranty data saved successfully', 'success')
    except Exception as e:
        flash(f'Error saving data: {str(e)}', 'danger')

    return redirect(url_for('warranty_period'))

# --- Inventory Trend Analysis ---
@app.route('/inventory_trend_analysis', methods=['GET'])
@login_required
@has_permission('Access_Inventory')
def inventory_trend_analysis():
    item_name = request.args.get('item_name')
    months_back = int(request.args.get('months', 6))

    items = db.execute_query("SELECT DISTINCT inventoy_name FROM inventoy_items WHERE inventoy_name IS NOT NULL AND inventoy_name != '' ORDER BY inventoy_name")

    trend_data = []
    trend_direction = "Stable"
    slope_val = 0
    forecast = 0

    if item_name:
        # Fetch Data
        raw_data = db.execute_query("""
            SELECT
                YEAR(inventory_recod_action_date) as Year,
                MONTH(inventory_recod_action_date) as Month,
                SUM(inventory_recod_movment_out) as MonthlySales
            FROM inventory_recod
            WHERE inventoy_name = %s
            AND inventory_recod_action_date >= DATE_SUB(NOW(), INTERVAL %s MONTH)
            AND inventory_recod_movment_out > 0
            GROUP BY YEAR(inventory_recod_action_date), MONTH(inventory_recod_action_date)
            ORDER BY Year, Month
        """, (item_name, months_back))

        if raw_data:
            # Prepare Lists
            sales = [float(r.get('MonthlySales', 0)) for r in raw_data]
            n = len(sales)

            # Moving Average (3 months)
            moving_avgs = []
            for i in range(n):
                if i < 2:
                    val = sum(sales[:i+1]) / (i+1)
                else:
                    val = sum(sales[i-2:i+1]) / 3
                moving_avgs.append(val)

            # Trend Analysis (Linear Regression)
            sumX = sum(range(n))
            sumY = sum(sales)
            sumXY = sum(i * sales[i] for i in range(n))
            sumX2 = sum(i * i for i in range(n))

            if n > 1 and (n * sumX2 - sumX * sumX) != 0:
                slope = (n * sumXY - sumX * sumY) / (n * sumX2 - sumX * sumX)
                intercept = (sumY - slope * sumX) / n
            else:
                slope = 0
                intercept = sales[0] if n > 0 else 0

            slope_val = round(slope, 2)
            trend_direction = "Increasing" if slope > 0 else "Decreasing" if slope < 0 else "Stable"
            forecast = round(intercept + slope * n, 1)

            # Combine for Template
            for i, r in enumerate(raw_data):
                trend_val = intercept + slope * i
                trend_data.append({
                    'Period': f"{r['Year']}-{r['Month']:02d}",
                    'Year': r['Year'],
                    'Month': r['Month'],
                    'SalesQuantity': r['MonthlySales'],
                    'MovingAverage': round(moving_avgs[i], 2),
                    'TrendValue': round(trend_val, 2)
                })

    return render_template('inventory_trend_analysis.html',
                           items=items,
                           trend_data=trend_data,
                           selected_item=item_name,
                           months=months_back,
                           trend_direction=trend_direction,
                           slope=slope_val,
                           next_month_forecast=forecast)

@app.route('/api/predict_account_type')
@login_required
def predict_account_type():
    account_name = request.args.get('name', '')
    if not account_name:
        return {'error': 'No name provided'}

    # Get matches
    matches = difflib.get_close_matches(account_name, knowledge_base.account_types.keys(), n=1, cutoff=0.6)

    if matches:
        match = matches[0]
        # Calculate a simple similarity ratio
        ratio = difflib.SequenceMatcher(None, account_name.lower(), match.lower()).ratio()

        # User requested > 75% probability logic (here mapped to ratio > 0.75)
        # Note: get_close_matches uses cutoff, but we check ratio manually for specific threshold if needed
        # The prompt said "if probability less than 75% connect the internet".
        # Since we don't have internet, we only return if confidence is high enough or return low confidence.

        predicted_type = knowledge_base.account_types[match]

        # Map knowledge base types to checkbox values
        # Knowledge Base: "Assets Account", "Cost Account", "Equity Accont", "Income Account", "Liabilities Account"
        # Checkbox values: "asset", "expense", "equity", "income", "liability"

        type_map = {
            "Assets Account": "asset",
            "Fixed Asset": "asset",
            "Intangible Assets": "asset",
            "Cost Account": "expense",
            "Equity Accont": "equity",
            "Equity Account": "equity",
            "Income Account": "income",
            "Liabilities Account": "liability",
            "Liabilities Accounts": "liability"
        }

        mapped_type = type_map.get(predicted_type, "")

        return {
            'match': match,
            'original_type': predicted_type,
            'mapped_type': mapped_type,
            'confidence': ratio
        }

    return {'confidence': 0}

@app.route('/api/get_customers')
@login_required
def api_get_customers():
    query = "SELECT id, customer_name as name FROM customer ORDER BY customer_name"
    rows = db.execute_query(query)
    return json.dumps(rows)

@app.route('/api/get_all_accounts')
@login_required
def api_get_all_accounts():
    """Return all active account names for dropdowns (e.g. TB correction modal)."""
    rows = db.execute_query(
        "SELECT account_name FROM new_account_table WHERE account_active = 1 ORDER BY account_name")
    return jsonify([{'account_name': r['account_name']} for r in rows] if rows else [])

@app.route('/api/get_sub_accounts')
@login_required
def api_get_sub_accounts():
    account_name = request.args.get('account_name')
    if not account_name:
        return json.dumps([])

    query = """
        SELECT sub_account_code as code, sub_sub_accaount_name as name
        FROM sub_accont_for_new_account
        WHERE sub_new_account = %s AND active = 1
        ORDER BY sub_sub_accaount_name
    """
    rows = db.execute_query(query, (account_name,))
    return json.dumps(rows)


@app.route('/sub_account_allocation', methods=['GET'])
@login_required
@has_permission('Access_Accounting')
def sub_account_allocation():
    """Tag existing transactions of an account with a sub-account, so historical
    amounts move out of '(Unallocated)' into the right sub on the P&L / BS."""
    accounts = db.execute_query("""
        SELECT DISTINCT s.sub_new_account AS account_name
        FROM sub_accont_for_new_account s
        JOIN new_account_table na ON na.account_name = s.sub_new_account
        WHERE s.active = 1 AND na.account_active = 1
        ORDER BY s.sub_new_account
    """) or []

    selected = (request.args.get('account') or '').strip()
    from_date = (request.args.get('from') or '').strip()
    to_date = (request.args.get('to') or '').strip()
    subs, txns = [], []
    if selected:
        subs = db.execute_query("""
            SELECT sub_account_code AS code, sub_sub_accaount_name AS name
            FROM sub_accont_for_new_account
            WHERE sub_new_account = %s AND active = 1
            ORDER BY sub_sub_accaount_name
        """, (selected,)) or []
        sub_name_by_code = {str(s['code']): s['name'] for s in subs}

        q = """
            SELECT id, entry_effective_date AS date, entry_naration AS narration,
                   COALESCE(enty_values_DR, 0) AS dr, COALESCE(enty_values_CR, 0) AS cr,
                   entry_jv AS jv, COALESCE(entry_sub_account_code, 0) AS sub_code
            FROM entry_details
            WHERE account_name = %s AND COALESCE(entry_deleted, 0) = 0
        """
        params = [selected]
        if from_date:
            q += " AND entry_effective_date >= %s"; params.append(from_date)
        if to_date:
            q += " AND entry_effective_date <= %s"; params.append(to_date)
        q += " ORDER BY entry_effective_date DESC, id DESC"
        for r in (db.execute_query(q, tuple(params)) or []):
            code = str(r['sub_code'])
            txns.append({'id': r['id'], 'date': r['date'], 'narration': r['narration'],
                         'dr': float(r['dr'] or 0), 'cr': float(r['cr'] or 0), 'jv': r['jv'],
                         'sub_code': '' if code == '0' else code})

    return render_template('sub_account_allocation.html', accounts=accounts, subs=subs,
                           txns=txns, selected=selected, from_date=from_date, to_date=to_date)


@app.route('/sub_account_allocation/save', methods=['POST'])
@login_required
@has_permission('Access_Accounting')
def sub_account_allocation_save():
    data = request.json or {}
    updated = 0
    for a in (data.get('allocations') or []):
        eid = a.get('id')
        if not eid:
            continue
        code = a.get('sub_code') or 0
        db.execute_query("UPDATE entry_details SET entry_sub_account_code = %s WHERE id = %s",
                         (code, eid), commit=True)
        updated += 1
    return jsonify({'success': True, 'updated': updated})


@app.route('/supplier_payments_report')
@login_required
@has_permission('Access_Reports')
def supplier_payments_report():
    """Payments made to suppliers: supplier, amount, method (Cash/Bank) and which
    bank/cash account. Combines cash_book_recode and bank_book_recod."""
    from_date = (request.args.get('from') or date.today().replace(day=1).strftime('%Y-%m-%d')).strip()
    to_date = (request.args.get('to') or date.today().strftime('%Y-%m-%d')).strip()
    supplier = (request.args.get('supplier') or '').strip()
    method = (request.args.get('method') or '').strip()  # '', 'Cash', 'Bank'
    download = request.args.get('download')

    rows = []

    if method in ('', 'Cash'):
        q = """
            SELECT cash_book_recode_suplier_name AS supplier, cash_book_recode_cr AS amount,
                   cash_book_recode_accont_name AS account, cash_book_recod_voucher_no AS voucher,
                   Payment_Date AS pdate, jv_numbers_jv_id AS jv
            FROM cash_book_recode
            WHERE COALESCE(cash_book_recode_cr, 0) > 0
              AND cash_book_recode_suplier_name IS NOT NULL AND cash_book_recode_suplier_name <> ''
              AND Payment_Date BETWEEN %s AND %s
        """
        params = [from_date, to_date]
        if supplier:
            q += " AND cash_book_recode_suplier_name = %s"; params.append(supplier)
        try:
            cash = db.execute_query(q + " AND (User_Revers IS NULL OR User_Revers = '' OR User_Revers = '0')",
                                    tuple(params))
        except Exception:
            cash = db.execute_query(q, tuple(params))
        for r in (cash or []):
            rows.append({'supplier': r['supplier'], 'amount': float(r['amount'] or 0),
                         'method': 'Cash', 'account': r['account'] or '', 'cheque': '',
                         'voucher': r['voucher'], 'date': r['pdate'], 'jv': r['jv']})

    if method in ('', 'Bank'):
        q = """
            SELECT bank_book__suplier_name AS supplier, bank_book__recode_cr AS amount,
                   bank_book__accont_name AS account, bank_book_recod_voucher_no AS voucher,
                   bank_book_chque_no AS cheque, Bank_Payment_Date AS pdate, jv_numbers_jv_id AS jv
            FROM bank_book_recod
            WHERE COALESCE(bank_book__recode_cr, 0) > 0
              AND bank_book__suplier_name IS NOT NULL AND bank_book__suplier_name <> ''
              AND Bank_Payment_Date BETWEEN %s AND %s
              AND COALESCE(bank_book_book_recode_dr, 0) = 0
        """
        params = [from_date, to_date]
        if supplier:
            q += " AND bank_book__suplier_name = %s"; params.append(supplier)
        bank = db.execute_query(q, tuple(params))
        for r in (bank or []):
            rows.append({'supplier': r['supplier'], 'amount': float(r['amount'] or 0),
                         'method': 'Bank', 'account': r['account'] or '', 'cheque': r['cheque'] or '',
                         'voucher': r['voucher'], 'date': r['pdate'], 'jv': r['jv']})

    rows.sort(key=lambda x: str(x['date'] or ''), reverse=True)
    total = sum(r['amount'] for r in rows)
    total_cash = sum(r['amount'] for r in rows if r['method'] == 'Cash')
    total_bank = sum(r['amount'] for r in rows if r['method'] == 'Bank')

    if download == 'csv':
        si = io.StringIO()
        cw = csv.writer(si)
        cw.writerow(['Supplier Payments', f'{from_date} to {to_date}'])
        cw.writerow(['Date', 'Supplier', 'Method', 'Bank/Cash Account', 'Cheque No', 'Voucher', 'JV', 'Amount'])
        for r in rows:
            cw.writerow([r['date'], r['supplier'], r['method'], r['account'], r['cheque'],
                         r['voucher'], r['jv'], f"{r['amount']:.2f}"])
        cw.writerow([])
        cw.writerow(['', '', '', '', '', '', 'Cash Total', f"{total_cash:.2f}"])
        cw.writerow(['', '', '', '', '', '', 'Bank Total', f"{total_bank:.2f}"])
        cw.writerow(['', '', '', '', '', '', 'Grand Total', f"{total:.2f}"])
        out = make_response(si.getvalue())
        out.headers["Content-Disposition"] = f"attachment; filename=Supplier_Payments_{from_date}_{to_date}.csv"
        out.headers["Content-type"] = "text/csv"
        return out

    suppliers = db.execute_query("SELECT supplier_name FROM suppliers WHERE Is_Suplier = 1 ORDER BY supplier_name") or []
    return render_template('supplier_payments_report.html', rows=rows, total=total,
                           total_cash=total_cash, total_bank=total_bank, from_date=from_date,
                           to_date=to_date, supplier=supplier, method=method, suppliers=suppliers)


@app.route('/grn_payment_method_report')
@login_required
@has_permission('Access_Reports')
def grn_payment_method_report():
    """Supplier purchase invoices report: invoice date, supplier name,
    amount, and the payment method (Cash/Cheque) — lets the user see at a
    glance which supplier invoices are meant to be paid by cash vs cheque.
    Covers EVERY supplier invoice (GRN, Direct Purchase, Service Entry,
    Opening Balance, ...), not just GRN — the payment method concept
    applies the same way regardless of how the invoice was created."""
    from_date = (request.args.get('from') or date.today().replace(day=1).strftime('%Y-%m-%d')).strip()
    to_date = (request.args.get('to') or date.today().strftime('%Y-%m-%d')).strip()
    supplier = (request.args.get('supplier') or '').strip()
    method = (request.args.get('method') or '').strip()  # '', 'Cash', 'Cheque'
    download = request.args.get('download')

    q = """
        SELECT s.suppliers_invoice_date AS inv_date, sup.supplier_name AS supplier,
               s.suppliers_invoice_number AS invoice_no, s.suppliers_invoice_total_oustanding AS amount,
               s.suppliers_invoice_payment_method AS method, s.suppliers_invoice_JV AS jv,
               jv.jv_user_code AS jv_user_code
        FROM suppliers_invoice_data s
        LEFT JOIN jv_numbers jv ON s.suppliers_invoice_JV = jv.jv_id
        LEFT JOIN suppliers sup ON s.suppliers_invoice_buinding_supplier = sup.sup_id
        WHERE COALESCE(s.suppliers_oustanding_delete, 0) = 0
          AND s.suppliers_invoice_date BETWEEN %s AND %s
    """
    params = [from_date, to_date]
    if supplier:
        q += " AND sup.supplier_name = %s"; params.append(supplier)
    if method:
        q += " AND s.suppliers_invoice_payment_method = %s"; params.append(method)
    q += " ORDER BY s.suppliers_invoice_date DESC, s.suppliers_invoice_JV DESC"

    rows = db.execute_query(q, tuple(params)) or []
    for r in rows:
        r['amount'] = float(r['amount'] or 0)
        r['method'] = r['method'] or 'Not Set'
        r['is_grn'] = r.get('jv_user_code') == 'JV FROM GRN'

    total = sum(r['amount'] for r in rows)
    total_cash = sum(r['amount'] for r in rows if r['method'] == 'Cash')
    total_cheque = sum(r['amount'] for r in rows if r['method'] == 'Cheque')

    if download == 'csv':
        si = io.StringIO()
        cw = csv.writer(si)
        cw.writerow(['GRN Payment Method Report', f'{from_date} to {to_date}'])
        cw.writerow(['Invoice Date', 'Supplier', 'Invoice No', 'Payment Method', 'Amount'])
        for r in rows:
            cw.writerow([r['inv_date'], r['supplier'], r['invoice_no'], r['method'], f"{r['amount']:.2f}"])
        cw.writerow([])
        cw.writerow(['', '', '', 'Cash Total', f"{total_cash:.2f}"])
        cw.writerow(['', '', '', 'Cheque Total', f"{total_cheque:.2f}"])
        cw.writerow(['', '', '', 'Grand Total', f"{total:.2f}"])
        out = make_response(si.getvalue())
        out.headers["Content-Disposition"] = f"attachment; filename=GRN_Payment_Method_{from_date}_{to_date}.csv"
        out.headers["Content-type"] = "text/csv"
        return out

    suppliers = db.execute_query("SELECT supplier_name FROM suppliers WHERE Is_Suplier = 1 ORDER BY supplier_name") or []
    return render_template('grn_payment_method_report.html', rows=rows, total=total,
                           total_cash=total_cash, total_cheque=total_cheque, from_date=from_date,
                           to_date=to_date, supplier=supplier, method=method, suppliers=suppliers)


def _supplier_purchasing_rows(from_date, to_date, supplier):
    """Every supplier invoice in the period (GRN, Direct Purchase, Service Entry,
    Opening Balance, ...) with Gross / VAT / Net split back out.

    suppliers_invoice_data only stores the VAT-inclusive grand total
    (suppliers_invoice_total_oustanding) plus the VAT RATE — the VAT amount
    itself is never stored — so Net/VAT are derived from the rate the same
    way the GRN posting logic (services.create_grn) already does it:
    Net = Gross / (1 + rate/100), VAT = Gross - Net."""
    q = """
        SELECT s.suppliers_invoice_date AS inv_date, sup.supplier_name AS supplier,
               s.suppliers_invoice_number AS invoice_no, s.suppliers_invoice_total_oustanding AS gross,
               s.suppliers_VAT_rate AS vat_rate, s.suppliers_invoice_JV AS jv,
               jv.jv_user_code AS jv_user_code
        FROM suppliers_invoice_data s
        LEFT JOIN jv_numbers jv ON s.suppliers_invoice_JV = jv.jv_id
        LEFT JOIN suppliers sup ON s.suppliers_invoice_buinding_supplier = sup.sup_id
        WHERE COALESCE(s.suppliers_oustanding_delete, 0) = 0
          AND s.suppliers_invoice_date BETWEEN %s AND %s
    """
    params = [from_date, to_date]
    if supplier:
        q += " AND sup.supplier_name = %s"; params.append(supplier)
    q += " ORDER BY s.suppliers_invoice_date DESC, s.suppliers_invoice_JV DESC"

    rows = db.execute_query(q, tuple(params)) or []
    for r in rows:
        gross = float(r['gross'] or 0)
        vat_rate = float(r['vat_rate'] or 0)
        if vat_rate > 0 and gross > 0:
            net = gross / (1 + (vat_rate / 100))
            vat = gross - net
        else:
            net = gross
            vat = 0.0
        r['gross'] = round(gross, 2)
        r['vat'] = round(vat, 2)
        r['net'] = round(net, 2)
        r['vat_rate'] = vat_rate
        r['is_grn'] = r.get('jv_user_code') == 'JV FROM GRN'

    totals = {
        'gross': sum(r['gross'] for r in rows),
        'vat': sum(r['vat'] for r in rows),
        'net': sum(r['net'] for r in rows),
    }
    return rows, totals


@app.route('/supplier_purchasing_report')
@login_required
@has_permission('Access_Reports')
def supplier_purchasing_report():
    """Supplier purchasing report — Gross / VAT / Net for every supplier
    invoice in a selected date period. Covers GRN, Direct Purchase, Service
    Entry and Opening Balance invoices alike (same coverage as the Supplier
    Payment Method report)."""
    from_date = (request.args.get('from') or date.today().replace(day=1).strftime('%Y-%m-%d')).strip()
    to_date = (request.args.get('to') or date.today().strftime('%Y-%m-%d')).strip()
    supplier = (request.args.get('supplier') or '').strip()
    download = request.args.get('download')

    rows, totals = _supplier_purchasing_rows(from_date, to_date, supplier)

    if download == 'csv':
        si = io.StringIO()
        cw = csv.writer(si)
        cw.writerow(['Supplier Purchasing Report', f'{from_date} to {to_date}'])
        cw.writerow(['Invoice Date', 'Supplier', 'Invoice No', 'VAT Rate %', 'Gross', 'VAT', 'Net'])
        for r in rows:
            cw.writerow([r['inv_date'], r['supplier'], r['invoice_no'], r['vat_rate'],
                         f"{r['gross']:.2f}", f"{r['vat']:.2f}", f"{r['net']:.2f}"])
        cw.writerow([])
        cw.writerow(['', '', '', 'Total', f"{totals['gross']:.2f}", f"{totals['vat']:.2f}", f"{totals['net']:.2f}"])
        out = make_response(si.getvalue())
        out.headers["Content-Disposition"] = f"attachment; filename=Supplier_Purchasing_Report_{from_date}_{to_date}.csv"
        out.headers["Content-type"] = "text/csv"
        return out

    suppliers = db.execute_query("SELECT supplier_name FROM suppliers WHERE Is_Suplier = 1 ORDER BY supplier_name") or []
    return render_template('supplier_purchasing_report.html', rows=rows, totals=totals,
                           from_date=from_date, to_date=to_date, supplier=supplier, suppliers=suppliers)


@app.route('/supplier_purchasing_report/export_xlsx')
@login_required
@has_permission('Access_Reports')
def supplier_purchasing_report_export_xlsx():
    from_date = (request.args.get('from') or date.today().replace(day=1).strftime('%Y-%m-%d')).strip()
    to_date = (request.args.get('to') or date.today().strftime('%Y-%m-%d')).strip()
    supplier = (request.args.get('supplier') or '').strip()

    try:
        import excel_export as xl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash("Excel export needs the 'openpyxl' package on the server — run: pip install openpyxl", 'warning')
        return redirect(url_for('supplier_purchasing_report', **{'from': from_date, 'to': to_date, 'supplier': supplier}))

    rows, totals = _supplier_purchasing_rows(from_date, to_date, supplier)

    wb, ws = xl.new_workbook('Supplier Purchasing Report')
    ncols = 7
    subtitle = f'{from_date} to {to_date}' + (f' — {supplier}' if supplier else '')
    row = xl.title_block(ws, ncols, _company_display_name(), 'SUPPLIER PURCHASING REPORT', subtitle)
    row = xl.header_row(ws, row, ['Invoice Date', 'Supplier', 'Invoice No', 'VAT Rate %', 'Gross', 'VAT', 'Net'])
    for r in rows:
        row = xl.data_row(ws, row,
                           [str(r['inv_date']), r['supplier'] or '', r['invoice_no'] or '',
                            r['vat_rate'], r['gross'], r['vat'], r['net']],
                           num_cols=(4, 5, 6, 7))
    if rows:
        # total_row() puts the label in column 1 and amounts starting at column
        # 2 — fine for a table where amounts ARE columns 2+, but here Gross/
        # VAT/Net are columns 5-7 (after Supplier/Invoice No/VAT Rate). Merge
        # the label across 1-4 and place the three totals in 5-7 by hand so
        # they land under the right headers instead of under Supplier etc.
        label_cell = ws.cell(row=row, column=1, value='TOTAL')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        label_cell.font = Font(bold=True, size=10, color=xl.BLUE_DARK)
        label_cell.fill = PatternFill('solid', fgColor='E3F2FD')
        label_cell.alignment = Alignment(horizontal='right')
        for i in range(2, 5):
            ws.cell(row=row, column=i).fill = PatternFill('solid', fgColor='E3F2FD')
        for i, v in enumerate([totals['gross'], totals['vat'], totals['net']], start=5):
            m = ws.cell(row=row, column=i, value=float(v or 0))
            m.number_format = xl.NUM_FMT
            m.font = Font(bold=True, size=10, color=xl.BLUE_DARK)
            m.fill = PatternFill('solid', fgColor='E3F2FD')
            m.alignment = Alignment(horizontal='right')
        row += 1
    xl.finish(ws, ncols, first_col_width=14, num_col_width=16)
    ws.column_dimensions['B'].width = 28
    return xl.workbook_response(wb, f'Supplier_Purchasing_Report_{from_date}_{to_date}.xlsx')


def _grn_payment_ready_rows(supplier, method, ready):
    """Shared query behind the Payment Ready List page and its Excel export,
    so both apply identical filters/exclusions and therefore identical
    totals. Covers EVERY outstanding supplier invoice regardless of how it
    was created (GRN, Direct Purchase, Service Entry, Opening Balance, ...)."""
    q = """
        SELECT s.s_i_id AS id, s.suppliers_invoice_date AS inv_date, sup.supplier_name AS supplier,
               s.suppliers_invoice_number AS invoice_no, s.suppliers_invoice_oustanding AS amount,
               s.suppliers_invoice_payment_method AS method, s.suppliers_invoice_payment_ready AS ready,
               s.suppliers_invoice_JV AS jv, jv.jv_user_code AS jv_user_code
        FROM suppliers_invoice_data s
        LEFT JOIN jv_numbers jv ON s.suppliers_invoice_JV = jv.jv_id
        JOIN suppliers sup ON s.suppliers_invoice_buinding_supplier = sup.sup_id AND sup.Is_Suplier = 1
        WHERE COALESCE(s.suppliers_oustanding_delete, 0) = 0
          AND s.suppliers_invoice_oustanding > 0.01
    """
    params = []
    if supplier:
        q += " AND sup.supplier_name = %s"; params.append(supplier)
    if method:
        q += " AND s.suppliers_invoice_payment_method = %s"; params.append(method)
    if ready == 'yes':
        q += " AND s.suppliers_invoice_payment_ready = 1"
    elif ready == 'no':
        q += " AND (s.suppliers_invoice_payment_ready = 0 OR s.suppliers_invoice_payment_ready IS NULL)"
    q += " ORDER BY s.suppliers_invoice_date ASC, s.suppliers_invoice_JV ASC"

    rows = db.execute_query(q, tuple(params)) or []

    # Drop invoices fully covered by a not-yet-cleared postdated cheque — the
    # same exclusion Supplier Aging applies, so the two reports' "outstanding"
    # totals reconcile. Those are tracked from the Postdated Cheques page
    # instead, and reappear here automatically once that cheque is Cancelled.
    try:
        pdc_rows = db.execute_query(
            "SELECT payload FROM postdated_cheques WHERE pdc_type = 'payment' AND status = 'pending'") or []
    except Exception:
        pdc_rows = []
    covered_ids = set()
    for pdc in pdc_rows:
        try:
            payload = json.loads(pdc['payload'] or '{}')
        except Exception:
            continue
        for p in payload.get('payments', []):
            covered_ids.add(str(p.get('id')))
    if covered_ids:
        rows = [r for r in rows if str(r['id']) not in covered_ids]

    for r in rows:
        r['amount'] = float(r['amount'] or 0)
        r['method'] = r['method'] or ''
        r['ready'] = bool(r['ready'])
        r['is_grn'] = r.get('jv_user_code') == 'JV FROM GRN'

    total = sum(r['amount'] for r in rows)
    ready_count = sum(1 for r in rows if r['ready'])
    return rows, total, ready_count


@app.route('/grn_payment_ready_list')
@login_required
@has_permission('Access_Reports')
def grn_payment_ready_list():
    """Checklist of outstanding supplier invoices (Cash/Cheque). Covers
    EVERY outstanding supplier invoice regardless of how it was created
    (GRN, Direct Purchase, Service Entry, Opening Balance, ...) — the cheque
    prep this tracks isn't specific to GRN purchases. The cheques
    themselves are drawn by a separate division working off their own
    document, not connected to this system -- when they send over their
    list of what's been drawn, the user checks off the matching invoices
    here as Ready. Purely a manual marker; it never touches the ledger,
    inventory, or the invoice's outstanding balance."""
    supplier = (request.args.get('supplier') or '').strip()
    method = (request.args.get('method') or '').strip()      # '', 'Cash', 'Cheque'
    ready = (request.args.get('ready') or '').strip()        # '', 'yes', 'no'

    rows, total, ready_count = _grn_payment_ready_rows(supplier, method, ready)

    suppliers = db.execute_query("SELECT supplier_name FROM suppliers WHERE Is_Suplier = 1 ORDER BY supplier_name") or []
    return render_template('grn_payment_ready_list.html', rows=rows, total=total,
                           ready_count=ready_count, supplier=supplier, method=method, ready=ready,
                           suppliers=suppliers)


@app.route('/grn_payment_ready_list/export_xlsx')
@login_required
@has_permission('Access_Reports')
def grn_payment_ready_list_export_xlsx():
    """Styled Excel export of the Payment Ready List, honoring whatever
    Supplier / Method / Ready Status filters are currently applied."""
    supplier = (request.args.get('supplier') or '').strip()
    method = (request.args.get('method') or '').strip()
    ready = (request.args.get('ready') or '').strip()

    try:
        import excel_export as xl
    except ImportError:
        flash("Excel export needs the 'openpyxl' package on the server — run: pip install openpyxl", 'warning')
        return redirect(url_for('grn_payment_ready_list', supplier=supplier, method=method, ready=ready))

    rows, total, ready_count = _grn_payment_ready_rows(supplier, method, ready)

    subtitle_bits = []
    if supplier: subtitle_bits.append(f'Supplier: {supplier}')
    if method: subtitle_bits.append(f'Method: {method}')
    if ready == 'yes': subtitle_bits.append('Ready only')
    elif ready == 'no': subtitle_bits.append('Not Ready only')
    subtitle = ' · '.join(subtitle_bits) if subtitle_bits else 'All outstanding supplier invoices'

    wb, ws = xl.new_workbook('Payment Ready List')
    ncols = 6
    row = xl.title_block(ws, ncols, _company_display_name(), 'PAYMENT READY LIST', subtitle)
    row = xl.header_row(ws, row, ['Invoice Date', 'Supplier', 'Invoice No', 'Payment Method', 'Outstanding', 'Ready'])
    for r in rows:
        row = xl.data_row(ws, row,
                          [str(r['inv_date']), r['supplier'] or '-', r['invoice_no'] or '-',
                           r['method'] or 'Not Set', r['amount'], 'Yes' if r['ready'] else 'No'],
                          num_cols=(5,))
    row = xl.total_row(ws, row, f'TOTAL — {len(rows)} invoice(s), {ready_count} marked ready',
                       [0, 0, 0, total, 0], bg='EAF6EA', color=xl.GREEN_DARK, size=11)
    # blank out the zero filler cells added just to line up the total under
    # the Outstanding column (col 5); the label itself covers cols 2-4 & 6.
    for col in (2, 3, 4, 6):
        ws.cell(row=row - 1, column=col).value = None

    xl.finish(ws, ncols, first_col_width=14, num_col_width=16)
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['F'].width = 10

    fname_bits = ['PaymentReadyList']
    if supplier:
        fname_bits.append(supplier.replace(' ', '_').replace('/', '-')[:20])
    fname = '_'.join(fname_bits) + f'_{date.today().strftime("%Y%m%d")}.xlsx'
    return xl.workbook_response(wb, fname)


@app.route('/grn_payment_ready_list/toggle', methods=['POST'])
@login_required
@has_permission('Access_Reports')
def grn_payment_ready_list_toggle():
    """AJAX endpoint: set the Ready flag on one outstanding GRN invoice."""
    data = request.get_json(silent=True) or request.form
    inv_id = data.get('id')
    ready_val = data.get('ready')
    if not inv_id:
        return jsonify({'success': False, 'error': 'Missing invoice id'}), 400
    try:
        ready_flag = 1 if str(ready_val) in ('1', 'true', 'True') else 0
        db.execute_query(
            "UPDATE suppliers_invoice_data SET suppliers_invoice_payment_ready = %s WHERE s_i_id = %s",
            (ready_flag, inv_id), commit=True)
        return jsonify({'success': True, 'ready': bool(ready_flag)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/grn_payment_ready_list/set_method', methods=['POST'])
@login_required
@has_permission('Access_Reports')
def grn_payment_ready_list_set_method():
    """AJAX endpoint: set the Payment Method on one outstanding invoice
    directly from the Ready List. Needed because invoices created outside
    the GRN screen (Direct Purchase, Service Entry, Opening Balance, ...)
    never had a chance to set a payment method at entry time."""
    data = request.get_json(silent=True) or request.form
    inv_id = data.get('id')
    method_val = (data.get('method') or '').strip()
    if not inv_id:
        return jsonify({'success': False, 'error': 'Missing invoice id'}), 400
    if method_val not in ('', 'Cash', 'Cheque'):
        return jsonify({'success': False, 'error': 'Invalid method'}), 400
    try:
        db.execute_query(
            "UPDATE suppliers_invoice_data SET suppliers_invoice_payment_method = %s WHERE s_i_id = %s",
            (method_val or None, inv_id), commit=True)
        return jsonify({'success': True, 'method': method_val})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/get_supplier_data')
@login_required
def get_supplier_data():
    supplier_name = request.args.get('name')
    if not supplier_name:
        return {'error': 'No supplier name'}, 400

    return _get_supplier_history_data(supplier_name)

def _post_bank_payment_entries(cursor, current_user, supplier_name, bank_account,
                                payment_date, narration, cheque_no, payments, wht_base=0.0,
                                manual_voucher=None):
    """Executes the actual GL/bank posting for a bank payment: settle
    invoices, allocate voucher numbers, create the JV + GL entries, and
    record the bank book rows. Shared by the immediate (same-day) flow and
    by clearing a postdated cheque — the caller owns the DB transaction
    (start_transaction / commit / rollback / close) and current_user is the
    numeric user id (get_current_user_id()), used for entry_create_user."""
    total_payment_base = round(sum(p['base'] for p in payments), 2)

    # 1. Update Invoices (Vendor Settle Logic)
    if payments:
        inv_ids = tuple(p['id'] for p in payments)
        format_strings = ','.join(['%s'] * len(inv_ids))
        cursor.execute(f"SELECT s_i_id, suppliers_invoice_oustanding, suppliers_invoice_total_payment FROM suppliers_invoice_data WHERE s_i_id IN ({format_strings})", inv_ids)
        res = cursor.fetchall()

        # Using list for invoice_data to allow updates to current_paid
        invoice_data = {str(r[0]): [float(r[1] or 0), float(r[2] or 0)] for r in res}

        update_args = []
        for p in payments:
            inv_d = invoice_data.get(str(p['id']))
            if not inv_d: continue

            current_outstanding = inv_d[0]
            current_paid = inv_d[1]

            if p['base'] > current_outstanding + 0.01:
                raise Exception(f"Payment amount {p['base']} (base) exceeds outstanding {current_outstanding} for invoice ID {p['id']}")

            new_total_paid = current_paid + p['base']

            # Update the cached data so subsequent duplicate payments use the new total
            inv_d[0] = current_outstanding - p['base'] # Reduce outstanding conceptually, although not checked directly here it is safer
            inv_d[1] = new_total_paid

            update_args.append((new_total_paid, p['id']))

        if update_args:
            # If there are duplicates, executemany might execute them sequentially or fail depending on the db driver and isolation.
            # It's better to aggregate updates per invoice ID.
            # Since update_args contains all sequential updates, the last one per ID is the one that matters.
            # Let's aggregate to ensure executemany works correctly if the driver batches them.
            final_updates = {}
            for paid, inv_id in update_args:
                final_updates[inv_id] = paid

            final_update_args = [(paid, inv_id) for inv_id, paid in final_updates.items()]
            cursor.executemany("UPDATE suppliers_invoice_data SET suppliers_invoice_total_payment = %s WHERE s_i_id = %s", final_update_args)

    # Check Workflow
    cursor.execute("SELECT setting_value FROM system_settings WHERE setting_key = 'enable_approval_workflow'")
    res_set = cursor.fetchone()
    workflow_enabled = res_set and res_set[0] == '1'
    status = 0 if workflow_enabled else 1

    # 2. Voucher Number — manual override if provided, else auto
    if manual_voucher:
        new_voucher = manual_voucher
    else:
        cursor.execute("SELECT MAX(bank_book_voucher_no) FROM bank_book_voucher_no WHERE bank_book_voucher_link = %s", (bank_account,))
        res = cursor.fetchone()
        max_voucher = res[0] if res and res[0] else 0
        new_voucher = max_voucher + 1

    cursor.execute("INSERT INTO bank_book_voucher_no (bank_book_voucher_link, bank_book_voucher_no, bank_book_chq_no) VALUES (%s, %s, %s)",
                   (bank_account, new_voucher, cheque_no))

    # 2a. Generate Master Voucher Number (Global Sequence)
    cursor.execute("INSERT INTO master_payment_voucher_no (voucher_no, create_date) VALUES (0, %s)", (date.today(),))

    try:
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS recent_activity (
            id INT AUTO_INCREMENT PRIMARY KEY,
            dot_color VARCHAR(20) DEFAULT 'blue',
            text_content TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
        ''')
    except mysql.connector.Error as e:
        if e.errno not in (1050, 1007, 1060, 1061, 1146, 1054):
            logging.error(f"Schema Migration Error: {e}")
    except Exception as e:
        logging.error(f"Schema Migration Error: {e}")

    master_voucher_no = cursor.lastrowid

    # 3. Create Journal Voucher (JV)
    cursor.execute("INSERT INTO jv_numbers (jv_user_code, jv_naration, status) VALUES (%s, %s, %s)",
                   ('JV FROM PAYMENT', narration, status))
    jv_no = cursor.lastrowid

    # Get Sub Account Code for Supplier
    cursor.execute("SELECT sub_account_code FROM sub_accont_for_new_account WHERE sub_sub_accaount_name = %s", (supplier_name,))
    res = cursor.fetchone()
    sub_ac_code = res[0] if res else 0

    # Debit AP (Full amount settled) — in base currency
    cursor.execute("""
        INSERT INTO entry_details (
            account_name, enty_values_DR, entry_effective_date, entry_create_date,
            entry_naration, entry_create_user, entry_jv, entry_sub_account_code
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
    """, ('Account Payable', total_payment_base, payment_date, date.today(), narration, current_user, jv_no, sub_ac_code))

    # Credit Bank (Net amount = Total - WHT) — in base currency
    net_payment = total_payment_base - wht_base
    cursor.execute("""
        INSERT INTO entry_details (
            account_name, enty_values_CR, entry_effective_date, entry_create_date,
            entry_naration, entry_create_user, entry_jv
        ) VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (bank_account, net_payment, payment_date, date.today(), narration, current_user, jv_no))

    # Credit WHT Payable (If any) — in base currency
    if wht_base > 0:
        cursor.execute("""
            INSERT INTO entry_details (
                account_name, enty_values_CR, entry_effective_date, entry_create_date,
                entry_naration, entry_create_user, entry_jv
            ) VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, ('WHT Payable', wht_base, payment_date, date.today(), f"WHT on {narration}", current_user, jv_no))

    # 4. Record Bank Transactions (Split proportionately if needed, or record full/net?)
    # Bank Book typically matches bank statement, so record NET payment.

    sup_id_res = db.execute_query("SELECT sup_id FROM suppliers WHERE supplier_name = %s", (supplier_name,))
    sup_id = sup_id_res[0]['sup_id'] if sup_id_res else 0

    for p in payments:
        net_item_amount = p['base']
        if total_payment_base > 0:
            net_item_amount = p['base'] * (net_payment / total_payment_base)

        cursor.execute("""
            INSERT INTO bank_book_recod (
                bank_book__accont_name, bank_book__recode_cr, bank_book__naration,
                bank_book__suplier_oustanding_id, bank_book__suplier_name, jv_numbers_jv_id,
                bank_book_recod_voucher_no, bank_book_chque_no, Bank_Sup_Code, Bank_User_Id, Bank_Payment_Date,
                master_voucher_no
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (bank_account, net_item_amount, narration, p['id'], supplier_name, jv_no, new_voucher, cheque_no, sup_id, current_user, payment_date, master_voucher_no))

    return jv_no, new_voucher, master_voucher_no, total_payment_base


@app.route('/bank_payment_submit', methods=['POST'])
@login_required
def bank_payment_submit():
    supplier_name = request.form.get('supplier')
    bank_account = request.form.get('bank_account')
    payment_date = request.form.get('payment_date')
    narration = request.form.get('narration')
    cheque_no = request.form.get('cheque_no')
    post_date = (request.form.get('post_date') or '').strip()
    wht_amount = float(request.form.get('wht_amount', 0))

    # Multi-currency: amounts are entered in this currency; rate converts to base (LKR)
    payment_currency = (request.form.get('payment_currency') or '').strip()
    exchange_rate = parse_float(request.form.get('exchange_rate', 1)) or 1.0
    if exchange_rate <= 0:
        exchange_rate = 1.0

    if not supplier_name or not bank_account:
        flash('Missing supplier or bank account', 'danger')
        return redirect(url_for('bank_payment'))

    # Collect payments. 'amount' is in the payment currency; 'base' is the LKR-equivalent.
    payments = []
    total_payment = 0
    inv_ids = request.form.getlist('inv_id[]')

    for inv_id in inv_ids:
        amount_str = request.form.get(f'payment_{inv_id}')
        if amount_str and float(amount_str) > 0:
            amount = float(amount_str)
            payments.append({'id': inv_id, 'amount': amount, 'base': round(amount * exchange_rate, 2)})
            total_payment += amount

    if not payments:
        flash('No payment amounts entered', 'warning')
        return redirect(url_for('bank_payment'))

    total_payment_base = round(total_payment * exchange_rate, 2)
    wht_base = round(wht_amount * exchange_rate, 2)
    if payment_currency and exchange_rate != 1.0:
        narration = (narration or '') + f" [{payment_currency} {total_payment:,.2f} @ {exchange_rate:.4f}]"

    manual_voucher = (request.form.get('manual_voucher') or '').strip()
    current_user = get_current_user_id()

    # Postdated cheque: a cheque dated for a future date is held in the PDC
    # register instead of posting to the books now — it only posts (invoice
    # settlement + JV + bank book) when cleared from the Postdated Cheques
    # page, on or after that post date.
    is_postdated = bool(cheque_no) and bool(post_date) and post_date > date.today().isoformat()

    if is_postdated:
        try:
            payload = json.dumps({
                'supplier_name': supplier_name, 'bank_account': bank_account,
                'payment_date': payment_date, 'narration': narration, 'cheque_no': cheque_no,
                'payments': payments, 'wht_base': wht_base, 'manual_voucher': manual_voucher,
            })
            db.execute_query("""
                INSERT INTO postdated_cheques (
                    pdc_type, party_name, account_name, cheque_no, post_date, issue_date,
                    amount, narration, payload, status, created_by
                ) VALUES ('payment', %s, %s, %s, %s, %s, %s, %s, %s, 'pending', %s)
            """, (supplier_name, bank_account, cheque_no, post_date, date.today(),
                  total_payment_base, narration, payload, get_current_user_pk()), commit=True)
            flash(f'Postdated cheque #{cheque_no} recorded for {supplier_name} — it will post to the books on '
                  f'{post_date}. Manage it from Postdated Cheques.', 'success')
        except Exception as e:
            flash(f'Error recording postdated cheque: {str(e)}', 'danger')
        return redirect(url_for('bank_payment'))

    conn = None
    cursor = None
    _last_jv = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        conn.start_transaction()

        jv_no, new_voucher, master_voucher_no, _ = _post_bank_payment_entries(
            cursor, current_user, supplier_name, bank_account, payment_date,
            narration, cheque_no, payments, wht_base, manual_voucher)

        conn.commit()
        flash(f'Payment processed successfully. Voucher No: {new_voucher}, Master Voucher: {master_voucher_no}', 'success')
        log_recent_activity('red', f'<strong>Bank Payment #{new_voucher}</strong> for <strong>{supplier_name}</strong> — LKR {total_payment:,.2f}')
        _last_jv = jv_no  # capture before finally block

    except Exception as e:
        if conn:
            conn.rollback()
        flash(f'Transaction failed: {str(e)}', 'danger')
        logging.error(f"Cash Payment Error: {e}")
        _last_jv = None
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

    if _last_jv:
        return redirect(url_for('bank_payment', last_jv=_last_jv))
    return redirect(url_for('bank_payment'))

# --- Postdated Cheques (PDC) ---
@app.route('/postdated_cheques')
@login_required
@has_permission('Access_Reversals')
def postdated_cheques():
    """List postdated cheques (Bank Payment + Customer Receipt) — pending,
    cleared or cancelled — with optional status/type filters."""
    status_filter = (request.args.get('status') or 'pending').strip()
    type_filter = (request.args.get('type') or '').strip()

    filters = []
    params = []
    if status_filter and status_filter != 'all':
        filters.append("status = %s")
        params.append(status_filter)
    if type_filter in ('payment', 'receipt'):
        filters.append("pdc_type = %s")
        params.append(type_filter)

    where_clause = (" WHERE " + " AND ".join(filters)) if filters else ""
    query = f"""
        SELECT id, pdc_type, party_name, account_name, cheque_no, post_date, issue_date,
               amount, narration, status, cleared_jv, cleared_date
        FROM postdated_cheques
        {where_clause}
        ORDER BY post_date ASC, id DESC
        LIMIT 200
    """
    rows = db.execute_query(query, tuple(params)) if params else db.execute_query(query)
    return render_template('postdated_cheques.html', rows=rows, status_filter=status_filter,
                           type_filter=type_filter, today_str=date.today().strftime('%Y-%m-%d'))


@app.route('/postdated_cheques/clear', methods=['POST'])
@login_required
@has_permission('Access_Reversals')
def postdated_cheques_clear():
    """Post a pending postdated cheque to the books now — runs the same
    invoice-settlement + JV + bank/cash book logic as an immediate payment
    or receipt, effective as of the cheque's post date."""
    pdc_id = request.form.get('id')
    if not pdc_id:
        flash('No postdated cheque selected', 'danger')
        return redirect(url_for('postdated_cheques'))

    conn = None
    cursor = None
    cursor2 = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor(dictionary=True)
        conn.start_transaction()

        cursor.execute("SELECT * FROM postdated_cheques WHERE id = %s AND status = 'pending' FOR UPDATE", (pdc_id,))
        pdc = cursor.fetchone()
        if not pdc:
            conn.rollback()
            flash('Postdated cheque not found or already processed.', 'danger')
            return redirect(url_for('postdated_cheques'))

        payload = json.loads(pdc['payload'] or '{}')
        effective_date = str(pdc['post_date'])
        current_user = get_current_user_id()
        cursor2 = conn.cursor()

        if pdc['pdc_type'] == 'payment':
            jv_no, new_voucher, master_voucher_no, _ = _post_bank_payment_entries(
                cursor2, current_user, payload['supplier_name'], payload['bank_account'],
                effective_date, payload['narration'], payload.get('cheque_no'),
                payload['payments'], payload.get('wht_base', 0.0), payload.get('manual_voucher'))
        else:
            jv_no, receipt_no, _ = _post_customer_receipt_entries(
                cursor2, current_user, payload['customer_id'], payload['account_type'],
                payload['account_name'], effective_date, payload['narration'], payload['payments'],
                payload.get('manual_receipt_no'), payload.get('online_payment_received', False),
                payload.get('transaction_code'), payload.get('card_last_digits'),
                payload.get('bank_transfer_confirmed', False), payload.get('transfer_id'),
                payload.get('cheque_no'))

        cursor2.execute(
            "UPDATE postdated_cheques SET status = 'cleared', cleared_jv = %s, cleared_date = %s WHERE id = %s",
            (jv_no, date.today(), pdc_id))

        conn.commit()
        flash(f"Postdated cheque #{pdc['cheque_no']} cleared and posted to the books (JV: {jv_no}).", 'success')

    except Exception as e:
        if conn:
            conn.rollback()
        flash(f'Error clearing postdated cheque: {str(e)}', 'danger')
        logging.error(f"PDC Clear Error: {e}")
    finally:
        if cursor2:
            cursor2.close()
        if cursor:
            cursor.close()
        if conn:
            conn.close()

    return redirect(url_for('postdated_cheques'))


@app.route('/postdated_cheques/reschedule', methods=['POST'])
@login_required
@has_permission('Access_Reversals')
def postdated_cheques_reschedule():
    """Change the post date on a pending postdated cheque — e.g. it bounced
    and a new date was agreed, or the party asked to push it out."""
    pdc_id = request.form.get('id')
    new_date = (request.form.get('new_post_date') or '').strip()
    if not pdc_id or not new_date:
        flash('A new post date is required', 'danger')
        return redirect(url_for('postdated_cheques'))

    try:
        res = db.execute_query("SELECT status FROM postdated_cheques WHERE id = %s", (pdc_id,))
        if not res:
            flash('Postdated cheque not found.', 'danger')
        elif res[0]['status'] != 'pending':
            flash('Only pending postdated cheques can be rescheduled.', 'warning')
        else:
            db.execute_query("UPDATE postdated_cheques SET post_date = %s WHERE id = %s",
                             (new_date, pdc_id), commit=True)
            flash(f'Post date updated to {new_date}.', 'success')
    except Exception as e:
        flash(f'Error rescheduling: {str(e)}', 'danger')

    return redirect(url_for('postdated_cheques'))


@app.route('/postdated_cheques/cancel', methods=['POST'])
@login_required
@has_permission('Access_Reversals')
def postdated_cheques_cancel():
    """Cancel a pending postdated cheque (e.g. returned/voided) — it was
    never posted to the books, so there is nothing to reverse."""
    pdc_id = request.form.get('id')
    if not pdc_id:
        flash('No postdated cheque selected', 'danger')
        return redirect(url_for('postdated_cheques'))

    try:
        res = db.execute_query("SELECT status FROM postdated_cheques WHERE id = %s", (pdc_id,))
        if not res:
            flash('Postdated cheque not found.', 'danger')
        elif res[0]['status'] != 'pending':
            flash('Only pending postdated cheques can be cancelled.', 'warning')
        else:
            db.execute_query("UPDATE postdated_cheques SET status = 'cancelled' WHERE id = %s",
                             (pdc_id,), commit=True)
            flash('Postdated cheque cancelled.', 'success')
    except Exception as e:
        flash(f'Error cancelling: {str(e)}', 'danger')

    return redirect(url_for('postdated_cheques'))


# --- Customer Loyalty ---
@app.route('/customer_loyalty', methods=['GET', 'POST'])
@login_required
def customer_loyalty():
    if request.method == 'POST':
        name = request.form.get('customer_name')
        billing = request.form.get('billing_address')
        delivery = request.form.get('delivery_address')
        email = request.form.get('email')
        mobile = request.form.get('mobile_no')
        paid = request.form.get('amount_paid')

        if not mobile:
            flash('Please enter the mobile number', 'danger')
            return redirect(url_for('customer_loyalty'))

        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            conn.start_transaction()

            # Determine next ID
            cursor.execute("SELECT MAX(id) FROM customer")
            res = cursor.fetchone()
            max_id = res[0] if res and res[0] else 0
            customer_code = max_id + 60001

            current_date = datetime.utcnow()

            query = """
                INSERT INTO customer (
                    id, customer_name, customer_code,
                    customer_Billing_Address, costomer_Delivery_Address,
                    e_mail, coustomer_credit_limit, Mobile_nimber,
                    Is_Loyality_Customer, Compay_Or_Not, Create_Date,
                    Paid_Amountl, Create_Cashiyer
                ) VALUES (0, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            params = (
                name, customer_code, billing, delivery, email,
                1, mobile, 1, 0, current_date, paid if paid else 0, 0
            )

            cursor.execute(query, params)
            conn.commit()
            flash('Updated ..', 'success')

        except Exception as e:
            conn.rollback()
            flash(f'Error: {str(e)}', 'danger')
        finally:
            cursor.close()
            conn.close()

        return redirect(url_for('customer_loyalty'))

    return render_template('customer_loyalty.html')

# --- Direct Purchasing ---
@app.route('/direct_purchasing', methods=['GET'])
@login_required
@has_permission('Access_Accounting')
def direct_purchasing():
    cash_accounts = db.execute_query("SELECT cash_book_account_name FROM cash_book")
    cost_accounts = db.execute_query("SELECT account_name FROM new_account_table WHERE account_active = 1 ORDER BY account_name")
    items = db.execute_query("SELECT inventoy_name FROM inventoy_items")
    try:
        jobs = db.execute_query("SELECT job_number, job_description FROM jobs_unit ORDER BY job_number")
    except Exception:
        jobs = []

    last_jv = request.args.get('last_jv')

    # ── History filters ──
    hf = {
        'search':     request.args.get('h_search', '').strip(),
        'date_from':  request.args.get('h_date_from', '').strip(),
        'date_to':    request.args.get('h_date_to', '').strip(),
        'amount_min': request.args.get('h_amount_min', '').strip(),
        'amount_max': request.args.get('h_amount_max', '').strip(),
        'status':     request.args.get('h_status', '').strip(),  # '', 'active', 'reversed'
    }

    # Default to the CURRENT MONTH on a fresh load so we don't scan/return everything.
    # Any submitted h_* param means the user is filtering — respect exactly what they set
    # (clearing both dates then filtering shows all records).
    hist_filter_submitted = any(
        k in request.args for k in
        ('h_search', 'h_date_from', 'h_date_to', 'h_amount_min', 'h_amount_max', 'h_status')
    )
    if not hist_filter_submitted:
        hf['date_from'] = date.today().replace(day=1).strftime('%Y-%m-%d')
        hf['date_to'] = date.today().strftime('%Y-%m-%d')

    h_filters = ["j.jv_user_code = 'JV FROM DIRECT CASH'", "c.cash_book_recode_cr > 0"]
    h_params = []
    if hf['search']:
        h_filters.append("(c.cash_book_recode_accont_name LIKE %s OR c.cash_book_recode_naration LIKE %s)")
        h_params += [f"%{hf['search']}%", f"%{hf['search']}%"]
    if hf['date_from']:
        h_filters.append("c.Payment_Date >= %s"); h_params.append(hf['date_from'])
    if hf['date_to']:
        h_filters.append("c.Payment_Date <= %s"); h_params.append(hf['date_to'])
    if hf['status'] == 'active':
        h_filters.append("c.User_Revers IS NULL")
    elif hf['status'] == 'reversed':
        h_filters.append("c.User_Revers IS NOT NULL")

    # amount filters applied via HAVING on the grouped sum
    having = []
    if hf['amount_min']:
        having.append("SUM(c.cash_book_recode_cr) >= %s");
    if hf['amount_max']:
        having.append("SUM(c.cash_book_recode_cr) <= %s")
    having_clause = (" HAVING " + " AND ".join(having)) if having else ""
    having_params = []
    if hf['amount_min']: having_params.append(hf['amount_min'])
    if hf['amount_max']: having_params.append(hf['amount_max'])

    where_clause = " AND ".join(h_filters)
    history = db.execute_query(f"""
        SELECT
            j.jv_id                          AS jv,
            MIN(c.cash_book_recod_voucher_no) AS voucher,
            MIN(c.Payment_Date)               AS date,
            MIN(c.cash_book_recode_accont_name) AS account,
            SUM(c.cash_book_recode_cr)        AS amount,
            CASE WHEN MAX(c.User_Revers) IS NOT NULL THEN 1 ELSE 0 END AS is_reversed
        FROM jv_numbers j
        JOIN cash_book_recode c ON c.jv_numbers_jv_id = j.jv_id
        WHERE {where_clause}
        GROUP BY j.jv_id
        {having_clause}
        ORDER BY j.jv_id DESC
        LIMIT 200
    """, tuple(h_params + having_params)) or []

    base_currency, currencies = _base_currency_and_list()
    return render_template('direct_purchasing.html',
                           cash_accounts=cash_accounts,
                           cost_accounts=cost_accounts,
                           inventory_items=items,
                           jobs=jobs,
                           today_date=datetime.now().strftime('%Y-%m-%d'),
                           session_payment_items=session.get('payment_items', []),
                           total_value=session.get('payment_total', 0),
                           last_jv=last_jv,
                           history=history,
                           hf=hf,
                           base_currency=base_currency,
                           currencies=currencies)


@app.route('/api/cost_accounts')
@login_required
def api_cost_accounts():
    """Live account list so a newly created account shows up in an open window
    (e.g. direct purchasing) without a full page refresh — all active accounts."""
    accounts = db.execute_query(
        "SELECT account_name FROM new_account_table WHERE account_active = 1 ORDER BY account_name") or []
    return jsonify([a['account_name'] for a in accounts])


@app.route('/direct_purchasing/add_item', methods=['POST'])
@login_required
def direct_purchasing_add_item():
    qty_str = request.form.get('qty')
    price_str = request.form.get('price')

    try:
        qty = float(qty_str) if qty_str and qty_str.strip() else 0.0
    except ValueError:
        qty = 0.0

    try:
        price = float(price_str) if price_str and price_str.strip() else 0.0
    except ValueError:
        price = 0.0

    raw_job = request.form.get('job_no', '').strip()
    item = {
        'account': request.form.get('cost_account'),
        'item_name': request.form.get('inventory_item'),
        'job_no': raw_job if raw_job else None,   # store None, never empty string
        'qty': qty,
        'price': price,
        'narration': request.form.get('narration', '')
    }
    item['total'] = round(item['qty'] * item['price'], 2)

    if 'payment_items' not in session:
        session['payment_items'] = []

    session['payment_items'].append(item)
    session.modified = True

    total = sum(i['total'] for i in session['payment_items'])
    session['payment_total'] = round(total, 2)

    return redirect(url_for('direct_purchasing'))

@app.route('/direct_purchasing/remove_item', methods=['POST'])
@login_required
def direct_purchasing_remove_item():
    try:
        idx = int(request.form.get('index', -1))
    except (ValueError, TypeError):
        idx = -1

    items = session.get('payment_items', [])
    if 0 <= idx < len(items):
        items.pop(idx)
        session['payment_items'] = items
        session['payment_total'] = round(sum(i['total'] for i in items), 2)
        session.modified = True

    return redirect(url_for('direct_purchasing'))

@app.route('/direct_purchasing/clear', methods=['POST'])
@login_required
def direct_purchasing_clear():
    session.pop('payment_items', None)
    session.pop('payment_total', None)
    session.modified = True
    return redirect(url_for('direct_purchasing'))

# ================================================================
# ── FINANCIAL PERIOD CLOSING ────────────────────────────────────
# ================================================================
# When a period is closed (locked up to period_end_date), any transaction
# whose effective date falls on or before that date can no longer be
# edited or reversed. Protects finalized financial periods.

_PERIOD_TABLE_READY = False

def _ensure_period_table():
    global _PERIOD_TABLE_READY
    if _PERIOD_TABLE_READY:
        return
    try:
        db.execute_query("""
            CREATE TABLE IF NOT EXISTS financial_period_close (
                id            INT AUTO_INCREMENT PRIMARY KEY,
                period_end_date DATE NOT NULL,
                closed_by     VARCHAR(100),
                closed_at     DATETIME DEFAULT CURRENT_TIMESTAMP,
                reopened      TINYINT DEFAULT 0,
                reopened_by   VARCHAR(100),
                reopened_at   DATETIME NULL,
                notes         VARCHAR(255)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """, commit=True)
        _PERIOD_TABLE_READY = True
    except Exception as e:
        logging.warning(f"_ensure_period_table: {e}")

def get_period_lock_date():
    """Return the latest closed period_end_date (date) still in force, or None."""
    _ensure_period_table()
    try:
        rows = db.execute_query(
            "SELECT MAX(period_end_date) AS d FROM financial_period_close WHERE reopened = 0")
        if rows and rows[0].get('d'):
            d = rows[0]['d']
            return d if isinstance(d, date) else date.fromisoformat(str(d)[:10])
    except Exception as e:
        logging.warning(f"get_period_lock_date: {e}")
    return None

def jv_in_closed_period(jv):
    """True if the JV's effective date is on/before the current lock date."""
    lock = get_period_lock_date()
    if not lock:
        return False
    try:
        rows = db.execute_query(
            "SELECT MIN(entry_effective_date) AS d FROM entry_details WHERE entry_jv = %s",
            (jv,))
        if rows and rows[0].get('d'):
            d = rows[0]['d']
            d = d if isinstance(d, date) else date.fromisoformat(str(d)[:10])
            return d <= lock
    except Exception as e:
        logging.warning(f"jv_in_closed_period: {e}")
    return False


@app.route('/financial_period_close', methods=['GET'])
@login_required
@has_permission('Access_Accounting')
def financial_period_close():
    _ensure_period_table()
    history = db.execute_query("""
        SELECT id, period_end_date, closed_by, closed_at, reopened, reopened_by, reopened_at, notes
        FROM financial_period_close
        ORDER BY period_end_date DESC, id DESC
        LIMIT 50
    """) or []
    lock = get_period_lock_date()
    return render_template('financial_period_close.html',
                           history=history,
                           lock_date=lock.strftime('%Y-%m-%d') if lock else None,
                           today_date=date.today().strftime('%Y-%m-%d'))


@app.route('/financial_period_close/close', methods=['POST'])
@login_required
@has_permission('Access_Accounting')
def financial_period_close_action():
    _ensure_period_table()
    end_date = (request.form.get('period_end_date') or '').strip()
    notes = (request.form.get('notes') or '').strip()
    if not end_date:
        flash('Please choose a period end date.', 'danger')
        return redirect(url_for('financial_period_close'))
    try:
        date.fromisoformat(end_date)
    except ValueError:
        flash('Invalid date.', 'danger')
        return redirect(url_for('financial_period_close'))

    try:
        db.execute_query(
            "INSERT INTO financial_period_close (period_end_date, closed_by, notes) VALUES (%s, %s, %s)",
            (end_date, get_current_user_id(), notes), commit=True)
        flash(f'Period closed through {end_date}. Transactions on or before this date are now locked.', 'success')
    except Exception as e:
        flash(f'Error closing period: {str(e)}', 'danger')
    return redirect(url_for('financial_period_close'))


@app.route('/financial_period_close/reopen', methods=['POST'])
@login_required
@has_permission('Access_Accounting')
def financial_period_reopen():
    _ensure_period_table()
    close_id = request.form.get('id')
    if not close_id:
        flash('No record specified.', 'danger')
        return redirect(url_for('financial_period_close'))
    try:
        db.execute_query(
            "UPDATE financial_period_close SET reopened=1, reopened_by=%s, reopened_at=NOW() WHERE id=%s",
            (get_current_user_id(), close_id), commit=True)
        flash('Period re-opened. Affected transactions can be edited again.', 'success')
    except Exception as e:
        flash(f'Error re-opening period: {str(e)}', 'danger')
    return redirect(url_for('financial_period_close'))


@app.route('/transaction/edit_metadata', methods=['POST'])
@login_required
@has_permission('Access_Accounting')
def transaction_edit_metadata():
    """Generic metadata-only editor for any posted transaction, keyed by JV.
    Updates date / voucher / narration across whichever tables hold rows for
    that JV. Amounts and GL accounts are never touched (ledger integrity)."""
    jv = (request.form.get('jv') or '').strip()
    new_date = (request.form.get('txn_date') or '').strip()
    new_voucher = (request.form.get('voucher_no') or '').strip()
    new_narration = (request.form.get('narration') or '').strip()

    if not jv:
        return jsonify({'success': False, 'error': 'No JV provided'}), 400

    eff_date = None
    if new_date:
        try:
            eff_date = date.fromisoformat(new_date)
        except ValueError:
            return jsonify({'success': False, 'error': 'Invalid date'}), 400

    # Period-close guard: block edits to transactions in a closed period,
    # and block moving a transaction INTO a closed period.
    lock = get_period_lock_date()
    if lock:
        if jv_in_closed_period(jv):
            return jsonify({'success': False,
                            'error': f'This transaction is in a closed period (locked through {lock}). Re-open the period to edit it.'}), 403
        if eff_date and eff_date <= lock:
            return jsonify({'success': False,
                            'error': f'The new date falls in a closed period (locked through {lock}).'}), 403

    if not (eff_date or new_voucher or new_narration):
        return jsonify({'success': False, 'error': 'Nothing to update'}), 400

    # Each statement is independent; a JV may not have rows in every table.
    # Missing table/column is caught and skipped so one type's schema never
    # blocks another. 0 rows affected is normal, not an error.
    stmts = []
    if eff_date:
        stmts += [
            ("UPDATE entry_details SET entry_effective_date=%s WHERE entry_jv=%s", (eff_date, jv)),
            ("UPDATE cash_book_recode SET Payment_Date=%s WHERE jv_numbers_jv_id=%s", (eff_date, jv)),
            ("UPDATE bank_book_recod SET Bank_Payment_Date=%s WHERE jv_numbers_jv_id=%s", (eff_date, jv)),
            ("UPDATE inventory_recod SET inventory_recod_action_date=%s WHERE JV_No=%s", (eff_date, jv)),
            ("UPDATE Invoice_Oustanding SET invoice_date=%s WHERE invoice_JV=%s", (eff_date, jv)),
            ("UPDATE suppliers_invoice_data SET suppliers_invoice_date=%s WHERE suppliers_invoice_JV=%s", (eff_date, jv)),
        ]
    if new_narration:
        stmts += [
            ("UPDATE entry_details SET entry_naration=%s WHERE entry_jv=%s", (new_narration, jv)),
            ("UPDATE jv_numbers SET jv_naration=%s WHERE jv_id=%s", (new_narration, jv)),
            ("UPDATE cash_book_recode SET cash_book_recode_naration=%s WHERE jv_numbers_jv_id=%s", (new_narration, jv)),
            ("UPDATE bank_book_recod SET bank_book__naration=%s WHERE jv_numbers_jv_id=%s", (new_narration, jv)),
        ]
    if new_voucher:
        stmts += [
            ("UPDATE cash_book_recode SET cash_book_recod_voucher_no=%s WHERE jv_numbers_jv_id=%s", (new_voucher, jv)),
            ("UPDATE bank_book_recod SET bank_book_recod_voucher_no=%s WHERE jv_numbers_jv_id=%s", (new_voucher, jv)),
        ]

    applied = 0
    errors = []
    for sql, params in stmts:
        try:
            db.execute_query(sql, params, commit=True)
            applied += 1
        except Exception as e:
            # Skip missing table/column for this transaction type
            errors.append(str(e).split('\n')[0])

    if applied == 0:
        return jsonify({'success': False, 'error': 'No matching records updated. ' + (errors[0] if errors else '')}), 500
    return jsonify({'success': True})


@app.route('/direct_purchasing/next_voucher')
@login_required
def direct_purchasing_next_voucher():
    """Return the next auto voucher number for a given cash account."""
    cash_account = request.args.get('cash_account', '')
    if not cash_account:
        return jsonify({'next': ''})
    res = db.execute_query(
        "SELECT MAX(cash_voucher_number) AS m FROM cash_voucher_no WHERE cash_voucher_link = %s",
        (cash_account,))
    mx = 0
    if res and res[0].get('m') is not None:
        try:
            mx = int(res[0]['m'])
        except (ValueError, TypeError):
            mx = 0
    return jsonify({'next': mx + 1})


@app.route('/direct_purchasing/edit', methods=['POST'])
@login_required
@has_permission('Access_Accounting')
def direct_purchasing_edit():
    """Edit metadata only (date, voucher no, narration) of a posted direct purchase.
    Amounts and GL accounts are NOT editable to protect ledger integrity."""
    jv = request.form.get('jv', '').strip()
    new_date = request.form.get('txn_date', '').strip()
    new_voucher = request.form.get('voucher_no', '').strip()
    new_narration = request.form.get('narration', '').strip()

    if not jv:
        return jsonify({'success': False, 'error': 'No JV provided'}), 400

    # Validate date
    eff_date = None
    if new_date:
        try:
            eff_date = date.fromisoformat(new_date)
        except ValueError:
            return jsonify({'success': False, 'error': 'Invalid date'}), 400

    # Period-close guard
    lock = get_period_lock_date()
    if lock:
        if jv_in_closed_period(jv):
            return jsonify({'success': False,
                            'error': f'This transaction is in a closed period (locked through {lock}). Re-open to edit.'}), 403
        if eff_date and eff_date <= lock:
            return jsonify({'success': False,
                            'error': f'The new date falls in a closed period (locked through {lock}).'}), 403

    conn = None
    cursor = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        conn.start_transaction()

        # Update effective date across GL, cash book, inventory (NOT create date)
        if eff_date:
            cursor.execute(
                "UPDATE entry_details SET entry_effective_date=%s WHERE entry_jv=%s",
                (eff_date, jv))
            cursor.execute(
                "UPDATE cash_book_recode SET Payment_Date=%s WHERE jv_numbers_jv_id=%s",
                (eff_date, jv))
            cursor.execute(
                "UPDATE inventory_recod SET inventory_recod_action_date=%s WHERE JV_No=%s",
                (eff_date, jv))

        # Update narration
        if new_narration:
            cursor.execute(
                "UPDATE entry_details SET entry_naration=%s WHERE entry_jv=%s",
                (new_narration, jv))
            cursor.execute(
                "UPDATE cash_book_recode SET cash_book_recode_naration=%s WHERE jv_numbers_jv_id=%s",
                (new_narration, jv))
            cursor.execute(
                "UPDATE jv_numbers SET jv_naration=%s WHERE jv_id=%s",
                (new_narration, jv))

        # Update voucher number on cash book records
        if new_voucher:
            cursor.execute(
                "UPDATE cash_book_recode SET cash_book_recod_voucher_no=%s WHERE jv_numbers_jv_id=%s",
                (new_voucher, jv))

        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        if conn: conn.rollback()
        logging.error(f"Direct Purchase edit error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if cursor: cursor.close()
        if conn: conn.close()


@app.route('/direct_purchasing/submit', methods=['POST'])
@login_required
def direct_purchasing_submit():
    if 'payment_items' not in session or not session['payment_items']:
        flash('No items to submit', 'warning')
        return redirect(url_for('direct_purchasing'))

    cash_account = request.form.get('cash_account')
    if not cash_account:
        flash('Please select a cash account', 'danger')
        return redirect(url_for('direct_purchasing'))

    items = session['payment_items']
    total_amount = session.get('payment_total', 0)
    current_user = get_current_user_id()
    today_date = date.today()                          # entry_create_date (when entered)
    # Use user's transaction date as entry_effective_date (when transaction occurred)
    raw_txn_date = request.form.get('date', '').strip()
    try:
        txn_date = date.fromisoformat(raw_txn_date) if raw_txn_date else today_date
    except ValueError:
        txn_date = today_date
    narration = (request.form.get('narration') or '').strip() or "Direct Purchase / Cash Payment"

    # Multi-currency: amounts are entered in this currency; rate converts to base (LKR)
    dp_currency = (request.form.get('payment_currency') or '').strip()
    exchange_rate = parse_float(request.form.get('exchange_rate', 1)) or 1.0
    if exchange_rate <= 0:
        exchange_rate = 1.0
    total_amount_base = round(parse_float(total_amount) * exchange_rate, 2)
    if dp_currency and exchange_rate != 1.0:
        narration = narration + f" [{dp_currency} {parse_float(total_amount):,.2f} @ {exchange_rate:.4f}]"

    # cash_voucher_number is a NUMERIC column — only accept a numeric manual
    # override. A non-numeric value would cause "1265 Data truncated" and abort.
    manual_voucher = (request.form.get('manual_voucher') or '').strip()
    manual_voucher_num = None
    if manual_voucher:
        try:
            manual_voucher_num = int(float(manual_voucher))
        except (ValueError, TypeError):
            manual_voucher_num = None  # ignore non-numeric; auto-generate instead

    conn = None
    cursor = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        conn.start_transaction()

        # 1. Voucher Number — use numeric manual override if valid, else auto-generate
        if manual_voucher_num is not None:
            new_voucher = manual_voucher_num
        else:
            cursor.execute("SELECT MAX(cash_voucher_number) FROM cash_voucher_no WHERE cash_voucher_link = %s", (cash_account,))
            res = cursor.fetchone()
            max_voucher = res[0] if res and res[0] else 0
            new_voucher = int(max_voucher) + 1

        cursor.execute("INSERT INTO cash_voucher_no (cash_voucher_link, cash_voucher_number) VALUES (%s, %s)", (cash_account, new_voucher))

        # 2. Create Journal Voucher (JV)
        cursor.execute("INSERT INTO jv_numbers (jv_user_code, jv_naration) VALUES (%s, %s)", ('JV FROM DIRECT CASH', narration))
        jv_no = cursor.lastrowid

        # 3. GL Entries (posted in base currency) — txn_date effective, today_date create
        # Credit Cash Account (Total)
        cursor.execute("""
            INSERT INTO entry_details (
                account_name, enty_values_CR, entry_effective_date, entry_create_date,
                entry_naration, entry_create_user, entry_jv
            ) VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (cash_account, total_amount_base, txn_date, today_date, narration, current_user, jv_no))

        # Debit Expense/Asset Accounts (Per Item)
        for item in items:
            # Ensure job_no is None (NULL) if empty — prevents FK constraint violation
            raw_job = item.get('job_no')
            job_no = raw_job if raw_job and str(raw_job).strip() else None
            item_total_base = round(parse_float(item['total']) * exchange_rate, 2)

            # Debit Entry
            cursor.execute("""
                INSERT INTO entry_details (
                    account_name, enty_values_DR, entry_effective_date, entry_create_date,
                    entry_naration, entry_create_user, entry_jv, entry_job_number
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (item['account'], item_total_base, txn_date, today_date,
                  item.get('narration') or 'Direct Purchase', current_user, jv_no, job_no))

            # 4. Cash Book Record
            cursor.execute("""
                INSERT INTO cash_book_recode (
                    cash_book_recode_dr, cash_book_recode_cr, cash_book_recode_accont_name,
                    cash_book_recode_naration, jv_numbers_jv_id, cash_book_recod_voucher_no,
                    User_Enter, Payment_Date
                ) VALUES (0, %s, %s, %s, %s, %s, %s, %s)
            """, (item_total_base, cash_account, item['narration'], jv_no, new_voucher, current_user, txn_date))

            # 5. Inventory Record (if Item Name is present)
            if item.get('item_name'):
                cursor.execute("SELECT inventoy_code, inventoy_items_messurment_unit FROM inventoy_items WHERE inventoy_name = %s", (item['item_name'],))
                inv_res = cursor.fetchone()

                if inv_res:
                    inv_code = inv_res[0]
                    inv_unit = inv_res[1]

                    cursor.execute("""
                        INSERT INTO inventory_recod (
                            inventoy_name, inventoy_code, inventory_recod_action_date,
                            inventory_recod_moument_in, inventory_recod_movment_out,
                            inventory_recod_mesrmet, inventory_recod_unit_price,
                            inventory_recod_account, inventory_recod_user_id,
                            inventory_recod_user_recod_date, JV_No
                        ) VALUES (%s, %s, %s, %s, 0, %s, %s, %s, %s, %s, %s)
                    """, (item['item_name'], inv_code, txn_date, item['qty'], inv_unit, round(parse_float(item['price']) * exchange_rate, 2), item['account'], current_user, today_date, jv_no))

        conn.commit()

        session.pop('payment_items', None)
        session.pop('payment_total', None)
        flash(f'Payment submitted successfully. JV No: {jv_no}, Voucher: {new_voucher}', 'success')
        _dp_jv = jv_no

    except Exception as e:
        if conn:
            try: conn.rollback()
            except Exception: pass
        flash(f'Error submitting payment: {str(e)}', 'danger')
        logging.error(f"Direct Payment Error: {e}")
        _dp_jv = None
    finally:
        if cursor:
            try: cursor.close()
            except Exception: pass
        if conn:
            try: conn.close()
            except Exception: pass

    if _dp_jv:
        return redirect(url_for('direct_purchasing', last_jv=_dp_jv))
    return redirect(url_for('direct_purchasing'))

# ================================================================
# ── DOCUMENT / ATTACHMENT SYSTEM (DB-stored, no filesystem) ────
# ================================================================
# All files are stored as base64 in `transaction_documents` table.
# Linked by ref_type (e.g. 'direct_payment') + ref_id (JV number)
# and optionally ref_line (line index within that transaction).

_DOCS_TABLE_CREATED = False

def _ensure_docs_table():
    """Create transaction_documents table if it doesn't exist yet."""
    global _DOCS_TABLE_CREATED
    if _DOCS_TABLE_CREATED:
        return
    try:
        db.execute_query("""
            CREATE TABLE IF NOT EXISTS transaction_documents (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                ref_type    VARCHAR(60)  NOT NULL,
                ref_id      VARCHAR(60)  NOT NULL,
                ref_line    INT          DEFAULT NULL,
                file_name   VARCHAR(255) NOT NULL,
                file_type   VARCHAR(100) DEFAULT NULL,
                file_data   LONGBLOB     NOT NULL,
                file_size   INT          DEFAULT 0,
                uploaded_by VARCHAR(100) DEFAULT NULL,
                uploaded_at DATETIME     DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_ref (ref_type, ref_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """, commit=True)
        _DOCS_TABLE_CREATED = True
    except Exception as e:
        logging.warning(f"_ensure_docs_table: {e}")

ALLOWED_DOC_EXTENSIONS = {
    'pdf', 'jpg', 'jpeg', 'png', 'gif', 'bmp', 'webp',
    'doc', 'docx', 'xls', 'xlsx', 'csv', 'txt'
}
MAX_DOC_SIZE_MB = 10

def _allowed_doc(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_DOC_EXTENSIONS


@app.route('/documents/upload', methods=['POST'])
@login_required
def document_upload():
    """Upload a document and store it in the DB. Returns JSON."""
    _ensure_docs_table()

    ref_type = request.form.get('ref_type', '').strip()
    ref_id   = request.form.get('ref_id',   '').strip()
    ref_line = request.form.get('ref_line')      # optional line index
    if not ref_type or not ref_id:
        return jsonify({'success': False, 'error': 'ref_type and ref_id are required'}), 400

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file provided'}), 400

    f = request.files['file']
    if not f or f.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    if not _allowed_doc(f.filename):
        return jsonify({'success': False,
                        'error': f'File type not allowed. Allowed: {", ".join(sorted(ALLOWED_DOC_EXTENSIONS))}'}), 400

    data = f.read()
    size_mb = len(data) / (1024 * 1024)
    if size_mb > MAX_DOC_SIZE_MB:
        return jsonify({'success': False,
                        'error': f'File too large ({size_mb:.1f} MB). Max {MAX_DOC_SIZE_MB} MB.'}), 400

    ref_line_val = int(ref_line) if ref_line and str(ref_line).isdigit() else None
    username = get_current_user_id()

    try:
        db.execute_query("""
            INSERT INTO transaction_documents
                (ref_type, ref_id, ref_line, file_name, file_type, file_data, file_size, uploaded_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (ref_type, ref_id, ref_line_val,
              f.filename, f.content_type, data, len(data), username),
             commit=True)

        # Get the new ID
        rows = db.execute_query(
            "SELECT id FROM transaction_documents WHERE ref_type=%s AND ref_id=%s "
            "ORDER BY id DESC LIMIT 1",
            (ref_type, ref_id))
        new_id = rows[0]['id'] if rows else None

        return jsonify({'success': True, 'id': new_id,
                        'file_name': f.filename, 'size': len(data)})
    except Exception as e:
        logging.error(f"Document upload error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/documents/list')
@login_required
def document_list():
    """Return JSON list of documents for a ref_type + ref_id (+ optional ref_line)."""
    _ensure_docs_table()
    ref_type = request.args.get('ref_type', '')
    ref_id   = request.args.get('ref_id',   '')
    ref_line = request.args.get('ref_line')

    if not ref_type or not ref_id:
        return jsonify([])

    try:
        if ref_line is not None:
            rows = db.execute_query(
                "SELECT id, ref_line, file_name, file_type, file_size, uploaded_by, uploaded_at "
                "FROM transaction_documents WHERE ref_type=%s AND ref_id=%s AND ref_line=%s "
                "ORDER BY id",
                (ref_type, ref_id, ref_line))
        else:
            rows = db.execute_query(
                "SELECT id, ref_line, file_name, file_type, file_size, uploaded_by, uploaded_at "
                "FROM transaction_documents WHERE ref_type=%s AND ref_id=%s "
                "ORDER BY ref_line, id",
                (ref_type, ref_id))

        result = []
        for r in rows:
            result.append({
                'id':          r['id'],
                'ref_line':    r['ref_line'],
                'file_name':   r['file_name'],
                'file_type':   r['file_type'] or '',
                'file_size':   r['file_size'] or 0,
                'uploaded_by': r['uploaded_by'] or '',
                'uploaded_at': str(r['uploaded_at'])[:16] if r['uploaded_at'] else ''
            })
        return jsonify(result)
    except Exception as e:
        logging.error(f"Document list error: {e}")
        return jsonify([])


@app.route('/documents/view/<int:doc_id>')
@login_required
def document_view(doc_id):
    """Stream a stored document (inline or download)."""
    _ensure_docs_table()
    try:
        rows = db.execute_query(
            "SELECT file_name, file_type, file_data FROM transaction_documents WHERE id=%s",
            (doc_id,))
        if not rows:
            return "Document not found", 404
        doc = rows[0]
        data  = doc['file_data']
        ctype = doc['file_type'] or 'application/octet-stream'
        fname = doc['file_name']

        from flask import Response as FlaskResponse
        disposition = 'inline' if ctype.startswith('image/') or ctype == 'application/pdf' else 'attachment'
        resp = FlaskResponse(data, mimetype=ctype)
        resp.headers['Content-Disposition'] = f'{disposition}; filename="{fname}"'
        resp.headers['Content-Length'] = len(data)
        return resp
    except Exception as e:
        return str(e), 500


@app.route('/documents/delete/<int:doc_id>', methods=['POST'])
@login_required
def document_delete(doc_id):
    """Delete a stored document."""
    _ensure_docs_table()
    try:
        db.execute_query(
            "DELETE FROM transaction_documents WHERE id=%s",
            (doc_id,), commit=True)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/documents/line_items')
@login_required
def document_line_items():
    """Return ordered line items for a direct_payment JV so the UI
    can render per-line upload zones indexed 0, 1, 2 …"""
    jv       = request.args.get('jv', '')
    ref_type = request.args.get('ref_type', 'direct_payment')
    if not jv:
        return jsonify([])

    try:
        # cash_book_recode rows for this JV represent each line item
        rows = db.execute_query("""
            SELECT cash_book_recode_accont_name  AS account,
                   cash_book_recode_naration     AS narration,
                   cash_book_recode_cr           AS amount,
                   (SELECT inventoy_name
                    FROM inventory_recod
                    WHERE JV_No = c.jv_numbers_jv_id
                    LIMIT 1)                     AS item_name
            FROM cash_book_recode c
            WHERE c.jv_numbers_jv_id = %s
            ORDER BY c.chash_book_recod_id
        """, (jv,))
        return jsonify([{
            'account':   r.get('account') or '',
            'narration': r.get('narration') or '',
            'amount':    float(r.get('amount') or 0),
            'item_name': r.get('item_name') or ''
        } for r in rows])
    except Exception as e:
        logging.error(f"document_line_items: {e}")
        return jsonify([])


# ================================================================
# ── AI DOCUMENT SCANNER (pytesseract + PyPDF2 + regex) ─────────
# ================================================================

def _configure_tesseract(pytesseract):
    """Point pytesseract at the tesseract binary and return its path (or None).

    On shared hosting (cPanel) you usually cannot `sudo apt-get install`, but you
    can install/build tesseract in your home dir. Allow an explicit override via
    the TESSERACT_CMD env var and probe common locations."""
    import shutil
    candidates = []
    env_cmd = os.getenv('TESSERACT_CMD')
    if env_cmd:
        candidates.append(env_cmd)
    which = shutil.which('tesseract')
    if which:
        candidates.append(which)
    candidates += [
        '/usr/bin/tesseract',
        '/usr/local/bin/tesseract',
        '/bin/tesseract',
        os.path.expanduser('~/bin/tesseract'),
        os.path.expanduser('~/.local/bin/tesseract'),
    ]
    for c in candidates:
        if c and os.path.exists(c):
            pytesseract.pytesseract.tesseract_cmd = c
            return c
    return None


_TESSERACT_HELP = (
    'No OCR is available on this server. Easiest fix (no server install): set the '
    'OCR_SPACE_API_KEY environment variable to a free key from https://ocr.space/ocrapi '
    'and the app will read images via cloud OCR. '
    'Alternatively install Tesseract in your account and set TESSERACT_CMD to its full path '
    '(e.g. TESSERACT_CMD=/home/USER/bin/tesseract), or on a server you control run '
    'sudo apt-get install tesseract-ocr. '
    'Tip: text-based PDFs are read without any OCR — upload a PDF instead of a photo/scan.'
)


def ocr_image_bytes(raw_bytes):
    """Return OCR text for image bytes.

    Order of preference:
      1. Local Tesseract binary (private, no data leaves the server).
      2. OCR.space cloud API — only if OCR_SPACE_API_KEY is set (opt-in; the image
         is sent to that third-party service).
    Raises RuntimeError('OCR_UNAVAILABLE') if neither is usable."""
    # 1. Local Tesseract
    try:
        import pytesseract
        from PIL import Image
        if _configure_tesseract(pytesseract):
            img = Image.open(io.BytesIO(raw_bytes))
            return pytesseract.image_to_string(img)
    except Exception as e:
        logging.warning(f"Local Tesseract OCR unavailable, will try cloud fallback: {e}")

    # 2. Cloud fallback (opt-in)
    api_key = os.getenv('OCR_SPACE_API_KEY')
    if api_key:
        try:
            resp = requests.post(
                'https://api.ocr.space/parse/image',
                files={'file': ('upload.png', raw_bytes)},
                data={'apikey': api_key, 'OCREngine': '2', 'scale': 'true'},
                timeout=60,
            )
            j = resp.json()
            if not j.get('IsErroredOnProcessing'):
                results = j.get('ParsedResults') or []
                return "\n".join(r.get('ParsedText', '') for r in results)
            logging.error(f"OCR.space error: {j.get('ErrorMessage')}")
        except Exception as e:
            logging.error(f"OCR.space request failed: {e}")

    raise RuntimeError('OCR_UNAVAILABLE')


@app.route('/documents/ai_extract', methods=['POST'])
@login_required
def documents_ai_extract():
    """
    Scan an uploaded file (image or PDF) with OCR / text extraction,
    then use regex heuristics to pull out:
      unit_price, quantity, total_amount, invoice_no, date
    Returns JSON with the extracted fields.
    """
    # If the browser already OCR'd the image (Tesseract.js runs client-side), use that
    # text directly — no server OCR needed and the image never leaves the user's machine.
    browser_text = (request.form.get('ocr_text') or '').strip()

    extracted_text = ''

    if browser_text:
        extracted_text = browser_text
    else:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file'}), 400

        f = request.files['file']
        if not f or f.filename == '':
            return jsonify({'success': False, 'error': 'Empty file'}), 400

        filename_lower = f.filename.lower()
        raw_data = f.read()

        try:
            if filename_lower.endswith('.pdf'):
                # ── PDF: use PyPDF2 text extraction ──
                import PyPDF2
                reader = PyPDF2.PdfReader(io.BytesIO(raw_data))
                for page in reader.pages:
                    t = page.extract_text()
                    if t:
                        extracted_text += t + '\n'

            elif any(filename_lower.endswith(ext) for ext in ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.tiff', '.tif')):
                # ── Image: server-side OCR fallback (used only if browser OCR wasn't sent) ──
                try:
                    extracted_text = ocr_image_bytes(raw_data)
                except RuntimeError:
                    return jsonify({'success': False, 'error': _TESSERACT_HELP}), 500
                except ImportError:
                    return jsonify({'success': False,
                                    'error': "The 'pytesseract'/'Pillow' Python packages are not installed. "
                                             "Run: pip install pytesseract Pillow. " + _TESSERACT_HELP}), 500
            else:
                # Try plain text
                try:
                    extracted_text = raw_data.decode('utf-8', errors='replace')
                except Exception:
                    extracted_text = ''

        except Exception as e:
            return jsonify({'success': False, 'error': f'Extraction error: {str(e)}'}), 500

    if not extracted_text.strip():
        return jsonify({'success': False,
                        'error': 'Could not extract text from this document. '
                                 'Try a clearer scan or a text-based PDF.'})

    # ── Regex extraction ──
    import re

    def first_match(patterns, text):
        for pat in patterns:
            m = re.search(pat, text, re.IGNORECASE)
            if m:
                # return first captured group that is not None
                for g in m.groups():
                    if g is not None:
                        return g.strip().replace(',', '')
        return None

    def clean_num(s):
        if s is None:
            return None
        s = re.sub(r'[^\d.]', '', s)
        try:
            return float(s)
        except Exception:
            return None

    # Unit price — look for "Unit Price", "Rate", "Price", "Unit Rate"
    unit_price_str = first_match([
        r'unit\s*(?:price|rate|cost)[:\s]*([0-9,]+\.?\d*)',
        r'rate[:\s]+([0-9,]+\.?\d*)',
        r'(?:price|cost)\s*per\s*(?:unit|item|piece)[:\s]*([0-9,]+\.?\d*)',
        r'up[:\s]+([0-9,]+\.?\d{2})',
    ], extracted_text)

    # Quantity — look for "Qty", "Quantity", "Units", "No. of", "Nos"
    qty_str = first_match([
        r'(?:qty|quantity|units?|nos?\.?|pieces?|pcs)[:\s.]*([0-9,]+\.?\d*)',
        r'(?:no\.|number)\s+of\s+(?:units?|items?|pieces?)[:\s]*([0-9,]+\.?\d*)',
    ], extracted_text)

    # Total amount — look for "Total", "Amount", "Grand Total", "Net Amount"
    total_str = first_match([
        r'(?:grand\s*)?total\s*(?:amount)?[:\s]*(?:lkr|rs\.?|usd|eur|gbp)?\s*([0-9,]+\.?\d{2})',
        r'(?:net\s*)?amount\s*(?:due|payable)?[:\s]*(?:lkr|rs\.?|usd|eur|gbp)?\s*([0-9,]+\.?\d{2})',
        r'(?:invoice\s*)?total[:\s]*([0-9,]+\.?\d{2})',
    ], extracted_text)

    # Invoice number
    inv_no_str = first_match([
        r'inv(?:oice)?\.?\s*(?:no\.?|number|#)[:\s]*([A-Z0-9\-/]+)',
        r'bill\s*(?:no\.?|number|#)[:\s]*([A-Z0-9\-/]+)',
        r'(?:ref|reference)\s*(?:no\.?|#)?[:\s]*([A-Z0-9\-/]{4,20})',
    ], extracted_text)

    # Date
    date_str = first_match([
        r'(?:invoice\s*|bill\s*|payment\s*)?date[:\s]*(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4})',
        r'(\d{4}[\/\-]\d{2}[\/\-]\d{2})',
        r'(\d{1,2}\s+(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\.?\s*\d{2,4})',
    ], extracted_text)

    unit_price = clean_num(unit_price_str)
    quantity   = clean_num(qty_str)
    total      = clean_num(total_str)

    # Fallback: if unit_price and quantity both found but no total, compute it
    if unit_price and quantity and not total:
        total = round(unit_price * quantity, 2)

    # Fallback: if total and quantity found but no unit_price, compute it
    if total and quantity and quantity > 0 and not unit_price:
        unit_price = round(total / quantity, 4)

    result = {
        'success':      True,
        'unit_price':   unit_price,
        'quantity':     quantity,
        'total_amount': total,
        'invoice_no':   inv_no_str,
        'date':         date_str,
        'raw_text':     extracted_text[:800],   # first 800 chars for debug preview
    }
    return jsonify(result)

# --- Inventory Price Editing ---
@app.route('/inventory_price_editing', methods=['GET'])
@login_required
@has_permission('Access_Inventory')
def inventory_price_editing():
    search = request.args.get('search', '')
    # Join only the LATEST price record per item so each item shows exactly once.
    # A plain LEFT JOIN duplicates items that have multiple price tiers, which made
    # category edits look like they "didn't save" (the duplicate row still showed None).
    query = """
        SELECT
            ii.id, ii.inventoy_name, ii.inventoy_code,
            ii.Main_Catogry, ii.Sub_Catogory,
            ipr.inventory_price_selling, ipr.inventory_price_profit_marging_comen,
            ipr.inventory_price_purcharsing, ipr.inventory_price_for_Loyality_customer
        FROM inventoy_items ii
        LEFT JOIN inventory_price_recod ipr
            ON ipr.id = (
                SELECT MAX(p2.id) FROM inventory_price_recod p2
                WHERE p2.inventory_price_link = ii.id
            )
    """
    params = None
    if search:
        query += " WHERE ii.inventoy_name LIKE %s OR ii.inventoy_code LIKE %s"
        search_pattern = f"%{search}%"
        params = (search_pattern, search_pattern)
    query += " ORDER BY ii.inventoy_name"

    items = db.execute_query(query, params)

    # Get all active items for the "Add New Price Tier" dropdown
    all_items = db.execute_query("SELECT id, inventoy_name, inventoy_code FROM inventoy_items WHERE active = 1 ORDER BY inventoy_name")

    # Category dropdowns — use the EXACT stored names (no TRIM): inventoy_items.Main_Catogry
    # is a foreign key into inventory_carogory.main_catogory, so option values must match
    # the parent keys character-for-character or the save fails with FK error 1452.
    main_categories = db.execute_query(
        "SELECT main_catogory FROM inventory_carogory WHERE main_catogory IS NOT NULL AND main_catogory != '' AND (dis_continue_main IS NULL OR dis_continue_main = 0) ORDER BY main_catogory"
    ) or []
    sub_categories = db.execute_query(
        "SELECT sub_catogory FROM inventory_carogory WHERE sub_catogory IS NOT NULL AND sub_catogory != '' AND (dis_continue_sub IS NULL OR dis_continue_sub = 0) ORDER BY sub_catogory"
    ) or []

    # Recent change history
    try:
        change_history = db.execute_query(
            "SELECT * FROM inventory_item_change_history ORDER BY changed_at DESC LIMIT 100"
        ) or []
    except Exception:
        change_history = []

    return render_template('inventory_price_editing.html',
                           items=items, all_items=all_items, search_query=search,
                           main_categories=main_categories, sub_categories=sub_categories,
                           change_history=change_history)

@app.route('/inventory_price_editing/update', methods=['POST'])
@login_required
def update_inventory_prices():
    item_ids = request.form.getlist('item_ids[]')
    market_prices = request.form.getlist('market_prices[]')
    spm_prices = request.form.getlist('spm_prices[]')
    loyalty_prices = request.form.getlist('loyalty_prices[]')
    item_names = request.form.getlist('item_names[]')
    main_cats = request.form.getlist('main_cats[]')
    sub_cats = request.form.getlist('sub_cats[]')
    old_names = request.form.getlist('old_names[]')
    old_main_cats = request.form.getlist('old_main_cats[]')
    old_sub_cats = request.form.getlist('old_sub_cats[]')

    if not item_ids:
        flash('No items to update', 'warning')
        return redirect(url_for('inventory_price_editing'))

    user_code = session.get('user_id', 'unknown')
    user_pk = session.get('user_pk', 0)

    # Check existence in bulk
    placeholders = ', '.join(['%s'] * len(item_ids))
    query = f"SELECT inventory_price_link FROM inventory_price_recod WHERE inventory_price_link IN ({placeholders})"
    existing_records = db.execute_query(query, tuple(item_ids))

    existing_links = set()
    if existing_records:
        if isinstance(existing_records[0], dict):
            existing_links = {str(r['inventory_price_link']) for r in existing_records}
        else:
            existing_links = {str(r[0]) for r in existing_records}

    for i in range(len(item_ids)):
        link_id = str(item_ids[i])
        new_name = item_names[i].strip() if i < len(item_names) else ''
        new_main = main_cats[i].strip() if i < len(main_cats) else ''
        new_sub = sub_cats[i].strip() if i < len(sub_cats) else ''
        prev_name = old_names[i] if i < len(old_names) else ''
        prev_main = old_main_cats[i] if i < len(old_main_cats) else ''
        prev_sub = old_sub_cats[i] if i < len(old_sub_cats) else ''

        # Update prices
        if link_id in existing_links:
            db.execute_query("""
                UPDATE inventory_price_recod SET
                inventory_price_selling = %s,
                inventory_price_profit_marging_comen = %s,
                inventory_price_for_Loyality_customer = %s
                WHERE inventory_price_link = %s
            """, (market_prices[i], spm_prices[i], loyalty_prices[i], link_id), commit=True)
        else:
            db.execute_query("""
                INSERT INTO inventory_price_recod
                (inventory_price_link, inventory_price_selling, inventory_price_profit_marging_comen, inventory_price_for_Loyality_customer)
                VALUES (%s, %s, %s, %s)
            """, (link_id, market_prices[i], spm_prices[i], loyalty_prices[i]), commit=True)

        # Detect changes in name / category and log them
        field_changes = []
        if new_name and new_name != prev_name:
            field_changes.append(('item_name', prev_name, new_name))
        if new_main != prev_main:
            field_changes.append(('main_category', prev_main, new_main))
        if new_sub != prev_sub:
            field_changes.append(('sub_category', prev_sub, new_sub))

        if field_changes:
            effective_name = new_name if new_name else prev_name
            db.execute_query("""
                UPDATE inventoy_items SET
                inventoy_name = %s, Main_Catogry = %s, Sub_Catogory = %s
                WHERE id = %s
            """, (effective_name, new_main, new_sub, link_id), commit=True)

            for field, old_val, new_val in field_changes:
                try:
                    db.execute_query("""
                        INSERT INTO inventory_item_change_history
                        (item_id, item_name, field_changed, old_value, new_value,
                         changed_by_user_code, changed_by_user_pk)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """, (link_id, effective_name, field, old_val, new_val, user_code, user_pk),
                    commit=True)
                except Exception:
                    pass

    flash('Changes saved successfully', 'success')
    clear_pos_items_cache()
    return redirect(url_for('inventory_price_editing'))

# --- Balance Sheet ---
@app.route('/balance_sheet', methods=['GET', 'POST'])
@login_required
@has_permission('Access_Reports')
def balance_sheet():
    # Comparative balance sheet: one or more "as at" dates (columns), like P&L periods.
    _ensure_sub_account_codes()
    dates = []
    src = request.form if request.method == 'POST' else request.args
    for d in src.getlist('as_at_date[]'):
        if d:
            dates.append(d)
    if not dates and request.args.get('as_at_date'):
        dates.append(request.args.get('as_at_date'))
    if not dates:
        dates.append(datetime.now().strftime('%Y-%m-%d'))
    dates = dates[:6]
    n = len(dates)
    as_at_date = dates[0]

    def _bs_section(flag_col, balance_expr, d):
        """One balance-sheet section as at date d, ordered by category Holding Level,
        then the per-account order (account_bs_sort), then name."""
        ordered = f"""
            SELECT
                na.account_name_of_catogory_Balace_sheet as category,
                na.account_name as name,
                {balance_expr} as balance
            FROM new_account_table na
            LEFT JOIN entry_details ed ON na.account_name = ed.account_name
                AND ed.entry_effective_date <= %s AND ed.entry_deleted = 0
            WHERE na.{flag_col} = 1
            GROUP BY na.account_name, na.account_name_of_catogory_Balace_sheet,
                     na.account_hold_possion_Balace_Sheet, na.account_bs_sort
            ORDER BY COALESCE(na.account_hold_possion_Balace_Sheet, 9999) + 0,
                     COALESCE(na.account_bs_sort, 9999), na.account_name
        """
        fallback = f"""
            SELECT
                na.account_name_of_catogory_Balace_sheet as category,
                na.account_name as name,
                {balance_expr} as balance
            FROM new_account_table na
            LEFT JOIN entry_details ed ON na.account_name = ed.account_name
                AND ed.entry_effective_date <= %s AND ed.entry_deleted = 0
            WHERE na.{flag_col} = 1
            GROUP BY na.account_name, na.account_name_of_catogory_Balace_sheet,
                     na.account_hold_possion_Balace_Sheet
            ORDER BY COALESCE(na.account_hold_possion_Balace_Sheet, 9999) + 0, na.account_name
        """
        try:
            return db.execute_query(ordered, (d,))
        except Exception:
            return db.execute_query(fallback, (d,))

    ASSET_EXPR = "COALESCE(SUM(ed.enty_values_DR), 0) - COALESCE(SUM(ed.enty_values_CR), 0)"
    CREDIT_EXPR = "COALESCE(SUM(ed.enty_values_CR), 0) - COALESCE(SUM(ed.enty_values_DR), 0)"

    def _retained(d):
        """Retained Earnings (Income - Expenses) as at date d."""
        rows = db.execute_query("""
            SELECT na.account_income, na.account_expenses,
                   COALESCE(SUM(ed.enty_values_DR), 0) as total_dr,
                   COALESCE(SUM(ed.enty_values_CR), 0) as total_cr
            FROM new_account_table na
            LEFT JOIN entry_details ed ON na.account_name = ed.account_name
                AND ed.entry_effective_date <= %s
                AND ed.entry_deleted = 0
            WHERE (na.account_income = 1 OR na.account_expenses = 1)
              AND na.account_active = 1
            GROUP BY na.account_name, na.account_income, na.account_expenses
        """, (d,)) or []
        inc, exp = 0.0, 0.0
        for row in rows:
            dr = float(row['total_dr'] or 0)
            cr = float(row['total_cr'] or 0)
            if row['account_income']:
                inc += cr - dr
            elif row['account_expenses']:
                exp += dr - cr
        return inc - exp

    def _bs_unclassified(d):
        """Accounts with NO classification (not asset/liability/equity/income/
        expense). Their balance would otherwise be dropped from both sides and
        the sheet wouldn't tally, so surface them under 'Unclassified' on the
        assets side with the natural DR-CR balance."""
        q = """
            SELECT 'Unclassified' as category, na.account_name as name,
                   COALESCE(SUM(ed.enty_values_DR), 0) - COALESCE(SUM(ed.enty_values_CR), 0) as balance
            FROM new_account_table na
            LEFT JOIN entry_details ed ON na.account_name = ed.account_name
                AND ed.entry_effective_date <= %s AND ed.entry_deleted = 0
            WHERE na.account_active = 1
              AND COALESCE(na.account_assets, 0) = 0 AND COALESCE(na.account_liabilities, 0) = 0
              AND COALESCE(na.account_equity, 0) = 0 AND COALESCE(na.account_income, 0) = 0
              AND COALESCE(na.account_expenses, 0) = 0
            GROUP BY na.account_name
        """
        try:
            return db.execute_query(q, (d,)) or []
        except Exception:
            return []

    asset_snaps = [_bs_section('account_assets', ASSET_EXPR, d) for d in dates]
    liab_snaps = [_bs_section('account_liabilities', CREDIT_EXPR, d) for d in dates]
    eq_snaps = [_bs_section('account_equity', CREDIT_EXPR, d) for d in dates]
    unclassified_snaps = [_bs_unclassified(d) for d in dates]
    retained_earnings = [_retained(d) for d in dates]

    def _merge(snaps):
        """Merge per-date rows into one row per account with an amounts list.
        Rows that are zero across every date are dropped."""
        seen, order = {}, []
        for i, rows in enumerate(snaps):
            for r in rows or []:
                key = (r['category'] or 'Uncategorized', r['name'])
                if key not in seen:
                    seen[key] = {'category': key[0], 'name': r['name'], 'amounts': [0.0] * n}
                    order.append(key)
                seen[key]['amounts'][i] = float(r['balance'] or 0)
        return [seen[k] for k in order if any(abs(v) >= 0.005 for v in seen[k]['amounts'])]

    merged_assets = _merge(asset_snaps)
    merged_liabs = _merge(liab_snaps)
    cleaned_equity = _merge(eq_snaps)
    # Fold unclassified accounts into the assets side so the sheet balances and
    # the user can see which accounts still need a proper classification.
    merged_assets.extend(_merge(unclassified_snaps))

    # ── Sub-account breakdown per account (drives the expandable + rows) ──
    # Show every DEFINED sub-account (linked by sub_new_account) under its main
    # account, with the value from any sub-coded postings (0 if none) — so the +
    # appears whenever a main account has sub-accounts defined.
    try:
        _bs_defined = {}   # main account -> [{code, name}]
        for r in (db.execute_query("SELECT sub_new_account, sub_account_code, sub_sub_accaount_name FROM sub_accont_for_new_account WHERE active = 1") or []):
            _bs_defined.setdefault((r['sub_new_account'] or '').strip(), []).append(
                {'code': str(r['sub_account_code']), 'name': r['sub_sub_accaount_name']})
    except Exception:
        _bs_defined = {}

    def _bs_sub_breakdown(flag_col, balance_expr):
        """{account_name: [{name, amounts[]}]} — all defined subs, valued per date."""
        vals_by = {}   # (acc, code) -> [amounts]
        for i, d in enumerate(dates):
            q = f"""
                SELECT ed.account_name AS acc, ed.entry_sub_account_code AS code, {balance_expr} AS balance
                FROM entry_details ed
                JOIN new_account_table na ON na.account_name = ed.account_name
                WHERE na.{flag_col} = 1 AND ed.entry_deleted = 0
                  AND ed.entry_effective_date <= %s
                  AND ed.entry_sub_account_code IS NOT NULL AND ed.entry_sub_account_code != 0
                GROUP BY ed.account_name, ed.entry_sub_account_code
            """
            try:
                rows = db.execute_query(q, (d,)) or []
            except Exception:
                rows = []
            for r in rows:
                key = ((r['acc'] or '').strip(), str(r['code']))
                vals_by.setdefault(key, [0.0] * n)[i] = float(r['balance'] or 0)
        out = {}
        for acc, subs_def in _bs_defined.items():
            subs = [{'name': s['name'], 'amounts': vals_by.get((acc, s['code']), [0.0] * n)}
                    for s in subs_def]
            subs.sort(key=lambda s: str(s['name']).lower())
            out[acc] = subs
        return out

    _asset_subs = _bs_sub_breakdown('account_assets', ASSET_EXPR)
    _liab_subs = _bs_sub_breakdown('account_liabilities', CREDIT_EXPR)
    _equity_subs = _bs_sub_breakdown('account_equity', CREDIT_EXPR)
    for r in merged_assets:
        r['sub_accounts'] = _asset_subs.get((r['name'] or '').strip(), [])
    for r in merged_liabs:
        r['sub_accounts'] = _liab_subs.get((r['name'] or '').strip(), [])
    for r in cleaned_equity:
        r['sub_accounts'] = _equity_subs.get((r['name'] or '').strip(), [])

    # "(Unallocated)" sub-row so the sub-accounts reconcile to the account total
    for r in (merged_assets + merged_liabs + cleaned_equity):
        if r.get('sub_accounts'):
            unalloc = [r['amounts'][i] - sum(s['amounts'][i] for s in r['sub_accounts'])
                       for i in range(n)]
            if any(abs(v) >= 0.005 for v in unalloc):
                r['sub_accounts'].append({'name': '(Unallocated)', 'amounts': unalloc})

    def _sum_cols(merged):
        t = [0.0] * n
        for r in merged:
            for i, v in enumerate(r['amounts']):
                t[i] += v
        return t

    total_assets = _sum_cols(merged_assets)
    total_liabilities = _sum_cols(merged_liabs)
    total_equity = _sum_cols(cleaned_equity)

    grouped_assets = {}
    for r in merged_assets:
        grouped_assets.setdefault(r['category'], []).append(r)
    grouped_liabilities = {}
    for r in merged_liabs:
        grouped_liabilities.setdefault(r['category'], []).append(r)

    # Order categories by the AUTHORITATIVE holding level from balance_sheet_category
    # (per-account level copies go stale when a category's level is edited later)
    bs_levels = _category_levels('balance_sheet_category')

    def _order_grouped(g):
        return dict(sorted(g.items(), key=lambda kv: (bs_levels.get(kv[0], 9999), kv[0])))

    grouped_assets = _order_grouped(grouped_assets)
    grouped_liabilities = _order_grouped(grouped_liabilities)

    # Custom subtotal rows: cumulative from the top of the side (assets or
    # liabilities) down to the chosen category
    bs_sub_defs = _report_subtotals('BS')

    def _cumulative_subtotals(grouped):
        subs = {}
        run = [0.0] * n
        for cat_name, accs in grouped.items():
            for a in accs:
                for i, v in enumerate(a['amounts']):
                    run[i] += v
            if cat_name in bs_sub_defs:
                subs[cat_name] = [{'label': lbl, 'amounts': list(run)} for lbl in bs_sub_defs[cat_name]]
        return subs

    bs_subtotals = {
        'assets': _cumulative_subtotals(grouped_assets),
        'liabilities': _cumulative_subtotals(grouped_liabilities)
    }

    report_data = {
        'assets': grouped_assets,
        'liabilities': grouped_liabilities,
        'equity': cleaned_equity
    }

    totals = {
        'assets': total_assets,
        'liabilities': total_liabilities,
        'equity': total_equity,
        'retained_earnings': retained_earnings
    }

    eq_liab_total = [totals['liabilities'][i] + totals['equity'][i] + totals['retained_earnings'][i]
                     for i in range(n)]

    # Styled Excel export (presentation format, one column per as-at date)
    if request.args.get('download') == 'xlsx':
        try:
            import excel_export as xl
        except ImportError:
            flash("Excel export needs the 'openpyxl' package on the server — run: pip install openpyxl", 'warning')
            return redirect(url_for('balance_sheet', as_at_date=as_at_date))

        wb, ws = xl.new_workbook('Balance Sheet')
        ncols = 1 + n
        row = xl.title_block(ws, ncols, _company_display_name(), 'BALANCE SHEET',
                             ' | '.join(f'As at {d}' for d in dates))
        row = xl.header_row(ws, row, ['Description'] + [f'As at {d}' for d in dates])

        row = xl.section_row(ws, row, ncols, 'ASSETS', color=xl.BLUE_DARK, bg='EFF6FF')
        for category, accounts in report_data['assets'].items():
            row = xl.section_row(ws, row, ncols, category, color='495057', bg='F8F9FA')
            for acc in accounts:
                row = xl.item_row(ws, row, acc['name'], acc['amounts'])
                for sub in acc.get('sub_accounts', []):
                    row = xl.item_row(ws, row, '        ↳ ' + sub['name'], sub['amounts'])
            for st in bs_subtotals['assets'].get(category, []):
                row = xl.total_row(ws, row, st['label'], st['amounts'], bg='EEF2FF', color=xl.INDIGO)
        row = xl.total_row(ws, row, 'TOTAL ASSETS', totals['assets'],
                           bg='EFF6FF', color=xl.BLUE_DARK, size=11)
        row += 1

        row = xl.section_row(ws, row, ncols, 'EQUITY AND LIABILITIES', color=xl.GREEN_DARK, bg='F0FFF0')
        row = xl.section_row(ws, row, ncols, 'Equity', color='495057', bg='F8F9FA')
        for acc in report_data['equity']:
            row = xl.item_row(ws, row, acc['name'], acc['amounts'])
            for sub in acc.get('sub_accounts', []):
                row = xl.item_row(ws, row, '        ↳ ' + sub['name'], sub['amounts'])
        row = xl.item_row(ws, row, 'Retained Earnings', totals['retained_earnings'])
        for category, accounts in report_data['liabilities'].items():
            row = xl.section_row(ws, row, ncols, category, color='495057', bg='F8F9FA')
            for acc in accounts:
                row = xl.item_row(ws, row, acc['name'], acc['amounts'])
                for sub in acc.get('sub_accounts', []):
                    row = xl.item_row(ws, row, '        ↳ ' + sub['name'], sub['amounts'])
            for st in bs_subtotals['liabilities'].get(category, []):
                row = xl.total_row(ws, row, st['label'], st['amounts'], bg='EEF2FF', color=xl.INDIGO)
        row = xl.total_row(ws, row, 'TOTAL EQUITY & LIABILITIES', eq_liab_total,
                           bg='F0FFF0', color=xl.GREEN_DARK, size=11)
        xl.finish(ws, ncols)
        return xl.workbook_response(wb, f'BalanceSheet_{dates[0]}.xlsx')

    return render_template('balance_sheet.html', as_at_date=as_at_date, dates=dates,
                           report_data=report_data, totals=totals, bs_subtotals=bs_subtotals,
                           eq_liab_total=eq_liab_total)

# --- Cash Flow ---

@app.route('/cash_flow', methods=['GET'])
@login_required
@has_permission('Access_Reports')
def cash_flow_view():
    from_date = request.args.get('from_date', date.today().replace(day=1).strftime('%Y-%m-%d'))
    to_date = request.args.get('to_date', date.today().strftime('%Y-%m-%d'))
    return render_template('cash_flow.html', from_date=from_date, to_date=to_date)

@app.route('/api/cash_flow/generate', methods=['POST'])
@login_required
@has_permission('Access_Reports')
def cash_flow_generate():
    data = request.json
    return _compute_cash_flow(data.get('from_date'), data.get('to_date'), data.get('rec_only', False))


def _compute_cash_flow(from_date, to_date, rec_only=False):
    """Compute the cash flow statement. Shared by the JSON API and the Excel export."""
    rec_filter = " AND ed.entry_Rec = 1 " if rec_only else ""

    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # 1. Net Profit
        net_profit_query = f'''
            SELECT
                (SELECT COALESCE(SUM(CASE
                    WHEN a.account_basment = 'CR' THEN (ed.enty_values_CR - ed.enty_values_DR)
                    WHEN a.account_basment = 'DR' THEN (ed.enty_values_DR - ed.enty_values_CR)
                    ELSE 0 END), 0)
                FROM entry_details ed
                JOIN new_account_table a ON ed.account_name = a.account_name
                WHERE a.account_income = 1
                AND ed.entry_effective_date BETWEEN %s AND %s
                AND ed.entry_deleted = 0 {rec_filter}) -
                (SELECT COALESCE(SUM(CASE
                    WHEN a.account_basment = 'CR' THEN (ed.enty_values_CR - ed.enty_values_DR)
                    WHEN a.account_basment = 'DR' THEN (ed.enty_values_DR - ed.enty_values_CR)
                    ELSE 0 END), 0)
                FROM entry_details ed
                JOIN new_account_table a ON ed.account_name = a.account_name
                WHERE a.account_expenses = 1
                AND ed.entry_effective_date BETWEEN %s AND %s
                AND ed.entry_deleted = 0 {rec_filter}) as NetProfit
        '''
        cursor.execute(net_profit_query, (from_date, to_date, from_date, to_date))
        net_profit_res = cursor.fetchone()
        net_profit = float(net_profit_res['NetProfit']) if net_profit_res and net_profit_res['NetProfit'] is not None else 0.0

        # 2. Adjustments
        adj_query = f'''
            SELECT a.account_name as Description,
                   cf.hold_level as HoldLevel,
                   SUM(CASE
                        WHEN a.account_basment = 'CR' THEN (ed.enty_values_CR - ed.enty_values_DR)
                        WHEN a.account_basment = 'DR' THEN (ed.enty_values_DR - ed.enty_values_CR)
                        ELSE 0 END) AS Amount
            FROM entry_details ed
            JOIN new_account_table a ON ed.account_name = a.account_name
            JOIN cf_catogory cf ON a.cf_catogory = cf.catogory_name
            WHERE cf.catogory_name = 'Adjustments'
            AND ed.entry_effective_date BETWEEN %s AND %s
            AND ed.entry_deleted = 0 {rec_filter}
            GROUP BY a.account_name, cf.hold_level
            ORDER BY cf.hold_level
        '''
        cursor.execute(adj_query, (from_date, to_date))
        adj_items = [dict(r) for r in cursor.fetchall()]

        # 3. Working Capital
        wc_query = f'''
            SELECT a.account_name,
                   a.account_assets,
                   a.account_liabilities,
                   cf.hold_level as HoldLevel,
                   SUM(CASE
                        WHEN a.account_basment = 'DR' THEN (ed.enty_values_CR - ed.enty_values_DR)
                        WHEN a.account_basment = 'CR' THEN (ed.enty_values_CR - ed.enty_values_DR)
                        ELSE 0 END) AS Amount
            FROM entry_details ed
            JOIN new_account_table a ON ed.account_name = a.account_name
            JOIN cf_catogory cf ON a.cf_catogory = cf.catogory_name
            WHERE cf.catogory_name = 'Changes In Working Capital'
            AND ed.entry_effective_date BETWEEN %s AND %s
            AND ed.entry_deleted = 0 {rec_filter}
            GROUP BY a.account_name, a.account_assets, a.account_liabilities, cf.hold_level
            ORDER BY cf.hold_level
        '''
        cursor.execute(wc_query, (from_date, to_date))
        wc_raw = cursor.fetchall()
        wc_items = []
        for r in wc_raw:
            prefix = ""
            if r['account_assets'] == 1:
                prefix = "(Increase)/Decrease In "
            elif r['account_liabilities'] == 1:
                prefix = "Increase/(Decrease) In "

            wc_items.append({
                'Description': prefix + r['account_name'],
                'Amount': float(r['Amount']),
                'HoldLevel': r['HoldLevel'] or 0
            })

        # 4. Investing
        inv_query = f'''
            SELECT a.account_name as Description,
                   cf.hold_level as HoldLevel,
                   SUM(CASE
                        WHEN a.account_basment = 'DR' THEN (ed.enty_values_DR - ed.enty_values_CR) * -1
                        WHEN a.account_basment = 'CR' THEN (ed.enty_values_CR - ed.enty_values_DR) * -1
                        ELSE 0 END) AS Amount
            FROM entry_details ed
            JOIN new_account_table a ON ed.account_name = a.account_name
            JOIN cf_catogory cf ON a.cf_catogory = cf.catogory_name
            WHERE cf.catogory_name = 'Investing Activities'
            AND ed.entry_effective_date BETWEEN %s AND %s
            AND ed.entry_deleted = 0 {rec_filter}
            GROUP BY a.account_name, cf.hold_level
            ORDER BY cf.hold_level
        '''
        cursor.execute(inv_query, (from_date, to_date))
        inv_items = [dict(r) for r in cursor.fetchall()]

        # 5. Financing
        fin_query = f'''
            SELECT a.account_name as Description,
                   cf.hold_level as HoldLevel,
                   SUM(CASE
                        WHEN a.account_basment = 'DR' THEN (ed.enty_values_DR - ed.enty_values_CR) * -1
                        WHEN a.account_basment = 'CR' THEN (ed.enty_values_CR - ed.enty_values_DR)
                        ELSE 0 END) AS Amount
            FROM entry_details ed
            JOIN new_account_table a ON ed.account_name = a.account_name
            JOIN cf_catogory cf ON a.cf_catogory = cf.catogory_name
            WHERE cf.catogory_name = 'Financing Activities'
            AND ed.entry_effective_date BETWEEN %s AND %s
            AND ed.entry_deleted = 0 {rec_filter}
            GROUP BY a.account_name, cf.hold_level
            ORDER BY cf.hold_level
        '''
        cursor.execute(fin_query, (from_date, to_date))
        fin_items = [dict(r) for r in cursor.fetchall()]

        # 6. Cash Balances
        cash_begin_query = f'''
            SELECT COALESCE(SUM(ed.enty_values_DR - ed.enty_values_CR), 0) as val
            FROM entry_details ed
            JOIN new_account_table a ON ed.account_name = a.account_name
            WHERE (a.account_name IN (SELECT bank_bookcol_account_number FROM bank_book)
                OR a.account_name IN (SELECT cash_book_account_name FROM cash_book))
            AND ed.entry_effective_date < %s
            AND ed.entry_deleted = 0 {rec_filter}
        '''
        cursor.execute(cash_begin_query, (from_date,))
        cash_begin = float(cursor.fetchone()['val'] or 0)

        cash_end_query = f'''
            SELECT COALESCE(SUM(ed.enty_values_DR - ed.enty_values_CR), 0) as val
            FROM entry_details ed
            JOIN new_account_table a ON ed.account_name = a.account_name
            WHERE (a.account_name IN (SELECT bank_bookcol_account_number FROM bank_book)
                OR a.account_name IN (SELECT cash_book_account_name FROM cash_book))
            AND ed.entry_effective_date <= %s
            AND ed.entry_deleted = 0 {rec_filter}
        '''
        cursor.execute(cash_end_query, (to_date,))
        cash_end = float(cursor.fetchone()['val'] or 0)

        # 7. Cash Accounts Breakdown
        cash_acc_query = f'''
            SELECT a.account_name as Description,
                   SUM(ed.enty_values_DR - ed.enty_values_CR) as Amount
            FROM entry_details ed
            JOIN new_account_table a ON ed.account_name = a.account_name
            WHERE (a.account_name IN (SELECT bank_bookcol_account_number FROM bank_book)
                OR a.account_name IN (SELECT cash_book_account_name FROM cash_book))
            AND ed.entry_effective_date <= %s
            AND ed.entry_deleted = 0 {rec_filter}
            GROUP BY a.account_name
        '''
        cursor.execute(cash_acc_query, (to_date,))
        cash_acc_items = [dict(r) for r in cursor.fetchall()]

        data = {
            'NetProfit': net_profit,
            'AdjustmentItems': adj_items,
            'WorkingCapitalItems': wc_items,
            'InvestingItems': inv_items,
            'FinancingItems': fin_items,
            'CashBeginning': cash_begin,
            'CashEnding': cash_end,
            'CashAccounts': cash_acc_items
        }

        return {'success': True, 'data': data}

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {'success': False, 'error': str(e)}

    finally:
        cursor.close()
        conn.close()


@app.route('/cash_flow/export_xlsx')
@login_required
@has_permission('Access_Reports')
def cash_flow_export_xlsx():
    """Styled Excel export of the cash flow statement (presentation format)."""
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    rec_only = request.args.get('rec_only') == '1'
    if not from_date or not to_date:
        flash('Select a date range first.', 'warning')
        return redirect(url_for('cash_flow_view'))
    try:
        import excel_export as xl
    except ImportError:
        flash("Excel export needs the 'openpyxl' package on the server — run: pip install openpyxl", 'warning')
        return redirect(url_for('cash_flow_view'))

    result = _compute_cash_flow(from_date, to_date, rec_only)
    if not result.get('success'):
        flash(f"Error generating cash flow: {result.get('error')}", 'danger')
        return redirect(url_for('cash_flow_view'))
    d = result['data']

    adj_total = sum(float(i['Amount'] or 0) for i in d['AdjustmentItems'])
    wc_total = sum(float(i['Amount'] or 0) for i in d['WorkingCapitalItems'])
    operating = float(d['NetProfit'] or 0) + adj_total + wc_total
    inv_total = sum(float(i['Amount'] or 0) for i in d['InvestingItems'])
    fin_total = sum(float(i['Amount'] or 0) for i in d['FinancingItems'])
    net_change = operating + inv_total + fin_total

    wb, ws = xl.new_workbook('Cash Flow')
    ncols = 2
    row = xl.title_block(ws, ncols, _company_display_name(), 'CASH FLOW STATEMENT',
                         f'For the period {from_date} to {to_date}')

    row = xl.section_row(ws, row, ncols, 'OPERATING ACTIVITIES', color=xl.GREEN_DARK, bg='EAF6EA')
    row = xl.item_row(ws, row, 'Net Profit / (Loss)', [d['NetProfit']])
    if d['AdjustmentItems']:
        row = xl.section_row(ws, row, ncols, 'Adjustments', color='495057', bg='F8F9FA')
        for i in d['AdjustmentItems']:
            row = xl.item_row(ws, row, i['Description'], [i['Amount']])
    if d['WorkingCapitalItems']:
        row = xl.section_row(ws, row, ncols, 'Changes in Working Capital', color='495057', bg='F8F9FA')
        for i in d['WorkingCapitalItems']:
            row = xl.item_row(ws, row, i['Description'], [i['Amount']])
    row = xl.total_row(ws, row, 'Net Cash from Operating Activities', [operating])

    row = xl.section_row(ws, row, ncols, 'INVESTING ACTIVITIES', color=xl.BLUE_DARK, bg='EFF6FF')
    for i in d['InvestingItems']:
        row = xl.item_row(ws, row, i['Description'], [i['Amount']])
    row = xl.total_row(ws, row, 'Net Cash from Investing Activities', [inv_total])

    row = xl.section_row(ws, row, ncols, 'FINANCING ACTIVITIES', color=xl.RED_DARK, bg='FDECEA')
    for i in d['FinancingItems']:
        row = xl.item_row(ws, row, i['Description'], [i['Amount']])
    row = xl.total_row(ws, row, 'Net Cash from Financing Activities', [fin_total])

    row += 1
    row = xl.total_row(ws, row, 'NET INCREASE / (DECREASE) IN CASH', [net_change],
                       bg='E3F2FD', color=xl.BLUE_DARK)
    row = xl.item_row(ws, row, 'Cash at Beginning of Period', [d['CashBeginning']])
    row = xl.total_row(ws, row, 'CASH AT END OF PERIOD', [d['CashEnding']],
                       bg='E3F2FD', color=xl.BLUE_DARK, size=11)
    if d['CashAccounts']:
        row += 1
        row = xl.section_row(ws, row, ncols, 'Cash & Bank Balances', color='495057', bg='F8F9FA')
        for i in d['CashAccounts']:
            row = xl.item_row(ws, row, i['Description'], [i['Amount']])
    xl.finish(ws, ncols)
    return xl.workbook_response(wb, f'CashFlow_{from_date}_{to_date}.xlsx')


@app.route('/inventory_item_save', methods=['POST'])
@login_required
def inventory_item_save():
    """Per-row AJAX save: prices + name + category, with change history logging."""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'No data'})

    item_id    = data.get('item_id')
    new_name   = (data.get('item_name') or '').strip()
    # Category values are foreign keys into inventory_carogory, so they must match a
    # parent key EXACTLY — do NOT strip/alter them. An empty selection ("None") maps
    # to NULL (FK columns allow NULL); "" would violate the constraint.
    new_main   = data.get('main_cat') or None
    new_sub    = data.get('sub_cat') or None
    prev_name  = data.get('old_name', '')
    prev_main  = data.get('old_main', '')
    prev_sub   = data.get('old_sub', '')
    market_p   = data.get('market_price', 0)
    spm_p      = data.get('spm_price', 0)
    loyalty_p  = data.get('loyalty_price', 0)

    user_code = session.get('user_id', 'unknown')
    user_pk   = session.get('user_pk', 0)

    try:
        existing = db.execute_query(
            "SELECT id FROM inventory_price_recod WHERE inventory_price_link = %s", (item_id,)
        )
        if existing:
            # Update only the latest price tier (matches what the table displays)
            db.execute_query("""
                UPDATE inventory_price_recod SET
                    inventory_price_selling = %s,
                    inventory_price_profit_marging_comen = %s,
                    inventory_price_for_Loyality_customer = %s
                WHERE inventory_price_link = %s
                ORDER BY id DESC LIMIT 1
            """, (market_p, spm_p, loyalty_p, item_id), commit=True)
        else:
            db.execute_query("""
                INSERT INTO inventory_price_recod
                (inventory_price_link, inventory_price_selling,
                 inventory_price_profit_marging_comen, inventory_price_for_Loyality_customer)
                VALUES (%s, %s, %s, %s)
            """, (item_id, market_p, spm_p, loyalty_p), commit=True)

        # Normalize for change detection only (None/"" and stray spaces treated equal)
        def _norm(v):
            return (v or '').strip()

        field_changes = []
        if new_name and new_name != prev_name:
            field_changes.append(('item_name', prev_name, new_name))
        if _norm(new_main) != _norm(prev_main):
            field_changes.append(('main_category', prev_main, new_main or ''))
        if _norm(new_sub) != _norm(prev_sub):
            field_changes.append(('sub_category', prev_sub, new_sub or ''))

        # Always write item name/category so a selection persists even if the page's
        # baseline ("old_*") was stale. This is the authoritative write for the item.
        effective_name = new_name or prev_name
        db.execute_query("""
            UPDATE inventoy_items SET
                inventoy_name = %s, Main_Catogry = %s, Sub_Catogory = %s
            WHERE id = %s
        """, (effective_name, new_main, new_sub, item_id), commit=True)

        for field, old_val, new_val in field_changes:
            try:
                db.execute_query("""
                    INSERT INTO inventory_item_change_history
                    (item_id, item_name, field_changed, old_value, new_value,
                     changed_by_user_code, changed_by_user_pk)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (item_id, effective_name, field,
                      old_val, new_val, user_code, user_pk), commit=True)
            except Exception:
                pass

        clear_pos_items_cache()

        # Verify what is actually stored now, and return it so the client can confirm
        check = db.execute_query(
            "SELECT inventoy_name, Main_Catogry, Sub_Catogory FROM inventoy_items WHERE id = %s",
            (item_id,)
        )
        stored = check[0] if check else {}
        logging.info(f"inventory_item_save item={item_id} -> stored={stored}")
        return jsonify({
            'success': True,
            'stored': {
                'item_name': stored.get('inventoy_name'),
                'main_cat':  stored.get('Main_Catogry'),
                'sub_cat':   stored.get('Sub_Catogory'),
            }
        })
    except Exception as e:
        logging.error(f"inventory_item_save error item={data.get('item_id')}: {e}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/add_new_price_tier', methods=['POST'])
@login_required
@has_permission('Access_Inventory')
def add_new_price_tier():
    item_id = request.form.get('item_id')
    cost_price = parse_float(request.form.get('cost_price', 0))
    selling_price = parse_float(request.form.get('selling_price', 0))
    special_price = parse_float(request.form.get('special_price', 0))
    loyalty_price = parse_float(request.form.get('loyalty_price', 0))

    if not item_id:
        flash('Must select an inventory item', 'danger')
        return redirect(url_for('inventory_price_editing'))

    try:
        query = '''
            INSERT INTO inventory_price_recod (
                id, inventory_price_link, inventory_price_purcharsing,
                inventory_price_selling, inventory_price_profit_marging_comen,
                inventory_price_for_Loyality_customer, created_date
            ) VALUES (0, %s, %s, %s, %s, %s, %s)
        '''
        db.execute_query(query, (item_id, cost_price, selling_price, special_price, loyalty_price, date.today()), commit=True)
        flash('New price tier added successfully!', 'success')
        clear_pos_items_cache()
    except Exception as e:
        flash(f'Error adding new price tier: {str(e)}', 'danger')

    return redirect(url_for('inventory_price_editing'))

# --- Inventory Image Endpoint ---

@app.route('/inventory_image/<int:item_id>')
@login_required
def inventory_image(item_id):
    """Serve inventory item image as JPEG. ?size=thumb resizes to 80×80."""
    row = db.execute_query(
        "SELECT inventoy_img FROM inventoy_items WHERE id = %s", (item_id,)
    )
    if not row or not row[0].get('inventoy_img'):
        return '', 404

    img_data = row[0]['inventoy_img']
    try:
        if isinstance(img_data, (bytes, bytearray)):
            raw = base64.b64decode(img_data.decode('utf-8'))
        else:
            raw = base64.b64decode(img_data)
    except Exception:
        return '', 404

    if request.args.get('size') == 'thumb':
        try:
            from PIL import Image as _PILImg
            import io as _io
            pil = _PILImg.open(_io.BytesIO(raw))
            pil.thumbnail((80, 80))
            buf = _io.BytesIO()
            pil.save(buf, format='JPEG', quality=75)
            raw = buf.getvalue()
        except Exception:
            pass

    from flask import Response as _Resp
    resp = _Resp(raw, mimetype='image/jpeg')
    resp.headers['Cache-Control'] = 'public, max-age=3600'
    return resp

# --- Inventory Balance ---

@app.route('/inventory_balance')
@login_required
@has_permission('Access_Inventory')
def inventory_balance():
    view = request.args.get('view', 'all')
    download = request.args.get('download')
    report_data = []

    try:
        if view == 'all':
            report_data = db.execute_query("""
                SELECT
                    ii.id,
                    ii.inventoy_name,
                    ii.inventoy_code,
                    ii.inventoy_bach_code,
                    ii.inventoy_items_messurment_unit AS unit,
                    ii.inventoy_img,
                    ii.Main_Catogry,
                    ii.Sub_Catogory,
                    COALESCE(ii.min_qty, 0) AS min_qty,
                    COALESCE(p.inventory_price_selling, 0) AS selling_price,
                    COALESCE(p.inventory_price_purcharsing, 0) AS cost_price,
                    COALESCE(SUM(r.inventory_recod_moument_in - r.inventory_recod_movment_out), 0) AS current_qty,
                    COALESCE(SUM(r.inventory_recod_total_value), 0) AS total_value
                FROM inventoy_items ii
                LEFT JOIN inventory_price_recod p ON ii.id = p.inventory_price_link
                LEFT JOIN inventory_recod r ON ii.inventoy_name = r.inventoy_name
                WHERE ii.active = 1
                GROUP BY ii.id, ii.inventoy_name, ii.inventoy_code, ii.inventoy_bach_code,
                         ii.inventoy_items_messurment_unit, ii.inventoy_img, ii.Main_Catogry,
                         ii.Sub_Catogory, ii.min_qty, p.inventory_price_selling, p.inventory_price_purcharsing
                ORDER BY ii.inventoy_name
            """)
        elif view == 'low':
            report_data = db.execute_query("""
                SELECT
                    ii.id,
                    ii.inventoy_name,
                    ii.inventoy_code,
                    ii.inventoy_items_messurment_unit,
                    ii.inventoy_img,
                    ii.Main_Catogry,
                    ii.Sub_Catogory,
                    COALESCE(ii.min_qty, 0) AS min_qty,
                    COALESCE(SUM(r.inventory_recod_moument_in - r.inventory_recod_movment_out), 0) AS current_balance
                FROM inventoy_items ii
                LEFT JOIN inventory_recod r ON ii.inventoy_name = r.inventoy_name
                WHERE ii.active = 1
                GROUP BY ii.id, ii.inventoy_name, ii.inventoy_code,
                         ii.inventoy_items_messurment_unit, ii.inventoy_img,
                         ii.Main_Catogry, ii.Sub_Catogory, ii.min_qty
                HAVING current_balance > 0 AND current_balance < COALESCE(ii.min_qty, 0)
                ORDER BY current_balance ASC
            """)
        elif view == 'out':
            report_data = db.execute_query("""
                SELECT
                    ii.id,
                    ii.inventoy_name,
                    ii.inventoy_code,
                    ii.inventoy_items_messurment_unit,
                    ii.inventoy_img,
                    ii.Main_Catogry,
                    ii.Sub_Catogory,
                    COALESCE(SUM(r.inventory_recod_moument_in - r.inventory_recod_movment_out), 0) AS current_balance
                FROM inventoy_items ii
                LEFT JOIN inventory_recod r ON ii.inventoy_name = r.inventoy_name
                WHERE ii.active = 1
                GROUP BY ii.id, ii.inventoy_name, ii.inventoy_code,
                         ii.inventoy_items_messurment_unit, ii.inventoy_img,
                         ii.Main_Catogry, ii.Sub_Catogory
                HAVING current_balance <= 0
                ORDER BY ii.inventoy_name
            """)
    except Exception as e:
        flash(f'Error loading inventory data: {str(e)}', 'danger')
        report_data = []

    # Images are now served via /inventory_image/<id> — just flag whether one exists.
    report_data = [
        dict(row, has_image=bool(row.get('inventoy_img')))
        for row in report_data
    ]

    if download == 'csv':
        si = io.StringIO()
        cw = csv.writer(si)

        if view == 'all':
            cw.writerow(['No', 'Item Name', 'Item Code', 'Barcode', 'Unit', 'Category Main', 'Category Sub',
                         'Min Qty', 'Selling Price', 'Cost Price', 'Current Qty', 'Total Value'])
            for i, r in enumerate(report_data):
                cw.writerow([
                    i + 1,
                    r.get('inventoy_name', ''),
                    r.get('inventoy_code', ''),
                    r.get('inventoy_bach_code', ''),
                    r.get('unit', ''),
                    r.get('Main_Catogry', ''),
                    r.get('Sub_Catogory', ''),
                    f"{r.get('min_qty', 0):.2f}",
                    f"{r.get('selling_price', 0):.2f}",
                    f"{r.get('cost_price', 0):.2f}",
                    f"{r.get('current_qty', 0):.2f}",
                    f"{r.get('total_value', 0):.2f}"
                ])
        elif view == 'low':
            cw.writerow(['No', 'Item Name', 'Item Code', 'Unit', 'Current Balance', 'Min Qty', 'Category Main', 'Category Sub'])
            for i, r in enumerate(report_data):
                cw.writerow([
                    i + 1,
                    r.get('inventoy_name', ''),
                    r.get('inventoy_code', ''),
                    r.get('inventoy_items_messurment_unit', ''),
                    f"{r.get('current_balance', 0):.2f}",
                    f"{r.get('min_qty', 0):.2f}",
                    r.get('Main_Catogry', ''),
                    r.get('Sub_Catogory', '')
                ])
        elif view == 'out':
            cw.writerow(['No', 'Item Name', 'Item Code', 'Unit', 'Category Main', 'Category Sub', 'Balance'])
            for i, r in enumerate(report_data):
                cw.writerow([
                    i + 1,
                    r.get('inventoy_name', ''),
                    r.get('inventoy_code', ''),
                    r.get('inventoy_items_messurment_unit', ''),
                    r.get('Main_Catogry', ''),
                    r.get('Sub_Catogory', ''),
                    f"{r.get('current_balance', 0):.2f}"
                ])

        output = make_response(si.getvalue())
        output.headers["Content-Disposition"] = f"attachment; filename=Inventory_Balance_{view}_{date.today()}.csv"
        output.headers["Content-type"] = "text/csv"
        return output

    # Pre-compute summary stats for the 'all' view
    stats = {}
    if view == 'all' and report_data:
        stats['total_items'] = len(report_data)
        stats['total_value'] = sum(float(r.get('total_value') or 0) for r in report_data)
        stats['out_stock']   = sum(1 for r in report_data if float(r.get('current_qty') or 0) <= 0)
        stats['low_stock']   = sum(
            1 for r in report_data
            if float(r.get('current_qty') or 0) > 0
            and float(r.get('min_qty') or 0) > 0
            and float(r.get('current_qty') or 0) < float(r.get('min_qty') or 0)
        )

    return render_template('inventory_balance.html', view=view, report_data=report_data, stats=stats)

# --- Inventory Reports ---
@app.route('/inventory_reports')
@login_required
@has_permission('Access_Reports')
def inventory_reports():
    report_type = request.args.get('report_type', 'balance')
    item_name = request.args.get('item_name')
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')

    items = db.execute_query("SELECT inventoy_name FROM inventoy_items")
    report_data = []
    opening_balance = 0

    if report_type == 'balance':
        # Inventory Balance
        rows = db.execute_query("""
            SELECT
                ii.inventoy_name as item_name, ii.inventoy_code as code, ii.inventoy_items_messurment_unit as unit,
                SUM(ir.inventory_recod_moument_in - ir.inventory_recod_movment_out) as qty,
                SUM((ir.inventory_recod_moument_in - ir.inventory_recod_movment_out) * ir.inventory_recod_unit_price) as value
            FROM inventoy_items ii
            LEFT JOIN inventory_recod ir ON ii.inventoy_name = ir.inventoy_name
            GROUP BY ii.id
        """)
        report_data = rows

    elif report_type == 'card' and item_name and from_date:
        # Opening Balance
        ob_res = db.execute_query("""
            SELECT SUM(inventory_recod_moument_in - inventory_recod_movment_out) as bal
            FROM inventory_recod
            WHERE inventoy_name = %s AND inventory_recod_action_date < %s
        """, (item_name, from_date))
        opening_balance = float(ob_res[0]['bal'] or 0) if ob_res else 0

        # Movements
        mvs = db.execute_query("""
            SELECT
                inventory_recod_action_date as date,
                inventory_recodcol_memo as description,
                inventory_recod_moument_in as in_qty,
                inventory_recod_movment_out as out_qty,
                inventory_recod_suplier_iv_no as iv_no,
                inventory_recod_unit_price as purchasing_price,
                inventory_recod_selling_price as selling_price
            FROM inventory_recod
            WHERE inventoy_name = %s AND inventory_recod_action_date BETWEEN %s AND %s
            ORDER BY inventory_recod_action_date
        """, (item_name, from_date, to_date))

        curr = opening_balance
        for m in mvs:
            curr += float(m['in_qty']) - float(m['out_qty'])
            m['balance'] = curr
            m['balance_value'] = curr * float(m['purchasing_price'] or 0)
            report_data.append(m)

    return render_template('inventory_reports.html', report_type=report_type, items=items, report_data=report_data, item_name=item_name, from_date=from_date, to_date=to_date, opening_balance=opening_balance)

# --- Bank Reconciliation ---
@app.route('/bank_reconciliation', methods=['GET'])
@login_required
def bank_reconciliation():
    bank_account = request.args.get('bank_account')
    rec_date = request.args.get('rec_date', date.today().strftime('%Y-%m-%d'))
    statement_balance = request.args.get('statement_balance', 0)

    bank_accounts = db.execute_query("SELECT bank_bookcol_account_number FROM bank_book")

    deposits = []
    payments = []
    book_balance = 0
    opening_balance = 0

    if bank_account:
        # Deposits (DR > 0, Not Reconciled, up to the reconciliation date).
        # GROUP BY ed.id: a JV can have several bank_book_recod rows (one per settled
        # invoice), which would duplicate the entry and double-count ticked amounts.
        deposits = db.execute_query("""
            SELECT
                ed.entry_save,
                ed.entry_date,
                ed.id,
                ed.entry_jv,
                ed.entry_effective_date,
                ed.entry_naration,
                MIN(bbr.bank_book_chque_no) AS bank_book_chque_no,
                ed.enty_values_DR
            FROM entry_details ed
            LEFT JOIN bank_book_recod bbr ON ed.entry_jv = bbr.jv_numbers_jv_id
            WHERE ed.account_name = %s
            AND ed.enty_values_DR > 0
            AND (ed.entry_Rec = 0 OR ed.entry_Rec IS NULL)
            AND ed.entry_deleted = 0
            AND DATE(COALESCE(ed.entry_effective_date, ed.entry_create_date)) <= %s
            GROUP BY ed.id, ed.entry_save, ed.entry_date, ed.entry_jv, ed.entry_effective_date,
                     ed.entry_naration, ed.enty_values_DR
            ORDER BY ed.entry_effective_date
        """, (bank_account, rec_date))

        # Payments (CR > 0, Not Reconciled, up to the reconciliation date)
        payments = db.execute_query("""
            SELECT
                ed.entry_save,
                ed.entry_date,
                ed.id,
                ed.entry_jv,
                ed.entry_effective_date,
                ed.entry_naration,
                MIN(bbr.bank_book_chque_no) AS bank_book_chque_no,
                ed.enty_values_CR
            FROM entry_details ed
            LEFT JOIN bank_book_recod bbr ON ed.entry_jv = bbr.jv_numbers_jv_id
            WHERE ed.account_name = %s
            AND ed.enty_values_CR > 0
            AND (ed.entry_Rec = 0 OR ed.entry_Rec IS NULL)
            AND ed.entry_deleted = 0
            AND DATE(COALESCE(ed.entry_effective_date, ed.entry_create_date)) <= %s
            GROUP BY ed.id, ed.entry_save, ed.entry_date, ed.entry_jv, ed.entry_effective_date,
                     ed.entry_naration, ed.enty_values_CR
            ORDER BY ed.entry_effective_date
        """, (bank_account, rec_date))

        # Book Balance — computed directly from the ledger (same basis as Ledger View).
        # The legacy bank_book_balance stored procedure returns no usable result on
        # some deployments, which made this silently show 0.
        try:
            bb_res = db.execute_query("""
                SELECT SUM(enty_values_DR) - SUM(enty_values_CR) as bal
                FROM entry_details
                WHERE account_name = %s
                  AND DATE(COALESCE(entry_effective_date, entry_create_date)) <= %s
                  AND entry_deleted = 0
            """, (bank_account, rec_date))
            book_balance = float(bb_res[0]['bal'] or 0) if bb_res else 0
        except Exception as e:
            logging.error(f"Book balance error: {e}")
            book_balance = 0

        # Opening Balance — net of everything already reconciled
        try:
            op_res = db.execute_query("""
                SELECT SUM(enty_values_DR) - SUM(enty_values_CR) as bal
                FROM entry_details
                WHERE account_name = %s AND entry_Rec = 1 AND entry_deleted = 0
            """, (bank_account,))
            opening_balance = float(op_res[0]['bal'] or 0) if op_res else 0
        except Exception as e:
            logging.error(f"Opening balance error: {e}")
            opening_balance = 0

    return render_template('bank_reconciliation.html',
                           bank_accounts=bank_accounts,
                           selected_account=bank_account,
                           rec_date=rec_date,
                           deposits=deposits,
                           payments=payments,
                           book_balance=book_balance,
                           opening_balance=opening_balance,
                           statement_balance=statement_balance)


@app.route('/bank_reconciliation/export_xlsx')
@login_required
def bank_reconciliation_export_xlsx():
    """Styled Excel export of the Bank Reconciliation working grid — the
    Uncleared Deposits / Uncleared Payments grids exactly as shown on screen
    for the selected account and as-at date. (This is the in-progress working
    paper; the completed-reconciliation report export already exists at
    /bank_reconciliation/report/<id> for a saved reconciliation.)"""
    bank_account = request.args.get('bank_account')
    rec_date = request.args.get('rec_date', date.today().strftime('%Y-%m-%d'))
    statement_balance = parse_float(request.args.get('statement_balance', 0))

    if not bank_account:
        flash('Select a bank account first.', 'warning')
        return redirect(url_for('bank_reconciliation'))

    try:
        import excel_export as xl
    except ImportError:
        flash("Excel export needs the 'openpyxl' package on the server — run: pip install openpyxl", 'warning')
        return redirect(url_for('bank_reconciliation', bank_account=bank_account, rec_date=rec_date,
                                 statement_balance=statement_balance))

    # Same queries as the page itself, so the export always matches what's on screen.
    deposits = db.execute_query("""
        SELECT
            ed.entry_save, ed.entry_date, ed.id, ed.entry_jv, ed.entry_effective_date,
            ed.entry_naration, MIN(bbr.bank_book_chque_no) AS bank_book_chque_no, ed.enty_values_DR
        FROM entry_details ed
        LEFT JOIN bank_book_recod bbr ON ed.entry_jv = bbr.jv_numbers_jv_id
        WHERE ed.account_name = %s
        AND ed.enty_values_DR > 0
        AND (ed.entry_Rec = 0 OR ed.entry_Rec IS NULL)
        AND ed.entry_deleted = 0
        AND DATE(COALESCE(ed.entry_effective_date, ed.entry_create_date)) <= %s
        GROUP BY ed.id, ed.entry_save, ed.entry_date, ed.entry_jv, ed.entry_effective_date,
                 ed.entry_naration, ed.enty_values_DR
        ORDER BY ed.entry_effective_date
    """, (bank_account, rec_date)) or []

    payments = db.execute_query("""
        SELECT
            ed.entry_save, ed.entry_date, ed.id, ed.entry_jv, ed.entry_effective_date,
            ed.entry_naration, MIN(bbr.bank_book_chque_no) AS bank_book_chque_no, ed.enty_values_CR
        FROM entry_details ed
        LEFT JOIN bank_book_recod bbr ON ed.entry_jv = bbr.jv_numbers_jv_id
        WHERE ed.account_name = %s
        AND ed.enty_values_CR > 0
        AND (ed.entry_Rec = 0 OR ed.entry_Rec IS NULL)
        AND ed.entry_deleted = 0
        AND DATE(COALESCE(ed.entry_effective_date, ed.entry_create_date)) <= %s
        GROUP BY ed.id, ed.entry_save, ed.entry_date, ed.entry_jv, ed.entry_effective_date,
                 ed.entry_naration, ed.enty_values_CR
        ORDER BY ed.entry_effective_date
    """, (bank_account, rec_date)) or []

    try:
        bb_res = db.execute_query("""
            SELECT SUM(enty_values_DR) - SUM(enty_values_CR) as bal
            FROM entry_details
            WHERE account_name = %s
              AND DATE(COALESCE(entry_effective_date, entry_create_date)) <= %s
              AND entry_deleted = 0
        """, (bank_account, rec_date))
        book_balance = float(bb_res[0]['bal'] or 0) if bb_res else 0
    except Exception:
        book_balance = 0

    try:
        op_res = db.execute_query("""
            SELECT SUM(enty_values_DR) - SUM(enty_values_CR) as bal
            FROM entry_details
            WHERE account_name = %s AND entry_Rec = 1 AND entry_deleted = 0
        """, (bank_account,))
        opening_balance = float(op_res[0]['bal'] or 0) if op_res else 0
    except Exception:
        opening_balance = 0

    wb, ws = xl.new_workbook('Bank Reconciliation')
    ncols = 5
    row = xl.title_block(ws, ncols, _company_display_name(),
                         'BANK RECONCILIATION — WORKING PAPER',
                         f'{bank_account}  ·  As at {rec_date}')

    row = xl.section_row(ws, row, ncols, 'UNCLEARED DEPOSITS / RECEIPTS', color=xl.GREEN_DARK, bg='EAF6EA')
    row = xl.header_row(ws, row, ['ID', 'Eff. Date', 'Description', 'Cheque No', 'Amount'])
    dep_total = 0.0
    for d in deposits:
        amt = float(d['enty_values_DR'] or 0)
        dep_total += amt
        row = xl.data_row(ws, row,
                          [d['id'], str(d['entry_effective_date']) if d['entry_effective_date'] else '',
                           d['entry_naration'] or '', d['bank_book_chque_no'] or '-', amt],
                          num_cols=(5,))
    row = xl.total_row(ws, row, f'Deposits Total ({len(deposits)})', [0, 0, 0, dep_total], bg='EAF6EA', color=xl.GREEN_DARK)
    for col in (2, 3, 4):
        ws.cell(row=row - 1, column=col).value = None
    row += 2

    row = xl.section_row(ws, row, ncols, 'UNCLEARED PAYMENTS', color=xl.RED_DARK, bg='FDEDEC')
    row = xl.header_row(ws, row, ['ID', 'Eff. Date', 'Description', 'Cheque No', 'Amount'])
    pay_total = 0.0
    for p in payments:
        amt = float(p['enty_values_CR'] or 0)
        pay_total += amt
        row = xl.data_row(ws, row,
                          [p['id'], str(p['entry_effective_date']) if p['entry_effective_date'] else '',
                           p['entry_naration'] or '', p['bank_book_chque_no'] or '-', amt],
                          num_cols=(5,))
    row = xl.total_row(ws, row, f'Payments Total ({len(payments)})', [0, 0, 0, pay_total], bg='FDEDEC', color=xl.RED_DARK)
    for col in (2, 3, 4):
        ws.cell(row=row - 1, column=col).value = None
    row += 2

    row = xl.section_row(ws, row, ncols, 'SUMMARY', color=xl.BLUE_DARK, bg='E3F2FD')
    row = xl.item_row(ws, row, 'Opening Balance (already reconciled)', [opening_balance])
    row = xl.item_row(ws, row, 'Uncleared Deposits (above)', [dep_total])
    row = xl.item_row(ws, row, 'Uncleared Payments (above)', [-pay_total])
    row = xl.total_row(ws, row, 'Book Balance', [book_balance], bg='F3F3F3', color='1A1A2E', size=11)
    if statement_balance:
        row = xl.item_row(ws, row, 'Statement Balance (entered)', [statement_balance])
        row = xl.total_row(ws, row, 'Difference (Book − Statement)', [book_balance - statement_balance],
                           bg='FFF8E1', color='8A6D00', size=11)

    xl.finish(ws, ncols, first_col_width=10, num_col_width=16)
    ws.column_dimensions['C'].width = 42
    ws.column_dimensions['D'].width = 16

    safe_acc = str(bank_account).replace(' ', '_').replace('/', '-')[:30]
    fname = f'BankReconciliation_{safe_acc}_{rec_date}.xlsx'
    return xl.workbook_response(wb, fname)


@app.route('/bank_reconciliation/process', methods=['POST'])
@login_required
def process_reconciliation():
    action = request.form.get('action') # 'save' or 'process'
    bank_account = request.form.get('bank_account')
    rec_date = request.form.get('rec_date')
    statement_balance = request.form.get('statement_balance', 0)

    if not bank_account:
        flash('Missing bank account', 'danger')
        return redirect(url_for('bank_reconciliation'))

    conn = None
    try:
        with db.get_connection() as conn:
            with conn.cursor(dictionary=True) as cursor:
                conn.start_transaction()
                # Parse cleared items and their dates
                # Format: 'id|date' for deposits and payments
                cleared_deposits = request.form.getlist('cleared_deposits[]')
                cleared_payments = request.form.getlist('cleared_payments[]')

                uncleared_deposits = request.form.getlist('uncleared_deposits[]')
                uncleared_payments = request.form.getlist('uncleared_payments[]')

                if action == 'save':
                    # Save Progress
                    # Process cleared deposits
                    for d in cleared_deposits:
                        parts = d.split('|')
                        d_id = parts[0]
                        d_date = parts[1] if len(parts) > 1 and parts[1] else datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        cursor.execute("UPDATE entry_details SET entry_save = 1, entry_date = %s WHERE id = %s", (d_date, d_id))

                    # Process cleared payments
                    for p in cleared_payments:
                        parts = p.split('|')
                        p_id = parts[0]
                        p_date = parts[1] if len(parts) > 1 and parts[1] else datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        cursor.execute("UPDATE entry_details SET entry_save = 1, entry_date = %s WHERE id = %s", (p_date, p_id))

                    # Mark uncleared as entry_save = 0
                    for d in uncleared_deposits:
                        cursor.execute("UPDATE entry_details SET entry_save = 0, entry_date = NULL WHERE id = %s", (d,))
                    for p in uncleared_payments:
                        cursor.execute("UPDATE entry_details SET entry_save = 0, entry_date = NULL WHERE id = %s", (p,))

                    conn.commit()
                    flash('Progress saved successfully!', 'success')

                elif action == 'process':
                    if not rec_date:
                        flash('Missing reconciliation date', 'danger')
                        return redirect(url_for('bank_reconciliation', bank_account=bank_account))

                    # Calculate Opening Balance (Simplified for Python version as in original code)
                    cursor.execute("SELECT SUM(enty_values_DR) - SUM(enty_values_CR) as bal FROM entry_details WHERE account_name = %s AND entry_Rec = 1", (bank_account,))
                    op_res = cursor.fetchone()
                    opening_balance = float(op_res['bal'] or 0) if op_res else 0

                    # Book Balance
                    cursor.execute("SELECT SUM(enty_values_DR) - SUM(enty_values_CR) as bal FROM entry_details WHERE account_name = %s AND entry_effective_date <= %s", (bank_account, rec_date))
                    bb_res = cursor.fetchone()
                    book_balance = float(bb_res['bal'] or 0) if bb_res else 0

                    statement_balance = float(statement_balance)

                    # Get sums of cleared deposits/payments to calculate closing balance
                    cleared_dep_sum = 0
                    cleared_pay_sum = 0

                    # Clear Deposits
                    for d in cleared_deposits:
                        parts = d.split('|')
                        d_id = parts[0]
                        d_date = parts[1] if len(parts) > 1 and parts[1] else rec_date
                        cursor.execute("UPDATE entry_details SET entry_Rec = 1, entry_effective_date = %s, entry_save = 1, entry_date = %s WHERE id = %s", (d_date, d_date + " 00:00:00", d_id))
                        cursor.execute("SELECT enty_values_DR FROM entry_details WHERE id = %s", (d_id,))
                        cleared_dep_sum += float(cursor.fetchone()['enty_values_DR'] or 0)

                    # Clear Payments
                    for p in cleared_payments:
                        parts = p.split('|')
                        p_id = parts[0]
                        p_date = parts[1] if len(parts) > 1 and parts[1] else rec_date
                        cursor.execute("UPDATE entry_details SET entry_Rec = 1, entry_effective_date = %s, entry_save = 1, entry_date = %s WHERE id = %s", (p_date, p_date + " 00:00:00", p_id))
                        cursor.execute("SELECT enty_values_CR FROM entry_details WHERE id = %s", (p_id,))
                        cleared_pay_sum += float(cursor.fetchone()['enty_values_CR'] or 0)

                    closing_balance = opening_balance + cleared_dep_sum - cleared_pay_sum

                    # Get last closing date for opene_date
                    cursor.execute("SELECT MAX(closing_date) as cd FROM bank_reconciliation_recodes WHERE bank_accont_no = %s", (bank_account,))
                    last_cd_res = cursor.fetchone()
                    last_closing_date = last_cd_res['cd'] if last_cd_res and last_cd_res['cd'] else '2000-01-01'

                    # Check tables
                    cursor.execute("SHOW TABLES LIKE 'bank_reconciliation_recodes'")
                    if not cursor.fetchone():
                        cursor.execute('''CREATE TABLE bank_reconciliation_recodes (
                            id INT AUTO_INCREMENT PRIMARY KEY,
                            opene_date DATE,
                            opene_balance DECIMAL(15,2),
                            closing_date DATE,
                            closing_balance DECIMAL(15,2),
                            save_user INT,
                            close_user INT,
                            last_save_user INT,
                            last_close_uer INT,
                            period_close_or_open INT,
                            bank_accont_no VARCHAR(100),
                            Book_Balance DECIMAL(15,2),
                            Bank_statment_Balance DECIMAL(15,2)
                        )''')

                    cursor.execute("SHOW TABLES LIKE 'bankreconciliiationditails'")
                    if not cursor.fetchone():
                        cursor.execute('''CREATE TABLE bankreconciliiationditails (
                            Id INT AUTO_INCREMENT PRIMARY KEY,
                            Key_to_Recode_Table INT,
                            Transaktion_Id INT,
                            Dr_Value DECIMAL(15,2),
                            Cr_Value DECIMAL(15,2),
                            Text TEXT,
                            Chq_No VARCHAR(100)
                        )''')

                    # Insert reconciliation record
                    cursor.execute('''
                        INSERT INTO bank_reconciliation_recodes
                        (opene_date, opene_balance, closing_date, closing_balance, save_user, close_user,
                         period_close_or_open, bank_accont_no, Book_Balance, Bank_statment_Balance)
                        VALUES (%s, %s, %s, %s, 0, %s, 1, %s, %s, %s)
                    ''', (last_closing_date, opening_balance, rec_date, closing_balance, get_current_user_id(), bank_account, book_balance, statement_balance))

                    rec_id = cursor.lastrowid

                    # Insert Uncleared Transactions into details table
                    for d in uncleared_deposits:
                        cursor.execute("SELECT id, enty_values_DR, entry_naration FROM entry_details WHERE id = %s", (d,))
                        detail = cursor.fetchone()
                        if detail:
                            cursor.execute("INSERT INTO bankreconciliiationditails (Key_to_Recode_Table, Transaktion_Id, Dr_Value, Cr_Value, Text, Chq_No) VALUES (%s, %s, %s, %s, %s, %s)",
                                         (rec_id, detail['id'], detail['enty_values_DR'], 0, detail['entry_naration'], ''))

                    for p in uncleared_payments:
                        cursor.execute("SELECT id, enty_values_CR, entry_naration FROM entry_details WHERE id = %s", (p,))
                        detail = cursor.fetchone()
                        if detail:
                            cursor.execute("INSERT INTO bankreconciliiationditails (Key_to_Recode_Table, Transaktion_Id, Dr_Value, Cr_Value, Text, Chq_No) VALUES (%s, %s, %s, %s, %s, %s)",
                                         (rec_id, detail['id'], 0, detail['enty_values_CR'], detail['entry_naration'], ''))

                    conn.commit()
                    flash(f'Reconciliation processed successfully! ID: {rec_id}', 'success')

    except Exception as e:
            conn.rollback()
            flash(f'Error processing reconciliation: {str(e)}', 'danger')

    return redirect(url_for('bank_reconciliation', bank_account=bank_account, rec_date=rec_date))

@app.route('/bank_reconciliation/history', methods=['GET'])
@login_required
def bank_reconciliation_history():
    bank_account = request.args.get('bank_account')

    bank_accounts = db.execute_query("SELECT bank_bookcol_account_number FROM bank_book")
    history = []

    if bank_account:
        try:
            history = db.execute_query('''
                SELECT id, opene_date, opene_balance, closing_date, closing_balance,
                       Book_Balance, Bank_statment_Balance, close_user, period_close_or_open
                FROM bank_reconciliation_recodes
                WHERE bank_accont_no = %s
                ORDER BY closing_date DESC
            ''', (bank_account,))
        except Exception:
            history = []

    return render_template('bank_reconciliation_history.html', bank_accounts=bank_accounts, selected_account=bank_account, history=history)

@app.route('/bank_reconciliation/report/<int:rec_id>', methods=['GET'])
@login_required
def bank_reconciliation_report(rec_id):
    rec = None
    deposits = []
    payments = []

    try:
        rec = db.execute_query("SELECT * FROM bank_reconciliation_recodes WHERE id = %s", (rec_id,))
        if rec:
            rec = rec[0]

            deposits = db.execute_query('''
                SELECT brd.Id, brd.Text, brd.Chq_No, brd.Dr_Value, ed.entry_effective_date
                FROM bankreconciliiationditails brd
                LEFT JOIN entry_details ed ON brd.Transaktion_Id = ed.id
                WHERE brd.Key_to_Recode_Table = %s AND brd.Dr_Value > 0
            ''', (rec_id,))

            payments = db.execute_query('''
                SELECT brd.Id, brd.Text, brd.Chq_No, brd.Cr_Value, ed.entry_effective_date
                FROM bankreconciliiationditails brd
                LEFT JOIN entry_details ed ON brd.Transaktion_Id = ed.id
                WHERE brd.Key_to_Recode_Table = %s AND brd.Cr_Value > 0
            ''', (rec_id,))

    except Exception as e:
        flash(f'Error loading report: {str(e)}', 'danger')

    # Styled Excel export (presentation format)
    if rec and request.args.get('download') == 'xlsx':
        try:
            import excel_export as xl
        except ImportError:
            flash("Excel export needs the 'openpyxl' package on the server — run: pip install openpyxl", 'warning')
            return redirect(url_for('bank_reconciliation_report', rec_id=rec_id))

        wb, ws = xl.new_workbook('Reconciliation')
        ncols = 5
        row = xl.title_block(ws, ncols, _company_display_name(),
                             f'BANK RECONCILIATION REPORT #{rec_id}',
                             f"Account {rec.get('bank_accont_no', '')} — "
                             f"{rec.get('opene_date', '')} to {rec.get('closing_date', '')}")

        row = xl.section_row(ws, row, ncols, 'BALANCES', color=xl.BLUE_DARK, bg='EFF6FF')
        row = xl.item_row(ws, row, 'Opening Balance', [rec.get('opene_balance', 0)])
        row = xl.item_row(ws, row, 'Closing Balance', [rec.get('closing_balance', 0)])
        row = xl.item_row(ws, row, 'Book Balance', [rec.get('Book_Balance', 0)])
        row = xl.item_row(ws, row, 'Bank Statement Balance', [rec.get('Bank_statment_Balance', 0)])
        diff = float(rec.get('closing_balance', 0) or 0) - float(rec.get('Bank_statment_Balance', 0) or 0)
        row = xl.total_row(ws, row, 'Difference', [diff],
                           bg=('E8F5E9' if abs(diff) < 0.01 else 'FFEBEE'),
                           color=(xl.GREEN_DARK if abs(diff) < 0.01 else xl.RED_DARK))
        row += 1

        row = xl.section_row(ws, row, ncols, 'UNCLEARED DEPOSITS', color=xl.GREEN_DARK, bg='EAF6EA')
        row = xl.header_row(ws, row, ['ID', 'Date', 'Description', 'Cheque No', 'Amount'])
        dep_total = 0.0
        for d in deposits:
            dep_total += float(d.get('Dr_Value', 0) or 0)
            row = xl.data_row(ws, row, [d.get('Id'), str(d.get('entry_effective_date') or ''),
                                        d.get('Text', ''), d.get('Chq_No') or '-',
                                        d.get('Dr_Value', 0)], num_cols=(5,))
        row = xl.total_row(ws, row, 'Total Uncleared Deposits', [0] * 3 + [dep_total])
        for col in range(2, 5):
            ws.cell(row=row - 1, column=col).value = None
        row += 1

        row = xl.section_row(ws, row, ncols, 'UNCLEARED PAYMENTS', color=xl.RED_DARK, bg='FDECEA')
        row = xl.header_row(ws, row, ['ID', 'Date', 'Description', 'Cheque No', 'Amount'])
        pay_total = 0.0
        for p in payments:
            pay_total += float(p.get('Cr_Value', 0) or 0)
            row = xl.data_row(ws, row, [p.get('Id'), str(p.get('entry_effective_date') or ''),
                                        p.get('Text', ''), p.get('Chq_No') or '-',
                                        p.get('Cr_Value', 0)], num_cols=(5,))
        row = xl.total_row(ws, row, 'Total Uncleared Payments', [0] * 3 + [pay_total])
        for col in range(2, 5):
            ws.cell(row=row - 1, column=col).value = None

        xl.finish(ws, ncols, first_col_width=12, num_col_width=16)
        ws.column_dimensions['C'].width = 46
        return xl.workbook_response(wb, f'BankReconciliation_{rec_id}.xlsx')

    return render_template('bank_reconciliation_report.html', rec=rec, deposits=deposits, payments=payments)

@app.route('/bank_reconciliation/reverse', methods=['POST'])
@login_required
def reverse_reconciliation():
    rec_id = request.form.get('rec_id')
    reason = request.form.get('reason')

    if not rec_id or not reason:
        flash('Missing reconciliation ID or reason', 'danger')
        return redirect(url_for('bank_reconciliation_history'))

    conn = None
    cursor = None
    account = ''
    try:
        conn = db.get_connection()
        cursor = conn.cursor(dictionary=True)
        conn.start_transaction()

        # Check tables
        cursor.execute("SHOW TABLES LIKE 'bank_reconciliation_reversal_log'")
        if not cursor.fetchone():
            cursor.execute('''CREATE TABLE bank_reconciliation_reversal_log (
                id INT AUTO_INCREMENT PRIMARY KEY,
                original_rec_id INT,
                bank_account VARCHAR(100),
                reversal_date DATETIME,
                reversed_by_user INT,
                opening_balance DECIMAL(15,2),
                closing_balance DECIMAL(15,2),
                reversal_reason TEXT
            )''')

        # Get Reconciliation details
        cursor.execute("SELECT * FROM bank_reconciliation_recodes WHERE id = %s", (rec_id,))
        rec = cursor.fetchone()
        if not rec:
            flash('Reconciliation record not found', 'danger')
            return redirect(url_for('bank_reconciliation_history'))

        account = rec['bank_accont_no']

        # Step 1: Get all transaction IDs from details table
        cursor.execute("SELECT Transaktion_Id FROM bankreconciliiationditails WHERE Key_to_Recode_Table = %s", (rec_id,))
        detail_trans_ids = [row['Transaktion_Id'] for row in cursor.fetchall() if row['Transaktion_Id']]

        # Step 2: Get cleared transaction IDs for this account and period
        cursor.execute('''
            SELECT id FROM entry_details
            WHERE account_name = %s
            AND entry_Rec = 1
            AND entry_effective_date BETWEEN %s AND %s
        ''', (account, rec['opene_date'], rec['closing_date']))
        cleared_trans_ids = [row['id'] for row in cursor.fetchall()]

        # Combine and remove duplicates
        all_trans_ids = list(set(detail_trans_ids + cleared_trans_ids))

        # Step 3: Un-reconcile all these transactions
        if all_trans_ids:
            format_strings = ','.join(['%s'] * len(all_trans_ids))
            cursor.execute(f'''
                UPDATE entry_details
                SET entry_Rec = 0, entry_save = 0, entry_date = NULL
                WHERE id IN ({format_strings})
            ''', tuple(all_trans_ids))

        # Step 4: Log Reversal
        cursor.execute('''
            INSERT INTO bank_reconciliation_reversal_log
            (original_rec_id, bank_account, reversal_date, reversed_by_user,
             opening_balance, closing_balance, reversal_reason)
            VALUES (%s, %s, NOW(), %s, %s, %s, %s)
        ''', (rec_id, account, get_current_user_id(), rec['opene_balance'], rec['closing_balance'], reason))

        # Step 5: Delete Details
        cursor.execute("DELETE FROM bankreconciliiationditails WHERE Key_to_Recode_Table = %s", (rec_id,))

        # Step 6: Delete Record
        cursor.execute("DELETE FROM bank_reconciliation_recodes WHERE id = %s", (rec_id,))

        conn.commit()
        flash('Reconciliation reversed successfully!', 'success')
    except Exception as e:
        if conn:
            conn.rollback()
        flash(f'Error reversing reconciliation: {str(e)}', 'danger')
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

    return redirect(url_for('bank_reconciliation_history', bank_account=account))


# --- Ledger View ---
@app.route('/ledger_view')
@login_required
@has_permission('Access_Reports')
def ledger_view():
    # Fetch accounts grouped/labeled like C# (P&L vs BS)
    # C# Logic:
    # 1. P&L: account_income=1 OR account_expenses=1
    # 2. BS: account_assets=1 OR account_liabilities=1 OR account_equity=1

    pl_accounts = db.execute_query("SELECT account_name FROM new_account_table WHERE account_income = 1 OR account_expenses = 1 ORDER BY account_name")
    bs_accounts = db.execute_query("SELECT account_name FROM new_account_table WHERE account_assets = 1 OR account_liabilities = 1 OR account_equity = 1 ORDER BY account_name")

    # All active sub-accounts, with which parent account each belongs to, so
    # the page can filter the Sub Account dropdown client-side as soon as an
    # Account is picked — no extra round trip needed.
    all_sub_accounts = db.execute_query(
        "SELECT sub_account_code, sub_sub_accaount_name, sub_new_account "
        "FROM sub_accont_for_new_account WHERE active = 1 ORDER BY sub_sub_accaount_name") or []

    # Structure for Select2 optgroups
    # However, HTML select structure is easier:
    return render_template('ledger_view.html',
                           pl_accounts=pl_accounts,
                           bs_accounts=bs_accounts,
                           all_sub_accounts=all_sub_accounts,
                           default_from=date.today().replace(day=1).strftime('%Y-%m-%d'),
                           default_to=date.today().strftime('%Y-%m-%d'))

@app.route('/api/ledger_data', methods=['POST'])
@login_required
def get_ledger_data():
    account_name = request.json.get('account_name')
    from_date = request.json.get('from_date')
    to_date = request.json.get('to_date')
    sub_account_code = request.json.get('sub_account_code') or None

    if not account_name or not from_date or not to_date:
        return {'error': 'Missing parameters'}, 400
    result = _ledger_data(account_name, from_date, to_date, sub_account_code)
    if 'error' in result:
        return result, 404
    return result


def _ledger_data(account_name, from_date, to_date, sub_account_code=None):
    """Ledger rows + running balance for one account, optionally narrowed to
    a single sub-account of it. Shared by the JSON API, the CSV export and
    the styled Excel export."""
    # 1. Get Account Basement (DR/CR)
    acc_res = db.execute_query("SELECT account_basment FROM new_account_table WHERE account_name = %s", (account_name,))
    if not acc_res:
        return {'error': 'Account not found'}
    basement = acc_res[0]['account_basment'] # 'DR' or 'CR'

    sub_filter_sql = " AND entry_sub_account_code = %s" if sub_account_code else ""
    sub_filter_params = (sub_account_code,) if sub_account_code else ()

    # 2. Calculate Opening Balance
    # Logic: Sum previous entries based on basement

    op_dr = 0
    op_cr = 0

    # Calculate Sums before from_date — filter strictly by EFFECTIVE date (when the
    # transaction occurred), falling back to create date only when effective is NULL.
    op_res = db.execute_query(f"""
        SELECT SUM(enty_values_DR), SUM(enty_values_CR)
        FROM entry_details
        WHERE account_name = %s
          AND DATE(COALESCE(entry_effective_date, entry_create_date)) < %s
          AND entry_deleted = 0
          {sub_filter_sql}
    """, (account_name, from_date) + sub_filter_params)

    if op_res:
        op_dr = float(op_res[0]['SUM(enty_values_DR)'] or 0)
        op_cr = float(op_res[0]['SUM(enty_values_CR)'] or 0)

    opening_balance = 0
    if basement == 'DR':
        opening_balance = op_dr - op_cr
    else: # CR
        opening_balance = op_cr - op_dr


    # 3. Fetch Transactions — show and filter by EFFECTIVE date (transaction date),
    # not the entry/create date. DATE() so datetime values on the last day are included.
    rows = db.execute_query(f"""
        SELECT
            COALESCE(ed.entry_effective_date, ed.entry_create_date) as date,
            ed.entry_naration as narration,
            ed.enty_values_DR as dr,
            ed.enty_values_CR as cr,
            ed.entry_jv as jv_no,
            ed.entry_sub_account_code as sub_code,
            s.sub_sub_accaount_name as sub_account
        FROM entry_details ed
        LEFT JOIN sub_accont_for_new_account s
            ON ed.entry_sub_account_code = s.sub_account_code AND ed.entry_sub_account_code != 0
        WHERE ed.account_name = %s
          AND DATE(COALESCE(ed.entry_effective_date, ed.entry_create_date)) BETWEEN %s AND %s
          AND ed.entry_deleted = 0
          {sub_filter_sql}
        ORDER BY COALESCE(ed.entry_effective_date, ed.entry_create_date), ed.id
    """, (account_name, from_date, to_date) + sub_filter_params)


    # 4. Process Running Balance
    data = []

    # Add Opening Balance Row
    data.append({
        'date': from_date,
        'narration': 'Opening Balance',
        'dr': 0,
        'cr': 0,
        'balance': opening_balance,
        'is_opening': True
    })

    current_bal = opening_balance

    # Sub-account breakdown for the period — only meaningful when looking at
    # the WHOLE account (not already narrowed to one sub-account), so a
    # "Sub Account Summary" can sit under the ledger showing how this
    # account's movement splits across its sub-accounts.
    sub_totals = {} if not sub_account_code else None

    for r in rows:
        dr = float(r['dr'] or 0)
        cr = float(r['cr'] or 0)

        if basement == 'DR':
            current_bal = current_bal + dr - cr
        else: # CR
            current_bal = current_bal + cr - dr


        narration_text = r['narration']
        # Skip the "[sub-account]" prefix when we've already filtered down to
        # one sub-account — every row would repeat the same label.
        if r.get('sub_account') and not sub_account_code:
            narration_text = f"[{r['sub_account']}] {narration_text}"

        data.append({
            'date': str(r['date']),
            'narration': narration_text,
            'dr': dr,
            'cr': cr,
            'balance': current_bal,
            'jv_no': r['jv_no'],
            'is_opening': False
        })

        if sub_totals is not None:
            key = r.get('sub_code') or 0
            if key not in sub_totals:
                sub_totals[key] = {'code': key, 'name': r.get('sub_account') or '(Unallocated)',
                                   'dr': 0.0, 'cr': 0.0}
            sub_totals[key]['dr'] += dr
            sub_totals[key]['cr'] += cr

    sub_summary = None
    if sub_totals is not None:
        sub_summary = []
        for v in sub_totals.values():
            net = (v['dr'] - v['cr']) if basement == 'DR' else (v['cr'] - v['dr'])
            sub_summary.append({'code': v['code'], 'name': v['name'],
                                'dr': v['dr'], 'cr': v['cr'], 'net': net})
        sub_summary.sort(key=lambda s: (s['code'] == 0, s['name'].lower()))

    return {'data': data, 'basement': basement, 'sub_summary': sub_summary}


@app.route('/api/jv_details/<jv_no>')
@login_required
def api_jv_details(jv_no):
    """Return all journal lines for a JV plus its header, for the ledger drill-down."""
    header = db.execute_query(
        "SELECT jv_id, jv_naration, jv_user_code, status FROM jv_numbers WHERE jv_id = %s", (jv_no,)
    )
    lines = db.execute_query("""
        SELECT ed.account_name,
               ed.enty_values_DR AS dr,
               ed.enty_values_CR AS cr,
               COALESCE(ed.entry_effective_date, ed.entry_create_date) AS date,
               ed.entry_naration AS narration,
               ed.entry_job_number AS job_no,
               ed.entry_create_user AS create_user,
               s.sub_sub_accaount_name AS sub_account
        FROM entry_details ed
        LEFT JOIN sub_accont_for_new_account s
            ON ed.entry_sub_account_code = s.sub_account_code AND ed.entry_sub_account_code != 0
        WHERE ed.entry_jv = %s AND ed.entry_deleted = 0
        ORDER BY ed.id
    """, (jv_no,))

    if not lines:
        return {'error': 'No journal details found for this JV.'}, 404

    # Resolve the actual user who entered it (jv_user_code is a source label like
    # 'JV FORM SEN INVOICE', not a person). entry_create_user may hold the Login id
    # or the User_Code depending on the transaction type — match either.
    entered_by = ''
    create_user = lines[0].get('create_user') if lines else None
    if create_user:
        try:
            u = db.execute_query(
                "SELECT User_Name FROM Login_Table WHERE id = %s OR User_Code = %s LIMIT 1",
                (create_user, create_user)
            )
            if u:
                entered_by = u[0].get('User_Name') or ''
        except Exception:
            pass
    if not entered_by:
        entered_by = (header[0]['jv_user_code'] if header else '') or ''

    total_dr = 0.0
    total_cr = 0.0
    for l in lines:
        l['dr'] = float(l['dr'] or 0)
        l['cr'] = float(l['cr'] or 0)
        l['date'] = str(l['date']) if l['date'] is not None else ''
        l['job_no'] = l['job_no'] if l['job_no'] else ''
        l['sub_account'] = l['sub_account'] or ''
        l.pop('create_user', None)
        total_dr += l['dr']
        total_cr += l['cr']

    return {
        'jv_no': jv_no,
        'narration': (header[0]['jv_naration'] if header else '') or '',
        'user': entered_by,
        'status': (header[0]['status'] if header else None),
        'lines': lines,
        'total_dr': total_dr,
        'total_cr': total_cr,
    }


@app.route('/ledger_view/export')
@login_required
@has_permission('Access_Reports')
def ledger_export():
    """CSV export of ledger for a single account."""
    account_name = request.args.get('account_name', '')
    from_date    = request.args.get('from_date', date.today().replace(day=1).strftime('%Y-%m-%d'))
    to_date      = request.args.get('to_date', date.today().strftime('%Y-%m-%d'))
    sub_account_code = request.args.get('sub_account_code') or None
    sub_account_name = request.args.get('sub_account_name') or ''
    if not account_name:
        flash('No account selected', 'warning')
        return redirect(url_for('ledger_view'))

    result = _ledger_data(account_name, from_date, to_date, sub_account_code)
    rows = result.get('data', [])
    si = io.StringIO()
    cw = csv.writer(si)
    cw.writerow([f'Ledger: {account_name}' + (f' — {sub_account_name}' if sub_account_name else '')])
    cw.writerow([f'Period: {from_date} to {to_date}'])
    cw.writerow([])
    cw.writerow(['Date', 'Narration', 'Debit', 'Credit', 'Balance'])
    for r in rows:
        cw.writerow([
            r.get('date', ''),
            r.get('narration', ''),
            f"{float(r.get('dr', 0)):.2f}" if not r.get('is_opening') else '',
            f"{float(r.get('cr', 0)):.2f}" if not r.get('is_opening') else '',
            f"{float(r.get('balance', 0)):.2f}"
        ])
    output = make_response(si.getvalue())
    safe_name = account_name.replace(' ', '_').replace('/', '-')[:40]
    output.headers["Content-Disposition"] = f"attachment; filename=Ledger_{safe_name}_{from_date}.csv"
    output.headers["Content-type"] = "text/csv"
    return output


@app.route('/ledger_view/export_xlsx')
@login_required
@has_permission('Access_Reports')
def ledger_export_xlsx():
    """Styled Excel export of the ledger for one account (presentation format)."""
    account_name = request.args.get('account_name', '')
    from_date = request.args.get('from_date', date.today().replace(day=1).strftime('%Y-%m-%d'))
    to_date = request.args.get('to_date', date.today().strftime('%Y-%m-%d'))
    sub_account_code = request.args.get('sub_account_code') or None
    sub_account_name = request.args.get('sub_account_name') or ''
    if not account_name:
        flash('No account selected', 'warning')
        return redirect(url_for('ledger_view'))
    try:
        import excel_export as xl
    except ImportError:
        flash("Excel export needs the 'openpyxl' package on the server — run: pip install openpyxl", 'warning')
        return redirect(url_for('ledger_view'))

    result = _ledger_data(account_name, from_date, to_date, sub_account_code)
    if 'error' in result:
        flash(f"Error: {result['error']}", 'danger')
        return redirect(url_for('ledger_view'))
    rows = result.get('data', [])

    wb, ws = xl.new_workbook('Ledger')
    ncols = 6
    row = xl.title_block(ws, ncols, _company_display_name(),
                         f'ACCOUNT LEDGER — {account_name}' + (f' / {sub_account_name}' if sub_account_name else ''),
                         f'{from_date} to {to_date}')
    row = xl.header_row(ws, row, ['Date', 'JV No', 'Narration', 'Debit', 'Credit', 'Balance'])
    for r in rows:
        if r.get('is_opening'):
            row = xl.total_row(ws, row, f"{r.get('date', '')}   Opening Balance",
                               [], bg='FFF8E1', color='8A6D00')
            m = ws.cell(row=row - 1, column=6, value=float(r.get('balance', 0)))
            m.number_format = xl.NUM_FMT
            from openpyxl.styles import Font as _F, Alignment as _A
            m.font = _F(bold=True, size=10, color='8A6D00')
            m.alignment = _A(horizontal='right')
            continue
        row = xl.data_row(ws, row,
                          [r.get('date', ''), r.get('jv_no', ''), r.get('narration', ''),
                           r.get('dr', 0), r.get('cr', 0), r.get('balance', 0)],
                          num_cols=(4, 5, 6))
    if rows:
        row = xl.total_row(ws, row, 'CLOSING BALANCE',
                           [0] * 4 + [rows[-1].get('balance', 0)],
                           bg='E3F2FD', color=xl.BLUE_DARK, size=11)
        # blank out the zero filler cells in the closing row
        for col in range(2, 6):
            ws.cell(row=row - 1, column=col).value = None
    xl.finish(ws, ncols, first_col_width=14, num_col_width=18)
    ws.column_dimensions['C'].width = 50
    safe_name = account_name.replace(' ', '_').replace('/', '-')[:40]
    return xl.workbook_response(wb, f'Ledger_{safe_name}_{from_date}_{to_date}.xlsx')


@app.route('/ledger_view/export_subaccount_summary')
@login_required
@has_permission('Access_Reports')
def ledger_export_subaccount_summary():
    """CSV export of the Sub Account Summary shown under the ledger — only
    meaningful for the whole account (not one sub-account already), same as
    the on-screen summary."""
    account_name = request.args.get('account_name', '')
    from_date = request.args.get('from_date', date.today().replace(day=1).strftime('%Y-%m-%d'))
    to_date = request.args.get('to_date', date.today().strftime('%Y-%m-%d'))
    if not account_name:
        flash('No account selected', 'warning')
        return redirect(url_for('ledger_view'))

    result = _ledger_data(account_name, from_date, to_date, None)
    if 'error' in result:
        flash(f"Error: {result['error']}", 'danger')
        return redirect(url_for('ledger_view'))
    summary = result.get('sub_summary') or []

    si = io.StringIO()
    cw = csv.writer(si)
    cw.writerow([f'Sub Account Summary: {account_name}'])
    cw.writerow([f'Period: {from_date} to {to_date}'])
    cw.writerow([])
    cw.writerow(['Sub Account', 'Debit', 'Credit', 'Net'])
    for s in summary:
        cw.writerow([s['name'], f"{s['dr']:.2f}", f"{s['cr']:.2f}", f"{s['net']:.2f}"])
    output = make_response(si.getvalue())
    safe_name = account_name.replace(' ', '_').replace('/', '-')[:40]
    output.headers["Content-Disposition"] = f"attachment; filename=SubAccountSummary_{safe_name}_{from_date}.csv"
    output.headers["Content-type"] = "text/csv"
    return output


@app.route('/ledger_view/export_subaccount_summary_xlsx')
@login_required
@has_permission('Access_Reports')
def ledger_export_subaccount_summary_xlsx():
    """Styled Excel export of the Sub Account Summary shown under the ledger."""
    account_name = request.args.get('account_name', '')
    from_date = request.args.get('from_date', date.today().replace(day=1).strftime('%Y-%m-%d'))
    to_date = request.args.get('to_date', date.today().strftime('%Y-%m-%d'))
    if not account_name:
        flash('No account selected', 'warning')
        return redirect(url_for('ledger_view'))
    try:
        import excel_export as xl
    except ImportError:
        flash("Excel export needs the 'openpyxl' package on the server — run: pip install openpyxl", 'warning')
        return redirect(url_for('ledger_view'))

    result = _ledger_data(account_name, from_date, to_date, None)
    if 'error' in result:
        flash(f"Error: {result['error']}", 'danger')
        return redirect(url_for('ledger_view'))
    summary = result.get('sub_summary') or []

    wb, ws = xl.new_workbook('Sub Account Summary')
    ncols = 4
    row = xl.title_block(ws, ncols, _company_display_name(),
                         f'SUB ACCOUNT SUMMARY — {account_name}', f'{from_date} to {to_date}')
    row = xl.header_row(ws, row, ['Sub Account', 'Debit', 'Credit', 'Net'])
    total_dr = total_cr = total_net = 0.0
    for s in summary:
        row = xl.data_row(ws, row, [s['name'], s['dr'], s['cr'], s['net']], num_cols=(2, 3, 4))
        total_dr += s['dr']; total_cr += s['cr']; total_net += s['net']
    if summary:
        row = xl.total_row(ws, row, 'TOTAL', [total_dr, total_cr, total_net], bg='E3F2FD', color=xl.BLUE_DARK)
    xl.finish(ws, ncols, first_col_width=14, num_col_width=18)
    ws.column_dimensions['A'].width = 40
    safe_name = account_name.replace(' ', '_').replace('/', '-')[:40]
    return xl.workbook_response(wb, f'SubAccountSummary_{safe_name}_{from_date}_{to_date}.xlsx')


# ================================================================
# ── CREDIT NOTE & DEBIT NOTE ────────────────────────────────────
# ================================================================
# Credit Note  → issued TO a customer  (reduces AR / sales return)
# Debit Note   → issued TO a supplier  (reduces AP / purchase return)

@app.route('/credit_note', methods=['GET'])
@login_required
@has_permission('Access_Accounting')
def credit_note():
    customers = db.execute_query(
        "SELECT sup_id, supplier_name FROM suppliers WHERE Is_Customer=1 ORDER BY supplier_name")
    ar_accounts = db.execute_query(
        "SELECT account_name FROM new_account_table WHERE account_assets=1 ORDER BY account_name")
    income_accounts = db.execute_query(
        "SELECT account_name FROM new_account_table WHERE account_income=1 ORDER BY account_name")
    today = date.today().strftime('%Y-%m-%d')
    return render_template('credit_note.html',
                           customers=customers, ar_accounts=ar_accounts,
                           income_accounts=income_accounts, today_date=today,
                           last_cn=request.args.get('last_cn'),
                           last_jv=request.args.get('last_jv'),
                           cn_date=request.args.get('cn_date'),
                           cust_name=request.args.get('cust_name'),
                           amount=request.args.get('amount'),
                           narration=request.args.get('narration'))


@app.route('/credit_note/submit', methods=['POST'])
@login_required
@has_permission('Access_Accounting')
def credit_note_submit():
    customer_id   = request.form.get('customer_id')
    cn_date       = request.form.get('cn_date') or str(date.today())
    narration     = request.form.get('narration', '').strip() or 'Credit Note'
    ar_account    = request.form.get('ar_account', 'Account Receivable')
    sales_account = request.form.get('sales_account', 'Sales')
    try:
        amount = float(request.form.get('amount', 0))
    except ValueError:
        amount = 0.0

    if not customer_id or amount <= 0:
        flash('Customer and a positive amount are required.', 'danger')
        return redirect(url_for('credit_note'))

    current_user_pk = get_current_user_pk()
    today = date.today()
    conn = None
    cursor = None

    try:
        # Get customer name
        cust = db.execute_query(
            "SELECT supplier_name FROM suppliers WHERE sup_id=%s", (customer_id,))
        cust_name = cust[0]['supplier_name'] if cust else ''

        conn = db.get_connection()
        cursor = conn.cursor()
        conn.start_transaction()

        # 1. Generate CN number
        cursor.execute("SELECT COUNT(*) FROM jv_numbers WHERE jv_user_code LIKE 'CN-%'")
        cn_seq = (cursor.fetchone()[0] or 0) + 1
        cn_code = f"CN-{cn_seq:05d}"

        # 2. Create JV
        cursor.execute(
            "INSERT INTO jv_numbers (jv_user_code, jv_naration, status) VALUES (%s,%s,1)",
            (cn_code, f'Credit Note: {narration}'))
        jv_no = cursor.lastrowid

        # 3. GL entries
        # DR: Sales / Revenue account (reverse the original sale)
        cursor.execute("""
            INSERT INTO entry_details
              (account_name, enty_values_DR, entry_effective_date, entry_create_date,
               entry_naration, entry_create_user, entry_jv)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        """, (sales_account, amount, cn_date, today, narration, current_user_pk, jv_no))

        # CR: Accounts Receivable (reduces what customer owes)
        cursor.execute("""
            INSERT INTO entry_details
              (account_name, enty_values_CR, entry_effective_date, entry_create_date,
               entry_naration, entry_create_user, entry_jv)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        """, (ar_account, amount, cn_date, today, narration, current_user_pk, jv_no))

        conn.commit()
        flash(f'Credit Note {cn_code} created successfully. JV: {jv_no}', 'success')
        return redirect(url_for('credit_note', last_cn=cn_code, last_jv=jv_no,
                                cn_date=cn_date, cust_name=cust_name,
                                amount=amount, narration=narration))
    except Exception as e:
        if conn: conn.rollback()
        logging.error(f"Credit Note error: {e}")
        flash(f'Error creating credit note: {str(e)}', 'danger')
        return redirect(url_for('credit_note'))
    finally:
        if cursor: cursor.close()
        if conn: conn.close()


@app.route('/debit_note', methods=['GET'])
@login_required
@has_permission('Access_Accounting')
def debit_note():
    suppliers = db.execute_query(
        "SELECT sup_id, supplier_name FROM suppliers WHERE Is_Suplier=1 ORDER BY supplier_name")
    ap_accounts = db.execute_query(
        "SELECT account_name FROM new_account_table WHERE account_liabilities=1 ORDER BY account_name")
    expense_accounts = db.execute_query(
        "SELECT account_name FROM new_account_table WHERE account_expenses=1 ORDER BY account_name")
    today = date.today().strftime('%Y-%m-%d')
    return render_template('debit_note.html',
                           suppliers=suppliers, ap_accounts=ap_accounts,
                           expense_accounts=expense_accounts, today_date=today,
                           last_dn=request.args.get('last_dn'),
                           last_jv=request.args.get('last_jv'),
                           dn_date=request.args.get('dn_date'),
                           sup_name=request.args.get('sup_name'),
                           amount=request.args.get('amount'),
                           narration=request.args.get('narration'))


@app.route('/debit_note/submit', methods=['POST'])
@login_required
@has_permission('Access_Accounting')
def debit_note_submit():
    supplier_id    = request.form.get('supplier_id')
    dn_date        = request.form.get('dn_date') or str(date.today())
    narration      = request.form.get('narration', '').strip() or 'Debit Note'
    ap_account     = request.form.get('ap_account', 'Account Payable')
    expense_account = request.form.get('expense_account', '')
    try:
        amount = float(request.form.get('amount', 0))
    except ValueError:
        amount = 0.0

    if not supplier_id or amount <= 0:
        flash('Supplier and a positive amount are required.', 'danger')
        return redirect(url_for('debit_note'))

    current_user_pk = get_current_user_pk()
    today = date.today()
    conn = None
    cursor = None

    try:
        sup = db.execute_query(
            "SELECT supplier_name FROM suppliers WHERE sup_id=%s", (supplier_id,))
        sup_name = sup[0]['supplier_name'] if sup else ''

        conn = db.get_connection()
        cursor = conn.cursor()
        conn.start_transaction()

        # 1. Generate DN number
        cursor.execute("SELECT COUNT(*) FROM jv_numbers WHERE jv_user_code LIKE 'DN-%'")
        dn_seq = (cursor.fetchone()[0] or 0) + 1
        dn_code = f"DN-{dn_seq:05d}"

        # 2. Create JV
        cursor.execute(
            "INSERT INTO jv_numbers (jv_user_code, jv_naration, status) VALUES (%s,%s,1)",
            (dn_code, f'Debit Note: {narration}'))
        jv_no = cursor.lastrowid

        # 3. GL entries
        # DR: Accounts Payable (reduces what we owe the supplier)
        cursor.execute("""
            INSERT INTO entry_details
              (account_name, enty_values_DR, entry_effective_date, entry_create_date,
               entry_naration, entry_create_user, entry_jv)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        """, (ap_account, amount, dn_date, today, narration, current_user_pk, jv_no))

        # CR: Expense / Purchase Return account
        if not expense_account:
            expense_account = ap_account  # fallback
        cursor.execute("""
            INSERT INTO entry_details
              (account_name, enty_values_CR, entry_effective_date, entry_create_date,
               entry_naration, entry_create_user, entry_jv)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        """, (expense_account, amount, dn_date, today, narration, current_user_pk, jv_no))

        conn.commit()
        flash(f'Debit Note {dn_code} created successfully. JV: {jv_no}', 'success')
        return redirect(url_for('debit_note', last_dn=dn_code, last_jv=jv_no,
                                dn_date=dn_date, sup_name=sup_name,
                                amount=amount, narration=narration))
    except Exception as e:
        if conn: conn.rollback()
        logging.error(f"Debit Note error: {e}")
        flash(f'Error creating debit note: {str(e)}', 'danger')
        return redirect(url_for('debit_note'))
    finally:
        if cursor: cursor.close()
        if conn: conn.close()


# ================================================================
# ── TRANSACTION SEARCH (powerful historical search) ─────────────
# ================================================================
@app.route('/transaction_search')
@login_required
@has_permission('Access_Reports')
def transaction_search():
    """Search ALL historical GL transactions by narration, amount, account,
    JV, voucher, date range and DR/CR — across every transaction type."""
    f = {
        'narration':  request.args.get('narration', '').strip(),
        'account':    request.args.get('account', '').strip(),
        'jv':         request.args.get('jv', '').strip(),
        'amount':     request.args.get('amount', '').strip(),
        'amount_min': request.args.get('amount_min', '').strip(),
        'amount_max': request.args.get('amount_max', '').strip(),
        'date_from':  request.args.get('date_from', '').strip(),
        'date_to':    request.args.get('date_to', '').strip(),
        'dr_cr':      request.args.get('dr_cr', '').strip(),
    }
    download = request.args.get('download')

    filters = ['ed.entry_deleted = 0']
    params = []
    if f['narration']:
        filters.append('ed.entry_naration LIKE %s'); params.append(f"%{f['narration']}%")
    if f['account']:
        filters.append('ed.account_name LIKE %s'); params.append(f"%{f['account']}%")
    if f['jv']:
        filters.append('(ed.entry_jv = %s OR j.jv_user_code LIKE %s)')
        params += [f['jv'], f"%{f['jv']}%"]
    if f['date_from']:
        filters.append('ed.entry_effective_date >= %s'); params.append(f['date_from'])
    if f['date_to']:
        filters.append('ed.entry_effective_date <= %s'); params.append(f['date_to'])
    if f['amount']:
        filters.append('(ed.enty_values_DR = %s OR ed.enty_values_CR = %s)')
        params += [f['amount'], f['amount']]
    if f['amount_min']:
        filters.append('(ed.enty_values_DR >= %s OR ed.enty_values_CR >= %s)')
        params += [f['amount_min'], f['amount_min']]
    if f['amount_max']:
        filters.append('GREATEST(COALESCE(ed.enty_values_DR,0), COALESCE(ed.enty_values_CR,0)) <= %s')
        params.append(f['amount_max'])
    if f['dr_cr'] == 'dr':
        filters.append('ed.enty_values_DR > 0')
    elif f['dr_cr'] == 'cr':
        filters.append('ed.enty_values_CR > 0')

    has_filter = any(v for v in f.values())
    rows = []
    if has_filter:
        where = ' AND '.join(filters)
        query = f"""
            SELECT ed.id, ed.entry_jv AS jv, j.jv_user_code AS code,
                   ed.entry_effective_date AS date, ed.account_name AS account,
                   ed.entry_naration AS narration,
                   COALESCE(ed.enty_values_DR,0) AS dr,
                   COALESCE(ed.enty_values_CR,0) AS cr
            FROM entry_details ed
            LEFT JOIN jv_numbers j ON ed.entry_jv = j.jv_id
            WHERE {where}
            ORDER BY ed.entry_effective_date DESC, ed.id DESC
            LIMIT 1000
        """
        rows = db.execute_query(query, tuple(params)) or []

    # CSV export
    if download == 'csv' and rows:
        si = io.StringIO()
        cw = csv.writer(si)
        cw.writerow(['Date', 'JV', 'Code', 'Account', 'Narration', 'Debit', 'Credit'])
        for r in rows:
            cw.writerow([r['date'], r['jv'], r['code'] or '', r['account'],
                         r['narration'] or '',
                         f"{float(r['dr']):.2f}", f"{float(r['cr']):.2f}"])
        out = make_response(si.getvalue())
        out.headers["Content-Disposition"] = "attachment; filename=transaction_search.csv"
        out.headers["Content-type"] = "text/csv"
        return out

    total_dr = sum(float(r['dr']) for r in rows)
    total_cr = sum(float(r['cr']) for r in rows)

    return render_template('transaction_search.html',
                           rows=rows, f=f, has_filter=has_filter,
                           total_dr=total_dr, total_cr=total_cr,
                           result_count=len(rows))


# --- Trial Balance ---
@app.route('/trial_balance')
@login_required
@has_permission('Access_Reports')
def trial_balance():
    as_at_date = request.args.get('as_at_date', datetime.now().strftime('%Y-%m-%d'))
    download = request.args.get('download')

    query = """
        SELECT
            a.account_name AS AccountName,
            a.account_basment AS Basement,
            SUM(COALESCE(e.enty_values_DR, 0)) AS Debit,
            SUM(COALESCE(e.enty_values_CR, 0)) AS Credit
        FROM new_account_table a
        LEFT JOIN entry_details e ON a.account_name = e.account_name
            AND e.entry_effective_date <= %s
            AND e.entry_deleted = 0
        GROUP BY a.account_name, a.account_basment
        HAVING Debit != 0 OR Credit != 0
        ORDER BY a.account_name ASC
    """

    rows = db.execute_query(query, (as_at_date,))

    # Calculate Totals
    total_dr = sum(float(r['Debit']) for r in rows)
    total_cr = sum(float(r['Credit']) for r in rows)
    diff = abs(total_dr - total_cr)
    status = "Balanced" if diff < 0.01 else f"Out of Balance: {diff:,.2f}"
    is_balanced = diff < 0.01

    if download == 'csv':
        si = io.StringIO()
        cw = csv.writer(si)
        cw.writerow(['Account Name', 'Type', 'Debit', 'Credit'])

        for r in rows:
            cw.writerow([
                r['AccountName'],
                r['Basement'],
                f"{float(r['Debit']):.2f}",
                f"{float(r['Credit']):.2f}"
            ])

        cw.writerow([])
        cw.writerow(['', 'TOTAL', f"{total_dr:.2f}", f"{total_cr:.2f}"])

        output = make_response(si.getvalue())
        output.headers["Content-Disposition"] = f"attachment; filename=Trial_Balance_{as_at_date}.csv"
        output.headers["Content-type"] = "text/csv"
        return output

    # Find unbalanced JVs (where total DR ≠ total CR within the JV)
    unbalanced_jvs = db.execute_query("""
        SELECT
            ed.entry_jv                          AS jv_id,
            j.jv_user_code                       AS jv_code,
            j.jv_naration                        AS narration,
            MIN(ed.entry_effective_date)         AS entry_date,
            SUM(COALESCE(ed.enty_values_DR, 0))  AS total_dr,
            SUM(COALESCE(ed.enty_values_CR, 0))  AS total_cr,
            ABS(SUM(COALESCE(ed.enty_values_DR,0)) - SUM(COALESCE(ed.enty_values_CR,0))) AS diff
        FROM entry_details ed
        LEFT JOIN jv_numbers j ON ed.entry_jv = j.jv_id
        WHERE ed.entry_deleted = 0
          AND ed.entry_effective_date <= %s
        GROUP BY ed.entry_jv, j.jv_user_code, j.jv_naration
        HAVING diff > 0.01
        ORDER BY diff DESC
        LIMIT 50
    """, (as_at_date,)) or []

    return render_template('trial_balance.html',
                           as_at_date=as_at_date,
                           rows=rows,
                           total_dr=total_dr,
                           total_cr=total_cr,
                           status=status,
                           is_balanced=is_balanced,
                           unbalanced_jvs=unbalanced_jvs)

@app.route('/trial_balance/correction', methods=['POST'])
@login_required
@has_permission('Access_Accounting')
def tb_correction():
    """Create a correcting GL entry to balance a specific JV."""
    jv_id        = request.form.get('jv_id')
    account_name = request.form.get('account_name', '').strip()
    narration    = request.form.get('narration', f'TB Correction for JV-{jv_id}').strip()

    if not jv_id or not account_name:
        return jsonify({'success': False, 'error': 'JV ID and account name required'}), 400

    try:
        # Get current imbalance for this JV
        rows = db.execute_query("""
            SELECT SUM(COALESCE(enty_values_DR,0)) AS dr, SUM(COALESCE(enty_values_CR,0)) AS cr,
                   MIN(entry_effective_date) AS edate
            FROM entry_details WHERE entry_jv = %s AND entry_deleted = 0
        """, (jv_id,))
        if not rows:
            return jsonify({'success': False, 'error': 'JV not found'}), 404

        total_dr = float(rows[0]['dr'] or 0)
        total_cr = float(rows[0]['cr'] or 0)
        diff     = total_dr - total_cr
        edate    = rows[0]['edate'] or date.today()

        if abs(diff) < 0.01:
            return jsonify({'success': False, 'error': 'This JV is already balanced'}), 400

        current_user_pk = get_current_user_pk()

        # Insert correcting entry on the same JV
        if diff > 0:
            # More DR than CR → add CR entry
            db.execute_query("""
                INSERT INTO entry_details
                    (account_name, enty_values_CR, entry_effective_date, entry_create_date,
                     entry_naration, entry_create_user, entry_jv)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (account_name, diff, edate, date.today(), narration, current_user_pk, jv_id),
                 commit=True)
        else:
            # More CR than DR → add DR entry
            db.execute_query("""
                INSERT INTO entry_details
                    (account_name, enty_values_DR, entry_effective_date, entry_create_date,
                     entry_naration, entry_create_user, entry_jv)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (account_name, abs(diff), edate, date.today(), narration, current_user_pk, jv_id),
                 commit=True)

        return jsonify({'success': True, 'corrected': round(abs(diff), 2)})
    except Exception as e:
        logging.error(f"TB Correction error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# Party-wise aging matrix with finer 15-day buckets (used by both aging reports)
FINE_AGING_BUCKETS = ['Current', '1-15', '16-30', '31-45', '46-60', '61-90', 'Over 90']


def _fine_aging_bucket(age_days):
    if age_days <= 0:  return 'Current'
    if age_days <= 15: return '1-15'
    if age_days <= 30: return '16-30'
    if age_days <= 45: return '31-45'
    if age_days <= 60: return '46-60'
    if age_days <= 90: return '61-90'
    return 'Over 90'


def _aging_matrix(rows, name_key):
    """Aggregate detailed invoice-level aging rows into one row per party with
    a column per fine bucket. Returns (matrix_rows, column_totals, grand_total)."""
    parties = {}
    for r in rows:
        name = r.get(name_key) or '(Unknown)'
        b = parties.setdefault(name, {bk: 0.0 for bk in FINE_AGING_BUCKETS})
        b[_fine_aging_bucket(r['AgeDays'])] += r['Outstanding']
    matrix_rows = []
    for n in sorted(parties, key=lambda x: str(x).lower()):
        b = parties[n]
        matrix_rows.append({'name': n, 'buckets': b, 'total': sum(b.values())})
    col_totals = {bk: sum(parties[n][bk] for n in parties) for bk in FINE_AGING_BUCKETS}
    grand_total = sum(col_totals.values())
    return matrix_rows, col_totals, grand_total


# --- Supplier Aging Report ---
@app.route('/supplier_aging')
@login_required
@has_permission('Access_Reports')
def supplier_aging():
    selected_supplier = request.args.get('supplier_id')
    download = request.args.get('download')

    # Load Suppliers for Dropdown
    suppliers = db.execute_query("SELECT sup_id, supplier_name FROM suppliers WHERE Is_Suplier = 1 ORDER BY supplier_name")

    # Aging Query
    query = """
        SELECT
            sid.s_i_id as InvoiceId,
            s.sup_id as SupplierId,
            s.supplier_name as SupplierName,
            sid.suppliers_invoice_number as InvoiceNumber,
            sid.suppliers_invoice_date as InvoiceDate,
            sid.suppliers_invoice_final_date as FinalDate,
            sid.suppliers_invoice_total_oustanding as InvoiceTotal,
            COALESCE(sid.suppliers_invoice_total_payment, 0) as PaidAmount,
            (sid.suppliers_invoice_total_oustanding - COALESCE(sid.suppliers_invoice_total_payment, 0)) as Outstanding
        FROM suppliers_invoice_data sid
        INNER JOIN suppliers s ON sid.suppliers_invoice_buinding_supplier = s.sup_id
        WHERE s.Is_Suplier = 1
        AND sid.suppliers_oustanding_delete = 0
        AND (sid.suppliers_invoice_total_oustanding - COALESCE(sid.suppliers_invoice_total_payment, 0)) > 0
    """

    params = []
    if selected_supplier:
        query += " AND s.sup_id = %s"
        params.append(selected_supplier)

    query += " ORDER BY s.supplier_name, sid.suppliers_invoice_final_date"

    rows = db.execute_query(query, tuple(params))

    # Drop invoices fully covered by a not-yet-cleared postdated cheque —
    # they're tracked from the Postdated Cheques page instead, and reappear
    # here automatically the moment that cheque is Cancelled.
    try:
        pdc_rows = db.execute_query(
            "SELECT payload FROM postdated_cheques WHERE pdc_type = 'payment' AND status = 'pending'") or []
    except Exception:
        pdc_rows = []
    covered_ids = set()
    for pdc in pdc_rows:
        try:
            payload = json.loads(pdc['payload'] or '{}')
        except Exception:
            continue
        for p in payload.get('payments', []):
            covered_ids.add(str(p.get('id')))

    if covered_ids:
        rows = [r for r in rows if str(r['InvoiceId']) not in covered_ids]

    # Process Aging
    aging_data = []
    today = date.today()

    buckets = {
        'Current': 0.0,
        '1-30 Days': 0.0,
        '31-60 Days': 0.0,
        '61-90 Days': 0.0,
        'Over 90 Days': 0.0
    }

    for r in rows:
        due_date = r['FinalDate']
        # Calculate days overdue (Today - Due Date)
        # Note: If due_date is None, treat as Current or handle error. Assuming valid date.
        if isinstance(due_date, datetime):
            due_date = due_date.date()

        age_days = (today - due_date).days

        bucket = "Current"
        if age_days > 90: bucket = "Over 90 Days"
        elif age_days > 60: bucket = "61-90 Days"
        elif age_days > 30: bucket = "31-60 Days"
        elif age_days > 0: bucket = "1-30 Days"
        else: bucket = "Current"

        # In C# code:
        # if (ageDays <= 0) return "Current";
        # if (ageDays <= 30) return "1-30 Days";
        # ...
        # This means positive ageDays are overdue.

        r['AgeDays'] = age_days
        r['AgingBucket'] = bucket
        r['Outstanding'] = float(r['Outstanding'])

        buckets[bucket] += r['Outstanding']
        aging_data.append(r)

    total_outstanding = sum(r['Outstanding'] for r in aging_data)
    total_invoices = len(aging_data)
    total_suppliers = len(set(r['SupplierId'] for r in aging_data))

    # Export to CSV
    if download == 'csv':
        si = io.StringIO()
        cw = csv.writer(si)
        cw.writerow(['Supplier ID', 'Supplier Name', 'Invoice No', 'Invoice Date', 'Due Date', 'Invoice Amount', 'Paid Amount', 'Outstanding', 'Age (Days)', 'Aging Bucket'])

        for r in aging_data:
            cw.writerow([
                r['SupplierId'],
                r['SupplierName'],
                r['InvoiceNumber'],
                r['InvoiceDate'],
                r['FinalDate'],
                f"{float(r['InvoiceTotal']):.2f}",
                f"{float(r['PaidAmount']):.2f}",
                f"{r['Outstanding']:.2f}",
                r['AgeDays'],
                r['AgingBucket']
            ])

        output = make_response(si.getvalue())
        output.headers["Content-Disposition"] = f"attachment; filename=Supplier_Aging_Report_{today}.csv"
        output.headers["Content-type"] = "text/csv"
        return output

    matrix_rows, matrix_totals, matrix_grand = _aging_matrix(aging_data, 'SupplierName')
    view = request.args.get('view', 'detail')

    return render_template('supplier_aging.html',
                           suppliers=suppliers,
                           selected_supplier=int(selected_supplier) if selected_supplier else None,
                           rows=aging_data,
                           buckets=buckets,
                           view=view,
                           fine_buckets=FINE_AGING_BUCKETS,
                           matrix_rows=matrix_rows,
                           matrix_totals=matrix_totals,
                           matrix_grand=matrix_grand,
                           summary={
                               'total_outstanding': total_outstanding,
                               'total_invoices': total_invoices,
                               'total_suppliers': total_suppliers,
                               'report_date': today
                           })

# --- Customer Aging Report ---
@app.route('/customer_aging')
@login_required
@has_permission('Access_Reports')
def customer_aging():
    selected_customer = request.args.get('customer_id')
    download = request.args.get('download')

    # Load Customers for Dropdown
    customers = db.execute_query("SELECT sup_id, supplier_name FROM suppliers WHERE Is_Customer = 1 ORDER BY supplier_name")

    # Aging Query
    query = """
        SELECT
            s.sup_id as CustomerId,
            s.supplier_name as CustomerName,
            io.invoice_number as InvoiceNumber,
            io.invoice_date as InvoiceDate,
            io.invoice_final_date as FinalDate,
            io.invoice_total_oustanding as InvoiceTotal,
            COALESCE(io.invoice_oustanding_Patment, 0) as PaidAmount,
            (io.invoice_total_oustanding - COALESCE(io.invoice_oustanding_Patment, 0)) as Outstanding
        FROM Invoice_Oustanding io
        INNER JOIN suppliers s ON io.invoice_buinding_Customer = s.sup_id
        WHERE s.Is_Customer = 1
        AND io.oustanding_delete = 0
        AND (io.invoice_total_oustanding - COALESCE(io.invoice_oustanding_Patment, 0)) > 0
    """

    params = []
    if selected_customer:
        query += " AND s.sup_id = %s"
        params.append(selected_customer)

    query += " ORDER BY s.supplier_name, io.invoice_final_date"

    rows = db.execute_query(query, tuple(params))

    # Process Aging
    aging_data = []
    today = date.today()

    buckets = {
        'Current': 0.0,
        '1-30 Days': 0.0,
        '31-60 Days': 0.0,
        '61-90 Days': 0.0,
        'Over 90 Days': 0.0
    }

    for r in rows:
        due_date = r['FinalDate']
        # Calculate days overdue (Today - Due Date)
        if isinstance(due_date, datetime):
            due_date = due_date.date()

        age_days = (today - due_date).days if due_date else 0

        bucket = "Current"
        if age_days > 90: bucket = "Over 90 Days"
        elif age_days > 60: bucket = "61-90 Days"
        elif age_days > 30: bucket = "31-60 Days"
        elif age_days > 0: bucket = "1-30 Days"
        else: bucket = "Current"

        r['AgeDays'] = age_days
        r['AgingBucket'] = bucket
        r['Outstanding'] = float(r['Outstanding'])

        buckets[bucket] += r['Outstanding']
        aging_data.append(r)

    total_outstanding = sum(r['Outstanding'] for r in aging_data)
    total_invoices = len(aging_data)
    total_customers = len(set(r['CustomerId'] for r in aging_data))

    # Export to CSV
    if download == 'csv':
        si = io.StringIO()
        cw = csv.writer(si)
        cw.writerow(['Customer ID', 'Customer Name', 'Invoice No', 'Invoice Date', 'Due Date', 'Invoice Amount', 'Paid Amount', 'Outstanding', 'Age (Days)', 'Aging Bucket'])

        for r in aging_data:
            cw.writerow([
                r['CustomerId'],
                r['CustomerName'],
                r['InvoiceNumber'],
                r['InvoiceDate'],
                r['FinalDate'],
                f"{float(r['InvoiceTotal']):.2f}",
                f"{float(r['PaidAmount']):.2f}",
                f"{r['Outstanding']:.2f}",
                r['AgeDays'],
                r['AgingBucket']
            ])

        output = make_response(si.getvalue())
        output.headers["Content-Disposition"] = f"attachment; filename=Customer_Aging_Report_{today}.csv"
        output.headers["Content-type"] = "text/csv"
        return output

    matrix_rows, matrix_totals, matrix_grand = _aging_matrix(aging_data, 'CustomerName')
    view = request.args.get('view', 'detail')

    return render_template('customer_aging.html',
                           customers=customers,
                           selected_customer=int(selected_customer) if selected_customer else None,
                           rows=aging_data,
                           buckets=buckets,
                           view=view,
                           fine_buckets=FINE_AGING_BUCKETS,
                           matrix_rows=matrix_rows,
                           matrix_totals=matrix_totals,
                           matrix_grand=matrix_grand,
                           summary={
                               'total_outstanding': total_outstanding,
                               'total_invoices': total_invoices,
                               'total_customers': total_customers,
                               'report_date': today
                           })


# --- Job Profit Analysis Report ---
@app.route('/job_profit_analysis', methods=['GET'])
@login_required
@has_permission('Access_Reports')
def job_profit_analysis():
    job_number = request.args.get('job_number')
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    download = request.args.get('download')

    # Defaults
    if not from_date:
        from_date = (date.today() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not to_date:
        to_date = date.today().strftime('%Y-%m-%d')

    # Load Jobs
    jobs = db.execute_query("""
        SELECT DISTINCT job_number
        FROM jobs_unit
        WHERE (job_finsh = 0 OR job_finsh IS NULL)
        AND (job_cancell = 0 OR job_cancell IS NULL)
        ORDER BY job_number
    """)

    profit_loss_data = []
    ratio_data = []
    summary = {
        'total_income': 0.0,
        'total_expenses': 0.0,
        'net_profit': 0.0,
        'profit_margin': 0.0,
    }

    if job_number and job_number.isdigit():
        query = """
            SELECT
                CASE
                    WHEN nat.account_income = 1 THEN 'Income'
                    WHEN nat.account_expenses = 1 THEN 'Expenses'
                    ELSE 'Other'
                END as AccountType,
                COALESCE(nat.account_name_of_catogory_PL, 'Uncategorized') as CategoryName,
                COALESCE(nat.account_hold_possion_PL, 999) as CategoryOrder,
                nat.account_name as AccountName,
                COALESCE(SUM(
                    CASE
                        WHEN nat.account_income = 1 THEN (ed.enty_values_CR - ed.enty_values_DR)
                        WHEN nat.account_expenses = 1 THEN (ed.enty_values_DR - ed.enty_values_CR)
                        ELSE 0
                    END
                ), 0) as Amount
            FROM entry_details ed
            INNER JOIN new_account_table nat ON ed.account_name = nat.account_name
            WHERE ed.entry_job_number = %s
                AND ed.entry_effective_date BETWEEN %s AND %s
                AND (nat.account_income = 1 OR nat.account_expenses = 1)
                AND ed.entry_deleted = 0
            GROUP BY AccountType, CategoryName, CategoryOrder, nat.account_name
            HAVING Amount != 0
            ORDER BY AccountType DESC, CategoryOrder, Amount DESC
        """

        rows = db.execute_query(query, (job_number, from_date, to_date))

        # Calculate totals
        total_income = sum(r['Amount'] for r in rows if r['AccountType'] == 'Income')
        total_expenses = sum(r['Amount'] for r in rows if r['AccountType'] == 'Expenses')
        net_profit = total_income - total_expenses
        profit_margin = (net_profit / total_income * 100) if total_income > 0 else 0

        summary = {
            'total_income': total_income,
            'total_expenses': total_expenses,
            'net_profit': net_profit,
            'profit_margin': profit_margin
        }

        # Build categorized P&L Tree
        categories = {}
        for r in rows:
            ctype = r['AccountType']
            cname = r['CategoryName']
            key = (ctype, cname)

            if key not in categories:
                categories[key] = {
                    'AccountType': ctype,
                    'CategoryName': cname,
                    'CategoryOrder': r['CategoryOrder'],
                    'CategoryTotal': 0.0,
                    'AccountCount': 0,
                    'Accounts': []
                }

            categories[key]['CategoryTotal'] += r['Amount']
            categories[key]['AccountCount'] += 1

            perc = 0
            if ctype == 'Income' and total_income > 0:
                perc = r['Amount'] / total_income * 100
            elif ctype == 'Expenses' and total_expenses > 0:
                perc = r['Amount'] / total_expenses * 100

            categories[key]['Accounts'].append({
                'AccountName': r['AccountName'],
                'Amount': r['Amount'],
                'Percentage': perc
            })

        # Sort and flatten categories for UI
        sorted_categories = sorted(categories.values(), key=lambda x: (x['AccountType'] == 'Expenses', x['CategoryOrder']))

        for cat in sorted_categories:
            profit_loss_data.append(cat)

        # Ratio Analysis Data
        if total_income > 0:
            ratio_data.append({
                'RatioType': "Profit Margin",
                'Value': profit_margin,
                'Description': "Net Profit as percentage of Total Income"
            })
            ratio_data.append({
                'RatioType': "Expense Ratio",
                'Value': (total_expenses / total_income) * 100,
                'Description': "Total Expenses as percentage of Total Income"
            })
            ratio_data.append({
                'RatioType': "Gross Profit Ratio",
                'Value': profit_margin, # Simplified same as margin in this context
                'Description': "Gross Profit as percentage of Total Income"
            })

        if total_expenses > 0:
            ratio_data.append({
                'RatioType': "Return on Investment",
                'Value': (net_profit / total_expenses) * 100,
                'Description': "Net Profit as percentage of Total Expenses"
            })

        for cat in sorted_categories:
            if cat['AccountType'] == 'Expenses' and cat['CategoryTotal'] > 0:
                if total_income > 0:
                    ratio_data.append({
                        'RatioType': f"{cat['CategoryName']} % of Income",
                        'Value': (cat['CategoryTotal'] / total_income) * 100,
                        'Description': f"{cat['CategoryName']} as percentage of Total Income"
                    })
                if total_expenses > 0:
                    ratio_data.append({
                        'RatioType': f"{cat['CategoryName']} % of Expenses",
                        'Value': (cat['CategoryTotal'] / total_expenses) * 100,
                        'Description': f"{cat['CategoryName']} as percentage of Total Expenses"
                    })
            elif cat['AccountType'] == 'Income' and cat['CategoryTotal'] > 0:
                if total_income > 0:
                    ratio_data.append({
                        'RatioType': f"{cat['CategoryName']} % of Income",
                        'Value': (cat['CategoryTotal'] / total_income) * 100,
                        'Description': f"{cat['CategoryName']} as percentage of Total Income"
                    })

        # CSV Export
        if download == 'csv':
            si = io.StringIO()
            cw = csv.writer(si)
            cw.writerow(['Account Type', 'Category', 'Account Name', 'Amount', 'Percentage'])

            for cat in sorted_categories:
                # Calculate category percentage
                if cat['AccountType'] == 'Income':
                    cat_perc = (cat['CategoryTotal'] / total_income) * 100 if total_income > 0 else 0
                else:
                    cat_perc = (cat['CategoryTotal'] / total_expenses) * 100 if total_expenses > 0 else 0

                cw.writerow([cat['AccountType'], cat['CategoryName'], 'CATEGORY TOTAL', f"{cat['CategoryTotal']:.2f}", f"{cat_perc:.2f}%"])

                for acc in cat['Accounts']:
                    cw.writerow([cat['AccountType'], cat['CategoryName'], acc['AccountName'], f"{acc['Amount']:.2f}", f"{acc['Percentage']:.2f}%"])
                cw.writerow([])

            cw.writerow([])
            cw.writerow(['SUMMARY'])
            cw.writerow(['Total Income', f"{total_income:.2f}"])
            cw.writerow(['Total Expenses', f"{total_expenses:.2f}"])
            cw.writerow(['Net Profit/Loss', f"{net_profit:.2f}"])
            cw.writerow(['Profit Margin', f"{profit_margin:.2f}%"])

            output = make_response(si.getvalue())
            output.headers["Content-Disposition"] = f"attachment; filename=Job_Profit_Analysis_{job_number}_{date.today()}.csv"
            output.headers["Content-type"] = "text/csv"
            return output

    safe_selected_job = None
    if job_number and job_number.isdigit():
        safe_selected_job = int(job_number)

    return render_template('job_profit_analysis.html',
                           jobs=jobs,
                           selected_job=safe_selected_job,
                           from_date=from_date,
                           to_date=to_date,
                           profit_loss_data=profit_loss_data,
                           ratio_data=ratio_data,
                           summary=summary)

# --- Sales Summary Cashier ---
@app.route('/sales_summary_cashier', methods=['GET'])
@login_required
@has_permission('Access_Reports')
def sales_summary_cashier():
    selected_date = request.args.get('date', date.today().strftime('%Y-%m-%d'))
    filter_type = request.args.get('filter', 'current')
    download = request.args.get('download')

    current_cashier_id = get_current_user_id()

    # 1. Fetch Current User PK (Session stores user_pk as 'id' from Login_Table)
    current_user_pk = session.get('user_pk')

    # Fetch cashier name from Pose_Setting_Table if possible, or Login_Table
    # The C# error suggests RecodeUserId in POS_Sales_Invoice_01 is INT (likely Login_Table.id or Pose_Setting_Table.Id)
    # But C# uses `control_variable.POS_User_ID` which implies it might be different from Login User.
    # However, given `current_cashier_id = get_current_user_id()` returns `session['user_id']` (User_Code e.g., 'ADM001'),
    # and the error says "Truncated incorrect DOUBLE value: 'ADM001'", it means `RecodeUserId` column is numeric.
    # We should use `session['user_pk']` (the auto-inc ID) for filtering if RecodeUserId stores the ID.

    res = db.execute_query("SELECT User_Name FROM Pose_Setting_Table WHERE Id = %s", (current_user_pk,))
    cashier_name = res[0]['User_Name'] if res else session.get('username', 'Unknown')

    # 2. Build Query
    query = """
        SELECT
            s.Invoice_No,
            s.ItemCoude,
            s.ItemName,
            s.PaymentMethord,
            s.QuntirySale,
            s.SllingPrice,
            s.ItemPriceComen,
            s.ItemLoyalityPrice,
            s.Total_Value,
            s.AcctionDate,
            s.Revers,
            s.jv,
            s.Sales_with_market_price_Active,
            s.Sales_with_Special_price_Active,
            s.Loyalty_Price_Active,
            s.Loyalty_No,
            s.RecodeUserId,
            lt.User_Name as CashierName
        FROM POS_Sales_Invoice_01 s
        LEFT JOIN Pose_Setting_Table lt ON s.RecodeUserId = lt.Id
        WHERE DATE(s.AcctionDate) = %s
        AND s.Revers = 0
    """
    params = [selected_date]

    if filter_type == 'current':
        # Use user_pk (INT) instead of User_Code (VARCHAR)
        query += " AND s.RecodeUserId = %s"
        params.append(current_user_pk)

    query += " ORDER BY s.AcctionDate DESC, s.jv DESC"

    rows = db.execute_query(query, tuple(params))

    # 3. Process Data
    sales_data = []
    total_cash = 0
    total_card = 0
    total_sales = 0

    for r in rows:
        # Calculate Actual Unit Price Logic from C#
        unit_price = r['SllingPrice']
        loyalty_no = r['Loyalty_No']

        if r['Loyalty_Price_Active'] == 1 and loyalty_no != "-1" and loyalty_no:
            unit_price = r['ItemLoyalityPrice']
        elif r['Sales_with_Special_price_Active'] == 1:
            unit_price = r['ItemPriceComen']
        elif r['Sales_with_market_price_Active'] == 1:
            unit_price = r['SllingPrice']

        r['UnitPrice'] = unit_price

        # Payment Method Text
        pm = r['PaymentMethord']
        r['PaymentMethodText'] = "Cash" if pm == 1 else "Card" if pm == 2 else "Other"

        # Aggregates
        val = float(r['Total_Value'] or 0)
        if pm == 1: total_cash += val
        elif pm == 2: total_card += val
        total_sales += val

        sales_data.append(r)

    transaction_count = len(set(r['jv'] for r in sales_data))
    cashier_count = len(set(r['RecodeUserId'] for r in sales_data))

    # 4. Export CSV
    if download == 'csv':
        si = io.StringIO()
        cw = csv.writer(si)
        cw.writerow(['Invoice No', 'Cashier ID', 'Cashier Name', 'Item Code', 'Item Name', 'Payment Method', 'Quantity', 'Unit Price', 'Total Value', 'Date', 'Time', 'Transaction No'])

        for r in sales_data:
            cw.writerow([
                r['Invoice_No'],
                r['RecodeUserId'],
                r['CashierName'],
                r['ItemCoude'],
                r['ItemName'],
                r['PaymentMethodText'],
                r['QuntirySale'],
                r['UnitPrice'],
                r['Total_Value'],
                r['AcctionDate'].strftime('%Y-%m-%d') if r['AcctionDate'] else '',
                r['AcctionDate'].strftime('%H:%M') if hasattr(r['AcctionDate'], 'strftime') else '00:00',
                r['jv']
            ])

        output = make_response(si.getvalue())
        output.headers["Content-Disposition"] = f"attachment; filename=Sales_Report_{selected_date}.csv"
        output.headers["Content-type"] = "text/csv"
        return output

    return render_template('sales_summary_cashier.html',
                           sales_data=sales_data,
                           selected_date=selected_date,
                           filter_type=filter_type,
                           cashier_name=cashier_name,
                           summary={
                               'cash': total_cash,
                               'card': total_card,
                               'total': total_sales,
                               'transactions': transaction_count,
                               'cashiers': cashier_count,
                               'record_count': len(sales_data)
                           })

# --- POS Reversal ---
@app.route('/pos_reversal')
@login_required
@has_permission('Access_Reversals')
def pos_reversal():
    # Use PK (INT) for RecodeUserId filtering
    current_cashier_id = session.get('user_pk')
    current_date = date.today().strftime('%Y-%m-%d')

    # 1. Fetch Sales History (Grouped by JV)
    # The C# code fetches grouped by JV for current user and date
    query = """
        SELECT jv, SUM(Total_Value) as Total_payment
        FROM POS_Sales_Invoice_01
        WHERE RecodeUserId = %s AND AcctionDate = %s AND Revers = 0
        GROUP BY jv
        ORDER BY jv
    """
    rows = db.execute_query(query, (current_cashier_id, current_date))

    jv_list = []
    for i, r in enumerate(rows):
        jv_list.append({
            'No': i + 1,
            'jv': r['jv'],
            'Total_payment': float(r['Total_payment'] or 0)
        })

    return render_template('pos_reversal.html', jv_list=jv_list)

@app.route('/pos_reversal/get_details')
@login_required
def pos_reversal_details():
    jv = request.args.get('jv')
    if not jv: return {'error': 'No JV provided'}, 400

    # Fetch details for the text box (similar to C# logic)
    query = """
        SELECT
            Invoice_No, ItemName, QuntirySale, Total_Value
        FROM POS_Sales_Invoice_01
        WHERE jv = %s
    """
    rows = db.execute_query(query, (jv,))

    details_text = f"Journal Voucher No: {jv} Reversal Impact\n"
    details_text += "-" * 40 + "\n"

    for r in rows:
        details_text += f"Inv: {r['Invoice_No']} | Item: {r['ItemName']} | Qty: {r['QuntirySale']} | Val: {r['Total_Value']}\n"

    details_text += "-" * 40 + "\n"
    details_text += "Do you want to reverse this entry?"

    return {'details': details_text}

@app.route('/pos_receipt/<int:jv_no>')
@login_required
def pos_receipt(jv_no):
    # Fetch invoice items for this JV
    query = """
        SELECT
            p.Invoice_No, p.ItemName, p.ItemMesurmet,
            p.SllingPrice, p.ItemPriceComen, p.ItemLoyalityPrice,
            p.Sales_with_market_price_Active, p.Sales_with_Special_price_Active, p.Loyalty_Price_Active,
            p.Loyalty_No, p.PaymentMethord, p.QuntirySale, p.Total_Value,
            p.RecodeUserId, p.AcctionDate, p.CashAccountName,
            ps.User_Name AS CashierName, ps.Footer_Message, ps.Top_Message
        FROM POS_Sales_Invoice_01 p
        LEFT JOIN Pose_Setting_Table ps ON p.RecodeUserId = ps.Id
        WHERE p.jv = %s
    """
    rows = db.execute_query(query, (jv_no,))

    if not rows:
        return "Receipt not found", 404

    # Company details
    comp = db.execute_query(
        "SELECT company_name, company_addras_1, company_addras_2, company_addras_3, "
        "company_land_line, vat_registered, company_vate_code FROM company LIMIT 1"
    )
    company = comp[0] if comp else {}

    items = []
    total_sales = 0
    total_savings = 0
    original_total = 0

    invoice_no   = rows[0]['Invoice_No']
    date_val     = rows[0]['AcctionDate']
    cashier      = rows[0].get('CashierName') or str(rows[0]['RecodeUserId'])
    loyalty_no   = rows[0]['Loyalty_No']
    payment_method = rows[0]['PaymentMethord']
    footer_msg   = rows[0].get('Footer_Message') or 'Thank you for your business!'
    top_msg      = rows[0].get('Top_Message') or ''
    is_loyalty   = bool(loyalty_no and str(loyalty_no) not in ('0', '-1', ''))

    for r in rows:
        qty           = float(r['QuntirySale'] or 0)
        market_price  = float(r['SllingPrice'] or 0)
        special_price = float(r['ItemPriceComen'] or 0)
        loyalty_price = float(r['ItemLoyalityPrice'] or 0)
        line_total    = float(r['Total_Value'] or 0)

        # Determine actual charged price from the active price type
        if r['Loyalty_Price_Active'] == 1 and is_loyalty and loyalty_price > 0:
            unit_price = loyalty_price
        elif r['Sales_with_Special_price_Active'] == 1 and special_price > 0:
            unit_price = special_price
        else:
            unit_price = market_price

        unit_charged_price = line_total / qty if qty > 0 else unit_price
        line_original      = market_price * qty
        line_saving        = max(0, line_original - line_total)

        total_sales    += line_total
        total_savings  += line_saving
        original_total += line_original

        items.append({
            'name':         r['ItemName'],
            'qty':          qty,
            'unit':         r['ItemMesurmet'] or '',
            'price':        unit_charged_price,
            'market_price': market_price,
            'total':        line_total,
            'saving':       line_saving
        })

    pm_text = {1: 'CASH', 2: 'CARD'}.get(payment_method, 'OTHER')

    return render_template('pos_receipt.html',
                           items=items,
                           invoice_no=invoice_no,
                           date=date_val,
                           cashier=cashier,
                           is_loyalty=is_loyalty,
                           loyalty_no=loyalty_no,
                           company=company,
                           top_msg=top_msg,
                           footer_msg=footer_msg,
                           totals={
                               'subtotal':  total_sales,
                               'savings':   total_savings,
                               'original':  original_total,
                               'final':     total_sales
                           },
                           payment_method=pm_text)

@app.route('/pos_reversal/process', methods=['POST'])
@login_required
def pos_reversal_process():
    jv = request.form.get('jv')
    if not jv:
        flash('No transaction selected', 'danger')
        return redirect(url_for('pos_reversal'))

    if jv_in_closed_period(jv):
        flash(f'Cannot reverse: this sale is in a closed period (locked through {get_period_lock_date()}).', 'danger')
        return redirect(url_for('pos_reversal'))

    current_user_pk = get_current_user_pk()
    today = date.today()

    conn = None
    cursor = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        conn.start_transaction()

        # ── Guard: check transaction exists and is not already reversed ──
        cursor.execute(
            "SELECT COUNT(*) FROM POS_Sales_Invoice_01 WHERE jv = %s", (jv,)
        )
        total_rows = cursor.fetchone()[0]
        if not total_rows:
            flash(f'Transaction JV-{jv} not found.', 'warning')
            conn.rollback()
            return redirect(url_for('pos_reversal'))

        cursor.execute(
            "SELECT COUNT(*) FROM POS_Sales_Invoice_01 WHERE jv = %s AND Revers = 1", (jv,)
        )
        already_reversed = cursor.fetchone()[0]
        if already_reversed:
            flash(f'Transaction JV-{jv} has already been reversed.', 'warning')
            conn.rollback()
            return redirect(url_for('pos_reversal'))

        # ── Step 1: Create a new reversal JV number ──
        cursor.execute(
            "INSERT INTO jv_numbers (jv_user_code, jv_naration, status) VALUES (%s, %s, 1)",
            (f'REV-{jv}', f'POS Reversal of JV-{jv}')
        )
        rev_jv = cursor.lastrowid

        # ── Step 2: Mark all POS sale lines for this JV as reversed ──
        # (replaces CALL POS_Customer_Delete — that SP used lowercase table name)
        cursor.execute(
            "UPDATE POS_Sales_Invoice_01 SET Revers = 1 WHERE jv = %s", (jv,)
        )

        # ── Step 3: Reverse GL entries — swap DR ↔ CR into the new reversal JV ──
        # (replaces CALL JV_Entry_Revers)
        cursor.execute("""
            INSERT INTO entry_details (
                account_name, enty_values_DR, enty_values_CR,
                entry_effective_date, entry_create_date,
                entry_naration, entry_create_user, entry_jv
            )
            SELECT
                account_name,
                COALESCE(enty_values_CR, 0),
                COALESCE(enty_values_DR, 0),
                %s, %s,
                %s, %s, %s
            FROM entry_details
            WHERE entry_jv = %s
        """, (today, today, f'POS Reversal of JV-{jv}', current_user_pk, rev_jv, jv))

        # ── Step 4: Restore inventory — convert OUT movements back to IN ──
        # (replaces CALL Inventory_Items_Revers_OUT — SP used wrong column casing)
        cursor.execute("""
            INSERT INTO inventory_recod (
                inventoy_name, inventoy_code,
                inventory_recod_action_date,
                inventory_recod_moument_in,
                inventory_recod_movment_out,
                inventory_recod_mesrmet,
                inventory_recod_unit_price,
                inventory_recod_account,
                inventory_recod_user_id,
                JV_No,
                inventory_recod_location
            )
            SELECT
                inventoy_name, inventoy_code,
                %s,
                inventory_recod_movment_out,   -- was OUT, now becomes IN
                0,
                inventory_recod_mesrmet,
                inventory_recod_unit_price,
                'POS Reversal',
                %s,
                %s,
                inventory_recod_location
            FROM inventory_recod
            WHERE JV_No = %s AND inventory_recod_movment_out > 0
        """, (today, current_user_pk, rev_jv, jv))

        conn.commit()
        flash(f'Transaction JV-{jv} reversed successfully. Reversal JV: {rev_jv}', 'success')

    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        flash(f'Error reversing transaction: {str(e)}', 'danger')
    finally:
        if cursor:
            try: cursor.close()
            except Exception: pass
        if conn:
            try: conn.close()
            except Exception: pass

    return redirect(url_for('pos_reversal'))

# --- Bank Payment Reversal ---
@app.route('/bank_payment_reversal')
@login_required
@has_permission('Access_Reversals')
def bank_payment_reversal():
    # Show ALL bank payments (including reversed) for history — is_reversed flag controls UI
    query = """
        SELECT
            b.id,
            b.bank_book_recod_voucher_no as Voucher,
            b.Bank_Payment_Date as Date,
            b.bank_book__accont_name as Account,
            b.bank_book__suplier_name as Supplier,
            b.bank_book__recode_cr as Amount,
            b.jv_numbers_jv_id as JV,
            CASE WHEN (b.bank_book_book_recode_dr IS NOT NULL AND b.bank_book_book_recode_dr > 0)
                 THEN 1 ELSE 0 END as is_reversed
        FROM bank_book_recod b
        WHERE b.bank_book__recode_cr > 0
        ORDER BY b.Bank_Payment_Date DESC, b.id DESC
        LIMIT 100
    """
    rows = db.execute_query(query)
    return render_template('bank_payment_reversal.html', rows=rows)

@app.route('/bank_payment_reversal/get_details')
@login_required
def bank_payment_reversal_details():
    jv_param = request.args.get('jv')
    if not jv_param: return {'error': 'No JV provided'}, 400

    jvs = [j.strip() for j in jv_param.split(',') if j.strip()]
    if not jvs: return {'error': 'No valid JVs provided'}, 400

    format_strings = ','.join(['%s'] * len(jvs))
    jv_tuple = tuple(jvs)

    # Fetch details text
    query = f"""
        SELECT
            jv_numbers_jv_id,
            suppliers_invoice_number as IV_No,
            suppliers_VAT_rate as VAT_Rate,
            cash_book_recode_cr as Paid_Amount
        FROM suppliers_invoice_data
        LEFT JOIN cash_book_recode ON suppliers_invoice_data.s_i_id = cash_book_recode_suplier_oustanding_id
        WHERE jv_numbers_jv_id IN ({format_strings})
    """
    inv_rows = db.execute_query(query, jv_tuple)

    # Pre-fetch Bank Book Record for bank payments specifically (schema variation handling)
    query_bank = f"""
        SELECT jv_numbers_jv_id, bank_book__naration, bank_book__recode_cr
        FROM bank_book_recod
        WHERE jv_numbers_jv_id IN ({format_strings})
    """
    bank_rows_all = db.execute_query(query_bank, jv_tuple)

    # Fetch GL Entries
    gl_query = f"SELECT entry_jv, account_name, enty_values_DR, enty_values_CR FROM entry_details WHERE entry_jv IN ({format_strings})"
    gl_rows = db.execute_query(gl_query, jv_tuple)

    # Fetch Inventory Movements
    inv_rec_query = f"""
        SELECT
            JV_No,
            inventoy_name,
            inventoy_code,
            inventory_recod_moument_in,
            inventory_recod_movment_out,
            inventory_recod_unit_price,
            inventory_recod_suplier_iv_no,
            inventory_recod_location
        FROM inventory_recod WHERE JV_No IN ({format_strings})
    """
    inv_rec_rows = db.execute_query(inv_rec_query, jv_tuple)

    text = ""
    for current_jv in jvs:
        jv_inv_rows = [r for r in inv_rows if str(r.get('jv_numbers_jv_id')) == current_jv]

        if not jv_inv_rows:
           bank_rows = [r for r in bank_rows_all if str(r.get('jv_numbers_jv_id')) == current_jv]
           if len(jvs) > 1 and text:
               text += f"\n"
           text += f"Bank Payment Reversal (JV: {current_jv})\n" + "-"*30 + "\n"
           for r in bank_rows:
               text += f"Narration: {r['bank_book__naration']} | Amount: {r['bank_book__recode_cr']}\n"
        else:
            if len(jvs) > 1 and text:
               text += f"\n"
            text += f"Journal Voucher {current_jv} Impact\n" + "-"*30 + "\n"
            for r in jv_inv_rows:
                text += f"Inv: {r['IV_No']} | VAT: {r['VAT_Rate']}% | Paid: {r['Paid_Amount']}\n"

        jv_gl_rows = [gl for gl in gl_rows if str(gl.get('entry_jv')) == current_jv]
        if jv_gl_rows:
            if len(jvs) > 1:
                text += f"\nGL Entries (JV: {current_jv}):\n"
            else:
                text += "\nGL Entries:\n"
            for gl in jv_gl_rows:
                text += f"Account Name:-  {gl['account_name']} Accout Dr:- {gl['enty_values_DR']} Accout Cr:- {gl['enty_values_CR']}\n"

        jv_inv_rec_rows = [ir for ir in inv_rec_rows if str(ir.get('JV_No')) == current_jv]
        if jv_inv_rec_rows:
            text += "------------------------\n"
            for ir in jv_inv_rec_rows:
                text += (
                    f"Inventory Name:-  {ir['inventoy_name']} "
                    f"Inventory Code:- {ir['inventoy_code']} "
                    f"Item Add:-  {ir['inventory_recod_moument_in']} "
                    f"Item Issue:-  {ir['inventory_recod_movment_out']} "
                    f"Item Price:-  {ir['inventory_recod_unit_price']} "
                    f"IV No:-  {ir['inventory_recod_suplier_iv_no']} "
                    f"Location:-  {ir['inventory_recod_location']}\n"
                )

    text += "\nDo you need to reverse this entry?"

    return {'details': text.strip()}

@app.route('/bank_payment_reversal/process', methods=['POST'])
@login_required
def bank_payment_reversal_process():
    jv = request.form.get('jv')
    if not jv:
        flash('No transaction selected', 'danger')
        return redirect(url_for('bank_payment_reversal'))

    if jv_in_closed_period(jv):
        flash(f'Cannot reverse: this transaction is in a closed period (locked through {get_period_lock_date()}).', 'danger')
        return redirect(url_for('bank_payment_reversal'))

    current_user_pk = get_current_user_pk()
    today = date.today()

    conn = None
    cursor = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        # Guard: check transaction exists
        cursor.execute(
            "SELECT COUNT(*) FROM bank_book_recod WHERE jv_numbers_jv_id = %s AND bank_book__recode_cr > 0",
            (jv,))
        if cursor.fetchone()[0] == 0:
            flash('Transaction not found.', 'danger')
            return redirect(url_for('bank_payment_reversal'))

        # Guard: already reversed (bank uses bank_book_book_recode_dr > 0 as reversal flag)
        cursor.execute(
            "SELECT COUNT(*) FROM bank_book_recod "
            "WHERE jv_numbers_jv_id = %s AND bank_book_book_recode_dr IS NOT NULL AND bank_book_book_recode_dr > 0",
            (jv,))
        if cursor.fetchone()[0] > 0:
            flash('This bank payment has already been reversed.', 'warning')
            return redirect(url_for('bank_payment_reversal'))

        conn.start_transaction()

        # Step 1: Create reversal JV
        cursor.execute(
            "INSERT INTO jv_numbers (jv_user_code, jv_naration, status) VALUES (%s, %s, 1)",
            (f'REV-BANK-{jv}', f'Bank Payment Reversal of JV-{jv}'))
        rev_jv = cursor.lastrowid

        # Step 2: Reverse GL entries (swap DR ↔ CR)
        cursor.execute("""
            INSERT INTO entry_details (
                account_name, enty_values_DR, enty_values_CR,
                entry_effective_date, entry_create_date,
                entry_naration, entry_create_user, entry_jv
            )
            SELECT account_name,
                   COALESCE(enty_values_CR, 0),
                   COALESCE(enty_values_DR, 0),
                   %s, %s, %s, %s, %s
            FROM entry_details WHERE entry_jv = %s
        """, (today, today, f'Bank Payment Reversal of JV-{jv}', current_user_pk, rev_jv, jv))

        # Step 3: Restore supplier outstanding — undo the payment amount per invoice
        # bank_book__suplier_oustanding_id links to suppliers_invoice_data.s_i_id
        cursor.execute("""
            UPDATE suppliers_invoice_data sid
            JOIN bank_book_recod bbr
                ON bbr.bank_book__suplier_oustanding_id = sid.s_i_id
               AND bbr.jv_numbers_jv_id = %s
               AND bbr.bank_book__recode_cr > 0
            SET
                sid.suppliers_invoice_total_payment = GREATEST(0,
                    COALESCE(sid.suppliers_invoice_total_payment, 0) - bbr.bank_book__recode_cr)
            WHERE bbr.jv_numbers_jv_id = %s
        """, (jv, jv))

        # Step 4: Mark bank book record as reversed
        # Set bank_book_book_recode_dr = payment amount — this is the existing reversal flag convention
        cursor.execute("""
            UPDATE bank_book_recod
            SET bank_book_book_recode_dr = bank_book__recode_cr
            WHERE jv_numbers_jv_id = %s
        """, (jv,))

        conn.commit()
        flash(f'Bank Payment (JV: {jv}) reversed successfully. Reversal JV: {rev_jv}', 'success')

    except Exception as e:
        if conn:
            conn.rollback()
        flash(f'Error reversing bank payment: {str(e)}', 'danger')
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

    return redirect(url_for('bank_payment_reversal'))

# --- Cash Payment Reversal (Supplier) ---
@app.route('/cash_payment_reversal')
@login_required
@has_permission('Access_Reversals')
def cash_payment_reversal():
    # Show ALL cash payments (including reversed) for history — is_reversed controls UI
    query = """
        SELECT
            c.chash_book_recod_id as id,
            c.cash_book_recod_voucher_no as Voucher,
            c.Payment_Date as Date,
            c.cash_book_recode_accont_name as Account,
            c.cash_book_recode_suplier_name as Supplier,
            c.cash_book_recode_cr as Amount,
            c.jv_numbers_jv_id as JV,
            CASE WHEN c.User_Revers IS NOT NULL THEN 1 ELSE 0 END as is_reversed
        FROM cash_book_recode c
        WHERE c.cash_book_recode_cr > 0
        AND c.cash_book_recode_suplier_name IS NOT NULL
        ORDER BY c.Payment_Date DESC, c.chash_book_recod_id DESC
        LIMIT 100
    """
    rows = db.execute_query(query)
    return render_template('cash_payment_reversal.html', rows=rows)

@app.route('/cash_payment_reversal/process', methods=['POST'])
@login_required
def cash_payment_reversal_process():
    jv = request.form.get('jv')
    if not jv:
        flash('No transaction selected', 'danger')
        return redirect(url_for('cash_payment_reversal'))

    if jv_in_closed_period(jv):
        flash(f'Cannot reverse: this transaction is in a closed period (locked through {get_period_lock_date()}).', 'danger')
        return redirect(url_for('cash_payment_reversal'))

    current_user_pk = get_current_user_pk()
    today = date.today()

    conn = None
    cursor = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        # Guard: check transaction exists
        cursor.execute(
            "SELECT COUNT(*) FROM cash_book_recode "
            "WHERE jv_numbers_jv_id = %s AND cash_book_recode_cr > 0",
            (jv,))
        if cursor.fetchone()[0] == 0:
            flash('Transaction not found.', 'danger')
            return redirect(url_for('cash_payment_reversal'))

        # Guard: already reversed
        cursor.execute(
            "SELECT COUNT(*) FROM cash_book_recode "
            "WHERE jv_numbers_jv_id = %s AND User_Revers IS NOT NULL",
            (jv,))
        if cursor.fetchone()[0] > 0:
            flash('This cash payment has already been reversed.', 'warning')
            return redirect(url_for('cash_payment_reversal'))

        conn.start_transaction()

        # Step 1: Create reversal JV
        cursor.execute(
            "INSERT INTO jv_numbers (jv_user_code, jv_naration, status) VALUES (%s, %s, 1)",
            (f'REV-CASH-{jv}', f'Cash Payment Reversal of JV-{jv}'))
        rev_jv = cursor.lastrowid

        # Step 2: Reverse GL entries (swap DR ↔ CR)
        cursor.execute("""
            INSERT INTO entry_details (
                account_name, enty_values_DR, enty_values_CR,
                entry_effective_date, entry_create_date,
                entry_naration, entry_create_user, entry_jv
            )
            SELECT account_name,
                   COALESCE(enty_values_CR, 0),
                   COALESCE(enty_values_DR, 0),
                   %s, %s, %s, %s, %s
            FROM entry_details WHERE entry_jv = %s
        """, (today, today, f'Cash Payment Reversal of JV-{jv}', current_user_pk, rev_jv, jv))

        # Step 3: Restore supplier outstanding — undo the payment amount per invoice
        # cash_book_recode_suplier_oustanding_id links to suppliers_invoice_data.s_i_id
        cursor.execute("""
            UPDATE suppliers_invoice_data sid
            JOIN cash_book_recode cbr
                ON cbr.cash_book_recode_suplier_oustanding_id = sid.s_i_id
               AND cbr.jv_numbers_jv_id = %s
               AND cbr.cash_book_recode_cr > 0
            SET
                sid.suppliers_invoice_total_payment = GREATEST(0,
                    COALESCE(sid.suppliers_invoice_total_payment, 0) - cbr.cash_book_recode_cr)
            WHERE cbr.jv_numbers_jv_id = %s
        """, (jv, jv))

        # Step 4: Mark cash book record as reversed
        cursor.execute("""
            UPDATE cash_book_recode SET User_Revers = %s
            WHERE jv_numbers_jv_id = %s
        """, (current_user_pk, jv))

        conn.commit()
        flash(f'Cash Payment (JV: {jv}) reversed successfully. Reversal JV: {rev_jv}', 'success')

    except Exception as e:
        if conn:
            conn.rollback()
        flash(f'Error reversing cash payment: {str(e)}', 'danger')
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

    return redirect(url_for('cash_payment_reversal'))

# --- Direct Payment Reversal (Inventory) ---
@app.route('/direct_payment_reversal')
@login_required
@has_permission('Access_Reversals')
def direct_payment_reversal():
    # Search / filter params
    search   = request.args.get('search', '').strip()
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')
    show_reversed = request.args.get('show_reversed', '0')

    params = []
    filters = ["j.jv_user_code = 'JV FROM DIRECT CASH'", "c.cash_book_recode_cr > 0"]

    if search:
        filters.append("(c.cash_book_recode_accont_name LIKE %s OR c.cash_book_recode_naration LIKE %s)")
        params += [f'%{search}%', f'%{search}%']
    if date_from:
        filters.append("c.Payment_Date >= %s"); params.append(date_from)
    if date_to:
        filters.append("c.Payment_Date <= %s"); params.append(date_to)
    if show_reversed != '1':
        filters.append("c.User_Revers IS NULL")

    where = " AND ".join(filters)
    query = f"""
        SELECT
            MIN(c.chash_book_recod_id)          AS id,
            MIN(c.cash_book_recod_voucher_no)    AS Voucher,
            MIN(c.Payment_Date)                  AS Date,
            MIN(c.cash_book_recode_accont_name)  AS Account,
            MIN(c.cash_book_recode_naration)     AS Narration,
            SUM(c.cash_book_recode_cr)           AS Amount,
            c.jv_numbers_jv_id                   AS JV,
            CASE WHEN MAX(c.User_Revers) IS NOT NULL THEN 1 ELSE 0 END AS is_reversed
        FROM cash_book_recode c
        JOIN jv_numbers j ON c.jv_numbers_jv_id = j.jv_id
        WHERE {where}
        GROUP BY c.jv_numbers_jv_id
        ORDER BY MIN(c.Payment_Date) DESC
        LIMIT 100
    """
    rows = db.execute_query(query, tuple(params)) or []
    return render_template('direct_payment_reversal.html', rows=rows,
                           search=search, date_from=date_from, date_to=date_to,
                           show_reversed=show_reversed)

@app.route('/direct_payment_reversal/process', methods=['POST'])
@login_required
def direct_payment_reversal_process():
    jv = request.form.get('jv')
    if not jv:
        flash('No transaction selected', 'danger')
        return redirect(url_for('direct_payment_reversal'))

    if jv_in_closed_period(jv):
        flash(f'Cannot reverse: this transaction is in a closed period (locked through {get_period_lock_date()}).', 'danger')
        return redirect(url_for('direct_payment_reversal'))

    current_user_pk = get_current_user_pk()
    today = date.today()

    conn = None
    cursor = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        # Guard: check transaction exists
        cursor.execute(
            "SELECT COUNT(*) FROM cash_book_recode "
            "WHERE jv_numbers_jv_id = %s AND cash_book_recode_cr > 0",
            (jv,))
        if cursor.fetchone()[0] == 0:
            flash('Transaction not found.', 'danger')
            return redirect(url_for('direct_payment_reversal'))

        # Guard: already reversed
        cursor.execute(
            "SELECT COUNT(*) FROM cash_book_recode "
            "WHERE jv_numbers_jv_id = %s AND User_Revers IS NOT NULL",
            (jv,))
        if cursor.fetchone()[0] > 0:
            flash('This direct payment has already been reversed.', 'warning')
            return redirect(url_for('direct_payment_reversal'))

        conn.start_transaction()

        # Step 1: Create reversal JV
        cursor.execute(
            "INSERT INTO jv_numbers (jv_user_code, jv_naration, status) VALUES (%s, %s, 1)",
            (f'REV-DIRECT-{jv}', f'Direct Payment Reversal of JV-{jv}'))
        rev_jv = cursor.lastrowid

        # Step 2: Reverse GL entries (swap DR ↔ CR)
        cursor.execute("""
            INSERT INTO entry_details (
                account_name, enty_values_DR, enty_values_CR,
                entry_effective_date, entry_create_date,
                entry_naration, entry_create_user, entry_jv
            )
            SELECT account_name,
                   COALESCE(enty_values_CR, 0),
                   COALESCE(enty_values_DR, 0),
                   %s, %s, %s, %s, %s
            FROM entry_details WHERE entry_jv = %s
        """, (today, today, f'Direct Payment Reversal of JV-{jv}', current_user_pk, rev_jv, jv))

        # Step 3: Reverse inventory IN movements — create matching OUT movements
        # This undoes the stock increase from the original direct purchase
        cursor.execute("""
            INSERT INTO inventory_recod (
                inventoy_name, inventoy_code,
                inventory_recod_action_date,
                inventory_recod_moument_in, inventory_recod_movment_out,
                inventory_recod_mesrmet, inventory_recod_unit_price,
                inventory_recod_account, inventory_recod_user_id,
                JV_No, inventory_recod_location
            )
            SELECT
                inventoy_name, inventoy_code,
                %s,
                0, inventory_recod_moument_in,
                inventory_recod_mesrmet, inventory_recod_unit_price,
                'Direct Payment Reversal',
                %s, %s,
                inventory_recod_location
            FROM inventory_recod
            WHERE JV_No = %s AND inventory_recod_moument_in > 0
        """, (today, current_user_pk, rev_jv, jv))

        # Step 4: Mark cash book record as reversed
        cursor.execute("""
            UPDATE cash_book_recode SET User_Revers = %s
            WHERE jv_numbers_jv_id = %s
        """, (current_user_pk, jv))

        conn.commit()
        flash(f'Direct Payment (JV: {jv}) reversed successfully. Reversal JV: {rev_jv}', 'success')

    except Exception as e:
        if conn:
            conn.rollback()
        flash(f'Error reversing direct payment: {str(e)}', 'danger')
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

    return redirect(url_for('direct_payment_reversal'))

@app.route('/get_reversal_details')
@login_required
def get_reversal_details():
    jv = request.args.get('jv')
    if not jv: return {'error': 'No JV'}, 400

    jv_list = [j.strip() for j in str(jv).split(',') if j.strip()]
    if not jv_list: return {'error': 'No valid JV'}, 400

    format_strings = ','.join(['%s'] * len(jv_list))

    query = f"SELECT entry_jv, account_name, enty_values_DR, enty_values_CR FROM entry_details WHERE entry_jv IN ({format_strings})"
    rows = db.execute_query(query, tuple(jv_list))

    inv_query = f"SELECT JV_No, inventoy_name, inventory_recod_moument_in FROM inventory_recod WHERE JV_No IN ({format_strings})"
    inv_rows = db.execute_query(inv_query, tuple(jv_list))

    entries_by_jv = {}
    items_by_jv = {}

    if rows:
        for r in rows:
            j = str(r.get('entry_jv', ''))
            if j not in entries_by_jv:
                entries_by_jv[j] = []
            entries_by_jv[j].append(r)

    if inv_rows:
        for r in inv_rows:
            j = str(r.get('JV_No', ''))
            if j not in items_by_jv:
                items_by_jv[j] = []
            items_by_jv[j].append(r)

    text = ""
    for j in jv_list:
        text += f"Journal Voucher {j} Details:\n" + "-"*30 + "\n"
        for r in entries_by_jv.get(j, []):
            text += f"{r['account_name']}: DR {r['enty_values_DR']} | CR {r['enty_values_CR']}\n"

        text += "\nInventory Items (if any):\n"
        for r in items_by_jv.get(j, []):
            text += f"{r['inventoy_name']}: Qty {r['inventory_recod_moument_in']}\n"
        text += "\n"

    return {'details': text.strip()}

# --- Reversal Category (GL Journal Reversal) ---
@app.route('/reversal_category')
@login_required
@has_permission('Access_Reversals')
def reversal_category():
    """Reverse general journal entries (JV entries not linked to cash/bank/POS)."""
    query = """
        SELECT
            j.jv_id as JV,
            j.jv_user_code as VoucherCode,
            MIN(e.entry_effective_date) as Date,
            j.jv_naration as Narration,
            SUM(COALESCE(e.enty_values_DR, 0)) as Amount,
            CASE WHEN j.jv_user_code LIKE 'REV-%' THEN 1 ELSE 0 END as is_reversed
        FROM jv_numbers j
        JOIN entry_details e ON j.jv_id = e.entry_jv
        WHERE j.jv_user_code LIKE 'JV %'
          AND j.jv_user_code NOT LIKE 'JV FROM PAYMENT%'
          AND j.jv_user_code NOT LIKE 'JV FROM RECEIPT%'
          AND j.jv_user_code NOT LIKE 'JV FORM SEN INVOICE%'
        GROUP BY j.jv_id, j.jv_user_code, j.jv_naration
        ORDER BY MIN(e.entry_effective_date) DESC
        LIMIT 100
    """
    rows = db.execute_query(query)
    return render_template('reversal_category.html', rows=rows)

@app.route('/reversal_category/process', methods=['POST'])
@login_required
@has_permission('Access_Reversals')
def reversal_category_process():
    jv = request.form.get('jv')
    if not jv:
        flash('No journal entry selected', 'danger')
        return redirect(url_for('reversal_category'))

    if jv_in_closed_period(jv):
        flash(f'Cannot reverse: this entry is in a closed period (locked through {get_period_lock_date()}).', 'danger')
        return redirect(url_for('reversal_category'))

    current_user_pk = get_current_user_pk()
    today = date.today()

    conn = None
    cursor = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        # Guard: transaction exists
        cursor.execute("SELECT COUNT(*) FROM entry_details WHERE entry_jv = %s", (jv,))
        if cursor.fetchone()[0] == 0:
            flash('Journal entry not found.', 'danger')
            return redirect(url_for('reversal_category'))

        # Guard: already reversed (a reversal JV with REV-JV-{jv} exists)
        cursor.execute(
            "SELECT COUNT(*) FROM jv_numbers WHERE jv_user_code = %s",
            (f'REV-JV-{jv}',))
        if cursor.fetchone()[0] > 0:
            flash('This journal entry has already been reversed.', 'warning')
            return redirect(url_for('reversal_category'))

        conn.start_transaction()

        # Create reversal JV
        cursor.execute(
            "INSERT INTO jv_numbers (jv_user_code, jv_naration, status) VALUES (%s, %s, 1)",
            (f'REV-JV-{jv}', f'Journal Reversal of JV-{jv}'))
        rev_jv = cursor.lastrowid

        # Reverse GL entries (swap DR ↔ CR)
        cursor.execute("""
            INSERT INTO entry_details (
                account_name, enty_values_DR, enty_values_CR,
                entry_effective_date, entry_create_date,
                entry_naration, entry_create_user, entry_jv
            )
            SELECT account_name,
                   COALESCE(enty_values_CR, 0),
                   COALESCE(enty_values_DR, 0),
                   %s, %s, %s, %s, %s
            FROM entry_details WHERE entry_jv = %s
        """, (today, today, f'Journal Reversal of JV-{jv}', current_user_pk, rev_jv, jv))

        conn.commit()
        flash(f'Journal Entry (JV: {jv}) reversed successfully. Reversal JV: {rev_jv}', 'success')

    except Exception as e:
        if conn:
            conn.rollback()
        flash(f'Error reversing journal entry: {str(e)}', 'danger')
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

    return redirect(url_for('reversal_category'))

# --- Customer Receipt (Accounts Receivable) ---
@app.route('/customer_receipt')
@login_required
@has_permission('Access_Accounting')
def customer_receipt():
    # Customers are stored in the `suppliers` table (Is_Customer=1).
    # Invoice_Oustanding.invoice_buinding_Customer holds suppliers.sup_id.
    query = """
        SELECT DISTINCT s.sup_id AS id, s.supplier_name AS customer_name
        FROM suppliers s
        JOIN Invoice_Oustanding io ON s.sup_id = io.invoice_buinding_Customer
        WHERE s.Is_Customer = 1
          AND io.oustanding_delete = 0
          AND (io.invoice_total_oustanding - COALESCE(io.invoice_oustanding_Patment, 0)) > 0
        ORDER BY s.supplier_name
    """
    customers = db.execute_query(query)
    cash_accounts = db.execute_query("SELECT cash_book_account_name FROM cash_book")
    bank_accounts = db.execute_query("SELECT bank_bookcol_account_number FROM bank_book")

    base_currency = 'LKR'
    try:
        res = db.execute_query("SELECT company_curency FROM company LIMIT 1")
        if res and res[0].get('company_curency'):
            base_currency = res[0]['company_curency']
    except Exception:
        pass
    try:
        currencies = db.execute_query("SELECT currency_code, currency_name FROM currency_table") or []
    except Exception:
        currencies = []

    return render_template('customer_receipt.html',
                           customers=customers,
                           cash_accounts=cash_accounts,
                           bank_accounts=bank_accounts,
                           today_date=date.today().strftime('%Y-%m-%d'),
                           base_currency=base_currency,
                           currencies=currencies,
                           last_jv=request.args.get('last_jv'))

@app.route('/customer_receipt/get_outstanding')
@login_required
def get_customer_outstanding():
    customer_id = request.args.get('customer_id')
    if not customer_id: return {'error': 'No customer ID'}, 400

    query = """
        SELECT
            Id, invoice_number, invoice_date, invoice_final_date,
            invoice_total_oustanding, invoice_oustanding_Patment,
            (invoice_total_oustanding - COALESCE(invoice_oustanding_Patment, 0)) AS Invoice_Oustanding,
            invoice_JV
        FROM Invoice_Oustanding
        WHERE invoice_buinding_Customer = %s
          AND oustanding_delete = 0
          AND (invoice_total_oustanding - COALESCE(invoice_oustanding_Patment, 0)) > 0
        ORDER BY invoice_date
    """
    rows = db.execute_query(query, (customer_id,))

    # Format for JSON
    data = []
    for r in rows:
        data.append({
            'id': r['Id'],
            'inv_no': r['invoice_number'],
            'jv_no': str(r['invoice_JV'] or ''),
            'date': str(r['invoice_date']),
            'due_date': str(r['invoice_final_date']),
            'total': float(r['invoice_total_oustanding'] or 0),
            'paid': float(r['invoice_oustanding_Patment'] or 0),
            'balance': float(r['Invoice_Oustanding'] or 0)
        })

    return {'invoices': data}

@app.route('/customer_receipt/get_history')
@login_required
def get_customer_receipt_history():
    customer_id = request.args.get('customer_id')
    if not customer_id: return {'error': 'No customer ID'}, 400

    # Get Customer Name (customers live in suppliers table with Is_Customer=1)
    res = db.execute_query("SELECT supplier_name FROM suppliers WHERE sup_id = %s", (customer_id,))
    if not res: return {'error': 'Customer not found'}, 404
    cust_name = res[0]['supplier_name']

    # Fetch History from Cash Book
    # Grouping by JV to show single line per receipt transaction if multiple invoices paid
    # However, C# grid shows individual lines. But for printing, we need to group by JV/Receipt.
    # We will return list of receipt headers (unique by JV/Voucher)

    query = """
        SELECT
            jv_numbers_jv_id as jv_no,
            cash_book_recod_voucher_no as voucher_no,
            Payment_Date as date,
            SUM(cash_book_recode_dr) as amount,
            cash_book_recode_accont_name as account,
            cash_book_recode_naration as narration
        FROM cash_book_recode
        WHERE TRIM(cash_book_recode_suplier_name) = TRIM(%s)
        GROUP BY jv_numbers_jv_id, cash_book_recod_voucher_no, Payment_Date, cash_book_recode_accont_name, cash_book_recode_naration
        ORDER BY Payment_Date DESC
    """
    cash_rows = db.execute_query(query, (cust_name,))

    # Fetch History from Bank Book
    query_bank = """
        SELECT
            jv_numbers_jv_id as jv_no,
            bank_book_recod_voucher_no as voucher_no,
            Bank_Payment_Date as date,
            SUM(bank_book_book_recode_dr) as amount,
            bank_book__accont_name as account,
            bank_book__naration as narration
        FROM bank_book_recod
        WHERE TRIM(bank_book__suplier_name) = TRIM(%s)
        GROUP BY jv_numbers_jv_id, bank_book_recod_voucher_no, Bank_Payment_Date, bank_book__accont_name, bank_book__naration
        ORDER BY Bank_Payment_Date DESC
    """
    bank_rows = db.execute_query(query_bank, (cust_name,))

    history = []

    for r in cash_rows:
        history.append({
            'type': 'Cash',
            'jv_no': r['jv_no'],
            'voucher_no': r['voucher_no'],
            'date': str(r['date']),
            'amount': float(r['amount'] or 0),
            'account': r['account'],
            'narration': r['narration']
        })

    for r in bank_rows:
        history.append({
            'type': 'Bank',
            'jv_no': r['jv_no'],
            'voucher_no': r['voucher_no'],
            'date': str(r['date']),
            'amount': float(r['amount'] or 0),
            'account': r['account'],
            'narration': r['narration']
        })

    # Sort combined history by date desc
    history.sort(key=lambda x: x['date'], reverse=True)

    return {'history': history}

@app.route('/receipt/print/<int:jv_no>')
@login_required
def print_receipt(jv_no):
    # Determine if Cash or Bank based on JV existence in tables

    # Try Cash Book
    cash_res = db.execute_query("""
        SELECT
            c.cash_book_recod_voucher_no as voucher_no,
            c.Payment_Date as date,
            c.cash_book_recode_suplier_name as received_from,
            c.cash_book_recode_accont_name as account,
            c.cash_book_recode_naration as narration,
            SUM(c.cash_book_recode_dr) as amount,
            c.User_Enter as user_id
        FROM cash_book_recode c
        WHERE c.jv_numbers_jv_id = %s
        GROUP BY c.cash_book_recod_voucher_no, c.Payment_Date, c.cash_book_recode_suplier_name,
                 c.cash_book_recode_accont_name, c.cash_book_recode_naration, c.User_Enter
    """, (jv_no,))

    # Try Bank Book
    bank_res = db.execute_query("""
        SELECT
            b.bank_book_recod_voucher_no as voucher_no,
            b.Bank_Payment_Date as date,
            b.bank_book__suplier_name as received_from,
            b.bank_book__accont_name as account,
            b.bank_book__naration as narration,
            SUM(b.bank_book_book_recode_dr) as amount,
            b.Bank_User_Id as user_id
        FROM bank_book_recod b
        WHERE b.jv_numbers_jv_id = %s
        GROUP BY b.bank_book_recod_voucher_no, b.Bank_Payment_Date, b.bank_book__suplier_name,
                 b.bank_book__accont_name, b.bank_book__naration, b.Bank_User_Id
    """, (jv_no,))

    receipt = None
    if cash_res: receipt = cash_res[0]
    elif bank_res: receipt = bank_res[0]

    if not receipt:
        return "Receipt Not Found", 404

    # Get Invoice Details (Invoices settled by this JV)
    # We join with Invoice_Oustanding or check `cash_book_suplier_oustanding_id` links
    # Note: `cash_book_recode` has `cash_book_recode_suplier_oustanding_id` which links to `Invoice_Oustanding.Id`

    invoices = []
    if cash_res:
        inv_query = """
            SELECT io.invoice_number, c.cash_book_recode_dr as amount_paid
            FROM cash_book_recode c
            JOIN Invoice_Oustanding io ON c.cash_book_recode_suplier_oustanding_id = io.Id
            WHERE c.jv_numbers_jv_id = %s
        """
        invoices = db.execute_query(inv_query, (jv_no,))
    elif bank_res:
        inv_query = """
            SELECT io.invoice_number, b.bank_book_book_recode_dr as amount_paid
            FROM bank_book_recod b
            JOIN Invoice_Oustanding io ON b.bank_book__suplier_oustanding_id = io.Id
            WHERE b.jv_numbers_jv_id = %s
        """
        invoices = db.execute_query(inv_query, (jv_no,))

    # Company Info
    company_res = db.execute_query("SELECT * FROM company LIMIT 1")
    company = company_res[0] if company_res else {}

    # Amount in words (Basic implementation or placeholder)
    # Ideally use a library like num2words

    return render_template('receipt_print.html',
                           receipt=receipt,
                           invoices=invoices,
                           company=company,
                           jv_no=jv_no)

def _post_customer_receipt_entries(cursor, current_user, customer_id, account_type, account_name,
                                    payment_date, narration, payments, manual_receipt_no=None,
                                    online_payment_received=False, transaction_code=None,
                                    card_last_digits=None, bank_transfer_confirmed=False,
                                    transfer_id=None, cheque_no=None):
    """Executes the actual GL/cash-or-bank posting for a customer receipt:
    settle outstanding invoices, allocate a receipt number, create the JV +
    GL entries, and record the cash/bank book rows. Shared by the immediate
    (same-day) flow and by clearing a postdated cheque — the caller owns the
    DB transaction (start_transaction / commit / rollback / close)."""
    total_receipt_base = round(sum(p['base'] for p in payments), 2)

    # 1. Update Invoice Outstanding (Settle)
    if payments:
        # Fetch all outstanding balances in a single query
        payment_ids = [p['id'] for p in payments]
        format_strings = ','.join(['%s'] * len(payment_ids))
        cursor.execute(f"SELECT Id, invoice_oustanding_Patment FROM Invoice_Oustanding WHERE Id IN ({format_strings})", tuple(payment_ids))

        # Map fetched balances to IDs
        res = cursor.fetchall()
        current_balances = {str(row[0]): float(row[1]) for row in res}

        # Aggregate BASE amounts per invoice (outstanding is stored in base currency)
        payment_totals = {}
        for p in payments:
            pid = str(p['id'])
            payment_totals[pid] = payment_totals.get(pid, 0.0) + p['base']

        # Prepare batch update data
        update_data = []
        for pid, total_amount in payment_totals.items():
            current_paid = current_balances.get(pid)
            if current_paid is not None:
                new_paid = current_paid + total_amount
                update_data.append((new_paid, pid))

        # Execute batch update
        if update_data:
            cursor.executemany("UPDATE Invoice_Oustanding SET invoice_oustanding_Patment = %s WHERE Id = %s", update_data)

    # 2. Generate Receipt No
    if account_type == 'cash':
        cursor.execute("SELECT MAX(reciept_no) FROM cash_recipt WHERE likn = %s", (account_name,))
        res = cursor.fetchone()
        receipt_no = (res[0] if res and res[0] else 0) + 1
        cursor.execute("INSERT INTO cash_recipt (likn, reciept_no) VALUES (%s, %s)", (account_name, receipt_no))
    else:
        # Bank Receipt logic
        cursor.execute("SELECT MAX(reciept_no) FROM bank_ecipt WHERE link = %s", (account_name,))
        res = cursor.fetchone()
        receipt_no = (res[0] if res and res[0] else 0) + 1
        cursor.execute("INSERT INTO bank_ecipt (link, reciept_no) VALUES (%s, %s)", (account_name, receipt_no))

    # 3. Create JV
    cursor.execute("INSERT INTO jv_numbers (jv_user_code, jv_naration) VALUES (%s, %s)", ('JV FROM RECEIPT', narration))
    jv_no = cursor.lastrowid

    # 4. GL Entries (posted in base currency)
    # Debit Cash/Bank
    cursor.execute("""
        INSERT INTO entry_details (
            account_name, enty_values_DR, entry_effective_date, entry_create_date,
            entry_naration, entry_create_user, entry_jv
        ) VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (account_name, total_receipt_base, payment_date, date.today(), narration, current_user, jv_no))

    # Credit Accounts Receivable
    # Need sub account code for customer if possible, usually stored in `sub_accont_for_new_account`
    # But simple version: just credit AR control account
    cursor.execute("""
        INSERT INTO entry_details (
            account_name, enty_values_CR, entry_effective_date, entry_create_date,
            entry_naration, entry_create_user, entry_jv
        ) VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, ('Account Receivable', total_receipt_base, payment_date, date.today(), narration, current_user, jv_no))

    # 5. Record Transaction (Cash/Bank Book)
    # Using `cash_book_recode` for cash receipts or `bank_book_recod` ??
    # The schema has `cash_book_recode` with `cash_book_recode_dr` (receipt)

    # Get customer name (customers stored in suppliers table with Is_Customer=1)
    cursor.execute("SELECT supplier_name FROM suppliers WHERE sup_id = %s", (customer_id,))
    row = cursor.fetchone()
    cust_name = row[0] if row else str(customer_id)

    if account_type == 'cash':
        # Create a record per invoice payment or summary?
        # `Recipt.xaml.cs` does foreach item in Accont_collections
        for p in payments:
             cursor.execute("""
                INSERT INTO cash_book_recode (
                    cash_book_recode_dr, cash_book_recode_cr, cash_book_recode_accont_name,
                    cash_book_recode_naration, cash_book_recode_suplier_oustanding_id,
                    cash_book_recode_suplier_name, jv_numbers_jv_id,
                    cash_book_recod_voucher_no, User_Enter, Payment_Date
                ) VALUES (%s, 0, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (p['base'], account_name, narration, p['id'], cust_name, jv_no, manual_receipt_no or receipt_no, current_user, payment_date))
    else:
        # Insert into cash_bank_payment_type first (mirroring WPF)
        cursor.execute("""
            INSERT INTO cash_bank_payment_type (
                manua_recipt_number, onlie_payment_recived, online_transaction_code,
                credit_card_no, bank_transfer, bank_transfer_id, bank_cheque, JV
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            manual_receipt_no,
            1 if online_payment_received else 0,
            transaction_code,
            card_last_digits,
            1 if bank_transfer_confirmed else 0,
            transfer_id,
            cheque_no,
            jv_no
        ))

        # Bank Recode
        for p in payments:
            cursor.execute("""
                INSERT INTO bank_book_recod (
                    bank_book_book_recode_dr, bank_book__recode_cr, bank_book__accont_name,
                    bank_book__naration, bank_book__suplier_oustanding_id,
                    bank_book__suplier_name, jv_numbers_jv_id,
                    bank_book_recod_voucher_no, bank_book_chque_no, Bank_User_Id, Bank_Payment_Date
                ) VALUES (%s, 0, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (p['base'], account_name, narration, p['id'], cust_name, jv_no, manual_receipt_no or receipt_no, cheque_no, current_user, payment_date))

    return jv_no, receipt_no, total_receipt_base


@app.route('/customer_receipt/submit', methods=['POST'])
@login_required
def submit_customer_receipt():
    customer_id = request.form.get('customer_id')
    account_type = request.form.get('account_type') # 'cash' or 'bank'

    if account_type == 'cash':
        account_name = request.form.get('cash_account')
    else:
        account_name = request.form.get('bank_account')

    payment_date = request.form.get('payment_date')
    narration = request.form.get('narration')

    # Multi-currency: amounts are entered in this currency; rate converts to base (LKR)
    receipt_currency = (request.form.get('receipt_currency') or '').strip()
    exchange_rate = parse_float(request.form.get('exchange_rate', 1)) or 1.0
    if exchange_rate <= 0:
        exchange_rate = 1.0

    # Optional extended WPF fields
    manual_receipt_no = request.form.get('manual_receipt_no')
    payment_method = request.form.get('payment_method')
    online_payment_received = request.form.get('online_payment_received') == 'on'
    transaction_code = request.form.get('transaction_code')
    card_last_digits = request.form.get('card_last_digits')
    bank_transfer_confirmed = request.form.get('bank_transfer_confirmed') == 'on'
    transfer_id = request.form.get('transfer_id')
    cheque_no = request.form.get('cheque_no')
    post_date = (request.form.get('post_date') or '').strip()

    if not customer_id or not account_name:
        flash('Missing required fields', 'danger')
        return redirect(url_for('customer_receipt'))

    # Get payments
    payments = []
    total_receipt = 0

    # Iterate form keys to find 'pay_{id}'
    for key in request.form:
        if key.startswith('pay_') and request.form[key]:
            inv_id = key.split('_')[1]
            try:
                amount = float(request.form[key])
                if amount > 0:
                    # 'amount' is in the receipt currency; 'base' is the LKR-equivalent posted to the ledger
                    payments.append({'id': inv_id, 'amount': amount, 'base': round(amount * exchange_rate, 2)})
                    total_receipt += amount
            except (ValueError, TypeError):
                pass

    if total_receipt <= 0:
        flash('No payment amount entered', 'warning')
        return redirect(url_for('customer_receipt'))

    total_receipt_base = round(total_receipt * exchange_rate, 2)
    if receipt_currency and exchange_rate != 1.0:
        narration = (narration or '') + f" [{receipt_currency} {total_receipt:,.2f} @ {exchange_rate:.4f}]"

    current_user = get_current_user_id()

    # Postdated cheque: a cheque dated for a future date is held in the PDC
    # register instead of posting to the books now — it only posts (invoice
    # settlement + JV + cash/bank book) when cleared from the Postdated
    # Cheques page, on or after that post date.
    is_postdated = (account_type == 'bank' and payment_method == 'cheque'
                     and bool(cheque_no) and bool(post_date) and post_date > date.today().isoformat())

    if is_postdated:
        try:
            row = db.execute_query("SELECT supplier_name FROM suppliers WHERE sup_id = %s", (customer_id,))
            cust_name = row[0]['supplier_name'] if row else str(customer_id)
            payload = json.dumps({
                'customer_id': customer_id, 'account_type': account_type, 'account_name': account_name,
                'payment_date': payment_date, 'narration': narration, 'payments': payments,
                'manual_receipt_no': manual_receipt_no, 'online_payment_received': online_payment_received,
                'transaction_code': transaction_code, 'card_last_digits': card_last_digits,
                'bank_transfer_confirmed': bank_transfer_confirmed, 'transfer_id': transfer_id,
                'cheque_no': cheque_no,
            })
            db.execute_query("""
                INSERT INTO postdated_cheques (
                    pdc_type, party_name, account_name, cheque_no, post_date, issue_date,
                    amount, narration, payload, status, created_by
                ) VALUES ('receipt', %s, %s, %s, %s, %s, %s, %s, %s, 'pending', %s)
            """, (cust_name, account_name, cheque_no, post_date, date.today(),
                  total_receipt_base, narration, payload, get_current_user_pk()), commit=True)
            flash(f'Postdated cheque #{cheque_no} recorded from {cust_name} — it will post to the books on '
                  f'{post_date}. Manage it from Postdated Cheques.', 'success')
        except Exception as e:
            flash(f'Error recording postdated cheque: {str(e)}', 'danger')
        return redirect(url_for('customer_receipt'))

    conn = None
    cursor = None
    _rcpt_jv = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        conn.start_transaction()

        jv_no, receipt_no, _ = _post_customer_receipt_entries(
            cursor, current_user, customer_id, account_type, account_name, payment_date, narration,
            payments, manual_receipt_no, online_payment_received, transaction_code, card_last_digits,
            bank_transfer_confirmed, transfer_id, cheque_no)

        conn.commit()
        flash(f'Receipt processed successfully. Receipt No: {receipt_no}', 'success')
        _rcpt_jv = jv_no

    except Exception as e:
        if conn:
            conn.rollback()
        flash(f'Error processing receipt: {str(e)}', 'danger')
        logging.error(f"Receipt Error: {e}")
        _rcpt_jv = None
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

    if _rcpt_jv:
        return redirect(url_for('customer_receipt', last_jv=_rcpt_jv))
    return redirect(url_for('customer_receipt'))

@app.route('/customer_receipt/delete', methods=['POST'])
@login_required
@has_permission('Access_Reversals')
def delete_customer_invoice():
    jv_no = request.form.get('jv_no')
    if not jv_no:
        return {'success': False, 'error': 'No JV Number provided'}, 400

    try:
        conn = db.get_connection()
        cursor = conn.cursor(dictionary=True)

        # Check if the invoice can be deleted (No partial payments made)
        cursor.execute("""
            SELECT invoice_total_oustanding, Invoice_Oustanding
            FROM Invoice_Oustanding
            WHERE invoice_JV = %s
        """, (jv_no,))
        inv = cursor.fetchone()

        if not inv:
            return {'success': False, 'error': 'Invoice not found'}, 404

        if inv['invoice_total_oustanding'] != inv['Invoice_Oustanding']:
            return {'success': False, 'error': 'Cannot delete invoice. Payments have already been made.'}, 400

        conn.start_transaction()

        current_user_pk = get_current_user_pk()
        today = date.today()

        # 1. Mark Invoice as deleted
        cursor.execute("UPDATE Invoice_Oustanding SET oustanding_delete = 1 WHERE invoice_JV = %s", (jv_no,))

        # 2. Reverse inventory records — mark IN movements as reversed by creating OUT movements
        cursor.execute("""
            INSERT INTO inventory_recod (
                inventoy_name, inventoy_code,
                inventory_recod_action_date,
                inventory_recod_moument_in, inventory_recod_movment_out,
                inventory_recod_mesrmet, inventory_recod_unit_price,
                inventory_recod_account, inventory_recod_user_id,
                JV_No, inventory_recod_location
            )
            SELECT inventoy_name, inventoy_code, %s,
                   0, inventory_recod_moument_in,
                   inventory_recod_mesrmet, inventory_recod_unit_price,
                   'Invoice Deletion Reversal', %s,
                   %s, inventory_recod_location
            FROM inventory_recod
            WHERE JV_No = %s AND inventory_recod_moument_in > 0
        """, (today, current_user_pk, jv_no, jv_no))

        # 3. Create reversal JV and reverse GL entries
        cursor.execute(
            "INSERT INTO jv_numbers (jv_user_code, jv_naration, status) VALUES (%s, %s, 1)",
            (f'REV-INV-{jv_no}', f'Customer Invoice Deletion Reversal of JV-{jv_no}'))
        rev_jv = cursor.lastrowid

        cursor.execute("""
            INSERT INTO entry_details (
                account_name, enty_values_DR, enty_values_CR,
                entry_effective_date, entry_create_date,
                entry_naration, entry_create_user, entry_jv
            )
            SELECT account_name,
                   COALESCE(enty_values_CR, 0),
                   COALESCE(enty_values_DR, 0),
                   %s, %s, %s, %s, %s
            FROM entry_details WHERE entry_jv = %s
        """, (today, today, f'Customer Invoice Deletion Reversal of JV-{jv_no}', current_user_pk, rev_jv, jv_no))

        conn.commit()
        return {'success': True}

    except Exception as e:
        if conn:
            conn.rollback()
        logging.error(f"Invoice Deletion Error: {e}")
        return {'success': False, 'error': str(e)}, 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/customer_receipt/reverse', methods=['POST'])
@login_required
@has_permission('Access_Reversals')
def reverse_customer_receipt():
    jv_no = request.form.get('jv_no')

    if not jv_no:
        return {'success': False, 'error': 'No JV Number provided'}, 400

    if jv_in_closed_period(jv_no):
        return {'success': False, 'error': f'This receipt is in a closed period (locked through {get_period_lock_date()}).'}, 403

    conn = None
    cursor = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        # Guard: already reversed? Check if a reversal JV for this already exists
        cursor.execute(
            "SELECT COUNT(*) FROM jv_numbers WHERE jv_user_code = %s",
            (f'REV-RCPT-{jv_no}',))
        if cursor.fetchone()[0] > 0:
            return {'success': False, 'error': 'Receipt already reversed'}, 400

        conn.start_transaction()

        current_user_pk = get_current_user_pk()
        today = date.today()

        # Step 1: Create reversal JV and reverse GL entries
        cursor.execute(
            "INSERT INTO jv_numbers (jv_user_code, jv_naration, status) VALUES (%s, %s, 1)",
            (f'REV-RCPT-{jv_no}', f'Customer Receipt Reversal of JV-{jv_no}'))
        rev_jv = cursor.lastrowid

        cursor.execute("""
            INSERT INTO entry_details (
                account_name, enty_values_DR, enty_values_CR,
                entry_effective_date, entry_create_date,
                entry_naration, entry_create_user, entry_jv
            )
            SELECT account_name,
                   COALESCE(enty_values_CR, 0),
                   COALESCE(enty_values_DR, 0),
                   %s, %s, %s, %s, %s
            FROM entry_details WHERE entry_jv = %s
        """, (today, today, f'Customer Receipt Reversal of JV-{jv_no}', current_user_pk, rev_jv, jv_no))

        # Step 2: Restore Invoice_Oustanding payment amounts
        # invoice_oustanding_Patment was increased when receipt was processed; undo that
        cursor.execute("""
            UPDATE Invoice_Oustanding io
            JOIN cash_book_recode cbr ON cbr.cash_book_recode_suplier_oustanding_id = io.Id
                                     AND cbr.jv_numbers_jv_id = %s
                                     AND cbr.cash_book_recode_dr > 0
            SET io.invoice_oustanding_Patment = GREATEST(0,
                    COALESCE(io.invoice_oustanding_Patment, 0) - cbr.cash_book_recode_dr)
            WHERE cbr.jv_numbers_jv_id = %s
        """, (jv_no, jv_no))

        # Also check bank book receipts
        cursor.execute("""
            UPDATE Invoice_Oustanding io
            JOIN bank_book_recod bbr ON bbr.bank_book__suplier_oustanding_id = io.Id
                                    AND bbr.jv_numbers_jv_id = %s
                                    AND bbr.bank_book_book_recode_dr > 0
            SET io.invoice_oustanding_Patment = GREATEST(0,
                    COALESCE(io.invoice_oustanding_Patment, 0) - bbr.bank_book_book_recode_dr)
            WHERE bbr.jv_numbers_jv_id = %s
        """, (jv_no, jv_no))

        conn.commit()
        return {'success': True}

    except Exception as e:
        if conn:
            conn.rollback()
        logging.error(f"Receipt Reversal Error: {e}")
        return {'success': False, 'error': str(e)}, 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

# --- Profit & Loss Report ---

import ast
import operator
import re

# Safe math evaluator
def safe_eval_math(expr):
    allowed_operators = {
        ast.Add: operator.add,
        ast.Sub: operator.sub,
        ast.Mult: operator.mul,
        ast.Div: operator.truediv,
        ast.USub: operator.neg
    }

    def _eval(node):
        if isinstance(node, ast.Constant):
            if isinstance(node.value, (int, float)):
                return node.value
            raise TypeError("Only numbers are allowed")
        elif getattr(ast, 'Num', None) and isinstance(node, getattr(ast, 'Num')):
            return node.n
        elif isinstance(node, ast.BinOp):
            return allowed_operators[type(node.op)](_eval(node.left), _eval(node.right))
        elif isinstance(node, ast.UnaryOp):
            return allowed_operators[type(node.op)](_eval(node.operand))
        else:
            raise TypeError(f"Unsupported mathematical operation: {node}")

    try:
        parsed = ast.parse(expr, mode='eval')
        return _eval(parsed.body)
    except Exception:
        return 0.0


def log_recent_activity(dot_color, text_content):
    try:
        # Check count
        res = db.execute_query("SELECT COUNT(*) as cnt FROM recent_activity")
        count = res[0]['cnt'] if res else 0
        if count >= 10:
            # Overwrite the oldest record
            oldest = db.execute_query("SELECT id FROM recent_activity ORDER BY created_at ASC LIMIT 1")
            if oldest:
                oldest_id = oldest[0]['id']
                db.execute_query("UPDATE recent_activity SET dot_color=%s, text_content=%s, created_at=CURRENT_TIMESTAMP WHERE id=%s", (dot_color, text_content, oldest_id))
                return
        # Otherwise insert new
        db.execute_query("INSERT INTO recent_activity (dot_color, text_content) VALUES (%s, %s)", (dot_color, text_content))
    except Exception as e:
        app.logger.error(f"Error logging recent activity: {e}")

def _safe_eval_expression(expr, context_vars):
    if not expr: return 0.0
    # Replace variable names (alphabetic strings) with their float values
    vars_in_expr = re.findall(r'[A-Za-z]+', expr)
    eval_str = expr
    for var in vars_in_expr:
        val = context_vars.get(var, 0.0)
        # Regex to safely replace whole words only
        eval_str = re.sub(fr'\b{var}\b', str(val), eval_str)

    # Use ast-based safe eval instead of Python's eval()
    return safe_eval_math(eval_str)

# Custom P&L Feature
@app.route('/profit_loss_custom', methods=['GET'])
@login_required
@has_permission('Access_Reports')
def profit_loss_custom():
    formats = db.execute_query("SELECT id, Description FROM New_PL_Format")
    accounts_rows = db.execute_query("SELECT account_name FROM new_account_table WHERE (account_income = 1 OR account_expenses = 1) AND account_active = 1")
    accounts = [row['account_name'] for row in accounts_rows]
    return render_template('profit_loss_custom.html', formats=formats, accounts=accounts)

@app.route('/api/pl_custom/format', methods=['POST'])
@login_required
def pl_custom_create_format():
    data = request.json
    desc = data.get('description')
    if not desc: return {'success': False, 'error': 'Description missing'}, 400
    db.execute_query("INSERT INTO New_PL_Format (Description) VALUES (%s)", (desc,), commit=True)
    return {'success': True}

@app.route('/api/pl_custom/setup/<int:format_id>', methods=['GET', 'POST', 'DELETE'])
@login_required
def pl_custom_setup(format_id):
    if request.method == 'GET':
        rows = db.execute_query("SELECT * FROM PL_Setup WHERE PL_Report_ID = %s ORDER BY LENGTH(PL_LIne_Number), PL_LIne_Number", (str(format_id),))
        return {'success': True, 'rows': rows}

    elif request.method == 'DELETE':
        db.execute_query("DELETE FROM PL_Setup WHERE PL_Report_ID = %s", (str(format_id),), commit=True)
        return {'success': True}

    elif request.method == 'POST':
        data = request.json
        rows = data.get('rows', [])

        # Reject duplicate line numbers up front — two rows with the same line
        # number overwrite each other at report time (wrong values and totals).
        seen = set()
        dupes = []
        for r in rows:
            ln = (r.get('PL_LIne_Number') or '').strip()
            if not ln:
                continue
            if ln in seen:
                dupes.append(ln)
            seen.add(ln)
        if dupes:
            uniq = ', '.join(sorted(set(dupes)))
            return {'success': False,
                    'error': f'Duplicate line number(s): {uniq}. Each row must have a unique line number.'}, 400

        conn = db.get_connection()
        cursor = conn.cursor()
        try:
            conn.start_transaction()
            cursor.execute("DELETE FROM PL_Setup WHERE PL_Report_ID = %s", (str(format_id),))
            for r in rows:
                cursor.execute('''
                    INSERT INTO PL_Setup (
                        PL_Report_ID, PL_LIne_Number, PL_Text_Description, PL_Text_Colom,
                        PL_Calqulation_instraction, PL_Rasior_instraction, PL_Text_Format,
                        PL_Text_line, PL_Text_Size
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                ''', (
                    str(format_id), r.get('PL_LIne_Number'), r.get('PL_Text_Description'), r.get('PL_Text_Colom'),
                    r.get('PL_Calqulation_instraction'), r.get('PL_Rasior_instraction'), r.get('PL_Text_Format'),
                    r.get('PL_Text_line'), r.get('PL_Text_Size')
                ))
            conn.commit()
            return {'success': True}
        except Exception as e:
            conn.rollback()
            return {'success': False, 'error': str(e)}, 500
        finally:
            cursor.close()
            conn.close()

@app.route('/api/pl_custom/generate', methods=['POST'])
@login_required
def pl_custom_generate():
    data = request.json
    format_id = data.get('format_id')
    prev_from = data.get('prev_from')
    prev_to = data.get('prev_to')
    curr_from = data.get('curr_from')
    curr_to = data.get('curr_to')

    rows = db.execute_query("SELECT * FROM PL_Setup WHERE PL_Report_ID = %s ORDER BY LENGTH(PL_LIne_Number), PL_LIne_Number", (str(format_id),))
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    results = []
    prev_vars = {}
    curr_vars = {}

    try:
        def get_balance(account_name, start, end):
            cursor.execute("SELECT account_basment, account_income, account_expenses FROM new_account_table WHERE account_name = %s", (account_name,))
            acc_info = cursor.fetchone()
            if not acc_info: return 0.0
            cursor.execute('''
                SELECT COALESCE(SUM(enty_values_DR), 0) as dr, COALESCE(SUM(enty_values_CR), 0) as cr
                FROM entry_details
                WHERE account_name = %s AND entry_effective_date BETWEEN %s AND %s AND entry_deleted = 0
            ''', (account_name, start, end))
            b = cursor.fetchone()
            dr = float(b['dr'] or 0)
            cr = float(b['cr'] or 0)
            if acc_info['account_expenses'] == 1: return dr - cr
            elif acc_info['account_income'] == 1: return cr - dr
            else:
                if acc_info['account_basment'] == 'DR': return dr - cr
                else: return cr - dr

        # Pass 1: Fetch account balances
        for r in rows:
            line_no = r['PL_LIne_Number']
            account = r['PL_Text_Colom']

            prev_val = 0.0
            curr_val = 0.0

            if account:
                prev_val = get_balance(account, prev_from, prev_to)
                curr_val = get_balance(account, curr_from, curr_to)

            if line_no:
                prev_vars[line_no] = prev_val
                curr_vars[line_no] = curr_val

        # Pass 2: Evaluate calc_instr
        for r in rows:
            line_no = r['PL_LIne_Number']
            calc_instr = r['PL_Calqulation_instraction']

            if calc_instr:
                prev_val = _safe_eval_expression(calc_instr, prev_vars)
                curr_val = _safe_eval_expression(calc_instr, curr_vars)

                if line_no:
                    prev_vars[line_no] = prev_val
                    curr_vars[line_no] = curr_val

        # Pass 3 & 4: Evaluate ratios and format results
        for r in rows:
            line_no = r['PL_LIne_Number']
            desc = r['PL_Text_Description']
            account = r['PL_Text_Colom']
            ratio_instr = r['PL_Rasior_instraction']

            prev_val = prev_vars.get(line_no, 0.0)
            curr_val = curr_vars.get(line_no, 0.0)

            diff_pct = ""
            if curr_val != 0 or prev_val != 0:
                if prev_val == 0:
                    val = ((curr_val - prev_val) / curr_val) * 100 if curr_val != 0 else 0
                else:
                    if curr_val == 0:
                        val = ((curr_val - prev_val) / prev_val) * 100
                    else:
                        val = ((curr_val - prev_val) / curr_val) * 100
                diff_pct = f"{val:.2f}"

            ratio_pct = ""
            if ratio_instr:
                r_val = _safe_eval_expression(ratio_instr, curr_vars)
                ratio_pct = f"{r_val:.2f}"

            results.append({
                'line': line_no,
                'description': desc,
                'account': account,
                'prev_val': f"{prev_val:,.2f}" if prev_val != 0 else "",
                'curr_val': f"{curr_val:,.2f}" if curr_val != 0 else "",
                'diff': diff_pct,
                'ratio': ratio_pct,
                'format': r.get('PL_Text_Format'),
                'line': r.get('PL_Text_line'),
                'size': r.get('PL_Text_Size')
            })

        # Detect duplicate line numbers — two rows sharing a line number overwrite
        # each other's value (one row shows the other's figure and totals go wrong).
        line_map = {}
        for r in rows:
            ln = (r.get('PL_LIne_Number') or '').strip()
            if not ln:
                continue
            label = (r.get('PL_Text_Description') or r.get('PL_Text_Colom') or '(unnamed)').strip()
            line_map.setdefault(ln, []).append(label or '(unnamed)')
        dupes = {ln: labels for ln, labels in line_map.items() if len(labels) > 1}
        warning = None
        if dupes:
            parts = ["line '{}' shared by {}".format(ln, ' & '.join(labels)) for ln, labels in dupes.items()]
            warning = ("Duplicate line numbers detected — these rows overwrite each other so "
                       "one shows the other's value and totals are wrong: " + "; ".join(parts) +
                       ". Give each row a unique line number in the Setup tab.")

        return {'success': True, 'results': results, 'warning': warning}
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {'success': False, 'error': str(e)}, 500
    finally:
        cursor.close()
        conn.close()



@app.route('/balance_sheet_custom', methods=['GET'])
@login_required
@has_permission('Access_Reports')
def balance_sheet_custom():
    formats = db.execute_query("SELECT id, Description FROM New_BS_Format")
    accounts_rows = db.execute_query("SELECT account_name FROM new_account_table WHERE (account_assets = 1 OR account_liabilities = 1 OR account_equity = 1) AND account_active = 1")
    accounts = [row['account_name'] for row in accounts_rows]
    accounts.append("Retained earnings")
    return render_template('balance_sheet_custom.html', formats=formats, accounts=accounts)

@app.route('/api/bs_custom/format', methods=['POST'])
@login_required
def bs_custom_create_format():
    data = request.json
    desc = data.get('description')
    if not desc: return {'success': False, 'error': 'Description missing'}, 400
    db.execute_query("INSERT INTO New_BS_Format (Description) VALUES (%s)", (desc,), commit=True)
    return {'success': True}

@app.route('/api/bs_custom/setup/<int:format_id>', methods=['GET', 'POST', 'DELETE'])
@login_required
def bs_custom_setup(format_id):
    if request.method == 'GET':
        rows = db.execute_query("SELECT * FROM BS_Setup WHERE BS_Report_ID = %s ORDER BY LENGTH(BS_LIne_Number), BS_LIne_Number", (str(format_id),))
        return {'success': True, 'rows': rows}

    elif request.method == 'DELETE':
        db.execute_query("DELETE FROM BS_Setup WHERE BS_Report_ID = %s", (str(format_id),), commit=True)
        return {'success': True}

    elif request.method == 'POST':
        data = request.json
        rows = data.get('rows', [])
        conn = db.get_connection()
        cursor = conn.cursor()
        try:
            conn.start_transaction()
            cursor.execute("DELETE FROM BS_Setup WHERE BS_Report_ID = %s", (str(format_id),))
            for r in rows:
                cursor.execute('''
                    INSERT INTO BS_Setup (
                        BS_Report_ID, BS_LIne_Number, BS_Text_Description, BS_Text_Colom,
                        BS_Col_A, BS_Col_B, BS_Calqulation_instraction, BS_Text_Format, BS_Text_line, BS_Text_Size
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ''', (
                    str(format_id), r.get('BS_LIne_Number'), r.get('BS_Text_Description'), r.get('BS_Text_Colom'),
                    r.get('BS_Col_A', ''), r.get('BS_Col_B', ''), r.get('BS_Calqulation_instraction'),
                    r.get('BS_Text_Format'), r.get('BS_Text_line'), r.get('BS_Text_Size')
                ))
            conn.commit()
            return {'success': True}
        except Exception as e:
            conn.rollback()
            return {'success': False, 'error': str(e)}, 500
        finally:
            cursor.close()
            conn.close()

@app.route('/api/bs_custom/generate', methods=['POST'])
@login_required
def bs_custom_generate():
    data = request.json
    format_id = data.get('format_id')
    as_at_date = data.get('as_at_date')

    rows = db.execute_query("SELECT * FROM BS_Setup WHERE BS_Report_ID = %s ORDER BY LENGTH(BS_LIne_Number), BS_LIne_Number", (str(format_id),))
    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    results = []
    vars_dict = {}

    try:
        # Pass 1: Gather database balances
        for r in rows:
            line_no = r['BS_LIne_Number']
            account = r['BS_Text_Colom']

            amount = 0.0

            if account:
                if account == "Retained earnings":
                    amount = calculate_retained_earnings(cursor, as_at_date)
                else:
                    cursor.execute("SELECT account_basment, account_assets, account_liabilities, account_equity FROM new_account_table WHERE account_name = %s", (account,))
                    acc_info = cursor.fetchone()
                    if acc_info:
                        cursor.execute('''
                            SELECT COALESCE(SUM(enty_values_DR), 0) as dr, COALESCE(SUM(enty_values_CR), 0) as cr
                            FROM entry_details
                            WHERE account_name = %s AND entry_effective_date <= %s AND entry_deleted = 0
                        ''', (account, as_at_date))
                        b = cursor.fetchone()
                        if b:
                            dr = float(b['dr'] or 0)
                            cr = float(b['cr'] or 0)

                            if acc_info['account_assets'] == 1:
                                amount = dr - cr
                            elif acc_info['account_liabilities'] == 1 or acc_info['account_equity'] == 1:
                                amount = cr - dr
                            else:
                                if acc_info['account_basment'] == 'DR':
                                    amount = dr - cr
                                else:
                                    amount = cr - dr

            if line_no:
                vars_dict[line_no] = amount

        # Pass 2: Evaluate Calculations
        for r in rows:
            line_no = r['BS_LIne_Number']
            calc_instr = r['BS_Calqulation_instraction']

            if calc_instr:
                amount = _safe_eval_expression(calc_instr, vars_dict)
                if line_no:
                    vars_dict[line_no] = amount

        # Pass 3: Format Output
        for r in rows:
            line_no = r['BS_LIne_Number']
            desc = r['BS_Text_Description']
            account = r['BS_Text_Colom']

            amount = vars_dict.get(line_no, 0.0)

            results.append({
                'line': line_no,
                'description': desc,
                'account': account,
                'amount': f"{amount:,.2f}" if amount != 0 else "",
                'format': r.get('BS_Text_Format'),
                'line': r.get('BS_Text_line'),
                'size': r.get('BS_Text_Size')
            })

        return {'success': True, 'results': results}
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {'success': False, 'error': str(e)}, 500
    finally:
        cursor.close()
        conn.close()

def calculate_retained_earnings(cursor, as_at_date, customer_sub_account_code=None):
    # Same logic as Balance Sheet endpoint for Retained earnings
    query = '''
        SELECT
            na.account_basment,
            COALESCE(SUM(ed.enty_values_DR), 0) as dr,
            COALESCE(SUM(ed.enty_values_CR), 0) as cr
        FROM new_account_table na
        JOIN entry_details ed ON na.account_name = ed.account_name
        WHERE (na.account_income = 1 OR na.account_expenses = 1)
          AND ed.entry_effective_date <= %s
          AND na.account_active = 1
          AND ed.entry_deleted = 0
    '''
    params = [as_at_date]

    if customer_sub_account_code is not None:
        query += " AND ed.entry_sub_account_code = %s"
        params.append(customer_sub_account_code)

    query += " GROUP BY na.account_basment"

    cursor.execute(query, tuple(params))

    rows = cursor.fetchall()

    total_retained_earnings = 0.0
    for row in rows:
        dr = float(row['dr'])
        cr = float(row['cr'])
        if row['account_basment'] == 'DR':
            total_retained_earnings -= (dr - cr)
        elif row['account_basment'] == 'CR':
            total_retained_earnings += (cr - dr)

    return total_retained_earnings


def _ensure_sub_account_codes():
    """Give every sub-account a unique code (id_sub + 10001). Sub-accounts made
    via the Add Sub Account page were saved with code 0, so postings to them
    couldn't be attributed to a specific sub. Idempotent (a no-op once coded)."""
    try:
        db.execute_query(
            "UPDATE sub_accont_for_new_account SET sub_account_code = id_sub + 10001 "
            "WHERE sub_account_code IS NULL OR sub_account_code = 0", commit=True)
    except Exception:
        pass


@app.route('/profit_loss', methods=['GET', 'POST'])
@login_required
@has_permission('Access_Reports')
def profit_loss():
    from profit_loss_report import ProfitLossReportGenerator

    _ensure_sub_account_codes()
    current_user_pk = get_current_user_pk()

    # Remember each user's last-used Period 1/2/3 ranges: on a fresh page
    # load with no periods submitted, fall back to whatever they last
    # generated instead of always resetting to "this month".
    saved_periods = None
    try:
        pref_row = db.execute_query(
            "SELECT periods_json FROM report_period_prefs WHERE user_id = %s AND report_type = 'PL'",
            (current_user_pk,))
        if pref_row and pref_row[0].get('periods_json'):
            saved_periods = json.loads(pref_row[0]['periods_json'])
    except Exception:
        saved_periods = None

    try:
        generator = ProfitLossReportGenerator(db)
        periods, report_data, default_start, default_end = generator.generate(request, fallback_periods=saved_periods)
    except Exception as e:
        flash(f'Error generating report: {str(e)}', 'danger')
        return redirect(url_for('index'))

    # Save the periods the user just generated with (only on an explicit
    # "Generate Report" submit, not on a GET like the CSV/Excel export links).
    if request.method == 'POST' and periods:
        try:
            db.execute_query("""
                INSERT INTO report_period_prefs (user_id, report_type, periods_json)
                VALUES (%s, 'PL', %s)
                ON DUPLICATE KEY UPDATE periods_json = VALUES(periods_json)
            """, (current_user_pk, json.dumps(periods)), commit=True)
        except Exception:
            pass

    # CSV export — walk the structured report explicitly (categories in statement order)
    if request.args.get('download') == 'csv':
        si = io.StringIO()
        cw = csv.writer(si)
        cw.writerow(['Profit & Loss Report'])
        cw.writerow(['Periods: ' + '; '.join(f"{p['start']} to {p['end']}" for p in periods)])
        cw.writerow([])
        n_periods = max(1, len(periods))
        cw.writerow(['Type', 'Category', 'Account'] + [f'Period {i+1}' for i in range(n_periods)])
        cats = report_data.get('all_categories') or \
               (report_data.get('income_categories', []) + report_data.get('expense_categories', []))
        for cat in cats:
            kind = 'Income' if cat.get('is_income') else 'Expense'
            for acc in cat.get('accounts', []):
                cw.writerow([kind, cat['name'], acc['name']] + [f"{float(v):.2f}" for v in acc['amounts']])
                for sub in acc.get('sub_accounts', []):
                    cw.writerow([kind, cat['name'], '    - ' + sub['name']] + [f"{float(v):.2f}" for v in sub['amounts']])
            cw.writerow([kind, cat['name'], f"{cat['name']} Total"] + [f"{float(v):.2f}" for v in cat.get('total', [])])
            for strow in cat.get('subtotal_rows', []):
                cw.writerow(['Subtotal', '', strow['label']] + [f"{float(v):.2f}" for v in strow['amounts']])
        cw.writerow([])
        cw.writerow(['', '', 'TOTAL INCOME'] + [f"{float(v):.2f}" for v in report_data.get('total_income', [])])
        cw.writerow(['', '', 'TOTAL EXPENSES'] + [f"{float(v):.2f}" for v in report_data.get('total_expense', [])])
        cw.writerow(['', '', 'NET PROFIT / (LOSS)'] + [f"{float(v):.2f}" for v in report_data.get('net_profit', [])])
        output = make_response(si.getvalue())
        fname_start = periods[0]['start'] if periods else default_start
        fname_end = periods[-1]['end'] if periods else default_end
        output.headers["Content-Disposition"] = f"attachment; filename=ProfitLoss_{fname_start}_{fname_end}.csv"
        output.headers["Content-type"] = "text/csv"
        return output

    # Styled Excel export (presentation format)
    if request.args.get('download') == 'xlsx':
        try:
            import excel_export as xl
        except ImportError:
            flash("Excel export needs the 'openpyxl' package on the server — run: pip install openpyxl", 'warning')
            return redirect(url_for('profit_loss'))

        wb, ws = xl.new_workbook('Profit & Loss')
        ncols = 1 + max(1, len(periods))
        sub = ' | '.join(f"{p['start']} to {p['end']}" for p in periods)
        row = xl.title_block(ws, ncols, _company_display_name(), 'PROFIT & LOSS STATEMENT', sub)
        row = xl.header_row(ws, row, ['Description'] +
                            [f"Period {i+1}\n{p['start']} to {p['end']}" for i, p in enumerate(periods)])
        for cat in report_data.get('all_categories') or []:
            if cat.get('is_income'):
                row = xl.section_row(ws, row, ncols, cat['name'], color=xl.GREEN_DARK, bg='EAF6EA')
            else:
                row = xl.section_row(ws, row, ncols, cat['name'], color=xl.RED_DARK, bg='FDECEA')
            for acc in cat['accounts']:
                row = xl.item_row(ws, row, acc['name'], acc['amounts'])
                for sub in acc.get('sub_accounts', []):
                    row = xl.item_row(ws, row, '        ↳ ' + sub['name'], sub['amounts'])
            row = xl.total_row(ws, row, f"{cat['name']} Total", cat.get('total', []))
            for st in cat.get('subtotal_rows', []):
                row = xl.total_row(ws, row, st['label'], st['amounts'], bg='EEF2FF', color=xl.INDIGO)
        row += 1
        row = xl.total_row(ws, row, 'TOTAL INCOME', report_data.get('total_income', []),
                           bg='E8F5E9', color=xl.GREEN_DARK)
        row = xl.total_row(ws, row, 'TOTAL EXPENSES', report_data.get('total_expense', []),
                           bg='FFEBEE', color=xl.RED_DARK)
        row = xl.total_row(ws, row, 'NET PROFIT / (LOSS)', report_data.get('net_profit', []),
                           bg='E3F2FD', color=xl.BLUE_DARK, size=11)
        xl.finish(ws, ncols)
        fs = periods[0]['start'] if periods else default_start
        fe = periods[-1]['end'] if periods else default_end
        return xl.workbook_response(wb, f'ProfitLoss_{fs}_{fe}.xlsx')

    return render_template('profit_loss.html',
                           periods=periods,
                           report_data=report_data,
                           default_start=default_start,
                           default_end=default_end)


# --- Custom subtotal rows for P&L / Balance Sheet (e.g. Gross Profit) ---
def _report_subtotals(report_type):
    """Return {after_category: [labels]} of custom subtotal rows for a report."""
    try:
        rows = db.execute_query(
            "SELECT after_category, label FROM report_subtotals WHERE report_type = %s ORDER BY id",
            (report_type,)) or []
    except Exception:
        return {}
    out = {}
    for r in rows:
        out.setdefault(r['after_category'], []).append(r['label'])
    return out


@app.route('/report_subtotals/add', methods=['POST'])
@login_required
@has_permission('Access_Accounting')
def report_subtotals_add():
    rt = (request.form.get('report_type') or '').strip().upper()
    label = (request.form.get('label') or '').strip()
    after_cat = (request.form.get('after_category') or '').strip()
    nxt = request.form.get('next') or url_for('pl_account_order')
    if rt in ('PL', 'BS') and label and after_cat:
        try:
            db.execute_query(
                "INSERT INTO report_subtotals (report_type, label, after_category) VALUES (%s, %s, %s)",
                (rt, label, after_cat), commit=True)
            flash(f'Subtotal "{label}" added after {after_cat}.', 'success')
        except Exception as e:
            flash(f'Error adding subtotal: {e}', 'danger')
    else:
        flash('A label and a category are required.', 'warning')
    return redirect(nxt)


@app.route('/report_subtotals/delete', methods=['POST'])
@login_required
@has_permission('Access_Accounting')
def report_subtotals_delete():
    sub_id = request.form.get('id')
    nxt = request.form.get('next') or url_for('pl_account_order')
    if sub_id:
        try:
            db.execute_query("DELETE FROM report_subtotals WHERE id = %s", (sub_id,), commit=True)
            flash('Subtotal removed.', 'success')
        except Exception as e:
            flash(f'Error removing subtotal: {e}', 'danger')
    return redirect(nxt)


# --- P&L Account Order (friendly arrangement of accounts within categories) ---
@app.route('/pl_account_order', methods=['GET', 'POST'])
@login_required
@has_permission('Access_Accounting')
def pl_account_order():
    if request.method == 'POST':
        try:
            items = json.loads(request.form.get('order_json') or '[]')
            for it in items:
                name = (it.get('name') or '').strip()
                sort = it.get('sort')
                if not name:
                    continue
                db.execute_query(
                    "UPDATE new_account_table SET account_pl_sort = %s WHERE account_name = %s",
                    (sort, name), commit=True
                )
            flash('P&L account order saved.', 'success')
        except Exception as e:
            flash(f'Error saving order: {e}', 'danger')
        return redirect(url_for('pl_account_order'))

    try:
        rows = db.execute_query("""
            SELECT account_name, account_name_of_catogory_PL AS cat,
                   account_hold_possion_PL AS pos, account_income, account_pl_sort
            FROM new_account_table
            WHERE (account_income = 1 OR account_expenses = 1) AND account_active = 1
            ORDER BY account_hold_possion_PL, COALESCE(account_pl_sort, 9999), account_name
        """) or []
    except Exception:
        rows = db.execute_query("""
            SELECT account_name, account_name_of_catogory_PL AS cat,
                   account_hold_possion_PL AS pos, account_income, NULL AS account_pl_sort
            FROM new_account_table
            WHERE (account_income = 1 OR account_expenses = 1) AND account_active = 1
            ORDER BY account_hold_possion_PL, account_name
        """) or []

    groups = []
    idx = {}
    for r in rows:
        cat = r['cat'] or 'Uncategorized'
        if cat not in idx:
            idx[cat] = len(groups)
            groups.append({'name': cat, 'pos': r['pos'],
                           'has_income': False, 'has_expense': False,
                           'accounts': []})
        g = groups[idx[cat]]
        if r['account_income'] == 1:
            g['has_income'] = True
        else:
            g['has_expense'] = True
        g['accounts'].append(r['account_name'])

    # Label reflects the actual contents: a category can hold both types (MIXED)
    for g in groups:
        if g['has_income'] and g['has_expense']:
            g['type_label'], g['color'] = 'MIXED', '#D67E1A'
        elif g['has_income']:
            g['type_label'], g['color'] = 'INCOME', '#107C10'
        else:
            g['type_label'], g['color'] = 'EXPENSE', '#C42B1C'

    # Holding level comes from the P&L Categories table (authoritative), not the
    # stale per-account copies; categories not in the table go last with no badge
    levels = _category_levels('`p&l_category`')
    for g in groups:
        g['pos'] = levels.get(g['name'])
    groups.sort(key=lambda g: (g['pos'] if g['pos'] is not None else 9999, g['name']))

    try:
        subtotal_list = db.execute_query(
            "SELECT id, label, after_category FROM report_subtotals WHERE report_type='PL' ORDER BY id") or []
    except Exception:
        subtotal_list = []

    return render_template('pl_account_order.html', groups=groups,
                           page_title='P&L Account Order',
                           page_subtitle='Arrange how accounts appear inside each category on the Profit & Loss',
                           back_url='/profit_loss', back_label='Back to P&L',
                           save_url='/pl_account_order',
                           categories_page='P&L Categories',
                           report_type='PL', subtotal_list=subtotal_list)


@app.route('/bs_account_order', methods=['GET', 'POST'])
@login_required
@has_permission('Access_Accounting')
def bs_account_order():
    if request.method == 'POST':
        try:
            items = json.loads(request.form.get('order_json') or '[]')
            for it in items:
                name = (it.get('name') or '').strip()
                sort = it.get('sort')
                if not name:
                    continue
                db.execute_query(
                    "UPDATE new_account_table SET account_bs_sort = %s WHERE account_name = %s",
                    (sort, name), commit=True
                )
            flash('Balance Sheet account order saved.', 'success')
        except Exception as e:
            flash(f'Error saving order: {e}', 'danger')
        return redirect(url_for('bs_account_order'))

    try:
        rows = db.execute_query("""
            SELECT account_name, account_name_of_catogory_Balace_sheet AS cat,
                   account_hold_possion_Balace_Sheet AS pos,
                   account_assets, account_liabilities, account_equity, account_bs_sort
            FROM new_account_table
            WHERE (account_assets = 1 OR account_liabilities = 1 OR account_equity = 1)
              AND account_active = 1
            ORDER BY COALESCE(account_hold_possion_Balace_Sheet, 9999) + 0,
                     COALESCE(account_bs_sort, 9999), account_name
        """) or []
    except Exception:
        rows = db.execute_query("""
            SELECT account_name, account_name_of_catogory_Balace_sheet AS cat,
                   account_hold_possion_Balace_Sheet AS pos,
                   account_assets, account_liabilities, account_equity, NULL AS account_bs_sort
            FROM new_account_table
            WHERE (account_assets = 1 OR account_liabilities = 1 OR account_equity = 1)
              AND account_active = 1
            ORDER BY COALESCE(account_hold_possion_Balace_Sheet, 9999) + 0, account_name
        """) or []

    groups = []
    idx = {}
    for r in rows:
        cat = r['cat'] or 'Uncategorized'
        if cat not in idx:
            idx[cat] = len(groups)
            if r['account_assets'] == 1:
                label, color = 'ASSET', '#107C10'
            elif r['account_liabilities'] == 1:
                label, color = 'LIABILITY', '#C42B1C'
            else:
                label, color = 'EQUITY', '#0d6efd'
            groups.append({'name': cat, 'pos': r['pos'],
                           'type_label': label, 'color': color, 'accounts': []})
        groups[idx[cat]]['accounts'].append(r['account_name'])

    # Authoritative holding level from the Balance Sheet Categories table
    levels = _category_levels('balance_sheet_category')
    for g in groups:
        g['pos'] = levels.get(g['name'])
    groups.sort(key=lambda g: (g['pos'] if g['pos'] is not None else 9999, g['name']))

    try:
        subtotal_list = db.execute_query(
            "SELECT id, label, after_category FROM report_subtotals WHERE report_type='BS' ORDER BY id") or []
    except Exception:
        subtotal_list = []

    return render_template('pl_account_order.html', groups=groups,
                           page_title='Balance Sheet Account Order',
                           page_subtitle='Arrange how accounts appear inside each category on the Balance Sheet',
                           back_url='/balance_sheet', back_label='Back to Balance Sheet',
                           save_url='/bs_account_order',
                           categories_page='Balance Sheet Categories',
                           report_type='BS', subtotal_list=subtotal_list)

@app.route('/api/dashboard/monthly_revenue')
@login_required
def dashboard_monthly_revenue():
    try:
        from datetime import datetime
        from dateutil.relativedelta import relativedelta
        import calendar

        conn = db.get_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500

        cursor = conn.cursor(dictionary=True)

        now = datetime.now()
        months_data = []

        # Calculate for the last 5 months
        for i in range(4, -1, -1):
            d = now - relativedelta(months=i)
            first_day = d.replace(day=1).strftime('%Y-%m-%d')
            last_day = d.replace(day=calendar.monthrange(d.year, d.month)[1]).strftime('%Y-%m-%d')
            label = d.strftime('%b')

            cursor.execute(
                "SELECT SUM(Total_Value) as total FROM POS_Sales_Invoice_01 WHERE DATE(AcctionDate) BETWEEN %s AND %s AND Revers = 0",
                (first_day, last_day)
            )
            row = cursor.fetchone()
            val = float(row['total'] if row and row['total'] is not None else 0)

            # Additional revenue from Invoices (if any)
            cursor.execute(
                "SELECT SUM(invoice_total_oustanding) as total FROM Invoice_Oustanding WHERE DATE(invoice_date) BETWEEN %s AND %s AND oustanding_delete = 0",
                (first_day, last_day)
            )
            inv_row = cursor.fetchone()
            if inv_row and inv_row['total'] is not None:
                val += float(inv_row['total'])

            months_data.append({
                'label': label,
                'val': val,
                'current': i == 0
            })

        # Calculate YTD
        ytd_start = now.replace(month=1, day=1).strftime('%Y-%m-%d')
        ytd_end = now.replace(day=calendar.monthrange(now.year, now.month)[1]).strftime('%Y-%m-%d')

        cursor.execute(
            "SELECT SUM(Total_Value) as total FROM POS_Sales_Invoice_01 WHERE DATE(AcctionDate) BETWEEN %s AND %s AND Revers = 0",
            (ytd_start, ytd_end)
        )
        row = cursor.fetchone()
        ytd_val = float(row['total'] if row and row['total'] is not None else 0)

        cursor.execute(
            "SELECT SUM(invoice_total_oustanding) as total FROM Invoice_Oustanding WHERE DATE(invoice_date) BETWEEN %s AND %s AND oustanding_delete = 0",
            (ytd_start, ytd_end)
        )
        inv_row = cursor.fetchone()
        if inv_row and inv_row['total'] is not None:
            ytd_val += float(inv_row['total'])

        # Calculate stats
        max_val = max(m['val'] for m in months_data) if months_data else 0
        best_month_data = max(months_data, key=lambda x: x['val']) if months_data else None

        total_5_months = sum(m['val'] for m in months_data)
        avg_monthly = total_5_months / 5 if months_data else 0

        # Inject max into month data for frontend pct calc
        # Avoid 0 division in frontend
        calc_max = max_val if max_val > 0 else 1
        for m in months_data:
            m['max'] = calc_max

        return jsonify({
            'success': True,
            'data': months_data,
            'stats': {
                'average': avg_monthly,
                'best_month_label': best_month_data['label'] if best_month_data else '',
                'best_month_val': best_month_data['val'] if best_month_data else 0,
                'ytd_total': ytd_val
            }
        })
    except Exception as e:
        logging.error(f"Error fetching monthly revenue: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/dashboard/recent_activity')
@login_required
def api_recent_activity():
    try:
        res = db.execute_query("SELECT dot_color, text_content, created_at FROM recent_activity ORDER BY created_at DESC LIMIT 10")

        # format time diff
        from datetime import datetime
        now = datetime.now()
        activities = []
        for r in (res or []):
            dt = r['created_at']
            if not dt:
                time_str = "just now"
            else:
                diff = (now - dt).total_seconds()
                if diff < 60:
                    time_str = f"{int(diff)}s ago"
                elif diff < 3600:
                    time_str = f"{int(diff//60)}m ago"
                elif diff < 86400:
                    time_str = f"{int(diff//3600)}h ago"
                else:
                    time_str = f"{int(diff//86400)}d ago"

            activities.append({
                'dot_color': r['dot_color'],
                'text_content': r['text_content'],
                'time_str': time_str
            })

        return jsonify({'success': True, 'data': activities})
    except Exception as e:
        app.logger.error(f"Error fetching recent activity: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/dashboard/kpis')
@login_required
def dashboard_kpis():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        import calendar
        from datetime import datetime
        now = datetime.now()
        start_date = now.replace(day=1).strftime('%Y-%m-%d')
        end_date = now.replace(day=calendar.monthrange(now.year, now.month)[1]).strftime('%Y-%m-%d')

    conn = db.get_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500

    try:
        cursor = conn.cursor(dictionary=True)

        # 1. Total Revenue
        cursor.execute(
            "SELECT SUM(Total_Value) as total FROM POS_Sales_Invoice_01 WHERE AcctionDate BETWEEN %s AND %s AND Revers = 0",
            (start_date, end_date)
        )
        row = cursor.fetchone()
        total_revenue = float(row['total'] or 0)

        # 2. Cash Receipts
        cursor.execute(
            "SELECT SUM(cash_book_recode_dr) as total FROM cash_book_recode WHERE Payment_Date BETWEEN %s AND %s AND cash_book_recode_dr IS NOT NULL",
            (start_date, end_date)
        )
        row = cursor.fetchone()
        cash_rcpt = float(row['total'] or 0)

        cursor.execute(
            "SELECT SUM(bank_book_book_recode_dr) as total FROM bank_book_recod WHERE Bank_Payment_Date BETWEEN %s AND %s AND bank_book_book_recode_dr IS NOT NULL",
            (start_date, end_date)
        )
        row = cursor.fetchone()
        bank_rcpt = float(row['total'] or 0)
        total_receipts = cash_rcpt + bank_rcpt

        # 3. Outstanding Payables (Current Snapshot, independent of date range, or we can filter by final date if required, but snapshot is standard)
        cursor.execute(
            "SELECT SUM(suppliers_invoice_oustanding) as total FROM suppliers_invoice_data WHERE suppliers_oustanding_delete = 0"
        )
        row = cursor.fetchone()
        outstanding_payables = float(row['total'] or 0)

        # 4. Pending Approvals
        cursor.execute(
            "SELECT COUNT(*) as count FROM OP_NO_Table WHERE status = 0 AND Delete_PO = 0"
        )
        row = cursor.fetchone()
        pending_approvals = int(row['count'] or 0)

        return jsonify({
            'success': True,
            'kpis': {
                'total_revenue': total_revenue,
                'cash_receipts': total_receipts,
                'outstanding_payables': outstanding_payables,
                'pending_approvals': pending_approvals
            }
        })

    except Exception as e:
        import logging
        logging.error(f"KPI fetch error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


# --- Dashboard Module Stats ---
@app.route('/api/dashboard/module_stats')
@login_required
def dashboard_module_stats():
    """Returns real-time counts for the Module Stats widget and Pending Approvals list."""
    from datetime import datetime as _dt
    today = _dt.now().date()

    conn = db.get_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'DB connection failed'}), 500

    try:
        cursor = conn.cursor(dictionary=True)

        # 1. Open (unpaid) customer invoices
        try:
            cursor.execute("""
                SELECT COUNT(*) as cnt
                FROM Invoice_Oustanding
                WHERE oustanding_delete = 0
                  AND (invoice_total_oustanding - COALESCE(invoice_oustanding_Patment, 0)) > 0
            """)
            open_invoices = int((cursor.fetchone() or {}).get('cnt', 0))
        except Exception:
            open_invoices = 0

        # 2. Receipts today (cash DR + bank DR rows created today)
        try:
            cursor.execute("""
                SELECT COUNT(*) as cnt FROM cash_book_recode
                WHERE DATE(Payment_Date) = %s AND cash_book_recode_dr > 0
            """, (today,))
            cash_rcpt = int((cursor.fetchone() or {}).get('cnt', 0))

            cursor.execute("""
                SELECT COUNT(*) as cnt FROM bank_book_recod
                WHERE DATE(Bank_Payment_Date) = %s AND bank_book_book_recode_dr > 0
            """, (today,))
            bank_rcpt = int((cursor.fetchone() or {}).get('cnt', 0))
            receipts_today = cash_rcpt + bank_rcpt
        except Exception:
            receipts_today = 0

        # 3. Low stock items (current balance < min_qty, where min_qty > 0)
        try:
            cursor.execute("""
                SELECT COUNT(*) as cnt FROM (
                    SELECT ir.inventoy_name,
                           SUM(COALESCE(ir.inventory_recod_movment_in,0))
                           - SUM(COALESCE(ir.inventory_recod_movment_out,0)) AS balance,
                           MAX(ii.min_qty) AS min_qty
                    FROM inventory_recod ir
                    JOIN inventoy_items ii ON ir.inventoy_name = ii.inventoy_name
                    WHERE ii.min_qty > 0 AND ii.active = 1
                    GROUP BY ir.inventoy_name
                    HAVING balance < MAX(ii.min_qty)
                ) t
            """)
            low_stock = int((cursor.fetchone() or {}).get('cnt', 0))
        except Exception:
            low_stock = 0

        # 4. GRN / Purchase Orders pending approval
        try:
            cursor.execute("""
                SELECT COUNT(*) as cnt FROM OP_NO_Table
                WHERE status = 0 AND Delete_PO = 0
            """)
            grn_pending = int((cursor.fetchone() or {}).get('cnt', 0))
        except Exception:
            grn_pending = 0

        # 5. POS sales today (transaction count)
        try:
            cursor.execute("""
                SELECT COUNT(DISTINCT Invoice_No) as cnt
                FROM POS_Sales_Invoice_01
                WHERE DATE(AcctionDate) = %s AND Revers = 0
            """, (today,))
            pos_today = int((cursor.fetchone() or {}).get('cnt', 0))
        except Exception:
            pos_today = 0

        # 6. Pending approvals — top 5 real items
        pending_items = []
        try:
            cursor.execute("""
                SELECT id, OP_NO_Other as ref_no, Create_Date as date,
                       Sup_Name as party, 'Purchase Order' as type,
                       (SELECT COALESCE(SUM(QTY*Unit_price),0)
                        FROM PO_Recode_Details
                        WHERE Link_OP_NO_Table = OP_NO_Table.id) as amount,
                       'po' as source
                FROM OP_NO_Table
                WHERE status = 0 AND Delete_PO = 0
                ORDER BY Create_Date DESC LIMIT 5
            """)
            for r in (cursor.fetchall() or []):
                pending_items.append({
                    'ref_no': r.get('ref_no') or f"PO-{r['id']}",
                    'party': r.get('party') or 'Purchase Order',
                    'type': 'Purchase Order',
                    'amount': float(r.get('amount') or 0),
                    'date': str(r.get('date') or ''),
                    'icon': 'fa-shopping-cart',
                    'color': '#7B3F9E'
                })
        except Exception:
            pass

        try:
            cursor.execute("""
                SELECT j.jv_id as id, j.jv_user_code as ref_no,
                       MIN(e.entry_effective_date) as date,
                       j.jv_naration as party,
                       SUM(COALESCE(e.enty_values_DR,0)) as amount
                FROM jv_numbers j
                LEFT JOIN entry_details e ON j.jv_id = e.entry_jv
                WHERE j.status = 0
                GROUP BY j.jv_id, j.jv_user_code, j.jv_naration
                ORDER BY j.jv_id DESC LIMIT 5
            """)
            for r in (cursor.fetchall() or []):
                pending_items.append({
                    'ref_no': f"JV-{r['id']}",
                    'party': r.get('party') or 'Journal Entry',
                    'type': 'Journal / Payment',
                    'amount': float(r.get('amount') or 0),
                    'date': str(r.get('date') or ''),
                    'icon': 'fa-book',
                    'color': '#0078D4'
                })
        except Exception:
            pass

        # Sort by amount desc, keep top 5
        pending_items.sort(key=lambda x: x['amount'], reverse=True)
        pending_items = pending_items[:5]

        return jsonify({
            'success': True,
            'stats': {
                'open_invoices': open_invoices,
                'receipts_today': receipts_today,
                'low_stock': low_stock,
                'grn_pending': grn_pending,
                'pos_today': pos_today
            },
            'pending_approvals': pending_items
        })

    except Exception as e:
        logging.error(f"Module stats error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        try:
            cursor.close()
            conn.close()
        except Exception:
            pass


# --- Dashboard VAT Export ---
@app.route('/dashboard_export_vat', methods=['GET'])
@login_required
def dashboard_export_vat():
    import io
    import csv
    from datetime import datetime
    import calendar
    from vat_helper import VATReportGenerator
    from flask import make_response

    export_type = request.args.get('type')
    if export_type not in ['input', 'output']:
        flash("Invalid export type.", "danger")
        return redirect(url_for('index'))

    now = datetime.now()
    first_day = now.replace(day=1).strftime('%Y-%m-%d')
    last_day = now.replace(day=calendar.monthrange(now.year, now.month)[1]).strftime('%Y-%m-%d')

    generator = VATReportGenerator(db, first_day, last_day)
    if not generator.check_vat_registered():
        flash("Company is not VAT Registered.", "warning")
        return redirect(url_for('index'))

    si = io.StringIO()
    cw = csv.writer(si)

    if export_type == 'output':
        data = generator.generate_schedule_01()
        cw.writerow(['Date', 'Invoice No', 'Purchaser', 'TIN', 'Total Value', 'VAT Rate', 'VAT Amount'])
        for r in data['rows']:
            cw.writerow([r['date'], r['invoice_no'], r['purchaser'], r['tin'], r['total'], r['rate'], r['vat_amount']])
        filename = f"Output_VAT_{now.strftime('%Y_%m')}.csv"
    else:
        data = generator.generate_schedule_02()
        cw.writerow(['Date', 'Invoice No', 'Supplier', 'TIN', 'Total Value', 'VAT Rate', 'VAT Amount', 'Disallowed VAT'])
        for r in data['rows']:
            cw.writerow([r['date'], r['invoice_no'], r['supplier'], r['tin'], r['total'], r['rate'], r['vat_amount'], r['disallowed_vat']])
        filename = f"Input_VAT_{now.strftime('%Y_%m')}.csv"

    output = make_response(si.getvalue())
    output.headers["Content-Disposition"] = f"attachment; filename={filename}"
    output.headers["Content-type"] = "text/csv"
    return output

# --- VAT Report (Sri Lanka Schedule 01 & 02) ---
@app.route('/vat_report', methods=['GET'])
@login_required
@has_permission('Access_Reports')
def vat_report():
    from vat_helper import VATReportGenerator

    from_date = request.args.get('from_date', date.today().replace(day=1).strftime('%Y-%m-%d'))
    to_date = request.args.get('to_date', date.today().strftime('%Y-%m-%d'))

    try:
        generator = VATReportGenerator(db, from_date, to_date)

        if not generator.check_vat_registered():
            flash("Company is not VAT Registered. Please enable VAT in Company Profile to view reports.", "warning")
            return render_template('vat_report.html', vat_enabled=False)

        report_data = generator.generate()
        return render_template('vat_report.html', **report_data)
    except Exception as e:
        import traceback
        traceback.print_exc()
        from markupsafe import escape
        return (f"<div style='font-family:sans-serif;max-width:840px;margin:40px auto;padding:24px;"
                f"border:1px solid #e2e8f0;border-radius:12px'>"
                f"<h3 style='color:#c0392b;margin-top:0'>VAT Report — could not load</h3>"
                f"<pre style='background:#f8f9fa;padding:14px;border-radius:8px;white-space:pre-wrap'>"
                f"{escape(type(e).__name__)}: {escape(str(e))}</pre>"
                f"<p style='color:#605E5C;font-size:.85rem'>If this names a missing table/column, that "
                f"object isn't set up in this company's database, or a table-name case differs on the server.</p>"
                f"<p><a href='/'>← Home</a></p></div>"), 500


# --- Cash Handover ---
@app.route('/cash_handover', methods=['GET', 'POST'])
@login_required
def cash_handover():
    if request.method == 'POST':
        # Retrieve form data
        amount = request.form.get('amount')
        notes = request.form.get('notes', '')
        handover_to = request.form.get('handover_to')

        if not amount or float(amount) <= 0:
            flash('Please enter a valid amount.', 'danger')
            return redirect(url_for('cash_handover'))

        try:
            # Create table if it doesn't exist
            try:
                with db.get_connection() as conn:
                    with conn.cursor(dictionary=True) as cursor:
                        conn.start_transaction()
                        cursor.execute("SHOW TABLES LIKE 'cash_handover_logs'")
                        if not cursor.fetchone():
                            cursor.execute('''
                                CREATE TABLE cash_handover_logs (
                                    id INT AUTO_INCREMENT PRIMARY KEY,
                                    user_id INT NOT NULL,
                                    handover_to VARCHAR(100),
                                    amount DECIMAL(15,2) NOT NULL,
                                    notes TEXT,
                                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                                )
                            ''')

                        # Insert the log
                        cursor.execute('''
                            INSERT INTO cash_handover_logs (user_id, handover_to, amount, notes)
                            VALUES (%s, %s, %s, %s)
                        ''', (get_current_user_id(), handover_to, float(amount), notes))

                        conn.commit()

            except Exception as inner_e:
                if conn: conn.rollback()
                raise inner_e

            flash('Cash handover recorded successfully.', 'success')
            return redirect(url_for('index'))

        except Exception as e:
            flash(f'Error recording cash handover: {str(e)}', 'danger')
            return redirect(url_for('cash_handover'))

    # GET request
    users = db.execute_query("SELECT id, username FROM Login_Table")
    return render_template('cash_handover.html', users=users)


# --- Cashier Day Sales Summary ---
@app.route('/cashier_day_sales')
@login_required
@has_permission('Access_POS')
def cashier_day_sales():
    user_id = get_current_user_id()
    today = datetime.now().strftime('%Y-%m-%d')

    # Simple summary query: total sales by this user today
    try:
        sales_summary = db.execute_query('''
            SELECT
                COUNT(*) as total_invoices,
                SUM(AcctionValue) as total_sales
            FROM POS_Sales_Invoice_01
            WHERE DATE(AcctionDate) = %s AND user_id = %s
        ''', (today, user_id))

        if not sales_summary or not sales_summary[0]:
            summary = {'total_invoices': 0, 'total_sales': 0.00}
        else:
            summary = {
                'total_invoices': sales_summary[0].get('total_invoices', 0) or 0,
                'total_sales': float(sales_summary[0].get('total_sales', 0.00) or 0.00)
            }

    except Exception as e:
        print(f"Error fetching day sales: {e}")
        summary = {'total_invoices': 0, 'total_sales': 0.00}

    return render_template('cashier_day_sales.html', summary=summary, date=today)


# --- POS Settings ---
@app.route('/pos_settings', methods=['GET', 'POST'])
@login_required
@has_permission('Access_POS')
def pos_settings():
    if request.method == 'POST':
        user_id = request.form.get('user_id')

        # General Settings
        location = request.form.get('location')
        card_ac = request.form.get('card_ac')
        cash_ac = request.form.get('cash_ac')

        # Pricing Settings
        market_price = 1 if request.form.get('market_price') else 0
        special_price = 1 if request.form.get('special_price') else 0
        loyalty_price = 1 if request.form.get('loyalty_price') else 0
        vat_enable = 1 if request.form.get('vat_enable') else 0

        # Messages
        footer = request.form.get('footer_msg')
        top = request.form.get('top_msg')

        # Image Handling

        img_data = None
        if 'receipt_logo' in request.files:
            file = request.files['receipt_logo']
            if file.filename != '':
                img_data = base64.b64encode(file.read()).decode('utf-8')

        try:
            if not user_id:
                flash('User ID missing', 'danger')
                return redirect(url_for('pos_settings'))

            # Update Query
            query = """
                UPDATE Pose_Setting_Table SET
                    Select_Inventry_Location=%s, Card_Control_AC=%s, Cash_Account=%s,
                    Sales_with_market_price=%s, Sales_with_Special_price=%s, Loyalty_Price=%s, VAT_Enable=%s,
                    Footer_Message=%s, Top_Message=%s
            """
            params = [location, card_ac, cash_ac, market_price, special_price, loyalty_price, vat_enable, footer, top]

            if img_data:
                query += ", Image=%s"
                params.append(img_data)

            query += " WHERE Id=%s"
            params.append(user_id)

            db.execute_query(query, tuple(params), commit=True)
            flash('POS Settings updated successfully.', 'success')

        except Exception as e:
            flash(f'Error updating settings: {str(e)}', 'danger')

        return redirect(url_for('pos_settings', user_id=user_id))

    # GET
    # 1. Fetch all POS users for dropdown
    pos_users = db.execute_query("SELECT Id, User_Name FROM Pose_Setting_Table")

    # 2. Determine Selected User
    selected_user_id = request.args.get('user_id')
    current_settings = {}

    if pos_users:
        if not selected_user_id:
            # Default to first user found or try to match current session user
            # Try matching session username first
            session_username = session.get('username')
            match = next((u for u in pos_users if u['User_Name'] == session_username), None)
            if match:
                selected_user_id = match['Id']
            else:
                selected_user_id = pos_users[0]['Id']

        # Fetch settings for selected user
        res = db.execute_query("SELECT * FROM Pose_Setting_Table WHERE Id = %s", (selected_user_id,))
        if res:
            current_settings = res[0]
            # Handle Image for Display (Convert bytes to base64)
            if current_settings.get('Image'):

                current_settings['ImageBase64'] = base64.b64encode(current_settings['Image']).decode('utf-8')

    locations = db.execute_query("SELECT inventory_locations_name FROM inventory_locations")
    accounts = db.execute_query("SELECT account_name FROM new_account_table") # For Card/Cash selection

    # Fetch SMS Delivery Logs
    sms_logs = []
    try:
        sms_logs = db.execute_query("SELECT * FROM sms_delivery_logs ORDER BY created_at DESC LIMIT 50")
    except Exception as e:
        print(f"Error fetching SMS logs: {e}")

    return render_template('pos_settings.html',
                           sms_logs=sms_logs,
                           settings=current_settings,
                           pos_users=pos_users,
                           selected_user_id=int(selected_user_id) if selected_user_id else 0,
                           locations=locations,
                           accounts=accounts)

@app.route('/add_pos_user', methods=['POST'])
@login_required
@has_permission('Access_POS')
def add_pos_user():
    username = request.form.get('username')
    password = request.form.get('password')
    mobile = request.form.get('mobile')

    if not username or not password:
        flash('Username and Password are required', 'danger')
        return redirect(url_for('pos_settings'))

    try:
        # Check duplicate
        exists = db.execute_query("SELECT Id FROM Pose_Setting_Table WHERE User_Name = %s", (username,))
        if exists:
            flash('Username already exists', 'danger')
            return redirect(url_for('pos_settings'))

        from werkzeug.security import generate_password_hash
        pw_hash = generate_password_hash(password)
        db.execute_query("""
            INSERT INTO Pose_Setting_Table (Id, User_Name, Password, Mobile_Number)
            VALUES (0, %s, %s, %s)
        """, (username, pw_hash, mobile), commit=True)

        flash(f'New Cashier {username} registered successfully', 'success')

    except Exception as e:
        flash(f'Error adding cashier: {str(e)}', 'danger')

    return redirect(url_for('pos_settings'))

# --- Point of Sale (POS) ---
@app.route('/pos', methods=['GET'])
@login_required
@has_permission('Access_POS')
def pos():
    return render_template('pos.html')
@app.route("/pos_customer_display", methods=["GET"])
@pos_login_required
def pos_customer_display():
    return render_template("pos_customer_display.html")


@app.route('/api/pos/login', methods=['POST'])
def pos_api_login():
    data = request.json or {}
    company_name = data.get('company_name')
    username = data.get('username')
    password = data.get('password')

    if not company_name or not username or not password:
        return {'success': False, 'error': 'Company Name, Username, and Password are required'}, 400

    try:
        master_user_res = master_db.execute_query('''
            SELECT t.db_name, t.is_active
            FROM tenants t
            WHERE t.company_name = %s
        ''', (company_name,))

        if master_user_res:
            if master_user_res[0]['is_active'] == 0:
                return {'success': False, 'error': 'Account suspended. Payment due.'}, 403
            tenant_db_name = master_user_res[0]['db_name']
        else:
            return {'success': False, 'error': 'Company not found'}, 404
    except Exception as e:
        tenant_db_name = db.db_name

    conn = None
    cursor = None

    try:
        conn = db.get_connection()
        cursor = conn.cursor(dictionary=True)
        if tenant_db_name and is_safe_db_name(tenant_db_name):
            cursor.execute(f"USE `{tenant_db_name}`")

        cursor.execute("SELECT * FROM Pose_Setting_Table WHERE User_Name = %s", (username,))
        users = cursor.fetchall()

        if not users:
            return {'success': False, 'error': 'Invalid username or password'}, 401

        user = users[0]

        if user.get('is_locked'):
            return {'success': False, 'error': 'Account locked due to too many failed attempts. Contact admin.'}, 403

        stored_password = user['Password']

        verified = False
        from werkzeug.security import check_password_hash, generate_password_hash
        try:
            if stored_password.startswith('scrypt:') or stored_password.startswith('pbkdf2:'):
                if check_password_hash(stored_password, password):
                    verified = True
            elif stored_password == password:
                verified = True
                new_hash = generate_password_hash(password)
                cursor.execute("UPDATE Pose_Setting_Table SET Password = %s WHERE Id = %s", (new_hash, user['Id']))
                conn.commit()
        except Exception:
            if stored_password == password:
                verified = True

        if not verified:
            fails = user.get('failed_attempts', 0) + 1
            if fails >= 3:
                cursor.execute("UPDATE Pose_Setting_Table SET failed_attempts = %s, is_locked = 1 WHERE Id = %s", (fails, user['Id']))
                conn.commit()
                return {'success': False, 'error': 'Account locked due to too many failed attempts. Contact admin.'}, 403
            else:
                cursor.execute("UPDATE Pose_Setting_Table SET failed_attempts = %s WHERE Id = %s", (fails, user['Id']))
                conn.commit()
                return {'success': False, 'error': 'Invalid username or password'}, 401

        # Successful login, reset attempts
        cursor.execute("UPDATE Pose_Setting_Table SET failed_attempts = 0 WHERE Id = %s", (user['Id'],))
        conn.commit()

        session['db_name'] = tenant_db_name
        session['username'] = username
        session['user_id'] = user['User_Name']
        session['user_pk'] = user['Id']

        cursor.execute('''
            SELECT
                i.id, i.inventoy_name, i.inventoy_code, i.inventoy_bach_code, i.inventoy_items_messurment_unit,
                p.inventory_price_selling, p.inventory_price_profit_marging_comen,
                p.inventory_price_for_Loyality_customer, p.inventory_price_purcharsing, i.expiry_date
            FROM inventoy_items i
            LEFT JOIN inventory_price_recod p ON i.id = p.inventory_price_link
        ''')
        items = cursor.fetchall()

        response_items = []
        for r in items:
            response_items.append({
                'id': r['id'],
                'name': r['inventoy_name'],
                'code': r['inventoy_code'],
                'barcode': r['inventoy_bach_code'],
                'unit': r['inventoy_items_messurment_unit'],
                'price_market': r['inventory_price_selling'],
                'price_special': r['inventory_price_profit_marging_comen'],
                'price_loyalty': r['inventory_price_for_Loyality_customer'],
                'price_cost': r['inventory_price_purcharsing'],
                'expiry_date': str(r.get('expiry_date')) if r.get('expiry_date') else None
            })

        return {
            'success': True,
            'settings': {
                'location': user['Select_Inventry_Location'],
                'card_ac': user['Card_Control_AC'],
                'cash_ac': user['Cash_Account'],
                'market_price': user['Sales_with_market_price'],
                'special_price': user['Sales_with_Special_price'],
                'loyalty_price': user['Loyalty_Price'],
                'vat_enable': user['VAT_Enable'],
                'footer': user['Footer_Message'],
                'top': user['Top_Message']
            },
            'items': response_items,
            'db_name': tenant_db_name
        }

    except Exception as e:
        import logging
        logging.error(f"POS Login Error: {e}")
        return {'success': False, 'error': 'Database error occurred'}, 500
    finally:
        if cursor: cursor.close()
        if conn: conn.close()




# --- POS Web Login with Device Fingerprinting & 2FA ---
def send_sms(mobile, message):
    """Sends an arbitrary SMS via the Notify.lk Gateway. Returns True on success."""
    settings = {}

    # Try to load credentials from active tenant DB site_settings if available
    conn = None
    try:
        with db.get_connection() as conn:
            with conn.cursor(dictionary=True) as cursor:
                # Handle cases where table doesn't exist yet gracefully
                try:
                    cursor.execute("SELECT setting_key, setting_value FROM site_settings WHERE setting_key IN ('sms_user_id', 'sms_api_key', 'sms_sender_id')")
                    settings = {r['setting_key']: r['setting_value'] for r in cursor.fetchall()}
                except Exception:
                    pass
    except Exception as e:
        logging.warning(f"Settings Load Error (Ignored): {e}")

    user_id = settings.get('sms_user_id') or os.getenv('NOTIFY_USER_ID')
    api_key = settings.get('sms_api_key') or os.getenv('NOTIFY_API_KEY')
    sender_id = settings.get('sms_sender_id') or os.getenv('NOTIFY_SENDER_ID', 'The Bunker')

    if not api_key or not user_id:
        logging.warning("NOTIFY_API_KEY or NOTIFY_USER_ID is not set. Skipping SMS delivery.")
        return False

    # Format number like the PHP script
    phone = str(mobile).strip().replace(" ", "").replace("-", "").replace("+", "")
    if phone.startswith("0"):
        phone = "94" + phone[1:]
    elif not phone.startswith("94"):
        phone = "94" + phone

    url = "https://app.notify.lk/api/v1/send"

    params = {
        'user_id': user_id,
        'api_key': api_key,
        'sender_id': sender_id,
        'to': phone,
        'message': message
    }

    try:
        logging.info(f"Sending SMS via Notify.lk to {phone} with sender {sender_id}")
        response = requests.get(url, params=params, timeout=10, verify=False)
        result = response.json()

        status_msg = result.get('status', 'failed')

        # Log to DB if table exists
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            cursor.execute("INSERT INTO sms_delivery_logs (mobile, message, status, api_response) VALUES (%s, %s, %s, %s)",
                           (phone, message, status_msg, response.text))
            conn.commit()
        except Exception as log_err:
            logging.error(f"Error logging SMS: {log_err}")
        finally:
            if 'cursor' in locals() and cursor: cursor.close()
            if 'conn' in locals() and conn: conn.close()

        if status_msg == 'success':
            logging.info(f"SMS delivered successfully to {phone}.")
            return True
        else:
            logging.error(f"NotifySMS API Error: {response.text}")
            return False
    except Exception as e:
        logging.error(f"Failed to send SMS: {e}")
        return False


def send_sms_otp(mobile, code):
    """Sends an OTP verification SMS via the Notify.lk Gateway."""
    return send_sms(mobile, f"Your SUWIN verification code is {code}.")
@app.route('/pos_login', methods=['GET', 'POST'])
def pos_web_login():
    if request.method == 'GET':
        return render_template('pos_login.html')

    company_name = request.form.get('company_name')
    username = request.form.get('username')
    password = request.form.get('password')

    if not company_name or not username or not password:
        flash('Company Name, Username, and Password are required', 'danger')
        return redirect(url_for('pos_web_login'))

    # Locate DB from company name in master DB
    try:
        master_user_res = master_db.execute_query('''
            SELECT t.db_name, t.is_active
            FROM tenants t
            WHERE t.company_name = %s
        ''', (company_name,))
        if master_user_res:
            if master_user_res[0]['is_active'] == 0:
                return redirect(url_for('payment_due'))
            tenant_db_name = master_user_res[0]['db_name']
        else:
            flash('Company not found', 'danger')
            return redirect(url_for('pos_web_login'))
    except Exception as e:
        tenant_db_name = db.db_name

    conn = None
    try:
        with db.get_connection() as conn:
            with conn.cursor(dictionary=True) as cursor:
                if tenant_db_name and is_safe_db_name(tenant_db_name):
                    cursor.execute(f"USE `{tenant_db_name}`")

                cursor.execute("SELECT * FROM Pose_Setting_Table WHERE User_Name = %s", (username,))
                users = cursor.fetchall()

                if not users:
                    flash('Invalid username or password', 'danger')
                    return redirect(url_for('pos_web_login'))

                user = users[0]

                if user.get('is_locked'):
                    flash('Account locked due to too many failed attempts. Contact admin.', 'danger')
                    return redirect(url_for('pos_web_login'))

                stored_password = user['Password']
                verified = False
                try:
                    if stored_password.startswith('scrypt:') or stored_password.startswith('pbkdf2:'):
                        if check_password_hash(stored_password, password):
                            verified = True
                    elif stored_password == password:
                        verified = True
                        new_hash = generate_password_hash(password)
                        cursor.execute("UPDATE Pose_Setting_Table SET Password = %s WHERE Id = %s", (new_hash, user['Id']))
                        conn.commit()
                except Exception:
                    if stored_password == password:
                        verified = True

                if not verified:
                    fails = user.get('failed_attempts', 0) + 1
                    if fails >= 3:
                        cursor.execute("UPDATE Pose_Setting_Table SET failed_attempts = %s, is_locked = 1 WHERE Id = %s", (fails, user['Id']))
                        conn.commit()
                        flash('Account locked due to too many failed attempts. Contact admin.', 'danger')
                    else:
                        cursor.execute("UPDATE Pose_Setting_Table SET failed_attempts = %s WHERE Id = %s", (fails, user['Id']))
                        conn.commit()
                        flash('Invalid username or password', 'danger')
                    return redirect(url_for('pos_web_login'))

                # Successful auth -> Reset failed attempts
                cursor.execute("UPDATE Pose_Setting_Table SET failed_attempts = 0 WHERE Id = %s", (user['Id'],))
                conn.commit()

                # Check Device Fingerprint
                ip_address = request.remote_addr
                user_agent = request.user_agent.string

                cursor.execute("SELECT * FROM pos_user_devices WHERE user_id = %s", (user['Id'],))
                all_devices = cursor.fetchall()

                device = next((d for d in all_devices if d['ip_address'] == ip_address and d['user_agent'] == user_agent), None)

                if not all_devices:
                    # Very first login - register device seamlessly
                    cursor.execute("INSERT INTO pos_user_devices (user_id, ip_address, user_agent, last_login) VALUES (%s, %s, %s, %s)",
                                   (user['Id'], ip_address, user_agent, datetime.now()))
                    conn.commit()
                elif not device:
                    # New device but user has previous devices - Require 2FA
                    otp = ''.join(random.choices(string.digits, k=6))
                    expires = datetime.now() + timedelta(minutes=10)
                    cursor.execute("INSERT INTO pos_2fa_codes (user_id, code, expires_at) VALUES (%s, %s, %s)", (user['Id'], otp, expires))
                    conn.commit()

                    # Send SMS
                    mobile = user.get('Mobile_Number')
                    if mobile:
                        sms_sent = send_sms_otp(mobile, otp)
                        if not sms_sent:
                            flash('Failed to send SMS verification code. Please check your SMS Gateway settings or contact support.', 'danger')
                            return redirect(url_for('pos_web_login'))
                    else:
                        logging.warning(f"Cannot send 2FA SMS for User {user['Id']} because Mobile_Number is NULL.")
                        flash('Your account requires SMS verification, but no mobile number is registered. Please contact support.', 'danger')
                        return redirect(url_for('pos_web_login'))

                    session['pending_pos_user_id'] = user['Id']
                    session['pending_pos_company'] = company_name
                    return render_template('pos_2fa.html')
                else:
                    # Existing verified device
                    cursor.execute("UPDATE pos_user_devices SET last_login = %s WHERE id = %s", (datetime.now(), device['id']))
                    conn.commit()

                if user.get('must_change_password'):
                    session['pending_pos_user_id'] = user['Id']
                    session['pending_pos_company'] = company_name
                    return render_template('pos_reset_password.html')

                # Final Login Success
                session['db_name'] = tenant_db_name
                session['username'] = username
                session['user_id'] = user['User_Name']
                session['user_pk'] = user['Id']
                run_schema_migrations()

                return redirect(url_for('pos'))
    except Exception as e:
        logging.error(f"POS Login DB Error: {e}")
        flash('An error occurred during login', 'danger')
        return redirect(url_for('pos_web_login'))

@app.route('/pos_verify_2fa', methods=['POST'])
def pos_verify_2fa():
    user_id = session.get('pending_pos_user_id')
    company_name = session.get('pending_pos_company')
    otp = request.form.get('otp')

    if not user_id or not company_name:
        flash('Session expired or invalid.', 'danger')
        return redirect(url_for('pos_web_login'))

    # Needs Tenant DB connection again to verify
    try:
        master_user_res = master_db.execute_query('SELECT db_name FROM tenants WHERE company_name = %s', (company_name,))
        tenant_db_name = master_user_res[0]['db_name'] if master_user_res else db.db_name
    except Exception:
        tenant_db_name = db.db_name

    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        if tenant_db_name and is_safe_db_name(tenant_db_name):
            cursor.execute(f"USE `{tenant_db_name}`")

        cursor.execute("SELECT * FROM pos_2fa_codes WHERE user_id = %s AND code = %s AND is_used = 0 AND expires_at > %s", (user_id, otp, datetime.now()))
        valid_code = cursor.fetchone()

        if valid_code:
            # Mark used
            cursor.execute("UPDATE pos_2fa_codes SET is_used = 1 WHERE id = %s", (valid_code['id'],))

            # Register device
            ip_address = request.remote_addr
            user_agent = request.user_agent.string
            cursor.execute("INSERT INTO pos_user_devices (user_id, ip_address, user_agent, last_login) VALUES (%s, %s, %s, %s)", (user_id, ip_address, user_agent, datetime.now()))

            # Fetch User to setup session
            cursor.execute("SELECT * FROM Pose_Setting_Table WHERE Id = %s", (user_id,))
            user = cursor.fetchone()
            conn.commit()

            # Since this is a new device, force password change immediately
            cursor.execute("UPDATE Pose_Setting_Table SET must_change_password = 1 WHERE Id = %s", (user_id,))
            conn.commit()

            return render_template('pos_reset_password.html')
        else:
            flash('Invalid or expired code.', 'danger')
            return render_template('pos_2fa.html')
    finally:
        cursor.close()
        conn.close()

@app.route('/pos_reset_password', methods=['POST'])
def pos_reset_password():
    from werkzeug.security import generate_password_hash
    user_id = session.get('pending_pos_user_id')
    company_name = session.get('pending_pos_company')
    new_password = request.form.get('new_password')
    confirm_password = request.form.get('confirm_password')

    if not user_id or not company_name:
        flash('Session expired or invalid.', 'danger')
        return redirect(url_for('pos_web_login'))

    if new_password != confirm_password:
        flash('Passwords do not match.', 'danger')
        return render_template('pos_reset_password.html')

    try:
        master_user_res = master_db.execute_query('SELECT db_name FROM tenants WHERE company_name = %s', (company_name,))
        tenant_db_name = master_user_res[0]['db_name'] if master_user_res else db.db_name
    except Exception:
        tenant_db_name = db.db_name

    conn = db.get_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        if tenant_db_name and is_safe_db_name(tenant_db_name):
            cursor.execute(f"USE `{tenant_db_name}`")

        new_hash = generate_password_hash(new_password)
        cursor.execute("UPDATE Pose_Setting_Table SET Password = %s, must_change_password = 0 WHERE Id = %s", (new_hash, user_id))

        cursor.execute("SELECT * FROM Pose_Setting_Table WHERE Id = %s", (user_id,))
        user = cursor.fetchone()
        conn.commit()

        session['db_name'] = tenant_db_name
        session['username'] = user['User_Name']
        session['user_id'] = user['User_Name']
        session['user_pk'] = user['Id']
        run_schema_migrations()

        session.pop('pending_pos_user_id', None)
        session.pop('pending_pos_company', None)

        flash('Password updated successfully.', 'success')
        return redirect(url_for('pos'))
    finally:
        cursor.close()
        conn.close()


@app.route('/api/pos/settings', methods=['GET'])
@login_required
def pos_api_settings():
    username = session.get('username')

    # Verify against Pose_Setting_Table
    users = db.execute_query("SELECT * FROM Pose_Setting_Table WHERE User_Name = %s", (username,))

    if users:
        settings = users[0]
        return {
            'success': True,
            'settings': {
                'location': settings['Select_Inventry_Location'],
                'card_ac': settings['Card_Control_AC'],
                'cash_ac': settings['Cash_Account'],
                'market_price': settings['Sales_with_market_price'],
                'special_price': settings['Sales_with_Special_price'],
                'loyalty_price': settings['Loyalty_Price'],
                'vat_enable': settings['VAT_Enable'],
                'footer': settings['Footer_Message'],
                'top': settings['Top_Message']
            }
        }

    return {'success': False, 'error': 'POS settings not found for current user'}

@app.route('/api/pos/items', methods=['GET'])
@pos_login_required
def pos_api_items():
    global _pos_items_cache
    db_name = get_session_db_name()
    current_time = time.time()

    # Check cache (5 minutes TTL = 300 seconds)
    if db_name in _pos_items_cache:
        cached_data = _pos_items_cache[db_name]
        if current_time - cached_data['timestamp'] < 300:
            return cached_data['data']

    # Fetch all active items with prices for caching
    query = '''
        SELECT
            i.id, p.id as price_recod_id, i.inventoy_name, i.inventoy_code, i.inventoy_bach_code, i.inventoy_items_messurment_unit,
            p.inventory_price_selling, p.inventory_price_profit_marging_comen,
            p.inventory_price_for_Loyality_customer, p.inventory_price_purcharsing, i.expiry_date
        FROM inventoy_items i
        LEFT JOIN inventory_price_recod p ON i.id = p.inventory_price_link
        WHERE i.active = 1
    '''
    rows = db.execute_query(query)

    items = []
    for r in rows:
        items.append({
            'id': r['id'],
            'price_id': r.get('price_recod_id'),
            'name': r['inventoy_name'],
            'code': r['inventoy_code'],
            'batch_code': r['inventoy_bach_code'],
            'unit': r['inventoy_items_messurment_unit'],
            'price_market': float(r['inventory_price_selling'] or 0),
            'price_special': float(r['inventory_price_profit_marging_comen'] or 0),
            'price_loyalty': float(r['inventory_price_for_Loyality_customer'] or 0),
            'cost': float(r['inventory_price_purcharsing'] or 0),
            'expiry_date': str(r.get('expiry_date')) if r.get('expiry_date') else None
        })

    result = json.dumps(items)
    _pos_items_cache[db_name] = {
        'timestamp': current_time,
        'data': result
    }

    return result

# Global cache for pos customers
pos_customers_cache = {'data': None, 'timestamp': 0}

@app.route('/api/pos/customers', methods=['GET'])
@pos_login_required
def pos_api_customers():
    global pos_customers_cache
    current_time = time.time()

    # Cache duration: 60 seconds
    if pos_customers_cache['data'] and (current_time - pos_customers_cache['timestamp'] < 60):
        return pos_customers_cache['data']

    # Fetch customers for caching
    query = "SELECT id, customer_name, Mobile_nimber FROM customer WHERE Compay_Or_Not = 0 OR Compay_Or_Not IS NULL"
    rows = db.execute_query(query)

    custs = []
    for r in rows:
        custs.append({
            'id': r['id'],
            'name': r['customer_name'],
            'mobile': r['Mobile_nimber']
        })

    cached_data = json.dumps(custs)
    pos_customers_cache['data'] = cached_data
    pos_customers_cache['timestamp'] = current_time

    return cached_data

@app.route('/api/pos/add_loyalty_customer', methods=['POST'])
@login_required
def pos_api_add_loyalty_customer():
    data = request.json
    name = data.get('name')
    mobile = data.get('mobile')
    email = data.get('email', '')
    billing_address = data.get('billing_address', '')
    delivery_address = data.get('delivery_address', '')
    amount_paid = data.get('amount_paid', 0)

    if not name or not mobile:
        return {'success': False, 'error': 'Name and Mobile Number are required'}

    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        # Determine max id to generate customer_code
        cursor.execute("SELECT COALESCE(MAX(id), 0) FROM customer")
        max_id = cursor.fetchone()[0]
        customer_code = max_id + 60001

        query = """
            INSERT INTO customer (
                customer_name, customer_code, customer_Billing_Address, costomer_Delivery_Address,
                e_mail, coustomer_credit_limit, Mobile_nimber, Is_Loyality_Customer, Compay_Or_Not,
                Create_Date, Paid_Amountl, Create_Cashiyer
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        # Note: Depending on the schema, Paid_Amountl might be a string or number,
        # and Is_Loyality_Customer might be a boolean/tinyint.
        cursor.execute(query, (
            name, str(customer_code), billing_address, delivery_address,
            email, 1, mobile, 1, 0,
            datetime.utcnow().strftime("%Y-%m-%d"), amount_paid, 0
        ))

        conn.commit()

        # Fetch the new customer details to return
        new_customer_id = cursor.lastrowid
        cursor.close()
        conn.close()

        return {
            'success': True,
            'customer': {
                'id': new_customer_id,
                'name': name,
                'mobile': mobile
            }
        }
    except Exception as e:
        if conn:
            conn.rollback()
        return {'success': False, 'error': str(e)}

def _generate_pos_invoice_number(cursor, today_date):
    """Helper to generate a new POS Invoice Number."""
    cursor.execute("INSERT INTO pos_invoice_no (IV_No) VALUES ('')")
    last_id = cursor.lastrowid
    invoice_no = f"{today_date.year}POS-{last_id}"
    cursor.execute("UPDATE pos_invoice_no SET IV_No = %s WHERE Id = %s", (invoice_no, last_id))
    return invoice_no

def _process_pos_cart_items(cursor, cart, settings, current_user, current_user_pk, payment, customer, invoice_no, jv_no, today_date):
    """Helper to process items, yielding total sales and total costs, and executing batch inserts."""
    total_sale_value = 0
    total_cost_value = 0
    pos_sales_params = []
    inventory_params = []
    action_date_str = today_date.strftime('%Y-%m-%d')

    for item in cart:
        qty = parse_float(item.get('qty', 0))
        cost = parse_float(item.get('cost', 0))
        total = parse_float(item.get('total', 0))

        # If the frontend passes 'cost' as the total cost already, multiplying it by qty squares it.
        # But we assume 'cost' is unit cost based on pos.html. Just to be safe and match legacy C#:
        # Wait, if pos.html sends unit cost, then unit_cost * qty is correct.
        total_item_cost = cost * qty

        total_sale_value += total
        total_cost_value += total_item_cost

        # Prepare POS_Sales_Invoice_01 params
        pos_sales_params.append((
            item.get('code'), item.get('name'), item.get('unit'),
            item.get('price_market'), item.get('price_special'), item.get('price_loyalty'),
            settings.get('market_active', 0), settings.get('special_active', 0), settings.get('loyalty_active', 0),
            current_user_pk, settings.get('location'), action_date_str, qty, cost,
            payment.get('method'), settings.get('cash_ac'), settings.get('bank_ac'),
            invoice_no, customer.get('loyalty_no', 0), total, jv_no
        ))

        # Prepare Inventory Movement OUT params.
        # Legacy C# app and submit_invoice BOTH expect total cost in inventory_recod_unit_price
        inventory_params.append((
            item.get('name'), item.get('code'), today_date, qty, item.get('unit'), total_item_cost,
            current_user_pk, jv_no, settings.get('location')
        ))

    # Batch Insert into POS_Sales_Invoice_01
    if pos_sales_params:
        cursor.executemany("""
            INSERT INTO POS_Sales_Invoice_01 (
                ItemCoude, ItemName, ItemMesurmet, SllingPrice, ItemPriceComen, ItemLoyalityPrice,
                Sales_with_market_price_Active, Sales_with_Special_price_Active, Loyalty_Price_Active,
                RecodeUserId, Location, AcctionDate, QuntirySale, InventoryCost, PaymentMethord,
                CashAccountName, BankAccountName, Invoice_No, Loyalty_No, Total_Value, jv, Revers
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 0)
        """, pos_sales_params)

    # Batch Insert into inventory_recod
    if inventory_params:
        cursor.executemany("""
            INSERT INTO inventory_recod (
                inventoy_name, inventoy_code, inventory_recod_action_date,
                inventory_recod_moument_in, inventory_recod_movment_out,
                inventory_recod_mesrmet, inventory_recod_unit_price,
                inventory_recod_account, inventory_recod_user_id, JV_No,
                inventory_recod_location
            ) VALUES (%s, %s, %s, 0, %s, %s, %s, 'Cost Of Goods Sold', %s, %s, %s)
        """, inventory_params)

    return total_sale_value, total_cost_value

def _post_pos_gl_entries(cursor, settings, payment, total_sale_value, total_cost_value, today_date, invoice_no, current_user_pk, jv_no):
    """Helper to post corresponding General Ledger entries for POS sale."""
    # Debit Cash/Bank
    ac_name = settings.get('cash_ac') if payment.get('method') == 1 else settings.get('bank_ac')
    cursor.execute("""
        INSERT INTO entry_details (
            account_name, enty_values_DR, entry_effective_date, entry_create_date,
            entry_naration, entry_create_user, entry_jv
        ) VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (ac_name, total_sale_value, today_date, today_date, f"POS Sale {invoice_no}", current_user_pk, jv_no))

    # Calculate VAT if enabled
    vat_enabled = settings.get('vat_enable') == 1
    vat_amount = 0
    net_sales = total_sale_value

    if vat_enabled:
        # Fetch VAT Rate from Tax Settings if exists, else 18%
        cursor.execute("SELECT rate FROM tax_rates WHERE tax_name LIKE '%VAT%' AND active=1 LIMIT 1")
        res_vat = cursor.fetchone()
        vat_rate = res_vat[0] if res_vat else 18.0

        net_sales = total_sale_value / (1 + (vat_rate / 100))
        vat_amount = total_sale_value - net_sales

    # Credit Sales Account (Net)
    cursor.execute("""
        INSERT INTO entry_details (
            account_name, enty_values_CR, entry_effective_date, entry_create_date,
            entry_naration, entry_create_user, entry_jv
        ) VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, ('Sales', net_sales, today_date, today_date, f"POS Sale {invoice_no}", current_user_pk, jv_no))

    # Credit VAT Control (If VAT > 0)
    if vat_amount > 0:
        cursor.execute("""
            INSERT INTO entry_details (
                account_name, enty_values_CR, entry_effective_date, entry_create_date,
                entry_naration, entry_create_user, entry_jv
            ) VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, ('VAT Control', vat_amount, today_date, today_date, f"VAT on POS Sale {invoice_no}", current_user_pk, jv_no))

    # Cost of Goods Sold (DR COGS, CR Inventory)
    if total_cost_value > 0:
        # Debit Cost Of Goods Sold
        cursor.execute("""
            INSERT INTO entry_details (
                account_name, enty_values_DR, entry_effective_date, entry_create_date,
                entry_naration, entry_create_user, entry_jv
            ) VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, ('Cost Of Goods Sold', total_cost_value, today_date, today_date, f"POS Sale {invoice_no} (COGS)", current_user_pk, jv_no))

        # Credit Inventory
        cursor.execute("""
            INSERT INTO entry_details (
                account_name, enty_values_CR, entry_effective_date, entry_create_date,
                entry_naration, entry_create_user, entry_jv
            ) VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, ('Inventory', total_cost_value, today_date, today_date, f"POS Sale {invoice_no} (COGS)", current_user_pk, jv_no))


@app.route('/pos/submit_sale', methods=['POST'])
@app.route('/api/pos/submit', methods=['POST'])
@login_required
def submit_pos_sale():
    data = request.json

    # Client is expected to save JSON locally as per requirement before sending

    cart = data.get('cart', [])
    payment = data.get('payment', {})
    customer = data.get('customer', {})
    settings = data.get('settings', {})

    if not cart: return {'error': 'Cart is empty'}, 400

    current_user = get_current_user_id()
    current_user_pk = get_current_user_pk()
    today_date = date.today()

    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        conn.start_transaction()

        # 1. Generate Invoice No
        invoice_no = _generate_pos_invoice_number(cursor, today_date)

        # 2. Create JV
        cursor.execute("INSERT INTO jv_numbers (jv_user_code, jv_naration) VALUES (%s, %s)", ('JV FROM POS', f"POS Sale {invoice_no}"))
        jv_no = cursor.lastrowid

        # 3. Process Cart Items
        total_sale_value, total_cost_value = _process_pos_cart_items(
            cursor, cart, settings, current_user, current_user_pk, payment, customer, invoice_no, jv_no, today_date
        )

        # 4. GL Entries
        _post_pos_gl_entries(
            cursor, settings, payment, total_sale_value, total_cost_value, today_date, invoice_no, current_user_pk, jv_no
        )

        conn.commit()
        return {'success': True, 'invoice_no': invoice_no, 'jv': jv_no}

    except Exception as e:
        conn.rollback()
        logging.error(f"POS Error: {e}")
        return {'error': str(e)}, 500
    finally:
        cursor.close()
        conn.close()

def run_schema_migrations(target_db_conn=None):
    """Checks and updates database schema for new features."""
    conn = target_db_conn if target_db_conn else db.get_connection()
    if not conn: return
    migrations.run_migrations(conn)
    try:
        cursor = conn.cursor()

        # 0. Migration Table
        try:
            cursor.execute("CREATE TABLE IF NOT EXISTS migrations (id INT AUTO_INCREMENT PRIMARY KEY, migration_name VARCHAR(255) UNIQUE NOT NULL, applied_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)")
        except mysql.connector.Error as e:
            if e.errno not in (1050, 1007, 1060, 1061, 1146, 1054):
                logging.error(f"Error creating migrations table: {e}")
        except Exception as e:
            logging.error(f"Error creating migrations table: {e}")

        # Helper to check/record migration
        def is_migration_applied(name):
            try:
                cursor.execute("SELECT id FROM migrations WHERE migration_name = %s", (name,))
                return cursor.fetchone() is not None
            except Exception:
                return False

        def record_migration(name):
            try:
                cursor.execute("INSERT INTO migrations (migration_name) VALUES (%s)", (name,))
                conn.commit()
            except Exception as e:
                logging.error(f"Error recording migration {name}: {e}")

        # 1. Password Hashing Migration (Modify VARCHAR Length)
        # Login_Table
        cursor.execute("SHOW COLUMNS FROM Login_Table LIKE 'Password'")
        res = cursor.fetchone()
        if res:
            # res structure depends on driver, usually tuple
            # Field, Type, Null, Key, Default, Extra
            # Type is index 1
            col_type = res[1]
            # Check if it is varchar(45) or similar short length
            if (isinstance(col_type, bytes) and b'varchar(45)' in col_type.lower()) or (isinstance(col_type, str) and 'varchar(45)' in col_type.lower()):
                pass
                cursor.execute("ALTER TABLE Login_Table MODIFY COLUMN Password VARCHAR(255)")

        # Pose_Setting_Table
        cursor.execute("SHOW TABLES LIKE 'Pose_Setting_Table'")
        if cursor.fetchone():
            cursor.execute("SHOW COLUMNS FROM Pose_Setting_Table LIKE 'Password'")
            res_pos = cursor.fetchone()
            if res_pos:
                col_type_pos = res_pos[1]
                if (isinstance(col_type_pos, bytes) and b'varchar(45)' in col_type_pos.lower()) or (isinstance(col_type_pos, str) and 'varchar(45)' in col_type_pos.lower()):
                    pass
                    cursor.execute("ALTER TABLE Pose_Setting_Table MODIFY COLUMN Password VARCHAR(255)")

        # 1b. User_Rights Columns
        cursor.execute("SHOW COLUMNS FROM User_Rights")
        columns = [row[0] for row in cursor.fetchall()]

        new_columns = [
            'Access_Inventory', 'Access_POS', 'Access_Accounting', 'Access_Reports',
            'Access_Reversals', 'Access_Fixed_Assets', 'Access_VAT', 'Access_HR', 'Access_CRM',
            'Access_Sales', 'Access_Purchase', 'Access_Settings'
        ]

        for col in new_columns:
            if col not in columns:
                logging.info(f"Migrating: Adding {col} to User_Rights")
                try:
                    cursor.execute(f"ALTER TABLE User_Rights ADD COLUMN {col} TINYINT DEFAULT 0")
                    conn.commit()
                except Exception as e:
                    logging.error(f"Migration add column {col}: {e}")

        # 1b. Password Column Expansion
        # Login_Table
        cursor.execute("SHOW COLUMNS FROM Login_Table LIKE 'Password'")
        res = cursor.fetchone()
        if res:
            # res format: ('Password', 'varchar(45)', ...)
            col_type = res[1].lower()
            if (isinstance(col_type, bytes) and b'varchar(45)' in col_type) or (isinstance(col_type, str) and 'varchar(45)' in col_type):
                pass
                cursor.execute("ALTER TABLE Login_Table MODIFY COLUMN Password VARCHAR(255)")

        # Pose_Setting_Table
        cursor.execute("SHOW TABLES LIKE 'Pose_Setting_Table'")
        if cursor.fetchone():
            cursor.execute("SHOW COLUMNS FROM Pose_Setting_Table LIKE 'Password'")
            res = cursor.fetchone()
            if res:
                col_type = res[1].lower()
                if (isinstance(col_type, bytes) and b'varchar(45)' in col_type) or (isinstance(col_type, str) and 'varchar(45)' in col_type):
                    pass
                    cursor.execute("ALTER TABLE Pose_Setting_Table MODIFY COLUMN Password VARCHAR(255)")

        # 2. Currency Table
        cursor.execute("SHOW TABLES LIKE 'currency_table'")
        if not cursor.fetchone():
            logging.info("Migrating: Creating currency_table")
            cursor.execute("""
                CREATE TABLE currency_table (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    currency_code VARCHAR(10) NOT NULL UNIQUE,
                    currency_name VARCHAR(100),
                    is_base_currency TINYINT DEFAULT 0
                )
            """)
            # Insert default if empty
            cursor.execute("INSERT INTO currency_table (currency_code, currency_name, is_base_currency) VALUES ('LKR', 'Sri Lankan Rupee', 1)")

        # 3. New Account Table Columns
        if not is_migration_applied('add_currency_code_to_new_account'):
            try:
                # Check column existence just in case, or use ADD COLUMN IF NOT EXISTS (MariaDB 10.2+)
                # Standard MySQL doesn't support IF NOT EXISTS in ALTER TABLE nicely without stored proc.
                # But since we track migration_name, we assume if not applied, we run it.
                # Wrap in try-except to handle "Duplicate column" if manually added.
                cursor.execute("ALTER TABLE new_account_table ADD COLUMN currency_code VARCHAR(10) DEFAULT 'LKR'")
                record_migration('add_currency_code_to_new_account')
                logging.info("Migrated: add_currency_code_to_new_account")
            except Exception as e:
                if "Duplicate column" in str(e) or "1060" in str(e):
                    record_migration('add_currency_code_to_new_account')
                else:
                    logging.error(f"Migration failed: {e}")

        # 4. Inventory Items Columns (UOM)
        cursor.execute("SHOW COLUMNS FROM inventoy_items")
        inv_columns = [row[0] for row in cursor.fetchall()]
        if 'uom_secondary' not in inv_columns:
            logging.info("Migrating: Adding uom_secondary to inventoy_items")
            cursor.execute("ALTER TABLE inventoy_items ADD COLUMN uom_secondary VARCHAR(45) NULL")

        if 'uom_conversion_rate' not in inv_columns:
            logging.info("Migrating: Adding uom_conversion_rate to inventoy_items")
            cursor.execute("ALTER TABLE inventoy_items ADD COLUMN uom_conversion_rate DOUBLE DEFAULT 1")

        # 5. Suppliers Table Columns (TIN, NIC)
        cursor.execute("SHOW COLUMNS FROM suppliers")
        sup_columns = [row[0] for row in cursor.fetchall()]

        if 'suppliers_TIN' not in sup_columns:
            logging.info("Migrating: Adding suppliers_TIN to suppliers")
            cursor.execute("ALTER TABLE suppliers ADD COLUMN suppliers_TIN VARCHAR(50) NULL")

        if 'suppliers_NIC' not in sup_columns:
            logging.info("Migrating: Adding suppliers_NIC to suppliers")
            cursor.execute("ALTER TABLE suppliers ADD COLUMN suppliers_NIC VARCHAR(20) NULL")

        # 5b. Company Table (VAT Registered)
        cursor.execute("SHOW COLUMNS FROM company")
        comp_columns = [row[0] for row in cursor.fetchall()]
        if 'vat_registered' not in comp_columns:
            logging.info("Migrating: Adding vat_registered to company")
            cursor.execute("ALTER TABLE company ADD COLUMN vat_registered TINYINT DEFAULT 0")

        # 6. Tax Rates Table
        cursor.execute("SHOW TABLES LIKE 'tax_rates'")
        if not cursor.fetchone():
            logging.info("Migrating: Creating tax_rates table")
            cursor.execute("""
                CREATE TABLE tax_rates (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    tax_name VARCHAR(100) NOT NULL,
                    rate DOUBLE NOT NULL,
                    description VARCHAR(255),
                    active TINYINT DEFAULT 1
                )
            """)
            # Default Data
            cursor.execute("INSERT INTO tax_rates (tax_name, rate, description) VALUES ('WHT - Interest', 10.0, 'Withholding Tax on Interest')")
            cursor.execute("INSERT INTO tax_rates (tax_name, rate, description) VALUES ('WHT - Rent', 10.0, 'Withholding Tax on Rent')")
            cursor.execute("INSERT INTO tax_rates (tax_name, rate, description) VALUES ('WHT - Professional Fees', 5.0, 'Withholding Tax on Professional Fees')")

        # 7. Cheque Print Settings Table
        cursor.execute("SHOW TABLES LIKE 'cheque_print_settings'")
        if not cursor.fetchone():
            logging.info("Migrating: Creating cheque_print_settings table")
            cursor.execute("""
                CREATE TABLE cheque_print_settings (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    bank_account VARCHAR(100) NULL,
                    paper_width_mm DOUBLE DEFAULT 175,
                    paper_height_mm DOUBLE DEFAULT 76,

                    date_x DOUBLE DEFAULT 140,
                    date_y DOUBLE DEFAULT 10,
                    date_font_size INT DEFAULT 10,

                    payee_x DOUBLE DEFAULT 20,
                    payee_y DOUBLE DEFAULT 25,
                    payee_font_size INT DEFAULT 11,

                    amount_words_x DOUBLE DEFAULT 25,
                    amount_words_y DOUBLE DEFAULT 35,
                    amount_words_font_size INT DEFAULT 10,
                    amount_words_width DOUBLE DEFAULT 130,

                    amount_digits_x DOUBLE DEFAULT 140,
                    amount_digits_y DOUBLE DEFAULT 35,
                    amount_digits_font_size INT DEFAULT 12,

                    is_cross_cheque TINYINT DEFAULT 1
                )
            """)

        # 8. Proforma Invoice Tables
        cursor.execute("SHOW TABLES LIKE 'proforma_invoice_header'")
        if not cursor.fetchone():
            logging.info("Migrating: Creating proforma_invoice_header table")
            cursor.execute("""
                CREATE TABLE proforma_invoice_header (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    pi_number VARCHAR(50) NOT NULL UNIQUE,
                    customer_name VARCHAR(200),
                    pi_date DATE,
                    expiry_date DATE,
                    subtotal DOUBLE,
                    vat_amount DOUBLE,
                    grand_total DOUBLE,
                    narration TEXT,
                    created_by INT,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            """)

        cursor.execute("SHOW TABLES LIKE 'proforma_invoice_details'")
        if not cursor.fetchone():
            logging.info("Migrating: Creating proforma_invoice_details table")
            cursor.execute("""
                CREATE TABLE proforma_invoice_details (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    pi_id INT,
                    item_name VARCHAR(200),
                    description TEXT,
                    qty DOUBLE,
                    unit_price DOUBLE,
                    total DOUBLE,
                    FOREIGN KEY (pi_id) REFERENCES proforma_invoice_header(id) ON DELETE CASCADE
                )
            """)

        # 9. Approval Workflow Updates
        # Add status columns to transaction tables
        # Status: 0=Parked, 1=Posted/Approved, 2=Rejected

        # OP_NO_Table (Purchase Orders)
        cursor.execute("SHOW COLUMNS FROM OP_NO_Table")
        op_cols = [row[0] for row in cursor.fetchall()]
        if 'status' not in op_cols:
            logging.info("Migrating: Adding status to OP_NO_Table")
            cursor.execute("ALTER TABLE OP_NO_Table ADD COLUMN status TINYINT DEFAULT 1")
            # Default 1 (Posted) for existing data to avoid breaking current flow

        # jv_numbers (Journal Vouchers - covers JV, Payments, Receipts)
        cursor.execute("SHOW COLUMNS FROM jv_numbers")
        jv_cols = [row[0] for row in cursor.fetchall()]
        if 'status' not in jv_cols:
            logging.info("Migrating: Adding status to jv_numbers")
            cursor.execute("ALTER TABLE jv_numbers ADD COLUMN status TINYINT DEFAULT 1")

        # System Settings Table (for toggles)
        cursor.execute("SHOW TABLES LIKE 'system_settings'")
        if not cursor.fetchone():
            logging.info("Migrating: Creating system_settings table")
            cursor.execute("""
                CREATE TABLE system_settings (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    setting_key VARCHAR(100) UNIQUE,
                    setting_value VARCHAR(255),
                    description VARCHAR(255)
                )
            """)
            # Default: Disable workflow initially (value '0') to match user request "if need it need to diasable"
            # User request: "post parking is doing ground workers and post is doing managing level if need it need to diasable"
            # I interpret this as "Enable it, but allow disabling". Let's default to DISABLED ('0') to not disrupt current flow immediately, or ENABLED ('1')?
            # User says "I Need to authentication proses... we cant bilt it like park and post".
            # "we cant bilt it like park and post" might mean "we can build it"? Or "we count bill it"?
            # Context: "parking is doing ground workers and post is doing managing level".
            # "if need it need to diasable" -> feature toggle.
            # I'll default to '0' (Disabled) so they can turn it on when ready.
            cursor.execute("INSERT INTO system_settings (setting_key, setting_value, description) VALUES ('enable_approval_workflow', '0', 'Enable Park & Post Workflow (0=Disabled, 1=Enabled)')")

        # 10. Master Payment Voucher Sequence
        cursor.execute("SHOW TABLES LIKE 'master_payment_voucher_no'")
        if not cursor.fetchone():
            logging.info("Migrating: Creating master_payment_voucher_no table")
            cursor.execute("""
                CREATE TABLE master_payment_voucher_no (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    voucher_no BIGINT NOT NULL,
                    create_date DATE
                )
            """)
            cursor.execute("INSERT INTO master_payment_voucher_no (voucher_no, create_date) VALUES (0, %s)", (date.today(),))

        try:
            cursor.execute('''
            CREATE TABLE IF NOT EXISTS recent_activity (
                id INT AUTO_INCREMENT PRIMARY KEY,
                dot_color VARCHAR(20) DEFAULT 'blue',
                text_content TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
            ''')
        except mysql.connector.Error as e:
            if e.errno not in (1050, 1007, 1060, 1061, 1146, 1054):
                logging.error(f"Schema Migration Error: {e}")
        except Exception as e:
            logging.error(f"Schema Migration Error: {e}")


        # 11. Add Master Voucher Column to Bank Book Record
        cursor.execute("SHOW COLUMNS FROM bank_book_recod")
        bbr_cols = [row[0] for row in cursor.fetchall()]
        if 'master_voucher_no' not in bbr_cols:
            logging.info("Migrating: Adding master_voucher_no to bank_book_recod")
            cursor.execute("ALTER TABLE bank_book_recod ADD COLUMN master_voucher_no BIGINT DEFAULT 0")

        cursor.execute("SHOW TABLES LIKE 'sms_delivery_logs'")
        if not cursor.fetchone():
            cursor.execute('''
                CREATE TABLE sms_delivery_logs (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    mobile VARCHAR(20),
                    message TEXT,
                    status VARCHAR(50),
                    api_response TEXT,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''')

        # Default Theme Setting
        cursor.execute("SELECT id FROM system_settings WHERE setting_key = 'system_theme'")
        if not cursor.fetchone():
            cursor.execute("INSERT INTO system_settings (setting_key, setting_value, description) VALUES ('system_theme', 'default', 'Active System Theme')")

        # ── HR / PAYROLL / CRM / EMAIL MODULES ────────────────────────────────
        if not is_migration_applied('hr_payroll_crm_email_v1'):
            try:
                # EMPLOYEES
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS employees (
                        id                INT AUTO_INCREMENT PRIMARY KEY,
                        emp_no            VARCHAR(20)  NOT NULL UNIQUE,
                        first_name        VARCHAR(100) NOT NULL,
                        last_name         VARCHAR(100) NOT NULL DEFAULT '',
                        nic               VARCHAR(20)  DEFAULT '',
                        email             VARCHAR(150) DEFAULT '',
                        mobile            VARCHAR(20)  DEFAULT '',
                        department        VARCHAR(100) DEFAULT '',
                        designation       VARCHAR(100) DEFAULT '',
                        date_of_joining   DATE         DEFAULT NULL,
                        date_of_birth     DATE         DEFAULT NULL,
                        gender            ENUM('Male','Female','Other') DEFAULT 'Male',
                        employment_type   ENUM('Permanent','Contract','Probation','Part-Time') DEFAULT 'Permanent',
                        basic_salary      DECIMAL(12,2) DEFAULT 0.00,
                        epf_no            VARCHAR(50)  DEFAULT '',
                        etf_no            VARCHAR(50)  DEFAULT '',
                        bank_name         VARCHAR(100) DEFAULT '',
                        bank_account_no   VARCHAR(50)  DEFAULT '',
                        bank_branch       VARCHAR(100) DEFAULT '',
                        is_active         TINYINT(1)   DEFAULT 1,
                        created_by        VARCHAR(100) DEFAULT '',
                        created_at        DATETIME     DEFAULT CURRENT_TIMESTAMP
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)

                # PAYROLL COMPONENTS
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS payroll_components (
                        id               INT AUTO_INCREMENT PRIMARY KEY,
                        component_name   VARCHAR(100) NOT NULL,
                        component_type   ENUM('Allowance','Deduction') NOT NULL DEFAULT 'Allowance',
                        is_taxable       TINYINT(1)   DEFAULT 0,
                        is_active        TINYINT(1)   DEFAULT 1,
                        created_at       DATETIME     DEFAULT CURRENT_TIMESTAMP
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)
                # Default components (safe with INSERT IGNORE)
                cursor.execute("""
                    INSERT IGNORE INTO payroll_components (id, component_name, component_type, is_taxable) VALUES
                    (1, 'Basic Salary',       'Allowance', 0),
                    (2, 'Transport Allowance','Allowance', 0),
                    (3, 'Medical Allowance',  'Allowance', 0),
                    (4, 'Overtime',           'Allowance', 1),
                    (5, 'Absence Deduction',  'Deduction', 0),
                    (6, 'Loan Deduction',     'Deduction', 0)
                """)

                # EMPLOYEE SALARY STRUCTURE
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS employee_salary_structure (
                        id            INT AUTO_INCREMENT PRIMARY KEY,
                        employee_id   INT           NOT NULL,
                        component_id  INT           NOT NULL,
                        amount        DECIMAL(12,2) DEFAULT 0.00,
                        effective_from DATE         DEFAULT NULL,
                        FOREIGN KEY (employee_id)  REFERENCES employees(id)           ON DELETE CASCADE,
                        FOREIGN KEY (component_id) REFERENCES payroll_components(id)  ON DELETE CASCADE
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)

                # PAYROLL RUNS
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS payroll_runs (
                        id                  INT AUTO_INCREMENT PRIMARY KEY,
                        pay_period_month    INT           NOT NULL,
                        pay_period_year     INT           NOT NULL,
                        run_date            DATE          DEFAULT NULL,
                        status              ENUM('Draft','Processed','Paid') DEFAULT 'Draft',
                        total_gross         DECIMAL(14,2) DEFAULT 0.00,
                        total_deductions    DECIMAL(14,2) DEFAULT 0.00,
                        total_net           DECIMAL(14,2) DEFAULT 0.00,
                        created_by          VARCHAR(100)  DEFAULT '',
                        created_at          DATETIME      DEFAULT CURRENT_TIMESTAMP,
                        UNIQUE KEY uq_run (pay_period_month, pay_period_year)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)

                # PAYROLL RUN LINES
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS payroll_run_lines (
                        id               INT AUTO_INCREMENT PRIMARY KEY,
                        payroll_run_id   INT           NOT NULL,
                        employee_id      INT           NOT NULL,
                        basic_salary     DECIMAL(12,2) DEFAULT 0.00,
                        total_allowances DECIMAL(12,2) DEFAULT 0.00,
                        total_deductions DECIMAL(12,2) DEFAULT 0.00,
                        epf_employee     DECIMAL(12,2) DEFAULT 0.00,
                        epf_employer     DECIMAL(12,2) DEFAULT 0.00,
                        etf_employer     DECIMAL(12,2) DEFAULT 0.00,
                        gross_salary     DECIMAL(12,2) DEFAULT 0.00,
                        net_salary       DECIMAL(12,2) DEFAULT 0.00,
                        status           ENUM('Pending','Paid') DEFAULT 'Pending',
                        FOREIGN KEY (payroll_run_id) REFERENCES payroll_runs(id) ON DELETE CASCADE,
                        FOREIGN KEY (employee_id)    REFERENCES employees(id)    ON DELETE CASCADE
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)

                # LEAVE TYPES
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS leave_types (
                        id            INT AUTO_INCREMENT PRIMARY KEY,
                        leave_name    VARCHAR(100) NOT NULL,
                        days_per_year INT          DEFAULT 0,
                        is_paid       TINYINT(1)   DEFAULT 1,
                        is_active     TINYINT(1)   DEFAULT 1,
                        created_at    DATETIME     DEFAULT CURRENT_TIMESTAMP
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)
                cursor.execute("""
                    INSERT IGNORE INTO leave_types (id, leave_name, days_per_year, is_paid) VALUES
                    (1, 'Annual Leave',    14, 1),
                    (2, 'Sick Leave',       7, 1),
                    (3, 'Casual Leave',     6, 1),
                    (4, 'Maternity Leave', 84, 1),
                    (5, 'No Pay Leave',     0, 0)
                """)

                # LEAVE APPLICATIONS
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS leave_applications (
                        id               INT AUTO_INCREMENT PRIMARY KEY,
                        employee_id      INT           NOT NULL,
                        leave_type_id    INT           NOT NULL,
                        start_date       DATE          NOT NULL,
                        end_date         DATE          NOT NULL,
                        days_requested   DECIMAL(4,1)  NOT NULL DEFAULT 1,
                        reason           TEXT,
                        status           ENUM('Pending','Approved','Rejected') DEFAULT 'Pending',
                        approved_by      VARCHAR(100)  DEFAULT NULL,
                        approved_at      DATETIME      DEFAULT NULL,
                        created_at       DATETIME      DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (employee_id)   REFERENCES employees(id)   ON DELETE CASCADE,
                        FOREIGN KEY (leave_type_id) REFERENCES leave_types(id) ON DELETE CASCADE
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)

                # CRM LEADS
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS crm_leads (
                        id                   INT AUTO_INCREMENT PRIMARY KEY,
                        lead_name            VARCHAR(200) NOT NULL,
                        company_name         VARCHAR(200) DEFAULT '',
                        email                VARCHAR(150) DEFAULT '',
                        mobile               VARCHAR(20)  DEFAULT '',
                        source               ENUM('Website','Referral','Cold Call','Walk-in','Social Media','Other') DEFAULT 'Other',
                        status               ENUM('New','Contacted','Qualified','Proposal','Won','Lost') DEFAULT 'New',
                        assigned_to          VARCHAR(100) DEFAULT '',
                        expected_value       DECIMAL(14,2) DEFAULT 0.00,
                        expected_close_date  DATE          DEFAULT NULL,
                        notes                TEXT,
                        created_by           VARCHAR(100)  DEFAULT '',
                        created_at           DATETIME      DEFAULT CURRENT_TIMESTAMP,
                        updated_at           DATETIME      DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)

                # CRM ACTIVITIES
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS crm_activities (
                        id             INT AUTO_INCREMENT PRIMARY KEY,
                        lead_id        INT           NOT NULL,
                        activity_type  ENUM('Call','Email','Meeting','Demo','Follow-up','Note') NOT NULL DEFAULT 'Note',
                        subject        VARCHAR(200)  DEFAULT '',
                        notes          TEXT,
                        activity_date  DATETIME      DEFAULT NULL,
                        next_follow_up DATE          DEFAULT NULL,
                        created_by     VARCHAR(100)  DEFAULT '',
                        created_at     DATETIME      DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (lead_id) REFERENCES crm_leads(id) ON DELETE CASCADE
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)

                # EMAIL SETTINGS
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS email_settings (
                        id             INT AUTO_INCREMENT PRIMARY KEY,
                        smtp_host      VARCHAR(200) DEFAULT '',
                        smtp_port      INT          DEFAULT 587,
                        smtp_username  VARCHAR(200) DEFAULT '',
                        smtp_password  VARCHAR(200) DEFAULT '',
                        sender_name    VARCHAR(200) DEFAULT '',
                        sender_email   VARCHAR(200) DEFAULT '',
                        use_tls        TINYINT(1)   DEFAULT 1,
                        is_active      TINYINT(1)   DEFAULT 0,
                        updated_at     DATETIME     DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)

                # EMAIL LOG
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS email_log (
                        id               INT AUTO_INCREMENT PRIMARY KEY,
                        recipient_email  VARCHAR(200) DEFAULT '',
                        subject          VARCHAR(500) DEFAULT '',
                        body             TEXT,
                        status           ENUM('Sent','Failed') DEFAULT 'Sent',
                        error_message    TEXT         DEFAULT NULL,
                        sent_at          DATETIME     DEFAULT CURRENT_TIMESTAMP,
                        related_type     VARCHAR(50)  DEFAULT NULL,
                        related_id       INT          DEFAULT NULL
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)

                record_migration('hr_payroll_crm_email_v1')
                logging.info("Migrated: hr_payroll_crm_email_v1 — HR, Payroll, CRM, Email tables created")

            except Exception as e:
                logging.error(f"Migration hr_payroll_crm_email_v1 error: {e}")
        # ── END HR / PAYROLL / CRM / EMAIL ────────────────────────────────────

        # ── POS TABLES ────────────────────────────────────────────────────────
        if not is_migration_applied('pos_tables_v1'):
            try:
                # Invoice number sequence table
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS pos_invoice_no (
                        Id     INT AUTO_INCREMENT PRIMARY KEY,
                        IV_No  VARCHAR(50) DEFAULT ''
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)

                # POS Sales table
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS POS_Sales_Invoice_01 (
                        Id                              INT AUTO_INCREMENT PRIMARY KEY,
                        ItemCoude                       VARCHAR(100),
                        ItemName                        VARCHAR(255),
                        ItemMesurmet                    VARCHAR(50),
                        SllingPrice                     DECIMAL(15,4) DEFAULT 0,
                        ItemPriceComen                  DECIMAL(15,4) DEFAULT 0,
                        ItemLoyalityPrice               DECIMAL(15,4) DEFAULT 0,
                        Sales_with_market_price_Active  TINYINT(1)    DEFAULT 0,
                        Sales_with_Special_price_Active TINYINT(1)    DEFAULT 0,
                        Loyalty_Price_Active            TINYINT(1)    DEFAULT 0,
                        RecodeUserId                    INT           DEFAULT 0,
                        Location                        VARCHAR(200),
                        AcctionDate                     DATE,
                        QuntirySale                     DECIMAL(15,4) DEFAULT 0,
                        InventoryCost                   DECIMAL(15,4) DEFAULT 0,
                        PaymentMethord                  INT           DEFAULT 1,
                        CashAccountName                 VARCHAR(255),
                        BankAccountName                 VARCHAR(255),
                        Invoice_No                      VARCHAR(50),
                        Loyalty_No                      VARCHAR(100)  DEFAULT '0',
                        Total_Value                     DECIMAL(15,4) DEFAULT 0,
                        jv                              INT           DEFAULT 0,
                        Revers                          TINYINT(1)    DEFAULT 0,
                        INDEX idx_pos_jv   (jv),
                        INDEX idx_pos_inv  (Invoice_No),
                        INDEX idx_pos_date (AcctionDate)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)

                record_migration('pos_tables_v1')
                logging.info("Migrated: pos_tables_v1 — pos_invoice_no + POS_Sales_Invoice_01 created")
            except Exception as e:
                logging.error(f"Migration pos_tables_v1 error: {e}")
        # ── END POS TABLES ────────────────────────────────────────────────────

        # ── DOCUMENT UPLOAD SYSTEM ────────────────────────────────────────────
        if not is_migration_applied('document_upload_v1'):
            try:
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS documents (
                        id            INT AUTO_INCREMENT PRIMARY KEY,
                        related_type  VARCHAR(50)  NOT NULL,
                        related_id    VARCHAR(200) NOT NULL,
                        file_name     VARCHAR(255) NOT NULL,
                        stored_name   VARCHAR(255) NOT NULL,
                        file_size     INT          DEFAULT 0,
                        file_type     VARCHAR(50)  DEFAULT '',
                        notes         TEXT         DEFAULT NULL,
                        uploaded_by   VARCHAR(100) DEFAULT '',
                        uploaded_at   DATETIME     DEFAULT CURRENT_TIMESTAMP,
                        INDEX idx_doc_related (related_type, related_id)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)
                record_migration('document_upload_v1')
                logging.info("Migrated: document_upload_v1 — documents table created")
            except Exception as e:
                logging.error(f"Migration document_upload_v1 error: {e}")
        # ── END DOCUMENT UPLOAD ───────────────────────────────────────────────

        conn.commit()
        cursor.close()
        conn.close()
    except Exception as e:
        logging.error(f"Schema Migration Error: {e}")

def ensure_default_accounts(target_db=None):
    """Ensures essential General Ledger accounts exist."""
    current_db = target_db if target_db else db
    try:
        defaults = [
            # Name, BS Position, BS Category, P&L Position, P&L Category, Type
            ('Account Payable', 6, 'Current liabilities', None, None, 'liabilities'),
            ('Account Receivable', 3, 'Current assets', None, None, 'assets'),
            ('Cost Of Goods Sold', None, None, 2, 'Cost Of Sales', 'expenses'),
            ('Sales', None, None, 1, 'Revenue', 'income'),
            ('Income', None, None, 1, 'Revenue', 'income'),
            ('Inventory', 3, 'Current assets', None, None, 'assets'),
            ('VAT Control', 6, 'Current liabilities', None, None, 'liabilities'),
            ('Cash In Hand', 3, 'Current assets', None, None, 'assets')
        ]

        current_user = 0 # System

        if not defaults:
            return

        # Extract all account names
        account_names = [acc[0] for acc in defaults]

        # Check existing accounts using a single batch query
        format_strings = ','.join(['%s'] * len(account_names))
        query = f"SELECT account_name FROM new_account_table WHERE account_name IN ({format_strings})"

        existing_rows = current_db.execute_query(query, tuple(account_names))

        # Store existing account names in a set for O(1) lookups
        existing_names = {row['account_name'] for row in (existing_rows or [])}

        for acc in defaults:
            name, bs_pos, bs_cat, pl_pos, pl_cat, acc_type = acc

            # Check against the set instead of making a DB query
            if name not in existing_names:
                logging.info(f"Creating default account: {name}")
                # Determine basement
                basement = 'DR' if acc_type in ['expenses', 'assets'] else 'CR'

                query = """
                    INSERT INTO new_account_table (
                        account_name, account_hold_possion_Balace_Sheet, account_name_of_catogory_Balace_sheet,
                        account_hold_possion_PL, account_name_of_catogory_PL,
                        account_income, account_expenses, account_assets, account_liabilities, account_equity,
                        accont_create_date, account_create_user, account_active, account_basment
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 1, %s)
                """
                current_db.execute_query(query, (
                    name, bs_pos, bs_cat, pl_pos, pl_cat,
                    1 if acc_type=='income' else 0, 1 if acc_type=='expenses' else 0,
                    1 if acc_type=='assets' else 0, 1 if acc_type=='liabilities' else 0, 0,
                    date.today(), current_user, basement
                ), commit=True)

        # 2. Add POS SALE sub-account under Income
        sub_query = "SELECT id_sub FROM sub_accont_for_new_account WHERE sub_sub_accaount_name = 'POS SALE' AND sub_new_account = 'Income'"
        if not current_db.execute_query(sub_query):
            current_db.execute_query("""
                INSERT INTO sub_accont_for_new_account (sub_sub_accaount_name, sub_new_account, creat_user, creat_date, active, sub_account_code)
                VALUES ('POS SALE', 'Income', %s, %s, 1, 0)
            """, (current_user, date.today()), commit=True)

        # 3. Add Common customer
        cust_query = "SELECT id FROM customer WHERE customer_code = '60001'"
        if not current_db.execute_query(cust_query):
            current_db.execute_query("""
                INSERT INTO customer (
                    customer_name, customer_code, customer_Billing_Address,
                    costomer_Delivery_Address, e_mail, coustomer_credit_limit
                ) VALUES ('Common customer', '60001', 'non', 'non', 'non', 0)
            """, commit=True)

        # 4. Add Direct Payment supplier
        sup_query = "SELECT sup_id FROM suppliers WHERE supplier_code = '70001'"
        if not db.execute_query(sup_query):
            db.execute_query("""
                INSERT INTO suppliers (supplier_name, supplier_code)
                VALUES ('Direct Payment', '70001')
            """, commit=True)

    except Exception as e:
        logging.error(f"Error ensuring default accounts/entities: {e}")

# --- Quotation Evaluation ---
@app.route('/quotation_evaluation', methods=['GET'])
@login_required
@has_permission('Access_Inventory')
def quotation_evaluation():
    return render_template('quotation_evaluation.html')

@app.route('/api/evaluate_quotations', methods=['POST'])
@login_required
def evaluate_quotations():
    data = request.json
    constraints = data.get('constraints', {})
    suppliers = data.get('suppliers', [])

    # Constraints
    max_days = float(constraints.get('max_days') or 9999)
    # Weights (Simple priority: Price, Speed, Quality)
    priority = constraints.get('priority', 'price')

    # 1. Filter Step
    filtered = []
    for s in suppliers:
        s_days = float(s.get('days', 0))
        if s_days <= max_days:
            filtered.append(s)

    if not filtered:
        return {'results': [], 'message': 'No suppliers meet the delivery deadline constraint.'}

    # 2. Scoring Step
    # Normalize values for scoring
    # Price: Lower is better
    # Days: Lower is better
    # Quality: Higher is better (1-5)

    # Avoid division by zero
    min_price = min(float(s['price']) for s in filtered) if filtered else 1
    min_days = min(float(s['days']) for s in filtered) if filtered else 1
    max_qual = max(float(s['quality']) for s in filtered) if filtered else 1

    scored = []
    for s in filtered:
        price = float(s['price'])
        days = float(s['days'])
        qual = float(s['quality'])

        # Calculate Scores (0 to 1 scale, 1 is best)
        # Price Score: (Min / Actual)
        s_price = min_price / price if price > 0 else 0
        # Days Score: (Min / Actual)
        s_days = min_days / days if days > 0 else 0
        # Quality Score: (Actual / Max)
        s_qual = qual / max_qual if max_qual > 0 else 0

        # Define Weights and Reasons based on Priority
        # Format: (Price Weight, Days Weight, Quality Weight, Reason)
        # s_price (min/actual), s_days (min/actual), s_qual (actual/max)
        strategies = {
            'speed':   (0.2, 0.6, 0.2, "Fastest delivery within constraints"),
            'quality': (0.2, 0.2, 0.6, "Best quality rating"),
            'price':   (0.6, 0.2, 0.2, "Best price")
        }

        # Get weights or default to 'price'
        w_price, w_days, w_qual, reason = strategies.get(priority, strategies['price'])

        # Calculate Score
        # Note: s_price and s_days logic was: Best/Actual (so higher is better, max 1)
        # s_qual logic was: Actual/Best (so higher is better, max 1)
        # However, the previous logic for s_qual in quality priority used s_qual * 0.6.
        # But for price/speed priority it also used s_qual * 0.2.
        # The key difference was which variable got the 0.6 weight.

        final_score = (s_price * w_price) + (s_days * w_days) + (s_qual * w_qual)

        s['score'] = round(final_score * 100, 1)
        scored.append(s)

    # Sort by Score Descending
    scored.sort(key=lambda x: x['score'], reverse=True)

    # Mark the winner
    if scored:
        scored[0]['is_winner'] = True
        scored[0]['win_reason'] = reason

    return {'results': scored}

# --- Proforma Invoice ---
@app.route('/proforma_invoice', methods=['GET', 'POST'])
@login_required
@has_permission('Access_Accounting') # or Sales
def proforma_invoice():
    if request.method == 'POST':
        cust_name = request.form.get('customer_name')
        pi_date = request.form.get('pi_date')
        exp_date = request.form.get('expiry_date')
        narration = request.form.get('narration')
        items_json = request.form.get('items_json')

        items = json.loads(items_json) if items_json else []
        if not items:
            flash('No items', 'danger')
            return redirect(url_for('proforma_invoice'))

        current_user = get_current_user_id()
        current_user_pk = get_current_user_pk()

        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            conn.start_transaction()

            # Generate PI Number
            # Simple timestamp based or sequence
            pi_no = f"PI-{datetime.now().strftime('%Y%m%d%H%M%S')}"

            # Calc Totals
            subtotal = sum(float(i['total']) for i in items)
            vat_rate = 0 # Can add VAT logic if needed, simple for now
            vat_amount = 0
            grand_total = subtotal + vat_amount

            # Insert Header
            cursor.execute("""
                INSERT INTO proforma_invoice_header (
                    pi_number, customer_name, pi_date, expiry_date,
                    subtotal, vat_amount, grand_total, narration, created_by
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (pi_no, cust_name, pi_date, exp_date, subtotal, vat_amount, grand_total, narration, current_user_pk))
            pi_id = cursor.lastrowid

            # Insert Details
            if items:
                cursor.executemany("""
                    INSERT INTO proforma_invoice_details (
                        pi_id, item_name, description, qty, unit_price, total
                    ) VALUES (%s, %s, %s, %s, %s, %s)
                """, [(pi_id, i['name'], i.get('desc', ''), i['qty'], i['price'], i['total']) for i in items])

            conn.commit()
            flash(f'Proforma Invoice {pi_no} created', 'success')
            return redirect(url_for('print_proforma', pi_id=pi_id))

        except Exception as e:
            conn.rollback()
            flash(f'Error: {str(e)}', 'danger')

        return redirect(url_for('proforma_invoice'))

    customers = db.execute_query("SELECT customer_name FROM customer")
    items = db.execute_query("SELECT inventoy_name, inventoy_items_messurment_unit, inventory_price_selling FROM inventoy_items LEFT JOIN inventory_price_recod ON inventoy_items.id = inventory_price_link")

    return render_template('proforma_invoice.html',
                           customers=customers,
                           items=items,
                           today_date=date.today().strftime('%Y-%m-%d'))

@app.route('/proforma/print/<int:pi_id>')
@login_required
def print_proforma(pi_id):
    header_res = db.execute_query("SELECT * FROM proforma_invoice_header WHERE id = %s", (pi_id,))
    if not header_res: return "Not Found", 404
    header = header_res[0]

    details = db.execute_query("SELECT * FROM proforma_invoice_details WHERE pi_id = %s", (pi_id,))
    company = db.execute_query("SELECT * FROM company LIMIT 1")[0]

    return render_template('proforma_print.html', header=header, details=details, company=company)

# --- Currency Setup ---
@app.route('/currency_setup', methods=['GET', 'POST'])
@login_required
@has_permission('Access_Accounting')
def currency_setup():
    if request.method == 'POST':
        code = request.form.get('code', '').upper()
        name = request.form.get('name')
        is_base = 1 if request.form.get('is_base') else 0

        if not code:
            flash('Code required', 'danger')
            return redirect(url_for('currency_setup'))

        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            conn.start_transaction()

            # If setting base, unset others
            if is_base:
                cursor.execute("UPDATE currency_table SET is_base_currency = 0")

            cursor.execute("INSERT INTO currency_table (currency_code, currency_name, is_base_currency) VALUES (%s, %s, %s)",
                           (code, name, is_base))

            conn.commit()
            cursor.close()
            conn.close()
            flash('Currency added', 'success')
        except Exception as e:
            flash(f'Error: {str(e)}', 'danger')

        return redirect(url_for('currency_setup'))

    currencies = db.execute_query("SELECT * FROM currency_table")
    return render_template('currency_setup.html', currencies=currencies)

@app.route('/currency_setup/delete', methods=['POST'])
@login_required
@has_permission('Access_Accounting')
def delete_currency():
    cid = request.form.get('id')
    try:
        db.execute_query("DELETE FROM currency_table WHERE id=%s", (cid,), commit=True)
        flash('Currency deleted', 'success')
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
    return redirect(url_for('currency_setup'))

# --- Exchange Rate API ---
@app.route('/api/get_exchange_rate')
@login_required
def get_exchange_rate():
    from_curr = request.args.get('from', '').upper()
    to_curr = request.args.get('to', '').upper()

    if not from_curr or not to_curr:
        return {'error': 'Missing currencies'}, 400

    if from_curr == to_curr:
        return {'rate': 1.0}

    # Check cache first
    cache_key = f"{from_curr}_{to_curr}"
    current_time = time.time()

    if cache_key in exchange_rate_cache:
        cached_data = exchange_rate_cache[cache_key]
        if current_time - cached_data['timestamp'] < CACHE_DURATION:
            return {'rate': cached_data['rate']}

    try:
        # Use urllib instead of requests to avoid external dependency if not installed
        url = f"https://api.exchangerate-api.com/v4/latest/{from_curr}"
        with urllib.request.urlopen(url) as response:
            data = json.loads(response.read().decode())
            rate = data.get('rates', {}).get(to_curr)

            if rate is not None:
                # Update cache
                exchange_rate_cache[cache_key] = {
                    'rate': float(rate),
                    'timestamp': current_time
                }

                # Also cache the reverse if possible (1/rate)
                exchange_rate_cache[f"{to_curr}_{from_curr}"] = {
                    'rate': 1.0 / float(rate),
                    'timestamp': current_time
                }

                return {'rate': float(rate)}

    except Exception as e:
        print(f"Exchange Rate API Error: {e}")
        # Fallback to hardcoded/mock values on error (e.g. no internet)
        pass

    # Fallback / Mocking Logic
    rate = 1.0
    if from_curr == 'USD' and to_curr == 'LKR':
        rate = 300.0 + random.uniform(-5, 5) # Fluctuation
    elif from_curr == 'LKR' and to_curr == 'USD':
        rate = 1 / 300.0
    elif from_curr == 'EUR' and to_curr == 'LKR':
        rate = 330.0

    return {'rate': rate}

# --- Journal Entry Management ---
@app.route('/journal_entry', methods=['GET'])
@login_required
@has_permission('Access_Accounting')
def journal_entry():
    # Fetch accounts with currency info and type classification
    accounts = db.execute_query("""
        SELECT
            account_name,
            currency_code,
            CASE
                WHEN account_income = 1 OR account_expenses = 1 THEN 'P&L Account'
                WHEN account_assets = 1 OR account_liabilities = 1 OR account_equity = 1 THEN 'BS Account'
                ELSE 'Other'
            END as account_type
        FROM new_account_table
        WHERE account_active = 1
    """)

    sub_accounts = db.execute_query("SELECT sub_account_code, sub_sub_accaount_name FROM sub_accont_for_new_account WHERE active = 1")
    jobs = db.execute_query("SELECT job_number FROM jobs_unit")

    # Auto-generate next system JV no if possible (simplified max + 1)
    jv_res = db.execute_query("SELECT MAX(jv_id) as max_id FROM jv_numbers")
    next_sys_jv = (jv_res[0]['max_id'] if jv_res and jv_res[0]['max_id'] else 0) + 1

    # Get Base Currency
    base_curr_res = db.execute_query("SELECT currency_code FROM currency_table WHERE is_base_currency = 1 LIMIT 1")
    base_currency = base_curr_res[0]['currency_code'] if base_curr_res else 'LKR'

    return render_template('journal_entry.html',
                           accounts=accounts,
                           sub_accounts=sub_accounts,
                           jobs=jobs,
                           next_sys_jv=next_sys_jv,
                           base_currency=base_currency,
                           today_date=date.today().strftime('%Y-%m-%d'),
                           last_jv=request.args.get('last_jv'))


@app.route('/api/journal_accounts')
@login_required
def api_journal_accounts():
    """Live account list for the journal-entry datalist so a newly created
    account appears without a full page refresh."""
    accounts = db.execute_query("""
        SELECT account_name, currency_code,
               CASE
                   WHEN account_income = 1 OR account_expenses = 1 THEN 'P&L Account'
                   WHEN account_assets = 1 OR account_liabilities = 1 OR account_equity = 1 THEN 'BS Account'
                   ELSE 'Other'
               END as account_type
        FROM new_account_table
        WHERE account_active = 1
        ORDER BY account_name
    """) or []
    return jsonify([{'name': a['account_name'], 'type': a['account_type'],
                     'currency': a['currency_code'] or ''} for a in accounts])

@app.route('/journal_entry/save', methods=['POST'])
@login_required
@has_permission('Access_Accounting')
def save_journal_entry():
    try:
        user_code = request.form.get('jv_user_code')
        entry_date = request.form.get('entry_date')
        main_narration = request.form.get('main_narration')
        entries_json = request.form.get('entries_json')

        entries = json.loads(entries_json) if entries_json else []

        if not entries:
            flash('No entries provided', 'danger')
            return redirect(url_for('journal_entry'))

        if not user_code or not main_narration:
            flash('JV Number and Main Narration are required', 'danger')
            return redirect(url_for('journal_entry'))

        # Verify balance again (Base Currency)
        total_dr = sum(parse_float(e['dr']) for e in entries)
        total_cr = sum(parse_float(e['cr']) for e in entries)

        if abs(total_dr - total_cr) > 0.01:
            flash(f'Entries not balanced. Diff: {total_dr - total_cr}', 'danger')
            return redirect(url_for('journal_entry'))

        current_user = get_current_user_id()
        current_user_pk = get_current_user_pk()

        conn = db.get_connection()
        cursor = conn.cursor()
        conn.start_transaction()

        try:
            # Check Workflow
            cursor.execute("SELECT setting_value FROM system_settings WHERE setting_key = 'enable_approval_workflow'")
            res_set = cursor.fetchone()
            workflow_enabled = res_set and res_set[0] == '1'
            status = 0 if workflow_enabled else 1

            # 1. Create JV Header
            cursor.execute("INSERT INTO jv_numbers (jv_user_code, jv_naration, status) VALUES (%s, %s, %s)",
                           (user_code, main_narration, status))
            jv_no = cursor.lastrowid

            # 2. Insert Entries
            for e in entries:
                # Handle sub account
                sub_code = 0
                if e.get('sub_account'):
                    # Format "Code - Name" -> split
                    parts = e['sub_account'].split(' - ')
                    if parts: sub_code = parts[0]

                # Handle Job No
                job_no = e.get('job_no') if e.get('job_no') else None

                # Currency Info
                curr_code = e.get('currency', 'LKR')
                fc_amt = parse_float(e.get('fc_amount', 0))
                rate = parse_float(e.get('rate', 1))

                cursor.execute("""
                    INSERT INTO entry_details (
                        account_name, enty_values_DR, enty_values_CR,
                        entry_effective_date, entry_create_date, entry_naration,
                        entry_create_user, entry_jv, entry_sub_account_code, entry_job_number,
                        currency_code, fc_amount, exchange_rate
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    e['account'], e['dr'], e['cr'],
                    entry_date, datetime.now().date(), e['narration'],
                    current_user, jv_no, sub_code, job_no,
                    curr_code, fc_amt, rate
                ))

            conn.commit()
            flash(f'Journal Entry created successfully. System JV: {jv_no}', 'success')
            _je_jv = jv_no

        except Exception as ex:
            conn.rollback()
            flash(f'Database Error: {str(ex)}', 'danger')
            _je_jv = None
        finally:
            cursor.close()
            conn.close()

        if _je_jv:
            return redirect(url_for('journal_entry', last_jv=_je_jv))

    except Exception as e:
        flash(f'System Error: {str(e)}', 'danger')

    return redirect(url_for('journal_entry'))

def _journal_entry_history_filters():
    """Read + validate the History filter querystring shared by the JSON
    endpoint and the Excel export, so both always agree on what's included."""
    return {
        'from_date':  request.args.get('from', '').strip(),
        'to_date':    request.args.get('to', '').strip(),
        'search':     request.args.get('search', '').strip(),
        'jv_q':       request.args.get('jv', '').strip(),
        'amount_min': request.args.get('amount_min', '').strip(),
        'amount_max': request.args.get('amount_max', '').strip(),
        'dr_cr':      request.args.get('dr_cr', '').strip(),
    }


def _journal_entry_history_rows(from_date, to_date, search, jv_q, amount_min, amount_max, dr_cr):
    """Shared query behind both the History JSON endpoint and its Excel
    export — keeps their filtering and scoping identical.

    Scopes history to JVs actually submitted from THIS window (manual
    Journal Entry). jv_user_code here is a free-text code the user types in
    the form, but every other module stamps its own fixed/prefixed code
    (or the numeric user id) when it posts a JV — so excluding those known
    patterns is what's left is genuinely manual journal entries, instead of
    this history dumping every JV in the whole system (GRN, Service Entry,
    Payments, POS, reversals, opening balances, etc.).
    NOTE: earlier this excluded purely-numeric jv_user_code values too (some
    other modules stamp the current user's numeric id as the code), but that
    risked hiding genuine manual entries if this account's own JV-code habit
    is also plain numbers — pulled back out to only the unambiguous,
    exact/prefix system codes below.

    Returns (data, total_dr, total_cr, has_filter) — has_filter is False
    when no filter was supplied at all, meaning callers should refuse to
    dump the whole ledger unbounded.
    """
    if not (from_date or to_date or search or jv_q or amount_min or amount_max or dr_cr):
        return [], 0.0, 0.0, False

    filters = [
        'ed.entry_deleted = 0',
        # NOTE: reversal entries (jv_user_code LIKE 'REV-JV-%') are deliberately
        # NOT excluded here — they're created from this same screen's "Reverse"
        # button, and the row template already renders them with a "Reversal"
        # badge (see is_reversal below). Excluding them used to make a
        # successful reversal look like it silently did nothing, since the
        # only visible trace of it (the new REV-JV- entry) never showed up
        # in this exact list.
        "jv.jv_user_code NOT IN ('JV FROM GRN','JV FORM SEN INVOICE','JV FROM PAYMENT',"
        "'JV FROM DIRECT CASH','JV FROM RECEIPT','JV FROM POS','JV FROM PRODUCTION ISSUE',"
        "'JV FROM PRODUCTION RECEIPT','JV FROM INVENTORY ISSUE','OB-UPLOAD')",
        "jv.jv_user_code NOT LIKE 'FA-PUR-%'",
        "jv.jv_user_code NOT LIKE 'FA-WO-%'",
        "jv.jv_user_code NOT LIKE 'DEP-%'",
        "jv.jv_user_code NOT LIKE 'CN-%'",
        "jv.jv_user_code NOT LIKE 'DN-%'",
    ]
    params = []
    if from_date:
        filters.append('ed.entry_effective_date >= %s'); params.append(from_date)
    if to_date:
        filters.append('ed.entry_effective_date <= %s'); params.append(to_date)
    if search:
        filters.append('(ed.account_name LIKE %s OR ed.entry_naration LIKE %s)')
        params += [f'%{search}%', f'%{search}%']
    if jv_q:
        filters.append('(ed.entry_jv = %s OR jv.jv_user_code LIKE %s)')
        params += [jv_q, f'%{jv_q}%']
    if amount_min:
        filters.append('(ed.enty_values_DR >= %s OR ed.enty_values_CR >= %s)')
        params += [amount_min, amount_min]
    if amount_max:
        filters.append('GREATEST(COALESCE(ed.enty_values_DR,0), COALESCE(ed.enty_values_CR,0)) <= %s')
        params.append(amount_max)
    if dr_cr == 'dr':
        filters.append('ed.enty_values_DR > 0')
    elif dr_cr == 'cr':
        filters.append('ed.enty_values_CR > 0')

    where = ' AND '.join(filters)
    query = f"""
        SELECT
            ed.entry_jv, ed.entry_effective_date, ed.account_name,
            ed.entry_naration, ed.enty_values_DR, ed.enty_values_CR,
            jv.jv_user_code
        FROM entry_details ed
        JOIN jv_numbers jv ON ed.entry_jv = jv.jv_id
        WHERE {where}
        ORDER BY ed.entry_jv DESC, ed.id ASC
        LIMIT 1000
    """
    rows = db.execute_query(query, tuple(params)) or []

    data = []
    for r in rows:
        data.append({
            'jv_no': r['entry_jv'],
            'date': str(r['entry_effective_date']),
            'account': r['account_name'],
            'narration': r['entry_naration'],
            'dr': float(r['enty_values_DR'] or 0),
            'cr': float(r['enty_values_CR'] or 0),
            'is_reversal': str(r.get('jv_user_code') or '').startswith('REV-JV-')
        })

    total_dr = sum(d['dr'] for d in data)
    total_cr = sum(d['cr'] for d in data)
    return data, total_dr, total_cr, True


@app.route('/journal_entry/history')
@login_required
def journal_entry_history():
    f = _journal_entry_history_filters()
    data, _total_dr, _total_cr, has_filter = _journal_entry_history_rows(
        f['from_date'], f['to_date'], f['search'], f['jv_q'],
        f['amount_min'], f['amount_max'], f['dr_cr'])
    if not has_filter:
        return json.dumps([])
    return json.dumps(data)


@app.route('/journal_entry/export_xlsx')
@login_required
@has_permission('Access_Reports')
def journal_entry_history_export_xlsx():
    """Styled Excel export of the Journal Entry History grid, honoring
    whatever Search / JV / Date / Amount / Type filters are currently
    applied — same filtering as the on-screen History list."""
    f = _journal_entry_history_filters()

    try:
        import excel_export as xl
    except ImportError:
        flash("Excel export needs the 'openpyxl' package on the server — run: pip install openpyxl", 'warning')
        return redirect(url_for('journal_entry'))

    data, total_dr, total_cr, has_filter = _journal_entry_history_rows(
        f['from_date'], f['to_date'], f['search'], f['jv_q'],
        f['amount_min'], f['amount_max'], f['dr_cr'])
    if not has_filter:
        flash('Set at least one History filter (date range, search, JV, or amount) before exporting.', 'warning')
        return redirect(url_for('journal_entry'))

    subtitle_bits = []
    if f['from_date'] or f['to_date']:
        subtitle_bits.append(f"{f['from_date'] or '…'} to {f['to_date'] or '…'}")
    if f['search']: subtitle_bits.append(f"Search: {f['search']}")
    if f['jv_q']: subtitle_bits.append(f"JV/Code: {f['jv_q']}")
    if f['amount_min']: subtitle_bits.append(f"Min: {f['amount_min']}")
    if f['amount_max']: subtitle_bits.append(f"Max: {f['amount_max']}")
    if f['dr_cr'] == 'dr': subtitle_bits.append('Debit only')
    elif f['dr_cr'] == 'cr': subtitle_bits.append('Credit only')
    subtitle = ' · '.join(subtitle_bits) if subtitle_bits else 'Manual Journal Entries'

    wb, ws = xl.new_workbook('JV History')
    ncols = 6
    row = xl.title_block(ws, ncols, _company_display_name(), 'JOURNAL ENTRY HISTORY', subtitle)
    row = xl.header_row(ws, row, ['JV No', 'Date', 'Account', 'Narration', 'Debit', 'Credit'])
    for r in data:
        row = xl.data_row(ws, row,
                          [r['jv_no'], r['date'], r['account'] or '-',
                           ('Reversal — ' + (r['narration'] or '')) if r['is_reversal'] else (r['narration'] or ''),
                           r['dr'], r['cr']],
                          num_cols=(5, 6))
    row = xl.total_row(ws, row, f'TOTAL — {len(data)} line(s)', [0, 0, 0, total_dr, total_cr],
                       bg='EAF6EA', color=xl.GREEN_DARK, size=11)
    for col in (2, 3, 4):
        ws.cell(row=row - 1, column=col).value = None

    xl.finish(ws, ncols, first_col_width=12, num_col_width=16)
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 34

    fname = f'JournalEntryHistory_{date.today().strftime("%Y%m%d")}.xlsx'
    return xl.workbook_response(wb, fname)


@app.route('/journal_entry/view/<int:jv_no>')
@login_required
def journal_entry_view(jv_no):
    """Quick on-screen detail of one JV's lines — a lightweight alternative
    to the formatted Print page for just checking what's in an entry."""
    header_res = db.execute_query("""
        SELECT j.jv_id, j.jv_user_code, j.jv_naration, MIN(e.entry_effective_date) as entry_date
        FROM jv_numbers j
        LEFT JOIN entry_details e ON j.jv_id = e.entry_jv
        WHERE j.jv_id = %s
        GROUP BY j.jv_id
    """, (jv_no,))
    if not header_res:
        return jsonify({'error': 'JV not found'}), 404
    header = header_res[0]

    details = db.execute_query("""
        SELECT id, account_name, entry_naration, enty_values_DR, enty_values_CR, entry_job_number
        FROM entry_details
        WHERE entry_jv = %s AND COALESCE(entry_deleted, 0) = 0
        ORDER BY id
    """, (jv_no,)) or []

    total_dr = sum(float(d['enty_values_DR'] or 0) for d in details)
    total_cr = sum(float(d['enty_values_CR'] or 0) for d in details)
    is_reversal = str(header['jv_user_code'] or '').startswith('REV-JV-')

    return jsonify({
        'jv_no': header['jv_id'],
        'user_code': header['jv_user_code'] or '',
        'narration': header['jv_naration'] or '',
        'date': str(header['entry_date']) if header['entry_date'] else '',
        'total_dr': total_dr,
        'total_cr': total_cr,
        'is_reversal': is_reversal,
        'lines': [{
            'id': d['id'],
            'account': d['account_name'],
            'narration': d['entry_naration'] or '',
            'dr': float(d['enty_values_DR'] or 0),
            'cr': float(d['enty_values_CR'] or 0),
            'job': d['entry_job_number'] or '',
        } for d in details],
    })


@app.route('/journal_entry/edit_amounts', methods=['POST'])
@login_required
@has_permission('Access_Accounting')
def journal_entry_edit_amounts():
    """Edit the Debit/Credit amounts on an existing manual Journal Entry's
    lines. Deliberately narrow in scope: the set of lines and their accounts
    stay exactly as they were — only each line's DR/CR number can change —
    and the entry must still balance (total DR == total CR) before the save
    is accepted. Reversal entries and entries in a closed period are
    blocked, same as the existing Reverse action."""
    jv_no = (request.form.get('jv_no') or '').strip()
    lines_json = request.form.get('lines_json')
    if not jv_no or not lines_json:
        return jsonify({'success': False, 'error': 'Missing JV or line data'}), 400

    try:
        lines = json.loads(lines_json)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'Invalid line data'}), 400
    if not lines:
        return jsonify({'success': False, 'error': 'No lines to update'}), 400

    if jv_in_closed_period(jv_no):
        return jsonify({'success': False,
                        'error': f'This entry is in a closed period (locked through {get_period_lock_date()}).'}), 403

    jvrow = db.execute_query("SELECT jv_user_code FROM jv_numbers WHERE jv_id = %s", (jv_no,))
    if not jvrow:
        return jsonify({'success': False, 'error': 'JV not found'}), 404
    if str(jvrow[0]['jv_user_code'] or '').startswith('REV-JV-'):
        return jsonify({'success': False, 'error': 'Reversal entries cannot be edited.'}), 400

    total_dr = 0.0
    total_cr = 0.0
    updates = []
    for ln in lines:
        try:
            line_id = int(ln.get('id'))
            dr = round(float(ln.get('dr') or 0), 2)
            cr = round(float(ln.get('cr') or 0), 2)
        except (TypeError, ValueError, AttributeError):
            return jsonify({'success': False, 'error': 'Invalid amount on one of the lines'}), 400
        if dr < 0 or cr < 0:
            return jsonify({'success': False, 'error': 'Amounts cannot be negative'}), 400
        if dr > 0 and cr > 0:
            return jsonify({'success': False, 'error': 'A line cannot have both a Debit and a Credit amount'}), 400
        total_dr += dr
        total_cr += cr
        updates.append((dr, cr, line_id))

    total_dr = round(total_dr, 2)
    total_cr = round(total_cr, 2)
    if total_dr != total_cr:
        return jsonify({'success': False,
                        'error': f'Entry is not balanced — Debit {total_dr:,.2f} vs Credit {total_cr:,.2f}.'}), 400
    if total_dr == 0:
        return jsonify({'success': False, 'error': 'Total amount cannot be zero'}), 400

    conn = None
    cursor = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        conn.start_transaction()
        for dr, cr, line_id in updates:
            cursor.execute(
                "UPDATE entry_details SET enty_values_DR=%s, enty_values_CR=%s "
                "WHERE id=%s AND entry_jv=%s AND COALESCE(entry_deleted,0)=0",
                (dr, cr, line_id, jv_no))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if cursor: cursor.close()
        if conn: conn.close()


@app.route('/journal_entry/reverse', methods=['POST'])
@login_required
@has_permission('Access_Accounting') # or Access_Reversals
def reverse_journal_entry():
    jv_no = request.form.get('jv_no')
    if not jv_no: return {'error': 'JV No required'}, 400

    if jv_in_closed_period(jv_no):
        return {'error': f'This entry is in a closed period (locked through {get_period_lock_date()}).'}, 403

    current_user_pk = get_current_user_pk()
    today = date.today()

    conn = None
    cursor = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        # Guard: a reversal JV can never itself be reversed — that re-creates the
        # original and allows endless back-and-forth duplication
        cursor.execute("SELECT jv_user_code FROM jv_numbers WHERE jv_id = %s", (jv_no,))
        row = cursor.fetchone()
        if row and str(row[0] or '').startswith('REV-JV-'):
            return {'error': 'This is a reversal entry — it cannot be reversed again. '
                             'Enter a new journal instead.'}, 400

        # Guard: check if already reversed (reversal JV exists with REV-JV- prefix)
        cursor.execute(
            "SELECT COUNT(*) FROM jv_numbers WHERE jv_user_code = %s",
            (f'REV-JV-{jv_no}',))
        if cursor.fetchone()[0] > 0:
            return {'error': 'This JV has already been reversed.'}, 400

        # Also check entry_deleted flag — a JV is only truly "deleted" once
        # EVERY one of its lines is. The old check picked one arbitrary row
        # (no ORDER BY) and looked only at that row's flag, so a multi-line
        # JV (e.g. one with a sub-account line) could get wrongly blocked
        # whenever that one row happened to be the one already flagged,
        # even though the rest of the entry was still fully active.
        cursor.execute(
            "SELECT COUNT(*) FROM entry_details WHERE entry_jv = %s AND COALESCE(entry_deleted, 0) = 0",
            (jv_no,))
        active_count = cursor.fetchone()[0]
        if active_count == 0:
            return {'error': 'Already reversed or deleted'}, 400

        conn.start_transaction()

        # Create reversal JV
        cursor.execute(
            "INSERT INTO jv_numbers (jv_user_code, jv_naration, status) VALUES (%s, %s, 1)",
            (f'REV-JV-{jv_no}', f'Journal Reversal of JV-{jv_no}'))
        rev_jv = cursor.lastrowid

        # Reverse GL entries (swap DR ↔ CR) on the ORIGINAL effective dates so
        # period reports (P&L / Balance Sheet) are reversed too — dating the
        # reversal "today" left old periods still showing the original amounts
        cursor.execute("""
            INSERT INTO entry_details (
                account_name, enty_values_DR, enty_values_CR,
                entry_effective_date, entry_create_date,
                entry_naration, entry_create_user, entry_jv,
                entry_sub_account_code, entry_job_number
            )
            SELECT account_name,
                   COALESCE(enty_values_CR, 0),
                   COALESCE(enty_values_DR, 0),
                   entry_effective_date, %s, %s, %s, %s,
                   entry_sub_account_code, entry_job_number
            FROM entry_details WHERE entry_jv = %s AND COALESCE(entry_deleted, 0) = 0
        """, (today, f'Journal Reversal of JV-{jv_no}', current_user_pk, rev_jv, jv_no))

        conn.commit()
        return {'success': True, 'rev_jv': rev_jv}

    except Exception as e:
        if conn:
            conn.rollback()
        return {'error': str(e)}, 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/fix_reversal_dates', methods=['GET', 'POST'])
@login_required
@has_permission('Access_Accounting')
def fix_reversal_dates():
    """One-time repair: reversal JVs used to be posted with TODAY's date instead of
    the original JV's effective date, leaving old periods un-reversed. This tool
    finds those reversals and re-dates them onto the original effective date."""
    if request.method == 'POST':
        ids = request.form.getlist('rev_ids[]')
        lock = get_period_lock_date()
        fixed, skipped = 0, 0
        for rev_jv in ids:
            row = db.execute_query("SELECT jv_user_code FROM jv_numbers WHERE jv_id = %s", (rev_jv,))
            code = (row[0]['jv_user_code'] if row else '') or ''
            if not code.startswith('REV-JV-'):
                continue
            try:
                orig_jv = int(code[7:])
            except (TypeError, ValueError):
                continue
            od = db.execute_query(
                "SELECT MIN(entry_effective_date) AS d FROM entry_details WHERE entry_jv = %s AND entry_deleted = 0",
                (orig_jv,))
            orig_date = od[0]['d'] if od and od[0]['d'] else None
            if not orig_date:
                continue
            if hasattr(orig_date, 'date'):
                orig_date = orig_date.date()
            if lock and orig_date <= lock:
                skipped += 1
                continue
            db.execute_query(
                "UPDATE entry_details SET entry_effective_date = %s WHERE entry_jv = %s",
                (orig_date, rev_jv), commit=True)
            fixed += 1
        msg = f'{fixed} reversal(s) re-dated to their original effective date.'
        if skipped:
            msg += f' {skipped} skipped (original date falls in a closed period).'
        flash(msg, 'success' if fixed else 'warning')
        return redirect(url_for('fix_reversal_dates'))

    # GET: list reversal JVs whose dates differ from their original JV's dates
    try:
        candidates = db.execute_query("""
            SELECT j.jv_id AS rev_jv,
                   CAST(SUBSTRING(j.jv_user_code, 8) AS UNSIGNED) AS orig_jv,
                   DATE(MIN(re.entry_effective_date)) AS rev_date,
                   (SELECT DATE(MIN(oe.entry_effective_date)) FROM entry_details oe
                    WHERE oe.entry_jv = CAST(SUBSTRING(j.jv_user_code, 8) AS UNSIGNED)
                      AND oe.entry_deleted = 0) AS orig_date,
                   SUM(re.enty_values_DR) AS total_dr
            FROM jv_numbers j
            JOIN entry_details re ON re.entry_jv = j.jv_id AND re.entry_deleted = 0
            WHERE j.jv_user_code LIKE 'REV-JV-%'
            GROUP BY j.jv_id, j.jv_user_code
            HAVING orig_date IS NOT NULL AND rev_date <> orig_date
            ORDER BY j.jv_id DESC
        """) or []
    except Exception as e:
        flash(f'Error scanning reversals: {e}', 'danger')
        candidates = []

    return render_template('fix_reversal_dates.html', candidates=candidates,
                           lock_date=get_period_lock_date())


def _find_orig_sub_account(orig_jv, account_name, rev_dr, rev_cr):
    """A reversal line's Debit/Credit is the ORIGINAL line's Credit/Debit
    swapped, so that swap (plus the account name) uniquely identifies which
    original line a reversal line came from — used to recover the
    sub-account code/job number an older reversal never carried over."""
    match = db.execute_query("""
        SELECT entry_sub_account_code, entry_job_number
        FROM entry_details
        WHERE entry_jv = %s AND account_name = %s
          AND COALESCE(enty_values_DR, 0) = %s AND COALESCE(enty_values_CR, 0) = %s
          AND COALESCE(entry_sub_account_code, 0) != 0
          AND COALESCE(entry_deleted, 0) = 0
        LIMIT 1
    """, (orig_jv, account_name, rev_cr or 0, rev_dr or 0))
    return match[0] if match else None


@app.route('/fix_reversal_sub_accounts', methods=['GET', 'POST'])
@login_required
@has_permission('Access_Accounting')
def fix_reversal_sub_accounts():
    """One-time repair: before this fix, reversing a JV did not carry the
    original line's sub-account code (or job number) onto the new reversal
    line — so a sub-account breakdown (e.g. in Profit & Loss or Ledger
    View) never saw the offsetting reversal and looked unchanged even
    though the reversal itself worked. This finds reversal lines missing a
    sub-account code where a matching original line has one, and backfills
    it onto the reversal."""
    if request.method == 'POST':
        rev_ids = request.form.getlist('rev_ids[]')
        fixed_jvs, fixed_lines = 0, 0
        for rev_jv in rev_ids:
            row = db.execute_query("SELECT jv_user_code FROM jv_numbers WHERE jv_id = %s", (rev_jv,))
            code = (row[0]['jv_user_code'] if row else '') or ''
            if not code.startswith('REV-JV-'):
                continue
            try:
                orig_jv = int(code[7:])
            except (TypeError, ValueError):
                continue
            lines = db.execute_query("""
                SELECT id, account_name, enty_values_DR, enty_values_CR
                FROM entry_details
                WHERE entry_jv = %s AND COALESCE(entry_sub_account_code, 0) = 0
                  AND COALESCE(entry_deleted, 0) = 0
            """, (rev_jv,)) or []
            any_fixed = False
            for ln in lines:
                match = _find_orig_sub_account(orig_jv, ln['account_name'],
                                               ln['enty_values_DR'], ln['enty_values_CR'])
                if not match:
                    continue
                db.execute_query(
                    "UPDATE entry_details SET entry_sub_account_code = %s, entry_job_number = %s WHERE id = %s",
                    (match['entry_sub_account_code'], match['entry_job_number'], ln['id']), commit=True)
                fixed_lines += 1
                any_fixed = True
            if any_fixed:
                fixed_jvs += 1
        msg = f'{fixed_jvs} reversal(s), {fixed_lines} line(s) backfilled with their original sub-account code.'
        flash(msg, 'success' if fixed_jvs else 'warning')
        return redirect(url_for('fix_reversal_sub_accounts'))

    # GET: scan reversal lines missing a sub-account code and check whether a
    # matching original line (same account, swapped DR/CR) actually has one —
    # only those are genuinely fixable, so nothing else shows up as a false lead.
    try:
        missing = db.execute_query("""
            SELECT ed.id, ed.entry_jv AS rev_jv, jv.jv_user_code, ed.account_name,
                   ed.enty_values_DR, ed.enty_values_CR, ed.entry_effective_date
            FROM entry_details ed
            JOIN jv_numbers jv ON ed.entry_jv = jv.jv_id
            WHERE jv.jv_user_code LIKE 'REV-JV-%'
              AND COALESCE(ed.entry_sub_account_code, 0) = 0
              AND COALESCE(ed.entry_deleted, 0) = 0
        """) or []
    except Exception as e:
        flash(f'Error scanning reversals: {e}', 'danger')
        missing = []

    by_jv = {}
    for m in missing:
        code = m['jv_user_code'] or ''
        try:
            orig_jv = int(code[7:])
        except (TypeError, ValueError):
            continue
        match = _find_orig_sub_account(orig_jv, m['account_name'], m['enty_values_DR'], m['enty_values_CR'])
        if not match:
            continue
        sub_name = None
        try:
            sn = db.execute_query(
                "SELECT sub_sub_accaount_name FROM sub_accont_for_new_account WHERE sub_account_code = %s",
                (match['entry_sub_account_code'],))
            sub_name = sn[0]['sub_sub_accaount_name'] if sn else None
        except Exception:
            pass
        entry = by_jv.setdefault(m['rev_jv'], {
            'rev_jv': m['rev_jv'], 'orig_jv': orig_jv,
            'date': str(m['entry_effective_date']), 'accounts': [], 'sub_names': set()
        })
        entry['accounts'].append(m['account_name'])
        entry['sub_names'].add(sub_name or f"code {match['entry_sub_account_code']}")

    candidates = sorted(by_jv.values(), key=lambda c: c['rev_jv'], reverse=True)
    for c in candidates:
        c['accounts'] = ', '.join(sorted(set(c['accounts'])))
        c['sub_names'] = ', '.join(sorted(c['sub_names']))

    return render_template('fix_reversal_sub_accounts.html', candidates=candidates)


@app.route('/journal_entry/print/<int:jv_no>')
@login_required
def print_journal_voucher(jv_no):
    # Fetch Header
    # Note: jv_numbers has jv_id, jv_user_code, jv_naration
    # Need date from entries or assume logic. entries have entry_effective_date.
    header_query = """
        SELECT j.jv_user_code, j.jv_naration, MIN(e.entry_effective_date) as entry_date,
               SUM(e.enty_values_DR) as total_amount
        FROM jv_numbers j
        LEFT JOIN entry_details e ON j.jv_id = e.entry_jv
        WHERE j.jv_id = %s
        GROUP BY j.jv_id
    """
    header_res = db.execute_query(header_query, (jv_no,))
    if not header_res:
        return "JV Not Found", 404
    header = header_res[0]

    # Fetch Details
    details_query = """
        SELECT account_name, entry_naration, enty_values_DR, enty_values_CR, entry_sub_account_code
        FROM entry_details
        WHERE entry_jv = %s
        ORDER BY id
    """
    details = db.execute_query(details_query, (jv_no,))

    # Fetch Company Info
    company_res = db.execute_query("SELECT * FROM company LIMIT 1")
    company = company_res[0] if company_res else {}

    return render_template('jv_print.html', header=header, details=details, company=company, jv_sys_id=jv_no)

def create_default_user():
    """Creates a default admin user if the Login_Table is empty."""
    try:
        # Check connection first
        conn = db.get_connection()
        if not conn:
            logging.warning("WARNING: Database connection failed. Cannot create default user.")
            return

        # Check for existing users
        result = db.execute_query("SELECT COUNT(*) as count FROM Login_Table")
        if result and result[0]['count'] == 0:
            print("No users found. Creating default admin user...")
            # Hash default password
            from werkzeug.security import generate_password_hash
            pw_hash = generate_password_hash('123')
            logging.info("No users found. Creating default admin user...")
            query = """
                INSERT INTO Login_Table (User_Name, Password, User_Code, User_Active, Mobile_No, Email)
                VALUES (%s, %s, %s, %s, %s, %s)
            """
            # Using 'admin' / '123' (Hashed)
            db.execute_query(query, ('admin', pw_hash, 'ADM001', 1, '0000000000', 'admin@example.com'), commit=True)
            logging.info("Default user created: admin / 123")

            # Create Default Rights for Admin
            last_id_res = db.execute_query("SELECT id FROM Login_Table WHERE User_Name = 'admin'")
            if last_id_res:
                uid = last_id_res[0]['id']
                db.execute_query("""
                    INSERT INTO User_Rights (Link_To_Loging_Tabke, Add_New_User, OP_Approved, Access_Inventory, Access_POS, Access_Accounting, Access_Reports, Access_Reversals)
                    VALUES (%s, 1, 1, 1, 1, 1, 1, 1)
                """, (uid,), commit=True)

        else:
            logging.info("Users exist in database. Skipping default user creation.")
    except Exception as e:
        logging.error("Error creating default user.")

def ensure_default_categories(target_db=None):
    """Ensures default Balance Sheet and P&L categories exist."""
    current_db = target_db if target_db else db
    try:
        conn = current_db.get_connection()
        if not conn: return
        cursor = conn.cursor()

        # Balance Sheet Categories
        bs_cats = [
            ('ASSETS', 1),
            ('Non-current assets', 2),
            ('Current assets', 3),
            ('EQUITY AND LIABILITIES', 4),
            ('Capital and reserves', 5),
            ('Current liabilities', 6)
        ]
        try:
            cursor.execute("SELECT holding_position FROM balance_sheet_category")
            existing_bs_positions = {row['holding_position'] if isinstance(row, dict) else row[0] for row in cursor.fetchall()}
        except Exception as e:
            logging.error(f"Error fetching BS categories: {e}")
            existing_bs_positions = set()

        bs_inserts = [(name, pos, date.today()) for name, pos in bs_cats if pos not in existing_bs_positions]
        if bs_inserts:
            try:
                cursor.executemany("INSERT INTO balance_sheet_category (name_of_category, holding_position, create_date_time) VALUES (%s, %s, %s)", bs_inserts)
            except Exception as e:
                logging.error(f"Error bulk inserting BS categories: {e}")

        # P&L Categories
        pl_cats = [
            ('Revenue', 1),
            ('Cost of sales', 2),
            ('Gross profit', 3),
            ('Distribution costs', 4),
            ('Administrative expenses', 5),
            ('Other operating expenses', 6),
            ('Finance cost', 7),
            ('Income from associates', 8),
            ('Income tax expenses', 9),
            ('Minority interest', 10),
            ('Extraordinary items', 11)
        ]
        try:
            cursor.execute("SELECT holding_position FROM `p&l_category`")
            existing_pl_positions = {row['holding_position'] if isinstance(row, dict) else row[0] for row in cursor.fetchall()}
        except Exception as e:
            logging.error(f"Error fetching PL categories: {e}")
            existing_pl_positions = set()

        pl_inserts = [(name, pos, date.today()) for name, pos in pl_cats if pos not in existing_pl_positions]
        if pl_inserts:
            try:
                cursor.executemany("INSERT INTO `p&l_category` (name_of_category, holding_position, create_date_time) VALUES (%s, %s, %s)", pl_inserts)
            except Exception as e:
                logging.error(f"Error bulk inserting PL categories: {e}")

        # CF Categories
        cf_cats = [
            ('Operating Activities', 1), ('Investing Activities', 2), ('Financing Activities', 3),
            ('Adjustments', 0), ('Changes In Working Capital', 0)
        ]
        try:
            cursor.execute("SELECT catogory_name FROM cf_catogory")
            existing_cf_names = {row['catogory_name'] if isinstance(row, dict) else row[0] for row in cursor.fetchall()}
        except Exception as e:
            logging.error(f"Error fetching CF categories: {e}")
            existing_cf_names = set()

        cf_inserts = [(name, pos) for name, pos in cf_cats if name not in existing_cf_names]
        if cf_inserts:
            try:
                cursor.executemany("INSERT INTO cf_catogory (catogory_name, hold_level) VALUES (%s, %s)", cf_inserts)
            except Exception as e:
                logging.error(f"Error bulk inserting CF categories: {e}")

        conn.commit()
        cursor.close()
        conn.close()
        clear_category_cache()
        logging.info("Default categories checked/created.")

    except Exception as e:
        logging.error(f"Error ensuring default categories: {e}")

def validate_db_config(config):
    """
    Validates database configuration to prevent command injection.
    Only allows alphanumeric characters, underscores, and hyphens in sensitive fields.
    Does not allow arguments starting with dash to prevent argument injection.
    """
    # Alphanumeric, underscore, hyphen.
    # Host might contain dots (IP/domain) and colons (IPv6/port although mysql cli uses -P for port).
    # MySQL CLI host (-h) expects hostname or IP.
    # Database and User usually don't have dots but to be safe we can stick to strict for them.

    strict_pattern = re.compile(r'^[a-zA-Z0-9_-]+$')
    host_pattern = re.compile(r'^[a-zA-Z0-9_.-]+$') # Allow dots for host

    # Fields to validate (user, host, database)
    # Password is treated differently (passed via env)

    for field in ['user', 'database']:
        value = config.get(field, '')
        if not value: continue
        if not isinstance(value, str) or value.startswith('-') or not strict_pattern.match(value):
            return False

    # Validate Host specifically
    host = config.get('host', '')
    if host:
        if not isinstance(host, str) or host.startswith('-') or not host_pattern.match(host):
            return False

    return True

# --- System Backup ---
@app.route('/system_backup')
@login_required
def system_backup():
    # Only allow admin or specific users? For now, login_required is minimal.
    # Ideally should be restricted.

    # Validate Config
    if not validate_db_config(db_config):
        flash('Invalid database configuration', 'danger')
        return redirect(url_for('index'))

    # Get the user's specific database name
    db_name = get_session_db_name()
    if not is_safe_db_name(db_name):
        flash('Invalid database name', 'danger')
        return redirect(url_for('index'))

    # Attempt to find the dump binary
    dump_cmd = shutil.which('mysqldump')
    if not dump_cmd:
        dump_cmd = shutil.which('mariadb-dump')

    # Fallback to string name if shutil.which fails due to PATH issues on some hosts
    if not dump_cmd:
        dump_cmd = 'mysqldump'

    try:
        filename = f"backup_{date.today().strftime('%Y%m%d')}.sql"

        # Create temporary defaults file to store credentials securely
        # Using tempfile with strict permissions
        defaults_file = tempfile.NamedTemporaryFile(mode='w+', delete=False)
        try:
            # Set permissions to 600 (owner read/write only) BEFORE writing data
            os.chmod(defaults_file.name, 0o600)

            # Write credentials
            defaults_file.write('[client]\n')
            defaults_file.write(f"user={db_config['user']}\n")
            if db_config['password']:
                defaults_file.write(f"password={db_config['password']}\n")
            defaults_file.write(f"host={db_config['host']}\n")
            defaults_file.flush()
            defaults_file.close()

            # Command construction
            # mysqldump --defaults-extra-file=... -- database > filename
            # Since we are in python, we can pipe output to string or file.

            cmd = [
                dump_cmd,
                f'--defaults-extra-file={defaults_file.name}',
                '--', # End of options
                db_name
            ]

            def generate():
                # Fix: We don't pipe stderr to avoid deadlocks when the buffer fills up.
                # We can either redirect it to DEVNULL or capture it to a temporary file.
                # To keep it simple and avoid deadlock, we pipe stderr to DEVNULL since
                # returning a streaming response means we can't easily send the error to the client anyway
                # once the stream has started, and a 64KB stderr buffer would hang the process.
                try:
                    process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.DEVNULL)
                except FileNotFoundError:
                    logging.error(f"Backup command not found: {cmd}")
                    yield b"-- Error: MySQL dump utility not found on server.\n"
                    return

                try:
                    while True:
                        chunk = process.stdout.read(8192)
                        if not chunk:
                            break
                        yield chunk
                    process.wait()
                    if process.returncode != 0:
                        logging.error(f"Backup failed with return code {process.returncode}")
                finally:
                    if process.poll() is None:
                        process.kill()
                    # Secure cleanup
                    if os.path.exists(defaults_file.name):
                        os.remove(defaults_file.name)

            # We don't remove defaults_file in the outer finally block anymore,
            # it is cleaned up by the generator when it completes or errors out.

            response = Response(stream_with_context(generate()), mimetype='application/sql')
            response.headers['Content-Disposition'] = f'attachment; filename={filename}'
            return response

        except Exception as e:
            if os.path.exists(defaults_file.name):
                os.remove(defaults_file.name)
            raise e

    except Exception as e:
        flash(f'Backup error: {str(e)}', 'danger')
        return redirect(url_for('index'))


# --- System Restore ---
@app.route('/system_restore', methods=['GET', 'POST'])
@login_required
def system_restore():
    if request.method == 'POST':
        # Check if file was uploaded
        if 'backup_file' not in request.files:
            flash('No file part', 'danger')
            return redirect(request.url)

        file = request.files['backup_file']

        if file.filename == '':
            flash('No selected file', 'danger')
            return redirect(request.url)

        if not file.filename.endswith('.sql'):
            flash('Invalid file format. Please upload a .sql file.', 'danger')
            return redirect(request.url)

        # Validate Config
        if not validate_db_config(db_config):
            flash('Invalid database configuration', 'danger')
            return redirect(url_for('index'))

        # Get the user's specific database name
        db_name = get_session_db_name()
        if not is_safe_db_name(db_name):
            flash('Invalid database name', 'danger')
            return redirect(url_for('index'))

        # Attempt to find the mysql binary
        mysql_cmd = shutil.which('mysql')
        if not mysql_cmd:
            mysql_cmd = shutil.which('mariadb')

        # Fallback
        if not mysql_cmd:
            mysql_cmd = 'mysql'

        try:
            # Create temporary defaults file to store credentials securely
            defaults_file = tempfile.NamedTemporaryFile(mode='w+', delete=False)
            try:
                # Set permissions to 600 (owner read/write only) BEFORE writing data
                os.chmod(defaults_file.name, 0o600)

                # Write credentials
                defaults_file.write('[client]\n')
                defaults_file.write(f"user={db_config['user']}\n")
                if db_config['password']:
                    defaults_file.write(f"password={db_config['password']}\n")
                defaults_file.write(f"host={db_config['host']}\n")
                defaults_file.flush()
                defaults_file.close()

                # Command construction
                cmd = [
                    mysql_cmd,
                    f'--defaults-extra-file={defaults_file.name}',
                    '--', # End of options
                    db_name
                ]

                try:
                    # Execute the mysql command
                    process = subprocess.Popen(cmd, stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE)

                    # Write the uploaded file content to stdin
                    file_content = file.read()
                    stdout, stderr = process.communicate(input=file_content)

                    if process.returncode == 0:
                        flash('Database restored successfully', 'success')
                    else:
                        error_msg = stderr.decode('utf-8') if stderr else "Unknown error"
                        logging.error(f"Restore failed with return code {process.returncode}. Error: {error_msg}")
                        flash(f'Restore failed. Error: {error_msg}', 'danger')

                except FileNotFoundError:
                    logging.error(f"Restore command not found: {cmd}")
                    flash('MySQL client utility not found on server.', 'danger')

            finally:
                if os.path.exists(defaults_file.name):
                    os.remove(defaults_file.name)

        except Exception as e:
            logging.error(f"Error during database restore: {str(e)}")
            flash(f'An error occurred during restore: {str(e)}', 'danger')

        return redirect(url_for('index'))

    return render_template('system_restore.html')


# --- Fixed Assets Module ---
@app.route('/fixed_assets')
@login_required
@has_permission('Access_Accounting')
def fixed_assets():
    # Fetch accounts for dropdowns
    accounts = db.execute_query("SELECT id, account_name FROM new_account_table WHERE account_active = 1 ORDER BY account_name")

    # Fetch existing Classes and Locations for suggestions
    classes = db.execute_query("SELECT DISTINCT asset_class FROM fixed_assets_register WHERE asset_class IS NOT NULL AND asset_class != ''")
    locations = db.execute_query("SELECT DISTINCT location FROM fixed_assets_register WHERE location IS NOT NULL AND location != ''")

    # Fetch suppliers
    suppliers = db.execute_query("SELECT sup_id, supplier_name, supplier_code FROM suppliers WHERE Is_Suplier = 1 ORDER BY supplier_name")

    return render_template('fixed_assets.html', accounts=accounts, classes=classes, locations=locations, suppliers=suppliers)

@app.route('/fixed_assets/add', methods=['POST'])
@login_required
@has_permission('Access_Accounting')
def add_fixed_asset():
    try:
        class_name = request.form.get('asset_class')
        desc = request.form.get('description')
        brand = request.form.get('brand_name')
        qty = int(request.form.get('quantity', 1))
        serial = request.form.get('serial_no')
        location = request.form.get('location')
        cost = float(request.form.get('cost_value', 0))
        p_date = request.form.get('purchasing_date')
        life = int(request.form.get('depreciable_life_months', 0))

        asset_acc = request.form.get('asset_account_id')
        exp_acc = request.form.get('expense_account_id')
        acc_dep_acc = request.form.get('accumulated_dep_account_id')
        supplier_id = request.form.get('supplier_id') or None

        if not asset_acc or not exp_acc or not acc_dep_acc:
             flash('Please select all GL accounts', 'warning')
             return redirect(url_for('fixed_assets'))

        # Post to GL if requested
        post_gl = request.form.get('post_gl')
        credit_acc_id = request.form.get('credit_account_id')

        jv_id = None

        if post_gl and credit_acc_id and cost > 0:
            current_user = get_current_user_id()
            current_user_pk = get_current_user_pk()

            conn = db.get_connection()
            if conn:
                cursor = conn.cursor(dictionary=True)
                try:
                    conn.start_transaction()

                    # Get account names
                    cursor.execute("SELECT account_name FROM new_account_table WHERE id = %s", (asset_acc,))
                    asset_acc_name = cursor.fetchone()['account_name']

                    cursor.execute("SELECT account_name FROM new_account_table WHERE id = %s", (credit_acc_id,))
                    credit_acc_name = cursor.fetchone()['account_name']

                    # Create JV
                    narration = f"Fixed Asset Purchase - {class_name} ({desc})"
                    cursor.execute("INSERT INTO jv_numbers (jv_user_code, jv_naration) VALUES (%s, %s)",
                                  (f"FA-PUR-{int(datetime.now().timestamp())}", narration))
                    jv_id = cursor.lastrowid

                    # DR Asset Account
                    cursor.execute("""
                        INSERT INTO entry_details (
                            account_name, enty_values_DR, enty_values_CR, entry_effective_date, entry_create_date,
                            entry_naration, entry_create_user, entry_jv
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """, (asset_acc_name, cost, 0, p_date, date.today(), narration, current_user_pk, jv_id))

                    # CR Payment/Liability Account
                    cursor.execute("""
                        INSERT INTO entry_details (
                            account_name, enty_values_DR, enty_values_CR, entry_effective_date, entry_create_date,
                            entry_naration, entry_create_user, entry_jv
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """, (credit_acc_name, 0, cost, p_date, date.today(), narration, current_user_pk, jv_id))

                    # Log Liability for Account Payable
                    if credit_acc_name.lower() == 'account payable' and supplier_id:
                        cursor.execute("SELECT supplier_code FROM suppliers WHERE sup_id = %s", (supplier_id,))
                        sup_res = cursor.fetchone()
                        supplier_code = sup_res['supplier_code'] if sup_res else ""

                        inv_num = f"FA-INV-{jv_id}"
                        cursor.execute("""
                            INSERT INTO suppliers_invoice_data (
                                suppliers_code, suppliers_invoice_number, suppliers_invoice_date,
                                suppliers_invoice_total_oustanding, suppliers_invoice_total_payment,
                                suppliers_invoice_final_date, suppliers_invoice_buinding_supplier,
                                suppliers_invoice_JV, suppliers_VAT_rate
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (supplier_code, inv_num, p_date, cost, 0, p_date, supplier_id, jv_id, 0))

                    # Finally insert Asset (now with JV)
                    query = """
                        INSERT INTO fixed_assets_register
                        (asset_class, description, brand_name, quantity, serial_no, location, cost_value, purchasing_date, depreciable_life_months, asset_account_id, expense_account_id, accumulated_dep_account_id, supplier_id, jv_id)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """
                    cursor.execute(query, (class_name, desc, brand, qty, serial, location, cost, p_date, life, asset_acc, exp_acc, acc_dep_acc, supplier_id, jv_id))

                    conn.commit()
                except Exception as e:
                    conn.rollback()
                    raise Exception(f"Failed to post to GL/Register: {str(e)}")
                finally:
                    cursor.close()
                    conn.close()
        else:
            # Standard Insert without GL
            query = """
                INSERT INTO fixed_assets_register
                (asset_class, description, brand_name, quantity, serial_no, location, cost_value, purchasing_date, depreciable_life_months, asset_account_id, expense_account_id, accumulated_dep_account_id, supplier_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            db.execute_query(query, (class_name, desc, brand, qty, serial, location, cost, p_date, life, asset_acc, exp_acc, acc_dep_acc, supplier_id), commit=True)

        flash('Asset added successfully', 'success')
    except Exception as e:
        flash(f'Error adding asset: {str(e)}', 'danger')
    return redirect(url_for('fixed_assets'))

@app.route('/fixed_assets/delete/<int:asset_id>', methods=['POST'])
@login_required
@has_permission('Access_Accounting')
def delete_fixed_asset(asset_id):
    try:
        conn = db.get_connection()
        if not conn:
            flash("Database connection error.", "danger")
            return redirect(url_for('fixed_assets'))

        cursor = conn.cursor(dictionary=True)
        conn.start_transaction()

        cursor.execute("SELECT * FROM fixed_assets_register WHERE id = %s", (asset_id,))
        asset = cursor.fetchone()

        if not asset:
            flash("Asset not found.", "danger")
            return redirect(url_for('fixed_assets'))

        if asset['is_written_off'] == 1:
            flash("Asset is already written off.", "warning")
            return redirect(url_for('fixed_assets'))

        action = request.form.get('action') # 'delete' or 'write_off'
        jv_id = asset.get('jv_id')

        # Check Payment Status if there is a JV linked
        payment_made = False
        if jv_id:
            cursor.execute("SELECT suppliers_invoice_total_payment FROM suppliers_invoice_data WHERE suppliers_invoice_JV = %s", (jv_id,))
            inv = cursor.fetchone()
            if inv and float(inv['suppliers_invoice_total_payment'] or 0) > 0:
                payment_made = True

        if action == 'delete':
            if payment_made:
                flash("Cannot delete asset: A payment has already been made against its purchase. Please use 'Write Off' instead.", "danger")
                return redirect(url_for('fixed_assets'))

            cursor.execute("SELECT COUNT(*) as dep_count FROM asset_depreciation_history WHERE asset_id = %s", (asset_id,))
            dep_res = cursor.fetchone()
            if dep_res and dep_res['dep_count'] > 0:
                flash("Cannot delete asset: It has already been depreciated. Please use 'Write Off' instead.", "danger")
                return redirect(url_for('fixed_assets'))

            # Safe to Delete: Remove GL entries and Asset
            if jv_id:
                cursor.execute("DELETE FROM entry_details WHERE entry_jv = %s", (jv_id,))
                cursor.execute("DELETE FROM suppliers_invoice_data WHERE suppliers_invoice_JV = %s", (jv_id,))
                cursor.execute("DELETE FROM jv_numbers WHERE jv_id = %s", (jv_id,))

            cursor.execute("DELETE FROM fixed_assets_register WHERE id = %s", (asset_id,))
            conn.commit()
            flash("Asset deleted successfully.", "success")

        elif action == 'write_off':
            # Write Off Process
            loss_acc_id = request.form.get('loss_account_id')
            if not loss_acc_id:
                flash("Write-Off requires a Loss/Expense account.", "danger")
                return redirect(url_for('fixed_assets'))

            cursor.execute("SELECT account_name FROM new_account_table WHERE id = %s", (loss_acc_id,))
            loss_acc = cursor.fetchone()
            if not loss_acc:
                flash("Invalid Loss Account.", "danger")
                return redirect(url_for('fixed_assets'))
            loss_acc_name = loss_acc['account_name']

            cursor.execute("SELECT account_name FROM new_account_table WHERE id = %s", (asset['asset_account_id'],))
            asset_acc_name = cursor.fetchone()['account_name']

            cursor.execute("SELECT account_name FROM new_account_table WHERE id = %s", (asset['accumulated_dep_account_id'],))
            acc_dep_name = cursor.fetchone()['account_name']

            # Calculate accumulated depreciation
            cursor.execute("SELECT SUM(amount) as total FROM asset_depreciation_history WHERE asset_id = %s", (asset_id,))
            dep_res = cursor.fetchone()
            acc_dep_total = float(dep_res['total'] or 0)

            nbv = float(asset['cost_value'] or 0) - acc_dep_total

            # Create Write Off JV
            current_user_pk = get_current_user_pk()
            today = date.today()
            narration = f"Write-Off Fixed Asset: {asset['asset_class']} - {asset['serial_no']}"
            cursor.execute("INSERT INTO jv_numbers (jv_user_code, jv_naration) VALUES (%s, %s)", (f"FA-WO-{int(datetime.now().timestamp())}", narration))
            wo_jv_id = cursor.lastrowid

            # 1. CR Asset Account (Full Cost)
            cursor.execute("""
                INSERT INTO entry_details (
                    account_name, enty_values_DR, enty_values_CR, entry_effective_date, entry_create_date, entry_naration, entry_create_user, entry_jv
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (asset_acc_name, 0, asset['cost_value'], today, today, narration, current_user_pk, wo_jv_id))

            # 2. DR Acc Dep Account (Reverse Accumulated Depreciation)
            if acc_dep_total > 0:
                cursor.execute("""
                    INSERT INTO entry_details (
                        account_name, enty_values_DR, enty_values_CR, entry_effective_date, entry_create_date, entry_naration, entry_create_user, entry_jv
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """, (acc_dep_name, acc_dep_total, 0, today, today, narration, current_user_pk, wo_jv_id))

            # 3. DR Loss Account (Net Book Value)
            if nbv > 0:
                cursor.execute("""
                    INSERT INTO entry_details (
                        account_name, enty_values_DR, enty_values_CR, entry_effective_date, entry_create_date, entry_naration, entry_create_user, entry_jv
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """, (loss_acc_name, nbv, 0, today, today, narration, current_user_pk, wo_jv_id))

            # Update Asset Register
            cursor.execute("UPDATE fixed_assets_register SET status = 'Written-Off', is_written_off = 1, write_off_amount = %s WHERE id = %s", (nbv, asset_id))

            conn.commit()
            flash(f"Asset written off successfully. Loss amount: {nbv}", "success")

    except Exception as e:
        if 'conn' in locals() and conn: conn.rollback()
        flash(f"Error processing asset deletion/write-off: {str(e)}", "danger")
    finally:
        if 'cursor' in locals() and cursor: cursor.close()
        if 'conn' in locals() and conn: conn.close()

    return redirect(url_for('fixed_assets'))

@app.route('/fixed_assets/calculate_depreciation', methods=['POST'])
@login_required
@has_permission('Access_Accounting')
def calculate_depreciation():
    month_str = request.form.get('month') # "YYYY-MM"
    if not month_str:
        return {'error': 'Month required'}, 400

    try:
        year, month = map(int, month_str.split('-'))
        # Last day of month
        import calendar
        last_day = calendar.monthrange(year, month)[1]
        dep_date = date(year, month, last_day)

        current_user = get_current_user_id()
        current_user_pk = get_current_user_pk()

        conn = db.get_connection()
        cursor = conn.cursor(dictionary=True)
        conn.start_transaction()

        try:
            # 1. Fetch Active Assets
            cursor.execute("SELECT * FROM fixed_assets_register WHERE status = 'Active'")
            assets = cursor.fetchall()

            processed_count = 0

            # Create a shared JV for this month's depreciation run
            cursor.execute("INSERT INTO jv_numbers (jv_user_code, jv_naration) VALUES (%s, %s)",
                           (f"DEP-{month_str}", f"Depreciation Run for {month_str}"))
            jv_id = cursor.lastrowid

            total_dr = 0
            total_cr = 0
            entries = []

            # Pre-fetch existing depreciations for this month
            cursor.execute("""
                SELECT asset_id FROM asset_depreciation_history
                WHERE YEAR(depreciation_date) = %s AND MONTH(depreciation_date) = %s
            """, (year, month))
            depreciated_assets = {row['asset_id'] for row in cursor.fetchall()}

            for asset in assets:
                # Check if already depreciated for this month
                if asset['id'] in depreciated_assets:
                    continue # Skip if already done

                # Check purchase date
                p_date = asset['purchasing_date']
                if not p_date: continue
                if isinstance(p_date, datetime): p_date = p_date.date()

                # If purchased this month or after, maybe skip or pro-rata?
                # Simple rule: Depreciate if purchased before end of month
                if p_date > dep_date:
                    continue

                # Calculate Amount (Straight Line Monthly)
                # Cost / Life Months
                life = asset['depreciable_life_months']
                if life <= 0: continue

                monthly_amount = asset['cost_value'] / life

                # Check if fully depreciated
                cursor.execute("SELECT SUM(amount) as total FROM asset_depreciation_history WHERE asset_id = %s", (asset['id'],))
                res = cursor.fetchone()
                acc_dep = res['total'] if res and res['total'] else 0

                remaining = asset['cost_value'] - acc_dep
                if remaining <= 0:
                    continue # Fully depreciated

                # Cap amount at remaining
                amount = min(monthly_amount, remaining)
                if amount <= 0: continue

                # Record History
                cursor.execute("""
                    INSERT INTO asset_depreciation_history (asset_id, depreciation_date, amount, jv_id)
                    VALUES (%s, %s, %s, %s)
                """, (asset['id'], dep_date, amount, jv_id))

                # Prepare GL Entries
                # Need account names for entry_details (it uses name, not ID, sadly)
                # Fetch account names
                cursor.execute("SELECT account_name FROM new_account_table WHERE id = %s", (asset['expense_account_id'],))
                exp_name = cursor.fetchone()['account_name']

                cursor.execute("SELECT account_name FROM new_account_table WHERE id = %s", (asset['accumulated_dep_account_id'],))
                acc_dep_name = cursor.fetchone()['account_name']

                entries.append({
                    'dr_acc': exp_name,
                    'cr_acc': acc_dep_name,
                    'amount': amount,
                    'narration': f"Depreciation {month_str} - {asset['asset_class']} - {asset['serial_no']}"
                })

                processed_count += 1

            # Aggregate Entries by Account to reduce lines?
            # Or insert per asset? Per asset is better for audit trail in narration

            insert_tuples = []
            for e in entries:
                # DR Tuple
                insert_tuples.append((e['dr_acc'], e['amount'], 0, dep_date, date.today(), e['narration'], current_user_pk, jv_id))
                # CR Tuple
                insert_tuples.append((e['cr_acc'], 0, e['amount'], dep_date, date.today(), e['narration'], current_user_pk, jv_id))

            if insert_tuples:
                cursor.executemany("""
                    INSERT INTO entry_details (
                        account_name, enty_values_DR, enty_values_CR, entry_effective_date, entry_create_date,
                        entry_naration, entry_create_user, entry_jv
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """, insert_tuples)

            conn.commit()
            return {'success': True, 'processed': processed_count, 'jv_id': jv_id}

        except Exception as e:
            conn.rollback()
            logging.error(f"Depreciation Error: {e}")
            return {'error': str(e)}, 500
        finally:
            cursor.close()
            conn.close()

    except Exception as e:
        return {'error': str(e)}, 500

@app.route('/fixed_assets/data')
@login_required
def fixed_assets_data():
    # Fetch all assets
    assets = db.execute_query("""
        SELECT f.*,
               COALESCE(s.suppliers_invoice_total_payment, 0) as jv_payment
        FROM fixed_assets_register f
        LEFT JOIN suppliers_invoice_data s ON f.jv_id = s.suppliers_invoice_JV
        ORDER BY f.id
    """)

    # Fetch all depreciation history
    history = db.execute_query("SELECT * FROM asset_depreciation_history ORDER BY depreciation_date")

    # Process history into a dict: asset_id -> { 'YYYY-MM': amount, 'total': sum }
    hist_map = {}
    months = set()

    for h in history:
        aid = h['asset_id']
        if aid not in hist_map: hist_map[aid] = {'total': 0, 'months': {}}

        d_date = h['depreciation_date']
        if isinstance(d_date, datetime): d_date = d_date.date()
        month_key = d_date.strftime('%Y-%b') # e.g., 2023-Jan

        hist_map[aid]['months'][month_key] = float(h['amount'])
        hist_map[aid]['total'] += float(h['amount'])
        months.add(month_key)

    # Sort months chronologically
    sorted_months = sorted(list(months), key=lambda x: datetime.strptime(x, '%Y-%b'))

    # Prepare Result
    result = {
        'columns': ['Class', 'Description', 'Brand', 'Qty', 'Serial', 'Location', 'Cost', 'Purchase Date', 'Life (M)'] + sorted_months + ['Total Dep', 'Net Book Value'],
        'data': []
    }

    for a in assets:
        row = {
            'id': a['id'],
            'class': a['asset_class'],
            'desc': a['description'],
            'brand': a['brand_name'],
            'qty': a['quantity'],
            'serial': a['serial_no'],
            'location': a['location'],
            'cost': float(a['cost_value']),
            'date': str(a['purchasing_date']),
            'life': a['depreciable_life_months']
        }

        h_data = hist_map.get(a['id'], {'total': 0, 'months': {}})

        # Add monthly columns
        for m in sorted_months:
            row[m] = h_data['months'].get(m, 0)

        row['total_dep'] = h_data['total']
        row['nbv'] = float(a['cost_value']) - h_data['total']
        row['status'] = a.get('status', 'Active')
        row['payment_made'] = 1 if float(a.get('jv_payment', 0)) > 0 else 0

        result['data'].append(row)

    result['month_headers'] = sorted_months
    return json.dumps(result)

# --- Inventory Transfer ---
@app.route('/inventory_transfer', methods=['GET'])
@login_required
@has_permission('Access_Inventory')
def inventory_transfer():
    locations = db.execute_query("SELECT inventory_locations_name FROM inventory_locations")
    jobs = db.execute_query("SELECT job_number FROM jobs_unit")
    # Fetch active items with cost for reference (though cost isn't changed in transfer)
    items = db.execute_query("""
        SELECT i.inventoy_name, i.inventoy_code, i.inventoy_items_messurment_unit, p.inventory_price_purcharsing
        FROM inventoy_items i
        LEFT JOIN inventory_price_recod p ON i.id = p.inventory_price_link
        WHERE i.active = 1
    """)
    return render_template('inventory_transfer.html', locations=locations, jobs=jobs, items=items, today_date=date.today().strftime('%Y-%m-%d'))

@app.route('/inventory_transfer/submit', methods=['POST'])
@login_required
@has_permission('Access_Inventory')
def submit_inventory_transfer():
    try:
        transfer_date = request.form.get('transfer_date')
        job_no = request.form.get('job_no')
        from_loc = request.form.get('from_location')
        to_loc = request.form.get('to_location')
        narration = request.form.get('narration')

        item_names = request.form.getlist('item_name[]')
        item_codes = request.form.getlist('item_code[]')
        item_units = request.form.getlist('item_unit[]')
        item_costs = request.form.getlist('item_cost[]')
        qtys = request.form.getlist('qty[]')

        if not item_names:
            flash('No items to transfer', 'danger')
            return redirect(url_for('inventory_transfer'))

        if from_loc == to_loc:
            flash('Source and Destination locations must be different', 'danger')
            return redirect(url_for('inventory_transfer'))

        current_user = get_current_user_id()
        current_user_pk = get_current_user_pk()

        conn = db.get_connection()
        cursor = conn.cursor()
        conn.start_transaction()

        try:
            # 1. Create JV for tracking (Transfer Note)
            cursor.execute("INSERT INTO jv_numbers (jv_user_code, jv_naration) VALUES (%s, %s)",
                           (str(current_user), f"Inventory Transfer: {narration}"))
            jv_no = cursor.lastrowid

            tf_note = f"TF-Note{jv_no}"

            batch_data = []
            today_date = datetime.now().date()

            for i in range(len(item_names)):
                qty = float(qtys[i])
                cost = float(item_costs[i] or 0)

                if qty <= 0: continue

                # 2. Record OUT from Source (moument_in = 0, movment_out = qty)
                batch_data.append((
                    item_names[i], item_codes[i], item_units[i], cost,
                    0, qty, # in, out
                    tf_note, current_user, today_date, from_loc,
                    transfer_date, narration, jv_no
                ))

                # 3. Record IN to Destination (moument_in = qty, movment_out = 0)
                batch_data.append((
                    item_names[i], item_codes[i], item_units[i], cost,
                    qty, 0, # in, out
                    tf_note, current_user, today_date, to_loc,
                    transfer_date, narration, jv_no
                ))

            if batch_data:
                query = """
                    INSERT INTO inventory_recod (
                        inventoy_name, inventoy_code, inventory_recod_mesrmet,
                        inventory_recod_unit_price, inventory_recod_moument_in, inventory_recod_movment_out,
                        inventory_recod_suplier_iv_no, inventory_recod_user_id,
                        inventory_recod_user_recod_date, inventory_recod_location,
                        inventory_recod_action_date, inventory_recodcol_memo, JV_No
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                cursor.executemany(query, batch_data)

            conn.commit()
            flash(f'Transfer successful. Tracking Ref: {tf_note}', 'success')

        except Exception as e:
            conn.rollback()
            flash(f'Error processing transfer: {str(e)}', 'danger')
        finally:
            cursor.close()
            conn.close()

    except Exception as e:
        flash(f'System Error: {str(e)}', 'danger')

    return redirect(url_for('inventory_transfer'))

# --- Invoice Creation ---
@app.route('/invoice_creating', methods=['GET'])
@login_required
@has_permission('Access_Accounting') # or Access_Sales if defined
def invoice_creating():
    customers = db.execute_query("SELECT supplier_name as customer_name FROM suppliers WHERE Is_Customer = 1")
    locations = db.execute_query("SELECT inventory_locations_name FROM inventory_locations")
    jobs = db.execute_query("SELECT job_number FROM jobs_unit")

    # Active items with cost (purcharsing price)
    items = db.execute_query("""
        SELECT i.id, i.inventoy_name, i.inventoy_code, i.inventoy_items_messurment_unit as unit, p.inventory_price_purcharsing as cost
        FROM inventoy_items i
        LEFT JOIN inventory_price_recod p ON i.id = p.inventory_price_link
        WHERE i.active = 1
    """)

    today_date = date.today().strftime('%Y-%m-%d')

    # Multi-currency support
    base_currency = 'LKR'
    try:
        res = db.execute_query("SELECT company_curency FROM company LIMIT 1")
        if res and res[0].get('company_curency'):
            base_currency = res[0]['company_curency']
    except Exception:
        pass
    try:
        currencies = db.execute_query("SELECT currency_code, currency_name FROM currency_table") or []
    except Exception:
        currencies = []

    return render_template('invoice_creating.html',
                           customers=customers,
                           locations=locations,
                           jobs=jobs,
                           inventory_items=items,
                           today_date=today_date,
                           base_currency=base_currency,
                           currencies=currencies,
                           last_jv=request.args.get('last_jv'),
                           last_inv=request.args.get('last_inv'))

def build_invoice_context(invoice_no):
    """Fetch and compute everything needed to render an invoice.
    Returns a context dict, or None if the invoice does not exist.
    Shared by the classic invoice print and the 2026 VAT Tax Invoice format."""
    # Fetch Invoice Header from Invoice_Oustanding
    header = db.execute_query("""
        SELECT o.invoice_number, o.invoice_date, o.invoice_final_date as due_date,
               o.invoice_buinding_Customer as customer_id, o.invoice_JV,
               o.VAT_rate as vat_rate
        FROM Invoice_Oustanding o
        WHERE o.invoice_number = %s
        LIMIT 1
    """, (invoice_no,))

    if not header:
        return None

    header = header[0]

    # Fetch Customer Info
    customer = db.execute_query("""
        SELECT supplier_name, supplier_address_1, supplier_address_2, supplier_address_3,
               suppliers_TIN, suppliers_vat_regidter_no, suppliers_teli_1
        FROM suppliers
        WHERE sup_id = %s
        LIMIT 1
    """, (header['customer_id'],))
    customer = customer[0] if customer else {}

    # Fetch Company Info
    company = db.execute_query("""
        SELECT company_name, company_land_line, company_addras_1, company_addras_2, company_addras_3, vat_registered, company_vate_code
        FROM company
        LIMIT 1
    """)
    company = company[0] if company else {}

    # Fetch Invoice Line Items
    items = db.execute_query("""
        SELECT Item_Name as invoice_item_name, Qty as invoice_qty, Pricing as invoice_price, mesurment as invoice_unit
        FROM Invoice_Recode
        WHERE JV_No = %s
    """, (header['invoice_JV'],))

    # Determine VAT Compliance & Invoice Type
    company_vat = company.get('company_vate_code') and str(company['company_vate_code']).strip()

    # Customer can have VAT in TIN or VAT NO field
    cust_vat = (customer.get('suppliers_TIN') and str(customer['suppliers_TIN']).strip()) or \
               (customer.get('suppliers_vat_regidter_no') and str(customer['suppliers_vat_regidter_no']).strip())

    vat_rate = header['vat_rate'] or 0.0
    subtotal = 0.0
    vat_amount = 0.0

    if company_vat and cust_vat and vat_rate > 0:
        # CASE 1: Both registered -> TAX INVOICE with explicit VAT
        invoice_title = "TAX INVOICE"
        for item in items:
            raw_total = item['invoice_qty'] * item['invoice_price']
            item['total'] = raw_total
            subtotal += item['total']

        vat_amount = (subtotal * vat_rate) / 100
        grand_total = subtotal + vat_amount

    elif company_vat and not cust_vat and vat_rate > 0:
        # CASE 2: Company registered, Customer NOT -> INVOICE with VAT rolled into items
        invoice_title = "INVOICE"
        for item in items:
            raw_total = item['invoice_qty'] * item['invoice_price']
            item_vat_amount = (raw_total * vat_rate) / 100

            # Roll VAT into line items
            item['total'] = raw_total + item_vat_amount
            item['invoice_price'] = item['invoice_price'] + ((item['invoice_price'] * vat_rate) / 100)

            subtotal += item['total']

        grand_total = subtotal
        vat_rate = 0.0 # Hide explicit VAT display
        vat_amount = 0.0

    else:
        # CASE 3: Company not registered -> Standard INVOICE, no VAT charged
        invoice_title = "INVOICE"
        vat_rate = 0.0
        for item in items:
            raw_total = item['invoice_qty'] * item['invoice_price']
            item['total'] = raw_total
            subtotal += item['total']

        grand_total = subtotal

    # Fetch Terms and Conditions
    invoice_terms = ""
    try:
        terms_res = db.execute_query("SELECT setting_value FROM system_settings WHERE setting_key = 'invoice_terms_conditions'")
        if terms_res and isinstance(terms_res, list) and len(terms_res) > 0 and 'setting_value' in terms_res[0]:
            invoice_terms = terms_res[0]['setting_value'] or ""
    except Exception as e:
        print(f"Error fetching invoice_terms for print: {e}")

    return dict(
        header=header,
        customer=customer,
        company=company,
        items=items,
        subtotal=subtotal,
        vat_rate=vat_rate,
        vat_amount=vat_amount,
        grand_total=grand_total,
        invoice_title=invoice_title,
        invoice_terms=invoice_terms,
    )


@app.route('/invoice_print/<string:invoice_no>')
@login_required
@has_permission('Access_Accounting')
def invoice_print(invoice_no):
    ctx = build_invoice_context(invoice_no)
    if ctx is None:
        flash("Invoice not found.", "danger")
        return redirect(url_for('invoice_creating'))
    return render_template('invoice_print.html', **ctx)


@app.route('/tax_invoice_2026/<string:invoice_no>')
@login_required
@has_permission('Access_Accounting')
def tax_invoice_2026(invoice_no):
    """Render an invoice in the new VAT Tax Invoice format mandated by
    Gazette Extraordinary No. 2463/05 (2025-11-17), effective 2026-01-01."""
    ctx = build_invoice_context(invoice_no)
    if ctx is None:
        flash("Invoice not found.", "danger")
        return redirect(url_for('invoice_creating'))

    from datetime import datetime as _dt
    import re as _re

    # Gazette requires values in LKR without cents
    net   = round(ctx['subtotal'] or 0)
    vat   = round(ctx['vat_amount'] or 0)
    total = round(ctx['grand_total'] or 0)

    # Resolve invoice date
    inv_date = ctx['header'].get('invoice_date')
    def _to_dt(v):
        try:
            if isinstance(v, str):
                return _dt.strptime(v[:10], '%Y-%m-%d')
            return _dt(v.year, v.month, v.day)
        except Exception:
            return None
    d = _to_dt(inv_date) or _dt.now()

    def _fmt_mmddyyyy(v):
        dd = _to_dt(v)
        return dd.strftime('%m/%d/%Y') if dd else (v or '')

    # Suggested gazette serial number: YYMMM_QQQQ_XXXXX (max 40 chars, no spaces)
    yy  = d.strftime('%y')
    mmm = d.strftime('%b').upper()
    digits = _re.sub(r'\D', '', str(ctx['header'].get('invoice_number') or '')) or '1'
    xxxxx = str(int(digits))
    branch_code = 'MAIN'
    gazette_serial = f"{yy}{mmm}_{branch_code}_{xxxxx}"[:40]

    ctx.update(
        net_amount=net,
        vat_display_amount=vat,
        total_incl_vat=total,
        total_in_words=amount_in_words(total),
        gazette_serial=gazette_serial,
        invoice_date_fmt=_fmt_mmddyyyy(inv_date),
        delivery_date_fmt=_fmt_mmddyyyy(inv_date),
        vat_percent=(ctx['vat_rate'] or 18),
    )
    return render_template('tax_invoice_2026.html', **ctx)


@app.route('/download/vat_invoice_format_2026')
@login_required
def download_vat_invoice_format_2026():
    """Download the official Gazette 2463/05 (2025-11-17) Tax Invoice specification PDF."""
    from flask import send_from_directory
    docs_dir = os.path.join(app.root_path, 'static', 'docs')
    return send_from_directory(
        docs_dir, 'vat_tax_invoice_format_2026.pdf',
        as_attachment=True,
        download_name='VAT_Tax_Invoice_Format_2026_Gazette_2463-05.pdf'
    )

@app.route('/api/get_item_prices/<string:item_ids>')
@login_required
def api_get_item_prices(item_ids):
    # Fetch all prices (selling, special, etc) for selection logic if multiple
    # Simplified: Returning selling price. If multiple pricing structure exists in `inventory_price_recod`, adjust here.
    # Accepts comma-separated item IDs to avoid N+1 query problems.

    ids = [i.strip() for i in item_ids.split(',') if i.strip().isdigit()]
    if not ids:
        return json.dumps({})

    placeholders = ', '.join(['%s'] * len(ids))
    query = f"SELECT inventory_price_link, inventory_price_selling FROM inventory_price_recod WHERE inventory_price_link IN ({placeholders})"

    prices = db.execute_query(query, tuple(ids))

    price_dict = {}
    for p in prices:
        link_id = str(p['inventory_price_link'])
        if link_id not in price_dict:
            price_dict[link_id] = []
        price_dict[link_id].append(p['inventory_price_selling'])

    return json.dumps(price_dict)

def calculate_invoice_totals(inv_items, non_inv_items, vat_rate, apply_vat):
    total_sales = 0
    total_cost = 0

    # Inventory Items
    for item in inv_items:
        qty = parse_float(item.get('qty', 0))
        price = parse_float(item.get('price', 0))
        cost = parse_float(item.get('cost', 0))
        total_sales += qty * price
        total_cost += qty * cost

    # Non-Inventory Items
    for item in non_inv_items:
        qty = parse_float(item.get('qty', 0))
        price = parse_float(item.get('price', 0))
        total_sales += qty * price

    vat_amount = 0
    grand_total = total_sales
    if apply_vat:
        vat_amount = (total_sales * vat_rate) / 100
        grand_total += vat_amount

    return {
        'total_sales': total_sales,
        'total_cost': total_cost,
        'vat_amount': vat_amount,
        'grand_total': grand_total
    }

def generate_invoice_number(cursor):
    cursor.execute("SELECT COALESCE(MAX(CAST(SUBSTRING(invoice_no, 5) AS UNSIGNED)), 0) FROM customer_outstanding")
    max_inv = cursor.fetchone()[0]
    return f"INV-{max_inv + 1:05d}"

def create_invoice_jv(cursor, current_user, narration):
    cursor.execute("INSERT INTO jv_numbers (jv_user_code, jv_naration) VALUES (%s, %s)",
                   (str(current_user), narration))
    return cursor.lastrowid

@dataclass
class OutstandingRecordContext:
    cursor: typing.Any
    invoice_no: str
    inv_date: str
    grand_total: float
    due_date: str
    cust_name: str
    jv_no: str
    vat_rate: float
    currency: str = ''
    exchange_rate: float = 1.0

def create_outstanding_record(ctx: OutstandingRecordContext):
    ctx.cursor.execute("SELECT sup_id FROM suppliers WHERE supplier_name = %s LIMIT 1", (ctx.cust_name,))
    res = ctx.cursor.fetchone()
    cust_id = res[0] if res else 0

    ctx.cursor.execute("""
        INSERT INTO Invoice_Oustanding (
            invoice_number, invoice_date, invoice_total_oustanding,
            invoice_oustanding_Patment, invoice_final_date,
            invoice_buinding_Customer, invoice_JV, VAT_rate, oustanding_delete,
            invoice_currency, invoice_exchange_rate
        ) VALUES (%s, %s, %s, 0, %s, %s, %s, %s, 0, %s, %s)
    """, (ctx.invoice_no, ctx.inv_date, ctx.grand_total, ctx.due_date, cust_id, ctx.jv_no,
          ctx.vat_rate, (ctx.currency or None), (ctx.exchange_rate or 1.0)))
    return ctx.cursor.lastrowid


@dataclass
class InvoiceBatchContext:
    cursor: typing.Any
    inv_items: list
    non_inv_items: list
    jv_no: str
    current_user: int
    customer_name: str
    outstanding_id: int
    location: str
    inv_date: str
    invoice_no: str

def process_invoice_items_batch(ctx: InvoiceBatchContext):
    # Prepare batch data
    invoice_recode_batch = []
    inventory_recode_batch = []

    current_date = datetime.now().strftime('%Y-%m-%d')

    # Pre-fetch warranty periods for all items in batch
    warranty_map = {}
    if ctx.inv_items:
        item_names = list(set([item.get('name') for item in ctx.inv_items if item.get('name')]))
        if item_names:
            format_strings = ','.join(['%s'] * len(item_names))
            ctx.cursor.execute(f"""
                SELECT name, yeas_, month, date_ FROM inventory_vorenty_period
                WHERE name IN ({format_strings})
            """, tuple(item_names))
            for row in ctx.cursor.fetchall():
                warranty_map[row[0]] = (row[1], row[2], row[3])

    # Inventory Items
    for item in ctx.inv_items:
        # Add to invoice_recode (Note: WPF code uses table `invoice_recode` - wait, schema says `Invoice_Recode`)
        # Check schema capitalization. Given previous tables, sticking to lowercase match if possible or schema name.
        # Schema: Invoice_Recode

        # Warranty Logic (Preserved but optimized to only run query)
        # Fetch warranty period for item
        w_end_date = None
        w_res = warranty_map.get(item.get('name'))
        if w_res:
            try:
                years, months, days = w_res
                # Logic retained from original code (pass)
                pass
            except Exception as e:
                logging.error(f"Error parsing warranty for item '{item.get('name')}': {e}")
                pass

        # Add to batch for Invoice_Recode
        invoice_recode_batch.append((
            item['name'], parse_float(item.get('qty', 0)), parse_float(item.get('price', 0)), 1, 'Being account of customer sales', ctx.jv_no, ctx.current_user,
            ctx.customer_name, 1, ctx.outstanding_id, item['unit'], current_date
        ))

        # Add to batch for Inventory_Recod
        inventory_recode_batch.append((
            item['name'], item['code'], item['unit'], item['cost'] * parse_float(item['qty']), parse_float(item['qty']),
            ctx.current_user, current_date, ctx.location, ctx.inv_date, ctx.jv_no, ctx.outstanding_id, ctx.invoice_no
        ))

    # Non-Inventory Items
    for item in ctx.non_inv_items:
        # Add to batch for Invoice_Recode
        invoice_recode_batch.append((
            item['name'], parse_float(item.get('qty', 0)), parse_float(item.get('price', 0)), 0, 'Being account of customer sales', ctx.jv_no, ctx.current_user,
            ctx.customer_name, 1, ctx.outstanding_id, item['unit'], current_date
        ))

    # Execute Batch Inserts
    if invoice_recode_batch:
        ctx.cursor.executemany("""
            INSERT INTO Invoice_Recode (
                Item_Name, Qty, Pricing, Inventory_Items_Or_Not, Natation, JV_No,
                User, Customer_Name, Save_Or_Not, Buinding_To_Oustanding, mesurment,
                recode_date
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, invoice_recode_batch)

    if inventory_recode_batch:
        ctx.cursor.executemany("""
            INSERT INTO inventory_recod (
                inventoy_name, inventoy_code, inventory_recod_mesrmet,
                inventory_recod_unit_price, inventory_recod_movment_out,
                inventory_recod_account, inventory_recod_user_id,
                inventory_recod_user_recod_date, inventory_recod_location,
                inventory_recod_action_date, inventory_recodcol_memo, JV_No,
                inventory_recod_link_invoice, inventory_recod_suplier_iv_no
            ) VALUES (%s, %s, %s, %s, %s, 'Inventoy', %s, %s, %s, %s, 'Credit Sales', %s, %s, %s)
        """, inventory_recode_batch)

@dataclass
class InvoiceGLContext:
    cursor: object
    current_user: int
    jv_no: str
    invoice_date: str
    job_no: str
    totals: dict

def post_invoice_gl_entries(ctx: InvoiceGLContext):
    job_no_val = ctx.job_no if ctx.job_no else None

    # DR Account Receivable (Total + VAT)
    ctx.cursor.execute("""
        INSERT INTO entry_details (
            account_name, enty_values_DR, entry_effective_date, entry_create_date,
            entry_naration, entry_create_user, entry_jv, entry_job_number
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
    """, ('Account Receivable', ctx.totals['grand_total'], ctx.invoice_date, datetime.now().date(), "Credit Sale", ctx.current_user, ctx.jv_no, job_no_val))

    # CR Income (Sales)
    ctx.cursor.execute("""
        INSERT INTO entry_details (
            account_name, enty_values_CR, entry_effective_date, entry_create_date,
            entry_naration, entry_create_user, entry_jv, entry_job_number
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
    """, ('Sales', ctx.totals['total_sales'], ctx.invoice_date, datetime.now().date(), "Credit Sale", ctx.current_user, ctx.jv_no, job_no_val))

    # CR VAT (If any)
    if ctx.totals['vat_amount'] > 0:
        ctx.cursor.execute("""
            INSERT INTO entry_details (
                account_name, enty_values_CR, entry_effective_date, entry_create_date,
                entry_naration, entry_create_user, entry_jv, entry_job_number
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, ('VAT Control', ctx.totals['vat_amount'], ctx.invoice_date, datetime.now().date(), "Credit Sale", ctx.current_user, ctx.jv_no, job_no_val))

    # Cost of Goods Sold (If inventory items exist)
    if ctx.totals['total_cost'] > 0:
            # DR COGS
        ctx.cursor.execute("""
            INSERT INTO entry_details (
                account_name, enty_values_DR, entry_effective_date, entry_create_date,
                entry_naration, entry_create_user, entry_jv, entry_job_number
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, ('Cost Of Goods Sold', ctx.totals['total_cost'], ctx.invoice_date, datetime.now().date(), "Credit Sale", ctx.current_user, ctx.jv_no, job_no_val))

        # CR Inventory
        ctx.cursor.execute("""
            INSERT INTO entry_details (
                account_name, enty_values_CR, entry_effective_date, entry_create_date,
                entry_naration, entry_create_user, entry_jv, entry_job_number
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, ('Inventory', ctx.totals['total_cost'], ctx.invoice_date, datetime.now().date(), "Credit Sale", ctx.current_user, ctx.jv_no, job_no_val))

def parse_invoice_form_data(form):
    customer_name = form.get('customer')
    inv_date = form.get('invoice_date')

    if not customer_name or not inv_date:
        return None, 'Customer and Invoice Date are required.'

    location = form.get('location')
    due_date = form.get('due_date')
    job_no = form.get('job_no')
    vat_rate = parse_float(form.get('vat_rate', 0))
    apply_vat = 1 if form.get('apply_vat') else 0

    # Multi-currency: the invoice is billed in this currency; exchange_rate converts to base
    invoice_currency = (form.get('invoice_currency') or '').strip()
    exchange_rate = parse_float(form.get('exchange_rate', 1)) or 1.0
    if exchange_rate <= 0:
        exchange_rate = 1.0

    inv_items_json = form.get('inventory_items_json')
    non_inv_items_json = form.get('non_inventory_items_json')

    try:
        inv_items = json.loads(inv_items_json) if inv_items_json else []
        non_inv_items = json.loads(non_inv_items_json) if non_inv_items_json else []
    except json.JSONDecodeError as e:
        return None, f'Error processing item list: {str(e)}'

    if not inv_items and not non_inv_items:
        return None, 'No items in invoice'

    return {
        'customer_name': customer_name,
        'inv_date': inv_date,
        'location': location,
        'due_date': due_date,
        'job_no': job_no,
        'vat_rate': vat_rate,
        'apply_vat': apply_vat,
        'inv_items': inv_items,
        'non_inv_items': non_inv_items,
        'invoice_currency': invoice_currency,
        'exchange_rate': exchange_rate
    }, None

@app.route('/invoice_creating/submit', methods=['POST'])
@login_required
def submit_invoice():
    # 1. Validation & Input Parsing
    parsed_data, error_msg = parse_invoice_form_data(request.form) if hasattr(request, 'form') else parse_invoice_form_data(request)

    if error_msg:
        flash(error_msg, 'danger')
        return redirect(url_for('invoice_creating'))

    customer_name = parsed_data['customer_name']
    inv_date = parsed_data['inv_date']
    location = parsed_data['location']
    due_date = parsed_data['due_date']
    job_no = parsed_data['job_no']
    vat_rate = parsed_data['vat_rate']
    apply_vat = parsed_data['apply_vat']
    inv_items = parsed_data['inv_items']
    non_inv_items = parsed_data['non_inv_items']

    current_user = get_current_user_id()
    # current_user_pk is unused in the transaction below

    # 2. Validate VAT Registration Rule (Warnings Only)
    if apply_vat == 'Yes' and vat_rate > 0:
        company_res = db.execute_query("SELECT company_vate_code FROM company LIMIT 1")
        company_vat = company_res[0]['company_vate_code'] if company_res else None

        cust_res = db.execute_query("SELECT suppliers_TIN, suppliers_vat_regidter_no FROM suppliers WHERE sup_id = %s", (customer_name,))
        cust_vat = None
        if cust_res:
            cust_vat = (cust_res[0].get('suppliers_TIN') and str(cust_res[0]['suppliers_TIN']).strip()) or \
                       (cust_res[0].get('suppliers_vat_regidter_no') and str(cust_res[0]['suppliers_vat_regidter_no']).strip())

        if not company_vat or not str(company_vat).strip():
            flash('Warning: Your Company is not VAT registered. The invoice will be processed at standard selling prices without adding a VAT component.', 'warning')
            vat_rate = 0.0
            apply_vat = 'No'
        elif not cust_vat:
            flash('Notice: The Customer is not VAT registered. A commercial invoice will be generated with VAT-inclusive selling prices.', 'info')

    # 3. Database Transaction
    conn = db.get_connection()
    if not conn:
        flash('Database connection failed.', 'danger')
        return redirect(url_for('invoice_creating'))

    cursor = conn.cursor()
    conn.start_transaction()

    try:
        # Check Payment Type
        payment_type = request.form.get('payment_type', 'Credit')

        # 3. Generate Invoice No (Credit_Invoice_No table) — manual override if provided
        cursor.execute("INSERT INTO Credit_Invoice_No (id) VALUES (0)")
        inv_id_seq = cursor.lastrowid
        manual_inv_no = (request.form.get('manual_invoice_no') or '').strip()
        invoice_no = manual_inv_no if manual_inv_no else f"IV-{datetime.now().year}{datetime.now().month}-{inv_id_seq}"

        # 4. Create JV Header
        cursor.execute("INSERT INTO jv_numbers (jv_user_code, jv_naration) VALUES (%s, %s)",
                       (str(current_user), "Credit Sales"))
        jv_no = cursor.lastrowid

        # 5. Calculate Totals (in the invoice currency) then convert to base for the GL
        inv_currency = parsed_data.get('invoice_currency') or ''
        rate = parsed_data.get('exchange_rate') or 1.0

        totals = calculate_invoice_totals(inv_items, non_inv_items, vat_rate, apply_vat)
        # Sales / VAT / grand total are in the invoice currency -> multiply by rate for base (LKR).
        # Cost (COGS) is already stored in base currency, so it is NOT converted.
        total_sales = round(totals['total_sales'] * rate, 2)
        vat_amount  = round(totals['vat_amount'] * rate, 2)
        grand_total = round(totals['grand_total'] * rate, 2)
        total_cost  = totals['total_cost']

        # 6. Insert Outstanding Record (stored in base currency, with FC info for reference)
        ctx = OutstandingRecordContext(
            cursor=cursor,
            invoice_no=invoice_no,
            inv_date=inv_date,
            grand_total=grand_total,
            due_date=due_date,
            cust_name=customer_name,
            jv_no=jv_no,
            vat_rate=vat_rate,
            currency=inv_currency,
            exchange_rate=rate
        )
        outstanding_id = create_outstanding_record(ctx)

        # 7. Insert Invoice Records (Details) & Update Inventory
        batch_ctx = InvoiceBatchContext(
            cursor=cursor,
            inv_items=inv_items,
            non_inv_items=non_inv_items,
            jv_no=jv_no,
            current_user=current_user,
            customer_name=customer_name,
            outstanding_id=outstanding_id,
            location=location,
            inv_date=inv_date,
            invoice_no=invoice_no
        )
        process_invoice_items_batch(batch_ctx)

        # 8. GL Entries
        gl_ctx = InvoiceGLContext(
            cursor=cursor,
            current_user=current_user,
            jv_no=jv_no,
            invoice_date=inv_date,
            job_no=job_no,
            totals={
                'grand_total': grand_total,
                'total_sales': total_sales,
                'vat_amount': vat_amount,
                'total_cost': total_cost
            }
        )
        post_invoice_gl_entries(gl_ctx)
        conn.commit()
        flash(f'Invoice {invoice_no} created successfully.|{invoice_no}', 'success')
        _inv_jv = jv_no
        _inv_no = invoice_no

    except Exception as e:
        conn.rollback()
        flash(f'Transaction failed: {str(e)}', 'danger')
        print(f"Invoice Error: {e}")
        _inv_jv = None
        _inv_no = None
    finally:
        cursor.close()
        conn.close()

    if _inv_jv:
        return redirect(url_for('invoice_creating', last_jv=_inv_jv, last_inv=_inv_no))
    return redirect(url_for('invoice_creating'))

# --- Inventory Production ---
@app.route('/inventory_production', methods=['GET'])
@login_required
@has_permission('Access_Inventory')
def inventory_production():
    locations = db.execute_query("SELECT inventory_locations_name FROM inventory_locations")
    jobs = db.execute_query("SELECT job_number FROM jobs_unit")

    # Accounts for Cost/Expense selection (PL accounts)
    expense_accounts = db.execute_query("SELECT account_name FROM new_account_table WHERE account_expenses = 1 OR account_assets = 1")

    items = db.execute_query("""
        SELECT i.inventoy_name, i.inventoy_code, i.inventoy_items_messurment_unit, p.inventory_price_purcharsing
        FROM inventoy_items i
        LEFT JOIN inventory_price_recod p ON i.id = p.inventory_price_link
        WHERE i.active = 1
    """)

    # Production history — identified by the TF-Prod- marker on inventory records
    history = db.execute_query("""
        SELECT
            r.JV_No                                   AS jv,
            j.jv_naration                             AS narration,
            MIN(r.inventory_recod_action_date)        AS date,
            MIN(r.inventory_recod_location)           AS location,
            SUM(COALESCE(r.inventory_recod_moument_in,0))  AS qty_in,
            SUM(COALESCE(r.inventory_recod_movment_out,0)) AS qty_out,
            SUM(COALESCE(r.inventory_recod_moument_in,0) * COALESCE(r.inventory_recod_unit_price,0)
              + COALESCE(r.inventory_recod_movment_out,0) * COALESCE(r.inventory_recod_unit_price,0)) AS value,
            CASE WHEN SUM(COALESCE(r.inventory_recod_moument_in,0)) > 0 THEN 'Receipt' ELSE 'Issue' END AS type,
            (SELECT COUNT(*) FROM jv_numbers rj WHERE rj.jv_user_code = CONCAT('REV-PROD-', r.JV_No)) AS rev_count
        FROM inventory_recod r
        LEFT JOIN jv_numbers j ON r.JV_No = j.jv_id
        WHERE r.inventory_recod_suplier_iv_no LIKE 'TF-Prod-%%'
        GROUP BY r.JV_No, j.jv_naration
        ORDER BY r.JV_No DESC
        LIMIT 100
    """) or []

    return render_template('inventory_production.html',
                           locations=locations, jobs=jobs, items=items,
                           expense_accounts=expense_accounts,
                           today_date=date.today().strftime('%Y-%m-%d'),
                           history=history)


@app.route('/inventory_production/reverse', methods=['POST'])
@login_required
@has_permission('Access_Reversals')
def inventory_production_reverse():
    """Reverse a production issue/receipt: opposite inventory movement +
    swapped GL entries. Blocked for closed periods and double-reversal."""
    jv = request.form.get('jv', '').strip()
    if not jv:
        flash('No production transaction selected', 'danger')
        return redirect(url_for('inventory_production'))

    if jv_in_closed_period(jv):
        flash(f'Cannot reverse: this transaction is in a closed period (locked through {get_period_lock_date()}).', 'danger')
        return redirect(url_for('inventory_production'))

    current_user_pk = get_current_user_pk()
    today = date.today()
    conn = None
    cursor = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        # Guard: already reversed?
        cursor.execute("SELECT COUNT(*) FROM jv_numbers WHERE jv_user_code = %s", (f'REV-PROD-{jv}',))
        if cursor.fetchone()[0] > 0:
            flash('This production transaction has already been reversed.', 'warning')
            return redirect(url_for('inventory_production'))

        # Guard: exists?
        cursor.execute("SELECT COUNT(*) FROM inventory_recod WHERE JV_No = %s AND inventory_recod_suplier_iv_no LIKE 'TF-Prod-%%'", (jv,))
        if cursor.fetchone()[0] == 0:
            flash('Production transaction not found.', 'danger')
            return redirect(url_for('inventory_production'))

        conn.start_transaction()

        # 1. Reversal JV
        cursor.execute("INSERT INTO jv_numbers (jv_user_code, jv_naration, status) VALUES (%s, %s, 1)",
                       (f'REV-PROD-{jv}', f'Production Reversal of JV-{jv}'))
        rev_jv = cursor.lastrowid

        # 2. Reverse inventory movements — swap IN<->OUT
        cursor.execute("""
            INSERT INTO inventory_recod (
                inventoy_name, inventoy_code, inventory_recod_mesrmet,
                inventory_recod_unit_price, inventory_recod_moument_in, inventory_recod_movment_out,
                inventory_recod_suplier_iv_no, inventory_recod_user_id,
                inventory_recod_user_recod_date, inventory_recod_location,
                inventory_recod_action_date, inventory_recodcol_memo, JV_No
            )
            SELECT
                inventoy_name, inventoy_code, inventory_recod_mesrmet,
                inventory_recod_unit_price,
                COALESCE(inventory_recod_movment_out,0),   -- OUT becomes IN
                COALESCE(inventory_recod_moument_in,0),    -- IN becomes OUT
                CONCAT('REV-', %s), %s, %s, inventory_recod_location,
                %s, 'Production Reversal', %s
            FROM inventory_recod
            WHERE JV_No = %s AND inventory_recod_suplier_iv_no LIKE 'TF-Prod-%%'
        """, (jv, current_user_pk, today, today, rev_jv, jv))

        # 3. Reverse GL entries — swap DR<->CR
        cursor.execute("""
            INSERT INTO entry_details (
                account_name, enty_values_DR, enty_values_CR,
                entry_effective_date, entry_create_date,
                entry_naration, entry_create_user, entry_jv
            )
            SELECT account_name,
                   COALESCE(enty_values_CR,0), COALESCE(enty_values_DR,0),
                   %s, %s, %s, %s, %s
            FROM entry_details WHERE entry_jv = %s
        """, (today, today, f'Production Reversal of JV-{jv}', current_user_pk, rev_jv, jv))

        conn.commit()
        flash(f'Production transaction (JV: {jv}) reversed successfully. Reversal JV: {rev_jv}', 'success')
    except Exception as e:
        if conn: conn.rollback()
        logging.error(f"Production reversal error: {e}")
        flash(f'Error reversing production transaction: {str(e)}', 'danger')
    finally:
        if cursor: cursor.close()
        if conn: conn.close()
    return redirect(url_for('inventory_production'))

@app.route('/inventory_production/issue', methods=['POST'])
@login_required
@has_permission('Access_Inventory')
def submit_production_issue():
    # Logic: Stock OUT, Dr Expense, Cr Inventory
    try:
        date_val = request.form.get('issue_date')
        job_no = request.form.get('job_no')
        source_loc = request.form.get('source_location')
        dr_account = request.form.get('cost_account') # User selected Expense Account
        narration = request.form.get('narration')

        item_names = request.form.getlist('item_name[]')
        item_codes = request.form.getlist('item_code[]')
        item_units = request.form.getlist('item_unit[]')
        item_costs = request.form.getlist('unit_cost[]') # or fetched from DB if readonly
        qtys = request.form.getlist('qty[]')

        if not item_names:
            flash('No items to issue', 'danger')
            return redirect(url_for('inventory_production'))

        current_user = get_current_user_id()
        current_user_pk = get_current_user_pk()

        conn = db.get_connection()
        cursor = conn.cursor()
        conn.start_transaction()

        try:
            # 1. Create JV
            cursor.execute("INSERT INTO jv_numbers (jv_user_code, jv_naration) VALUES (%s, %s)",
                           ('JV FROM PRODUCTION ISSUE', f"Production Issue: {narration}"))
            jv_no = cursor.lastrowid

            total_value = 0

            # 2. Process Items (Stock OUT)
            for i in range(len(item_names)):
                qty = float(qtys[i])
                # Note: WPF code allows user to select price/cost or takes it from grid.
                # Here we take from form (which defaults to DB cost but is editable or hidden)
                cost = float(item_costs[i] or 0)

                if qty <= 0: continue

                line_val = qty * cost
                total_value += line_val

                # Stock OUT
                cursor.execute("""
                    INSERT INTO inventory_recod (
                        inventoy_name, inventoy_code, inventory_recod_mesrmet,
                        inventory_recod_unit_price, inventory_recod_movment_out,
                        inventory_recod_suplier_iv_no, inventory_recod_user_id,
                        inventory_recod_user_recod_date, inventory_recod_location,
                        inventory_recod_action_date, inventory_recodcol_memo, JV_No
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    item_names[i], item_codes[i], item_units[i], cost, qty,
                    f"TF-Prod-{jv_no}", current_user_pk, datetime.now().date(), source_loc,
                    date_val, narration, jv_no
                ))

            # 3. GL Entries
            # Dr User Selected Account (Expense/WIP)
            cursor.execute("""
                INSERT INTO entry_details (
                    account_name, enty_values_DR, entry_effective_date, entry_create_date,
                    entry_naration, entry_create_user, entry_jv, entry_job_number
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (dr_account, total_value, date_val, datetime.now().date(), narration, current_user_pk, jv_no, job_no if job_no else None))

            # Cr Inventory Control Account
            cursor.execute("""
                INSERT INTO entry_details (
                    account_name, enty_values_CR, entry_effective_date, entry_create_date,
                    entry_naration, entry_create_user, entry_jv, entry_job_number
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, ('Inventory', total_value, date_val, datetime.now().date(), narration, current_user_pk, jv_no, job_no if job_no else None))

            conn.commit()
            flash(f'Production Issue recorded successfully. JV: {jv_no}', 'success')

        except Exception as e:
            conn.rollback()
            flash(f'Error processing issue: {str(e)}', 'danger')
        finally:
            cursor.close()
            conn.close()

    except Exception as e:
        flash(f'System Error: {str(e)}', 'danger')

    return redirect(url_for('inventory_production'))

@app.route('/inventory_production/receive', methods=['POST'])
@login_required
@has_permission('Access_Inventory')
def submit_production_receive():
    # Logic: Stock IN, Dr Inventory, Cr Expense/WIP
    try:
        date_val = request.form.get('receive_date')
        job_no = request.form.get('job_no')
        target_loc = request.form.get('target_location')
        cr_account = request.form.get('credit_account') # User selected Cost Source
        narration = request.form.get('narration')

        item_names = request.form.getlist('item_name[]')
        item_codes = request.form.getlist('item_code[]')
        item_units = request.form.getlist('item_unit[]')
        item_costs = request.form.getlist('unit_cost[]')
        qtys = request.form.getlist('qty[]')

        if not item_names:
            flash('No items to receive', 'danger')
            return redirect(url_for('inventory_production'))

        current_user = get_current_user_id()
        current_user_pk = get_current_user_pk()

        conn = db.get_connection()
        cursor = conn.cursor()
        conn.start_transaction()

        try:
            # 1. Create JV
            cursor.execute("INSERT INTO jv_numbers (jv_user_code, jv_naration) VALUES (%s, %s)",
                           ('JV FROM PRODUCTION RECEIPT', f"Production Receipt: {narration}"))
            jv_no = cursor.lastrowid

            total_value = 0

            # 2. Process Items (Stock IN)
            for i in range(len(item_names)):
                qty = float(qtys[i])
                cost = float(item_costs[i] or 0)

                if qty <= 0: continue

                line_val = qty * cost
                total_value += line_val

                # Stock IN
                cursor.execute("""
                    INSERT INTO inventory_recod (
                        inventoy_name, inventoy_code, inventory_recod_mesrmet,
                        inventory_recod_unit_price, inventory_recod_moument_in,
                        inventory_recod_suplier_iv_no, inventory_recod_user_id,
                        inventory_recod_user_recod_date, inventory_recod_location,
                        inventory_recod_action_date, inventory_recodcol_memo, JV_No
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    item_names[i], item_codes[i], item_units[i], cost, qty,
                    f"TF-Prod-{jv_no}", current_user_pk, datetime.now().date(), target_loc,
                    date_val, narration, jv_no
                ))

            # 3. GL Entries
            # Dr Inventory Control Account
            cursor.execute("""
                INSERT INTO entry_details (
                    account_name, enty_values_DR, entry_effective_date, entry_create_date,
                    entry_naration, entry_create_user, entry_jv, entry_job_number
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, ('Inventory', total_value, date_val, datetime.now().date(), narration, current_user_pk, jv_no, job_no if job_no else None))

            # Cr User Selected Account (Expense/WIP)
            cursor.execute("""
                INSERT INTO entry_details (
                    account_name, enty_values_CR, entry_effective_date, entry_create_date,
                    entry_naration, entry_create_user, entry_jv, entry_job_number
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (cr_account, total_value, date_val, datetime.now().date(), narration, current_user_pk, jv_no, job_no if job_no else None))

            # 4. Log to inventory_productions (from WPF Manufacturing_Inventory logic)
            cursor.execute("""
                INSERT INTO inventory_productions (ID, JV_No, Delete_Or_Note, Effective_Date)
                VALUES (0, %s, 0, %s)
            """, (jv_no, date_val))

            conn.commit()
            flash(f'Production Receipt recorded successfully. JV: {jv_no}', 'success')

        except Exception as e:
            conn.rollback()
            flash(f'Error processing receipt: {str(e)}', 'danger')
        finally:
            cursor.close()
            conn.close()

    except Exception as e:
        flash(f'System Error: {str(e)}', 'danger')

    return redirect(url_for('inventory_production'))

def create_db_if_missing():
    """Attempts to create the database if it does not exist."""
    try:
        # Check connection
        try:
            conn = mysql.connector.connect(**db_config)
            conn.close()
            return # Connected successfully
        except mysql.connector.Error:
            pass # Failed, proceed to create

        # Connect without DB name
        temp_config = db_config.copy()
        if 'database' in temp_config:
            del temp_config['database']

        try:
            conn_root = mysql.connector.connect(**temp_config)
            cursor = conn_root.cursor()

            db_name = db_config.get('database', 'Book_keeping')
            if not is_safe_db_name(db_name):
                raise ValueError(f"Invalid database name: {db_name}")
            logging.warning(f"Database '{db_name}' not found or connection failed. Attempting to create...")
            if not is_safe_db_name(db_name):
                raise ValueError("Invalid database name")
            cursor.execute(f"CREATE DATABASE IF NOT EXISTS `{db_name}`")
            conn_root.commit()
            cursor.close()
            conn_root.close()
            logging.info(f"Database '{db_name}' checked/created.")
        except mysql.connector.Error as e:
            if e.errno in (1007, 1044, 1045):
                logging.warning(f"Ignored DB creation error {e.errno} in create_db_if_missing: {e.msg}")
            else:
                raise e
    except Exception as e:
        logging.warning(f"Warning: Could not check/create database: {e}")

def execute_sql_file(cursor, filepath, db_name=None):
    """Parses and executes a MySQL dump file with DELIMITER support."""
    import re
    logging.info(f"Executing SQL file: {filepath}")
    with open(filepath, 'r', encoding='utf-8') as f:
        # Read lines to handle DELIMITER command which is line-based
        content = f.read()
        if db_name:
            content = re.sub(r'(?i)Book_keeping', db_name, content)
        lines = content.split('\n')

    delimiter = ';'
    statement = ""

    for line in lines:
        stripped = line.strip()

        # Handle DELIMITER command
        # Syntax: DELIMITER $$
        if stripped.upper().startswith('DELIMITER '):
            delimiter = stripped.split()[1]
            continue

        # Skip comments (simple check)
        if stripped.startswith('--') or stripped.startswith('#'):
            continue

        # Skip empty lines if statement is empty
        if not statement and not stripped:
            continue

        statement += line + "\n"

        # Check if statement ends with delimiter
        if statement.strip().endswith(delimiter):
            # Clean up statement
            sql_to_run = statement.strip()
            # Remove the delimiter from the end
            if sql_to_run.endswith(delimiter):
                 sql_to_run = sql_to_run[:-len(delimiter)]

            if sql_to_run.strip():
                try:
                    cursor.execute(sql_to_run)
                    # consume results if any
                    while cursor.nextset(): pass
                except mysql.connector.Error as e:
                    # Ignore "Table already exists" or similar non-critical if robust
                    # But for initial schema, we usually want to know.
                    # Warning: USE command might fail if user doesn't have perm, but we are inside DB context usually
                    logging.warning(f"SQL Execution Warning: {e}\nStatement partial: {sql_to_run[:100]}...")

            statement = ""

def import_initial_schema():
    """Imports database_schema.sql if Login_Table is missing, using Python parser."""
    try:
        conn = db.get_connection()
        if not conn: return
        cursor = conn.cursor()

        cursor.execute("SHOW TABLES LIKE 'Login_Table'")
        if cursor.fetchone():
            cursor.close()
            conn.close()
            return # Schema likely exists

        logging.info("Login_Table missing. Attempting to import initial schema...")

        default_db_name = db_config.get('database')

        if os.path.exists('database_schema.sql'):
            try:
                execute_sql_file(cursor, 'database_schema.sql', db_name=default_db_name)
                logging.info("Schema imported successfully.")

                if os.path.exists('fixed_assets.sql'):
                    execute_sql_file(cursor, 'fixed_assets.sql', db_name=default_db_name)
                    logging.info("Fixed Assets schema imported.")

                conn.commit()
            except Exception as ex:
                logging.error(f"Failed to execute SQL file: {ex}")
                if conn: conn.rollback()
        else:
            logging.warning("database_schema.sql not found.")

        cursor.close()
        conn.close()

    except Exception as e:
        logging.error(f"Error importing initial schema: {e}")

@app.route('/sales_summary_customer', methods=['GET'])
@login_required
@has_permission('Access_Reports')
def sales_summary_customer():
    from_date = request.args.get('from_date', date.today().replace(day=1).strftime('%Y-%m-%d'))
    to_date = request.args.get('to_date', date.today().strftime('%Y-%m-%d'))

    query = """
        SELECT
            c.customer_code,
            c.customer_name,
            COUNT(io.Id) as invoice_count,
            SUM(io.invoice_total_oustanding) as total_sales,
            SUM(io.invoice_oustanding_Patment) as total_paid,
            SUM(io.invoice_total_oustanding - io.invoice_oustanding_Patment) as balance_due
        FROM Invoice_Oustanding io
        JOIN customer c ON io.invoice_buinding_Customer = c.id
        WHERE io.invoice_date BETWEEN %s AND %s
        AND io.oustanding_delete = 0
        GROUP BY c.id, c.customer_code, c.customer_name
        ORDER BY total_sales DESC
    """

    rows = db.execute_query(query, (from_date, to_date))

    # Calculate Grand Totals
    totals = {
        'sales': sum(float(r['total_sales'] or 0) for r in rows),
        'paid': sum(float(r['total_paid'] or 0) for r in rows),
        'balance': sum(float(r['balance_due'] or 0) for r in rows),
        'count': sum(int(r['invoice_count'] or 0) for r in rows)
    }

    return render_template('sales_summary_customer.html',
                           rows=rows,
                           totals=totals,
                           from_date=from_date,
                           to_date=to_date)

# Ensure initialization runs once regardless of startup method
app_initialized = False

@app.before_request
def initialize_app():
    global app_initialized
    if not app_initialized:
        create_db_if_missing()
        setup_master_db()
        import_initial_schema()
        run_schema_migrations()
        ensure_default_categories()
        create_default_user()
        ensure_default_accounts()
        app_initialized = True


@app.route('/service_entry', methods=['GET'])
@login_required
@has_permission('Access_Accounting')
def service_entry():
    suppliers = db.execute_query("SELECT sup_id, supplier_name, supplier_code FROM suppliers WHERE Is_Suplier = 1")
    accounts = db.execute_query("""
        SELECT account_name
        FROM new_account_table
        WHERE account_active = 1 AND (account_income = 1 OR account_expenses = 1)
    """)
    sub_accounts = db.execute_query("SELECT sub_account_code, sub_sub_accaount_name FROM sub_accont_for_new_account WHERE active = 1")
    jobs = db.execute_query("SELECT job_number FROM jobs_unit WHERE job_finsh = 0 AND job_cancell = 0")

    # ── Service Entry History (with advanced filters, mirrors Direct Purchase) ──
    hf = {
        'search':     request.args.get('h_search', '').strip(),
        'date_from':  request.args.get('h_date_from', '').strip(),
        'date_to':    request.args.get('h_date_to', '').strip(),
        'amount_min': request.args.get('h_amount_min', '').strip(),
        'amount_max': request.args.get('h_amount_max', '').strip(),
        'status':     request.args.get('h_status', '').strip(),  # '', 'active', 'reversed'
    }

    # Default to the CURRENT MONTH on a fresh load so we don't scan/return everything.
    # If the user submits the filter (any h_* param present), respect exactly what they set
    # (clearing both dates then filtering shows all records).
    hist_filter_submitted = any(
        k in request.args for k in
        ('h_search', 'h_date_from', 'h_date_to', 'h_amount_min', 'h_amount_max', 'h_status')
    )
    if not hist_filter_submitted:
        hf['date_from'] = date.today().replace(day=1).strftime('%Y-%m-%d')
        hf['date_to'] = date.today().strftime('%Y-%m-%d')

    h_filters = ["j.jv_user_code LIKE %s"]
    h_params = ['JV FORM SEN INVOICE%']
    if hf['search']:
        h_filters.append("(sup.supplier_name LIKE %s OR s.suppliers_invoice_number LIKE %s OR j.jv_naration LIKE %s)")
        like = f"%{hf['search']}%"
        h_params += [like, like, like]
    if hf['date_from']:
        h_filters.append("s.suppliers_invoice_date >= %s"); h_params.append(hf['date_from'])
    if hf['date_to']:
        h_filters.append("s.suppliers_invoice_date <= %s"); h_params.append(hf['date_to'])
    if hf['amount_min']:
        h_filters.append("s.suppliers_invoice_total_oustanding >= %s"); h_params.append(hf['amount_min'])
    if hf['amount_max']:
        h_filters.append("s.suppliers_invoice_total_oustanding <= %s"); h_params.append(hf['amount_max'])
    if hf['status'] == 'active':
        h_filters.append("s.suppliers_oustanding_delete = 0")
    elif hf['status'] == 'reversed':
        h_filters.append("s.suppliers_oustanding_delete = 1")

    where_clause = " AND ".join(h_filters)
    history = db.execute_query(f"""
        SELECT
            s.suppliers_invoice_JV               AS jv,
            s.suppliers_invoice_number           AS invoice_no,
            s.suppliers_invoice_date             AS date,
            sup.supplier_name                    AS supplier,
            s.suppliers_invoice_total_oustanding AS amount,
            s.suppliers_oustanding_delete        AS is_reversed,
            j.jv_naration                        AS narration
        FROM suppliers_invoice_data s
        JOIN jv_numbers j ON s.suppliers_invoice_JV = j.jv_id
        LEFT JOIN suppliers sup ON s.suppliers_invoice_buinding_supplier = sup.sup_id
        WHERE {where_clause}
        ORDER BY s.s_i_id DESC
        LIMIT 200
    """, tuple(h_params)) or []

    base_currency, currencies = _base_currency_and_list()
    return render_template('service_entry.html',
                           suppliers=suppliers,
                           accounts=accounts,
                           sub_accounts=sub_accounts,
                           jobs=jobs,
                           today_date=date.today().strftime('%Y-%m-%d'),
                           last_jv=request.args.get('last_jv'),
                           history=history,
                           hf=hf,
                           base_currency=base_currency,
                           currencies=currencies)

@app.route('/service_entry/save', methods=['POST'])
@login_required
@has_permission('Access_Accounting')
def save_service_entry():
    try:
        supplier_id = request.form.get('supplier_id')
        effective_date = request.form.get('effective_date')
        invoice_number = request.form.get('invoice_number')
        invoice_date = request.form.get('invoice_date')
        due_date = request.form.get('due_date')
        main_narration = request.form.get('main_narration')
        header_job = request.form.get('header_job_number')
        include_vat = request.form.get('include_vat') == '1'
        vat_rate = parse_float(request.form.get('vat_rate')) if include_vat else 0.0
        entries_json = request.form.get('entries_json')
        total_amount = parse_float(request.form.get('total_amount'))

        # Multi-currency: amounts are entered in this currency; rate converts to base (LKR)
        entry_currency = (request.form.get('entry_currency') or '').strip()
        exchange_rate = parse_float(request.form.get('exchange_rate', 1)) or 1.0
        if exchange_rate <= 0:
            exchange_rate = 1.0
        total_amount_base = round(total_amount * exchange_rate, 2)
        if entry_currency and exchange_rate != 1.0:
            main_narration = (main_narration or '') + f" [{entry_currency} {total_amount:,.2f} @ {exchange_rate:.4f}]"

        entries = json.loads(entries_json) if entries_json else []

        if not entries:
            flash('No entries provided', 'danger')
            return redirect(url_for('service_entry'))

        if not supplier_id or not invoice_number:
            flash('Supplier and Invoice Number are required', 'danger')
            return redirect(url_for('service_entry'))

        current_user = get_current_user_id()
        current_user_pk = get_current_user_pk()

        conn = db.get_connection()
        cursor = conn.cursor()
        conn.start_transaction()

        try:
            # 1. Generate JV Number
            cursor.execute("INSERT INTO jv_numbers (jv_user_code, jv_naration, status) VALUES (%s, %s, %s)",
                           ("JV FORM SEN INVOICE", main_narration, 1))
            jv_no = cursor.lastrowid

            # Get Supplier Details
            cursor.execute("SELECT supplier_code FROM suppliers WHERE sup_id = %s", (supplier_id,))
            sup_res = cursor.fetchone()
            supplier_code = sup_res[0] if sup_res else ""

            # 2. Insert into suppliers_invoice_data
            cursor.execute("""
                INSERT INTO suppliers_invoice_data (
                    suppliers_code, suppliers_invoice_number, suppliers_invoice_date,
                    suppliers_invoice_total_oustanding, suppliers_invoice_total_payment,
                    suppliers_invoice_final_date, suppliers_invoice_buinding_supplier,
                    suppliers_invoice_JV, suppliers_VAT_rate
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                supplier_code, invoice_number, invoice_date,
                total_amount_base, 0,
                due_date, supplier_id,
                jv_no, vat_rate
            ))

            # 3. Journal Entries (posted in base currency)
            # Credit Account Payable
            cursor.execute("""
                INSERT INTO entry_details (
                    account_name, account_code, enty_values_DR, enty_values_CR,
                    entry_effective_date, entry_create_date, entry_naration, entry_create_user,
                    entry_job_number, entry_sub_account_code, entry_jv
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                'Account Payable', 0, 0, total_amount_base,
                effective_date, date.today(), main_narration, current_user_pk,
                header_job if header_job else None, 0, jv_no
            ))

            # Debit VAT Control if applicable (converted to base)
            total_dr_base = sum(parse_float(e['dr']) for e in entries) * exchange_rate
            if include_vat and vat_rate > 0:
                vat_amount = total_dr_base * (vat_rate / 100.0)
                cursor.execute("""
                    INSERT INTO entry_details (
                        account_name, account_code, enty_values_DR, enty_values_CR,
                        entry_effective_date, entry_create_date, entry_naration, entry_create_user,
                        entry_job_number, entry_sub_account_code, entry_jv, entry_VAT
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    'VAT Control', 0, vat_amount, 0,
                    effective_date, date.today(), main_narration, current_user_pk,
                    header_job if header_job else None, 0, jv_no, vat_rate
                ))

            # Debit Expense/P&L Accounts
            for e in entries:
                sub_code = 0
                if e.get('sub_account'):
                    parts = e['sub_account'].split(' - ')
                    if parts: sub_code = parts[0]

                job_no = e.get('job_no') if e.get('job_no') else None

                cursor.execute("""
                    INSERT INTO entry_details (
                        account_name, account_code, enty_values_DR, enty_values_CR,
                        entry_effective_date, entry_create_date, entry_naration, entry_create_user,
                        entry_job_number, entry_sub_account_code, entry_jv
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    e['account'], 0, round(parse_float(e['dr']) * exchange_rate, 2), 0,
                    effective_date, date.today(), e['memo'], current_user_pk,
                    job_no, sub_code, jv_no
                ))

            conn.commit()
            flash(f'Service Entry / Supplier Liability saved successfully. JV: {jv_no}', 'success')
            _se_jv = jv_no

        except Exception as e:
            conn.rollback()
            flash(f'Database error: {str(e)}', 'danger')
            _se_jv = None
        finally:
            cursor.close()
            conn.close()

        if _se_jv:
            return redirect(url_for('service_entry', last_jv=_se_jv))

    except Exception as e:
        flash(f'Error saving Service Entry: {str(e)}', 'danger')

    return redirect(url_for('service_entry'))



@app.route('/service_entry_reversal')
@login_required
@has_permission('Access_Reversals')
def service_entry_reversal():
    # Fetch recent Service Entries (from suppliers_invoice_data & jv_numbers)
    # We look for outstanding invoices that have not been deleted
    query = """
        SELECT
            s.suppliers_invoice_JV as jv,
            j.jv_user_code,
            sup.supplier_name as SupplierName,
            s.suppliers_invoice_number as InvoiceNo,
            s.suppliers_invoice_date as Date,
            s.suppliers_invoice_total_oustanding as Amount
        FROM suppliers_invoice_data s
        JOIN jv_numbers j ON s.suppliers_invoice_JV = j.jv_id
        JOIN suppliers sup ON s.suppliers_invoice_buinding_supplier = sup.sup_id
        WHERE s.suppliers_oustanding_delete = 0
        AND j.jv_user_code LIKE 'JV FORM SEN INVOICE%'
        ORDER BY s.s_i_id DESC
        LIMIT 50
    """
    rows = db.execute_query(query)
    return render_template('service_entry_reversal.html', rows=rows)

@app.route('/service_entry_reversal/process', methods=['POST'])
@login_required
def service_entry_reversal_process():
    jv = request.form.get('jv')
    if not jv:
        flash('No transaction selected', 'danger')
        return redirect(url_for('service_entry_reversal'))

    if jv_in_closed_period(jv):
        flash(f'Cannot reverse: this entry is in a closed period (locked through {get_period_lock_date()}).', 'danger')
        return redirect(url_for('service_entry_reversal'))

    current_user = get_current_user_id()

    conn = None
    cursor = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        # 1. Security Check: Bank Reconciled?
        cursor.execute("SELECT COUNT(*) FROM entry_details WHERE entry_jv = %s AND entry_Rec = 1", (jv,))
        if cursor.fetchone()[0] > 0:
            flash('Cannot reverse: Transaction has been Bank Reconciled.', 'danger')
            return redirect(url_for('service_entry_reversal'))

        # 2. Security Check: Payments Made?
        cursor.execute("SELECT suppliers_invoice_total_payment FROM suppliers_invoice_data WHERE suppliers_invoice_JV = %s AND suppliers_oustanding_delete = 0", (jv,))
        inv_res = cursor.fetchone()
        if not inv_res:
            flash('Service Entry not found or already deleted.', 'danger')
            return redirect(url_for('service_entry_reversal'))

        if inv_res[0] > 0:
            flash('Cannot reverse: Payments have been made against this invoice. Reverse payments first.', 'danger')
            return redirect(url_for('service_entry_reversal'))

        conn.start_transaction()

        # 3. Create reversal JV
        current_user_pk = get_current_user_pk()
        today = date.today()
        cursor.execute(
            "INSERT INTO jv_numbers (jv_user_code, jv_naration, status) VALUES (%s, %s, 1)",
            (f'REV-SEN-{jv}', f'Service Entry Reversal of JV-{jv}'))
        rev_jv = cursor.lastrowid

        # 4. Reverse GL entries (swap DR ↔ CR)
        cursor.execute("""
            INSERT INTO entry_details (
                account_name, enty_values_DR, enty_values_CR,
                entry_effective_date, entry_create_date,
                entry_naration, entry_create_user, entry_jv
            )
            SELECT account_name,
                   COALESCE(enty_values_CR, 0),
                   COALESCE(enty_values_DR, 0),
                   %s, %s, %s, %s, %s
            FROM entry_details WHERE entry_jv = %s
        """, (today, today, f'Service Entry Reversal of JV-{jv}', current_user_pk, rev_jv, jv))

        # 5. Reverse Supplier Liability (Mark outstanding as deleted). Also
        # rename the invoice number with a -REV suffix — suppliers_invoice_number
        # is UNIQUE at the DB level, so leaving it unchanged would block any
        # future entry (e.g. a Service Entry Edit) that reuses the same number.
        cursor.execute(
            """UPDATE suppliers_invoice_data
               SET suppliers_oustanding_delete = 1,
                   suppliers_invoice_number = LEFT(CONCAT(COALESCE(suppliers_invoice_number, ''), '-REV', %s), 200)
               WHERE suppliers_invoice_JV = %s""",
            (rev_jv, jv))

        conn.commit()
        flash(f'Service Entry (JV: {jv}) reversed successfully. Reversal JV: {rev_jv}', 'success')

    except Exception as e:
        if conn:
            conn.rollback()
        flash(f'Error reversing Service Entry: {str(e)}', 'danger')
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

    return redirect(url_for('service_entry_reversal'))


@app.route('/service_entry/edit/<int:jv_no>', methods=['GET', 'POST'])
@login_required
@has_permission('Access_Accounting')
def service_entry_edit(jv_no):
    """Edit a previously-created Service Entry (SRN). Since a Service Entry is
    already posted to the GL and creates a supplier liability, 'editing' is
    done the same audit-safe way as everywhere else in this app (see
    service_entry_reversal_process): the original entry is reversed (GL
    entries + supplier liability) and a brand new, corrected entry is created
    in its place — both inside one DB transaction, so it's all-or-nothing.
    Blocked by the same guard rails as reversal: payments already made
    against it, bank reconciliation, or a closed accounting period."""
    jvrow = db.execute_query("SELECT jv_naration, jv_user_code FROM jv_numbers WHERE jv_id = %s", (jv_no,))
    if not jvrow or not (jvrow[0]['jv_user_code'] or '').startswith('JV FORM SEN INVOICE'):
        flash('No Service Entry found to edit.', 'danger')
        return redirect(url_for('service_entry_reversal'))

    inv_row = db.execute_query("""
        SELECT s.suppliers_invoice_number, s.suppliers_invoice_date, s.suppliers_invoice_final_date,
               s.suppliers_invoice_total_payment, s.suppliers_invoice_buinding_supplier,
               s.suppliers_VAT_rate, s.suppliers_oustanding_delete, sup.supplier_name
        FROM suppliers_invoice_data s
        LEFT JOIN suppliers sup ON s.suppliers_invoice_buinding_supplier = sup.sup_id
        WHERE s.suppliers_invoice_JV = %s
    """, (jv_no,))
    if not inv_row:
        flash('Service Entry invoice record not found.', 'danger')
        return redirect(url_for('service_entry_reversal'))
    inv_row = inv_row[0]

    if inv_row['suppliers_oustanding_delete']:
        flash('This Service Entry has already been deleted.', 'danger')
        return redirect(url_for('service_entry_reversal'))

    if inv_row['suppliers_invoice_total_payment'] and float(inv_row['suppliers_invoice_total_payment']) > 0:
        flash('Cannot edit: Payments have been made against this Service Entry. Reverse those payments first.', 'danger')
        return redirect(url_for('service_entry_reversal'))

    if jv_in_closed_period(jv_no):
        flash(f'Cannot edit: this Service Entry is in a closed period (locked through {get_period_lock_date()}).', 'danger')
        return redirect(url_for('service_entry_reversal'))

    bank_rec = db.execute_query("SELECT COUNT(*) as c FROM entry_details WHERE entry_jv = %s AND entry_Rec = 1", (jv_no,))
    if bank_rec and bank_rec[0]['c'] > 0:
        flash('Cannot edit: Transaction has been Bank Reconciled.', 'danger')
        return redirect(url_for('service_entry_reversal'))

    if request.method == 'POST':
        supplier_id = request.form.get('supplier_id')
        effective_date = request.form.get('effective_date')
        invoice_number = request.form.get('invoice_number')
        invoice_date = request.form.get('invoice_date')
        due_date = request.form.get('due_date')
        main_narration = request.form.get('main_narration')
        header_job = request.form.get('header_job_number') or None
        include_vat = request.form.get('include_vat') == '1'
        vat_rate = parse_float(request.form.get('vat_rate')) if include_vat else 0.0
        entries_json = request.form.get('entries_json')
        total_amount = parse_float(request.form.get('total_amount'))

        entries = json.loads(entries_json) if entries_json else []
        if not entries:
            flash('No entries provided', 'warning')
            return redirect(url_for('service_entry_edit', jv_no=jv_no))

        if not supplier_id or not invoice_number:
            flash('Supplier and Invoice Number are required', 'danger')
            return redirect(url_for('service_entry_edit', jv_no=jv_no))

        sup_res = db.execute_query("SELECT supplier_code FROM suppliers WHERE sup_id = %s", (supplier_id,))
        if not sup_res:
            flash('Invalid Supplier', 'danger')
            return redirect(url_for('service_entry_edit', jv_no=jv_no))
        supplier_code = sup_res[0]['supplier_code']

        conn = None
        cursor = None
        try:
            conn = db.get_connection()
            cursor = conn.cursor()

            # Re-check guards inside the transaction (state may have changed
            # between loading the edit form and submitting it).
            cursor.execute("SELECT COUNT(*) FROM entry_details WHERE entry_jv = %s AND entry_Rec = 1", (jv_no,))
            if cursor.fetchone()[0] > 0:
                flash('Cannot edit: Transaction has been Bank Reconciled.', 'danger')
                return redirect(url_for('service_entry_reversal'))

            cursor.execute(
                "SELECT suppliers_invoice_total_payment FROM suppliers_invoice_data "
                "WHERE suppliers_invoice_JV = %s AND suppliers_oustanding_delete = 0", (jv_no,))
            chk = cursor.fetchone()
            if not chk:
                flash('Service Entry not found or already deleted.', 'danger')
                return redirect(url_for('service_entry_reversal'))
            if chk[0] and float(chk[0]) > 0:
                flash('Cannot edit: Payments have been made against this Service Entry. Reverse those payments first.', 'danger')
                return redirect(url_for('service_entry_reversal'))

            conn.start_transaction()
            current_user = get_current_user_id()
            current_user_pk = get_current_user_pk()
            today = date.today()

            # 1. Reverse the original Service Entry (mirrors service_entry_reversal_process step-for-step)
            cursor.execute(
                "INSERT INTO jv_numbers (jv_user_code, jv_naration, status) VALUES (%s, %s, 1)",
                (f'REV-SEN-{jv_no}', f'Service Entry Edit Reversal of JV-{jv_no}'))
            rev_jv = cursor.lastrowid

            cursor.execute("""
                INSERT INTO entry_details (
                    account_name, enty_values_DR, enty_values_CR,
                    entry_effective_date, entry_create_date,
                    entry_naration, entry_create_user, entry_jv
                )
                SELECT account_name,
                       COALESCE(enty_values_CR, 0),
                       COALESCE(enty_values_DR, 0),
                       %s, %s, %s, %s, %s
                FROM entry_details WHERE entry_jv = %s
            """, (today, today, f'Service Entry Edit Reversal of JV-{jv_no}', current_user_pk, rev_jv, jv_no))

            # Same rename as service_entry_reversal_process: free up the invoice
            # number (UNIQUE at the DB level) so the corrected entry below can
            # reuse it — this is exactly what caused the duplicate-key error
            # when this fix was missing on the GRN side.
            cursor.execute(
                """UPDATE suppliers_invoice_data
                   SET suppliers_oustanding_delete = 1,
                       suppliers_invoice_number = LEFT(CONCAT(COALESCE(suppliers_invoice_number, ''), '-REV', %s), 200)
                   WHERE suppliers_invoice_JV = %s""",
                (rev_jv, jv_no))

            # 2. Create the corrected Service Entry (mirrors save_service_entry() step-for-step).
            # Amounts here are always in base currency — the edit form only offers
            # base-currency entry, avoiding any ambiguity about re-converting
            # already-converted historical figures.
            cursor.execute("INSERT INTO jv_numbers (jv_user_code, jv_naration, status) VALUES (%s, %s, %s)",
                           ("JV FORM SEN INVOICE", main_narration, 1))
            new_jv = cursor.lastrowid

            cursor.execute("""
                INSERT INTO suppliers_invoice_data (
                    suppliers_code, suppliers_invoice_number, suppliers_invoice_date,
                    suppliers_invoice_total_oustanding, suppliers_invoice_total_payment,
                    suppliers_invoice_final_date, suppliers_invoice_buinding_supplier,
                    suppliers_invoice_JV, suppliers_VAT_rate
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                supplier_code, invoice_number, invoice_date,
                total_amount, 0,
                due_date, supplier_id,
                new_jv, vat_rate
            ))

            cursor.execute("""
                INSERT INTO entry_details (
                    account_name, account_code, enty_values_DR, enty_values_CR,
                    entry_effective_date, entry_create_date, entry_naration, entry_create_user,
                    entry_job_number, entry_sub_account_code, entry_jv
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                'Account Payable', 0, 0, total_amount,
                effective_date, today, main_narration, current_user_pk,
                header_job, 0, new_jv
            ))

            total_dr = sum(parse_float(e['dr']) for e in entries)
            if include_vat and vat_rate > 0:
                vat_amount = total_dr * (vat_rate / 100.0)
                cursor.execute("""
                    INSERT INTO entry_details (
                        account_name, account_code, enty_values_DR, enty_values_CR,
                        entry_effective_date, entry_create_date, entry_naration, entry_create_user,
                        entry_job_number, entry_sub_account_code, entry_jv, entry_VAT
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    'VAT Control', 0, vat_amount, 0,
                    effective_date, today, main_narration, current_user_pk,
                    header_job, 0, new_jv, vat_rate
                ))

            for e in entries:
                sub_code = 0
                if e.get('sub_account'):
                    parts = e['sub_account'].split(' - ')
                    if parts: sub_code = parts[0]

                job_no = e.get('job_no') if e.get('job_no') else None

                cursor.execute("""
                    INSERT INTO entry_details (
                        account_name, account_code, enty_values_DR, enty_values_CR,
                        entry_effective_date, entry_create_date, entry_naration, entry_create_user,
                        entry_job_number, entry_sub_account_code, entry_jv
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    e['account'], 0, round(parse_float(e['dr']), 2), 0,
                    effective_date, today, e['memo'], current_user_pk,
                    job_no, sub_code, new_jv
                ))

            conn.commit()
            flash(f'Service Entry updated. Old JV-{jv_no} reversed; corrected entry created as JV-{new_jv}.', 'success')
            return redirect(url_for('service_entry_reversal'))

        except Exception as e:
            if conn:
                conn.rollback()
            flash(f'Error updating Service Entry: {str(e)}', 'danger')
            return redirect(url_for('service_entry_edit', jv_no=jv_no))
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()

    # GET: Load form data + prefill values from the existing Service Entry
    suppliers = db.execute_query("SELECT sup_id, supplier_name, supplier_code FROM suppliers WHERE Is_Suplier = 1")
    accounts = db.execute_query("""
        SELECT account_name
        FROM new_account_table
        WHERE account_active = 1 AND (account_income = 1 OR account_expenses = 1)
    """)
    jobs = db.execute_query("SELECT job_number FROM jobs_unit WHERE job_finsh = 0 AND job_cancell = 0")

    line_rows = db.execute_query("""
        SELECT account_name, enty_values_DR, entry_naration, entry_job_number, entry_sub_account_code
        FROM entry_details
        WHERE entry_jv = %s AND account_name NOT IN ('Account Payable', 'VAT Control')
    """, (jv_no,)) or []

    sub_code_names = {}
    codes = [r['entry_sub_account_code'] for r in line_rows if r.get('entry_sub_account_code')]
    if codes:
        placeholders = ','.join(['%s'] * len(codes))
        name_rows = db.execute_query(
            f"SELECT sub_account_code, sub_sub_accaount_name FROM sub_accont_for_new_account WHERE sub_account_code IN ({placeholders})",
            tuple(codes)) or []
        sub_code_names = {str(r['sub_account_code']): r['sub_sub_accaount_name'] for r in name_rows}

    prefill_entries = [{
        'account': r['account_name'],
        'sub_account_code': str(r['entry_sub_account_code']) if r.get('entry_sub_account_code') else '',
        'sub_account_name': sub_code_names.get(str(r['entry_sub_account_code']), '') if r.get('entry_sub_account_code') else '',
        'job_no': r['entry_job_number'] or '',
        'memo': r['entry_naration'] or '',
        'dr': float(r['enty_values_DR'] or 0),
    } for r in line_rows]

    vat_row = db.execute_query(
        "SELECT enty_values_DR AS amt FROM entry_details WHERE entry_jv = %s AND account_name = 'VAT Control'",
        (jv_no,))
    prefill_include_vat = bool(vat_row)
    prefill_vat_rate = float(inv_row['suppliers_VAT_rate'] or 0)

    ap_row = db.execute_query(
        "SELECT entry_effective_date, entry_job_number FROM entry_details WHERE entry_jv = %s AND account_name = 'Account Payable' LIMIT 1",
        (jv_no,))
    prefill_effective_date = str(ap_row[0]['entry_effective_date']) if ap_row and ap_row[0]['entry_effective_date'] else str(inv_row['suppliers_invoice_date'])
    prefill_header_job = (ap_row[0]['entry_job_number'] if ap_row else '') or ''

    hf = {'search': '', 'date_from': date.today().replace(day=1).strftime('%Y-%m-%d'),
          'date_to': date.today().strftime('%Y-%m-%d'), 'amount_min': '', 'amount_max': '', 'status': ''}
    history = db.execute_query("""
        SELECT
            s.suppliers_invoice_JV               AS jv,
            s.suppliers_invoice_number           AS invoice_no,
            s.suppliers_invoice_date             AS date,
            sup.supplier_name                    AS supplier,
            s.suppliers_invoice_total_oustanding AS amount,
            s.suppliers_oustanding_delete        AS is_reversed,
            j.jv_naration                        AS narration
        FROM suppliers_invoice_data s
        JOIN jv_numbers j ON s.suppliers_invoice_JV = j.jv_id
        LEFT JOIN suppliers sup ON s.suppliers_invoice_buinding_supplier = sup.sup_id
        WHERE j.jv_user_code LIKE %s AND s.suppliers_invoice_date >= %s AND s.suppliers_invoice_date <= %s
        ORDER BY s.s_i_id DESC
        LIMIT 200
    """, ('JV FORM SEN INVOICE%', hf['date_from'], hf['date_to'])) or []

    base_currency, currencies = _base_currency_and_list()
    return render_template('service_entry.html',
                           suppliers=suppliers,
                           accounts=accounts,
                           jobs=jobs,
                           today_date=date.today().strftime('%Y-%m-%d'),
                           last_jv=None,
                           history=history,
                           hf=hf,
                           base_currency=base_currency,
                           currencies=currencies,
                           edit_mode=True,
                           edit_jv=jv_no,
                           prefill_supplier_id=inv_row['suppliers_invoice_buinding_supplier'],
                           prefill_invoice_no=inv_row['suppliers_invoice_number'],
                           prefill_invoice_date=str(inv_row['suppliers_invoice_date']),
                           prefill_due_date=str(inv_row['suppliers_invoice_final_date']),
                           prefill_effective_date=prefill_effective_date,
                           prefill_narration=jvrow[0]['jv_naration'],
                           prefill_header_job=prefill_header_job,
                           prefill_include_vat=prefill_include_vat,
                           prefill_vat_rate=prefill_vat_rate,
                           prefill_entries=prefill_entries)


# ================================================================
# ── HR MODULE ───────────────────────────────────────────────────
# ================================================================

@app.route('/employees')
@login_required
def employees():
    emps = db.execute_query("SELECT * FROM employees WHERE is_active=1 ORDER BY emp_no")
    stats = {
        'total':    len(emps),
        'permanent': sum(1 for e in emps if e.get('employment_type') == 'Permanent'),
        'contract':  sum(1 for e in emps if e.get('employment_type') == 'Contract'),
    }
    return render_template('employees.html', employees=emps, stats=stats)


@app.route('/employee/add', methods=['GET', 'POST'])
@login_required
def add_employee():
    if request.method == 'POST':
        fields = ['emp_no','first_name','last_name','nic','email','mobile',
                  'department','designation','date_of_joining','date_of_birth',
                  'gender','employment_type','epf_no','etf_no',
                  'bank_name','bank_account_no','bank_branch']
        data   = {f: (request.form.get(f) or '').strip() for f in fields}
        data['date_of_joining'] = data['date_of_joining'] or None
        data['date_of_birth']   = data['date_of_birth']   or None
        data['basic_salary']    = float(request.form.get('basic_salary') or 0)
        data['created_by']      = session.get('username', '')
        if not data['emp_no'] or not data['first_name']:
            flash('Employee number and first name are required.', 'danger')
        else:
            try:
                conn   = db.get_connection()
                cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO employees (emp_no,first_name,last_name,nic,email,mobile,
                        department,designation,date_of_joining,date_of_birth,gender,
                        employment_type,epf_no,etf_no,bank_name,bank_account_no,bank_branch,
                        basic_salary,created_by)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (data['emp_no'],data['first_name'],data['last_name'],data['nic'],
                      data['email'],data['mobile'],data['department'],data['designation'],
                      data['date_of_joining'],data['date_of_birth'],data['gender'],
                      data['employment_type'],data['epf_no'],data['etf_no'],
                      data['bank_name'],data['bank_account_no'],data['bank_branch'],
                      data['basic_salary'],data['created_by']))
                conn.commit()
                cursor.close(); conn.close()
                flash(f"Employee {data['emp_no']} — {data['first_name']} {data['last_name']} added.", 'success')
                return redirect(url_for('employees'))
            except Exception as e:
                flash(f'Error: {str(e)}', 'danger')
    return render_template('employee_form.html', employee=None, mode='add')


@app.route('/employee/edit/<int:emp_id>', methods=['GET', 'POST'])
@login_required
def edit_employee(emp_id):
    if request.method == 'POST':
        try:
            conn   = db.get_connection()
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE employees SET first_name=%s,last_name=%s,nic=%s,email=%s,mobile=%s,
                    department=%s,designation=%s,date_of_joining=%s,date_of_birth=%s,
                    gender=%s,employment_type=%s,basic_salary=%s,epf_no=%s,etf_no=%s,
                    bank_name=%s,bank_account_no=%s,bank_branch=%s
                WHERE id=%s
            """, (
                request.form.get('first_name','').strip(),
                request.form.get('last_name','').strip(),
                request.form.get('nic','').strip(),
                request.form.get('email','').strip(),
                request.form.get('mobile','').strip(),
                request.form.get('department','').strip(),
                request.form.get('designation','').strip(),
                request.form.get('date_of_joining') or None,
                request.form.get('date_of_birth')   or None,
                request.form.get('gender',''),
                request.form.get('employment_type','Permanent'),
                float(request.form.get('basic_salary') or 0),
                request.form.get('epf_no','').strip(),
                request.form.get('etf_no','').strip(),
                request.form.get('bank_name','').strip(),
                request.form.get('bank_account_no','').strip(),
                request.form.get('bank_branch','').strip(),
                emp_id,
            ))
            conn.commit()
            cursor.close(); conn.close()
            flash('Employee record updated.', 'success')
            return redirect(url_for('employees'))
        except Exception as e:
            flash(f'Error: {str(e)}', 'danger')
    emp = db.execute_query("SELECT * FROM employees WHERE id=%s", (emp_id,))
    if not emp:
        flash('Employee not found.', 'danger')
        return redirect(url_for('employees'))
    return render_template('employee_form.html', employee=emp[0], mode='edit')


@app.route('/employee/deactivate/<int:emp_id>', methods=['POST'])
@login_required
def deactivate_employee(emp_id):
    try:
        conn = db.get_connection(); cursor = conn.cursor()
        cursor.execute("UPDATE employees SET is_active=0 WHERE id=%s", (emp_id,))
        conn.commit(); cursor.close(); conn.close()
        flash('Employee deactivated.', 'success')
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
    return redirect(url_for('employees'))


# ── Leave Types ─────────────────────────────────────────────

@app.route('/leave_types', methods=['GET', 'POST'])
@login_required
def leave_types():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            name    = request.form.get('leave_name','').strip()
            days    = int(request.form.get('days_per_year') or 0)
            is_paid = 1 if request.form.get('is_paid') else 0
            if name:
                try:
                    conn = db.get_connection(); cursor = conn.cursor()
                    cursor.execute(
                        "INSERT INTO leave_types (leave_name,days_per_year,is_paid) VALUES (%s,%s,%s)",
                        (name, days, is_paid))
                    conn.commit(); cursor.close(); conn.close()
                    flash('Leave type added.', 'success')
                except Exception as e:
                    flash(f'Error: {str(e)}', 'danger')
        elif action == 'delete':
            lt_id = request.form.get('lt_id')
            try:
                conn = db.get_connection(); cursor = conn.cursor()
                cursor.execute("UPDATE leave_types SET is_active=0 WHERE id=%s", (lt_id,))
                conn.commit(); cursor.close(); conn.close()
                flash('Leave type removed.', 'success')
            except Exception as e:
                flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('leave_types'))
    types = db.execute_query("SELECT * FROM leave_types WHERE is_active=1 ORDER BY leave_name")
    return render_template('leave_types.html', leave_types=types)


# ── Leave Applications ──────────────────────────────────────

@app.route('/leave_application', methods=['GET', 'POST'])
@login_required
def leave_application():
    if request.method == 'POST':
        emp_id  = int(request.form.get('employee_id') or 0)
        lt_id   = int(request.form.get('leave_type_id') or 0)
        start   = request.form.get('start_date')
        end     = request.form.get('end_date')
        reason  = request.form.get('reason','').strip()
        if emp_id and lt_id and start and end:
            s    = datetime.strptime(start,'%Y-%m-%d').date()
            e    = datetime.strptime(end,'%Y-%m-%d').date()
            days = (e - s).days + 1
            try:
                conn = db.get_connection(); cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO leave_applications
                        (employee_id,leave_type_id,start_date,end_date,days_requested,reason)
                    VALUES (%s,%s,%s,%s,%s,%s)
                """, (emp_id, lt_id, start, end, days, reason))
                conn.commit(); cursor.close(); conn.close()
                flash(f'Leave application submitted for {days} day(s).', 'success')
            except Exception as e:
                flash(f'Error: {str(e)}', 'danger')
        else:
            flash('All fields are required.', 'danger')
        return redirect(url_for('leave_application'))

    employees_list = db.execute_query(
        "SELECT id,emp_no,first_name,last_name FROM employees WHERE is_active=1 ORDER BY first_name")
    types_list = db.execute_query("SELECT * FROM leave_types WHERE is_active=1 ORDER BY leave_name")
    applications = db.execute_query("""
        SELECT la.*,e.first_name,e.last_name,e.emp_no,lt.leave_name
        FROM leave_applications la
        JOIN employees e  ON la.employee_id  = e.id
        JOIN leave_types lt ON la.leave_type_id = lt.id
        ORDER BY la.created_at DESC LIMIT 100
    """)
    return render_template('leave_application.html',
                           employees=employees_list,
                           leave_types=types_list,
                           applications=applications)


# ── Leave Approvals ─────────────────────────────────────────

@app.route('/leave_approvals', methods=['GET', 'POST'])
@login_required
def leave_approvals():
    if request.method == 'POST':
        app_id  = request.form.get('application_id')
        action  = request.form.get('action')
        status  = 'Approved' if action == 'approve' else 'Rejected'
        try:
            conn = db.get_connection(); cursor = conn.cursor()
            cursor.execute("""
                UPDATE leave_applications SET status=%s, approved_by=%s, approved_at=NOW()
                WHERE id=%s
            """, (status, session.get('username',''), app_id))
            conn.commit(); cursor.close(); conn.close()
            flash(f'Leave {status.lower()}.', 'success')
        except Exception as e:
            flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('leave_approvals'))

    pending = db.execute_query("""
        SELECT la.*,e.first_name,e.last_name,e.emp_no,lt.leave_name,lt.is_paid
        FROM leave_applications la
        JOIN employees e  ON la.employee_id  = e.id
        JOIN leave_types lt ON la.leave_type_id = lt.id
        WHERE la.status='Pending' ORDER BY la.created_at ASC
    """)
    history = db.execute_query("""
        SELECT la.*,e.first_name,e.last_name,e.emp_no,lt.leave_name
        FROM leave_applications la
        JOIN employees e  ON la.employee_id  = e.id
        JOIN leave_types lt ON la.leave_type_id = lt.id
        WHERE la.status!='Pending' ORDER BY la.approved_at DESC LIMIT 50
    """)
    return render_template('leave_approvals.html', pending=pending, history=history)


# ================================================================
# ── PAYROLL MODULE ──────────────────────────────────────────────
# ================================================================

@app.route('/payroll_components', methods=['GET', 'POST'])
@login_required
def payroll_components():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            name    = request.form.get('component_name','').strip()
            ctype   = request.form.get('component_type','Allowance')
            taxable = 1 if request.form.get('is_taxable') else 0
            if name:
                try:
                    conn = db.get_connection(); cursor = conn.cursor()
                    cursor.execute(
                        "INSERT INTO payroll_components (component_name,component_type,is_taxable) VALUES (%s,%s,%s)",
                        (name, ctype, taxable))
                    conn.commit(); cursor.close(); conn.close()
                    flash('Component added.', 'success')
                except Exception as e:
                    flash(f'Error: {str(e)}', 'danger')
        elif action == 'delete':
            pc_id = request.form.get('pc_id')
            try:
                conn = db.get_connection(); cursor = conn.cursor()
                cursor.execute("UPDATE payroll_components SET is_active=0 WHERE id=%s", (pc_id,))
                conn.commit(); cursor.close(); conn.close()
                flash('Component removed.', 'success')
            except Exception as e:
                flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('payroll_components'))
    components = db.execute_query(
        "SELECT * FROM payroll_components WHERE is_active=1 ORDER BY component_type,component_name")
    return render_template('payroll_components.html', components=components)


@app.route('/payroll_run', methods=['GET', 'POST'])
@login_required
def payroll_run_list():
    if request.method == 'POST':
        month = int(request.form.get('pay_month') or 0)
        year  = int(request.form.get('pay_year')  or 0)
        if month and year:
            existing = db.execute_query(
                "SELECT id FROM payroll_runs WHERE pay_period_month=%s AND pay_period_year=%s",
                (month, year))
            if existing:
                flash(f'Payroll for {month}/{year} already exists.', 'warning')
                return redirect(url_for('payroll_run_detail', run_id=existing[0]['id']))
            try:
                conn = db.get_connection(); cursor = conn.cursor()
                cursor.execute(
                    "INSERT INTO payroll_runs (pay_period_month,pay_period_year,run_date,created_by) VALUES (%s,%s,%s,%s)",
                    (month, year, date.today(), session.get('username','')))
                run_id = cursor.lastrowid
                conn.commit(); cursor.close(); conn.close()
                flash('Payroll run created. Click Process to calculate.', 'success')
                return redirect(url_for('payroll_run_detail', run_id=run_id))
            except Exception as e:
                flash(f'Error: {str(e)}', 'danger')
    runs = db.execute_query(
        "SELECT * FROM payroll_runs ORDER BY pay_period_year DESC,pay_period_month DESC")
    return render_template('payroll_run.html', runs=runs)


@app.route('/payroll_run/<int:run_id>', methods=['GET', 'POST'])
@login_required
def payroll_run_detail(run_id):
    run = db.execute_query("SELECT * FROM payroll_runs WHERE id=%s", (run_id,))
    if not run:
        flash('Payroll run not found.', 'danger')
        return redirect(url_for('payroll_run_list'))
    run = run[0]

    if request.method == 'POST' and run['status'] == 'Draft':
        try:
            emp_list = db.execute_query("SELECT * FROM employees WHERE is_active=1")
            conn = db.get_connection(); cursor = conn.cursor()
            cursor.execute("DELETE FROM payroll_run_lines WHERE payroll_run_id=%s", (run_id,))
            total_gross = total_ded_sum = total_net = 0.0

            for emp in emp_list:
                basic = float(emp.get('basic_salary') or 0)
                struct = db.execute_query("""
                    SELECT ess.amount, pc.component_type
                    FROM employee_salary_structure ess
                    JOIN payroll_components pc ON ess.component_id=pc.id
                    WHERE ess.employee_id=%s AND pc.is_active=1
                """, (emp['id'],))
                extra_allow = sum(float(s['amount'] or 0) for s in struct if s['component_type']=='Allowance')
                deductions  = sum(float(s['amount'] or 0) for s in struct if s['component_type']=='Deduction')

                gross    = basic + extra_allow
                epf_emp  = round(basic * 0.08, 2)
                epf_er   = round(basic * 0.12, 2)
                etf_er   = round(basic * 0.03, 2)
                total_d  = deductions + epf_emp
                net      = gross - total_d

                cursor.execute("""
                    INSERT INTO payroll_run_lines
                        (payroll_run_id,employee_id,basic_salary,total_allowances,total_deductions,
                         epf_employee,epf_employer,etf_employer,gross_salary,net_salary)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (run_id, emp['id'], basic, extra_allow, deductions,
                      epf_emp, epf_er, etf_er, gross, net))

                total_gross += gross
                total_ded_sum += total_d
                total_net += net

            cursor.execute("""
                UPDATE payroll_runs SET status='Processed',total_gross=%s,total_deductions=%s,total_net=%s
                WHERE id=%s
            """, (total_gross, total_ded_sum, total_net, run_id))
            conn.commit(); cursor.close(); conn.close()
            flash(f'Payroll processed for {len(emp_list)} employee(s). Net payable: {total_net:,.2f}', 'success')
            return redirect(url_for('payroll_run_detail', run_id=run_id))
        except Exception as e:
            flash(f'Error processing payroll: {str(e)}', 'danger')

    lines = db.execute_query("""
        SELECT prl.*,e.emp_no,e.first_name,e.last_name,e.designation,e.bank_account_no,e.bank_name
        FROM payroll_run_lines prl
        JOIN employees e ON prl.employee_id=e.id
        WHERE prl.payroll_run_id=%s ORDER BY e.emp_no
    """, (run_id,))
    return render_template('payroll_run_detail.html', run=run, lines=lines)


@app.route('/payslip/print/<int:run_id>/<int:emp_id>')
@login_required
def payslip_print(run_id, emp_id):
    run  = db.execute_query("SELECT * FROM payroll_runs WHERE id=%s", (run_id,))
    line = db.execute_query("""
        SELECT prl.*,e.*
        FROM payroll_run_lines prl JOIN employees e ON prl.employee_id=e.id
        WHERE prl.payroll_run_id=%s AND prl.employee_id=%s
    """, (run_id, emp_id))
    if not run or not line:
        flash('Payslip not found.', 'danger')
        return redirect(url_for('payroll_run_list'))
    company_info = db.execute_query("SELECT * FROM company_profile LIMIT 1")
    components   = db.execute_query("""
        SELECT ess.amount,pc.component_name,pc.component_type
        FROM employee_salary_structure ess
        JOIN payroll_components pc ON ess.component_id=pc.id
        WHERE ess.employee_id=%s AND pc.is_active=1
        ORDER BY pc.component_type,pc.component_name
    """, (emp_id,))
    return render_template('payslip_print.html',
                           run=run[0], line=line[0],
                           company=company_info[0] if company_info else {},
                           components=components)


# ================================================================
# ── CRM MODULE ──────────────────────────────────────────────────
# ================================================================

@app.route('/crm')
@login_required
def crm_pipeline():
    statuses = ['New','Contacted','Qualified','Proposal','Won','Lost']
    leads_by_status = {
        st: db.execute_query("SELECT * FROM crm_leads WHERE status=%s ORDER BY updated_at DESC", (st,))
        for st in statuses
    }
    summary = db.execute_query("""
        SELECT status, COUNT(*) AS cnt, COALESCE(SUM(expected_value),0) AS total_value
        FROM crm_leads GROUP BY status
    """)
    total_pipeline = sum(
        float(r.get('total_value') or 0) for r in summary if r['status'] not in ('Won','Lost'))
    return render_template('crm_pipeline.html',
                           leads_by_status=leads_by_status,
                           statuses=statuses,
                           summary=summary,
                           total_pipeline=total_pipeline)


@app.route('/crm/lead/add', methods=['GET', 'POST'])
@login_required
def add_crm_lead():
    if request.method == 'POST':
        lead_name = request.form.get('lead_name','').strip()
        if not lead_name:
            flash('Lead name is required.', 'danger')
        else:
            try:
                conn = db.get_connection(); cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO crm_leads (lead_name,company_name,email,mobile,source,status,
                        expected_value,expected_close_date,notes,assigned_to,created_by)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    lead_name,
                    request.form.get('company_name','').strip(),
                    request.form.get('email','').strip(),
                    request.form.get('mobile','').strip(),
                    request.form.get('source','Other'),
                    request.form.get('status','New'),
                    float(request.form.get('expected_value') or 0),
                    request.form.get('expected_close_date') or None,
                    request.form.get('notes','').strip(),
                    request.form.get('assigned_to','').strip() or session.get('username',''),
                    session.get('username',''),
                ))
                lead_id = cursor.lastrowid
                conn.commit(); cursor.close(); conn.close()
                flash(f'Lead "{lead_name}" added.', 'success')
                return redirect(url_for('crm_lead_detail', lead_id=lead_id))
            except Exception as e:
                flash(f'Error: {str(e)}', 'danger')
    return render_template('crm_lead_form.html', lead=None, mode='add')


@app.route('/crm/lead/<int:lead_id>', methods=['GET', 'POST'])
@login_required
def crm_lead_detail(lead_id):
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'update':
            try:
                conn = db.get_connection(); cursor = conn.cursor()
                cursor.execute("""
                    UPDATE crm_leads SET lead_name=%s,company_name=%s,email=%s,mobile=%s,
                        source=%s,status=%s,expected_value=%s,expected_close_date=%s,
                        notes=%s,assigned_to=%s
                    WHERE id=%s
                """, (
                    request.form.get('lead_name','').strip(),
                    request.form.get('company_name','').strip(),
                    request.form.get('email','').strip(),
                    request.form.get('mobile','').strip(),
                    request.form.get('source','Other'),
                    request.form.get('status','New'),
                    float(request.form.get('expected_value') or 0),
                    request.form.get('expected_close_date') or None,
                    request.form.get('notes','').strip(),
                    request.form.get('assigned_to','').strip(),
                    lead_id,
                ))
                conn.commit(); cursor.close(); conn.close()
                flash('Lead updated.', 'success')
            except Exception as e:
                flash(f'Error: {str(e)}', 'danger')
        elif action == 'activity':
            try:
                conn = db.get_connection(); cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO crm_activities
                        (lead_id,activity_type,subject,notes,activity_date,next_follow_up,created_by)
                    VALUES (%s,%s,%s,%s,%s,%s,%s)
                """, (
                    lead_id,
                    request.form.get('activity_type','Note'),
                    request.form.get('subject','').strip(),
                    request.form.get('act_notes','').strip(),
                    request.form.get('activity_date') or datetime.now().strftime('%Y-%m-%d %H:%M'),
                    request.form.get('next_follow_up') or None,
                    session.get('username',''),
                ))
                conn.commit(); cursor.close(); conn.close()
                flash('Activity logged.', 'success')
            except Exception as e:
                flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('crm_lead_detail', lead_id=lead_id))

    lead = db.execute_query("SELECT * FROM crm_leads WHERE id=%s", (lead_id,))
    if not lead:
        flash('Lead not found.', 'danger')
        return redirect(url_for('crm_pipeline'))
    activities = db.execute_query(
        "SELECT * FROM crm_activities WHERE lead_id=%s ORDER BY activity_date DESC", (lead_id,))
    return render_template('crm_lead_detail.html', lead=lead[0], activities=activities)


@app.route('/crm/lead/<int:lead_id>/delete', methods=['POST'])
@login_required
def delete_crm_lead(lead_id):
    try:
        conn = db.get_connection(); cursor = conn.cursor()
        cursor.execute("DELETE FROM crm_activities WHERE lead_id=%s", (lead_id,))
        cursor.execute("DELETE FROM crm_leads WHERE id=%s", (lead_id,))
        conn.commit(); cursor.close(); conn.close()
        flash('Lead deleted.', 'success')
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
    return redirect(url_for('crm_pipeline'))


# ================================================================
# ── EMAIL INTEGRATION ───────────────────────────────────────────
# ================================================================

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText as _MIMEText


def send_email_via_smtp(to_email, subject, html_body):
    """Returns (True, None) on success or (False, error_string) on failure."""
    cfg = db.execute_query("SELECT * FROM email_settings WHERE is_active=1 LIMIT 1")
    if not cfg:
        return False, 'Email not configured. Go to Settings → Email Settings.'
    cfg = cfg[0]
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = f"{cfg.get('sender_name','')} <{cfg.get('sender_email','')}>"
        msg['To']      = to_email
        msg.attach(_MIMEText(html_body, 'html'))
        port = int(cfg.get('smtp_port') or 587)
        if cfg.get('use_tls'):
            srv = smtplib.SMTP(cfg['smtp_host'], port, timeout=15)
            srv.starttls()
        else:
            srv = smtplib.SMTP_SSL(cfg['smtp_host'], port, timeout=15)
        srv.login(cfg['smtp_username'], cfg['smtp_password'])
        srv.sendmail(cfg['sender_email'], to_email, msg.as_string())
        srv.quit()
        _log_email_record(to_email, subject, html_body, 'Sent', None)
        return True, None
    except Exception as e:
        _log_email_record(to_email, subject, html_body, 'Failed', str(e))
        return False, str(e)


def _log_email_record(recipient, subject, body, status, error):
    try:
        conn = db.get_connection(); cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO email_log (recipient_email,subject,body,status,error_message) VALUES (%s,%s,%s,%s,%s)",
            (recipient, subject, (body or '')[:5000], status, error))
        conn.commit(); cursor.close(); conn.close()
    except Exception:
        pass


@app.route('/email_settings', methods=['GET', 'POST'])
@login_required
def email_settings():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'save':
            try:
                conn = db.get_connection(); cursor = conn.cursor()
                cursor.execute("DELETE FROM email_settings")
                cursor.execute("""
                    INSERT INTO email_settings
                        (smtp_host,smtp_port,smtp_username,smtp_password,sender_name,sender_email,use_tls,is_active)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    request.form.get('smtp_host','').strip(),
                    int(request.form.get('smtp_port') or 587),
                    request.form.get('smtp_username','').strip(),
                    request.form.get('smtp_password','').strip() or None,
                    request.form.get('sender_name','').strip(),
                    request.form.get('sender_email','').strip(),
                    1 if request.form.get('use_tls') else 0,
                    1 if request.form.get('is_active') else 0,
                ))
                conn.commit(); cursor.close(); conn.close()
                flash('Email settings saved.', 'success')
            except Exception as e:
                flash(f'Error: {str(e)}', 'danger')
        elif action == 'test':
            test_to = request.form.get('test_email','').strip()
            if test_to:
                ok, err = send_email_via_smtp(test_to,
                    'Suwin ERP — Test Email',
                    '<p style="font-family:Inter,sans-serif;">This is a test email from <b>Suwin Accounting ERP</b>. ✅</p>')
                flash('Test email sent successfully!' if ok else f'Test failed: {err}',
                      'success' if ok else 'danger')
        return redirect(url_for('email_settings'))

    cfg  = db.execute_query("SELECT * FROM email_settings LIMIT 1")
    logs = db.execute_query("SELECT * FROM email_log ORDER BY sent_at DESC LIMIT 30")
    return render_template('email_settings.html',
                           settings=cfg[0] if cfg else {},
                           email_logs=logs)


# ════════════════════════════════════════════════════════════════════════════
# DOCUMENT UPLOAD SYSTEM
# ════════════════════════════════════════════════════════════════════════════
import uuid as _uuid
from werkzeug.utils import secure_filename as _secure_filename

UPLOAD_FOLDER   = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads', 'documents')
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'gif', 'doc', 'docx', 'xls', 'xlsx', 'csv', 'txt', 'zip'}
MAX_UPLOAD_MB   = 20  # MB

def _allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def _get_upload_dir():
    """Returns upload dir for the current tenant, creates if missing."""
    db_name = session.get('db_name', 'default')
    safe_name = re.sub(r'[^a-z0-9_]', '_', db_name.lower())
    path = os.path.join(UPLOAD_FOLDER, safe_name)
    os.makedirs(path, exist_ok=True)
    return path


@app.route('/api/documents/upload', methods=['POST'])
@login_required
def api_document_upload():
    related_type = request.form.get('related_type', '').strip()
    related_id   = request.form.get('related_id', '').strip()
    notes        = request.form.get('notes', '').strip()

    if not related_type or not related_id:
        return jsonify({'success': False, 'error': 'related_type and related_id are required'}), 400

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file provided'}), 400

    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    if not _allowed_file(file.filename):
        return jsonify({'success': False, 'error': f'File type not allowed. Allowed: {", ".join(sorted(ALLOWED_EXTENSIONS))}'}), 400

    # Size check
    file.seek(0, 2)
    size_bytes = file.tell()
    file.seek(0)
    if size_bytes > MAX_UPLOAD_MB * 1024 * 1024:
        return jsonify({'success': False, 'error': f'File too large. Max {MAX_UPLOAD_MB} MB'}), 400

    original_name = _secure_filename(file.filename)
    ext          = original_name.rsplit('.', 1)[1].lower() if '.' in original_name else ''
    stored_name  = f"{_uuid.uuid4().hex}.{ext}"
    upload_dir   = _get_upload_dir()
    file_path    = os.path.join(upload_dir, stored_name)

    try:
        file.save(file_path)
    except Exception as e:
        logging.error(f"File save error: {e}")
        return jsonify({'success': False, 'error': 'Could not save file on server'}), 500

    doc_id = db.execute_query(
        """INSERT INTO documents (related_type, related_id, file_name, stored_name,
               file_size, file_type, notes, uploaded_by)
           VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
        (related_type, related_id, original_name, stored_name,
         size_bytes, ext, notes, session.get('user_name', '')),
        commit=True
    )
    if not doc_id:
        return jsonify({'success': False, 'error': 'DB insert failed'}), 500

    return jsonify({
        'success': True,
        'doc': {
            'id': doc_id,
            'file_name': original_name,
            'file_size': size_bytes,
            'file_type': ext,
            'notes': notes,
            'uploaded_by': session.get('user_name', ''),
            'uploaded_at': 'just now'
        }
    })


@app.route('/api/documents/list/<string:related_type>/<path:related_id>')
@login_required
def api_document_list(related_type, related_id):
    docs = db.execute_query(
        """SELECT id, file_name, file_size, file_type, notes, uploaded_by,
                  DATE_FORMAT(uploaded_at, '%%d %%b %%Y %%H:%%i') as uploaded_at
           FROM documents
           WHERE related_type = %s AND related_id = %s
           ORDER BY uploaded_at DESC""",
        (related_type, related_id)
    ) or []
    return jsonify({'success': True, 'docs': docs})


@app.route('/api/documents/view/<int:doc_id>')
@login_required
def api_document_view(doc_id):
    row = db.execute_query(
        "SELECT file_name, stored_name, file_type FROM documents WHERE id = %s", (doc_id,)
    )
    if not row:
        return "File not found", 404
    row = row[0]
    upload_dir = _get_upload_dir()
    file_path  = os.path.join(upload_dir, row['stored_name'])
    if not os.path.exists(file_path):
        return "File missing on server", 404

    from flask import send_file as _send_file
    mime_map = {
        'pdf': 'application/pdf', 'png': 'image/png', 'jpg': 'image/jpeg',
        'jpeg': 'image/jpeg', 'gif': 'image/gif',
        'doc': 'application/msword',
        'docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        'xls': 'application/vnd.ms-excel',
        'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'csv': 'text/csv', 'txt': 'text/plain', 'zip': 'application/zip'
    }
    mime = mime_map.get(row['file_type'], 'application/octet-stream')
    # PDFs and images display inline; others download
    disposition = 'inline' if row['file_type'] in ('pdf', 'png', 'jpg', 'jpeg', 'gif') else 'attachment'
    return _send_file(file_path, mimetype=mime,
                      as_attachment=(disposition == 'attachment'),
                      download_name=row['file_name'])


@app.route('/api/documents/delete/<int:doc_id>', methods=['POST'])
@login_required
def api_document_delete(doc_id):
    row = db.execute_query(
        "SELECT stored_name, uploaded_by FROM documents WHERE id = %s", (doc_id,)
    )
    if not row:
        return jsonify({'success': False, 'error': 'Document not found'}), 404
    row = row[0]

    # Delete physical file
    upload_dir = _get_upload_dir()
    file_path  = os.path.join(upload_dir, row['stored_name'])
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
    except Exception as e:
        logging.warning(f"Could not delete file {file_path}: {e}")

    db.execute_query("DELETE FROM documents WHERE id = %s", (doc_id,), commit=True)
    return jsonify({'success': True})


@app.route('/api/email/send_invoice', methods=['POST'])
@login_required
def api_send_invoice_email():
    data     = request.get_json() or {}
    to_email = (data.get('to_email') or '').strip()
    inv_no   = (data.get('invoice_no') or '').strip()
    if not to_email or not inv_no:
        return jsonify({'success': False, 'error': 'Email and invoice number required.'})
    company_info = db.execute_query("SELECT * FROM company_profile LIMIT 1")
    co = company_info[0] if company_info else {}
    subject = f"Invoice {inv_no} — {co.get('company_name','Suwin ERP')}"
    body = f"""
    <div style="font-family:Inter,sans-serif;max-width:620px;margin:auto;padding:32px;border:1px solid #e2e8f0;border-radius:12px;">
      <div style="border-top:4px solid #0078D4;padding-bottom:20px;margin-bottom:24px;">
        <h2 style="color:#0078D4;margin:16px 0 4px;">{co.get('company_name','')}</h2>
        <p style="color:#605E5C;margin:0;">{co.get('company_address','')}</p>
      </div>
      <p>Dear Customer,</p>
      <p>Please find your invoice <strong>{inv_no}</strong> ready for viewing.</p>
      <a href="{request.host_url}invoice_print/{inv_no}"
         style="display:inline-block;background:#0078D4;color:#fff;padding:12px 28px;border-radius:8px;text-decoration:none;font-weight:600;margin:16px 0;">
        View Invoice →
      </a>
      <p style="color:#605E5C;font-size:13px;margin-top:28px;border-top:1px solid #f3f2f1;padding-top:16px;">
        {co.get('company_phone','')} &nbsp;|&nbsp; {co.get('company_email','')}
      </p>
    </div>"""
    ok, err = send_email_via_smtp(to_email, subject, body)
    if ok:
        return jsonify({'success': True, 'message': f'Invoice emailed to {to_email}'})
    return jsonify({'success': False, 'error': err})


# ─────────────────────────────────────────────────────────────────────────────
# Inventory Issue — issue stock to a cost account (Dr cost account, Cr Inventory),
# deduct stock, with an issue report. Mirrors the production-issue GL pattern.
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/inventory_issue')
@login_required
@has_permission('Access_Inventory')
def inventory_issue():
    items = db.execute_query("""
        SELECT i.id, i.inventoy_name AS name, i.inventoy_code AS code,
               i.inventoy_items_messurment_unit AS unit,
               (SELECT COALESCE(p.inventory_price_purcharsing, 0)
                  FROM inventory_price_recod p
                 WHERE p.inventory_price_link = i.id
                 ORDER BY p.id DESC LIMIT 1) AS cost,
               COALESCE((SELECT SUM(r.inventory_recod_moument_in - r.inventory_recod_movment_out)
                           FROM inventory_recod r
                          WHERE r.inventoy_name = i.inventoy_name), 0) AS stock
        FROM inventoy_items i
        WHERE i.active = 1
        ORDER BY i.inventoy_name
    """) or []
    locations = db.execute_query("SELECT inventory_locations_name FROM inventory_locations") or []
    accounts = db.execute_query("""
        SELECT account_name FROM new_account_table
        WHERE (account_expenses = 1 OR account_assets = 1) AND account_active = 1
        ORDER BY account_name
    """) or []
    jobs = db.execute_query("SELECT job_number FROM jobs_unit") or []
    return render_template('inventory_issue.html', items=items, locations=locations,
                           accounts=accounts, jobs=jobs,
                           today_date=date.today().strftime('%Y-%m-%d'))


@app.route('/inventory_issue/submit', methods=['POST'])
@login_required
@has_permission('Access_Inventory')
def inventory_issue_submit():
    """Stock OUT + Dr selected cost account, Cr Inventory control account."""
    try:
        date_val = request.form.get('issue_date') or date.today().strftime('%Y-%m-%d')
        job_no = request.form.get('job_no')
        source_loc = request.form.get('source_location')
        dr_account = request.form.get('cost_account')
        narration = request.form.get('narration') or 'Inventory Issue'

        item_names = request.form.getlist('item_name[]')
        item_codes = request.form.getlist('item_code[]')
        item_units = request.form.getlist('item_unit[]')
        item_costs = request.form.getlist('unit_cost[]')
        qtys = request.form.getlist('qty[]')

        if not item_names:
            flash('No items to issue.', 'danger')
            return redirect(url_for('inventory_issue'))
        if not dr_account:
            flash('Please select a cost account to charge.', 'danger')
            return redirect(url_for('inventory_issue'))
        if not source_loc:
            flash('Please select the source location.', 'danger')
            return redirect(url_for('inventory_issue'))

        current_user_pk = get_current_user_pk()

        conn = db.get_connection()
        cursor = conn.cursor()
        conn.start_transaction()
        try:
            cursor.execute("INSERT INTO jv_numbers (jv_user_code, jv_naration) VALUES (%s, %s)",
                           ('JV FROM INVENTORY ISSUE', f"Inventory Issue: {narration}"))
            jv_no = cursor.lastrowid

            total_value = 0.0
            for i in range(len(item_names)):
                qty = float(qtys[i] or 0)
                cost = float(item_costs[i] or 0)
                if qty <= 0:
                    continue
                total_value += qty * cost
                cursor.execute("""
                    INSERT INTO inventory_recod (
                        inventoy_name, inventoy_code, inventory_recod_mesrmet,
                        inventory_recod_unit_price, inventory_recod_moument_in, inventory_recod_movment_out,
                        inventory_recod_suplier_iv_no, inventory_recod_user_id,
                        inventory_recod_user_recod_date, inventory_recod_location,
                        inventory_recod_action_date, inventory_recodcol_memo, JV_No
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (item_names[i], item_codes[i], item_units[i], cost, 0, qty,
                      f"ISS-{jv_no}", current_user_pk, datetime.now().date(), source_loc,
                      date_val, narration, jv_no))

            if total_value <= 0:
                conn.rollback()
                flash('Nothing to issue — check quantities.', 'warning')
                return redirect(url_for('inventory_issue'))

            # GL: Dr selected cost account, Cr Inventory control account
            cursor.execute("""
                INSERT INTO entry_details (
                    account_name, enty_values_DR, entry_effective_date, entry_create_date,
                    entry_naration, entry_create_user, entry_jv, entry_job_number
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
            """, (dr_account, total_value, date_val, datetime.now().date(), narration,
                  current_user_pk, jv_no, job_no if job_no else None))
            cursor.execute("""
                INSERT INTO entry_details (
                    account_name, enty_values_CR, entry_effective_date, entry_create_date,
                    entry_naration, entry_create_user, entry_jv, entry_job_number
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
            """, ('Inventory', total_value, date_val, datetime.now().date(), narration,
                  current_user_pk, jv_no, job_no if job_no else None))

            conn.commit()
            flash(f'Inventory issued successfully. JV: {jv_no}. Dr {dr_account} / Cr Inventory {total_value:,.2f}.', 'success')
        except Exception as e:
            conn.rollback()
            flash(f'Error processing issue: {str(e)}', 'danger')
        finally:
            cursor.close()
            conn.close()
    except Exception as e:
        flash(f'System Error: {str(e)}', 'danger')
    return redirect(url_for('inventory_issue'))


@app.route('/inventory_issue/report')
@login_required
@has_permission('Access_Reports')
def inventory_issue_report():
    from_date = (request.args.get('from') or date.today().replace(day=1).strftime('%Y-%m-%d')).strip()
    to_date = (request.args.get('to') or date.today().strftime('%Y-%m-%d')).strip()

    lines = db.execute_query("""
        SELECT inventory_recod_action_date AS date, JV_No AS jv,
               inventoy_name AS name, inventoy_code AS code,
               inventory_recod_mesrmet AS unit,
               inventory_recod_movment_out AS qty,
               inventory_recod_unit_price AS cost,
               inventory_recod_location AS location,
               inventory_recodcol_memo AS narration
        FROM inventory_recod
        WHERE inventory_recod_suplier_iv_no LIKE 'ISS-%%'
          AND inventory_recod_action_date BETWEEN %s AND %s
        ORDER BY inventory_recod_action_date DESC, JV_No DESC, id DESC
    """, (from_date, to_date)) or []

    # cost account (DR side) per JV
    jv_account = {}
    jvs = list({l['jv'] for l in lines if l['jv'] is not None})
    if jvs:
        ph = ','.join(['%s'] * len(jvs))
        accs = db.execute_query(
            f"SELECT entry_jv, account_name FROM entry_details WHERE enty_values_DR > 0 AND entry_jv IN ({ph})",
            tuple(jvs)) or []
        for a in accs:
            jv_account.setdefault(a['entry_jv'], a['account_name'])

    rows, total_value = [], 0.0
    for l in lines:
        qty = float(l['qty'] or 0)
        cost = float(l['cost'] or 0)
        val = qty * cost
        total_value += val
        rows.append({**l, 'qty': qty, 'cost': cost, 'value': val,
                     'cost_account': jv_account.get(l['jv'], '-')})

    return render_template('inventory_issue_report.html', rows=rows,
                           from_date=from_date, to_date=to_date, total_value=total_value)


# ─────────────────────────────────────────────────────────────────────────────
# Food Costing module (Phase 1: cost cards / recipe costing)
# Isolated feature — new tables, gated by the 'food_costing' menu key. Reuses the
# existing inventory items, per-item purchase cost and per-location stock.
# ─────────────────────────────────────────────────────────────────────────────

def _fc_ensure_schema():
    """Create the food-costing tables on demand (idempotent). Each CREATE is
    guarded like the rest of the app's runtime DDL: on shared hosting a no-op
    CREATE (table already exists) or a DDL privilege quirk must not 500 the
    page — if the tables exist the reads/writes still work."""
    for ddl in (
        """
        CREATE TABLE IF NOT EXISTS fc_menu_item (
            id            INT AUTO_INCREMENT PRIMARY KEY,
            outlet        VARCHAR(50)  NOT NULL DEFAULT 'Restaurant',
            section       VARCHAR(100) DEFAULT 'A la carte',
            item_code     VARCHAR(50)  DEFAULT NULL,
            item_name     VARCHAR(200) NOT NULL,
            selling_price DOUBLE       DEFAULT 0,
            active        TINYINT(1)   DEFAULT 1,
            created_at    DATETIME     DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_fc_item_name (item_name)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """,
        """
        CREATE TABLE IF NOT EXISTS fc_recipe_line (
            id                  INT AUTO_INCREMENT PRIMARY KEY,
            menu_item_id        INT NOT NULL,
            inventory_item_id   INT NOT NULL,
            inventory_item_name VARCHAR(200),
            portion_qty         DOUBLE NOT NULL DEFAULT 0,
            portion_unit        VARCHAR(45) DEFAULT NULL,
            stock_factor        DOUBLE DEFAULT 1,
            INDEX idx_fc_recipe_item (menu_item_id)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """,
        """
        CREATE TABLE IF NOT EXISTS fc_sale (
            id           INT AUTO_INCREMENT PRIMARY KEY,
            sale_date    DATE NOT NULL,
            outlet       VARCHAR(50)  DEFAULT NULL,
            menu_item_id INT          DEFAULT NULL,
            item_code    VARCHAR(50)  DEFAULT NULL,
            item_name    VARCHAR(200) DEFAULT NULL,
            qty          DOUBLE       DEFAULT 0,
            unit_price   DOUBLE       DEFAULT 0,
            plate_cost   DOUBLE       DEFAULT 0,
            cost_method  VARCHAR(20)  DEFAULT NULL,
            depleted     TINYINT(1)   DEFAULT 0,
            created_at   DATETIME     DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_fc_sale_date (sale_date)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """,
    ):
        try:
            db.execute_query(ddl, commit=True)
        except Exception as e:
            logging.warning(f"food_costing schema DDL skipped: {e}")


FC_COSTING_METHODS = ('weighted_avg', 'fifo', 'lifo', 'latest')


def _fc_costing_method():
    row = db.execute_query(
        "SELECT setting_value FROM system_settings WHERE setting_key = 'fc_costing_method'")
    m = (row[0]['setting_value'] if row else '') or 'fifo'
    return m if m in FC_COSTING_METHODS else 'fifo'


def _fc_ingredient_costs(method=None, location=None):
    """Map inventory item NAME -> unit cost, using the chosen inventory costing
    method computed from inventory_recod cost layers (date + qty in/out + unit
    price). weighted_avg / fifo / lifo / latest are all supported because each
    stock-in movement is a cost layer."""
    method = (method or _fc_costing_method()).lower()
    params = []
    loc_clause = ""
    if location:
        loc_clause = " AND inventory_recod_location = %s"
        params.append(location)
    rows = db.execute_query(f"""
        SELECT inventoy_name AS name,
               COALESCE(inventory_recod_action_date, '1900-01-01') AS dt,
               COALESCE(inventory_recod_moument_in, 0)  AS qin,
               COALESCE(inventory_recod_movment_out, 0) AS qout,
               COALESCE(inventory_recod_unit_price, 0)  AS price,
               id AS mid
        FROM inventory_recod
        WHERE inventoy_name IS NOT NULL {loc_clause}
        ORDER BY inventoy_name, dt, id
    """, tuple(params)) or []

    ins = {}          # name -> [[dt, qty, price], ...] in date order (cost layers)
    out_qty = {}      # name -> total quantity consumed
    for r in rows:
        name = r['name']
        qin = float(r['qin'] or 0)
        qout = float(r['qout'] or 0)
        price = float(r['price'] or 0)
        if qin > 0:
            ins.setdefault(name, []).append([r['dt'], qin, price])
        if qout > 0:
            out_qty[name] = out_qty.get(name, 0.0) + qout

    costs = {}
    for name, layers in ins.items():
        if not layers:
            costs[name] = 0.0
            continue
        if method == 'weighted_avg':
            tot_qty = sum(l[1] for l in layers)
            tot_val = sum(l[1] * l[2] for l in layers)
            costs[name] = (tot_val / tot_qty) if tot_qty else 0.0
        elif method == 'latest':
            costs[name] = layers[-1][2]
        elif method in ('fifo', 'lifo'):
            # Walk the layers in consumption order and find the unit cost of the
            # next layer still holding stock after total consumption is applied.
            seq = layers if method == 'fifo' else list(reversed(layers))
            remaining = out_qty.get(name, 0.0)
            nxt = None
            for _dt, qty, price in seq:
                if remaining >= qty:
                    remaining -= qty
                    continue
                nxt = price
                break
            if nxt is None:  # everything consumed -> fall back to a sensible layer
                nxt = layers[-1][2] if method == 'fifo' else layers[0][2]
            costs[name] = nxt
        else:
            costs[name] = layers[-1][2]
    return costs


def _fc_plate_cost(menu_item_id, cost_map=None):
    """Plate cost = sum(portion_qty * stock_factor * ingredient_unit_cost)."""
    if cost_map is None:
        cost_map = _fc_ingredient_costs()
    lines = db.execute_query(
        "SELECT inventory_item_name, portion_qty, stock_factor FROM fc_recipe_line WHERE menu_item_id = %s",
        (menu_item_id,))
    total = 0.0
    for ln in (lines or []):
        unit_cost = cost_map.get(ln['inventory_item_name'], 0.0)
        total += float(ln['portion_qty'] or 0) * float(ln['stock_factor'] or 1) * unit_cost
    return total


def _fc_error_page(title, e):
    """Render a readable error instead of a blank 500 so the cause is visible
    (e.g. a missing template after a partial deploy)."""
    import traceback
    traceback.print_exc()
    from markupsafe import escape
    hint = ""
    if type(e).__name__ == 'TemplateNotFound':
        hint = ("<p style='color:#b25e00'>A template file is missing on the server. "
                "Deploy the latest <code>templates/</code> folder and restart.</p>")
    return (f"<div style='font-family:sans-serif;max-width:760px;margin:40px auto;padding:24px;"
            f"border:1px solid #e2e8f0;border-radius:12px'>"
            f"<h3 style='color:#c0392b;margin-top:0'>{escape(title)} — could not load</h3>{hint}"
            f"<pre style='background:#f8f9fa;padding:14px;border-radius:8px;white-space:pre-wrap'>"
            f"{escape(type(e).__name__)}: {escape(str(e))}</pre>"
            f"<p><a href='/food_costing'>← Back to Food Costing</a></p></div>"), 500


def _fc_safe(title):
    """Decorator: turn an unhandled view error into a readable diagnostic page."""
    def deco(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            try:
                return f(*args, **kwargs)
            except Exception as e:
                return _fc_error_page(title, e)
        return wrapper
    return deco


@app.route('/food_costing')
@login_required
@has_permission('Access_Reports')
@_fc_safe('Food Costing')
def food_costing():
    _fc_ensure_schema()
    items = db.execute_query("SELECT * FROM fc_menu_item ORDER BY outlet, section, item_name") or []
    cost_map = _fc_ingredient_costs()
    cards = []
    for it in items:
        plate = _fc_plate_cost(it['id'], cost_map)
        sp = float(it.get('selling_price') or 0)
        cards.append({
            **it,
            'plate_cost': plate,
            'gross_profit': sp - plate,
            'food_cost_pct': (plate / sp * 100) if sp > 0 else 0,
        })
    kitchen_row = db.execute_query(
        "SELECT setting_value FROM system_settings WHERE setting_key = 'fc_kitchen_location'")
    kitchen_location = kitchen_row[0]['setting_value'] if kitchen_row else ''
    locations = db.execute_query("SELECT inventory_locations_name FROM inventory_locations") or []
    # Existing outlets (defaults + any already used) for the New Cost Card picker.
    outlets = ['Restaurant', 'Bar']
    for r in (db.execute_query("SELECT DISTINCT outlet FROM fc_menu_item WHERE outlet IS NOT NULL AND outlet <> ''") or []):
        if r['outlet'] not in outlets:
            outlets.append(r['outlet'])
    return render_template('food_costing.html', cards=cards,
                           kitchen_location=kitchen_location, locations=locations,
                           outlets=outlets, costing_method=_fc_costing_method())


@app.route('/api/food_costing/ingredients')
@login_required
def food_costing_ingredients():
    """Inventory items for the recipe picker: id, name, code, unit, unit cost
    (costed with the tenant's selected inventory costing method)."""
    cost_map = _fc_ingredient_costs()
    rows = db.execute_query("""
        SELECT ii.id AS id, ii.inventoy_name AS name, ii.inventoy_code AS code,
               ii.inventoy_items_messurment_unit AS unit
        FROM inventoy_items ii
        WHERE ii.active = 1
        ORDER BY ii.inventoy_name
    """) or []
    return jsonify([{'id': r['id'], 'name': r['name'], 'code': r['code'],
                     'unit': r['unit'] or '', 'cost': float(cost_map.get(r['name'], 0.0))} for r in rows])


@app.route('/api/food_costing/item/<int:item_id>')
@login_required
def food_costing_get_item(item_id):
    _fc_ensure_schema()
    it = db.execute_query("SELECT * FROM fc_menu_item WHERE id = %s", (item_id,))
    if not it:
        return jsonify({'success': False, 'error': 'Not found'}), 404
    lines = db.execute_query(
        "SELECT * FROM fc_recipe_line WHERE menu_item_id = %s ORDER BY id", (item_id,)) or []
    return jsonify({'success': True, 'item': it[0], 'lines': lines})


@app.route('/api/food_costing/item', methods=['POST'])
@login_required
def food_costing_save_item():
    _fc_ensure_schema()
    data = request.json or {}
    item_id = data.get('id')
    outlet = (data.get('outlet') or 'Restaurant').strip()
    section = (data.get('section') or '').strip()
    item_code = (data.get('item_code') or '').strip()
    item_name = (data.get('item_name') or '').strip()
    selling_price = float(data.get('selling_price') or 0)
    lines = data.get('lines') or []

    if not item_name:
        return jsonify({'success': False, 'error': 'Item name is required.'}), 400

    if item_id:
        db.execute_query("""UPDATE fc_menu_item SET outlet=%s, section=%s, item_code=%s,
                            item_name=%s, selling_price=%s WHERE id=%s""",
                         (outlet, section, item_code, item_name, selling_price, item_id), commit=True)
    else:
        item_id = db.execute_query("""INSERT INTO fc_menu_item
                            (outlet, section, item_code, item_name, selling_price)
                            VALUES (%s,%s,%s,%s,%s)""",
                         (outlet, section, item_code, item_name, selling_price), commit=True)

    db.execute_query("DELETE FROM fc_recipe_line WHERE menu_item_id = %s", (item_id,), commit=True)
    for ln in lines:
        inv_id = ln.get('inventory_item_id')
        if not inv_id:
            continue
        db.execute_query("""INSERT INTO fc_recipe_line
                            (menu_item_id, inventory_item_id, inventory_item_name,
                             portion_qty, portion_unit, stock_factor)
                            VALUES (%s,%s,%s,%s,%s,%s)""",
                         (item_id, inv_id, (ln.get('inventory_item_name') or '').strip(),
                          float(ln.get('portion_qty') or 0), (ln.get('portion_unit') or '').strip(),
                          float(ln.get('stock_factor') or 1)), commit=True)

    return jsonify({'success': True, 'id': item_id, 'plate_cost': _fc_plate_cost(item_id)})


@app.route('/api/food_costing/item/<int:item_id>/delete', methods=['POST'])
@login_required
def food_costing_delete_item(item_id):
    _fc_ensure_schema()
    db.execute_query("DELETE FROM fc_recipe_line WHERE menu_item_id = %s", (item_id,), commit=True)
    db.execute_query("DELETE FROM fc_menu_item WHERE id = %s", (item_id,), commit=True)
    return jsonify({'success': True})


def _fc_save_setting(key, value):
    existing = db.execute_query("SELECT id FROM system_settings WHERE setting_key = %s", (key,))
    if existing:
        db.execute_query("UPDATE system_settings SET setting_value=%s WHERE setting_key=%s",
                         (value, key), commit=True)
    else:
        db.execute_query("INSERT INTO system_settings (setting_key, setting_value) VALUES (%s, %s)",
                         (key, value), commit=True)


@app.route('/api/food_costing/kitchen_location', methods=['POST'])
@login_required
def food_costing_set_kitchen():
    loc = (request.json or {}).get('location', '').strip()
    _fc_save_setting('fc_kitchen_location', loc)
    return jsonify({'success': True})


@app.route('/api/food_costing/costing_method', methods=['POST'])
@login_required
def food_costing_set_method():
    method = (request.json or {}).get('method', '').strip().lower()
    if method not in FC_COSTING_METHODS:
        return jsonify({'success': False, 'error': 'Invalid method.'}), 400
    _fc_save_setting('fc_costing_method', method)
    return jsonify({'success': True})


def _fc_kitchen_location():
    row = db.execute_query("SELECT setting_value FROM system_settings WHERE setting_key = 'fc_kitchen_location'")
    return (row[0]['setting_value'] if row else '') or ''


def _fc_deplete_kitchen(sale_id, menu_item_id, sale_qty, sale_date, kitchen, cost_map=None):
    """Write OUT movements at the kitchen location for one sale, tagged 'FCS-<id>'
    so they can be reversed on edit/delete. Mirrors the inventory-transfer OUT."""
    if not kitchen or not menu_item_id or float(sale_qty or 0) <= 0:
        return
    if cost_map is None:
        cost_map = _fc_ingredient_costs()
    lines = db.execute_query("""
        SELECT rl.inventory_item_name AS name, rl.portion_qty AS pq, rl.stock_factor AS sf,
               ii.inventoy_code AS code, ii.inventoy_items_messurment_unit AS unit
        FROM fc_recipe_line rl
        LEFT JOIN inventoy_items ii ON ii.id = rl.inventory_item_id
        WHERE rl.menu_item_id = %s
    """, (menu_item_id,)) or []
    uid = get_current_user_id()
    ref = f"FCS-{sale_id}"
    today = datetime.now().date()
    for ln in lines:
        consumed = float(sale_qty) * float(ln['pq'] or 0) * float(ln['sf'] or 1)
        if consumed <= 0:
            continue
        unit_cost = cost_map.get(ln['name'], 0.0)
        db.execute_query("""
            INSERT INTO inventory_recod (
                inventoy_name, inventoy_code, inventory_recod_mesrmet,
                inventory_recod_unit_price, inventory_recod_moument_in, inventory_recod_movment_out,
                inventory_recod_suplier_iv_no, inventory_recod_user_id,
                inventory_recod_user_recod_date, inventory_recod_location,
                inventory_recod_action_date, inventory_recodcol_memo
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (ln['name'], ln['code'], ln['unit'], unit_cost, 0, consumed,
              ref, uid, today, kitchen, sale_date, 'Food costing sale'), commit=True)


def _fc_reverse_depletion(sale_id):
    db.execute_query("DELETE FROM inventory_recod WHERE inventory_recod_suplier_iv_no = %s",
                     (f"FCS-{sale_id}",), commit=True)


def _fc_parse_sales_file(fs):
    """Parse an uploaded CSV/XLSX export: Item Code / Item Name / Quantity."""
    fname = (fs.filename or '').lower()
    if fname.endswith('.xlsx') or fname.endswith('.xlsm'):
        from openpyxl import load_workbook
        wb = load_workbook(fs, read_only=True, data_only=True)
        ws = wb.active
        data = [list(r) for r in ws.iter_rows(values_only=True)]
    else:
        import csv, io
        text = fs.read().decode('utf-8-sig', errors='ignore')
        data = list(csv.reader(io.StringIO(text)))

    data = [r for r in data if r and any(c not in (None, '') for c in r)]
    if not data:
        return []

    header = [str(c).strip().lower() if c is not None else '' for c in data[0]]

    def find(keys):
        for i, h in enumerate(header):
            if any(k in h for k in keys):
                return i
        return None

    ci, ni, qi = find(['code']), find(['name', 'item', 'description']), find(['qty', 'quantity', 'sold'])
    start = 1
    if ci is None and ni is None and qi is None:
        ci, ni, qi, start = 0, 1, 2, 0  # no header -> positional

    out = []
    for r in data[start:]:
        def cell(idx):
            if idx is None or idx >= len(r) or r[idx] is None:
                return ''
            return str(r[idx]).strip()
        code, nm, qraw = cell(ci), cell(ni), cell(qi)
        if not code and not nm:
            continue
        try:
            q = float(str(qraw).replace(',', '')) if qraw else 0.0
        except ValueError:
            q = 0.0
        out.append({'code': code, 'name': nm, 'qty': q})
    return out


@app.route('/api/food_costing/import', methods=['POST'])
@login_required
def food_costing_import():
    _fc_ensure_schema()
    sale_date = (request.form.get('sale_date') or '').strip()
    if not sale_date:
        return jsonify({'success': False, 'error': 'Please choose the sale date.'}), 400
    fs = request.files.get('file')
    if not fs or not fs.filename:
        return jsonify({'success': False, 'error': 'Please choose a file.'}), 400
    try:
        parsed = _fc_parse_sales_file(fs)
    except Exception as e:
        return jsonify({'success': False, 'error': f'Could not read file: {e}'}), 400
    if not parsed:
        return jsonify({'success': False, 'error': 'No rows found in the file.'}), 400

    cards = db.execute_query(
        "SELECT id, outlet, item_code, item_name, selling_price FROM fc_menu_item") or []
    by_code, by_name = {}, {}
    for c in cards:
        if c.get('item_code'):
            by_code[str(c['item_code']).strip().lower()] = c
        by_name[str(c['item_name']).strip().lower()] = c

    kitchen = _fc_kitchen_location()
    method = _fc_costing_method()
    cost_map = _fc_ingredient_costs()

    mapped, unmapped, total_sales, total_cost, plate_cache = 0, [], 0.0, 0.0, {}
    for row in parsed:
        card = by_code.get(row['code'].lower()) if row['code'] else None
        if not card and row['name']:
            card = by_name.get(row['name'].lower())
        qty = row['qty']
        if card:
            mid = card['id']
            if mid not in plate_cache:
                plate_cache[mid] = _fc_plate_cost(mid, cost_map)
            plate = plate_cache[mid]
            price = float(card.get('selling_price') or 0)
            sale_id = db.execute_query("""INSERT INTO fc_sale
                (sale_date, outlet, menu_item_id, item_code, item_name, qty, unit_price, plate_cost, cost_method, depleted)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                (sale_date, card.get('outlet'), mid, row['code'] or card.get('item_code'),
                 card.get('item_name'), qty, price, plate, method, 1 if kitchen else 0), commit=True)
            if kitchen:
                _fc_deplete_kitchen(sale_id, mid, qty, sale_date, kitchen, cost_map)
            mapped += 1
            total_sales += qty * price
            total_cost += qty * plate
        else:
            db.execute_query("""INSERT INTO fc_sale
                (sale_date, outlet, menu_item_id, item_code, item_name, qty, unit_price, plate_cost, cost_method, depleted)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                (sale_date, None, None, row['code'], row['name'], qty, 0, 0, method, 0), commit=True)
            unmapped.append(row['name'] or row['code'])

    return jsonify({'success': True, 'mapped': mapped, 'unmapped': unmapped,
                    'total_sales': total_sales, 'total_cost': total_cost, 'kitchen_set': bool(kitchen)})


@app.route('/food_costing/sales')
@login_required
@has_permission('Access_Reports')
def food_costing_sales():
    try:
        _fc_ensure_schema()
        from_date = (request.args.get('from') or '').strip()
        to_date = (request.args.get('to') or '').strip()
        q = "SELECT * FROM fc_sale WHERE 1=1"
        params = []
        if from_date:
            q += " AND sale_date >= %s"; params.append(from_date)
        if to_date:
            q += " AND sale_date <= %s"; params.append(to_date)
        q += " ORDER BY sale_date DESC, id DESC"
        sales = db.execute_query(q, tuple(params)) or []
        rows, tot_sales, tot_cost = [], 0.0, 0.0
        for s in sales:
            rev = float(s['qty'] or 0) * float(s['unit_price'] or 0)
            cost = float(s['qty'] or 0) * float(s['plate_cost'] or 0)
            tot_sales += rev; tot_cost += cost
            rows.append({**s, 'revenue': rev, 'cost': cost, 'gp': rev - cost,
                         'fc_pct': (cost / rev * 100) if rev > 0 else 0})
        return render_template('food_costing_sales.html', sales=rows, from_date=from_date, to_date=to_date,
                               tot_sales=tot_sales, tot_cost=tot_cost, tot_gp=tot_sales - tot_cost)
    except Exception as e:
        return _fc_error_page('Food Sales History', e)


@app.route('/api/food_costing/sale/<int:sale_id>', methods=['POST'])
@login_required
def food_costing_edit_sale(sale_id):
    _fc_ensure_schema()
    data = request.json or {}
    s = db.execute_query("SELECT * FROM fc_sale WHERE id = %s", (sale_id,))
    if not s:
        return jsonify({'success': False, 'error': 'Not found'}), 404
    s = s[0]
    new_qty = float(data.get('qty', s['qty']) or 0)
    new_date = data.get('sale_date') or s['sale_date']
    if hasattr(new_date, 'strftime'):
        new_date = new_date.strftime('%Y-%m-%d')
    db.execute_query("UPDATE fc_sale SET qty=%s, sale_date=%s WHERE id=%s",
                     (new_qty, new_date, sale_id), commit=True)
    if s.get('menu_item_id') and s.get('depleted'):
        _fc_reverse_depletion(sale_id)
        kitchen = _fc_kitchen_location()
        if kitchen:
            _fc_deplete_kitchen(sale_id, s['menu_item_id'], new_qty, new_date, kitchen)
    return jsonify({'success': True})


@app.route('/api/food_costing/sale/<int:sale_id>/delete', methods=['POST'])
@login_required
def food_costing_delete_sale(sale_id):
    _fc_ensure_schema()
    _fc_reverse_depletion(sale_id)
    db.execute_query("DELETE FROM fc_sale WHERE id = %s", (sale_id,), commit=True)
    return jsonify({'success': True})


@app.route('/api/food_costing/opening_stock', methods=['POST'])
@login_required
def food_costing_opening_stock():
    data = request.json or {}
    item_id = data.get('item_id')
    location = (data.get('location') or '').strip()
    qty = float(data.get('qty') or 0)
    unit_cost = float(data.get('unit_cost') or 0)
    as_date = (data.get('date') or datetime.now().date().strftime('%Y-%m-%d'))
    if not item_id or not location or qty <= 0:
        return jsonify({'success': False, 'error': 'Item, location and quantity are required.'}), 400
    it = db.execute_query("""SELECT inventoy_name AS name, inventoy_code AS code,
                             inventoy_items_messurment_unit AS unit FROM inventoy_items WHERE id = %s""",
                          (item_id,))
    if not it:
        return jsonify({'success': False, 'error': 'Item not found.'}), 404
    it = it[0]
    # NB: inventory_recod_total_value is a GENERATED column — never insert it.
    db.execute_query("""
        INSERT INTO inventory_recod (
            inventoy_name, inventoy_code, inventory_recod_mesrmet,
            inventory_recod_unit_price, inventory_recod_moument_in, inventory_recod_movment_out,
            inventory_recod_suplier_iv_no, inventory_recod_user_id,
            inventory_recod_user_recod_date, inventory_recod_location, inventory_recod_action_date,
            inventory_recodcol_memo
        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (it['name'], it['code'], it['unit'], unit_cost, qty, 0,
          'FC-OPEN', get_current_user_id(), datetime.now().date(), location, as_date, 'Opening stock'),
        commit=True)
    return jsonify({'success': True})


@app.route('/food_costing/report')
@login_required
@has_permission('Access_Reports')
@_fc_safe('Food Costing Report')
def food_costing_report():
    """Phase 3 reports: menu-item profitability, outlet summary and theoretical
    ingredient consumption over a date range."""
    _fc_ensure_schema()
    from_date = (request.args.get('from') or '').strip()
    to_date = (request.args.get('to') or '').strip()

    def date_filter(col):
        clause, params = "", []
        if from_date:
            clause += f" AND {col} >= %s"; params.append(from_date)
        if to_date:
            clause += f" AND {col} <= %s"; params.append(to_date)
        return clause, params

    # 1. Menu item profitability
    f, p = date_filter('sale_date')
    items = db.execute_query(f"""
        SELECT COALESCE(item_name, '(unmapped)') AS item_name,
               MAX(outlet) AS outlet, MAX(item_code) AS item_code,
               MAX(menu_item_id) AS menu_item_id,
               SUM(qty) AS qty,
               SUM(qty * unit_price) AS revenue,
               SUM(qty * plate_cost) AS cost
        FROM fc_sale WHERE 1=1 {f}
        GROUP BY item_name
        ORDER BY revenue DESC
    """, tuple(p)) or []
    for it in items:
        it['qty'] = float(it['qty'] or 0)
        it['revenue'] = float(it['revenue'] or 0)
        it['cost'] = float(it['cost'] or 0)
        it['gp'] = it['revenue'] - it['cost']
        it['fc_pct'] = (it['cost'] / it['revenue'] * 100) if it['revenue'] > 0 else 0

    # 2. Outlet summary
    outlets = db.execute_query(f"""
        SELECT COALESCE(outlet, 'Unassigned') AS outlet,
               SUM(qty * unit_price) AS revenue, SUM(qty * plate_cost) AS cost
        FROM fc_sale WHERE 1=1 {f}
        GROUP BY outlet ORDER BY revenue DESC
    """, tuple(p)) or []
    for o in outlets:
        o['revenue'] = float(o['revenue'] or 0)
        o['cost'] = float(o['cost'] or 0)
        o['gp'] = o['revenue'] - o['cost']
        o['fc_pct'] = (o['cost'] / o['revenue'] * 100) if o['revenue'] > 0 else 0

    # 3. Ingredient consumption (theoretical: sales x recipe portions)
    f2, p2 = date_filter('s.sale_date')
    ingredients = db.execute_query(f"""
        SELECT rl.inventory_item_name AS name,
               MAX(rl.portion_unit) AS unit,
               SUM(s.qty * rl.stock_factor * rl.portion_qty) AS stock_qty
        FROM fc_sale s
        JOIN fc_recipe_line rl ON rl.menu_item_id = s.menu_item_id
        WHERE s.menu_item_id IS NOT NULL {f2}
        GROUP BY rl.inventory_item_name
        ORDER BY stock_qty DESC
    """, tuple(p2)) or []
    cost_map = _fc_ingredient_costs()
    for ing in ingredients:
        ing['stock_qty'] = float(ing['stock_qty'] or 0)
        ing['unit_cost'] = cost_map.get(ing['name'], 0.0)
        ing['value'] = ing['stock_qty'] * ing['unit_cost']

    totals = {'revenue': sum(i['revenue'] for i in items),
              'cost': sum(i['cost'] for i in items)}
    totals['gp'] = totals['revenue'] - totals['cost']
    totals['fc_pct'] = (totals['cost'] / totals['revenue'] * 100) if totals['revenue'] > 0 else 0
    totals['ingredient_value'] = sum(i['value'] for i in ingredients)

    return render_template('food_costing_report.html', items=items, outlets=outlets,
                           ingredients=ingredients, totals=totals,
                           from_date=from_date, to_date=to_date,
                           method=_fc_costing_method())


if __name__ == '__main__':
    app.run(port=5000)
application = app
